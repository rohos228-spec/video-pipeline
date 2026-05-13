"""Импорт v8-xlsx → БД для проекта.

Что делает:
  1. Лист «Общий план» — реконструирует project.general_plan как текст
     (label/value пары + блоки смысловых частей).
  2. Лист «план», строка 49 («закадровый текст»), колонки 3..N —
     извлекает все непустые блоки. Объединяет их в project.script_text
     (через пробелы), и для каждого создаёт запись Frame в БД с
     автоматическим распределением start_ts/end_ts/duration в окне
     60-75 сек (та же логика, что в шаге 3 split_frames.py).
  3. После импорта запускает recompute_status — status автоматически
     поднимется до правильного *_ready по новым данным.

Безопасность:
  - Если в БД уже есть Frame-записи, скрипт НЕ дублирует — он
    обновляет существующие по `number` (или просто пропускает, если
    --no-overwrite-frames).
  - Если у проекта уже есть general_plan/script_text, скрипт ПЕРЕЗАПИСЫВАЕТ
    их содержимым из xlsx. Чтобы оставить — флаг --keep-fields.

Запуск:
    python -m import_from_xlsx 3                         # импорт для проекта #3
    python -m import_from_xlsx 3 --dry-run               # только показать
    python -m import_from_xlsx 3 --keep-fields           # не трогать general_plan/script_text
"""

from __future__ import annotations

import argparse
import asyncio
import sys
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select

from app.db import session_scope
from app.models import Frame, Project
from app.services.project_state import recompute_status
from app.settings import settings

# --- константы под v8-шаблон ---------------------------------------------
SHEET_GENERAL_V8 = "Общий план"
SHEET_PLAN_V8 = "план"
ROW_VOICEOVER_V8 = 49

# Служебные ряды, которые не входят в general_plan (вроде пути к bgm).
_SKIP_LABEL_NEEDLES = ("фоновая музыка", "background music", "bgm")

# Длительность кадра — проп. длине voiceover-блока (русская речь ~14 симв/сек).
# Не пытаемся уложиться в 60-75 сек суммарно, как делает split_frames.py для
# короткой нарезки ChatGPT — у пользователя в xlsx может быть много блоков
# (длинный ролик). Каждый блок просто получает разумную длительность.
CHARS_PER_SEC = 14.0
MIN_FRAME = 1.5
MAX_FRAME = 6.0


def _distribute_durations(cells: list[str]) -> list[float]:
    if not cells:
        return []
    return [
        round(min(max(len(c) / CHARS_PER_SEC, MIN_FRAME), MAX_FRAME), 2)
        for c in cells
    ]


def _read_general_plan(wb) -> str | None:
    """Реконструируем general_plan как текст из листа «Общий план».

    Структура (по факту v8): пары (label, value) в колонках A/B на верхних
    строках + блоки в строках 7+ с колонками: 'Название блока',
    'Основная мысль', 'Подтемы / элементы', 'Функция блока',
    'Как подводит к следующему'.
    """
    if SHEET_GENERAL_V8 not in wb.sheetnames:
        return None
    ws = wb[SHEET_GENERAL_V8]
    lines: list[str] = []
    block_header_row: int | None = None

    for r in range(1, min(ws.max_row, 200) + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        a_s = str(a).strip() if a is not None else ""
        b_s = str(b).strip() if b is not None else ""

        # Заголовок секции в колонке A (без значения в B) — оформляем как ##
        if a_s and not b_s:
            lines.append(f"\n## {a_s}\n")
            # Запоминаем строку — следующая может быть header колонок.
            block_header_row = r + 1
            continue

        # Заголовок таблицы блоков — пять колонок
        if block_header_row == r:
            headers = [
                ws.cell(row=r, column=c).value for c in range(1, 6)
            ]
            if all(h for h in headers):
                # Это header — пропускаем, начнём читать со следующего ряда.
                block_header_row = -1
                continue
            block_header_row = None

        # Обычная пара label / value
        if a_s and b_s:
            a_lower = a_s.lower()
            if any(n in a_lower for n in _SKIP_LABEL_NEEDLES):
                # служебный параметр — в general_plan не попадает
                continue
            lines.append(f"**{a_s}:** {b_s}")
            continue

        # Строка блока (5 колонок, без A-only заголовка)
        if block_header_row == -1 and any(
            ws.cell(row=r, column=c).value for c in range(1, 6)
        ):
            row_cells = [
                str(ws.cell(row=r, column=c).value or "").strip()
                for c in range(1, 6)
            ]
            if row_cells[0]:
                lines.append(f"\n### {row_cells[0]}")
            for label, idx in [
                ("Основная мысль", 1),
                ("Подтемы", 2),
                ("Функция", 3),
                ("Как подводит к следующему", 4),
            ]:
                if row_cells[idx]:
                    lines.append(f"- **{label}:** {row_cells[idx]}")

    text = "\n".join(line for line in lines if line.strip()).strip()
    return text if text else None


def _read_voiceover_blocks(wb) -> list[str]:
    """Читаем закадровые блоки с листа «план», строки ROW_VOICEOVER_V8 (49),
    колонки 3..N. Возвращаем список непустых строк по порядку.
    """
    if SHEET_PLAN_V8 not in wb.sheetnames:
        return []
    ws = wb[SHEET_PLAN_V8]
    out: list[str] = []
    for col in range(3, ws.max_column + 1):
        v = ws.cell(row=ROW_VOICEOVER_V8, column=col).value
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        # Многострочные ячейки — в одну строку.
        s = " ".join(s.split())
        out.append(s)
    return out


async def _do_import(project_id: int, dry_run: bool, keep_fields: bool) -> int:
    async with session_scope() as session:
        project = (
            await session.execute(select(Project).where(Project.id == project_id))
        ).scalar_one_or_none()
        if project is None:
            print(f"[!] проект #{project_id} не найден")
            return 1

        # Путь к xlsx — `data/videos/<slug>/project.xlsx`
        xlsx_path = (
            Path(settings.data_dir) / "videos" / project.slug / "project.xlsx"
        )
        if not xlsx_path.exists():
            print(f"[!] xlsx не найден: {xlsx_path}")
            return 2

        print(f"проект #{project.id} '{project.topic or project.slug}'")
        print(f"  xlsx = {xlsx_path}")

        wb = load_workbook(filename=str(xlsx_path), data_only=True)
        print(f"  sheets = {wb.sheetnames}")

        # 1) general_plan
        new_plan = _read_general_plan(wb)
        if new_plan:
            print(f"  general_plan: {len(new_plan)} симв")
        else:
            print("  general_plan: НЕ найдено в xlsx")

        # 2) voiceover blocks
        blocks = _read_voiceover_blocks(wb)
        print(f"  закадровые блоки на листе «план» R49: {len(blocks)}")
        if blocks:
            print(
                f"    char counts: min={min(len(b) for b in blocks)}, "
                f"max={max(len(b) for b in blocks)}, "
                f"avg={sum(len(b) for b in blocks)//len(blocks)}"
            )

        # 3) script_text — конкатенация блоков через пробел
        new_script = " ".join(blocks) if blocks else None
        if new_script:
            print(f"  script_text (конкатенация): {len(new_script)} симв")

        if dry_run:
            print("\n[dry-run] изменений в БД не делаем")
            return 0

        # --- применяем ---
        changed: list[str] = []
        if new_plan and not keep_fields and project.general_plan != new_plan:
            project.general_plan = new_plan
            changed.append("general_plan")
        if new_script and not keep_fields and project.script_text != new_script:
            project.script_text = new_script
            changed.append("script_text")

        # Frames: создаём, если нет; обновляем по number, если есть
        existing_frames = (
            await session.execute(
                select(Frame).where(Frame.project_id == project.id).order_by(Frame.number)
            )
        ).scalars().all()
        by_number = {f.number: f for f in existing_frames}

        durations = _distribute_durations(blocks)
        t = 0.0
        n_created = 0
        n_updated = 0
        for i, (cell, dur) in enumerate(zip(blocks, durations, strict=True), start=1):
            start_ts = t
            end_ts = t + dur
            fr = by_number.get(i)
            if fr is None:
                session.add(
                    Frame(
                        project_id=project.id,
                        number=i,
                        voiceover_text=cell,
                        start_ts=start_ts,
                        end_ts=end_ts,
                        duration_seconds=dur,
                    )
                )
                n_created += 1
            else:
                # Обновляем только voiceover_text, остальное оставляем как было
                # (вдруг шаги 5/7 уже что-то заполнили — image_prompt/anim_prompt).
                if fr.voiceover_text != cell:
                    fr.voiceover_text = cell
                    n_updated += 1
            t = end_ts

        await session.flush()

        # 4) recompute_status — поднять status до правильного уровня
        old, new, recomputed = await recompute_status(
            session, project, log_prefix="import_from_xlsx"
        )
        if recomputed:
            changed.append(f"status: {old.value} → {new.value}")

        await session.commit()

        print()
        print(f"[ok] frames: создано {n_created}, обновлено {n_updated}")
        if changed:
            print(f"[ok] поля: {', '.join(changed)}")
        else:
            print("[=] поля не изменились")

        # 5) подсказка
        hints = {
            "new": "В TG жми «1. План» — ChatGPT напишет план.",
            "plan_ready": "В TG жми «2. Закадровый текст».",
            "script_ready": "В TG жми «3. Разбивка на блоки».",
            "frames_ready": "В TG жми «4. Hero-картинка».",
            "hero_ready": "В TG жми «5. Промты картинок».",
            "image_prompts_ready": "В TG жми «6. Картинки».",
            "images_ready": "В TG жми «7. Промты анимации».",
        }
        h = hints.get(project.status.value)
        if h:
            print()
            print(f"СЛЕДУЮЩИЙ ШАГ: {h}")
        return 0


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("project_id", type=int)
    p.add_argument("--dry-run", action="store_true")
    p.add_argument(
        "--keep-fields",
        action="store_true",
        help="не перезаписывать project.general_plan и script_text",
    )
    args = p.parse_args()
    return asyncio.run(_do_import(args.project_id, args.dry_run, args.keep_fields))


if __name__ == "__main__":
    sys.exit(main())
