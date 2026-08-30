"""Экспорт таблицы шаблонов сцена→кадры (T/X) в exports/ для скачивания.

SoT — ``templates/shot_templates/shot_templates.json``. После правки таблицы:

    python3 scripts/export_shot_templates.py

Пересобирает ``exports/SCENE_SHOT_TEMPLATES.xlsx`` и
``exports/scene_shot_templates_table.html`` (структура листов — как в
эталонной «Сцены кадры база.xlsx»: Как читать / Каталог / Кадры / Сжатие /
Выбор / Цепи / Parent + rules_ai).
"""

from __future__ import annotations

import html
import json
import sys
from pathlib import Path

_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_ROOT))

_JSON = _ROOT / "templates" / "shot_templates" / "shot_templates.json"
_XLSX = _ROOT / "exports" / "SCENE_SHOT_TEMPLATES.xlsx"
_HTML = _ROOT / "exports" / "scene_shot_templates_table.html"


def _load() -> dict:
    return json.loads(_JSON.read_text(encoding="utf-8"))


def _compression_text(tpl: dict) -> str:
    rows = []
    for c in tpl.get("compression") or []:
        rows.append(f"{c.get('variant')}: {c.get('keep')} — {c.get('when')}")
    return "\n".join(rows)


def _catalog_rows(data: dict) -> list[list[str]]:
    rows = []
    for t in data.get("templates") or []:
        parent_axis = f"Parent: {t.get('parent_default')}\nОсь: {t.get('axis')}"
        if t.get("note"):
            parent_axis += f"\n{t.get('note')}"
        rows.append(
            [
                str(t.get("id") or ""),
                str(t.get("name") or ""),
                str(t.get("when") or ""),
                str(t.get("shots_full") or "").replace(";", "\n"),
                parent_axis,
                _compression_text(t),
                str(t.get("example") or ""),
            ]
        )
    return rows


def _shot_rows(data: dict) -> list[list[str]]:
    rows = []
    for s in data.get("shots") or []:
        rows.append(
            [
                str(s.get("template_id") or ""),
                str(s.get("n") or ""),
                str(s.get("plan") or ""),
                str(s.get("angle") or ""),
                str(s.get("place") or ""),
                str(s.get("action") or ""),
                str(s.get("parent") or ""),
                str(s.get("role") or ""),
                "да" if s.get("required") == 1 else "",
            ]
        )
    return rows


def _compression_rows(data: dict) -> list[list[str]]:
    rows = []
    for t in data.get("templates") or []:
        for c in t.get("compression") or []:
            rows.append(
                [
                    str(t.get("id") or ""),
                    str(c.get("variant") or ""),
                    str(c.get("keep") or ""),
                    str(c.get("when") or ""),
                ]
            )
    return rows


def _select_rows(data: dict) -> list[list[str]]:
    rows = []
    for r in data.get("select") or []:
        step = "—" if str(r.get("step")) == "99" else str(r.get("step") or "")
        rows.append([step, str(r.get("question") or r.get("if") or ""), str(r.get("then_template") or "")])
    return rows


def _chain_rows(data: dict) -> list[list[str]]:
    return [
        [
            str(c.get("id") or "").replace("_", " "),
            str(c.get("chain") or ""),
            str(c.get("what") or ""),
        ]
        for c in data.get("chains") or []
    ]


def _parent_rows(data: dict) -> list[list[str]]:
    return [
        [str(p.get("situation") or ""), str(p.get("parent") or "")]
        for p in data.get("parent_rules") or []
    ]


def _rules_rows(data: dict) -> list[list[str]]:
    return [
        [
            str(r.get("тема") or ""),
            str(r.get("значение") or ""),
            str(r.get("как понять") or ""),
            str(r.get("команда_для_ии") or ""),
        ]
        for r in data.get("rules_ai") or []
    ]


def _readme_rows(data: dict) -> list[list[str]]:
    rows = [
        ["Универсальные шаблоны: сцена → кадры", ""],
        ["", ""],
        ["Лист", "Что внутри"],
        ["Каталог", "все типы T0–T10 и хвосты X1–X2: когда, лестница кадров, parent, сжатие, эталон"],
        ["Кадры", "каждый кадр каждой схемы отдельной строкой"],
        ["Сжатие", "все варианты укорочения лестницы"],
        ["Выбор", "дерево решения: первое «да» → шаблон"],
        ["Цепи", "как склеивать шаблоны в ролике"],
        ["Parent", "сводка правил родителя"],
        ["rules_ai", "машинные правила axis/parent/сжатия для ИИ-агента"],
        ["", ""],
    ]
    for rule in data.get("rules_general") or []:
        rows.append([rule, ""])
    rows.append(["", ""])
    for note in data.get("select_notes") or []:
        rows.append([note, ""])
    for ex in data.get("example_chains") or []:
        rows.append(["", ""])
        rows.append([f"Эталонная цепь «{ex.get('id')}»", ""])
        rows.append(["закадр", str(ex.get("vo") or "")])
        for sc in ex.get("scenes") or []:
            rows.append([str(sc), ""])
        if ex.get("note"):
            rows.append([str(ex.get("note")), ""])
    return rows


_SHEETS: list[tuple[str, list[str], object]] = []


def _build_sheets(data: dict) -> list[tuple[str, list[str], list[list[str]]]]:
    return [
        ("Как читать", [], _readme_rows(data)),
        (
            "Каталог",
            ["id", "Шаблон", "Когда брать", "Кадры (полная лестница)", "Parent / ось", "Сжатие", "Эталон"],
            _catalog_rows(data),
        ),
        (
            "Кадры",
            ["Шаблон", "№", "План", "Ракурс", "Слот", "Действие", "Parent", "Роль", "required"],
            _shot_rows(data),
        ),
        (
            "Сжатие",
            ["Шаблон", "Вариант", "Какие кадры остаются", "Когда"],
            _compression_rows(data),
        ),
        ("Выбор", ["Шаг", "Вопрос (первое «да»)", "Шаблон"], _select_rows(data)),
        ("Цепи", ["История", "Цепь шаблонов", "Что происходит"], _chain_rows(data)),
        ("Parent", ["Ситуация", "Parent"], _parent_rows(data)),
        ("rules_ai", ["тема", "значение", "как понять", "команда_для_ии"], _rules_rows(data)),
    ]


def _write_xlsx(sheets: list[tuple[str, list[str], list[list[str]]]]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True)
    wrap = Alignment(vertical="top", wrap_text=True)
    for name, header, rows in sheets:
        ws = wb.create_sheet(title=name[:31])
        all_rows = ([header] if header else []) + [list(r) for r in rows]
        for r_i, row in enumerate(all_rows, start=1):
            for c_i, val in enumerate(row, start=1):
                cell = ws.cell(row=r_i, column=c_i, value=val)
                cell.alignment = wrap
                if r_i == 1 and header:
                    cell.font = header_font
        width = max((len(r) for r in all_rows), default=0)
        for c_i in range(1, width + 1):
            longest = max(
                (len(str(r[c_i - 1])) if c_i <= len(r) else 0) for r in all_rows[:40]
            )
            ws.column_dimensions[get_column_letter(c_i)].width = min(
                max(12, longest * 0.9), 72
            )
        ws.freeze_panes = "A2" if header else None
    _XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(_XLSX)


def _write_html(sheets: list[tuple[str, list[str], list[list[str]]]]) -> None:
    parts = [
        "<!doctype html><html lang=ru><head><meta charset=utf-8>"
        "<title>Шаблоны сцена → кадры</title><style>",
        "body{font:15px/1.45 system-ui,Segoe UI,sans-serif;margin:20px;color:#111;background:#fff;max-width:1500px}",
        "h1{font-size:24px;margin:0 0 8px}",
        "h2{font-size:18px;margin:32px 0 10px;padding:8px 0;border-bottom:2px solid #222;position:sticky;top:0;background:#fff;z-index:2}",
        ".nav{position:sticky;top:0;background:#fff;padding:10px 0;border-bottom:1px solid #ddd;z-index:3;margin:0 0 16px}",
        ".nav a{margin-right:10px;color:#1a4d8c;font-weight:600}",
        "table{border-collapse:collapse;width:100%;margin:0 0 12px;font-size:13px}",
        "th,td{border:1px solid #ccc;vertical-align:top;padding:8px 10px}",
        "th{background:#eee;text-align:left;position:sticky;top:48px;z-index:1}",
        "td{white-space:pre-wrap}",
        "tr:nth-child(even) td{background:#fafafa}",
        "</style></head><body><h1>Шаблоны сцена → кадры (T0–T10, X1/X2)</h1>",
    ]
    nav = " ".join(
        f'<a href="#s{i}">{html.escape(name)}</a>' for i, (name, _h, _r) in enumerate(sheets)
    )
    parts.append(f'<p class=nav>{nav}</p>')
    for i, (name, header, rows) in enumerate(sheets):
        parts.append(f'<h2 id="s{i}">{html.escape(name)}</h2><table>')
        if header:
            parts.append(
                "<thead><tr>" + "".join(f"<th>{html.escape(h)}</th>" for h in header) + "</tr></thead>"
            )
        parts.append("<tbody>")
        for row in rows:
            parts.append(
                "<tr>" + "".join(f"<td>{html.escape(str(c))}</td>" for c in row) + "</tr>"
            )
        parts.append("</tbody></table>")
    parts.append("</body></html>")
    _HTML.parent.mkdir(parents=True, exist_ok=True)
    _HTML.write_text("".join(parts), encoding="utf-8")


def main() -> None:
    data = _load()
    sheets = _build_sheets(data)
    _write_xlsx(sheets)
    _write_html(sheets)
    print(f"ok: {_XLSX.relative_to(_ROOT)} + {_HTML.relative_to(_ROOT)} "
          f"({len(sheets)} листов, templates={len(data.get('templates') or [])})")


if __name__ == "__main__":
    main()
