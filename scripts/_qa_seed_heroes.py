"""Fill Персонажи sheet for #50 so Outsee hero can run for real."""
from __future__ import annotations

import json
import sqlite3
from pathlib import Path

from openpyxl import load_workbook

XLSX = Path("data/videos/testovyy-trukraym/project.xlsx")
# rows: ID=1, character_id=2?, name=3, look=4, clothes=5, char=6, rules=7
# From excel_characters.py: ROW_ID=1, ROW_NAME=3, ROW_LOOK=4, ROW_CLOTHES=5, ROW_CHAR=6, ROW_RULES=7

HEROES = [
    # col, id, name, look, clothes, char
    (
        2,
        "c01",
        "Римский инженер",
        "мужчина 35 лет, коротко стриженные тёмные волосы, сильный профиль, спокойный взгляд",
        "туника цвета сырого льна, кожаный пояс, плащ-сагум, сандалии-калиги",
        "практичный, уверенный, говорит коротко и по делу",
    ),
    (
        3,
        "c02",
        "Сенатор",
        "мужчина 55 лет, седина на висках, глубокие морщины, тяжёлый взгляд",
        "белая тога с пурпурной каймой, туника под ней",
        "властный, сдержанный, привык командовать",
    ),
    (
        4,
        "c03",
        "Современный горожанин",
        "мужчина 30 лет, обычная современная стрижка, нейтральное лицо",
        "тёмная куртка, джинсы, кроссовки — без логотипов",
        "задумчивый наблюдатель, узнаёт Рим в своём городе",
    ),
]


def main() -> None:
    wb = load_workbook(XLSX)
    ws = wb["Персонажи"]
    for col, cid, name, look, clothes, char in HEROES:
        ws.cell(1, col).value = cid
        ws.cell(2, col).value = cid
        ws.cell(3, col).value = name
        ws.cell(4, col).value = look
        ws.cell(5, col).value = clothes
        ws.cell(6, col).value = char
        ws.cell(7, col).value = ""
    # clear unused template cols 5+
    for col in range(5, 8):
        for row in range(1, 8):
            if col > 4:
                pass
    wb.save(XLSX)
    print("saved", XLSX)

    conn = sqlite3.connect("data/state.db")
    row = conn.execute("select meta from projects where id=50").fetchone()
    meta = json.loads(row[0] or "{}")
    meta.pop("hero_skipped_empty", None)
    meta["excel_hero_enabled"] = True
    conn.execute(
        "update projects set meta=? where id=50",
        (json.dumps(meta, ensure_ascii=False),),
    )
    conn.commit()
    print("cleared hero_skipped_empty in meta")


if __name__ == "__main__":
    main()
