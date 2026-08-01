"""E2E: полный проект по пайплайну с заглушками GPT/медиа.

Реальное: модели, apply-ops, экспорт в xlsx, backfill, шаг anim_pr целиком
(с подменённым GPT-клиентом), harness-гейты. Проверка после каждого шага.
"""

import asyncio
import re
import sys
from pathlib import Path

sys.path.insert(0, "/workspace")

from sqlalchemy import select
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from app.models import Base, Frame, FrameStatus, Project, ProjectStatus
from app.services import db_apply, db_v2
from app.settings import settings

DATA = Path("/tmp/e2e-data")
DATA.mkdir(exist_ok=True)
settings.data_dir = DATA
settings.sqlite_path = DATA / "state.db"

REPORT: list[tuple[str, bool, str]] = []


def step(name: str, ok: bool, detail: str = "") -> None:
    REPORT.append((name, ok, detail))
    print(f"{'OK  ' if ok else 'FAIL'} {name}: {detail}")


def make_png(path: Path, color: tuple[int, int, int]) -> None:
    from PIL import Image

    Image.new("RGB", (96, 128), color).save(path)


async def main() -> int:
    for old in (DATA / "state.db",):
        if old.exists():
            old.unlink()
    engine = create_async_engine(f"sqlite+aiosqlite:///{DATA / 'state.db'}")
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
        await db_v2.migrate_db_v2_schema(conn)
    factory = async_sessionmaker(engine, expire_on_commit=False)

    async with factory() as s:
        from openpyxl import Workbook

        p = Project(slug="e2e-full", title="E2E полный", topic="Рачок в неоне", status=ProjectStatus.new, meta={})
        s.add(p)
        await s.flush()
        p.data_dir.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "план"
        wb.create_sheet("Общий план")
        wb.save(p.data_dir / "project.xlsx")
        wb.close()
        await s.commit()
        pid = p.id

        # ── 1. plan через apply-ops (target=project) ──
        r = await db_apply.apply_ops(
            s, p, [{"target": "project", "fields": {"общий_план": "Эпизод 1: рачок идёт в неоновый город."}}]
        )
        await s.commit()
        p.status = ProjectStatus.plan_ready
        await s.commit()
        step("plan", r["updated"] == 1 and r["exported"]["cells"] > 0, f"exported={r['exported']}")

        # ── 2. script (voiceover.txt) ──
        vo_blocks = [
            "Рачок просыпается в бочке с ржавчиной.",
            "Он выбирается наружу и видит неоновый город.",
            "Где-то внизу мигает вывеска «ЕДА».",
        ]
        (p.data_dir / "voiceover.txt").write_text("\n\n".join(vo_blocks), encoding="utf-8")
        p.status = ProjectStatus.script_ready
        await s.commit()
        step("script", (p.data_dir / "voiceover.txt").is_file(), "voiceover.txt")

        # ── 3. split: кадры + закадр/длительность через apply-ops ──
        for i in range(1, 4):
            s.add(Frame(project_id=pid, number=i, voiceover_text="", attrs={}))
        await s.commit()
        await s.refresh(p)
        await db_v2.backfill_project_v2(s, p)
        await s.commit()
        frames = list(
            (await s.execute(select(Frame).where(Frame.project_id == pid).order_by(Frame.number))).scalars()
        )
        ops = [
            {
                "frame_uuid": f.uuid,
                "fields": {"закадр": vo_blocks[f.number - 1], "длительность": 2.0 + f.number},
            }
            for f in frames
        ]
        r = await db_apply.apply_ops(s, p, ops)
        await s.commit()
        p.status = ProjectStatus.frames_ready
        await s.commit()
        step("split", r["updated"] == 3, f"frames=3 updated={r['updated']}")

        # ── 4. img_pr через apply-ops ──
        ops = [
            {"frame_uuid": f.uuid, "fields": {"промт_картинки": f"knitted style, neon crab, frame {f.number}, 9:16 vertical"}}
            for f in frames
        ]
        r = await db_apply.apply_ops(s, p, ops)
        await s.commit()
        p.status = ProjectStatus.image_prompts_ready
        await s.commit()
        step("img_pr", r["updated"] == 3, f"updated={r['updated']}")

        # ── 5. img (заглушка медиа: реальные PNG) ──
        scenes = p.data_dir / "scenes"
        scenes.mkdir(exist_ok=True)
        colors = [(200, 60, 60), (60, 200, 120), (60, 90, 220)]
        for f in frames:
            make_png(scenes / f"frame_{f.number:03d}_abcd1234.png", colors[f.number - 1])
            f.status = FrameStatus.image_generated
        await s.commit()
        p.status = ProjectStatus.images_ready
        await s.commit()
        step("img", len(list(scenes.glob("*.png"))) == 3, "3 png (stub media)")

        # ── 6. anim_pr — РЕАЛЬНЫЙ шаг с подменённым GPT ──
        p.status = ProjectStatus.generating_animation_prompts
        await s.commit()

        class FakeGpt:
            async def new_conversation(self) -> None:
                return None

            async def ask_anim_pr_initial(self, text, file, **kw):
                return "принято"

            async def ask_anim_pr_batch(self, msg, images, **kw):
                ids = re.findall(r"ID изображения: (\[ID: [^\]]+\])", msg)
                return "\n\n".join(
                    f"ID изображения: {i}\nтекст анимации: slow neon push-in, dust motes, cinematic"
                    for i in ids
                )

        import app.orchestrator.steps.make_animation_prompts as anim_step

        anim_step.get_gpt_client = lambda: FakeGpt()
        gate_err = ""
        try:
            await anim_step.run(s, p, None)
        except Exception as e:  # noqa: BLE001
            gate_err = str(e)
        await s.commit()
        await s.refresh(p)
        frames = list(
            (await s.execute(select(Frame).where(Frame.project_id == pid).order_by(Frame.number))).scalars()
        )
        anim_filled = sum(1 for f in frames if (f.animation_prompt or "").strip())
        vers = 0
        from app.models import PromptVersion

        vers = (
            await s.execute(
                select(PromptVersion).where(PromptVersion.project_id == pid, PromptVersion.kind == "video")
            )
        ).scalars().all()
        step(
            "anim_pr (real step + apply-ops + gate)",
            p.status is ProjectStatus.animation_prompts_ready and anim_filled == 3 and not gate_err,
            f"status={p.status.value} filled={anim_filled}/3 video_versions={len(vers)} gate_err={gate_err or '—'}",
        )

        # ── 7. video/audio/music (заглушки файлов) ──
        vids = p.data_dir / "videos"
        vids.mkdir(exist_ok=True)
        for f in frames:
            (vids / f"clip_{f.number:03d}_abcd1234.mp4").write_bytes(b"\x00\x00\x00\x18ftypmp42" + b"\x00" * 50)
        (p.data_dir / "audio").mkdir(exist_ok=True)
        (p.data_dir / "audio" / "voice.mp3").write_bytes(b"ID3" + b"\x00" * 50)
        (p.data_dir / "music").mkdir(exist_ok=True)
        (p.data_dir / "music" / "bgm.mp3").write_bytes(b"ID3" + b"\x00" * 50)
        p.status = ProjectStatus.videos_ready
        await s.commit()
        step("video/audio/music", len(list(vids.glob("*.mp4"))) == 3, "3 mp4 + audio + music (stub media)")

        # ── 8. assemble + ПОЛНЫЙ harness ──
        final = p.data_dir / "final"
        final.mkdir(exist_ok=True)
        (final / "final_abcd1234.mp4").write_bytes(b"\x00\x00\x00\x18ftypmp42" + b"\x00" * 100)
        p.status = ProjectStatus.assembled
        await s.commit()
        from app.services.agent_harness import run_harness_verify

        rep = await run_harness_verify(s, p, allow_repair=False, include_http=False)
        await s.commit()
        bad = [f"{c.name}({c.detail})" for c in rep.checks if not c.ok]
        step("assemble + full harness", rep.ok, f"ok={rep.ok} bad={bad or '—'}")

        # ── 9. xlsx: итоговая консистентность DB↔Excel ──
        from openpyxl import load_workbook

        wb2 = load_workbook(p.data_dir / "project.xlsx")
        try:
            ws2 = wb2["план"]
            row_checks = {
                "R45": all(ws2.cell(row=45, column=3 + i).value for i in range(3)),
                "R48": all(ws2.cell(row=48, column=3 + i).value for i in range(3)),
                "R49": all(ws2.cell(row=49, column=3 + i).value for i in range(3)),
            }
            general = wb2["Общий план"]["B2"].value
        finally:
            wb2.close()
        step(
            "xlsx consistency",
            all(row_checks.values()) and bool(general),
            f"{row_checks} общий_план={'да' if general else 'нет'}",
        )

    await engine.dispose()
    print()
    failed = [x for x in REPORT if not x[1]]
    print(f"RESULT: {len(REPORT) - len(failed)}/{len(REPORT)} шагов OK")
    return 1 if failed else 0


sys.exit(asyncio.run(main()))
