"""Deep artifact compare #50 vs sekty for LIVE QA."""
from __future__ import annotations

import json
import sqlite3
from collections import Counter
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]


def scan_dir(p: Path) -> dict:
    out: dict = {}
    if not p.exists():
        return {"missing": True}
    for child in sorted(p.iterdir()):
        if child.is_dir():
            files = [f for f in child.rglob("*") if f.is_file()]
            out[child.name] = {
                "files": len(files),
                "kb": sum(f.stat().st_size for f in files) // 1024,
                "sample": [str(f.relative_to(p)) for f in sorted(files)[:8]],
            }
        else:
            out[child.name] = child.stat().st_size
    return out


def main() -> None:
    truk = ROOT / "data/videos/testovyy-trukraym"
    sekty = ROOT / "data/videos/sekty"
    print("=== TRUKRAYM TREE ===")
    print(json.dumps(scan_dir(truk), ensure_ascii=False, indent=2))
    print("=== SEKTY TREE (summary) ===")
    s = scan_dir(sekty)
    for k, v in s.items():
        if isinstance(v, dict) and "files" in v:
            print(f"  {k}: {v['files']} files {v['kb']}KB")
        else:
            print(f"  {k}: {v}")

    db = sqlite3.connect(ROOT / "data/state.db")
    db.row_factory = sqlite3.Row
    arts = db.execute(
        "SELECT kind, path, frame_id FROM artifacts WHERE project_id=50 ORDER BY kind, frame_id"
    ).fetchall()
    print("ARTIFACTS_DB", dict(Counter(r["kind"] for r in arts)))
    for kind in sorted({r["kind"] for r in arts}):
        rows = [r for r in arts if r["kind"] == kind]
        print(f"  {kind}: {len(rows)}")
        for r in rows[:3]:
            print(f"    f={r['frame_id']} {r['path']}")

    # frames
    fcols = {c[1] for c in db.execute("PRAGMA table_info(frames)").fetchall()}
    num_col = "number" if "number" in fcols else ("idx" if "idx" in fcols else "id")
    frames = db.execute(
        f"SELECT id, {num_col} AS num, "
        "image_prompt IS NOT NULL as has_ip, "
        "animation_prompt IS NOT NULL as has_ap "
        f"FROM frames WHERE project_id=50 ORDER BY {num_col}"
    ).fetchall()
    print(f"FRAMES {len(frames)}")
    for f in frames:
        print(
            f"  #{f['num']} id={f['id']} img_pr={f['has_ip']} anim_pr={f['has_ap']}"
        )

    # xlsx via openpyxl from venv if available
    try:
        from openpyxl import load_workbook

        for name in ("testovyy-trukraym", "sekty"):
            p = ROOT / f"data/videos/{name}/project.xlsx"
            wb = load_workbook(p, read_only=True, data_only=True)
            print(f"XLSX {name}: {wb.sheetnames}")
            wb.close()
    except Exception as e:
        print("XLSX ERR", type(e).__name__, e)


if __name__ == "__main__":
    main()
