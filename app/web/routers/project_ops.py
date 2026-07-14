"""Операции проекта для веб-студии: xlsx, ассеты, пауза, сброс шага."""

from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse
from loguru import logger
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Artifact, ArtifactKind, Project, ProjectStatus
from app.services.event_bus import publish_project_event
from app.services.project_control import pause_project as pause_project_svc
from app.services.project_control import resume_project as resume_project_svc
from app.services.project_control import stop_project_running
from app.services.reset_step import reset_step
from app.services.run_sync import (
    ensure_run_for_project,
    reset_nodes_from_step,
    sync_run_for_project,
    _get_default_workflow_id,
)
from app.services.chatgpt_xlsx import sync_project_xlsx
from app.settings import settings
from app.storage import ProjectSheet
from app.web.deps import get_session
from app.web.project_dto import project_to_detail
from app.web.schemas import ProjectDetail

router = APIRouter(prefix="/projects", tags=["project-ops"])


def _project_or_404(project: Project | None) -> Project:
    if project is None:
        raise HTTPException(status_code=404, detail="project not found")
    return project


@router.post("/{project_id}/pause", response_model=ProjectDetail)
async def pause_project(
    project_id: int, session: AsyncSession = Depends(get_session)
) -> Project:
    p = _project_or_404(await session.get(Project, project_id))
    await pause_project_svc(session, p)
    await session.commit()
    await sync_run_for_project(project_id)
    await session.refresh(p)
    await publish_project_event(project_id, event_type="project_updated", payload={"paused": True})
    return p


@router.post("/{project_id}/resume", response_model=ProjectDetail)
async def resume_project(
    project_id: int, session: AsyncSession = Depends(get_session)
) -> Project:
    p = _project_or_404(await session.get(Project, project_id))
    await resume_project_svc(session, p)
    await session.commit()
    await session.refresh(p)
    await publish_project_event(project_id, event_type="project_updated", payload={"resumed": True})
    return p


@router.post("/{project_id}/continue")
async def continue_project(
    project_id: int, session: AsyncSession = Depends(get_session)
) -> dict:
    """Снять stop/паузу и продвинуть проект на следующий шаг (если *_ready)."""
    from app.orchestrator.auto_advance import continue_project_pipeline

    p = _project_or_404(await session.get(Project, project_id))
    info = await continue_project_pipeline(session, p, bot=None)
    await session.commit()
    await session.refresh(p)
    await publish_project_event(
        project_id,
        event_type="project_updated",
        payload={"continue": True, **info},
    )
    return {"project": project_to_detail(p), **info}


@router.post("/{project_id}/stop")
async def stop_project(
    project_id: int, session: AsyncSession = Depends(get_session)
) -> dict:
    p = _project_or_404(await session.get(Project, project_id))
    info = await stop_project_running(session, p)
    if not info["ok"]:
        raise HTTPException(status_code=400, detail=info["message"])
    await session.commit()
    await sync_run_for_project(project_id)
    await session.refresh(p)
    await publish_project_event(
        project_id,
        event_type="project_updated",
        payload={"stopped": True, "message": info["message"]},
    )
    return {
        "project": project_to_detail(p),
        "message": info["message"],
        "generation_still_active": info["generation_still_active"],
        "xlsx_stopped": info["xlsx_stopped"],
    }


@router.post("/{project_id}/finish/images")
async def finish_missing_images(
    project_id: int, session: AsyncSession = Depends(get_session)
) -> dict:
    """Доделка картинок: frame_NNN_*.png без файла → generating_images."""
    from app.services.finish_missing import trigger_finish_missing_images

    p = _project_or_404(await session.get(Project, project_id))
    info = await trigger_finish_missing_images(session, p)
    await session.commit()
    await sync_run_for_project(project_id)
    await session.refresh(p)
    await publish_project_event(
        project_id,
        event_type="project_updated",
        payload={"finish_missing": "images", **info},
    )
    return {**info, "project": project_to_detail(p)}


@router.post("/{project_id}/finish/animation-prompts")
async def resume_animation_prompts(
    project_id: int, session: AsyncSession = Depends(get_session)
) -> dict:
    """Догонка промтов анимации: plan R48 → БД → generating_animation_prompts."""
    from app.services.finish_missing import trigger_resume_animation_prompts

    p = _project_or_404(await session.get(Project, project_id))
    info = await trigger_resume_animation_prompts(session, p)
    await session.commit()
    await sync_run_for_project(project_id)
    await session.refresh(p)
    await publish_project_event(
        project_id,
        event_type="project_updated",
        payload={"finish_missing": "animation_prompts", **info},
    )
    return {**info, "project": project_to_detail(p)}


@router.post("/{project_id}/finish/videos")
async def finish_missing_videos(
    project_id: int, session: AsyncSession = Depends(get_session)
) -> dict:
    """Доделка видео: clip_NNN_*.mp4 без файла → generating_videos."""
    from app.services.finish_missing import trigger_finish_missing_videos

    p = _project_or_404(await session.get(Project, project_id))
    info = await trigger_finish_missing_videos(session, p)
    await session.commit()
    await sync_run_for_project(project_id)
    await session.refresh(p)
    await publish_project_event(
        project_id,
        event_type="project_updated",
        payload={"finish_missing": "videos", **info},
    )
    return {**info, "project": project_to_detail(p)}


@router.post("/{project_id}/mass-lanes/parse-topics")
async def parse_mass_topics_xlsx(
    project_id: int,
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_session),
) -> dict:
    """Парсит любой xlsx (построчно темы) и сохраняет очередь на родителе (правило B)."""
    import tempfile

    from app.storage.mass_topics import parse_topics_xlsx
    from app.services.mass_factory import apply_topics_upload

    parent = _project_or_404(await session.get(Project, project_id))
    suffix = Path(file.filename or "topics.xlsx").suffix or ".xlsx"
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp_path = Path(tmp.name)
    try:
        content = await file.read()
        if not content:
            raise HTTPException(status_code=400, detail="пустой файл")
        tmp_path.write_bytes(content)
        try:
            topics = parse_topics_xlsx(tmp_path)
        except Exception as exc:
            logger.warning("mass-lanes/parse-topics: parse failed for #{}: {}", project_id, exc)
            raise HTTPException(
                status_code=400,
                detail=f"не удалось прочитать Excel: {exc}",
            ) from exc
    finally:
        tmp_path.unlink(missing_ok=True)

    if not topics:
        raise HTTPException(status_code=400, detail="в Excel не найдено тем (построчно)")

    try:
        result = await apply_topics_upload(
            session,
            parent,
            topics=topics,
            filename=file.filename or "topics.xlsx",
        )
        await session.commit()
    except HTTPException:
        raise
    except Exception as exc:
        await session.rollback()
        logger.exception("mass-lanes/parse-topics failed for project #{}", project_id)
        raise HTTPException(
            status_code=500,
            detail=f"ошибка сохранения Excel: {exc}",
        ) from exc

    await publish_project_event(
        project_id,
        event_type="project_updated",
        payload={"mass_excel_upload": result.get("count"), "revision": result.get("revision")},
    )
    return {"topics": topics, "cards": [], "count": len(topics), **result}


@router.get("/{project_id}/mass-factory/status")
async def mass_factory_status_endpoint(
    project_id: int,
    session: AsyncSession = Depends(get_session),
) -> dict:
    from app.services.mass_factory import mass_factory_status

    parent = _project_or_404(await session.get(Project, project_id))
    return await mass_factory_status(session, parent)


@router.post("/{project_id}/mass-lanes/start")
async def start_mass_lanes(
    project_id: int,
    payload: dict,
    session: AsyncSession = Depends(get_session),
) -> dict:
    """Очередь видео: родитель-шаблон, генерация в дочерних проектах."""
    from app.services.mass_factory import start_mass_queue
    from app.web.routers.projects import _slugify

    template = _project_or_404(await session.get(Project, project_id))
    meta_template = dict(template.meta or {})
    topics: list[str] = [str(t).strip() for t in (payload.get("topics") or []) if str(t).strip()]
    if not topics:
        meta_topics = meta_template.get("mass_queue_topics") or meta_template.get("mass_excel_topics")
        if isinstance(meta_topics, list):
            topics = [str(t).strip() for t in meta_topics if str(t).strip()]
    if not topics:
        bindings = meta_template.get("excel_lane_bindings")
        if isinstance(bindings, list):
            ordered = sorted(
                [b for b in bindings if isinstance(b, dict)],
                key=lambda b: int(b.get("topic_index") or 0),
            )
            for b in ordered:
                t = str(b.get("topic") or "").strip()
                if t:
                    topics.append(t)
    try:
        result = await start_mass_queue(session, template, topics=topics or None, slugify=_slugify)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        logger.exception("mass-lanes/start failed for project #{}", project_id)
        raise HTTPException(status_code=500, detail=f"ошибка запуска очереди: {exc}") from exc

    await session.commit()
    return {
        "created": [{"id": result["started_id"], "topic": result["topic"]}],
        "count": 1,
        "queue_size": result["queue_size"],
        "remaining": result["remaining"],
        "started_id": result["started_id"],
    }


@router.post("/{project_id}/steps/{step_code}/reset", response_model=ProjectDetail)
async def reset_project_step(
    project_id: int,
    step_code: str,
    session: AsyncSession = Depends(get_session),
) -> Project:
    p = _project_or_404(await session.get(Project, project_id))
    try:
        summary = await reset_step(session, p, step_code)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e
    if summary.get("error"):
        raise HTTPException(status_code=400, detail=str(summary["error"]))
    wf_id = await _get_default_workflow_id()
    if wf_id is not None:
        await ensure_run_for_project(project_id, wf_id)
    await reset_nodes_from_step(session, project_id, step_code)
    await session.flush()
    await session.commit()
    await session.refresh(p)
    await sync_run_for_project(project_id)
    await session.refresh(p)
    await publish_project_event(
        project_id,
        event_type="project_updated",
        payload={"reset_step": step_code},
    )
    return p


@router.get("/{project_id}/excel-hero")
async def get_excel_hero(
    project_id: int, session: AsyncSession = Depends(get_session)
) -> dict:
    """Текущее состояние excel-hero в project.meta (если есть)."""
    p = _project_or_404(await session.get(Project, project_id))
    meta = dict(p.meta or {})
    cfg = meta.get("excel_hero") or {}
    chars = cfg.get("characters") if isinstance(cfg, dict) else None
    return {"loaded": bool(chars), "characters": chars or []}


@router.post("/{project_id}/excel-hero/load")
async def load_excel_hero(
    project_id: int, session: AsyncSession = Depends(get_session)
) -> dict:
    """Перечитать лист «Персонажи» из project.xlsx → project.meta['excel_hero'].

    После этого шаг hero пойдёт по excel-ветке (`_run_excel`), беря данные
    из meta, без необходимости заполнять hero_descriptions/hero_count.
    """
    from app.services.excel_characters import parse_persons_sheet

    p = _project_or_404(await session.get(Project, project_id))
    xlsx = p.data_dir / "project.xlsx"
    if not xlsx.exists():
        raise HTTPException(
            status_code=404,
            detail=f"project.xlsx не найден по пути {xlsx}",
        )
    try:
        chars = parse_persons_sheet(xlsx)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(
            status_code=400,
            detail=f"не удалось распарсить лист «Персонажи»: {e}",
        ) from e
    if not chars:
        raise HTTPException(
            status_code=400,
            detail="на листе «Персонажи» нет ни одного заполненного персонажа",
        )
    meta = dict(p.meta or {})
    meta["excel_hero"] = {"characters": [c.to_dict() for c in chars]}
    p.meta = meta
    p.updated_at = datetime.utcnow()
    await session.commit()
    await publish_project_event(
        project_id,
        event_type="project_updated",
        payload={"excel_hero": len(chars)},
    )
    return {
        "loaded": True,
        "count": len(chars),
        "characters": [c.to_dict() for c in chars],
    }


@router.delete("/{project_id}/excel-hero", status_code=204)
async def clear_excel_hero(
    project_id: int, session: AsyncSession = Depends(get_session)
) -> None:
    """Убрать excel_hero — hero пойдёт по обычной ветке (hero_descriptions)."""
    p = _project_or_404(await session.get(Project, project_id))
    meta = dict(p.meta or {})
    if "excel_hero" in meta:
        del meta["excel_hero"]
        p.meta = meta
        p.updated_at = datetime.utcnow()
        await session.commit()
        await publish_project_event(
            project_id, event_type="project_updated", payload={"excel_hero": 0}
        )


@router.get("/{project_id}/xlsx")
async def download_xlsx(
    project_id: int, session: AsyncSession = Depends(get_session)
) -> FileResponse:
    p = _project_or_404(await session.get(Project, project_id))
    xlsx = p.data_dir / "project.xlsx"
    if not xlsx.exists():
        sheet = ProjectSheet(file_path=xlsx)
        sheet.ensure_initialized(project_id=p.id, slug=p.slug)
    if not xlsx.exists():
        raise HTTPException(status_code=404, detail="project.xlsx not found")
    return FileResponse(
        path=str(xlsx),
        filename=f"{p.slug}-project.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.post("/{project_id}/xlsx/reload", response_model=ProjectDetail)
async def reload_xlsx(
    project_id: int, session: AsyncSession = Depends(get_session)
) -> Project:
    p = _project_or_404(await session.get(Project, project_id))
    xlsx = p.data_dir / "project.xlsx"
    if not xlsx.exists():
        raise HTTPException(status_code=404, detail="project.xlsx not found")
    await sync_project_xlsx(session, p, xlsx)
    await session.commit()
    await session.refresh(p)
    await publish_project_event(project_id, event_type="project_updated", payload={"xlsx": "reloaded"})
    return p


@router.post("/{project_id}/xlsx/upload", response_model=ProjectDetail)
async def upload_xlsx(
    project_id: int,
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_session),
) -> Project:
    p = _project_or_404(await session.get(Project, project_id))
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="need .xlsx file")
    dest = p.data_dir / "project.xlsx"
    dest.parent.mkdir(parents=True, exist_ok=True)
    content = await file.read()
    import tempfile

    from app.services.xlsx_versioning import validate_xlsx

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(content)
        tmp_path = Path(tmp.name)
    validation_err = validate_xlsx(tmp_path)
    if validation_err is not None:
        tmp_path.unlink(missing_ok=True)
        raise HTTPException(status_code=400, detail=validation_err)
    tmp_path.replace(dest)
    await sync_project_xlsx(session, p, dest)
    await session.commit()
    await session.refresh(p)
    await publish_project_event(project_id, event_type="project_updated", payload={"xlsx": "uploaded"})
    return p


@router.get("/{project_id}/xlsx/preview")
async def preview_xlsx(
    project_id: int,
    sheet: str | None = Query(None),
    max_rows: int = Query(40, ge=1, le=500),
    max_cols: int = Query(80, ge=1, le=200),
    start_row: int = Query(1, ge=1, le=500),
    row: int | None = Query(None, ge=1, le=500),
    raw: bool = Query(False),
    session: AsyncSession = Depends(get_session),
) -> dict:
    p = _project_or_404(await session.get(Project, project_id))
    xlsx = p.data_dir / "project.xlsx"
    if not xlsx.exists():
        sheet = ProjectSheet(file_path=xlsx)
        sheet.ensure_initialized(project_id=p.id, slug=p.slug)
    if not xlsx.exists():
        return {
            "path": str(xlsx),
            "sheets": [],
            "active_sheet": "",
            "headers": [],
            "rows": [],
            "cells": [],
        }
    from openpyxl import load_workbook

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    sheets = wb.sheetnames
    active = sheet if sheet in sheets else (sheets[0] if sheets else "")

    if row is not None and active:
        ws = wb[active]
        row_iter = ws.iter_rows(
            min_row=row, max_row=row, max_col=max_cols, values_only=True
        )
        row_vals = next(row_iter, None)
        cells = (
            ["" if c is None else str(c) for c in row_vals]
            if row_vals is not None
            else []
        )
        wb.close()
        return {
            "path": str(xlsx),
            "sheets": sheets,
            "active_sheet": active,
            "row": row,
            "cells": cells,
        }

    headers: list[str] = []
    rows: list[list[str]] = []
    if active:
        ws = wb[active]
        end_row = min(start_row + max_rows - 1, ws.max_row or start_row)
        for i, row_vals in enumerate(
            ws.iter_rows(
                min_row=start_row,
                max_row=end_row,
                max_col=max_cols,
                values_only=True,
            )
        ):
            cells = ["" if c is None else str(c) for c in row_vals]
            if len(cells) > max_cols:
                cells = cells[:max_cols]
            elif len(cells) < max_cols:
                cells = cells + [""] * (max_cols - len(cells))
            if raw:
                rows.append(cells)
            elif i == 0:
                headers = cells
                continue
            else:
                rows.append(cells)
            if len(rows) >= max_rows:
                break
    wb.close()
    return {
        "path": str(xlsx),
        "sheets": sheets,
        "active_sheet": active,
        "headers": headers if not raw else [],
        "rows": rows,
    }


@router.get("/{project_id}/montage-board")
async def montage_board(
    project_id: int,
    session: AsyncSession = Depends(get_session),
) -> dict:
    """Сетка монтажа: озвучка, персонажи, shot1/2 картинки и видео по кадрам."""
    p = _project_or_404(await session.get(Project, project_id))
    from app.services.montage_board import build_montage_board

    return await build_montage_board(session, p)


@router.get("/{project_id}/assets")
async def list_project_assets(
    project_id: int,
    kind: str = Query("all", pattern="^(all|hero|items|images|videos|audio|final|text)$"),
    session: AsyncSession = Depends(get_session),
) -> list[dict]:
    p = _project_or_404(await session.get(Project, project_id))
    out: list[dict] = []

    # DB artifacts
    if kind in ("all", "hero", "items", "images", "videos", "audio", "final"):
        from sqlalchemy import select

        arts = (
            await session.execute(select(Artifact).where(Artifact.project_id == project_id))
        ).scalars().all()
        kind_map = {
            "hero": {ArtifactKind.hero_reference},
            "items": {ArtifactKind.item_reference},
            "images": {ArtifactKind.scene_image},
            "videos": {ArtifactKind.scene_video},
            "audio": {ArtifactKind.audio, ArtifactKind.subtitle},
            "final": {ArtifactKind.final_video},
        }
        for a in arts:
            if kind != "all" and a.kind not in kind_map.get(kind, set()):
                continue
            rel = _rel_path(a.path) if a.path else None
            out.append(
                {
                    "source": "artifact",
                    "id": a.uuid,
                    "kind": a.kind.value if hasattr(a.kind, "value") else str(a.kind),
                    "path": a.path,
                    "preview_url": f"/api/artifacts/{a.uuid}/file" if a.uuid else None,
                    "frame_id": a.frame_id,
                    "meta": a.meta or {},
                }
            )

    # Disk scan (fallback)
    base = p.data_dir
    subdirs = {
        "hero": ["characters", "hero"],
        "items": ["items", "objects"],
        "images": ["scenes", "images"],
        "videos": ["videos", "clips"],
        "audio": ["audio", "subs"],
        "final": ["final"],
        "text": [],
    }
    if kind in ("all", "text"):
        for name in ("voiceover.txt", "script.txt", "general_plan.txt"):
            fp = base / name
            if fp.exists():
                out.append(
                    {
                        "source": "file",
                        "id": name,
                        "kind": "text",
                        "path": str(fp),
                        "preview_url": f"/api/files?path={fp}",
                        "label": name,
                    }
                )
        plan = base / "project.xlsx"
        if plan.exists():
            out.append(
                {
                    "source": "file",
                    "id": "project.xlsx",
                    "kind": "xlsx",
                    "path": str(plan),
                    "preview_url": None,
                    "label": "Таблица проекта",
                }
            )

    scan_kind = kind if kind in subdirs else None
    if scan_kind or kind == "all":
        keys = [scan_kind] if scan_kind else list(subdirs.keys())
        for k in keys:
            for sub in subdirs.get(k, []):
                d = base / sub
                if not d.is_dir():
                    continue
                for fp in sorted(d.rglob("*")):
                    if not fp.is_file():
                        continue
                    if fp.suffix.lower() not in {".png", ".jpg", ".jpeg", ".webp", ".mp4", ".webm", ".wav", ".mp3"}:
                        continue
                    rel = _rel_path(str(fp))
                    out.append(
                        {
                            "source": "file",
                            "id": rel,
                            "kind": k,
                            "path": str(fp),
                            "preview_url": f"/api/files?path={fp}",
                            "label": fp.name,
                        }
                    )
    return out


@router.post("/{project_id}/assets/hero/replace")
async def replace_hero_image(
    project_id: int,
    file: UploadFile = File(...),
    replace_path: str | None = Query(None),
    session: AsyncSession = Depends(get_session),
) -> dict:
    """Заменить reference-картинку персонажа (файл в data/.../characters/)."""
    p = _project_or_404(await session.get(Project, project_id))
    if not file.filename:
        raise HTTPException(status_code=400, detail="empty filename")
    ext = Path(file.filename).suffix.lower()
    if ext not in {".png", ".jpg", ".jpeg", ".webp"}:
        raise HTTPException(status_code=400, detail="need image file (.png, .jpg, .webp)")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="empty file")

    chars_dir = p.data_dir / "characters"
    chars_dir.mkdir(parents=True, exist_ok=True)

    dest: Path
    if replace_path:
        candidate = Path(replace_path)
        if not candidate.is_absolute():
            candidate = Path(settings.data_dir) / replace_path
        try:
            candidate.resolve().relative_to(p.data_dir.resolve())
        except ValueError as e:
            raise HTTPException(status_code=400, detail="replace_path outside project") from e
        dest = candidate
        dest.parent.mkdir(parents=True, exist_ok=True)
    else:
        stem = Path(file.filename).stem or "hero"
        dest = chars_dir / f"{stem}{ext}"

    dest.write_bytes(content)
    await publish_project_event(
        project_id,
        event_type="project_updated",
        payload={"hero_replaced": str(dest.name)},
    )
    rel = _rel_path(str(dest))
    return {
        "path": str(dest),
        "preview_url": f"/api/files?path={dest}",
        "id": rel,
    }


def _rel_path(path: str | None) -> str:
    if not path:
        return ""
    try:
        p = Path(path)
        data = Path(settings.data_dir).resolve()
        return str(p.resolve().relative_to(data))
    except Exception:
        return path


@router.patch("/{project_id}/excel-gpt/{node_key}")
async def patch_excel_gpt_config(
    project_id: int,
    node_key: str,
    payload: dict,
    session: AsyncSession = Depends(get_session),
) -> dict:
    from sqlalchemy.orm.attributes import flag_modified

    p = _project_or_404(await session.get(Project, project_id))
    meta = dict(p.meta or {})
    configs = dict(meta.get("excel_gpt_nodes") or {})
    cur = dict(configs.get(node_key) or {})
    for key in ("label", "inputSource", "uploadedFileName", "slotIndex"):
        if key in payload:
            cur[key] = payload[key]
    configs[node_key] = cur
    meta["excel_gpt_nodes"] = configs
    meta["active_excel_gpt_node_key"] = node_key
    p.meta = meta
    flag_modified(p, "meta")
    await session.commit()
    return {"ok": True, "config": cur}


@router.post("/{project_id}/excel-gpt/{node_key}/upload")
async def upload_excel_gpt_file(
    project_id: int,
    node_key: str,
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_session),
) -> dict:
    from sqlalchemy.orm.attributes import flag_modified

    from app.services.excel_gpt_node import upload_dir, upload_file_path

    p = _project_or_404(await session.get(Project, project_id))
    if not file.filename:
        raise HTTPException(status_code=400, detail="need filename")
    safe_name = Path(file.filename).name
    dest_dir = upload_dir(p, node_key)
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest = upload_file_path(p, node_key, safe_name)
    content = await file.read()
    dest.write_bytes(content)
    meta = dict(p.meta or {})
    configs = dict(meta.get("excel_gpt_nodes") or {})
    cur = dict(configs.get(node_key) or {})
    cur["inputSource"] = "upload"
    cur["uploadedFileName"] = safe_name
    cur["label"] = safe_name
    configs[node_key] = cur
    meta["excel_gpt_nodes"] = configs
    meta["active_excel_gpt_node_key"] = node_key
    p.meta = meta
    flag_modified(p, "meta")
    await session.commit()
    return {"ok": True, "fileName": safe_name, "path": str(dest)}


@router.post("/{project_id}/excel-gpt/remap-keys")
async def remap_excel_gpt_keys(
    project_id: int,
    payload: dict,
    session: AsyncSession = Depends(get_session),
) -> dict:
    from sqlalchemy.orm.attributes import flag_modified

    from app.services.excel_gpt_node import remap_node_keys_in_meta

    p = _project_or_404(await session.get(Project, project_id))
    raw = payload.get("mapping")
    if not isinstance(raw, dict):
        raise HTTPException(status_code=400, detail="mapping dict required")
    mapping = {str(k): str(v) for k, v in raw.items() if k and v}
    remapped = remap_node_keys_in_meta(p, mapping)
    flag_modified(p, "meta")
    await session.commit()
    return {"ok": True, "remapped": remapped}


@router.post("/parents/disable-auto-mode")
async def disable_auto_mode_all_parents(
    session: AsyncSession = Depends(get_session),
) -> dict:
    """Выключить автопродвижение у всех родительских проектов (воркер перестанет сам жать шаги)."""
    from app.services.mass_factory import mass_parent_id

    rows = (await session.execute(select(Project))).scalars().all()
    disabled = 0
    for p in rows:
        if mass_parent_id(p) is not None:
            continue
        if p.auto_mode:
            p.auto_mode = False
            disabled += 1
            await publish_project_event(
                p.id,
                event_type="project_updated",
                payload={"auto_mode": False},
            )
    await session.commit()
    return {"parents_total": sum(1 for p in rows if mass_parent_id(p) is None), "disabled": disabled}


@router.post("/{project_id}/remount-video")
async def remount_project_video(
    project_id: int,
    audio_only: bool = Query(False),
    session: AsyncSession = Depends(get_session),
) -> dict:
    """Перемонтаж: синхрон xlsx→кадры, Whisper по озвучке, новая сборка (видеоклипы не трогаем)."""
    from app.services.remount_video import remount_video

    p = _project_or_404(await session.get(Project, project_id))
    result = await remount_video(session, p, run_assemble=not audio_only)
    await session.commit()
    await publish_project_event(
        project_id,
        event_type="project_updated",
        payload={
            "remount_video": True,
            "done": result.get("done"),
            "error": result.get("error"),
            "final_video": result.get("final_video"),
        },
    )
    if result.get("error") and not result.get("done"):
        raise HTTPException(status_code=400, detail=result)
    return result


@router.post("/restore-original-voiceover")
async def restore_all_parents_voiceover(
    dry_run: bool = Query(False),
    force: bool = Query(False),
    session: AsyncSession = Depends(get_session),
) -> dict:
    """Восстановить исходный voiceover у всех родительских проектов."""
    from app.services.voiceover_recovery import restore_all_parent_voiceovers

    summary = await restore_all_parent_voiceovers(
        session, dry_run=dry_run, force=force
    )
    if summary.get("restored"):
        for row in summary.get("results", []):
            if row.get("restored"):
                await publish_project_event(
                    int(row["project_id"]),
                    event_type="project_updated",
                    payload={"voiceover_restored": True, "source": row.get("source")},
                )
    return summary


@router.post("/{project_id}/restore-original-voiceover")
async def restore_project_voiceover(
    project_id: int,
    dry_run: bool = Query(False),
    force: bool = Query(False),
    session: AsyncSession = Depends(get_session),
) -> dict:
    """Восстановить исходный voiceover одного проекта."""
    from app.services.voiceover_recovery import restore_original_voiceover

    from app.services.mass_factory import mass_parent_id
    from app.services.voiceover_recovery import is_parent_project

    p = _project_or_404(await session.get(Project, project_id))
    if not is_parent_project(p):
        raise HTTPException(
            status_code=400,
            detail={
                "reason": "child_project_skipped",
                "mass_parent_id": mass_parent_id(p),
                "hint": "восстановление только для родительских проектов",
            },
        )
    result = await restore_original_voiceover(
        session, p, dry_run=dry_run, force=force
    )
    if result.get("restored"):
        await session.commit()
        await publish_project_event(
            project_id,
            event_type="project_updated",
            payload={"voiceover_restored": True, "source": result.get("source")},
        )
    return result


@router.get("/{project_id}/original-voiceover-preview")
async def preview_original_voiceover(
    project_id: int,
    session: AsyncSession = Depends(get_session),
) -> dict:
    """Показать, откуда будет взят исходный voiceover (без записи)."""
    from app.services.voiceover_recovery import find_original_voiceover

    from app.services.voiceover_recovery import is_parent_project

    p = _project_or_404(await session.get(Project, project_id))
    if not is_parent_project(p):
        raise HTTPException(
            status_code=400,
            detail="preview только для родительских проектов",
        )
    cand = await find_original_voiceover(session, p)
    if cand is None:
        return {
            "project_id": project_id,
            "found": False,
            "preview": None,
            "source": None,
            "chars": 0,
        }
    return {
        "project_id": project_id,
        "found": True,
        "source": cand.source,
        "chars": len(cand.text),
        "preview": cand.text[:500],
    }

