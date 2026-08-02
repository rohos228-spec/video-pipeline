"""REST для визуализации DB v2 (кнопка «База» в Studio).

Чтение графа + настройки: правка кадра, вставка между кадрами, тексты,
версии промтов, сущности, связи. Всё по id/uuid карточек — без адресов
вида «строка 48».
"""

from __future__ import annotations

import asyncio

from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel, Field
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import (
    Entity,
    Frame,
    FrameEdge,
    FrameStatus,
    FrameText,
    Project,
    PromptVersion,
    Scene,
)
from app.orchestrator.node_registry import STEP_CODE_TO_NODE_TYPE
from app.services import db_apply, db_v2
from app.services.xlsx_v8_import import (
    ROW_DURATION_V8,
    ROW_IMAGE_PROMPT_2_V8,
    ROW_IMAGE_PROMPT_V8,
    ROW_TIMECODE_V8,
    ROW_VIDEO_PROMPT_2_V8,
    ROW_VIDEO_PROMPT_V8,
    ROW_VOICEOVER_V8,
    _cell_text,
    _resolve_plan_sheet,
)
from app.web.deps import get_session

router = APIRouter(prefix="/db", tags=["db"])

# Лист «план» v8: строки, которые оператор правит в Excel и на которые
# ссылаются промты. Показываем их в «Базе» рядом с карточкой кадра.
_ROWS_PERSONS = (8, 23, 38)
_ROWS_ITEMS = (9, 24, 39)


async def _project(session: AsyncSession, project_id: int) -> Project:
    p = await session.get(Project, project_id)
    if p is None:
        raise HTTPException(404, f"проект {project_id} не найден")
    return p


async def _frame(session: AsyncSession, frame_id: int) -> Frame:
    fr = await session.get(Frame, frame_id)
    if fr is None:
        raise HTTPException(404, f"кадр {frame_id} не найден")
    return fr


def _merged_plan_ids(ws, col: int, rows: tuple[int, ...]) -> str | None:
    seen: list[str] = []
    for row in rows:
        text = _cell_text(ws, row, col)
        if not text:
            continue
        for part in text.replace(";", ",").split(","):
            item = part.strip()
            if item and item not in seen:
                seen.append(item)
    return ", ".join(seen) if seen else None


def _excel_rows_for_project(project: Project) -> dict[int, dict[str, str | int | None]]:
    """Строки листа «план» project.xlsx по номеру кадра (для UI «Базы»).

    Excel пока остаётся источником правды для промтов — карточка кадра
    обязана показывать те же R45/R48/R49, что видел оператор в книге.
    """
    path = project.data_dir / "project.xlsx"
    if not path.is_file():
        return {}
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {}
    try:
        wb = load_workbook(path, data_only=True)
    except Exception:  # noqa: BLE001
        return {}
    try:
        ws = _resolve_plan_sheet(wb)
        if ws is None:
            return {}
        out: dict[int, dict[str, str | int | None]] = {}
        max_col = ws.max_column or 0
        for col in range(3, max_col + 1):
            frame_no = col - 2
            values = {
                "column": col,
                "r15_timecode": _cell_text(ws, ROW_TIMECODE_V8, col),
                "r45_image_prompt": _cell_text(ws, ROW_IMAGE_PROMPT_V8, col),
                "r46_image_prompt_2": _cell_text(ws, ROW_IMAGE_PROMPT_2_V8, col),
                "r48_video_prompt": _cell_text(ws, ROW_VIDEO_PROMPT_V8, col),
                "r64_video_prompt_2": _cell_text(ws, ROW_VIDEO_PROMPT_2_V8, col),
                "r49_voiceover": _cell_text(ws, ROW_VOICEOVER_V8, col),
                "r50_duration": _cell_text(ws, ROW_DURATION_V8, col),
                "persons": _merged_plan_ids(ws, col, _ROWS_PERSONS),
                "items": _merged_plan_ids(ws, col, _ROWS_ITEMS),
            }
            if any(v not in (None, "") for k, v in values.items() if k != "column"):
                out[frame_no] = values
        return out
    finally:
        wb.close()


@router.get("/overview")
async def db_overview(session: AsyncSession = Depends(get_session)) -> dict:
    projects = list((await session.execute(select(Project).order_by(Project.id))).scalars())
    out = []
    for p in projects:
        frames_n = (
            await session.execute(
                select(func.count(Frame.id)).where(Frame.project_id == p.id)
            )
        ).scalar_one()
        scenes_n = (
            await session.execute(
                select(func.count(Scene.id)).where(Scene.project_id == p.id)
            )
        ).scalar_one()
        entities_n = (
            await session.execute(
                select(func.count(Entity.id)).where(Entity.project_id == p.id)
            )
        ).scalar_one()
        edges_n = (
            await session.execute(
                select(func.count(FrameEdge.id)).where(FrameEdge.project_id == p.id)
            )
        ).scalar_one()
        out.append(
            {
                "id": p.id,
                "slug": p.slug,
                "title": p.title,
                "topic": p.topic,
                "status": p.status.value if p.status else None,
                "frames": frames_n,
                "scenes": scenes_n,
                "entities": entities_n,
                "edges": edges_n,
            }
        )
    return {"projects": out}


@router.get("/projects/{project_id}/graph")
async def db_graph(
    project_id: int, session: AsyncSession = Depends(get_session)
) -> dict:
    project = await _project(session, project_id)
    await db_v2.backfill_project_v2(session, project)
    await session.commit()
    graph = await db_v2.project_graph(session, project)
    graph["excel_rows"] = _excel_rows_for_project(project)
    # Сводка последних проверок — чип «Проверки» в «Базе».
    from app.services.agent_harness import read_ops_telemetry

    tel = read_ops_telemetry(project.meta if isinstance(project.meta, dict) else {})
    checks = tel.get("checks") or []
    graph["harness"] = {
        "outcome": tel.get("last_outcome"),
        "updated_at": tel.get("updated_at"),
        "total": len(checks),
        "failed": [c.get("name") for c in checks if not c.get("ok")],
        "next_action": tel.get("next_action"),
    }
    return graph


@router.post("/projects/{project_id}/export-xlsx")
async def export_xlsx(
    project_id: int, session: AsyncSession = Depends(get_session)
) -> dict:
    """Кнопка «Экспорт в Excel»: переписать строки листа «план» из DB."""
    project = await _project(session, project_id)
    frames = list(
        (
            await session.execute(
                select(Frame).where(Frame.project_id == project.id).order_by(Frame.number)
            )
        ).scalars()
    )
    try:
        return db_apply.export_project_xlsx(project, frames)
    except db_apply.ApplyOpsError as e:
        raise HTTPException(400, str(e)) from None


class ApplyOp(BaseModel):
    target: str = Field(default="frame", max_length=16)
    frame_uuid: str | None = Field(default=None, max_length=64)
    fields: dict = Field(default_factory=dict)


class ApplyOpsBody(BaseModel):
    ops: list[ApplyOp]
    export_xlsx: bool = True


@router.post("/projects/{project_id}/apply-ops")
async def apply_ops(
    project_id: int,
    body: ApplyOpsBody,
    session: AsyncSession = Depends(get_session),
) -> dict:
    """Fail-closed запись от GPT: JSON-операции (логика в `services/db_apply`).

    ``target="frame"`` (по умолчанию) — правка кадра по ``frame_uuid``;
    ``target="project"`` — поля уровня проекта (общий план). Поля можно
    писать по-человечески («закадр», «промт_картинки») — сервер сам
    переводит в канонические ключи. Любая ошибка — 400, откат транзакции.
    """
    project = await _project(session, project_id)
    ops = [
        {"target": op.target, "frame_uuid": op.frame_uuid, "fields": op.fields}
        for op in body.ops
    ]
    try:
        result = await db_apply.apply_ops(
            session, project, ops, export_xlsx=body.export_xlsx
        )
    except db_apply.ApplyOpsError as e:
        await session.rollback()
        raise HTTPException(400, str(e)) from None
    await session.commit()
    return result


_ORCHESTRATOR_SYSTEM = (
    "Ты — оркестратор видеопроекта. Источник правды — база (НЕ Excel).\n"
    "Тебе дан граф проекта (frames[] с uuid и полями), состояние шагов "
    "пайплайна, РЕЗУЛЬТАТЫ ПРОВЕРОК (harness: «ок»/«НЕ ОК» по каждой "
    "проверке + итог) и СТАТУСЫ НОД последнего прогона (включая hitl-ноды; "
    "«ждёт аппрува человека» = HITL-гейт ждёт решения пользователя).\n"
    "На вопросы про проверки/проверку/статусы нод отвечай СТРОГО по этим "
    "разделам контекста — если прогона проверок не было, скажи это прямо.\n"
    "ЗАПРЕТ ОТКАЗОВ: не пиши «не поддерживается / нет такого действия / "
    "не умею», если действие есть в списке ниже. Сначала выдай JSON action, "
    "потом коротко по-русски. Если не хватает файла — read_file, не отказ.\n"
    "Если есть раздел ДИАГНОСТИКА / КОД РАБОЧЕЙ ОБЛАСТИ или пользователь "
    "просит починить/исправить баг в КОДЕ программы — отвечай как кодинг-агент: "
    "read_file → причина → edit_files → run_tests → git_commit_push "
    "(действия 16–20). Без воды. "
    "Правь только app/, tests/, prompts/. Не трогай .env и data/.\n"
    "ПОРЯДОК ШАГОВ — ТОЛЬКО ТАКОЙ (другого не существует, не выдумывай):\n"
    "plan → script → split → hero → items → enrich_1..5 → img_pr → img → "
    "anim_pr → video → audio → music → assemble → publish.\n"
    "Отвечая про «следующий шаг», называй шаги из этого списка и опирайся "
    "на состояние шагов из контекста.\n"
    "ДЕЙСТВИЯ (можно совместить в одном JSON: {\"ops\":[...], \"actions\":[...]}):\n"
    "1) ИЗМЕНИТЬ данные кадра/проекта → "
    '{"ops":[{"frame_uuid":"<uuid из графа>","fields":{...}}]}. '
    "Поля по-человечески: закадр, промт_картинки, промт_видео, смысл, "
    "длительность, промт_картинки_2, промт_видео_2; общий план: "
    '{"target":"project","fields":{"общий_план":"…"}}.\n'
    "2) ЗАПУСТИТЬ шаг → {\"actions\":[{\"run_step\":\"<код>\"}]} "
    "(img_pr, img, anim_pr, video, audio, music, assemble, publish…).\n"
    "3) ОСТАНОВИТЬ генерацию → {\"actions\":[{\"stop_step\":true}]}.\n"
    "4) НАСТРОЙКА генерации → "
    "{\"actions\":[{\"set_option\":{\"key\":\"<ключ>\",\"value\":\"<id>\"}}]} — "
    "ключи и допустимые id в разделе НАСТРОЙКИ контекста "
    "(image_generator, aspect_ratio, image_resolution, image_quality, "
    "video_generator, video_resolution, hero_mode, auto_mode).\n"
    "5) ВАРИАНТ промта шага → "
    "{\"actions\":[{\"set_prompt\":{\"step\":\"<шаг>\",\"variant\":\"<имя>\"}}]} — "
    "варианты в разделе ПРОМТЫ контекста.\n"
    "6) ТЕКСТОВАЯ LLM → "
    "{\"actions\":[{\"set_text_llm\":{\"provider\":\"kie|tokenrouter\"}}]}.\n"
    "7) ОТКРЫТЬ окна программы → {\"actions\":[{\"open_ui\":{...}}]} — kinds: "
    "step_prompts (плюс step; ОБЯЗАТЕЛЬНО при выборе/сравнении вариантов "
    "промтов — человек выбирает в окне), node_studio/prompt_builder/hitl "
    "(плюс node_type), topic (тема), baza, gpt_chat, create (Create/Outsee), "
    "fleet, settings. "
    "ТОЧНЕЕ — с node_key из раздела НОДЫ КАНВАСА: "
    "{\"open_ui\":{\"kind\":\"node_studio\",\"node_key\":\"n_plan\"}}.\n"
    "НАЗВАНИЯ НОД: маппи русские названия пользователя на node_key по "
    "разделу НОДЫ КАНВАСА (название в «»). ВНИМАНИЕ: «Сценарий» — это нода "
    "plan (n_plan); «Закадровый текст» — script (n_script); «Работа с GPT» — "
    "ноды excel_gpt (n_excel_gpt_1, n_excel_gpt_2…).\n"
    "8) HITL-РЕШЕНИЕ (когда нода ждёт аппрува) → "
    "{\"actions\":[{\"hitl_decision\":\"approve|regenerate|reject|edit_prompt\"}]}.\n"
    "9) ТЕМА проекта → {\"actions\":[{\"set_topic\":\"…\"}]}.\n"
    "ВАЖНО: ноды topic и storage — конфиг, у них НЕТ промтов и студии. "
    "Для темы используй set_topic или open_ui topic; не открывай для них "
    "step_prompts/node_studio/prompt_builder.\n"
    "10) СОЗДАТЬ проект → "
    "{\"actions\":[{\"create_project\":{\"title\":\"…\",\"hero_mode\":\"auto\"}}]} — "
    "hero_mode ∈ hero|no_hero|auto (по умолчанию auto). "
    "Создаёт проект и открывает его (open_project).\n"
    "10b) ДОЧЕРНИЙ проект из родителя → "
    "{\"actions\":[{\"create_child\":{\"parent_id\":<id>}}]} или "
    "{\"create_child\":{\"parent_title\":\"История ведьм\"}} или "
    "{\"create_child\":true} (родитель = текущий проект). "
    "Копирует настройки генерации и промты (prompt_overrides / gpt_text); "
    "НЕ копирует кадры, Excel-данные, PNG/MP4. id/title бери из раздела "
    "ПРОЕКТЫ В СТУДИИ. После создания открой ребёнка (ui open_project).\n"
    "11) ДОБАВИТЬ ноду в граф → "
    "{\"actions\":[{\"add_node\":{\"node_type\":\"hitl_gate\",\"after\":\"each\"}}]} — "
    "after=\"each\" (после каждой рабочей ноды, без дублей) или "
    "after=\"<node_key>\". «Нода проверки»: hitl_gate = аппрув человеком, "
    "excel_gpt = автопроверка GPT — уточни у пользователя, если неясно. "
    "В ответе называй, что именно добавил и после каких нод (см. СВЯЗИ).\n"
    "12) УДАЛИТЬ ноды → {\"actions\":[{\"remove_node\":{\"node_type\":\"X\"}}]} "
    "(все добавленные оркестратором типа X — ТОЛЬКО если прямо просят «удали "
    "ВСЕ»), {\"remove_node\":{\"node_type\":\"X\",\"only\":\"duplicates\"}} — "
    "ТОЛЬКО дубли (две проверки подряд): «лишние удали» = всегда duplicates, "
    "или {\"remove_node\":{\"node_key\":\"n_...\"}} (конкретную). Удаление НЕ "
    "выполняется сразу: человек подтверждает кнопкой в чате.\n"
    "ПЕРЕИМЕНОВАТЬ ноду → {\"actions\":[{\"rename_node\":{\"node_key\":\"n_...\",\"label\":\"…\"}}]}. "
    "Новые проверки подписываются по родителю автоматически («Проверка — Сценарий»).\n"
    "13) ПОЧИНИТЬ граф (после удалений/разрывов) → "
    "{\"actions\":[{\"repair_graph\":true}]} — цепочка пересобирается слева "
    "направо (проверки встают за своими родителями). ПОРЯДОК ВАЖЕН: если "
    "граф сломан — repair_graph идёт в actions ПЕРВЫМ, до add_node.\n"
    "14) ПРОГНАТЬ проверки (harness) → {\"actions\":[{\"run_harness\":true}]}.\n"
    "15) Вопрос/обсуждение без изменений — обычный текст.\n"
    "16) ПРОЧИТАТЬ КОД (allowlist app/ tests/ prompts/) → "
    "{\"actions\":[{\"read_file\":{\"path\":\"app/...py\","
    "\"start_line\":1,\"end_line\":120}}]} — "
    "сначала читай, потом правь. Содержимое вернётся в ответе Studio.\n"
    "17) ПРАВКА КОДА → "
    "{\"actions\":[{\"edit_files\":[{\"path\":\"app/...py\","
    "\"old_string\":\"…\",\"new_string\":\"…\"}]}]} — "
    "old_string должен встречаться ровно 1 раз. "
    "Новый мелкий файл: {\"path\":\"tests/....py\",\"content\":\"...\"}.\n"
    "18) ПРОГНАТЬ pytest → "
    "{\"actions\":[{\"run_tests\":[\"tests/test_....py\"]}]}. "
    "После edit_files — обязательно, перед push. FAIL не отменяет другие "
    "действия — смотри вывод и чини дальше.\n"
    "19) COMMIT+PUSH в origin/<ветка этого ПК> → "
    "{\"actions\":[{\"git_commit_push\":{\"message\":\"fix: …\","
    "\"files\":[\"app/....py\"],\"auto\":false}}]} — "
    "ветка = из раздела КОД РАБОЧЕЙ ОБЛАСТИ / ORCHESTRATOR_GIT_BRANCH. "
    "по умолчанию auto=false: человек жмёт «Подтвердить push» в чате. "
    "auto=true — сразу push (только если пользователь явно просит "
    "«сразу запушь» / «без подтверждения»). Без force-push.\n"
    "20) УДАЛИТЬ ПРОЕКТЫ → "
    "{\"actions\":[{\"delete_projects\":{\"ids\":[52,50]}}]} или "
    "{\"delete_projects\":{\"titles\":[\"тест оркестра\",\"Тестовый трукрайм\"]}}. "
    "id/title бери из раздела ПРОЕКТЫ В СТУДИИ. "
    "НИКОГДА не пиши «удаление не поддерживается». "
    "Удаление НЕ сразу: человек жмёт «Подтвердить удаление проектов» в чате.\n"
    "ТИПЫ НОД по-человечески (объясняй так, если спрашивают «что это»): "
    "hitl_gate — нода проверки: пайплайн встаёт и ждёт аппрува человека; "
    "excel_gpt — работа/автопроверка через GPT; plan — «Сценарий»; "
    "script — «Закадровый текст»; split — «Разбивка»; hero — «Персонажи»; "
    "items — «Предметы»; image_prompts — «Промты картинок»; "
    "images — «Картинки»; animation_prompts — «Промты анимации»; "
    "videos — «Видео»; audio — «Озвучка»; music — «Музыка»; "
    "assemble — «Сборка»; publish — «Публикация»; topic — «Тема»; "
    "storage — хранилище.\n"
    "Не выдумывай uuid, поля, коды шагов, ключи и значения настроек — "
    "неизвестное = отклонено с ошибкой. Одна операция = один кадр."
)

_STEP_RU: dict[str, str] = {
    "plan": "план",
    "script": "сценарий",
    "split": "разбивка на кадры",
    "hero": "персонажи (референсы)",
    "items": "предметы",
    "enrich_1": "обогащение 1",
    "enrich_2": "обогащение 2",
    "enrich_3": "обогащение 3",
    "enrich_4": "обогащение 4",
    "enrich_5": "обогащение 5",
    "img_pr": "промты картинок",
    "img": "генерация картинок",
    "anim_pr": "промты видео",
    "video": "генерация видео",
    "audio": "озвучка",
    "music": "музыка",
    "assemble": "монтаж",
    "publish": "публикация",
}


def _pipeline_state_lines(project: Project) -> list[str]:
    """Состояние шагов по реальному порядку пайплайна (для контекста модели)."""
    from app.orchestrator.node_registry import (
        LINEAR_RUNNING_PIPELINE,
        NODE_TYPE_TO_READY,
        NODE_TYPE_TO_RUNNING,
        NODE_TYPE_TO_STEP_CODE,
    )

    ordered: list[str] = []
    for _run, node_type in LINEAR_RUNNING_PIPELINE:
        ordered.append(NODE_TYPE_TO_RUNNING[node_type].value)
        ordered.append(NODE_TYPE_TO_READY[node_type].value)
    current = project.status.value if project.status else ""
    pos = ordered.index(current) if current in ordered else 0

    lines = []
    for i, (_run, node_type) in enumerate(LINEAR_RUNNING_PIPELINE):
        ready_idx = 2 * i + 1
        if ready_idx < pos:
            state = "готово"
        elif 2 * i <= pos:
            state = "текущий"
        else:
            state = "ждёт"
        step_code = NODE_TYPE_TO_STEP_CODE.get(node_type, node_type)
        lines.append(f"- {step_code} ({_STEP_RU.get(step_code, step_code)}): {state}")
    return lines


async def _checks_context(session: AsyncSession, project: Project) -> list[str]:
    """Результаты проверок (harness-telemetry) + статусы нод последнего прогона."""
    from app.models import NodeRun, WorkflowRun
    from app.services.agent_harness import read_ops_telemetry

    lines: list[str] = []
    tel = read_ops_telemetry(project.meta if isinstance(project.meta, dict) else {})
    checks = tel.get("checks") or []
    if checks:
        outcome = "ОК" if tel.get("last_outcome") == "verify_pass" else "НЕ ОК"
        lines.append(
            f"ПРОВЕРКИ HARNESS (последняя {tel.get('updated_at', '?')}, итог: {outcome}):"
        )
        for c in checks:
            mark = "ок" if c.get("ok") else "НЕ ОК"
            detail = str(c.get("detail") or "")[:100]
            suffix = f" — {detail}" if detail and not c.get("ok") else ""
            lines.append(f"- {c.get('name')}: {mark}{suffix}")
        if tel.get("next_action") and tel["next_action"] != "none":
            lines.append(f"  next_action: {tel['next_action']}")
    else:
        lines.append(
            "ПРОВЕРКИ HARNESS: прогонов не было (свежие данные — через harness/verify)."
        )

    run = (
        await session.execute(
            select(WorkflowRun)
            .where(WorkflowRun.project_id == project.id)
            .order_by(WorkflowRun.id.desc())
        )
    ).scalars().first()
    if run is None:
        lines.append("СТАТУСЫ НОД: прогонов workflow ещё не было.")
        return lines
    nodes = list(
        (
            await session.execute(
                select(NodeRun).where(NodeRun.workflow_run_id == run.id).order_by(NodeRun.id)
            )
        ).scalars()
    )
    if nodes:
        ru = {
            "pending": "ждёт",
            "queued": "в очереди",
            "running": "выполняется",
            "waiting_hitl": "ждёт аппрува человека",
            "done": "готово",
            "failed": "ОШИБКА",
            "skipped": "пропущена",
        }
        lines.append(f"СТАТУСЫ НОД (прогон #{run.id}, статус прогона {run.status.value}):")
        for n in nodes:
            st = ru.get(n.status.value, n.status.value)
            extra = ""
            if n.status.value == "failed" and n.error:
                extra = f" — {str(n.error)[:120]}"
            elif n.progress_text:
                extra = f" ({n.progress_text})"
            lines.append(f"- {n.node_type}: {st}{extra}")
    return lines


def _log_tail_context(project: Project, max_lines: int = 25) -> list[str]:
    """Хвост ошибок/варнингов проекта из свежего backend-лога."""
    from app.settings import settings

    try:
        logs = sorted(
            settings.data_dir.glob("backend*.log"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
    except Exception:  # noqa: BLE001
        return []
    if not logs:
        return []
    marker = f"[#{project.id}]"
    out: list[str] = []
    try:
        for ln in logs[0].read_text(encoding="utf-8", errors="replace").splitlines():
            if marker in ln and ("ERROR" in ln or "Traceback" in ln or "WARNING" in ln):
                out.append(ln.strip()[:200])
    except Exception:  # noqa: BLE001
        return []
    return out[-max_lines:]


async def _diagnostics_context(session: AsyncSession, project: Project) -> list[str]:
    """Ошибки нод + проваленные проверки + хвост лога (для режима кодинг-агента)."""
    from app.models import NodeRun, NodeRunStatus, WorkflowRun
    from app.services.agent_harness import read_ops_telemetry

    lines: list[str] = []
    run = (
        await session.execute(
            select(WorkflowRun)
            .where(WorkflowRun.project_id == project.id)
            .order_by(WorkflowRun.id.desc())
        )
    ).scalars().first()
    if run is not None:
        failed = list(
            (
                await session.execute(
                    select(NodeRun).where(
                        NodeRun.workflow_run_id == run.id,
                        NodeRun.status == NodeRunStatus.failed,
                    )
                )
            ).scalars()
        )
        if failed:
            lines.append("ОШИБКИ НОД (последний прогон):")
            for n in failed:
                lines.append(f"- {n.node_type}: {str(n.error or '')[:200]}")
    tel = read_ops_telemetry(project.meta if isinstance(project.meta, dict) else {})
    bad_checks = [c for c in (tel.get("checks") or []) if not c.get("ok")]
    if bad_checks:
        lines.append("ПРОВАЛЕННЫЕ ПРОВЕРКИ:")
        for c in bad_checks:
            lines.append(f"- {c.get('name')}: {str(c.get('detail') or '')[:120]}")
    tail = _log_tail_context(project)
    if tail:
        lines.append("ЛОГ BACKEND (хвост ошибок проекта):")
        lines.extend(f"  {ln}" for ln in tail)
    if lines:
        return ["ДИАГНОСТИКА:", *lines]
    return []


def _code_workspace_context() -> list[str]:
    """Ветка git + грязные allowlist-файлы — чтобы оркестратор не гадал."""
    lines = ["КОД РАБОЧЕЙ ОБЛАСТИ (для read_file / edit_files / git):"]
    try:
        from app.services.code_autofix import ALLOWED_PREFIXES
        from app.services.git_ops import changed_paths, current_branch, push_branch

        want = push_branch()
        br = current_branch()
        lines.append(f"- push-ветка (ORCHESTRATOR_GIT_BRANCH): {want}")
        lines.append(f"- текущий HEAD: {br}" + (" ✓" if br == want else " ← СНАЧАЛА checkout нужной ветки"))
        lines.append(f"- allowlist путей: {', '.join(ALLOWED_PREFIXES)}")
        dirty = [
            p
            for p in changed_paths()
            if any(p.startswith(pref) for pref in ALLOWED_PREFIXES)
        ][:40]
        if dirty:
            lines.append("- изменённые файлы (allowlist):")
            lines.extend(f"  · {p}" for p in dirty)
        else:
            lines.append("- изменённых allowlist-файлов нет")
    except Exception as e:  # noqa: BLE001
        lines.append(f"- git недоступен: {e}")
    return lines


_NODE_STATUS_RU = {
    "pending": "ждёт",
    "queued": "в очереди",
    "running": "выполняется",
    "waiting_hitl": "ждёт аппрува человека",
    "done": "готово",
    "failed": "ОШИБКА",
    "skipped": "пропущена",
}


async def _canvas_nodes_context(
    session: AsyncSession, project: Project
) -> tuple[list[str], dict[str, dict]]:
    """Ноды канваса проекта: ключ → название (data.label) + тип + статус.

    Возвращает (строки для контекста, keymap node_key → {type, label}).
    Названия — те, что пользователь видит на канвасе («Сценарий» = plan!).
    """
    from app.models import NodeRun, Workflow, WorkflowRun

    run = (
        await session.execute(
            select(WorkflowRun)
            .where(WorkflowRun.project_id == project.id)
            .order_by(WorkflowRun.id.desc())
        )
    ).scalars().first()
    nodes: list[dict] = []
    edges_src: list[dict] = []
    status_by_key: dict[str, str] = {}
    if run is not None:
        nodes = list(run.nodes_snapshot or [])
        edges_src = list(run.edges_snapshot or [])
        for nr in (
            await session.execute(
                select(NodeRun).where(NodeRun.workflow_run_id == run.id)
            )
        ).scalars():
            status_by_key[nr.node_key] = nr.status.value
    if not nodes:
        wf = (
            await session.execute(
                select(Workflow).where(Workflow.is_default.is_(True))
            )
        ).scalars().first()
        nodes = list((wf.nodes if wf else []) or [])
        edges_src = list((wf.edges if wf else []) or [])

    lines: list[str] = []
    keymap: dict[str, dict] = {}
    for n in nodes:
        key = str(n.get("id") or "")
        if not key:
            continue
        typ = str(n.get("type") or "")
        label = str((n.get("data") or {}).get("label") or typ)
        keymap[key] = {"type": typ, "label": label}
        st = status_by_key.get(key)
        suffix = f" — {_NODE_STATUS_RU.get(st, st)}" if st else ""
        lines.append(f"- {key} → «{label}» (тип {typ}){suffix}")
    if not lines:
        return [], keymap
    # Связи графа — чтобы оркестратор видел, что откуда следует.
    chain = ""
    try:
        order = _linear_order(nodes, edges_src)
        chain = " → ".join(order)
    except Exception:  # noqa: BLE001
        pairs = [f"{e.get('source')}→{e.get('target')}" for e in edges_src]
        chain = ", ".join(pairs)
    if chain:
        lines.append(f"СВЯЗИ: {chain}")
    return ["НОДЫ КАНВАСА (ключ → «название» (тип) — статус):", *lines], keymap


def _settings_context(project: Project) -> list[str]:
    """Настройки генерации, варианты промтов, текстовая LLM (для контекста)."""
    from app import generation_options as go
    from app.services import prompt_library as pl
    from app.services.text_llm_catalog import catalog_status

    def ids(cat) -> str:
        return ", ".join(c.id for c in cat)

    overrides = project.prompt_overrides or {}
    lines = [
        "НАСТРОЙКИ ГЕНЕРАЦИИ (текущее → допустимые id):",
        f"- image_generator: {project.image_generator or 'default'} → {ids(go.IMAGE_GENERATORS)}",
        f"- aspect_ratio: {project.aspect_ratio or 'default'} → {ids(go.ASPECT_RATIOS)}",
        f"- image_resolution: {project.image_resolution or 'default'} → {ids(go.IMAGE_RESOLUTIONS)} (зависит от генератора)",
        f"- image_quality: {project.image_quality or 'default'} → {ids(go.IMAGE_QUALITIES)}",
        f"- video_generator: {project.video_generator or 'default'} → {ids(go.VIDEO_GENERATORS)}",
        f"- video_resolution: {project.video_resolution or 'default'} → {ids(go.VIDEO_RESOLUTIONS)}",
        f"- hero_mode: {project.hero_mode} → hero, no_hero, auto",
        f"- auto_mode: {'вкл' if project.auto_mode else 'выкл'} → true/false",
        "ПРОМТЫ (шаг → выбранный вариант → доступные):",
    ]
    for step_code in ("plan", "script", "split", "img_pr", "anim_pr", "hero_style"):
        try:
            variants = pl.list_prompts(step_code)
        except Exception:  # noqa: BLE001
            variants = []
        if variants:
            lines.append(
                f"- {step_code}: {overrides.get(step_code, 'default')} → {', '.join(variants)}"
            )
    st = catalog_status()
    lines.append(
        f"ТЕКСТОВАЯ LLM: активна {st['active_label']} → провайдеры: kie, tokenrouter"
    )
    return lines


def _apply_set_option(project: Project, key: object, value: object) -> str:
    from app import generation_options as go

    k = str(key or "").strip()
    v = str(value or "").strip()
    catalogs = {
        "image_generator": go.IMAGE_GENERATORS_BY_ID,
        "aspect_ratio": go.ASPECT_RATIOS_BY_ID,
        "image_resolution": go.IMAGE_RESOLUTIONS_BY_ID,
        "image_quality": go.IMAGE_QUALITIES_BY_ID,
        "video_generator": go.VIDEO_GENERATORS_BY_ID,
        "video_resolution": go.VIDEO_RESOLUTIONS_BY_ID,
    }
    if k in catalogs:
        if v not in catalogs[k]:
            raise db_apply.ApplyOpsError(
                f"set_option {k}: неизвестное значение {v!r}; "
                f"допустимые: {sorted(catalogs[k])}"
            )
        if k == "image_resolution":
            v = go.clamp_image_resolution_id(project.image_generator, v)
        setattr(project, k, v)
        return f"{k}={v}"
    if k == "hero_mode":
        if v not in ("hero", "no_hero", "auto"):
            raise db_apply.ApplyOpsError("hero_mode: допустимые hero, no_hero, auto")
        project.hero_mode = v
        return f"hero_mode={v}"
    if k == "auto_mode":
        project.auto_mode = v.lower() in ("1", "true", "да", "вкл", "on")
        return f"auto_mode={'вкл' if project.auto_mode else 'выкл'}"
    raise db_apply.ApplyOpsError(
        f"set_option: неизвестный ключ {k!r}; "
        f"допустимые: {sorted([*catalogs, 'hero_mode', 'auto_mode'])}"
    )


def _apply_set_prompt(project: Project, step: object, variant: object) -> str:
    from app.services import prompt_library as pl

    s = str(step or "").strip()
    v = str(variant or "").strip()
    try:
        variants = pl.list_prompts(s)
    except Exception:  # noqa: BLE001
        variants = []
    if not variants:
        raise db_apply.ApplyOpsError(f"set_prompt: для шага {s!r} нет вариантов промтов")
    if v not in variants:
        raise db_apply.ApplyOpsError(
            f"set_prompt {s}: неизвестный вариант {v!r}; доступные: {variants}"
        )
    overrides = dict(project.prompt_overrides or {})
    overrides[s] = v
    project.prompt_overrides = overrides
    return f"{s}→{v}"


_UI_NODE_KINDS = ("node_studio", "prompt_builder", "hitl")
_UI_SIMPLE_KINDS = ("topic", "baza", "gpt_chat", "create", "fleet", "settings")


def _all_node_types() -> set[str]:
    from app.orchestrator.node_registry import (
        CONFIG_NODE_TYPES,
        HITL_NODE_TYPES,
        WORK_NODES,
    )

    return set(WORK_NODES) | set(HITL_NODE_TYPES) | set(CONFIG_NODE_TYPES) | {"excel_gpt"}


async def _pending_hitl_id(session: AsyncSession, project: Project) -> int | None:
    from app.models import HITLDecision, HITLRequest

    row = (
        await session.execute(
            select(HITLRequest)
            .where(
                HITLRequest.project_id == project.id,
                HITLRequest.decision == HITLDecision.pending,
            )
            .order_by(HITLRequest.id.desc())
        )
    ).scalars().first()
    return row.id if row else None


async def _resolve_open_ui(
    session: AsyncSession, project: Project, spec: dict, keymap: dict[str, dict]
) -> dict:
    """open_ui → ui_action для фронта (валидация fail-closed).

    Приоритет — ``node_key`` из раздела НОДЫ КАНВАСА (точное окно ноды);
    иначе — step/node_type как раньше.
    """
    from app.orchestrator.node_registry import (
        CONFIG_NODE_TYPES,
        NODE_TYPE_TO_STEP_CODE,
    )

    kind = str((spec or {}).get("kind") or "").strip()
    node_key = str((spec or {}).get("node_key") or "").strip() or None
    key_type = (keymap.get(node_key) or {}).get("type") if node_key else None
    if node_key and not key_type:
        raise db_apply.ApplyOpsError(
            f"open_ui: неизвестный node_key {node_key!r}; ключи — в разделе НОДЫ КАНВАСА"
        )

    if kind == "step_prompts":
        step = str(spec.get("step") or "").strip()
        if not step and key_type:
            step = NODE_TYPE_TO_STEP_CODE.get(key_type, "")
        if step not in STEP_CODE_TO_NODE_TYPE:
            raise db_apply.ApplyOpsError(f"open_ui step_prompts: неизвестный шаг {step!r}")
        out = {"kind": kind, "step": step, "node_type": STEP_CODE_TO_NODE_TYPE[step]}
        if node_key:
            out["node_key"] = node_key
        return out
    if kind in _UI_NODE_KINDS:
        node_type = str(spec.get("node_type") or "").strip() or (key_type or "")
        if node_type in CONFIG_NODE_TYPES:
            raise db_apply.ApplyOpsError(
                f"open_ui {kind}: у ноды {node_type!r} нет промтов/студии — "
                "это настройка. Тема меняется через set_topic "
                "(или open_ui topic — редактор в инспекторе)."
            )
        if node_type not in _all_node_types():
            raise db_apply.ApplyOpsError(
                f"open_ui {kind}: неизвестный тип ноды {node_type!r}"
            )
        item: dict = {"kind": kind, "node_type": node_type}
        if node_key:
            item["node_key"] = node_key
        if kind == "hitl":
            hitl_id = await _pending_hitl_id(session, project)
            if hitl_id is None:
                raise db_apply.ApplyOpsError("open_ui hitl: нет ожидающих HITL-запросов")
            item["hitl_id"] = hitl_id
        return item
    if kind in _UI_SIMPLE_KINDS:
        return {"kind": kind}
    raise db_apply.ApplyOpsError(
        f"open_ui: неизвестный kind {kind!r}; есть: step_prompts, "
        f"node_studio, prompt_builder, hitl, {', '.join(_UI_SIMPLE_KINDS)}"
    )


async def _apply_hitl_decision(
    session: AsyncSession, project: Project, decision: object
) -> dict:
    """HITL-решение из чата — та же логика, что кнопки Studio/TG."""
    from datetime import datetime

    from app.models import HITLDecision, HITLRequest
    from app.services.event_bus import publish_hitl_event
    from app.services.hitl_apply import apply_hitl_side_effects

    mapping = {
        "approve": HITLDecision.approved,
        "regenerate": HITLDecision.regenerate,
        "reject": HITLDecision.rejected,
        "edit_prompt": HITLDecision.edit_prompt,
    }
    new_decision = mapping.get(str(decision or "").strip())
    if new_decision is None:
        raise db_apply.ApplyOpsError(
            f"hitl_decision: неизвестно {decision!r}; есть: {sorted(mapping)}"
        )
    req = (
        await session.execute(
            select(HITLRequest)
            .where(
                HITLRequest.project_id == project.id,
                HITLRequest.decision == HITLDecision.pending,
            )
            .order_by(HITLRequest.id.desc())
        )
    ).scalars().first()
    if req is None:
        raise db_apply.ApplyOpsError("hitl_decision: нет ожидающих HITL-запросов")
    req.decision = new_decision
    req.decided_at = datetime.utcnow()
    await apply_hitl_side_effects(session, req, new_decision)
    await session.commit()
    await publish_hitl_event(
        project.id,
        req.id,
        event_type="hitl_decided",
        payload={"decision": new_decision.value, "kind": req.kind.value},
    )
    return {"hitl_decision": f"{new_decision.value} → {req.kind.value}"}


async def _apply_create_project(session: AsyncSession, spec: dict) -> dict:
    """Создать проект — тот же код, что кнопка «Новый проект» (projects router)."""
    from app.web.routers.projects import create_project as _create_endpoint
    from app.web.schemas import CreateProjectRequest

    title = str((spec or {}).get("title") or "").strip()
    if not title:
        raise db_apply.ApplyOpsError("create_project: нужен title")
    hero_mode = str((spec or {}).get("hero_mode") or "auto").strip()
    if hero_mode not in ("hero", "no_hero", "auto"):
        raise db_apply.ApplyOpsError("create_project: hero_mode ∈ hero|no_hero|auto")
    detail = await _create_endpoint(
        CreateProjectRequest(title=title, hero_mode=hero_mode), session
    )
    return {"id": detail.id, "title": detail.title or title, "slug": detail.slug}


async def _resolve_delete_project_ids(session: AsyncSession, spec: dict) -> list[tuple[int, str]]:
    """Вернуть [(id, title), ...] по ids и/или titles из каталога."""
    raw_ids = spec.get("ids") or spec.get("project_ids") or []
    if isinstance(raw_ids, (int, str)):
        raw_ids = [raw_ids]
    titles = spec.get("titles") or spec.get("names") or []
    if isinstance(titles, str):
        titles = [titles]
    if not raw_ids and not titles:
        raise db_apply.ApplyOpsError(
            "delete_projects: нужен ids:[…] и/или titles:[…] из ПРОЕКТЫ В СТУДИИ"
        )

    want_ids: set[int] = set()
    for x in raw_ids:
        try:
            want_ids.add(int(x))
        except (TypeError, ValueError) as e:
            raise db_apply.ApplyOpsError(f"delete_projects: плохой id {x!r}") from e

    rows = (
        await session.execute(select(Project.id, Project.title, Project.slug, Project.topic))
    ).all()
    by_id = {int(r[0]): r for r in rows}
    for tid in list(want_ids):
        if tid not in by_id:
            raise db_apply.ApplyOpsError(f"delete_projects: проект #{tid} не найден")

    for title in titles:
        q = str(title or "").strip().lower()
        if not q:
            continue
        matches = [
            r
            for r in rows
            if q in ((r[1] or "").strip().lower())
            or q in ((r[2] or "").strip().lower())
            or q in ((r[3] or "")[:80].strip().lower())
        ]
        if not matches:
            raise db_apply.ApplyOpsError(
                f"delete_projects: не найден проект по названию {title!r}"
            )
        if len(matches) > 1:
            opts = ", ".join(f"#{m[0]} {(m[1] or m[2])}" for m in matches[:8])
            raise db_apply.ApplyOpsError(
                f"delete_projects: «{title}» неоднозначно ({opts}) — укажи ids"
            )
        want_ids.add(int(matches[0][0]))

    out: list[tuple[int, str]] = []
    for pid in sorted(want_ids):
        r = by_id[pid]
        label = (r[1] or "").strip() or (r[2] or "") or f"#{pid}"
        out.append((pid, label))
    if not out:
        raise db_apply.ApplyOpsError("delete_projects: пустой список")
    return out


async def _apply_delete_projects(session: AsyncSession, ids: list[int]) -> dict:
    """Удалить проекты (тот же путь, что DELETE /api/projects/{id})."""
    from app.services.event_bus import publish_project_event
    from app.services.sidebar_layout import remove_project_from_layout

    deleted: list[str] = []
    for pid in ids:
        p = await session.get(Project, int(pid))
        if p is None:
            continue
        title = (p.title or p.slug or str(pid)).strip()
        remove_project_from_layout(int(pid))
        await session.delete(p)
        deleted.append(f"#{pid} {title}")
    await session.commit()
    for pid in ids:
        try:
            await publish_project_event(int(pid), event_type="project_deleted")
        except Exception:  # noqa: BLE001
            pass
    if not deleted:
        raise db_apply.ApplyOpsError("delete_projects: нечего удалять (уже нет)")
    return {"delete_projects": f"удалено: {', '.join(deleted)}", "ids": ids}


async def _projects_catalog_context(session: AsyncSession) -> list[str]:
    """Краткий каталог проектов Studio — чтобы оркестратор находил чужие id."""
    rows = (
        await session.execute(
            select(Project.id, Project.slug, Project.title, Project.status, Project.topic)
            .order_by(Project.id.desc())
            .limit(80)
        )
    ).all()
    if not rows:
        return ["ПРОЕКТЫ В СТУДИИ: (пусто)"]
    lines = [
        "ПРОЕКТЫ В СТУДИИ (id | slug | title | status) — ищи по названию отсюда:",
    ]
    for pid, slug, title, status, topic in rows:
        st = getattr(status, "value", status)
        label = (title or "").strip() or (topic or "")[:48].strip() or "—"
        lines.append(f"- #{pid} | {slug} | {label} | {st}")
    return lines


async def _resolve_parent_for_child(
    session: AsyncSession, current: Project, spec: dict | bool | None
) -> Project:
    """parent_id / parent_title / true(=текущий) → Project."""
    if spec is True or spec is None or spec == {}:
        return current
    if not isinstance(spec, dict):
        raise db_apply.ApplyOpsError(
            "create_child: ожидай true | {parent_id} | {parent_title}"
        )
    parent_id = spec.get("parent_id")
    parent_title = str(spec.get("parent_title") or spec.get("title") or "").strip()
    if parent_id is not None:
        try:
            pid = int(parent_id)
        except (TypeError, ValueError) as e:
            raise db_apply.ApplyOpsError("create_child: parent_id должен быть числом") from e
        parent = await session.get(Project, pid)
        if parent is None:
            raise db_apply.ApplyOpsError(f"create_child: проект #{pid} не найден")
        return parent
    if parent_title:
        needle = parent_title.casefold()
        rows = (
            await session.execute(select(Project).order_by(Project.id.desc()).limit(200))
        ).scalars().all()
        hits = [
            p
            for p in rows
            if needle in (p.title or "").casefold()
            or needle in (p.slug or "").casefold()
            or needle in (p.topic or "").casefold()
        ]
        if not hits:
            raise db_apply.ApplyOpsError(
                f"create_child: проект «{parent_title}» не найден — "
                "смотри раздел ПРОЕКТЫ В СТУДИИ"
            )
        if len(hits) > 1:
            # точное совпадение title предпочтительнее
            exact = [p for p in hits if (p.title or "").casefold() == needle]
            hits = exact or hits
        if len(hits) > 1:
            ids = ", ".join(f"#{p.id}" for p in hits[:8])
            raise db_apply.ApplyOpsError(
                f"create_child: несколько проектов «{parent_title}» ({ids}) — укажи parent_id"
            )
        return hits[0]
    return current


async def _apply_create_child(
    session: AsyncSession, current: Project, spec: dict | bool | None
) -> dict:
    """Дочерний проект — тот же код, что POST /api/projects/{id}/child."""
    from app.services.project_child import (
        create_child_from_parent,
        finalize_child_data_dir,
    )
    from app.web.routers.projects import _slugify

    parent = await _resolve_parent_for_child(session, current, spec)
    child = await create_child_from_parent(session, parent, slugify=_slugify)
    await session.commit()
    await session.refresh(child)
    try:
        await finalize_child_data_dir(parent, child)
    except Exception as exc:  # noqa: BLE001
        raise db_apply.ApplyOpsError(f"create_child: data_dir: {exc}") from exc
    return {
        "id": child.id,
        "title": child.title or child.slug,
        "slug": child.slug,
        "parent_id": parent.id,
    }


def _linear_order(nodes: list[dict], edges: list[dict]) -> list[str]:
    """Порядок нод линейной цепочки по edges. Нелинейный граф → ApplyOpsError."""
    by_id = {str(n.get("id")): n for n in nodes}
    out_map: dict[str, list[str]] = {}
    in_deg: dict[str, int] = {nid: 0 for nid in by_id}
    for e in edges:
        src, tgt = str(e.get("source")), str(e.get("target"))
        if src in by_id and tgt in by_id:
            out_map.setdefault(src, []).append(tgt)
            in_deg[tgt] = in_deg.get(tgt, 0) + 1
    heads = [nid for nid, d in in_deg.items() if d == 0]
    if len(heads) != 1:
        raise db_apply.ApplyOpsError(
            "граф нелинейный — сначала {\"repair_graph\":true} (пересоберёт цепочку), "
            "потом повтори add_node"
        )
    order: list[str] = []
    cur = heads[0]
    seen: set[str] = set()
    while cur:
        if cur in seen:
            raise db_apply.ApplyOpsError("цикл в графе — отказ")
        seen.add(cur)
        order.append(cur)
        nxt = out_map.get(cur) or []
        if len(nxt) > 1:
            raise db_apply.ApplyOpsError(
                "граф нелинейный — добавляй по одной ноде (after=<node_key>)"
            )
        cur = nxt[0] if nxt else ""
    if len(order) != len(by_id):
        raise db_apply.ApplyOpsError("в графе есть несвязанные ноды — отказ")
    return order


async def _apply_add_node(session: AsyncSession, project: Project, spec: dict) -> dict:
    """Вставить ноду в граф проекта (project.meta.canvas_graph), цепочка перешивается.

    ``after="each"`` — после каждой рабочей ноды (не hitl/config); иначе
    ``after=<node_key>`` — после конкретной. Тип — любой из реестра нод
    (hitl_gate = нода проверки-аппрува).
    """
    from app.models import Workflow, WorkflowRun
    from app.orchestrator.node_registry import CONFIG_NODE_TYPES, HITL_NODE_TYPES

    node_type = str((spec or {}).get("node_type") or "").strip()
    if node_type not in _all_node_types():
        raise db_apply.ApplyOpsError(
            f"add_node: неизвестный тип {node_type!r}; есть: {sorted(_all_node_types())}"
        )
    after = str((spec or {}).get("after") or "each").strip()
    base_label = str((spec or {}).get("label") or "").strip() or (
        "Проверка"
        if node_type.startswith("hitl")
        else ("Работа с GPT" if node_type == "excel_gpt" else node_type)
    )

    meta = dict(project.meta or {})
    graph = meta.get("canvas_graph")
    if not isinstance(graph, dict) or not graph.get("nodes"):
        wf = (
            await session.execute(
                select(Workflow).where(Workflow.is_default.is_(True))
            )
        ).scalars().first()
        if wf is None:
            raise db_apply.ApplyOpsError("add_node: нет workflow для проекта")
        graph = {"nodes": list(wf.nodes or []), "edges": list(wf.edges or [])}
    nodes = [dict(n) for n in graph["nodes"]]
    edges = [dict(e) for e in (graph.get("edges") or [])]
    order = _linear_order(nodes, edges)
    by_id = {str(n.get("id")): n for n in nodes}

    if after == "each":
        skip = set(CONFIG_NODE_TYPES) | set(HITL_NODE_TYPES)
        next_of_pre = {str(e.get("source")): str(e.get("target")) for e in edges}

        def _is_check_of_type(nid: str) -> bool:
            """Нода — ПРОВЕРКА этого типа, не рабочая.

            hitl-нода — всегда проверка (тип=роль). excel_gpt — по маркеру
            «добавлено оркестратором» (рабочие «Работа с GPT» — не проверки).
            """
            n = by_id.get(nid)
            if n is None or str(n.get("type")) != node_type:
                return False
            if node_type in HITL_NODE_TYPES:
                return True
            return (
                str((n.get("data") or {}).get("description") or "")
                == "добавлено оркестратором"
            )

        def _already_has(nid: str) -> bool:
            nxt = next_of_pre.get(nid)
            return bool(nxt and _is_check_of_type(nxt))

        targets = [
            nid
            for nid in order
            if str(by_id[nid].get("type")) not in skip
            and not _is_check_of_type(nid)  # проверки не получают проверок
            and not _already_has(nid)
        ]
    else:
        if after not in by_id:
            raise db_apply.ApplyOpsError(
                f"add_node: неизвестный after {after!r}; ключи — в разделе НОДЫ КАНВАСА"
            )
        targets = [after]
    if not targets:
        if after == "each":
            return {"add_node": f"{node_type}: уже есть после каждой ноды — дублей нет"}
        raise db_apply.ApplyOpsError("add_node: некуда вставлять (targets пусты)")

    existing_ids = set(by_id)
    new_nodes: list[dict] = []
    k = 0
    # Новые ноды — отдельным рядом ПОД пайплайном, разнесены по x:
    # никогда не накладываются на существующие и друг на друга.
    max_y = max(
        (float((n.get("position") or {}).get("y", 0.0)) for n in nodes), default=0.0
    )
    min_x = min(
        (float((n.get("position") or {}).get("x", 0.0)) for n in nodes), default=0.0
    )
    # excel_gpt-ноды обязаны иметь slotIndex (иначе шаги enrich не различит).
    next_slot = 1
    if node_type == "excel_gpt":
        existing_slots = [
            int((n.get("data") or {}).get("slotIndex") or 0)
            for n in nodes
            if str(n.get("type")) == "excel_gpt"
        ]
        next_slot = max(existing_slots, default=0) + 1

    for nid in targets:
        k += 1
        new_id = f"n_{node_type}_{k}"
        while new_id in existing_ids:
            k += 1
            new_id = f"n_{node_type}_{k}"
        existing_ids.add(new_id)
        # Подпись по родителю: «Проверка — Сценарий», «Работа с GPT — Картинки».
        parent_label = str(
            (by_id[nid].get("data") or {}).get("label") or by_id[nid].get("type") or nid
        )
        node_label = f"{base_label} — {parent_label}"
        new_nodes.append(
            {
                "id": new_id,
                "type": node_type,
                "position": {
                    "x": min_x + (k - 1) * 200.0,
                    "y": max_y + 160.0,
                },
                "data": {
                    "label": node_label,
                    "description": "добавлено оркестратором",
                    **(
                        {"slotIndex": next_slot + len(new_nodes)}
                        if node_type == "excel_gpt"
                        else {}
                    ),
                },
            }
        )

    # Перешивка цепочки: для каждого target: target→next становится target→new→next.
    next_of = {}
    for e in edges:
        next_of[str(e.get("source"))] = str(e.get("target"))
    new_node_for = dict(zip(targets, new_nodes, strict=True))
    out_edges: list[dict] = []
    for e in edges:
        src, tgt = str(e.get("source")), str(e.get("target"))
        if src in new_node_for:
            nn = new_node_for[src]
            out_edges.append(
                {"id": f"e_{src}_{nn['id']}", "source": src, "target": nn["id"]}
            )
            out_edges.append(
                {"id": f"e_{nn['id']}_{tgt}", "source": nn["id"], "target": tgt}
            )
        else:
            out_edges.append(e)
    # target без исходящего ребра (хвост цепочки) — просто хвост target→new.
    for nid, nn in new_node_for.items():
        if nid not in next_of:
            out_edges.append({"id": f"e_{nid}_{nn['id']}", "source": nid, "target": nn["id"]})

    all_nodes = nodes + new_nodes
    meta["canvas_graph"] = {"nodes": all_nodes, "edges": out_edges}
    project.meta = meta

    run = (
        await session.execute(
            select(WorkflowRun)
            .where(WorkflowRun.project_id == project.id)
            .order_by(WorkflowRun.id.desc())
        )
    ).scalars().first()
    if run is not None:
        run.nodes_snapshot = list(all_nodes)
        run.edges_snapshot = list(out_edges)
    await session.commit()
    return {"add_node": f"{node_type} ×{len(new_nodes)} ({'каждой' if after == 'each' else after})"}


async def _remove_node_targets(
    session: AsyncSession, project: Project, spec: dict
) -> tuple[list[dict], list[dict], set[str]]:
    """Граф + целевые node ids для удаления (БЕЗ записи — для preview/confirm)."""
    from app.models import Workflow

    node_key = str((spec or {}).get("node_key") or "").strip()
    node_type = str((spec or {}).get("node_type") or "").strip()
    if not node_key and not node_type:
        raise db_apply.ApplyOpsError("remove_node: нужен node_key или node_type")

    meta = dict(project.meta or {})
    graph = meta.get("canvas_graph")
    if not isinstance(graph, dict) or not graph.get("nodes"):
        wf = (
            await session.execute(
                select(Workflow).where(Workflow.is_default.is_(True))
            )
        ).scalars().first()
        if wf is None:
            raise db_apply.ApplyOpsError("remove_node: нет workflow для проекта")
        graph = {"nodes": list(wf.nodes or []), "edges": list(wf.edges or [])}
    nodes = [dict(n) for n in graph["nodes"]]
    edges = [dict(e) for e in (graph.get("edges") or [])]
    by_id = {str(n.get("id")): n for n in nodes}

    if node_key:
        if node_key not in by_id:
            raise db_apply.ApplyOpsError(
                f"remove_node: неизвестный node_key {node_key!r} (см. НОДЫ КАНВАСА)"
            )
        targets = {node_key}
    else:
        only = str((spec or {}).get("only") or "").strip()
        if only == "duplicates":
            # Только «две проверки подряд»: вторую (оркестраторскую) — в удаление.
            order = _linear_order(nodes, edges)
            type_by_id = {nid: str(by_id[nid].get("type")) for nid in order}
            targets = set()
            prev = ""
            for nid in order:
                if (
                    type_by_id[nid] == node_type
                    and prev
                    and type_by_id.get(prev) == node_type
                ):
                    desc = str((by_id[nid].get("data") or {}).get("description") or "")
                    if desc == "добавлено оркестратором":
                        targets.add(nid)
                prev = nid
            if not targets:
                raise db_apply.ApplyOpsError(
                    f"remove_node: дублей типа {node_type!r} подряд нет"
                )
        elif only:
            raise db_apply.ApplyOpsError(
                f"remove_node: неизвестный only {only!r} (есть: duplicates)"
            )
        else:
            targets = {
                str(n.get("id"))
                for n in nodes
                if str(n.get("type")) == node_type
                and str((n.get("data") or {}).get("description") or "")
                == "добавлено оркестратором"
            }
            if not targets:
                raise db_apply.ApplyOpsError(
                    f"remove_node: нод типа {node_type!r}, добавленных оркестратором, нет"
                )
    return nodes, edges, targets


async def _apply_remove_node(session: AsyncSession, project: Project, spec: dict) -> dict:
    """Удалить ноды из графа проекта (meta.canvas_graph + снапшот прогона).

    ``node_key`` — конкретная нода. ``node_type`` — все ноды этого типа,
    ДОБАВЛЕННЫЕ оркестратором (маркер data.description); чужие/родные ноды
    этого типа не трогаем. Рёбра перешиваются: prev → next.
    """
    from app.models import WorkflowRun

    nodes, edges, targets = await _remove_node_targets(session, project, spec)

    # Перешивка: выжившие ноды связываются в порядке исходной цепочки —
    # корректно и при удалении подряд идущих нод (стопки проверок).
    try:
        order = _linear_order(nodes, edges)
    except db_apply.ApplyOpsError:
        order = None
    if order is not None:
        survivors = [nid for nid in order if nid not in targets]
        out_edges = [
            {"id": f"e_{a}_{b}", "source": a, "target": b}
            for a, b in zip(survivors, survivors[1:], strict=False)
        ]
    else:
        out_edges = [
            e
            for e in edges
            if str(e.get("source")) not in targets and str(e.get("target")) not in targets
        ]
        for tid in targets:
            preds = [str(e.get("source")) for e in edges if str(e.get("target")) == tid]
            succs = [str(e.get("target")) for e in edges if str(e.get("source")) == tid]
            for pr in preds:
                for sc in succs:
                    if pr not in targets and sc not in targets:
                        out_edges.append(
                            {"id": f"e_{pr}_{sc}", "source": pr, "target": sc}
                        )
    left_nodes = [n for n in nodes if str(n.get("id")) not in targets]

    meta = dict(project.meta or {})
    meta["canvas_graph"] = {"nodes": left_nodes, "edges": out_edges}
    project.meta = meta
    run = (
        await session.execute(
            select(WorkflowRun)
            .where(WorkflowRun.project_id == project.id)
            .order_by(WorkflowRun.id.desc())
        )
    ).scalars().first()
    if run is not None:
        run.nodes_snapshot = list(left_nodes)
        run.edges_snapshot = list(out_edges)
    await session.commit()
    return {"remove_node": f"удалено {len(targets)}"}


async def _apply_repair_graph(session: AsyncSession, project: Project) -> dict:
    """Пересобрать цепочку графа слева направо (после разрывов/удалений).

    Порядок — по координате x на канвасе (так его видит человек). Ноды и
    позиции не трогаем, только рёбра.
    """
    from app.models import Workflow, WorkflowRun

    meta = dict(project.meta or {})
    graph = meta.get("canvas_graph")
    if not isinstance(graph, dict) or not graph.get("nodes"):
        wf = (
            await session.execute(
                select(Workflow).where(Workflow.is_default.is_(True))
            )
        ).scalars().first()
        if wf is None:
            raise db_apply.ApplyOpsError("repair_graph: нет workflow для проекта")
        graph = {"nodes": list(wf.nodes or []), "edges": list(wf.edges or [])}
    nodes = [dict(n) for n in graph["nodes"]]
    if len(nodes) < 2:
        raise db_apply.ApplyOpsError("repair_graph: нод меньше двух — нечего чинить")

    def _pos(n: dict) -> tuple[float, float]:
        pos = n.get("position") or {}
        return float(pos.get("x", 0.0)), float(pos.get("y", 0.0))

    # Основной ряд = y с наибольшим числом нод; проверки (ряды ниже)
    # вставляются сразу после своей родительской ноды — по подписи
    # «… — <родитель>» (так их пишет add_node) или по ближайшему x.
    from collections import Counter

    y_counts = Counter(_pos(n)[1] for n in nodes)
    main_y = y_counts.most_common(1)[0][0]
    main_row = sorted((n for n in nodes if _pos(n)[1] == main_y), key=_pos)
    check_rows = [n for n in nodes if _pos(n)[1] != main_y]

    def _label(n: dict) -> str:
        return str((n.get("data") or {}).get("label") or "")

    checks_by_parent: dict[str, list[dict]] = {}
    unmatched: list[dict] = []
    main_labels = {_label(n).strip().casefold() for n in main_row}
    for c in check_rows:
        lab = _label(c)
        parent = lab.split("—")[-1].strip().casefold() if "—" in lab else ""
        if parent and parent in main_labels:
            checks_by_parent.setdefault(parent, []).append(c)
        else:
            unmatched.append(c)

    ordered: list[dict] = []
    for n in main_row:
        ordered.append(n)
        ordered.extend(checks_by_parent.get(_label(n).strip().casefold(), []))
    ordered.extend(sorted(unmatched, key=_pos))

    new_edges = [
        {"id": f"e_{a['id']}_{b['id']}", "source": a["id"], "target": b["id"]}
        for a, b in zip(ordered, ordered[1:], strict=False)
    ]
    meta["canvas_graph"] = {"nodes": nodes, "edges": new_edges}
    project.meta = meta
    run = (
        await session.execute(
            select(WorkflowRun)
            .where(WorkflowRun.project_id == project.id)
            .order_by(WorkflowRun.id.desc())
        )
    ).scalars().first()
    if run is not None:
        run.nodes_snapshot = list(nodes)
        run.edges_snapshot = list(new_edges)
    await session.commit()
    return {"repair_graph": f"цепочка пересобрана: {len(nodes)} нод слева направо"}


async def _apply_rename_node(session: AsyncSession, project: Project, spec: dict) -> dict:
    """Переименовать ноду графа (data.label) — подпись проверок и т.п."""
    from app.models import Workflow, WorkflowRun

    node_key = str((spec or {}).get("node_key") or "").strip()
    label = str((spec or {}).get("label") or "").strip()
    if not node_key or not label:
        raise db_apply.ApplyOpsError("rename_node: нужны node_key и label")

    meta = dict(project.meta or {})
    graph = meta.get("canvas_graph")
    if not isinstance(graph, dict) or not graph.get("nodes"):
        wf = (
            await session.execute(
                select(Workflow).where(Workflow.is_default.is_(True))
            )
        ).scalars().first()
        if wf is None:
            raise db_apply.ApplyOpsError("rename_node: нет workflow для проекта")
        graph = {"nodes": list(wf.nodes or []), "edges": list(wf.edges or [])}
    nodes = [dict(n) for n in graph["nodes"]]
    edges = list(graph.get("edges") or [])
    hit = None
    for n in nodes:
        if str(n.get("id")) == node_key:
            hit = n
            break
    if hit is None:
        raise db_apply.ApplyOpsError(
            f"rename_node: неизвестный node_key {node_key!r} (см. НОДЫ КАНВАСА)"
        )
    data = dict(hit.get("data") or {})
    old_label = str(data.get("label") or "")
    data["label"] = label
    hit["data"] = data

    meta["canvas_graph"] = {"nodes": nodes, "edges": edges}
    project.meta = meta
    run = (
        await session.execute(
            select(WorkflowRun)
            .where(WorkflowRun.project_id == project.id)
            .order_by(WorkflowRun.id.desc())
        )
    ).scalars().first()
    if run is not None:
        run.nodes_snapshot = list(nodes)
        run.edges_snapshot = list(edges)
    await session.commit()
    return {"rename_node": f"{node_key}: «{old_label}» → «{label}»"}


def _cut(text: object, limit: int = 160) -> str:
    s = " ".join(str(text or "").split())
    return s if len(s) <= limit else s[: limit - 1] + "…"


def _orchestrator_context(graph: dict) -> str:
    proj = graph.get("project") or {}
    lines = [
        f"Проект #{proj.get('id')} «{proj.get('title') or proj.get('slug')}», "
        f"статус {proj.get('status')}",
        "Кадры:",
    ]
    for f in graph.get("frames") or []:
        lines.append(
            f"- uuid={f.get('uuid')} кадр {f.get('number')}: "
            f"закадр={_cut(f.get('voiceover_text'))} | "
            f"промт_картинки={_cut(f.get('image_prompt'))} | "
            f"промт_видео={_cut(f.get('animation_prompt'))} | "
            f"смысл={_cut(f.get('meaning'))} | "
            f"длительность={f.get('duration_seconds')}"
        )
    return "\n".join(lines)


class OrchestratorChatBody(BaseModel):
    message: str = Field(min_length=1)
    history: list[dict] = Field(default_factory=list)


@router.post("/projects/{project_id}/orchestrator/chat")
async def orchestrator_chat(
    project_id: int,
    body: OrchestratorChatBody,
    session: AsyncSession = Depends(get_session),
) -> dict:
    """Диалог с оркестратором по текущему проекту.

    Контекст = граф DB (uuid кадров). Ответ модели с ``{"ops": [...]}``
    применяется через ``db_apply.apply_ops`` (fail-closed) + авто-экспорт
    в project.xlsx. Текстовый ответ (вопрос/обсуждение) — без записи.
    """
    from app.services.gpt_client import GptApiUnavailable, get_gpt_client

    project = await _project(session, project_id)
    await db_v2.backfill_project_v2(session, project)
    await session.commit()
    graph = await db_v2.project_graph(session, project)

    history_txt = "\n".join(
        f"{m.get('role', '?')}: {m.get('content', '')}"
        for m in (body.history or [])[-8:]
    )
    diag_lines = await _diagnostics_context(session, project)
    canvas_lines, canvas_keymap = await _canvas_nodes_context(session, project)
    catalog_lines = await _projects_catalog_context(session)
    code_lines = await asyncio.to_thread(_code_workspace_context)
    topic_snip = (project.topic or "").strip().replace("\n", " ")[:160]
    prompt = (
        _ORCHESTRATOR_SYSTEM
        + "\n\n"
        + "\n".join(catalog_lines)
        + "\n\nТЕКУЩИЙ ПРОЕКТ В ЧАТЕ: "
        f"#{project.id} | {project.slug} | {(project.title or '') or '—'} | "
        f"{getattr(project.status, 'value', project.status)}"
        + (f"\nТЕМА: {topic_snip}" if topic_snip else "")
        + "\n"
        + "\nШАГИ ПАЙПЛАЙНА (реальный порядок и состояние):\n"
        + "\n".join(_pipeline_state_lines(project))
        + "\n\n"
        + "\n".join(await _checks_context(session, project))
        + ("\n\n" + "\n".join(diag_lines) if diag_lines else "")
        + "\n\n"
        + "\n".join(code_lines)
        + "\n\n"
        + "\n".join(_settings_context(project))
        + ("\n\n" + "\n".join(canvas_lines) if canvas_lines else "")
        + "\n\nГРАФ ПРОЕКТА:\n"
        + _orchestrator_context(graph)
        + ("\n\nИСТОРИЯ ДИАЛОГА:\n" + history_txt if history_txt else "")
        + "\n\nПОЛЬЗОВАТЕЛЬ: "
        + body.message
    )
    try:
        reply = await get_gpt_client().ask_fresh(
            prompt, timeout=600, project_id=project.id
        )
    except GptApiUnavailable as e:
        raise HTTPException(503, str(e)) from None

    applied = None
    actions_run: list[dict] = []
    ui_actions: list[dict] = []
    pending_confirm: list[dict] = []
    error = None
    ops_data = db_apply.extract_apply_ops_json(reply or "")
    if ops_data:
        ops = ops_data.get("ops") or []
        if ops:
            try:
                applied = await db_apply.apply_ops(
                    session,
                    project,
                    ops,
                    export_xlsx=bool(ops_data.get("export_xlsx", True)),
                )
                await session.commit()
            except db_apply.ApplyOpsError as e:
                await session.rollback()
                error = str(e)
        for act in ops_data.get("actions") or []:
            act = act or {}
            try:
                if "run_step" in act:
                    step = str(act.get("run_step") or "").strip()
                    if not step:
                        continue
                    if step not in STEP_CODE_TO_NODE_TYPE:
                        raise db_apply.ApplyOpsError(
                            f"неизвестный шаг {step!r}; "
                            f"разрешены: {sorted(STEP_CODE_TO_NODE_TYPE)}"
                        )
                    from app.services.project_steps import start_step

                    new_status = await start_step(
                        session, project, step, explicit_ui_start=True
                    )
                    await session.commit()
                    actions_run.append({"run_step": step, "status": new_status.value})
                elif act.get("stop_step"):
                    from app.services.project_control import stop_project_running

                    await stop_project_running(session, project)
                    await session.commit()
                    actions_run.append({"stop_step": True})
                elif "set_option" in act:
                    spec = act.get("set_option") or {}
                    label = _apply_set_option(project, spec.get("key"), spec.get("value"))
                    await session.commit()
                    actions_run.append({"set_option": label})
                elif "set_prompt" in act:
                    spec = act.get("set_prompt") or {}
                    label = _apply_set_prompt(
                        project, spec.get("step"), spec.get("variant")
                    )
                    await session.commit()
                    actions_run.append({"set_prompt": label})
                elif "set_text_llm" in act:
                    spec = act.get("set_text_llm") or {}
                    from app.services.text_llm_catalog import write_choice

                    try:
                        payload = write_choice(
                            provider=str(spec.get("provider") or ""),
                            model_id=spec.get("model_id"),
                        )
                    except ValueError as e:
                        raise db_apply.ApplyOpsError(str(e)) from None
                    actions_run.append(
                        {"set_text_llm": f"{payload['provider']} ({payload['model_id']})"}
                    )
                elif "open_ui" in act:
                    ui_actions.append(
                        await _resolve_open_ui(
                            session, project, act.get("open_ui") or {}, canvas_keymap
                        )
                    )
                elif "hitl_decision" in act:
                    actions_run.append(await _apply_hitl_decision(session, project, act.get("hitl_decision")))
                elif "set_topic" in act:
                    text = str(act.get("set_topic") or "").strip()
                    if not text:
                        raise db_apply.ApplyOpsError("set_topic: пустая тема")
                    project.topic = text
                    await session.commit()
                    actions_run.append({"set_topic": text[:60]})
                elif "create_project" in act:
                    spec = act.get("create_project") or {}
                    created = await _apply_create_project(session, spec)
                    actions_run.append(
                        {"create_project": f"#{created['id']} {created['title']}"}
                    )
                    ui_actions.append({"kind": "open_project", "project_id": created["id"]})
                elif "delete_projects" in act:
                    # Удаление проектов — только после кнопки подтверждения в чате.
                    spec = act.get("delete_projects") or {}
                    targets = await _resolve_delete_project_ids(session, spec)
                    pending_confirm.append(
                        {
                            "kind": "delete_projects",
                            "count": len(targets),
                            "nodes": [str(pid) for pid, _ in targets],
                            "files": [f"#{pid} {title}" for pid, title in targets],
                            "message": "удалить проекты: "
                            + ", ".join(f"#{pid} {title}" for pid, title in targets),
                        }
                    )
                elif "create_child" in act:
                    created = await _apply_create_child(
                        session, project, act.get("create_child")
                    )
                    actions_run.append(
                        {
                            "create_child": (
                                f"#{created['id']} {created['title']} "
                                f"← parent #{created['parent_id']}"
                            )
                        }
                    )
                    ui_actions.append({"kind": "open_project", "project_id": created["id"]})
                elif "add_node" in act:
                    actions_run.append(
                        await _apply_add_node(session, project, act.get("add_node") or {})
                    )
                elif act.get("repair_graph"):
                    actions_run.append(await _apply_repair_graph(session, project))
                elif "rename_node" in act:
                    actions_run.append(
                        await _apply_rename_node(
                            session, project, act.get("rename_node") or {}
                        )
                    )
                elif "remove_node" in act:
                    # Удаление — только с подтверждением человеком (кнопка в чате).
                    spec = act.get("remove_node") or {}
                    _nodes, _edges, targets = await _remove_node_targets(
                        session, project, spec
                    )
                    pending_confirm.append(
                        {
                            "kind": "remove_node",
                            "node_key": spec.get("node_key"),
                            "node_type": spec.get("node_type"),
                            "only": spec.get("only"),
                            "count": len(targets),
                            "nodes": sorted(targets)[:20],
                        }
                    )
                elif act.get("run_harness"):
                    from app.services.agent_harness import run_harness_verify

                    rep = await run_harness_verify(
                        session, project, allow_repair=False, include_http=False
                    )
                    await session.commit()
                    bad = [c.name for c in rep.checks if not c.ok]
                    actions_run.append(
                        {"run_harness": "ок" if rep.ok else f"НЕОК: {bad}"}
                    )
                elif "read_file" in act:
                    from app.services.code_autofix import CodeAutofixError, read_file

                    spec = act.get("read_file") or {}
                    if isinstance(spec, str):
                        spec = {"path": spec}
                    try:
                        rfile = await asyncio.to_thread(
                            lambda: read_file(
                                str(spec.get("path") or ""),
                                start_line=spec.get("start_line"),
                                end_line=spec.get("end_line"),
                            )
                        )
                    except CodeAutofixError as e:
                        raise db_apply.ApplyOpsError(str(e)) from None
                    actions_run.append(
                        {
                            "read_file": (
                                f"{rfile['path']} L{rfile['start_line']}-"
                                f"{rfile['end_line']}/{rfile['total_lines']}"
                                + (" …" if rfile["truncated"] else "")
                                + "\n"
                                + rfile["content"]
                            )
                        }
                    )
                elif "edit_files" in act:
                    from app.services.code_autofix import (
                        CodeAutofixError,
                        apply_edits,
                    )

                    try:
                        # sync FS — в thread, иначе блокирует весь Studio UI
                        result = await asyncio.to_thread(
                            apply_edits, act.get("edit_files") or []
                        )
                    except CodeAutofixError as e:
                        raise db_apply.ApplyOpsError(str(e)) from None
                    actions_run.append(
                        {
                            "edit_files": (
                                f"{result['count']} файл(ов): "
                                + ", ".join(result["changed"][:8])
                            )
                        }
                    )
                elif "run_tests" in act:
                    from app.services.code_autofix import (
                        CodeAutofixError,
                        run_tests,
                    )

                    spec = act.get("run_tests")
                    paths = spec if isinstance(spec, list) else None
                    try:
                        # pytest sync — не блокировать event loop (зависание всего UI)
                        tres = await asyncio.to_thread(
                            lambda: run_tests(paths, timeout=90.0)
                        )
                    except CodeAutofixError as e:
                        raise db_apply.ApplyOpsError(str(e)) from None
                    # FAIL не рвёт весь батч — модель видит вывод и чинит дальше
                    tag = "ok" if tres["ok"] else "FAIL"
                    out = (tres.get("output") or "")[:500]
                    actions_run.append(
                        {
                            "run_tests": (
                                f"{tag} {', '.join(tres['tests'][:6])}"
                                + (f"\n{out}" if out and not tres["ok"] else "")
                            )
                        }
                    )
                elif "git_commit_push" in act:
                    from app.services.code_autofix import CodeAutofixError
                    from app.services.git_ops import GitOpsError, commit_and_push

                    spec = act.get("git_commit_push") or {}
                    if not isinstance(spec, dict):
                        raise db_apply.ApplyOpsError(
                            "git_commit_push: ожидается объект {message, files?, auto?}"
                        )
                    message = str(spec.get("message") or "").strip()
                    files = spec.get("files") or []
                    if files is not None and not isinstance(files, list):
                        raise db_apply.ApplyOpsError(
                            "git_commit_push.files: список путей"
                        )
                    auto = bool(spec.get("auto"))
                    if not message:
                        raise db_apply.ApplyOpsError("git_commit_push: пустой message")
                    if auto:
                        try:
                            pushed = await asyncio.to_thread(
                                commit_and_push,
                                [str(x) for x in files],
                                message,
                            )
                        except (GitOpsError, CodeAutofixError) as e:
                            raise db_apply.ApplyOpsError(str(e)) from None
                        actions_run.append(
                            {
                                "git_commit_push": (
                                    f"pushed {pushed['sha']} "
                                    f"({len(pushed['files'])} files)"
                                )
                            }
                        )
                    else:
                        pending_confirm.append(
                            {
                                "kind": "git_commit_push",
                                "message": message[:500],
                                "files": [str(x) for x in files][:40],
                                "count": len(files) if files else 0,
                                "nodes": [str(x) for x in files][:20],
                            }
                        )
                else:
                    keys = sorted(str(k) for k in act.keys())
                    raise db_apply.ApplyOpsError(
                        f"неизвестное действие {keys}; есть: run_step, stop_step, "
                        "set_option, set_prompt, set_text_llm, open_ui, hitl_decision, "
                        "set_topic, create_project, create_child, delete_projects, "
                        "add_node, remove_node, rename_node, repair_graph, run_harness, "
                        "read_file, edit_files, run_tests, git_commit_push"
                    )
            except db_apply.ApplyOpsError as e:
                error = str(e)
            except Exception as e:  # noqa: BLE001
                await session.rollback()
                error = f"действие {act}: {e}"
    # Контур самообучения: каждый диалог — в ops-журнал проекта.
    try:
        from app.services.agent_harness import append_ops_event

        append_ops_event(
            project.data_dir,
            {
                "type": "orchestrator_chat",
                "message": body.message[:500],
                "reply": (reply or "")[:500],
                "applied": bool(applied),
                "actions": [sorted(a.keys()) for a in actions_run],
                "ui_actions": [u.get("kind") for u in ui_actions],
                "pending_confirm": bool(pending_confirm),
                "error": error,
            },
        )
    except Exception:  # noqa: BLE001
        pass

    return {
        "reply": reply,
        "applied": applied,
        "actions_run": actions_run,
        "ui_actions": ui_actions,
        "pending_confirm": pending_confirm,
        "error": error,
    }


@router.get("/projects/{project_id}/orchestrator/feedback")
async def orchestrator_feedback(
    project_id: int,
    limit: int = 50,
    session: AsyncSession = Depends(get_session),
) -> dict:
    """Хвост журнала диалогов оркестратора (для разбора/самообучения)."""
    import json as _json

    project = await _project(session, project_id)
    path = project.data_dir / "ops" / "events.jsonl"
    out: list[dict] = []
    if path.is_file():
        for ln in path.read_text(encoding="utf-8", errors="replace").splitlines():
            try:
                evt = _json.loads(ln)
            except _json.JSONDecodeError:
                continue
            if evt.get("type") == "orchestrator_chat":
                out.append(evt)
    return {"project_id": project_id, "count": len(out), "events": out[-limit:]}


class ConfirmRemoveBody(BaseModel):
    node_key: str | None = None
    node_type: str | None = None
    only: str | None = None


@router.post("/projects/{project_id}/orchestrator/confirm-remove")
async def orchestrator_confirm_remove(
    project_id: int,
    body: ConfirmRemoveBody,
    session: AsyncSession = Depends(get_session),
) -> dict:
    """Подтверждённое удаление нод (кнопка «Подтвердить» в чате оркестратора)."""
    project = await _project(session, project_id)
    try:
        return await _apply_remove_node(
            session,
            project,
            {"node_key": body.node_key, "node_type": body.node_type, "only": body.only},
        )
    except db_apply.ApplyOpsError as e:
        raise HTTPException(400, str(e)) from None


class ConfirmDeleteProjectsBody(BaseModel):
    ids: list[int] = Field(default_factory=list)


@router.post("/projects/{project_id}/orchestrator/confirm-delete-projects")
async def orchestrator_confirm_delete_projects(
    project_id: int,
    body: ConfirmDeleteProjectsBody,
    session: AsyncSession = Depends(get_session),
) -> dict:
    """Подтверждённое удаление проектов (кнопка в чате оркестратора)."""
    await _project(session, project_id)  # auth / existence of chat context
    ids = [int(x) for x in (body.ids or [])]
    if not ids:
        raise HTTPException(400, "delete_projects: пустой ids")
    try:
        return await _apply_delete_projects(session, ids)
    except db_apply.ApplyOpsError as e:
        raise HTTPException(400, str(e)) from None


class ConfirmGitPushBody(BaseModel):
    message: str
    files: list[str] | None = None


@router.post("/projects/{project_id}/orchestrator/confirm-git-push")
async def orchestrator_confirm_git_push(
    project_id: int,
    body: ConfirmGitPushBody,
    session: AsyncSession = Depends(get_session),
) -> dict:
    """Подтверждённый commit+push в origin/main (кнопка в чате оркестратора)."""
    await _project(session, project_id)  # auth / existence
    from app.services.code_autofix import CodeAutofixError
    from app.services.git_ops import GitOpsError, commit_and_push

    try:
        pushed = await asyncio.to_thread(
            commit_and_push, list(body.files or []), body.message.strip()
        )
    except (GitOpsError, CodeAutofixError) as e:
        raise HTTPException(400, str(e)) from None
    return {
        "git_commit_push": (
            f"git HEAD: {pushed['sha']} уже в {pushed['branch']} "
            f"({len(pushed['files'])} files)"
        ),
        "sha": pushed["sha"],
        "branch": pushed["branch"],
        "files": pushed["files"],
    }


class FramePatch(BaseModel):
    status: str | None = None
    duration_seconds: float | None = None
    meaning: str | None = None
    voiceover_text: str | None = None
    attrs: dict | None = None


@router.patch("/frames/{frame_id}")
async def patch_frame(
    frame_id: int,
    body: FramePatch,
    session: AsyncSession = Depends(get_session),
) -> dict:
    fr = await _frame(session, frame_id)
    if body.status is not None:
        try:
            fr.status = FrameStatus(body.status)
        except ValueError:
            raise HTTPException(400, f"неизвестный статус: {body.status}") from None
    if body.duration_seconds is not None:
        fr.duration_seconds = body.duration_seconds
    if body.meaning is not None:
        fr.meaning = body.meaning
    if body.voiceover_text is not None:
        fr.voiceover_text = body.voiceover_text
    if body.attrs is not None:
        fr.attrs = body.attrs
    await session.commit()
    return {"ok": True, "id": fr.id, "uuid": fr.uuid, "sort_key": fr.sort_key}


class InsertFrameBody(BaseModel):
    after_frame_id: int | None = None
    scene_id: int | None = None


@router.post("/projects/{project_id}/frames/insert", status_code=201)
async def insert_frame(
    project_id: int,
    body: InsertFrameBody,
    session: AsyncSession = Depends(get_session),
) -> dict:
    project = await _project(session, project_id)
    try:
        fr = await db_v2.insert_frame_after(
            session,
            project,
            after_frame_id=body.after_frame_id,
            scene_id=body.scene_id,
        )
    except ValueError as e:
        raise HTTPException(404, str(e)) from None
    await session.commit()
    return {"id": fr.id, "uuid": fr.uuid, "sort_key": fr.sort_key, "scene_id": fr.scene_id}


class TextBody(BaseModel):
    kind: str = Field(default="extra", max_length=32)
    text: str


@router.post("/frames/{frame_id}/texts", status_code=201)
async def add_text(
    frame_id: int,
    body: TextBody,
    session: AsyncSession = Depends(get_session),
) -> dict:
    fr = await _frame(session, frame_id)
    t = FrameText(
        project_id=fr.project_id, frame_id=fr.id, kind=body.kind, text=body.text
    )
    session.add(t)
    await session.commit()
    return {"id": t.id, "kind": t.kind}


@router.delete("/texts/{text_id}")
async def delete_text(
    text_id: int, session: AsyncSession = Depends(get_session)
) -> dict:
    t = await session.get(FrameText, text_id)
    if t is None:
        raise HTTPException(404, "текст не найден")
    await session.delete(t)
    await session.commit()
    return {"ok": True}


class PromptBody(BaseModel):
    kind: str = Field(default="img", max_length=24)
    text: str
    set_active: bool = True


@router.post("/frames/{frame_id}/prompts", status_code=201)
async def add_prompt(
    frame_id: int,
    body: PromptBody,
    session: AsyncSession = Depends(get_session),
) -> dict:
    fr = await _frame(session, frame_id)
    pv = await db_v2.add_prompt_version(
        session,
        fr.project_id,
        fr.id,
        kind=body.kind,
        text=body.text,
        set_active=body.set_active,
    )
    if body.set_active and body.kind == "img":
        fr.image_prompt = body.text
    elif body.set_active and body.kind == "video":
        fr.animation_prompt = body.text
    await session.commit()
    return {"id": pv.id, "version": pv.version, "is_active": pv.is_active}


@router.post("/prompts/{prompt_id}/activate")
async def activate_prompt(
    prompt_id: int, session: AsyncSession = Depends(get_session)
) -> dict:
    pv = await session.get(PromptVersion, prompt_id)
    if pv is None:
        raise HTTPException(404, "версия промта не найдена")
    siblings = (
        await session.execute(
            select(PromptVersion).where(
                PromptVersion.frame_id == pv.frame_id,
                PromptVersion.kind == pv.kind,
                PromptVersion.is_active.is_(True),
            )
        )
    ).scalars().all()
    for s in siblings:
        s.is_active = False
    pv.is_active = True
    fr = await session.get(Frame, pv.frame_id)
    if fr is not None and pv.kind == "img":
        fr.image_prompt = pv.text
    elif fr is not None and pv.kind == "video":
        fr.animation_prompt = pv.text
    await session.commit()
    return {"ok": True, "id": pv.id}


class EntityBody(BaseModel):
    type: str = Field(max_length=24)
    code: str | None = Field(default=None, max_length=24)
    name: str | None = None
    attrs: dict = Field(default_factory=dict)


@router.post("/projects/{project_id}/entities", status_code=201)
async def add_entity(
    project_id: int,
    body: EntityBody,
    session: AsyncSession = Depends(get_session),
) -> dict:
    project = await _project(session, project_id)
    max_key = (
        await session.execute(
            select(func.max(Entity.sort_key)).where(Entity.project_id == project.id)
        )
    ).scalar_one() or 0.0
    en = Entity(
        project_id=project.id,
        type=body.type,
        code=body.code,
        name=body.name,
        attrs=body.attrs,
        sort_key=float(max_key) + 10.0,
    )
    session.add(en)
    await session.commit()
    return {"id": en.id}


@router.patch("/entities/{entity_id}")
async def patch_entity(
    entity_id: int,
    body: EntityBody,
    session: AsyncSession = Depends(get_session),
) -> dict:
    en = await session.get(Entity, entity_id)
    if en is None:
        raise HTTPException(404, "сущность не найдена")
    en.type = body.type
    en.code = body.code
    en.name = body.name
    en.attrs = body.attrs
    await session.commit()
    return {"ok": True}


@router.delete("/entities/{entity_id}")
async def delete_entity(
    entity_id: int, session: AsyncSession = Depends(get_session)
) -> dict:
    en = await session.get(Entity, entity_id)
    if en is None:
        raise HTTPException(404, "сущность не найдена")
    await session.delete(en)
    await session.commit()
    return {"ok": True}


class EdgeBody(BaseModel):
    to_frame_id: int
    type: str = Field(default="next", max_length=24)


@router.post("/frames/{frame_id}/edges", status_code=201)
async def add_edge(
    frame_id: int,
    body: EdgeBody,
    session: AsyncSession = Depends(get_session),
) -> dict:
    fr = await _frame(session, frame_id)
    target = await _frame(session, body.to_frame_id)
    if target.project_id != fr.project_id:
        raise HTTPException(400, "связь между разными проектами запрещена")
    e = FrameEdge(
        project_id=fr.project_id,
        from_frame_id=fr.id,
        to_frame_id=target.id,
        type=body.type,
    )
    session.add(e)
    await session.commit()
    return {"id": e.id}


@router.delete("/edges/{edge_id}")
async def delete_edge(
    edge_id: int, session: AsyncSession = Depends(get_session)
) -> dict:
    e = await session.get(FrameEdge, edge_id)
    if e is None:
        raise HTTPException(404, "связь не найдена")
    await session.delete(e)
    await session.commit()
    return {"ok": True}


class SceneBody(BaseModel):
    title: str | None = None
    place: str | None = None
    meaning: str | None = None
    scene_type: str | None = None
    after_scene_id: int | None = None


@router.post("/projects/{project_id}/scenes", status_code=201)
async def add_scene(
    project_id: int,
    body: SceneBody,
    session: AsyncSession = Depends(get_session),
) -> dict:
    project = await _project(session, project_id)
    scenes = list(
        (
            await session.execute(
                select(Scene).where(Scene.project_id == project.id).order_by(Scene.sort_key)
            )
        ).scalars()
    )
    before: float | None = None
    after: float | None = None
    if body.after_scene_id is not None:
        for i, sc in enumerate(scenes):
            if sc.id == body.after_scene_id:
                before = sc.sort_key
                after = scenes[i + 1].sort_key if i + 1 < len(scenes) else None
                break
    elif scenes:
        before = scenes[-1].sort_key
    sc = Scene(
        project_id=project.id,
        sort_key=db_v2.sort_key_between(before, after),
        title=body.title,
        place=body.place,
        meaning=body.meaning,
        scene_type=body.scene_type,
    )
    session.add(sc)
    await session.commit()
    return {"id": sc.id, "sort_key": sc.sort_key}
