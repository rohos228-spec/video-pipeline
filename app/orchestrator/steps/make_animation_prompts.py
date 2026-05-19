"""Шаг 8: «Промты анимации» — batched-images flow в ОДНОМ чате ChatGPT.

Логика (полностью переделана с per-frame `ask_fresh` на batched-images):

1. Открываем НОВЫЙ чат ChatGPT (`new_conversation`). Это единственное
   место во всём шаге, где разрешено создание нового диалога.
2. Первое сообщение в этом чате: прикладываем `project.xlsx` + файл
   мастер-промта (`prompt_anim_pr.md`) + текст «сопр. сообщения»
   (редактируется юзером в picker'е, см. `gpt_text_builder.anim_pr`).
3. Получаем подтверждающий ответ от GPT (обычно короткий «готов»).
4. Дальше БАТЧАМИ по ~10 картинок отправляем в ТОТ ЖЕ чат
   сгенерированные ранее изображения кадров
   (`<project>/scenes/frame_NNN_<uuid>.png`).
5. В ответ GPT даёт по картинке её анимационный промт. Парсим ответ:
   ищем `frame_NNN` в тексте и забираем до следующего `frame_NNN`
   (или до конца сообщения). Полученные промты:
     - пишем в xlsx: лист «план», колонка кадра (N+2), строка 48
       (см. `xlsx_v8_import.ROW_VIDEO_PROMPT_V8`);
     - пишем в БД: `Frame.animation_prompt` + `FrameStatus`.
6. Переходим к следующему батчу — в ТОТ ЖЕ чат, без `new_conversation`.
7. В случае ЛЮБОЙ ошибки на батче: ждём 10 минут и повторяем тот же
   батч (без нового чата). Ретраи бесконечные — продолжаем пока не
   обработаем все кадры. Юзер всегда может ⏹ остановить шаг.

Контракт со старой логикой:
  - Входной статус: `generating_animation_prompts` (выставляется
    `on_prompt_picker_cb` при ▶ Запустить шаг).
  - Выходной статус: `animation_prompts_ready`.
  - Если все кадры уже имеют `animation_prompt` — сразу переводим в
    ready без обращения к GPT.
"""

from __future__ import annotations

import asyncio
import re
from datetime import datetime
from pathlib import Path

from aiogram import Bot  # noqa: F401
from loguru import logger
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.bots.browser import browser_session
from app.bots.chatgpt import ChatGPTBot
from app.models import Frame, FrameStatus, Project, ProjectStatus
from app.services import gpt_text_builder as gtb
from app.services.gpt_check import (
    GptCheckDecision,
    gpt_check_file_artifact,
    load_check_prompt,
)
from app.services.prompt_library import get_project_prompt
from app.services.step_cancel import StepCancelledError, raise_if_cancelled
from app.services.xlsx_v8_import import (
    ROW_VIDEO_PROMPT_V8,
    SHEET_PLAN_V8,
)
from app.storage import for_project as _sheet_for_project

# Размер батча картинок, отправляемых в одном сообщении ChatGPT.
# Не жёстко 10 — если осталось меньше, последний батч будет короче.
_BATCH_SIZE_DEFAULT = 10

# Пауза перед повтором батча в случае ошибки (10 минут — по требованию
# пользователя).
_RETRY_SLEEP_SECONDS = 600.0

# Тайм-аут одного запроса к ChatGPT внутри батча (картинки могут
# обрабатываться долго, особенно при батче в 10 фото).
_GPT_TIMEOUT = 900.0

# Регэксп для парсинга ответа GPT — ищем `frame_NNN` (с опциональным
# `_<uuid>.png` суффиксом). Группа 1 — номер кадра (3 цифры).
_FRAME_RE = re.compile(
    r"frame[_\-\s]*0*(\d+)(?:[_\-][a-f0-9]+)?(?:\.png)?",
    re.IGNORECASE,
)


def _parse_batch_reply(reply: str) -> dict[int, str]:
    """Парсит ответ GPT и возвращает mapping `frame_number → animation_prompt`.

    Стратегия: ищем все вхождения `frame_NNN` в ответе. Текст между
    очередным `frame_NNN` и следующим (или концом строки) считаем
    промтом этого кадра. Чистим ведущие двоеточия / тире / стрелки.

    Если в ответе нет ни одного `frame_NNN` — возвращаем пустой dict
    (вызывающий код кинет ошибку, чтобы запустить retry).
    """
    matches = list(_FRAME_RE.finditer(reply or ""))
    if not matches:
        return {}

    out: dict[int, str] = {}
    for i, m in enumerate(matches):
        try:
            frame_no = int(m.group(1))
        except (ValueError, TypeError):
            continue
        start = m.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(reply)
        chunk = reply[start:end].strip()
        # Срезаем ведущие разделители (двоеточие, тире, стрелка, кавычки).
        chunk = re.sub(r"^[\s\:\-\u2014\u2013=>\"'`]+", "", chunk).strip()
        # Срезаем трейлинг кавычек/запятых от JSON-подобных форматов.
        chunk = re.sub(r"[\s\,\"'`]+$", "", chunk).strip()
        if chunk and frame_no not in out:
            out[frame_no] = chunk
    return out


def _frame_image_path(project: Project, frame_number: int) -> Path | None:
    """Находит существующий файл картинки кадра в `<data_dir>/scenes/`.

    Возвращает самый свежий по mtime файл, имя которого начинается с
    `frame_<NNN>_`. Если файлов нет — `None`.
    """
    scenes_dir = project.data_dir / "scenes"
    if not scenes_dir.exists():
        return None
    candidates = list(scenes_dir.glob(f"frame_{frame_number:03d}_*.png"))
    if not candidates:
        return None
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0]


def _write_anim_prompt_to_xlsx_v8(
    project: Project, frame_number: int, animation_prompt: str
) -> bool:
    """Записывает промт анимации в xlsx — лист «план», колонка `frame_number+2`,
    строка 48 (см. `ROW_VIDEO_PROMPT_V8`).

    Это «прямой» путь записи (мимо `ProjectSheet.write_frame`) — нужен
    потому что ProjectSheet работает с листом «Кадры» (v7 layout),
    а user'ский xlsx использует лист «план» (v8 layout).
    Возвращает True при успехе.
    """
    xlsx_path = project.data_dir / "project.xlsx"
    if not xlsx_path.exists():
        logger.warning(
            "[#{}] write_anim_prompt: project.xlsx не найден: {}",
            project.id, xlsx_path,
        )
        return False
    try:
        wb = load_workbook(filename=str(xlsx_path))
        if SHEET_PLAN_V8 not in wb.sheetnames:
            logger.warning(
                "[#{}] write_anim_prompt: листа «{}» нет в xlsx, пропускаю",
                project.id, SHEET_PLAN_V8,
            )
            return False
        ws = wb[SHEET_PLAN_V8]
        # column 3 = frame 1, column 4 = frame 2, ...
        col = frame_number + 2
        ws.cell(row=ROW_VIDEO_PROMPT_V8, column=col, value=animation_prompt)
        wb.save(str(xlsx_path))
        return True
    except Exception as e:  # noqa: BLE001
        logger.warning(
            "[#{}] write_anim_prompt: openpyxl упал: {}", project.id, e,
        )
        return False


async def _send_first_message(
    gpt: ChatGPTBot,
    project: Project,
    prompt_file: Path,
    xlsx_path: Path,
    accompanying: str,
) -> None:
    """Открывает новый чат и отправляет первое сообщение шага.

    Это единственное место во всём шаге, где вызывается `new_conversation()`.
    Дальше — только `ask_with_files` без открытия нового чата.

    При ошибке кидает наружу — вызывающий код решает, retry'ить с новым
    чатом (можно, поскольку ответа ещё не было) или нет.
    """
    logger.info(
        "[#{}] anim_pr: открываю новый чат + первое сообщение "
        "(prompt_file={}, xlsx={}, accompanying_len={})",
        project.id,
        prompt_file.name,
        xlsx_path.name,
        len(accompanying),
    )
    await gpt.new_conversation()
    reply = await gpt.ask_with_files(
        accompanying.strip(),
        [prompt_file, xlsx_path],
        timeout=_GPT_TIMEOUT,
    )
    reply_text = (reply or "").strip()
    if not reply_text:
        # Это значит, что send из chatgpt.py не сработал (либо сработал, но
        # GPT не успел ответить за timeout). В любом случае нельзя считать
        # это успехом — иначе следующий батч уплывёт в застрявший композер.
        raise RuntimeError(
            "anim_pr: первое сообщение — пустой ответ GPT "
            "(вероятно, send в композере не прошёл). Пойду в retry."
        )
    logger.info(
        "[#{}] anim_pr: первое сообщение принято, GPT ответил {} симв.",
        project.id, len(reply_text),
    )


async def _process_batch(
    gpt: ChatGPTBot,
    project: Project,
    session: AsyncSession,
    frames_to_process: list[Frame],
    batch_index: int,
    total_pending: int,
) -> dict[int, str]:
    """Отправляет один батч картинок в текущий чат, парсит ответ и
    возвращает mapping `frame_number → animation_prompt`.

    Сами картинки берутся из `<data_dir>/scenes/`. Кадры, для которых
    файл не найден, пропускаются в этом батче (юзеру вылетит warning).
    Если ни одного файла не нашлось — кидаем RuntimeError, потому что
    батч пустой и присылать в чат нечего.

    Запись в xlsx / БД делает вызывающий код — после успешного парсинга.
    """
    image_paths: list[Path] = []
    pairs: list[tuple[Frame, Path]] = []
    for fr in frames_to_process:
        img = _frame_image_path(project, fr.number)
        if img is None:
            logger.warning(
                "[#{}] anim_pr batch #{}: для кадра {} нет файла в scenes/, "
                "пропускаю в этом батче",
                project.id, batch_index, fr.number,
            )
            continue
        image_paths.append(img)
        pairs.append((fr, img))

    if not image_paths:
        raise RuntimeError(
            f"anim_pr batch #{batch_index}: в scenes/ нет файлов для кадров "
            f"{[f.number for f in frames_to_process]}"
        )

    frame_nums = [f.number for f, _ in pairs]
    batch_text = (
        f"Партия #{batch_index}. Прикладываю {len(image_paths)} картинок "
        f"(кадры {frame_nums[0]}–{frame_nums[-1]} из общей очереди в "
        f"{total_pending} штук).\n\n"
        "Для КАЖДОЙ картинки верни её промт анимации в формате, который "
        "ты получил в инструкции (мастер-промт). Имя файла обязательно "
        "включай в ответ — по нему я найду нужную колонку в xlsx.\n\n"
        "Никаких других файлов прикладывать не нужно — только текстовый "
        "ответ. Если для какой-то картинки промт получить не получилось — "
        "так и напиши, я повторно её пришлю в следующей итерации."
    )

    logger.info(
        "[#{}] anim_pr batch #{}: отправляю {} картинок (кадры {})",
        project.id, batch_index, len(image_paths), frame_nums,
    )
    reply = await gpt.ask_with_files(
        batch_text, image_paths, timeout=_GPT_TIMEOUT,
    )
    reply_text = (reply or "").strip()
    logger.info(
        "[#{}] anim_pr batch #{}: получил ответ {} симв.",
        project.id, batch_index, len(reply_text),
    )
    if not reply_text:
        # Аналогично _send_first_message: пустой ответ = либо send в композере
        # не прошёл, либо GPT не успел ответить. Не пишем «успех» — кидаем
        # ошибку, чтобы retry-loop ушёл в 10-мин паузу и попробовал заново.
        raise RuntimeError(
            f"anim_pr batch #{batch_index}: пустой ответ GPT "
            f"(вероятно, send в композере не прошёл)"
        )

    parsed = _parse_batch_reply(reply_text)
    # Оставляем только промты для кадров из этого батча.
    valid = {n: p for n, p in parsed.items() if n in frame_nums}
    missing = [n for n in frame_nums if n not in valid]
    if not valid:
        raise RuntimeError(
            f"anim_pr batch #{batch_index}: GPT не вернул ни одного "
            f"промта для кадров {frame_nums}; ответ (первые 500 симв): "
            f"{(reply or '')[:500]!r}"
        )
    if missing:
        logger.warning(
            "[#{}] anim_pr batch #{}: GPT не вернул промт для кадров {}; "
            "они останутся в очереди на следующий проход.",
            project.id, batch_index, missing,
        )
    return valid


async def run(session: AsyncSession, project: Project, bot: Bot) -> None:
    """Основная точка входа шага 8 «Промты анимации»."""
    if project.status is not ProjectStatus.generating_animation_prompts:
        return
    logger.info("[#{}] make_animation_prompts starting (batched-images flow)", project.id)

    # 1. Загружаем все кадры проекта по порядку.
    all_frames = (
        await session.execute(
            select(Frame).where(Frame.project_id == project.id).order_by(Frame.number)
        )
    ).scalars().all()
    if not all_frames:
        logger.warning("[#{}] anim_pr: нет ни одного кадра в БД", project.id)
        project.status = ProjectStatus.animation_prompts_ready
        await session.flush()
        return

    # 2. Отбираем кадры, у которых ещё нет animation_prompt.
    pending = [fr for fr in all_frames if not (fr.animation_prompt or "").strip()]
    if not pending:
        logger.info(
            "[#{}] anim_pr: у всех {} кадров уже есть animation_prompt — "
            "сразу ready", project.id, len(all_frames),
        )
        project.status = ProjectStatus.animation_prompts_ready
        await session.flush()
        try:
            _sheet_for_project(project).write_general(status=project.status.value)
        except Exception as e:  # noqa: BLE001
            logger.warning(
                "[#{}] xlsx write_general(status) failed: {}", project.id, e,
            )
        return

    logger.info(
        "[#{}] anim_pr: обрабатываю {} кадров (всего в проекте {})",
        project.id, len(pending), len(all_frames),
    )

    # 3. Готовим мастер-промт как файл + сопровождающий текст.
    proj_xlsx = project.data_dir / "project.xlsx"
    if not proj_xlsx.exists():
        raise RuntimeError(
            f"anim_pr: project.xlsx не найден ({proj_xlsx}); нечего отправлять в GPT"
        )

    try:
        master = get_project_prompt(project, "anim_pr")
    except FileNotFoundError:
        master = (
            "# anim_pr\n\nМастер-промт не настроен. Открой "
            "prompts/07_animation/default.md и опиши там задачу: "
            "по каждой присланной картинке вернуть промт анимации, "
            "обязательно с именем файла."
        )

    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    out_dir = proj_xlsx.parent / "tmp_gpt"
    out_dir.mkdir(parents=True, exist_ok=True)
    prompt_file = out_dir / f"prompt_anim_pr_{ts}.md"
    prompt_file.write_text((master or "").strip(), encoding="utf-8")

    accompanying = gtb.get_effective_text(
        project, "anim_pr", n_frames=len(pending),
    )

    # 4. Открываем браузер и весь шаг идёт в ОДНОМ чате.
    async with browser_session() as bs:
        gpt = ChatGPTBot(bs)
        try:
            # 4a. ПЕРВОЕ СООБЩЕНИЕ — пока не доставлено, можно
            # пере-открывать чат (юзер сказал «никаких новых диалогов
            # ПОСЛЕ ответа GPT»; до первого ответа — ещё можно).
            first_done = False
            while not first_done:
                raise_if_cancelled(project.id)
                try:
                    await _send_first_message(
                        gpt, project, prompt_file, proj_xlsx, accompanying,
                    )
                    first_done = True
                except StepCancelledError:
                    raise
                except Exception as e:  # noqa: BLE001
                    logger.exception(
                        "[#{}] anim_pr: первое сообщение упало ({}). "
                        "Жду {} сек и пробую заново (с новым чатом).",
                        project.id, e, int(_RETRY_SLEEP_SECONDS),
                    )
                    await asyncio.sleep(_RETRY_SLEEP_SECONDS)

            # 4b. БАТЧИ КАРТИНОК — в ТОТ ЖЕ чат, без new_conversation.
            batch_index = 0
            while True:
                raise_if_cancelled(project.id)

                # Перечитываем список pending каждый круг — на случай,
                # если кто-то параллельно проставил animation_prompt
                # (через xlsx-импорт, например).
                fresh_pending = [
                    fr for fr in all_frames
                    if not (fr.animation_prompt or "").strip()
                ]
                if not fresh_pending:
                    logger.info(
                        "[#{}] anim_pr: все кадры обработаны", project.id,
                    )
                    break

                batch_index += 1
                batch = fresh_pending[:_BATCH_SIZE_DEFAULT]

                # Внутренний retry — пока батч не отработает.
                while True:
                    raise_if_cancelled(project.id)
                    try:
                        parsed = await _process_batch(
                            gpt, project, session, batch,
                            batch_index, len(fresh_pending),
                        )
                    except StepCancelledError:
                        raise
                    except Exception as e:  # noqa: BLE001
                        logger.exception(
                            "[#{}] anim_pr batch #{}: упало ({}). "
                            "Жду {} сек и повторяю в ТОТ ЖЕ чат.",
                            project.id, batch_index, e,
                            int(_RETRY_SLEEP_SECONDS),
                        )
                        await asyncio.sleep(_RETRY_SLEEP_SECONDS)
                        continue

                    # 4c. Применяем результаты: xlsx + БД.
                    applied = 0
                    sheet = None
                    try:
                        sheet = _sheet_for_project(project)
                    except Exception as e:  # noqa: BLE001
                        logger.warning(
                            "[#{}] anim_pr: _sheet_for_project упал: {}",
                            project.id, e,
                        )

                    for fr in batch:
                        prm = parsed.get(fr.number)
                        if not prm:
                            # GPT не вернул для этого кадра — оставим в
                            # очереди на следующий проход (попадёт в
                            # следующий батч в новом круге while).
                            continue
                        fr.animation_prompt = prm
                        fr.status = FrameStatus.animation_prompt_ready
                        await session.flush()
                        # Пишем в v8-xlsx (лист «план», R48).
                        _write_anim_prompt_to_xlsx_v8(
                            project, fr.number, prm,
                        )
                        # Зеркалим во v7-xlsx (лист «Кадры», write_frame),
                        # на случай если у проекта старый формат и тоже
                        # ждёт animation_prompt.
                        if sheet is not None:
                            try:
                                sheet.write_frame(
                                    fr.number,
                                    animation_prompt=prm,
                                    frame_status=fr.status.value,
                                )
                            except Exception as e:  # noqa: BLE001
                                logger.warning(
                                    "[#{}] xlsx write_frame(animation_prompt) "
                                    "failed for frame {}: {}",
                                    project.id, fr.number, e,
                                )
                        applied += 1

                    logger.info(
                        "[#{}] anim_pr batch #{}: применил {} промтов "
                        "(из {} запрошенных)",
                        project.id, batch_index, applied, len(batch),
                    )
                    # Батч закончен (даже если не все промты получены —
                    # оставшиеся кадры всплывут в следующем круге).
                    break

        except StepCancelledError as e:
            logger.info(
                "[#{}] make_animation_prompts: {} — выхожу из цикла",
                project.id, e,
            )
            try:
                await session.refresh(project)
            except Exception:  # noqa: BLE001
                logger.warning("[#{}] не смог refresh project после ⏹", project.id)
            return

    # 5. (Фаза 3) GPT-проверка xlsx после записи anim-промтов.
    try:
        check_prompt = load_check_prompt("animation_prompts")
    except FileNotFoundError:
        check_prompt = None
        logger.warning("[#{}] промт check_animation_prompts не найден, пропускаю GPT-check", project.id)
    sheet_for_check = _sheet_for_project(project)
    xlsx_path = sheet_for_check.ensure_initialized(project_id=project.id, slug=project.slug)
    if check_prompt and xlsx_path.exists():
        tmp_dir = xlsx_path.parent / "tmp_gpt"
        tmp_dir.mkdir(parents=True, exist_ok=True)
        async with browser_session() as bs:
            gpt_chk = ChatGPTBot(bs)
            check_result = await gpt_check_file_artifact(
                chatgpt_bot=gpt_chk,
                check_prompt=check_prompt,
                artifact_path=xlsx_path,
                new_conversation=True,
                timeout=1200.0,
                download_replacement_to=tmp_dir / "animation_prompts_replaced.xlsx",
            )
            logger.info(
                "[#{}] animation_prompts GPT-check: decision={}",
                project.id, check_result.decision.value,
            )
            if check_result.decision is GptCheckDecision.replace_artifact:
                if check_result.replaced_path and check_result.replaced_path.exists():
                    import shutil
                    shutil.copy2(str(check_result.replaced_path), str(xlsx_path))
                    logger.info("[#{}] animation_prompts: GPT заменил xlsx", project.id)
                    from app.services.xlsx_v8_import import import_v8_xlsx
                    try:
                        await import_v8_xlsx(session, project, xlsx_path, keep_fields=False, update_frames_voiceover=False)
                    except Exception as e:  # noqa: BLE001
                        logger.warning("[#{}] animation_prompts resync after replace failed: {}", project.id, e)

    # 6. Все кадры обработаны — статус ready.
    project.status = ProjectStatus.animation_prompts_ready
    await session.flush()
    try:
        _sheet_for_project(project).write_general(status=project.status.value)
    except Exception as e:  # noqa: BLE001
        logger.warning("[#{}] xlsx write_general(status) failed: {}", project.id, e)
