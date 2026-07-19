"""Шаг 6: генерация картинок по уже готовым промтам (outsee nano-banana-2).

Промты должны быть подготовлены на шаге 5 (generate_image_prompts).
Этот шаг только генерит и валидирует картинки.

Входной статус: generating_images.
Выходной статус: images_ready.

Алгоритм (НЕ БЛОКИРУЕТСЯ на ожидании approve пользователя):
  Фаза A — все первые кадры (shot_01, строка 45 xlsx):
  1. Берёт следующий кадр в статусе image_prompt_ready.
  2. Генерит картинку в outsee, сохраняет ``frame_NNN_<uuid>.png``.

  Фаза B — вторые кадры (shot_02, строка 46), только где enrich заполнил
  блок 16–29 / есть промт:
  3. После завершения фазы A — для каждой сцены с shot_02 генерит
     ``frame_NNN_s2_<uuid>.png`` с референсом = PNG первого кадра той же колонки.

  HITL-карточки шлются, но воркер не ждёт approve между кадрами.
"""

from __future__ import annotations

import asyncio
import re
import uuid
from pathlib import Path

from aiogram import Bot
from loguru import logger
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.bots.browser import browser_session
from app.bots.chatgpt import ChatGPTBot
from app.bots.outsee import (
    OutseeBot,
    OutseeContentRejectedError,
    OutseeImageError,
    outsee_error_kind,
    outsee_error_kind_label,
)
from app.generation_options import (
    ASPECT_RATIOS_BY_ID,
    DEFAULTS,
    IMAGE_GENERATORS_BY_ID,
    IMAGE_RESOLUTIONS_BY_ID,
    OUTSEE_PROMPT_MAX_CHARS,
    build_gen_id_prefix,
    is_skippable_empty_prompt,
    prepend_gen_id,
    resolve_image_quality_slug,
)
from app.models import (
    Artifact,
    ArtifactKind,
    Frame,
    FrameStatus,
    HITLDecision,
    HITLKind,
    HITLRequest,
    Project,
    ProjectStatus,
)
from app.services.hitl import send_hitl_photo
from app.services.plan_shot2 import (
    SHOT2_PROMPT_ATTR,
    SHOT2_STATUS_ATTR,
    disk_has_shot2_image,
    find_shot1_image,
    read_shot2_columns,
)
from app.services.scan_frames import (
    disk_has_valid_frame_image,
    frame_needs_shot1_image,
    is_valid_scene_image,
    newest_frame_image_path,
)
from app.services.outsee_retry import generate_image_with_retries
from app.services.step_cancel import (
    StepCancelledError,
    consume_stop,
    raise_if_cancelled,
    sleep_cancellable,
)
from app.settings import settings
from app.storage import for_project as _sheet_for_project

# Лист «план» v8 — какие строки в столбце кадра используются для рефов.
_XLSX_SHEET_PLAN = "план"
# v8-шаблон «план» дублирует лейблы «персонажи» / «предметы» в нескольких
# блоках (под заголовками «кадр1», «кадр2», «кадр3»). Юзер может вписать
# id в ЛЮБОЙ из этих строк. Раньше код смотрел только row=38/39 (3-й
# блок), и если юзер вписал в row=8 — рефы не подгружались. Теперь
# читаем ВСЕ три строки и сливаем (с dedupe сохраняя порядок).
_XLSX_ROWS_PERSONS = (8, 23, 38)   # «персонажи» — id c01..c05
_XLSX_ROWS_ITEMS = (9, 24, 39)     # «предметы» — id i01 / predmet1
_OUTSEE_MAX_REFS = 2  # лимит Outsee на одну генерацию картинки

_REF_ID_RE = re.compile(r"^(c\d+|i\d+|predmet\d+)$", re.IGNORECASE)


def normalize_ref_id(token: str) -> str | None:
    """Нормализует id из ячейки «план» (c02:, I01 → c02/i01). Мусор отбрасывает."""
    t = (token or "").strip().lower()
    t = t.rstrip(":;,.)]}»\"'")
    if not t or not _REF_ID_RE.match(t):
        return None
    return t


def ref_id_file_aliases(ref_id: str) -> list[str]:
    """Варианты имён файлов: i01 ↔ predmet1 (новый/старый лист «Предметы»)."""
    rid = normalize_ref_id(ref_id)
    if not rid:
        return []
    aliases: list[str] = [rid]
    if rid.startswith("i") and rid[1:].isdigit():
        aliases.append(f"predmet{int(rid[1:])}")
    elif rid.startswith("predmet") and rid[7:].isdigit():
        n = int(rid[7:])
        aliases.append(f"i{n:02d}")
    return list(dict.fromkeys(aliases))


def _parse_ref_ids(cell_value: object) -> list[str]:
    """Парсит строку из xlsx-ячейки в список ID. Поддерживает разделители:
    запятая, пробел, точка с запятой, знак «+». Пустые токены и whitespace
    игнорируются. Регистр приводим к lower-case (id хранится как c01/i01).
    """
    if cell_value is None:
        return []
    s = str(cell_value).strip()
    if not s:
        return []
    # Заменим разделители на запятые и split.
    for ch in (";", "+", "/", "|", " "):
        s = s.replace(ch, ",")
    out: list[str] = []
    for tok in s.split(","):
        norm = normalize_ref_id(tok)
        if norm:
            out.append(norm)
    return out


def _resolve_plan_sheet(wb):  # noqa: ANN001
    """Лист «план» (v8), без учёта регистра."""
    if _XLSX_SHEET_PLAN in wb.sheetnames:
        return wb[_XLSX_SHEET_PLAN]
    low = _XLSX_SHEET_PLAN.casefold()
    for name in wb.sheetnames:
        if name.casefold() == low:
            return wb[name]
    return None


def _find_ref_file(base_dir: Path, ref_id: str) -> Path | None:
    """Ищет файл вида `<ref_id>_<anything>.png` в указанной папке.
    Возвращает САМЫЙ СВЕЖИЙ (по mtime) — если у юзера несколько
    регенераций одного персонажа/предмета. None если ничего нет.
    """
    if not base_dir.is_dir():
        return None
    candidates: list[Path] = []
    for ext in ("png", "jpg", "jpeg", "webp"):
        candidates.extend(base_dir.glob(f"{ref_id}_*.{ext}"))
        # Бывают legacy-имена для hero: `hero_<N>_v1_<uuid>.png`. Для
        # ref_id="c01" это не подойдёт — но если юзер положил «c01.png»
        # без суффикса, тоже подберём.
        for p in base_dir.glob(f"{ref_id}.{ext}"):
            candidates.append(p)
    if not candidates:
        return None
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0]


def _find_ref_file_any(base_dir: Path, ref_id: str) -> Path | None:
    """Ищет файл по id и синонимам (i01 / predmet1)."""
    for alias in ref_id_file_aliases(ref_id):
        found = _find_ref_file(base_dir, alias)
        if found is not None:
            return found
    return None


async def _artifact_ref_path(
    session: AsyncSession | None,
    project_id: int,
    ref_id: str,
    *,
    kind: str,
) -> Path | None:
    """Fallback: hero_reference / item_reference из БД, если файл не в папке."""
    if session is None:
        return None
    from sqlalchemy import desc

    aliases = set(ref_id_file_aliases(ref_id))
    if not aliases:
        return None
    if kind == "character":
        kind_filter = ArtifactKind.hero_reference
    elif kind == "item":
        kind_filter = ArtifactKind.item_reference
    else:
        return None
    rows = (
        await session.execute(
            select(Artifact)
            .where(
                Artifact.project_id == project_id,
                Artifact.kind == kind_filter,
            )
            .order_by(desc(Artifact.id))
        )
    ).scalars().all()
    for art in rows:
        meta = art.meta or {}
        candidates: set[str] = set()
        for key in ("excel_id", "item_id"):
            raw = meta.get(key)
            if raw:
                norm = normalize_ref_id(str(raw))
                if norm:
                    candidates.add(norm)
        if kind == "item":
            idx = meta.get("item_index")
            if isinstance(idx, int) and idx > 0:
                candidates.add(f"predmet{idx}")
                candidates.add(f"i{idx:02d}")
        if kind == "character":
            hero_idx = meta.get("hero_index")
            if isinstance(hero_idx, int) and hero_idx > 0:
                candidates.add(f"c{hero_idx:02d}")
        if not candidates & aliases:
            continue
        if art.path:
            path = Path(art.path)
            if path.is_file():
                return path
    return None


def _hero_legacy_ref(project_data_dir: Path, persons_id: str) -> Path | None:
    """Fallback для старых проектов: реф персонажа c0X лежит как
    `hero_X_v1_<uuid>.png` (нумерация по old hero_index, X = int(ID[1:])).
    Возвращает самый свежий v1 — если c0X парсится в число.
    """
    if not persons_id.startswith("c"):
        return None
    try:
        idx = int(persons_id[1:])
    except ValueError:
        return None
    chars_dir = project_data_dir / "characters"
    if not chars_dir.is_dir():
        return None
    candidates = sorted(
        chars_dir.glob(f"hero_{idx}_v1_*.png"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    return candidates[0] if candidates else None


async def _collect_ref_paths(
    session: AsyncSession | None,
    project: Project,
    ref_ids: list[str],
    *,
    kind: str,
    base_dir: Path,
    frame_number: int,
    max_count: int,
) -> list[Path]:
    """Резолвит id из xlsx в файлы. До max_count штук, порядок как в ячейке."""
    if max_count <= 0:
        return []
    refs: list[Path] = []
    for rid in ref_ids:
        if len(refs) >= max_count:
            break
        if kind == "character":
            found = (
                _find_ref_file_any(base_dir, rid)
                or _hero_legacy_ref(project.data_dir, rid)
                or await _artifact_ref_path(
                    session, project.id, rid, kind="character"
                )
            )
            label = "персонаж"
            missing_hint = (
                "запусти шаг «Персонажи» или положи cNN.png в characters/"
            )
        elif kind == "item":
            found = _find_ref_file_any(
                base_dir, rid
            ) or await _artifact_ref_path(session, project.id, rid, kind="item")
            label = "предмет"
            missing_hint = (
                "запусти шаг «Предметы» или положи iNN.png / predmetN.png в items/"
            )
        else:
            continue
        if found is not None:
            refs.append(found)
            logger.info(
                "[#{}] frame {} ref {} '{}' → {}",
                project.id, frame_number, label, rid, found,
            )
        else:
            logger.warning(
                "[#{}] frame {} ref {} '{}' не найден "
                "(папка {}) — {}",
                project.id, frame_number, label, rid, base_dir, missing_hint,
            )
    return refs


async def _load_refs_for_frame(
    session: AsyncSession | None,
    project: Project,
    frame_number: int,
) -> list[Path]:
    """Читает xlsx-ячейки «персонажи» / «предметы» для столбца кадра.

    Outsee — максимум 2 рефа на генерацию. Порядок заполнения слотов:
      1) персонажи из ячейки (c01, c02 через запятую — до 2 найденных);
      2) предметы — в оставшиеся слоты;
      3) постоянный продукт массового — если остался свободный слот.
    """
    refs: list[Path] = []
    xlsx_path = (
        project.data_dir / "project.xlsx"
    )
    persons_ids: list[str] = []
    items_ids: list[str] = []
    if xlsx_path.exists():
        try:
            from openpyxl import load_workbook  # ленивый импорт
            wb = load_workbook(xlsx_path, data_only=True, read_only=True)
            ws = _resolve_plan_sheet(wb)
            if ws is not None:
                # В v8 столбцы кадров — с 3 (1=label, 2=зарезервировано).
                col = frame_number + 2

                # Читаем ВСЕ три «persons» строки и сливаем с dedupe,
                # сохраняя порядок: row=8 (под кадр1) первой имеет
                # приоритет, потом 23, потом 38. Так юзер может вписать
                # id в ЛЮБУЮ из них.
                def _merged(rows: tuple[int, ...]) -> list[str]:
                    merged: list[str] = []
                    seen: set[str] = set()
                    for r in rows:
                        for x in _parse_ref_ids(
                            ws.cell(row=r, column=col).value
                        ):
                            if x not in seen:
                                seen.add(x)
                                merged.append(x)
                    return merged

                persons_ids = _merged(_XLSX_ROWS_PERSONS)
                items_ids = _merged(_XLSX_ROWS_ITEMS)
            wb.close()
        except ImportError:
            logger.warning(
                "openpyxl не установлен — не могу прочитать xlsx-рефы"
            )
        except Exception as e:  # noqa: BLE001
            logger.warning(
                "[#{}] frame {}: ошибка чтения xlsx-рефов: {}",
                project.id, frame_number, e,
            )

    chars_dir = project.data_dir / "characters"
    items_dir = project.data_dir / "items"

    refs.extend(
        await _collect_ref_paths(
            session,
            project,
            persons_ids,
            kind="character",
            base_dir=chars_dir,
            frame_number=frame_number,
            max_count=_OUTSEE_MAX_REFS,
        )
    )

    slots_left = _OUTSEE_MAX_REFS - len(refs)
    if slots_left > 0:
        refs.extend(
            await _collect_ref_paths(
                session,
                project,
                items_ids,
                kind="item",
                base_dir=items_dir,
                frame_number=frame_number,
                max_count=slots_left,
            )
        )

    # Постоянный продукт массового (если есть). Подставляем как
    # дополнительный ref, если в кадре остался свободный слот (< 2 рефов).
    # Outsee лимит — 2 ref'а на генерацию, поэтому если уже занято обоими
    # (char + item) — продукт не помещается, оставляем кадр без него.
    meta = getattr(project, "meta", None) or {}
    prod = meta.get("permanent_product") or {}
    prod_ref_path = prod.get("reference_image_path")
    if prod_ref_path and len(refs) < 2:
        prod_path = Path(prod_ref_path)
        if prod_path.exists():
            refs.append(prod_path)
            logger.info(
                "[#{}] frame {} ref продукт '{}' → {} (slot {})",
                project.id, frame_number,
                prod.get("name") or "?", prod_path, len(refs),
            )
        else:
            logger.warning(
                "[#{}] frame {}: продукт-референс {} не найден на диске",
                project.id, frame_number, prod_ref_path,
            )
    elif prod_ref_path and len(refs) >= _OUTSEE_MAX_REFS:
        logger.warning(
            "[#{}] frame {}: у кадра уже {} ref'ов, продукт-референс "
            "не помещается — Outsee лимит. Кадр уйдёт без продукта.",
            project.id, frame_number, _OUTSEE_MAX_REFS,
        )

    return refs[:_OUTSEE_MAX_REFS]


async def run(session: AsyncSession, project: Project, bot: Bot) -> None:
    if project.status is not ProjectStatus.generating_images:
        return
    logger.info("[#{}] generate_images starting", project.id)

    xlsx_path = project.data_dir / "project.xlsx"
    from app.services.xlsx_v8_import import bootstrap_frames_for_image_step

    if not xlsx_path.is_file():
        logger.error(
            "[#{}] generate_images: project.xlsx не найден: {}",
            project.id,
            xlsx_path,
        )
    else:
        boot = await bootstrap_frames_for_image_step(session, project, xlsx_path)
        logger.info(
            "[#{}] generate_images: bootstrap R45={} R46={} created={} "
            "shot1={} shot2={}",
            project.id,
            boot.prompts_in_xlsx,
            boot.shot2_in_xlsx,
            boot.frames_created,
            boot.frames_prompt_updated,
            boot.frames_shot2_updated,
        )

    frames = (
        await session.execute(
            select(Frame).where(Frame.project_id == project.id).order_by(Frame.number)
        )
    ).scalars().all()
    if not frames and xlsx_path.exists():
        await bootstrap_frames_for_image_step(session, project, xlsx_path)
        frames = (
            await session.execute(
                select(Frame)
                .where(Frame.project_id == project.id)
                .order_by(Frame.number)
            )
        ).scalars().all()

    if not frames:
        from app.services.xlsx_v8_import import (
            describe_image_prompts_xlsx_scan,
            read_image_prompts_from_project_xlsx,
        )

        scan = (
            describe_image_prompts_xlsx_scan(xlsx_path)
            if xlsx_path.exists()
            else "project.xlsx не найден"
        )
        n_xlsx = (
            len(read_image_prompts_from_project_xlsx(xlsx_path))
            if xlsx_path.exists()
            else 0
        )
        if n_xlsx:
            raise RuntimeError(
                f"в project.xlsx {n_xlsx} промтов, но кадры в БД не созданы. "
                f"Диагностика: {scan}"
            )
        raise RuntimeError(
            f"нет кадров и нет промтов в project.xlsx. Диагностика: {scan}"
        )

    # Промты с диска → БД (до проверки missing / failed).
    if xlsx_path.is_file():
        from app.services.xlsx_v8_import import (
            apply_image_prompts_from_xlsx_to_frames,
            read_image_prompts_from_project_xlsx,
        )

        n = apply_image_prompts_from_xlsx_to_frames(frames, xlsx_path)
        if n:
            logger.info(
                "[#{}] generate_images: image_prompt с диска xlsx → {} кадров",
                project.id,
                n,
            )
            await session.flush()
            frames = (
                await session.execute(
                    select(Frame)
                    .where(Frame.project_id == project.id)
                    .order_by(Frame.number)
                )
            ).scalars().all()

    xlsx_prompts: dict[int, str] = (
        read_image_prompts_from_project_xlsx(xlsx_path)
        if xlsx_path.is_file()
        else {}
    )

    # Доп. sync v7/v8 (voiceover, animation) — после bootstrap по R45.
    missing_prompts = [
        fr.number
        for fr in frames
        if fr.number in xlsx_prompts
        and is_skippable_empty_prompt(xlsx_prompts.get(fr.number) or "")
    ]
    if missing_prompts and xlsx_path.exists():
        try:
            from app.services.chatgpt_xlsx import sync_project_xlsx

            logger.info(
                "[#{}] generate_images: у {} кадров нет image_prompt после "
                "bootstrap — sync_project_xlsx",
                project.id,
                len(missing_prompts),
            )
            await sync_project_xlsx(session, project, xlsx_path, keep_fields=False)
            await bootstrap_frames_for_image_step(session, project, xlsx_path)
            apply_image_prompts_from_xlsx_to_frames(frames, xlsx_path)
            await session.flush()
            frames = (
                await session.execute(
                    select(Frame)
                    .where(Frame.project_id == project.id)
                    .order_by(Frame.number)
                )
            ).scalars().all()
            xlsx_prompts = read_image_prompts_from_project_xlsx(xlsx_path)
            missing_prompts = [
                fr.number
                for fr in frames
                if fr.number in xlsx_prompts
                and is_skippable_empty_prompt(xlsx_prompts.get(fr.number) or "")
            ]
        except Exception as e:  # noqa: BLE001
            logger.warning(
                "[#{}] generate_images: доп. sync xlsx failed: {}",
                project.id,
                e,
            )

    if missing_prompts:
        logger.warning(
            "[#{}] generate_images: у {} кадров в xlsx R45 заглушка/пусто — "
            "пропускаю (failed): {}",
            project.id,
            len(missing_prompts),
            missing_prompts[:12],
        )
        for fr in frames:
            if fr.number not in missing_prompts:
                continue
            if fr.status not in (
                FrameStatus.image_approved,
                FrameStatus.image_generated,
            ):
                fr.status = FrameStatus.failed
                attrs = dict(fr.attrs or {})
                attrs["fail_reason"] = "no_image_prompt"
                fr.attrs = attrs
        await session.flush()

    out_dir = project.data_dir / "scenes"

    sheet = _sheet_for_project(project)
    try:
        sheet.ensure_frame_columns(len(frames))
    except Exception as e:  # noqa: BLE001
        logger.warning("[#{}] project_sheet ensure_frame_columns failed: {}", project.id, e)

    # Очередь: источник истины — валидный PNG на диске, не статус в БД.
    # Иначе image_generated без файла / без outsee → шаг «завершён», кадры
    # так и не генерировались.
    queued = 0
    for fr in frames:
        if disk_has_valid_frame_image(out_dir, fr.number):
            if fr.status not in (
                FrameStatus.image_approved,
                FrameStatus.image_generated,
            ):
                fr.status = FrameStatus.image_generated
            continue
        bad = newest_frame_image_path(out_dir, fr.number)
        if bad is not None and not is_valid_scene_image(bad):
            logger.warning(
                "[#{}] frame {}: на диске невалидная картинка {} ({} B) — "
                "в outsee",
                project.id,
                fr.number,
                bad.name,
                bad.stat().st_size,
            )
        if not frame_needs_shot1_image(fr, out_dir):
            if (
                fr.number in xlsx_prompts
                and (xlsx_prompts[fr.number] or "").strip()
                and not is_skippable_empty_prompt(xlsx_prompts[fr.number])
                and not disk_has_valid_frame_image(out_dir, fr.number)
            ):
                fr.image_prompt = xlsx_prompts[fr.number]
                fr.status = FrameStatus.image_prompt_ready
                attrs = dict(fr.attrs or {})
                if attrs.pop("fail_reason", None) is not None:
                    fr.attrs = attrs
            else:
                continue
        fr.status = FrameStatus.image_prompt_ready
        queued += 1
    await session.flush()
    logger.info(
        "[#{}] generate_images: в очередь outsee — {} кадров (scenes={})",
        project.id,
        queued,
        out_dir,
    )
    if queued == 0:
        from app.services.scan_frames import scan_missing_frames

        with_prompt = sum(
            1
            for fr in frames
            if not is_skippable_empty_prompt(fr.image_prompt or "")
        )
        missing = await scan_missing_frames(session, project)
        on_disk = sum(
            1 for fr in frames if disk_has_valid_frame_image(out_dir, fr.number)
        )
        logger.warning(
            "[#{}] generate_images: очередь пуста — кадров в БД={}, "
            "с image_prompt={}, валидных PNG на диске={}, без PNG но с промтом={}. "
            "Проверь project.xlsx R45 и «Перечитать xlsx».",
            project.id,
            len(frames),
            with_prompt,
            on_disk,
            missing,
        )
        if with_prompt == 0:
            from app.services.xlsx_v8_import import describe_image_prompts_xlsx_scan

            scan = describe_image_prompts_xlsx_scan(xlsx_path)
            in_xlsx = len(xlsx_prompts)
            skippable = sum(
                1 for p in xlsx_prompts.values() if is_skippable_empty_prompt(p or "")
            )
            raise RuntimeError(
                f"в project.xlsx {in_xlsx} ячеек R45, из них заглушек {skippable}, "
                f"в БД с промтом 0. Диагностика: {scan}"
            )
        if missing and xlsx_path.exists():
            await bootstrap_frames_for_image_step(session, project, xlsx_path)
            await session.flush()
            frames = (
                await session.execute(
                    select(Frame)
                    .where(Frame.project_id == project.id)
                    .order_by(Frame.number)
                )
            ).scalars().all()
            queued = 0
            for fr in frames:
                if disk_has_valid_frame_image(out_dir, fr.number):
                    continue
                if not frame_needs_shot1_image(fr, out_dir):
                    continue
                fr.status = FrameStatus.image_prompt_ready
                queued += 1
            await session.flush()
            logger.info(
                "[#{}] generate_images: повторная очередь после bootstrap — {}",
                project.id,
                queued,
            )
            if queued == 0:
                raise RuntimeError(
                    f"в xlsx есть промты, на диске нет картинок, но в outsee "
                    f"0 кадров (без PNG: {missing}). Проверь scenes/ и статусы кадров"
                )
        if not missing and on_disk >= with_prompt:
            logger.info(
                "[#{}] generate_images: все {} кадров с промтом уже на диске — "
                "outsee не нужен",
                project.id,
                with_prompt,
            )

    async with browser_session() as bs:
        outsee = OutseeBot(bs)
        # `gpt` нужен для GPT-rewrite внутри generate_image_with_retries —
        # после 3 неудачных попыток в outsee он попросит ChatGPT переписать
        # промт без триггеров модерации, потом ещё 3 попытки.
        gpt = ChatGPTBot(bs)
        phase = "shot1"
        shot2_queued = 0
        try:
            while True:
                raise_if_cancelled(project.id)

                if phase == "shot1":
                    await _apply_pending_regens(session, project.id)
                    target = await _next_frame_to_process(
                        session, project.id, out_dir
                    )
                    if target is not None:
                        await _generate_and_send(
                            session, bot, outsee, gpt, project, target, out_dir,
                            shot=1,
                        )
                        continue
                    if await _all_frames_have_image_or_failed(
                        session, project.id, out_dir
                    ):
                        xlsx_path = project.data_dir / "project.xlsx"
                        shot2_queued = await _init_shot2_queue(
                            session, project, frames, out_dir, xlsx_path
                        )
                        if shot2_queued:
                            logger.info(
                                "[#{}] generate_images: фаза shot_02 — {} сцен",
                                project.id, shot2_queued,
                            )
                            phase = "shot2"
                            continue
                        logger.info(
                            "[#{}] generate_images: shot_02 нет — завершаю",
                            project.id,
                        )
                        break
                    pending = await _pending_shot1_numbers(
                        session, project.id, out_dir
                    )
                    if pending:
                        logger.warning(
                            "[#{}] generate_images: нет image_prompt_ready, "
                            "но {} кадров без PNG — жду: {}{}",
                            project.id,
                            len(pending),
                            pending[:8],
                            "…" if len(pending) > 8 else "",
                        )
                    await sleep_cancellable(3.0, project.id)
                    continue

                # phase == "shot2"
                target2 = await _next_shot2_frame_to_process(session, project.id)
                if target2 is not None:
                    ref1 = find_shot1_image(out_dir, target2.number)
                    if ref1 is None:
                        logger.error(
                            "[#{}] frame {} shot_02: нет PNG shot_01 — skip",
                            project.id, target2.number,
                        )
                        attrs = dict(target2.attrs or {})
                        attrs[SHOT2_STATUS_ATTR] = "failed"
                        target2.attrs = attrs
                        await session.flush()
                        continue
                    await _generate_and_send(
                        session, bot, outsee, gpt, project, target2, out_dir,
                        shot=2,
                        shot1_reference=ref1,
                    )
                    continue
                if await _all_shot2_done(session, project.id):
                    break
                await sleep_cancellable(3.0, project.id)
        except StepCancelledError as e:
            consume_stop(project.id)
            # ⏹ Остановить — статус уже откачен обработчиком кнопки в
            # другой сессии. Обновляем наш ORM-объект, чтобы worker'овый
            # commit() не перезаписал откат старым running-статусом.
            # НЕ ставим images_ready.
            logger.info("[#{}] generate_images: {} — выхожу из цикла",
                        project.id, e)
            try:
                await session.refresh(project)
            except Exception:  # noqa: BLE001
                logger.warning("[#{}] не смог refresh project после ⏹", project.id)
            return
        except asyncio.CancelledError:
            logger.info("[#{}] generate_images: hard-cancel (⏹)", project.id)
            try:
                await session.refresh(project)
            except Exception:  # noqa: BLE001
                pass
            raise

    raise_if_cancelled(project.id)
    await session.refresh(project)
    if project.status is not ProjectStatus.generating_images:
        logger.info(
            "[#{}] generate_images: статус уже {} — не ставлю images_ready (⏹?)",
            project.id,
            project.status.value,
        )
        return

    from app.services.post_step_validate import finalize_or_retry

    if not await finalize_or_retry(
        session,
        project,
        step="images",
        ready_status=ProjectStatus.images_ready,
        running_status=ProjectStatus.generating_images,
    ):
        return

    project.status = ProjectStatus.images_ready
    await session.flush()
    logger.info("[#{}] generate_images complete", project.id)


# ---------------------------------------------------------------------------


async def _pending_shot1_numbers(
    session: AsyncSession, project_id: int, out_dir: Path
) -> list[int]:
    frames = (
        await session.execute(
            select(Frame)
            .where(Frame.project_id == project_id)
            .order_by(Frame.number)
        )
    ).scalars().all()
    return [fr.number for fr in frames if frame_needs_shot1_image(fr, out_dir)]


async def _next_frame_to_process(
    session: AsyncSession, project_id: int, out_dir: Path
) -> Frame | None:
    """Следующий кадр для outsee: промт есть, валидного PNG на диске нет."""
    frames = (
        await session.execute(
            select(Frame)
            .where(Frame.project_id == project_id)
            .order_by(Frame.number)
        )
    ).scalars().all()
    for fr in frames:
        if not frame_needs_shot1_image(fr, out_dir):
            continue
        if fr.status is not FrameStatus.image_prompt_ready:
            fr.status = FrameStatus.image_prompt_ready
        return fr
    return None


async def _all_frames_have_image_or_failed(
    session: AsyncSession, project_id: int, out_dir: Path
) -> bool:
    """True когда у каждого кадра с промтом есть валидный PNG или failed."""
    frames = (
        await session.execute(
            select(Frame)
            .where(Frame.project_id == project_id)
            .order_by(Frame.number)
        )
    ).scalars().all()
    for fr in frames:
        if is_skippable_empty_prompt(fr.image_prompt or ""):
            continue
        if fr.status is FrameStatus.failed:
            continue
        if fr.status is FrameStatus.image_approved:
            continue
        if disk_has_valid_frame_image(out_dir, fr.number):
            continue
        return False
    return True


async def _init_shot2_queue(
    session: AsyncSession,
    project: Project,
    frames: list[Frame],
    out_dir: Path,
    xlsx_path: Path,
) -> int:
    """Подготовить очередь shot_02 из xlsx → ``frame.attrs``."""
    by_num = read_shot2_columns(xlsx_path)
    queued = 0
    for fr in frames:
        info = by_num.get(fr.number)
        attrs = dict(fr.attrs or {})
        if info is None or not info.has_shot2:
            if SHOT2_PROMPT_ATTR in attrs or SHOT2_STATUS_ATTR in attrs:
                attrs.pop(SHOT2_PROMPT_ATTR, None)
                attrs.pop(SHOT2_STATUS_ATTR, None)
                fr.attrs = attrs
            continue
        attrs[SHOT2_PROMPT_ATTR] = info.prompt
        if disk_has_shot2_image(out_dir, fr.number):
            attrs[SHOT2_STATUS_ATTR] = "image_generated"
        else:
            attrs[SHOT2_STATUS_ATTR] = "image_prompt_ready"
            queued += 1
        fr.attrs = attrs
    await session.flush()
    return queued


async def _next_shot2_frame_to_process(
    session: AsyncSession, project_id: int
) -> Frame | None:
    frames = (
        await session.execute(
            select(Frame)
            .where(Frame.project_id == project_id)
            .order_by(Frame.number)
        )
    ).scalars().all()
    for fr in frames:
        attrs = fr.attrs or {}
        if attrs.get(SHOT2_STATUS_ATTR) != "image_prompt_ready":
            continue
        if is_skippable_empty_prompt(attrs.get(SHOT2_PROMPT_ATTR) or ""):
            skip_attrs = dict(attrs)
            skip_attrs[SHOT2_STATUS_ATTR] = "skipped"
            fr.attrs = skip_attrs
            await session.flush()
            continue
        return fr
    return None


async def _all_shot2_done(session: AsyncSession, project_id: int) -> bool:
    frames = (
        await session.execute(
            select(Frame)
            .where(Frame.project_id == project_id)
            .order_by(Frame.number)
        )
    ).scalars().all()
    for fr in frames:
        attrs = fr.attrs or {}
        prompt = attrs.get(SHOT2_PROMPT_ATTR) or ""
        if not prompt or is_skippable_empty_prompt(prompt):
            continue
        st = attrs.get(SHOT2_STATUS_ATTR)
        if st not in ("image_generated", "image_approved", "failed", "skipped"):
            return False
    return True


async def _apply_pending_regens(session: AsyncSession, project_id: int) -> None:
    """Находит HITL-решения regenerate/edit_prompt, которые ещё не
    «потреблены», возвращает соответствующие кадры в image_prompt_ready
    и помечает HITL как consumed."""
    hitls = (
        await session.execute(
            select(HITLRequest)
            .where(HITLRequest.project_id == project_id)
            .where(HITLRequest.kind == HITLKind.approve_images)
            .where(
                HITLRequest.decision.in_(
                    [HITLDecision.regenerate, HITLDecision.edit_prompt]
                )
            )
            .order_by(HITLRequest.id.desc())
        )
    ).scalars().all()
    for h in hitls:
        payload = dict(h.payload or {})
        if payload.get("consumed"):
            continue
        if h.frame_id is None:
            payload["consumed"] = True
            h.payload = payload
            continue
        frame = (
            await session.execute(select(Frame).where(Frame.id == h.frame_id))
        ).scalar_one_or_none()
        if frame is None:
            payload["consumed"] = True
            h.payload = payload
            continue
        # Возвращаем кадр в очередь на outsee. Выбор «Повторить» vs
        # заполнение промта делается в _generate_and_send на основе
        # последнего решения пользователя.
        frame.status = FrameStatus.image_prompt_ready
        payload["consumed"] = True
        h.payload = payload
        logger.info(
            "[#{}] frame {}: повторная генерация по решению '{}' (HITL #{})",
            project_id,
            frame.number,
            h.decision.value,
            h.id,
        )
    await session.flush()


async def _generate_and_send(
    session: AsyncSession,
    bot: Bot,
    outsee: OutseeBot,
    gpt: ChatGPTBot,
    project: Project,
    frame: Frame,
    out_dir: Path,
    *,
    shot: int = 1,
    shot1_reference: Path | None = None,
) -> None:
    """Один прогон outsee → сохранение артефакта → HITL-карточка."""
    raise_if_cancelled(project.id)
    is_shot2 = shot == 2
    attrs = dict(frame.attrs or {})
    if is_shot2:
        prompt_text = (attrs.get(SHOT2_PROMPT_ATTR) or "").strip()
    else:
        prompt_text = (frame.image_prompt or "").strip()
    if is_skippable_empty_prompt(prompt_text):
        logger.warning(
            "[#{}] frame {} shot_{}: пустой промт — skip",
            project.id,
            frame.number,
            f"{shot:02d}" if is_shot2 else "01",
        )
        if is_shot2:
            attrs[SHOT2_STATUS_ATTR] = "skipped"
            frame.attrs = attrs
        await session.flush()
        return
    # Проверяем последний HITL: если последнее решение было regenerate —
    # используем кнопку «Повторить» (без перезаполнения промта); иначе —
    # обычная генерация с текущим image_prompt.
    last_hitl = (
        await session.execute(
            select(HITLRequest)
            .where(HITLRequest.frame_id == frame.id)
            .where(HITLRequest.kind == HITLKind.approve_images)
            .order_by(HITLRequest.id.desc())
            .limit(1)
        )
    ).scalar_one_or_none()
    use_regen_button = (
        not is_shot2
        and last_hitl is not None
        and last_hitl.decision is HITLDecision.regenerate
    )

    attempt = (
        await session.execute(
            select(HITLRequest)
            .where(HITLRequest.frame_id == frame.id)
            .where(HITLRequest.kind == HITLKind.approve_images)
        )
    ).scalars().all()
    attempt_number = len(attempt) + 1

    gen_id = uuid.uuid4().hex
    short_uuid = gen_id[:8]
    if is_shot2:
        file_path = out_dir / f"frame_{frame.number:03d}_s2_{short_uuid}.png"
        prompt_id_prefix = build_gen_id_prefix(
            project.id, frame.number, short_uuid
        ) + "-S2"
    else:
        file_path = out_dir / f"frame_{frame.number:03d}_{short_uuid}.png"
        prompt_id_prefix = build_gen_id_prefix(project.id, frame.number, short_uuid)

    full_prompt_len = len(prepend_gen_id(prompt_text, prompt_id_prefix))
    if full_prompt_len > OUTSEE_PROMPT_MAX_CHARS:
        logger.warning(
            "[#{}] frame {}: image_prompt {} симв > outsee {} — "
            "GPT сожмёт перед отправкой",
            project.id,
            frame.number,
            full_prompt_len,
            OUTSEE_PROMPT_MAX_CHARS,
        )

    # Настройки картинки из проекта (с дефолтами).
    img_gen = IMAGE_GENERATORS_BY_ID.get(
        project.image_generator or DEFAULTS["image_generator"]
    )
    ar = ASPECT_RATIOS_BY_ID.get(
        project.aspect_ratio or DEFAULTS["aspect_ratio"]
    )
    ir = IMAGE_RESOLUTIONS_BY_ID.get(
        project.image_resolution or DEFAULTS["image_resolution"]
    )
    aspect_slug = ar.outsee_slug if ar else "9:16"
    model_slug = img_gen.outsee_slug if img_gen else None
    res_slug = ir.outsee_slug if ir else None
    quality_slug = resolve_image_quality_slug(
        project.image_generator, project.image_quality
    )
    logger.info(
        "[#{}] frame {} shot_{} attempt {} gen_id={}: outsee {}",
        project.id,
        frame.number,
        shot,
        attempt_number,
        gen_id[:8],
        "regenerate" if use_regen_button else "generate",
    )
    sheet = _sheet_for_project(project)
    try:
        sheet.write_frame(
            frame.number,
            image_gen_id=gen_id,
            attempt=attempt_number,
            frame_status="image_generating",
            last_error="",
        )
    except Exception as e:  # noqa: BLE001
        logger.warning("[#{}] xlsx write_frame(gen_id) failed: {}", project.id, e)

    # Референсы: shot_02 — всегда PNG shot_01 той же колонки; shot_01 — персонажи/предметы из xlsx.
    if is_shot2:
        refs: list[Path] = [shot1_reference] if shot1_reference else []
    else:
        refs = await _load_refs_for_frame(session, project, frame.number)
    if refs:
        logger.info(
            "[#{}] frame {}: {} ref(ов) подгружено: {}",
            project.id, frame.number, len(refs), [str(r) for r in refs],
        )

    try:
        if use_regen_button:
            try:
                result = await outsee.regenerate_image(
                    file_path, gen_id=gen_id, project_id=project.id
                )
            except OutseeImageError:
                # Если на странице нет предыдущего результата (или другая
                # «структурная» ошибка regenerate) — падаем на полноценный
                # generate с тем же gen_id, чтобы не плодить ложных файлов.
                logger.warning(
                    "[#{}] frame {}: «Повторить» не сработала — падаю на generate",
                    project.id,
                    frame.number,
                )
                result = await generate_image_with_retries(
                    outsee, gpt,
                    prompt=prompt_text,
                    out_path=file_path,
                    max_attempts_per_prompt=3,
                    gpt_rewrite=True,
                    aspect_ratio=aspect_slug,
                    gen_id=gen_id,
                    model_slug=model_slug,
                    resolution=res_slug,
                    quality=quality_slug,
                    relax=bool(project.image_relax),
                    prompt_id_prefix=prompt_id_prefix,
                    reference_image=refs if refs else None,
                    project_id=project.id,
                )
        else:
            # До 3 попыток с исходным image_prompt; если все 3 провалились —
            # GPT-rewrite промта (убирает триггеры модерации) + ещё 3 попытки.
            result = await generate_image_with_retries(
                outsee, gpt,
                prompt=prompt_text,
                out_path=file_path,
                max_attempts_per_prompt=3,
                gpt_rewrite=True,
                aspect_ratio=aspect_slug,
                gen_id=gen_id,
                model_slug=model_slug,
                resolution=res_slug,
                quality=quality_slug,
                relax=bool(project.image_relax),
                prompt_id_prefix=prompt_id_prefix,
                reference_image=refs if refs else None,
                project_id=project.id,
            )
    except StepCancelledError:
        raise
    except OutseeImageError as e:
        # Не «возьму последнюю картинку», не silent retry: помечаем кадр
        # failed и шлём в TG понятное описание ошибки (с gen_id, baseline-ом
        # и тем что нашли). Пайплайн пойдёт к следующему кадру; общая логика
        # анти-зацикливания (MAX_FAIL=3) защитит проект целиком.
        logger.exception(
            "[#{}] frame {}: outsee fail (gen_id={})",
            project.id,
            frame.number,
            gen_id[:8],
        )
        frame.status = FrameStatus.failed
        if is_shot2:
            attrs = dict(frame.attrs or {})
            attrs[SHOT2_STATUS_ATTR] = "failed"
            frame.attrs = attrs
        try:
            sheet.write_frame(
                frame.number,
                frame_status=frame.status.value,
                last_error=e.format_text()[:1500],
            )
        except Exception:  # noqa: BLE001
            pass
        await session.flush()
        try:
            kind = outsee_error_kind_label(outsee_error_kind(e))
            if isinstance(e, OutseeContentRejectedError):
                head = (
                    f"🚫 Кадр #{frame.number} проекта #{project.id}: "
                    f"outsee отклонил промт (модерация). "
                    f"6 попыток + GPT-rewrite не помогли.\n\n"
                )
            else:
                head = (
                    f"⚠️ Кадр #{frame.number} проекта #{project.id}: "
                    f"картинку поймать не удалось ({kind}).\n\n"
                )
            await bot.send_message(
                settings.telegram_owner_chat_id,
                (head + f"<pre>{_html_escape(e.format_text())}</pre>")[:3800],
                parse_mode="HTML",
            )
        except Exception:  # noqa: BLE001
            pass
        await session.commit()
        return

    art = Artifact(
        project_id=project.id,
        frame_id=frame.id,
        kind=ArtifactKind.scene_image,
        uuid=uuid.uuid4().hex,
        path=str(result.file_path),
        meta={
            "gen_id": gen_id,
            "raw_url": result.raw_url or "",
            "shot": shot,
        },
    )
    session.add(art)
    if is_shot2:
        attrs = dict(frame.attrs or {})
        attrs[SHOT2_STATUS_ATTR] = "image_generated"
        frame.attrs = attrs
    else:
        frame.status = FrameStatus.image_generated
    await session.flush()

    try:
        if not is_shot2:
            sheet.write_frame(
                frame.number,
                image_path=str(result.file_path),
                image_url=result.raw_url,
                frame_status=frame.status.value,
            )
    except Exception as e:  # noqa: BLE001
        logger.warning("[#{}] xlsx write_frame(image_path) failed: {}", project.id, e)

    shot_label = f" (shot_0{shot})" if is_shot2 else ""
    caption = (
        f"{prompt_id_prefix}\n"
        f"Кадр #{frame.number}{shot_label} / P{project.id}. Попытка {attempt_number}.\n"
        f"{(frame.voiceover_text or '')[:600]}"
    )
    payload = {
        "step": "image",
        "frame_id": frame.id,
        "attempt": attempt_number,
        "gen_id": gen_id,
        "prompt_id_prefix": prompt_id_prefix,
        "photo_path": str(result.file_path),
        "shot": shot,
    }
    await send_hitl_photo(
        bot,
        session,
        project,
        kind=HITLKind.approve_images,
        photo_path=str(result.file_path),
        caption=caption,
        payload=payload,
        frame_id=frame.id,
        allow_edit=True,
    )
    # Коммитим сразу, чтобы callback-хендлер в другом таске видел HITL.
    await session.commit()


def _html_escape(s: str) -> str:
    import html as _h

    return _h.escape(s)
