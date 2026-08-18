"""GPT текстом через HTTP API (OpenAI-совместимый `/v1/chat/completions`).

Замена браузерного ChatGPT (CDP) для excel_gpt / проверочных нод:
  * без Playwright/Chrome — прямой httpx-запрос к шлюзу (GRSAI / OpenAI-совм.);
  * полная обработка ошибок: 401/403 (ключ), 429 (лимит, ретрай), 5xx/сеть/таймаут
    (ретрай с экспоненциальной паузой), пустой/битый ответ;
  * приложенные файлы (xlsx/txt) сворачиваются в текстовый контекст;
  * скачивание/копирование контента: `download_content` тянет файл по URL,
    `collect_result_urls` достаёт ссылки из ответа модели.

Настройки — `settings.gpt_*` (ключ/база можно переиспользовать из GRSAI_*).
"""

from __future__ import annotations

import asyncio
import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import httpx
from loguru import logger

from app.settings import settings

# Коды, при которых повтор бессмысленен (не ретраим).
_FATAL_STATUS = frozenset({400, 401, 403, 404, 422})
# Коды, при которых повтор имеет смысл (лимиты / временные сбои шлюза).
_RETRY_STATUS = frozenset({408, 409, 425, 429, 500, 502, 503, 504})

_URL_RE = re.compile(r"https?://[^\s\)\]\}<>\"']+", re.IGNORECASE)
_TEXT_SUFFIXES = frozenset({
    ".txt",
    ".md",
    ".csv",
    ".tsv",
    ".json",
    ".yaml",
    ".yml",
    ".xml",
    ".html",
    ".htm",
    ".log",
    ".srt",
    ".vtt",
})
_IMAGE_SUFFIXES = frozenset({".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp"})
# Лимит картинок и размера (base64 раздувает ~4/3) — чтобы не упереться в payload.
_MAX_VISION_IMAGES = 8
_MAX_VISION_BYTES = 4_000_000
# Мелкий бинарь (docx/zip/…) можно отдать как base64 в тексте — НЕ для PDF.
_MAX_BINARY_INLINE_BYTES = 400_000
# PDF в Responses API: multimodal input_file на kie.ai часто → code=500.
# По умолчанию шлём только извлечённый текст (pypdf); file_data — опционально.
_MAX_PDF_INLINE_BYTES = 12_000_000
_MAX_PDF_FILES = 4
_PDF_CONTEXT_MAX_CHARS = 80_000
# kie.ai: полный PDF-текст (~20–40k) на «переведи» → code=500 или hang/timeout.
# Короткие куски (~4–6k) стабильны — режем по страницам.
_PDF_PROVIDER_SAFE_CHARS = 5_500
_PDF_PAGE_HEADER_RE = re.compile(
    r"(?m)^---\s*стр\.\s*\d+/\d+\s*---\s*$"
)
# XLSX → TSV: длинный «Общий план» / «план» легко >280k — раньше резали книгу
# и модель «заполняла» только видимый кусок (~часть строк). Бюджет под GPT-5.x.
_XLSX_CONTEXT_MAX_CHARS = 900_000
_XLSX_DEFAULT_MAX_ROWS = 5_000
_XLSX_PRIORITY_MAX_ROWS = 8_000
# Лист «план»: кадры = столбцы (B..). Раньше max_cols=40 — модель не видела
# кадры 41+ и «не могла» заполнить строку на 100–150 столбцов, хотя writeback
# TSV столбцы не режет. 256 хватает на short/series без раздувания бюджета.
_XLSX_DEFAULT_MAX_COLS = 256
# Листы плана — первыми и почти без обрезки; «план» (кадры) огромный — в хвост.
_XLSX_PRIORITY_SHEET_RE = re.compile(
    r"общий\s*план",
    re.IGNORECASE,
)
# Критичные строки листа «план» — всегда в context pack до остального бюджета.
_XLSX_PINNED_PLAN_ROWS: tuple[int, ...] = (15, 45, 46, 48, 49, 50, 64)


class GptApiError(Exception):
    """Ошибка GPT API с контекстом для логов/повторов."""

    def __init__(self, message: str, *, context: dict[str, Any] | None = None) -> None:
        super().__init__(message)
        self.context = context or {}

    @property
    def retryable(self) -> bool:
        return bool(self.context.get("retryable"))


@dataclass
class GptChatResult:
    text: str
    model: str
    finish_reason: str = ""
    usage: dict[str, Any] = field(default_factory=dict)
    raw: dict[str, Any] = field(default_factory=dict)
    # kie Codex / OpenAI Responses: id вида resp_… (в UI kie = Task ID)
    response_id: str = ""


def gpt_api_enabled() -> bool:
    return bool(settings.gpt_api_enabled)


_PROXY_LOGGED = False
_RELAY_BASE_LOGGED = False


def _mask_proxy_url(url: str) -> str:
    """user:pass@host → user:***@host (для логов)."""
    try:
        if "://" not in url:
            return url
        scheme, rest = url.split("://", 1)
        if "@" not in rest:
            return f"{scheme}://{rest}"
        creds, host = rest.rsplit("@", 1)
        user = creds.split(":", 1)[0] if creds else ""
        return f"{scheme}://{user}:***@{host}"
    except Exception:  # noqa: BLE001
        return "<proxy>"


def _gpt_proxy_url() -> str | None:
    raw = (getattr(settings, "gpt_proxy_url", None) or "").strip()
    if not raw:
        return None
    if raw.startswith("socks5h://"):
        return "socks5://" + raw[len("socks5h://") :]
    return raw


def _async_client(**kwargs: Any) -> httpx.AsyncClient:
    """httpx-клиент; при GPT_PROXY_URL — весь трафик GPT/kie через прокси."""
    global _PROXY_LOGGED
    proxy = _gpt_proxy_url()
    if proxy:
        if proxy.lower().startswith(("socks4://", "socks5://", "socks5h://")):
            try:
                import socksio  # noqa: F401
            except ImportError as e:
                raise GptApiError(
                    "GPT_PROXY_URL=socks… нужен пакет: pip install 'httpx[socks]'",
                    context={"error_kind": "proxy_socks_missing", "retryable": False},
                ) from e
        kwargs.setdefault("proxy", proxy)
        # Явный прокси — не подмешивать системные HTTP(S)_PROXY.
        kwargs.setdefault("trust_env", False)
        if not _PROXY_LOGGED:
            logger.info("GPT API: using proxy {}", _mask_proxy_url(proxy))
            _PROXY_LOGGED = True
    return httpx.AsyncClient(**kwargs)


def _node_vibecode_override() -> bool:
    try:
        from app.services.llm_override import current_override
    except Exception:  # noqa: BLE001
        return False
    ov = current_override()
    return bool(ov and ov.kind == "text" and ov.provider == "vibecode")


def _override_vibecode_base_url() -> str:
    relay = (getattr(settings, "gpt_relay_token", None) or "").strip()
    gbase = (settings.gpt_base_url or "").strip().rstrip("/")
    low = gbase.lower()
    if relay and gbase and "kie.ai" not in low and "vibecode.moe" not in low:
        return gbase
    return (settings.vibecode_base_url or "https://vibecode.moe/v1").strip().rstrip("/")


def _headers() -> dict[str, str]:
    vibe_ov = _node_vibecode_override()
    if vibe_ov:
        key = (settings.vibecode_api_key or "").strip() or settings.gpt_api_effective_key
    else:
        key = settings.gpt_api_effective_key
    if not key:
        if vibe_ov or settings.text_llm_is_vibecode:
            raise GptApiError(
                "VIBECODE_API_KEY пуст — задай ключ vibecode.moe (vk-…) в .env",
                context={"error_kind": "no_key", "provider": "vibecode"},
            )
        if settings.text_llm_is_tokenrouter:
            raise GptApiError(
                "TOKENROUTER_API_KEY пуст — задай ключ TokenRouter (Kimi K3) в .env",
                context={"error_kind": "no_key", "provider": "tokenrouter"},
            )
        raise GptApiError(
            "GPT_API_KEY пуст (и GRSAI_API_KEY тоже) — задай ключ в .env",
            context={"error_kind": "no_key", "provider": "kie"},
        )
    headers = {
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
    }
    relay = (getattr(settings, "gpt_relay_token", None) or "").strip()
    if relay:
        headers["X-VP-Relay-Token"] = relay
    return headers


def _chat_url(model: str) -> str:
    global _RELAY_BASE_LOGGED
    base = (
        _override_vibecode_base_url()
        if _node_vibecode_override()
        else settings.gpt_api_effective_base_url
    )
    if not base:
        raise GptApiError(
            "База текстового LLM пуста — задай TOKENROUTER_BASE_URL, VIBECODE_BASE_URL или GPT_BASE_URL",
            context={"error_kind": "no_base"},
        )
    if not _RELAY_BASE_LOGGED and "kie.ai" not in base.lower():
        logger.info(
            "GPT API: base_url={} (VPS-relay / non-kie host)",
            base,
        )
        _RELAY_BASE_LOGGED = True
    if _node_vibecode_override():
        path = "/chat/completions" if base.lower().endswith("/v1") else "/v1/chat/completions"
    else:
        path = (settings.gpt_chat_path_effective or "/v1/chat/completions").strip()
    if not path.startswith("/"):
        path = "/" + path
    # kie.ai: путь зависит от модели (/{model}/v1/chat/completions).
    if "{model}" in path:
        path = path.replace("{model}", model)
    return f"{base}{path}"


# ─────────────────────────── файлы → контекст ───────────────────────────


def _xlsx_sheet_priority(title: str) -> int:
    """Меньше = раньше в контексте. «Общий план» важнее огромного «план»."""
    t = (title or "").strip()
    if _XLSX_PRIORITY_SHEET_RE.search(t):
        return 0
    low = t.casefold()
    if low in {"сценарий", "script", "персонажи", "герой"}:
        return 1
    if low == "план" or low.startswith("план "):
        return 9
    return 5


def _is_plan_sheet_title(title: str) -> bool:
    low = (title or "").strip().casefold()
    return low == "план" or low.startswith("план ")


def _row_to_tsv_line(excel_row: int, cells: list[str]) -> str:
    return "\t".join([f"@row={excel_row}", *cells])


def xlsx_to_text(
    path: Path,
    *,
    max_rows: int = _XLSX_DEFAULT_MAX_ROWS,
    max_cols: int = _XLSX_DEFAULT_MAX_COLS,
    max_chars: int | None = None,
    write_contract: str = "tsv",
) -> str:
    """Свернуть xlsx в читаемый TSV-контекст (непустые строки листов).

    Приоритетные листы («Общий план») идут первыми.
    На листе «план» сначала пинятся R15/R45–50/R48/R64, затем остаток бюджета.
    ``write_contract``:
      - ``tsv`` — check/fix: правки через `# Лист:` (legacy);
      - ``apply_ops`` — DB SoT: запрет TSV, ответ только JSON apply-ops.
    """
    budget = _XLSX_CONTEXT_MAX_CHARS if max_chars is None else max(4_000, int(max_chars))
    try:
        from openpyxl import load_workbook
    except Exception as e:  # noqa: BLE001
        return f"[xlsx: openpyxl недоступен: {e}]"
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        return f"[xlsx: не удалось открыть {path.name}: {e}]"
    sheets = sorted(list(wb.worksheets), key=lambda ws: (_xlsx_sheet_priority(ws.title), ws.title))
    apply_ops_mode = str(write_contract or "").strip().lower() == "apply_ops"
    if apply_ops_mode:
        # Не писать «бинарный / недоступен / нет файла» — модель копирует это в отказ.
        banner = (
            "[DB SoT READ SNAPSHOT] Ниже текстовый снимок зеркала Excel. "
            "Источник правды для записи — frame_uuid + attrs (см. db_frames.json). "
            "Ответ = ТОЛЬКО JSON apply-ops "
            '{"ops":[...],"characters":[...],"scenes":[...]} без markdown. '
            "Не возвращай TSV и не требуй вложений: пайплайн сам запишет JSON в DB.]"
        )
    else:
        banner = (
            "[xlsx text-export: это текстовый снимок книги для API; "
            "бинарный .xlsx в песочнице модели недоступен — это норма. "
            "Обрезка по лимиту контекста ≠ отсутствие файла. "
            "Проверяй и правь по TSV ниже; для записи верни `# Лист:` блоки. "
            "Каждая строка начинается с `@row=<номер Excel>` — сохрани префикс, "
            "чтобы запись попала в ту же строку шаблона (не уплотняй пустые ряды).]"
        )
    out: list[str] = [banner]
    used = len(out[0])
    truncated_sheets: list[str] = []
    for ws in sheets:
        # apply_ops: не использовать `# Лист:` — иначе модель уходит в TSV/отказ.
        header = (
            f"[SHEET: {ws.title}]"
            if apply_ops_mode
            else f"# Лист: {ws.title}"
        )
        if used + len(header) + 32 > budget:
            truncated_sheets.append(ws.title)
            skip = (
                f"[SHEET: {ws.title}]\n… (лист пропущен: лимит контекста)"
                if apply_ops_mode
                else f"# Лист: {ws.title}\n… (лист пропущен: лимит контекста)"
            )
            out.append(skip)
            continue
        sheet_lines = [header]
        sheet_used = len(header)
        rows_written = 0
        row_cap = (
            _XLSX_PRIORITY_MAX_ROWS
            if _XLSX_PRIORITY_SHEET_RE.search(ws.title or "")
            else max_rows
        )
        sheet_truncated = False
        pin_plan = _is_plan_sheet_title(ws.title or "")
        pinned_set = set(_XLSX_PINNED_PLAN_ROWS) if pin_plan else set()
        emitted_rows: set[int] = set()

        # Materialize rows once (read_only iterator is single-pass).
        all_rows: list[tuple[int, list[str]]] = []
        for excel_row, row in enumerate(ws.iter_rows(values_only=True), start=1):
            cells = ["" if c is None else str(c) for c in row[:max_cols]]
            if not any(c.strip() for c in cells):
                continue
            all_rows.append((excel_row, cells))

        def _append_line(excel_row: int, cells: list[str]) -> bool:
            """Return False if budget exhausted (caller marks truncated)."""
            nonlocal sheet_used, rows_written, sheet_truncated
            if excel_row in emitted_rows:
                return True
            line = _row_to_tsv_line(excel_row, cells)
            if used + sheet_used + len(line) + 1 > budget:
                sheet_lines.append("… (обрезано: лимит контекста — файл на диске есть)")
                sheet_truncated = True
                return False
            if rows_written >= row_cap:
                sheet_lines.append("… (обрезано: лимит строк листа — файл на диске есть)")
                sheet_truncated = True
                return False
            sheet_lines.append(line)
            sheet_used += len(line) + 1
            rows_written += 1
            emitted_rows.add(excel_row)
            return True

        # 1) Priority pack: pinned plan rows first
        if pin_plan:
            by_num = {r: c for r, c in all_rows}
            for pr in _XLSX_PINNED_PLAN_ROWS:
                if pr in by_num:
                    if not _append_line(pr, by_num[pr]):
                        break

        # 2) Remaining nonempty rows
        if not sheet_truncated:
            for excel_row, cells in all_rows:
                if excel_row in emitted_rows:
                    continue
                # Skip non-pinned until pins done (already); just fill rest
                if not _append_line(excel_row, cells):
                    break

        block = "\n".join(sheet_lines)
        out.append(block)
        used += len(block) + 1
        if sheet_truncated:
            truncated_sheets.append(ws.title)
        # silence unused
        del pinned_set
    import contextlib

    with contextlib.suppress(Exception):
        wb.close()
    if truncated_sheets:
        if apply_ops_mode:
            out.append(
                "[DB SoT READ SNAPSHOT: частично обрезаны листы: "
                + ", ".join(truncated_sheets)
                + " — пиши JSON apply-ops по видимым кадрам/db_frames.json]"
            )
        else:
            out.append(
                "[xlsx text-export: частично обрезаны листы (это НЕ «нет project.xlsx»): "
                + ", ".join(truncated_sheets)
                + " — ЗАПРЕЩЕНО считать книгу полной; запроси продолжение экспорта "
                "или заполни только видимые @row, пометив remaining: truncated]"
            )
        logger.warning(
            "xlsx_to_text: обрезаны листы {} path={} chars~{}",
            truncated_sheets,
            path.name,
            used,
        )
    return "\n".join(out).strip() or "[xlsx: пустой]"


def is_pdf_path(path: Path) -> bool:
    return path.suffix.lower() == ".pdf"


def pdf_to_text(path: Path, *, max_chars: int = _PDF_CONTEXT_MAX_CHARS) -> str:
    """Извлечь текст из PDF (pypdf). Без сырого бинаря / base64 в тексте."""
    try:
        from pypdf import PdfReader
    except ImportError:
        return (
            f"[pdf {path.name}: нужен пакет pypdf — "
            f"pip install pypdf; текст не извлечён]"
        )
    try:
        reader = PdfReader(str(path))
    except Exception as e:  # noqa: BLE001
        return f"[pdf {path.name}: не открыт ({e})]"

    pages_out: list[str] = []
    total = 0
    n_pages = len(reader.pages)
    for i, page in enumerate(reader.pages, start=1):
        try:
            chunk = (page.extract_text() or "").strip()
        except Exception as e:  # noqa: BLE001
            chunk = f"[стр.{i}: ошибка извлечения: {e}]"
        if not chunk:
            continue
        header = f"--- стр. {i}/{n_pages} ---"
        block = f"{header}\n{chunk}"
        if total + len(block) + 2 > max_chars:
            remain = max_chars - total - len(header) - 20
            if remain > 80:
                pages_out.append(f"{header}\n{chunk[:remain]}\n… (обрезано)")
            else:
                pages_out.append(f"… (ещё страницы обрезаны, всего {n_pages})")
            break
        pages_out.append(block)
        total += len(block) + 2

    if not pages_out:
        return (
            f"[pdf {path.name}: {n_pages} стр., текста почти нет "
            f"(скан/картинки) — смотри multimodal input_file если приложен]"
        )
    return "\n\n".join(pages_out)


def split_pdf_text_chunks(
    text: str,
    *,
    max_chars: int = _PDF_PROVIDER_SAFE_CHARS,
) -> list[str]:
    """Разрезать вывод pdf_to_text на куски ≤ max_chars (по границам страниц)."""
    raw = (text or "").strip()
    if not raw:
        return []
    budget = max(1_500, int(max_chars))
    if len(raw) <= budget:
        return [raw]

    pages: list[str] = []
    matches = list(_PDF_PAGE_HEADER_RE.finditer(raw))
    if matches:
        for i, m in enumerate(matches):
            start = m.start()
            end = matches[i + 1].start() if i + 1 < len(matches) else len(raw)
            page = raw[start:end].strip()
            if page:
                pages.append(page)
    else:
        pages = [raw]

    chunks: list[str] = []
    buf = ""
    for page in pages:
        if len(page) > budget:
            if buf:
                chunks.append(buf)
                buf = ""
            for i in range(0, len(page), budget):
                piece = page[i : i + budget].strip()
                if piece:
                    chunks.append(piece)
            continue
        candidate = f"{buf}\n\n{page}".strip() if buf else page
        if len(candidate) <= budget:
            buf = candidate
            continue
        if buf:
            chunks.append(buf)
        buf = page
    if buf:
        chunks.append(buf)
    return chunks or [raw[:budget]]


def pdf_paths_need_chunking(
    paths: list[Path] | None,
    *,
    threshold: int = _PDF_PROVIDER_SAFE_CHARS,
) -> bool:
    """True если суммарный текст PDF больше безопасного лимита провайдера."""
    pdfs = [p for p in (paths or []) if is_pdf_path(p) and p.is_file()]
    if not pdfs:
        return False
    total = 0
    for pdf in pdfs:
        total += len(pdf_to_text(pdf))
        if total > threshold:
            return True
    return False


def pdf_to_input_file_part(path: Path) -> dict[str, Any] | None:
    """Responses API part: input_file + file_data (data-URL). None если слишком большой."""
    import base64

    try:
        size = path.stat().st_size
    except OSError:
        return None
    if size < 32 or size > _MAX_PDF_INLINE_BYTES:
        logger.warning(
            "gpt_api: pdf {} size={} — skip input_file (limit={})",
            path.name,
            size,
            _MAX_PDF_INLINE_BYTES,
        )
        return None
    try:
        raw = path.read_bytes()
    except OSError as e:
        logger.warning("gpt_api: pdf {} read failed: {}", path.name, e)
        return None
    b64 = base64.b64encode(raw).decode("ascii")
    return {
        "type": "input_file",
        "filename": path.name,
        "file_data": f"data:application/pdf;base64,{b64}",
    }


def file_to_context(
    path: Path,
    *,
    max_chars: int = 60_000,
    xlsx_write_contract: str = "tsv",
) -> str:
    """Текстовое представление файла для вложения в prompt (как attach в ChatGPT)."""
    import base64

    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm", ".xls"):
        # Для книг — отдельный большой бюджет; max_chars не режем до 60k.
        xlsx_budget = max(max_chars, _XLSX_CONTEXT_MAX_CHARS)
        body = xlsx_to_text(
            path, max_chars=xlsx_budget, write_contract=xlsx_write_contract
        )
    elif suffix in _TEXT_SUFFIXES:
        try:
            body = path.read_text(encoding="utf-8", errors="replace")
        except Exception as e:  # noqa: BLE001
            body = f"[не удалось прочитать {path.name}: {e}]"
    elif suffix in _IMAGE_SUFFIXES:
        try:
            size = path.stat().st_size
        except OSError:
            size = -1
        return f"[изображение {path.name}: {suffix}, {size} байт — см. vision-вложение]"
    elif suffix == ".pdf":
        # Раньше: printable из бинаря + data:application/pdf;base64 в тексте →
        # kie Responses отдавал code=500 Server exception. Текст — через pypdf;
        # сам файл в Responses — input_file (см. build_input).
        try:
            size = path.stat().st_size
        except OSError:
            size = -1
        extracted = pdf_to_text(
            path, max_chars=max(max_chars, _PDF_CONTEXT_MAX_CHARS)
        )
        body = f"[pdf {path.name}: {size} байт]\n{extracted}"
    elif suffix in {".docx"}:
        # docx = zip+xml; вытащим word/document.xml текст грубо.
        try:
            import zipfile
            from xml.etree import ElementTree as ET

            with zipfile.ZipFile(path) as zf:
                xml = zf.read("word/document.xml")
            root = ET.fromstring(xml)
            texts = [
                (n.text or "")
                for n in root.iter()
                if n.tag.endswith("}t") and (n.text or "").strip()
            ]
            body = "\n".join(texts) or f"[docx {path.name}: пустой текст]"
        except Exception as e:  # noqa: BLE001
            body = f"[docx {path.name}: не разобрал ({e})]"
    else:
        # Прочий бинарь — метаданные; мелкий файл инлайним base64 (модель может вернуть).
        try:
            raw = path.read_bytes()
            size = len(raw)
        except OSError as e:
            return f"[файл {path.name}: не прочитан ({e})]"
        if size <= _MAX_BINARY_INLINE_BYTES:
            mime = {
                ".zip": "application/zip",
                ".mp4": "video/mp4",
                ".webm": "video/webm",
            }.get(suffix, "application/octet-stream")
            body = (
                f"[файл {path.name}: {suffix or 'bin'}, {size} байт]\n"
                f"data:{mime};base64,{base64.b64encode(raw).decode('ascii')}"
            )
        else:
            return (
                f"[файл {path.name}: {suffix or 'bin'}, {size} байт — "
                f"слишком большой для inline (>{_MAX_BINARY_INLINE_BYTES}), "
                f"только имя; скачай исходник кнопкой ↓ в Studio]"
            )
    # XLSX уже укладывается в свой бюджет внутри xlsx_to_text — не резать до 60k.
    # JSON батчи img_pr (db_frames_*.json) тоже не должны молча обрезаться:
    # модель тогда отказывается («файл обрезан после кадра N»).
    soft_cap = (
        max(max_chars, _XLSX_CONTEXT_MAX_CHARS)
        if suffix in (".xlsx", ".xlsm", ".xls")
        else max(max_chars, _PDF_CONTEXT_MAX_CHARS)
        if suffix == ".pdf"
        else max(max_chars, 900_000)
        if suffix == ".json"
        else max_chars
    )
    if len(body) > soft_cap and "base64," not in body:
        body = body[:soft_cap] + "\n… (обрезано)"
    return f"===== {path.name} =====\n{body}"


def is_image_path(path: Path) -> bool:
    return path.suffix.lower() in _IMAGE_SUFFIXES


def _compress_image_for_vision(path: Path, *, max_bytes: int = _MAX_VISION_BYTES) -> tuple[bytes, str]:
    """Ужать картинку до max_bytes (JPEG). Возвращает (bytes, mime)."""
    import io

    from PIL import Image

    with Image.open(path) as src:
        img = src.convert("RGB")
        for max_side in (2048, 1536, 1280, 1024, 768, 640, 512):
            w, h = img.size
            scale = min(1.0, float(max_side) / float(max(w, h)))
            work = img
            if scale < 1.0:
                work = img.resize(
                    (max(1, int(w * scale)), max(1, int(h * scale))),
                    Image.Resampling.LANCZOS,
                )
            for quality in (85, 75, 65, 55, 45, 35):
                buf = io.BytesIO()
                work.save(buf, format="JPEG", quality=quality, optimize=True)
                data = buf.getvalue()
                if len(data) <= max_bytes:
                    if work is not img:
                        work.close()
                    return data, "image/jpeg"
            if work is not img:
                work.close()
    raise GptApiError(
        f"картинка слишком большая для vision даже после сжатия: {path.name}",
        context={"error_kind": "media_too_large", "retryable": False},
    )


def image_to_data_url(path: Path) -> str:
    """data:image/...;base64,... для multimodal / vision."""
    import base64

    suffix = path.suffix.lower()
    mime = {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".webp": "image/webp",
        ".gif": "image/gif",
    }.get(suffix, "application/octet-stream")
    data = path.read_bytes()
    if len(data) > _MAX_VISION_BYTES:
        orig = len(data)
        data, mime = _compress_image_for_vision(path)
        logger.info(
            "gpt_api vision compress {}: {} → {} bytes ({})",
            path.name,
            orig,
            len(data),
            mime,
        )
    b64 = base64.b64encode(data).decode("ascii")
    return f"data:{mime};base64,{b64}"


def split_input_paths(
    input_paths: list[Path] | None,
) -> tuple[list[Path], list[Path]]:
    """(text_or_other_files, image_files) с лимитом vision."""
    files = [p for p in (input_paths or []) if p and p.is_file()]
    images = [p for p in files if is_image_path(p)][:_MAX_VISION_IMAGES]
    others = [p for p in files if p not in images]
    return others, images


def is_responses_mode() -> bool:
    """API формата Responses (input/output) вместо chat/completions?"""
    if _node_vibecode_override():
        return False
    mode = (settings.gpt_api_mode_effective or "auto").strip().lower()
    if mode == "responses":
        return True
    if mode == "chat":
        return False
    return "responses" in (settings.gpt_chat_path_effective or "").lower()


def _compose_user_text(
    *,
    prompt: str,
    accompanying: str = "",
    text_paths: list[Path] | None = None,
    image_names: list[str] | None = None,
    xlsx_write_contract: str = "tsv",
) -> str:
    parts: list[str] = []
    if (prompt or "").strip():
        parts.append(prompt.strip())
    if (accompanying or "").strip():
        parts.append("## Сопроводительный текст\n" + accompanying.strip())
    files = list(text_paths or [])
    if files:
        parts.append("## Приложенные файлы")
        for p in files:
            parts.append(
                file_to_context(p, xlsx_write_contract=xlsx_write_contract)
            )
    if image_names:
        parts.append(
            "## Изображения (vision)\n"
            + "\n".join(f"- {n}" for n in image_names)
        )
    return "\n\n".join(parts) or "(пусто)"


def normalize_history(
    history: list[dict[str, Any]] | None,
    *,
    max_messages: int = 40,
    max_chars: int = 120_000,
) -> list[dict[str, str]]:
    """Оставить user/assistant реплики для multi-turn (хвост, лимит символов)."""
    if not history:
        return []
    cleaned: list[dict[str, str]] = []
    for raw in history:
        if not isinstance(raw, dict):
            continue
        role = str(raw.get("role") or "").strip().lower()
        if role not in ("user", "assistant"):
            continue
        content = raw.get("content")
        if isinstance(content, list):
            # multimodal leftover — берём только текст
            text = "".join(
                str(p.get("text") or "")
                for p in content
                if isinstance(p, dict)
            ).strip()
        else:
            text = str(content or "").strip()
        if not text:
            continue
        cleaned.append({"role": role, "content": text})
    if max_messages > 0 and len(cleaned) > max_messages:
        cleaned = cleaned[-max_messages:]
    # с хвоста: уложиться в max_chars, выкидывая самые старые
    if max_chars > 0:
        total = 0
        kept_rev: list[dict[str, str]] = []
        for item in reversed(cleaned):
            n = len(item["content"])
            if kept_rev and total + n > max_chars:
                break
            kept_rev.append(item)
            total += n
        cleaned = list(reversed(kept_rev))
    return cleaned


def build_input(
    *,
    prompt: str,
    accompanying: str = "",
    input_paths: list[Path] | None = None,
    system: str | None = None,
    history: list[dict[str, Any]] | None = None,
    xlsx_write_contract: str = "tsv",
) -> list[dict[str, Any]]:
    """Ввод для Responses API: всегда input[] (kie.ai codex падает на голой строке).

    Голый ``input: \"text\"`` → HTTP 200 + ``{code:500, msg: Server exception}``.
    Нужен ``[{role, content}]`` или multimodal content[].
    PDF: текст через pypdf в input_text + файл как input_file (file_data),
    не data:application/pdf;base64 внутри текста (тоже даёт code=500).
    """
    prior = normalize_history(history)
    others, images = split_input_paths(input_paths)
    pdfs = [p for p in others if is_pdf_path(p)][:_MAX_PDF_FILES]
    text_paths = [p for p in others if not is_pdf_path(p)] + pdfs
    text = _compose_user_text(
        prompt=prompt,
        accompanying=accompanying,
        text_paths=text_paths,
        image_names=[p.name for p in images],
        xlsx_write_contract=xlsx_write_contract,
    )
    if system:
        text = f"[Инструкция]\n{system}\n\n{text}"

    items: list[dict[str, Any]] = [
        {"role": m["role"], "content": m["content"]} for m in prior
    ]

    file_parts: list[dict[str, Any]] = []
    # kie.ai: input_file(PDF) → code=500 Server exception. Текст уже в input_text.
    # Включать file_data только явно: GPT_PDF_INPUT_FILE=1
    attach_pdf_file = str(
        getattr(settings, "gpt_pdf_input_file", None)
        or ""
    ).strip().lower() in {"1", "true", "yes", "on"}
    if not attach_pdf_file:
        try:
            from os import environ

            attach_pdf_file = environ.get("GPT_PDF_INPUT_FILE", "").strip().lower() in {
                "1",
                "true",
                "yes",
                "on",
            }
        except Exception:  # noqa: BLE001
            attach_pdf_file = False
    if attach_pdf_file:
        for pdf in pdfs:
            part = pdf_to_input_file_part(pdf)
            if part is not None:
                file_parts.append(part)
                logger.info(
                    "gpt_api: attach pdf as input_file name={} bytes~{}",
                    pdf.name,
                    pdf.stat().st_size if pdf.is_file() else "?",
                )
    elif pdfs:
        logger.info(
            "gpt_api: pdf as text-only (no input_file) files={}",
            [p.name for p in pdfs],
        )

    if not images and not file_parts:
        items.append({"role": "user", "content": text})
        return items

    content: list[dict[str, Any]] = [{"type": "input_text", "text": text}]
    content.extend(file_parts)
    for img in images:
        try:
            content.append(
                {"type": "input_image", "image_url": image_to_data_url(img)}
            )
        except GptApiError as e:
            logger.warning("gpt_api vision skip {}: {}", img.name, e)
            content.append(
                {
                    "type": "input_text",
                    "text": f"[не удалось вложить изображение {img.name}: {e}]",
                }
            )
    items.append({"role": "user", "content": content})
    return items


def build_messages(
    *,
    prompt: str,
    accompanying: str = "",
    input_paths: list[Path] | None = None,
    system: str | None = None,
    history: list[dict[str, Any]] | None = None,
    xlsx_write_contract: str = "tsv",
) -> list[dict[str, Any]]:
    """Собрать messages для chat/completions (текст + optional vision + история)."""
    messages: list[dict[str, Any]] = []
    if system:
        messages.append({"role": "system", "content": system})
    for m in normalize_history(history):
        messages.append({"role": m["role"], "content": m["content"]})

    others, images = split_input_paths(input_paths)
    text = _compose_user_text(
        prompt=prompt,
        accompanying=accompanying,
        text_paths=others,
        image_names=[p.name for p in images],
        xlsx_write_contract=xlsx_write_contract,
    )
    if not images:
        messages.append({"role": "user", "content": text})
        return messages

    content: list[dict[str, Any]] = [{"type": "text", "text": text}]
    for img in images:
        try:
            content.append(
                {
                    "type": "image_url",
                    "image_url": {"url": image_to_data_url(img)},
                }
            )
        except GptApiError as e:
            logger.warning("gpt_api vision skip {}: {}", img.name, e)
            content.append(
                {
                    "type": "text",
                    "text": f"[не удалось вложить изображение {img.name}: {e}]",
                }
            )
    messages.append({"role": "user", "content": content})
    return messages


# ─────────────────────────── HTTP вызов ───────────────────────────


def _check_provider_envelope(payload: dict[str, Any]) -> None:
    """Некоторые шлюзы (kie.ai) отдают ошибку как HTTP 200 с {code,msg,data}.

    Если code не успешный и нет choices — поднимаем понятную ошибку.
    """
    if not isinstance(payload, dict) or payload.get("choices"):
        return
    code = payload.get("code")
    if code is None:
        return
    try:
        code_int = int(code)
    except (TypeError, ValueError):
        return
    if code_int in (0, 200):
        return
    msg = str(payload.get("msg") or payload.get("message") or "ошибка провайдера")
    retryable = code_int in _RETRY_STATUS or code_int >= 500
    hint = ""
    low = msg.lower()
    if "not authorized" in low or "apikey" in low or code_int in (401, 403):
        hint = " — ключ не авторизован на эту модель (проверь GPT_API_KEY/модель у провайдера)"
    raise GptApiError(
        f"GPT провайдер code={code_int}: {msg}{hint}",
        context={"provider_code": code_int, "retryable": retryable},
    )


def _parse_responses(payload: dict[str, Any]) -> tuple[str, str]:
    """(text, status) из ответа Responses API (kie.ai codex, gpt-5.6)."""
    _check_provider_envelope(payload)
    status = str(payload.get("status") or "")
    text_parts: list[str] = []
    output = payload.get("output")
    if isinstance(output, list):
        for item in output:
            if not isinstance(item, dict):
                continue
            content = item.get("content")
            if isinstance(content, list):
                for c in content:
                    if (
                        isinstance(c, dict)
                        and c.get("type") in ("output_text", "text")
                        and c.get("text")
                    ):
                        text_parts.append(str(c["text"]))
    text = "".join(text_parts).strip()
    if not text:
        ot = payload.get("output_text")
        if isinstance(ot, str):
            text = ot.strip()
    if not text:
        raise GptApiError(
            "GPT(responses): пустой output",
            context={"status": status, "payload": payload},
        )
    return text, status


def _http_timeout(total_s: float) -> httpx.Timeout:
    """Общий таймаут + короткий connect (не ждать connect = GPT_TIMEOUT)."""
    total = max(1.0, float(total_s))
    return httpx.Timeout(total, connect=min(30.0, total))


def _stream_timeout(total_s: float) -> httpx.Timeout:
    """Таймаут для SSE: без ``total`` (он рвал бы весь длинный ответ целиком).

    Важно: ``read`` НЕ берём из GPT_TIMEOUT / scene_design 280s.
    Иначе httpx ReadTimeout сам закрывает сокет, пока модель ещё
    reasoning'ит без дельт — на kie задача потом success, у нас обрывок.

    По умолчанию ``GPT_STREAM_READ_TIMEOUT_S=0`` → read=None: ждём, пока
    сервер сам закроет SSE (completed / EOF). >0 — только аварийный idle.
    """
    configured = float(getattr(settings, "gpt_stream_read_timeout_s", 0.0) or 0.0)
    if configured > 0:
        read: float | None = max(configured, 120.0)
    else:
        read = None
    # upload большого input (xlsx/db_frames) тоже не режем 60с.
    write: float | None = None
    _ = total_s  # оставлен для совместимости вызова; на SSE idle не влияет
    return httpx.Timeout(None, connect=30.0, read=read, write=write, pool=30.0)


def _raise_http_status(
    status: int,
    body_text: str,
    *,
    use_model: str,
) -> None:
    """Общая обработка HTTP-кодов GPT (fatal / retryable)."""
    if status in _FATAL_STATUS:
        low = body_text.lower()
        hint = ""
        if "apikey error" in low:
            hint = (
                f" — ключ не авторизован на модель {use_model!r} "
                "(добавь модель в whitelist ключа grsai)"
            )
        elif "model not register" in low or "not register" in low:
            hint = f" — модель {use_model!r} не существует у провайдера (проверь GPT_MODEL)"
        raise GptApiError(
            f"GPT HTTP {status}: {body_text[:400]}{hint}",
            context={"status_code": status, "retryable": False, "model": use_model},
        )
    if status in _RETRY_STATUS or status >= 500:
        raise GptApiError(
            f"GPT HTTP {status}: {body_text[:300]}",
            context={"status_code": status, "retryable": True, "model": use_model},
        )
    if status >= 400:
        raise GptApiError(
            f"GPT HTTP {status}: {body_text[:400]}",
            context={"status_code": status, "retryable": False, "model": use_model},
        )


def parse_responses_sse_lines(
    lines: list[str],
) -> tuple[str, str, dict[str, Any], str]:
    """Собрать (text, status, final_payload, response_id) из SSE lines kie Codex.

    Важно: при ``stream:false`` Cloudflare/прокси рвёт длинный ответ
    (у kie задача уже success). SSE держит соединение дельтами.

    Если stream оборвался после дельт / ``output_text.done``, но до
    целого ``response.completed`` — берём уже накопленный текст (salvage).
    """
    delta_parts: list[str] = []
    done_text = ""
    response_id = ""
    final_response: dict[str, Any] | None = None
    status = ""
    event_counts: dict[str, int] = {}
    bad_json_n = 0

    for raw in lines:
        line = (raw or "").strip()
        if not line or line.startswith("event:"):
            continue
        if not line.startswith("data:"):
            continue
        data_str = line[5:].strip()
        if not data_str or data_str == "[DONE]":
            continue
        try:
            event = json.loads(data_str)
        except json.JSONDecodeError:
            bad_json_n += 1
            continue
        if not isinstance(event, dict):
            continue
        etype = str(event.get("type") or "")
        if etype:
            event_counts[etype] = event_counts.get(etype, 0) + 1
        if etype in {"response.created", "response.in_progress"}:
            resp = event.get("response")
            if isinstance(resp, dict) and resp.get("id"):
                response_id = str(resp["id"])
        elif etype == "response.output_text.delta":
            delta = event.get("delta")
            if delta:
                delta_parts.append(str(delta))
        elif etype == "response.output_text.done":
            # Полный текст куска — страховка, если completed обрезан CF.
            t = event.get("text")
            if isinstance(t, str) and t.strip():
                done_text = t
        elif etype == "response.completed":
            resp = event.get("response")
            if isinstance(resp, dict):
                final_response = resp
                if resp.get("id"):
                    response_id = str(resp["id"])
                status = str(resp.get("status") or "completed")

    deltas_joined = "".join(delta_parts).strip()
    completed_text = ""
    if final_response is not None:
        try:
            completed_text, status = _parse_responses(final_response)
        except GptApiError:
            status = status or str(final_response.get("status") or "")

    # Берём самый полный источник: completed часто рвётся на длинных файлах,
    # а дельты / done уже целые.
    candidates = [completed_text, done_text.strip(), deltas_joined]
    text = max(candidates, key=lambda s: len(s or "")).strip()

    if not text:
        raise GptApiError(
            "GPT(responses/stream): пустой output",
            context={
                "retryable": True,
                "response_id": response_id,
                "error_kind": "empty_stream",
                "sse_lines": len(lines),
                "sse_bad_json": bad_json_n,
                "sse_events": event_counts,
                "delta_chars": len(deltas_joined),
            },
        )
    if not status:
        status = "completed" if final_response is not None else "stream_partial"
    return text, status, final_response or {}, response_id


def looks_truncated_llm_text(text: str) -> bool:
    """True если ответ похож на обрезанный CF/прокси (незакрытый JSON/строка)."""
    t = (text or "").rstrip()
    if not t:
        return True
    if t.count("```") % 2 == 1:
        return True
    depth = 0
    in_str = False
    esc = False
    for ch in t:
        if in_str:
            if esc:
                esc = False
            elif ch == "\\":
                esc = True
            elif ch == '"':
                in_str = False
            continue
        if ch == '"':
            in_str = True
        elif ch in "{[":
            depth += 1
        elif ch in "}]":
            depth -= 1
    if in_str or depth > 0:
        return True
    return t.endswith((",", ":", "\\"))


def stitch_llm_continuation(base: str, cont: str) -> str:
    """Склеить base + continuation; убрать повтор хвоста, если модель пересказала."""
    a = base or ""
    b = (cont or "").lstrip()
    if not b:
        return a
    if not a:
        return b
    # Ищем максимальный overlap хвоста a с началом b (до 2k).
    # Порог ≥8: короче — ложные склейки; длиннее — CF-continue часто
    # повторяет только `"]}` / пару слов и тогда дублируется.
    max_ov = min(len(a), len(b), 2000)
    for ov in range(max_ov, 7, -1):
        if a.endswith(b[:ov]):
            return a + b[ov:]
    return a + b


def responses_retrieve_urls(post_url: str, response_id: str) -> list[str]:
    """URL для готового ответа kie по Task ID (resp_…)."""
    rid = (response_id or "").strip()
    if not rid:
        return []
    post = (post_url or "").rstrip("/")
    origin = ""
    try:
        parsed = httpx.URL(post)
        origin = f"{parsed.scheme}://{parsed.host}"
        if parsed.port:
            origin = f"{origin}:{parsed.port}"
    except Exception:  # noqa: BLE001
        origin = ""
    urls = [f"{post}/{rid}"]
    if origin:
        urls.append(f"{origin}/api/v1/jobs/recordInfo?taskId={rid}")
        urls.append(f"{origin}/v1/responses/{rid}")
    # без дублей, порядок: тот же path что POST, потом jobs, потом /v1/responses
    seen: set[str] = set()
    out: list[str] = []
    for u in urls:
        if u not in seen:
            seen.add(u)
            out.append(u)
    return out


def parse_retrieved_response_payload(payload: dict[str, Any]) -> tuple[str, str]:
    """Текст и status из GET Responses / jobs.recordInfo."""
    if not isinstance(payload, dict):
        return "", ""
    data = payload.get("data")
    if isinstance(data, dict):
        raw_json = data.get("resultJson")
        if isinstance(raw_json, str) and raw_json.strip():
            try:
                inner = json.loads(raw_json)
            except json.JSONDecodeError:
                inner = None
            if isinstance(inner, dict):
                text, status = parse_retrieved_response_payload(inner)
                if text:
                    return text, status or str(data.get("state") or "")
            if raw_json.strip().startswith("{") or raw_json.strip().startswith("["):
                pass
            else:
                return raw_json.strip(), str(data.get("state") or "success")
        nested = data.get("response") if isinstance(data.get("response"), dict) else data
        if nested is not payload:
            text, status = parse_retrieved_response_payload(nested)
            if text:
                return text, status
    try:
        text, status = _parse_responses(payload)
        return text, status
    except GptApiError:
        pass
    ot = payload.get("output_text")
    if isinstance(ot, str) and ot.strip():
        return ot.strip(), str(payload.get("status") or "completed")
    return "", str(payload.get("status") or payload.get("state") or "")


async def fetch_completed_response(
    *,
    post_url: str,
    headers: dict[str, str],
    response_id: str,
    timeout: float = 90.0,
    polls: int = 12,
    poll_s: float = 3.0,
) -> tuple[str, str, dict[str, Any]]:
    """GET готового ответа kie по resp_… после обрыва SSE.

    ConnectTimeout / сеть — retry на том же URL (раньше сразу break).
    Returns:
        (text, status, raw_payload). Пустой text — забрать не удалось.
    """
    urls = responses_retrieve_urls(post_url, response_id)
    if not urls:
        return "", "", {}
    last_status = ""
    last_raw: dict[str, Any] = {}
    # connect отдельно длиннее: CF/kie часто рвут SSE, а GET потом долго открывается
    sto = httpx.Timeout(
        max(30.0, float(timeout)),
        connect=max(30.0, min(60.0, float(timeout))),
    )
    try:
        async with _async_client(timeout=sto) as client:
            if not hasattr(client, "get"):
                return "", "", {}
            for url in urls:
                for attempt in range(max(1, int(polls))):
                    try:
                        resp = await client.get(url, headers=headers)
                    except Exception as e:  # noqa: BLE001
                        logger.warning(
                            "gpt_api.retrieve {} failed attempt={}/{}: {}: {}",
                            url,
                            attempt + 1,
                            polls,
                            type(e).__name__,
                            e,
                        )
                        # таймаут/обрыв — ждём и пробуем снова, не бросаем URL
                        await asyncio.sleep(poll_s * (1 + attempt // 3))
                        continue
                    if resp.status_code == 404:
                        break
                    if resp.status_code >= 400:
                        logger.warning(
                            "gpt_api.retrieve HTTP {} url={} attempt={}",
                            resp.status_code,
                            url,
                            attempt + 1,
                        )
                        await asyncio.sleep(poll_s)
                        continue
                    try:
                        payload = resp.json()
                    except Exception:  # noqa: BLE001
                        await asyncio.sleep(poll_s)
                        continue
                    if not isinstance(payload, dict):
                        await asyncio.sleep(poll_s)
                        continue
                    last_raw = payload
                    text, status = parse_retrieved_response_payload(payload)
                    last_status = status
                    st = (status or "").lower()
                    if text and st in ("", "completed", "success", "complete"):
                        logger.info(
                            "gpt_api.retrieve OK id={} chars={} url={}",
                            response_id,
                            len(text),
                            url,
                        )
                        return text, status or "completed", payload
                    if st in ("in_progress", "queued", "pending", "generating", "waiting"):
                        await asyncio.sleep(poll_s)
                        continue
                    if text:
                        return text, status or "completed", payload
                    # success без текста (resultJson null) — ещё poll, вдруг допишется
                    if st in ("", "completed", "success", "complete"):
                        await asyncio.sleep(poll_s)
                        continue
                    break
    except Exception as e:  # noqa: BLE001
        logger.warning("gpt_api.retrieve client failed: {}: {}", type(e).__name__, e)
    if last_status:
        logger.warning(
            "gpt_api.retrieve no text id={} last_status={}",
            response_id,
            last_status,
        )
    return "", last_status, last_raw


async def _chat_responses_stream(
    *,
    url: str,
    headers: dict[str, str],
    body: dict[str, Any],
    timeout: float,
    use_model: str,
) -> GptChatResult:
    """POST Responses API с stream=true и сборкой текста из SSE.

    При обрыве после дельт: salvage, затем GET готового ответа по resp_ id
    (kie UI уже success, стрим Cloudflare оборвал).
    """
    stream_body = {**body, "stream": True}
    lines: list[str] = []
    cf_ray = ""
    logged_task_id = False
    saw_text_done = False
    saw_completed = False
    stream_err: BaseException | None = None
    sto = _stream_timeout(timeout)
    logger.info(
        "gpt_api.chat SSE start model={} read_timeout={} (None=wait server EOF)",
        use_model,
        sto.read,
    )
    async with _async_client(timeout=sto) as client:
        try:
            async with client.stream(
                "POST", url, headers=headers, json=stream_body
            ) as resp:
                cf_ray = resp.headers.get("cf-ray") or ""
                if resp.status_code >= 400:
                    err_body = (await resp.aread()).decode("utf-8", errors="replace")
                    _raise_http_status(resp.status_code, err_body, use_model=use_model)
                # Читаем до EOF сервера — сами цикл не прерываем.
                async for line in resp.aiter_lines():
                    lines.append(line)
                    if not line.startswith("data:") or "response." not in line:
                        continue
                    try:
                        ev = json.loads(line[5:].strip())
                    except json.JSONDecodeError:
                        continue
                    if not isinstance(ev, dict):
                        continue
                    et = str(ev.get("type") or "")
                    if et == "response.output_text.done":
                        saw_text_done = True
                    elif et == "response.completed":
                        saw_completed = True
                    if (
                        logged_task_id
                        or "resp_" not in line
                    ):
                        continue
                    resp_obj = ev.get("response")
                    if isinstance(resp_obj, dict) and resp_obj.get("id"):
                        logger.info(
                            "gpt_api.chat task_id={} cf_ray={} model={}",
                            resp_obj["id"],
                            cf_ray or "-",
                            use_model,
                        )
                        logged_task_id = True
                logger.info(
                    "gpt_api.chat SSE EOF lines={} text_done={} completed={} "
                    "cf_ray={}",
                    len(lines),
                    saw_text_done,
                    saw_completed,
                    cf_ray or "-",
                )
        except (httpx.TimeoutException, httpx.HTTPError) as e:
            stream_err = e
            logger.warning(
                "gpt_api.chat SSE interrupted lines={} text_done={} completed={} "
                "err={}: {} — try salvage",
                len(lines),
                saw_text_done,
                saw_completed,
                type(e).__name__,
                e,
            )

    text = ""
    finish = ""
    final_payload: dict[str, Any] = {}
    response_id = ""
    try:
        text, finish, final_payload, response_id = parse_responses_sse_lines(lines)
    except GptApiError as e:
        response_id = str((e.context or {}).get("response_id") or "")
        if not response_id:
            if stream_err is not None:
                raise GptApiError(
                    f"GPT сетевая ошибка: {type(stream_err).__name__}: {stream_err}",
                    context={
                        "error_kind": "network",
                        "retryable": True,
                        "model": use_model,
                        "sse_lines": len(lines),
                        "cf_ray": cf_ray,
                    },
                ) from stream_err
            raise

    if stream_err is not None:
        logger.warning(
            "gpt_api.chat SSE salvaged chars={} task_id={} after {}",
            len(text),
            response_id or "-",
            type(stream_err).__name__,
        )
        finish = "stream_salvaged"

    fetched = False
    need_fetch = bool(response_id) and (
        stream_err is not None
        or not saw_completed
        or looks_truncated_llm_text(text)
        or not (text or "").strip()
    )
    if need_fetch:
        got, got_status, got_raw = await fetch_completed_response(
            post_url=url,
            headers=headers,
            response_id=response_id,
            timeout=max(90.0, min(float(timeout), 180.0)),
            polls=12,
            poll_s=3.0,
        )
        if got and len(got) > len(text or ""):
            text = got
            finish = got_status or "completed"
            if got_raw:
                final_payload = got_raw
            fetched = True
            stream_err = None

    if not (text or "").strip():
        raise GptApiError(
            "GPT(responses/stream): пустой output",
            context={
                "retryable": True,
                "response_id": response_id,
                "error_kind": "empty_stream",
                "sse_lines": len(lines),
                "cf_ray": cf_ray,
            },
        )

    usage = (
        final_payload.get("usage")
        if isinstance(final_payload.get("usage"), dict)
        else {}
    )
    raw = dict(final_payload) if final_payload else {"stream_lines": len(lines)}
    if response_id:
        raw.setdefault("id", response_id)
    if cf_ray:
        raw["cf_ray"] = cf_ray
    if fetched:
        raw["retrieved_by_id"] = True
    elif stream_err is not None:
        raw["sse_salvaged"] = True
        raw["sse_interrupt"] = type(stream_err).__name__
    return GptChatResult(
        text=text,
        model=use_model,
        finish_reason=finish,
        usage=usage,
        raw=raw,
        response_id=response_id,
    )


def parse_chat_completions_sse_lines(lines: list[str]) -> tuple[str, str, dict[str, Any]]:
    """Собрать текст из OpenAI-совместимого SSE ``chat/completions``."""
    chunks: list[str] = []
    finish = ""
    last: dict[str, Any] = {}
    for raw in lines:
        line = (raw or "").strip()
        if not line.startswith("data:"):
            continue
        data = line[5:].strip()
        if not data or data == "[DONE]":
            continue
        try:
            obj = json.loads(data)
        except json.JSONDecodeError:
            continue
        if not isinstance(obj, dict):
            continue
        last = obj
        choices = obj.get("choices")
        if not isinstance(choices, list) or not choices:
            continue
        first = choices[0] if isinstance(choices[0], dict) else {}
        fr = first.get("finish_reason")
        if isinstance(fr, str) and fr:
            finish = fr
        delta = first.get("delta") if isinstance(first.get("delta"), dict) else {}
        piece = delta.get("content")
        if isinstance(piece, str) and piece:
            chunks.append(piece)
        elif isinstance(piece, list):
            chunks.append(
                "".join(
                    str(p.get("text") or "")
                    for p in piece
                    if isinstance(p, dict)
                )
            )
        msg = first.get("message") if isinstance(first.get("message"), dict) else {}
        mc = msg.get("content")
        if isinstance(mc, str) and mc and not delta:
            chunks.append(mc)
    return "".join(chunks), finish or "stop", last


async def _chat_completions_stream(
    *,
    url: str,
    headers: dict[str, str],
    body: dict[str, Any],
    timeout: float,
    use_model: str,
) -> GptChatResult:
    """POST chat/completions с stream=true (длинные ответы vibecode / OpenAI)."""
    stream_body = {**body, "stream": True}
    sto = _stream_timeout(timeout)
    lines: list[str] = []
    stream_err: BaseException | None = None
    async with _async_client(timeout=sto) as client:
        try:
            async with client.stream(
                "POST", url, headers=headers, json=stream_body
            ) as resp:
                if resp.status_code >= 400:
                    err_txt = (await resp.aread()).decode("utf-8", errors="replace")
                    _raise_http_status(
                        resp.status_code, err_txt, use_model=use_model
                    )
                async for raw in resp.aiter_lines():
                    if raw:
                        lines.append(raw)
        except GptApiError:
            raise
        except BaseException as e:
            stream_err = e
            logger.warning(
                "GPT(chat/stream) interrupt after {} lines: {}: {}",
                len(lines),
                type(e).__name__,
                e,
            )
    text, finish, last = parse_chat_completions_sse_lines(lines)
    if not (text or "").strip():
        if stream_err is not None:
            raise GptApiError(
                f"GPT(chat/stream) пустой output после {type(stream_err).__name__}: {stream_err}",
                context={
                    "retryable": True,
                    "error_kind": "empty_stream",
                    "sse_lines": len(lines),
                },
            ) from stream_err
        raise GptApiError(
            "GPT(chat/stream): пустой output",
            context={"retryable": True, "error_kind": "empty_stream", "sse_lines": len(lines)},
        )
    usage = last.get("usage") if isinstance(last.get("usage"), dict) else {}
    raw = dict(last) if last else {"stream_lines": len(lines)}
    if stream_err is not None:
        raw["sse_salvaged"] = True
        raw["sse_interrupt"] = type(stream_err).__name__
    return GptChatResult(
        text=text,
        model=use_model,
        finish_reason=finish,
        usage=usage,
        raw=raw,
        response_id=str(last.get("id") or ""),
    )


def _parse_choice(payload: dict[str, Any]) -> tuple[str, str]:
    """(text, finish_reason) из ответа chat/completions."""
    _check_provider_envelope(payload)
    choices = payload.get("choices")
    if not isinstance(choices, list) or not choices:
        raise GptApiError("GPT: пустой choices в ответе", context={"payload": payload})
    first = choices[0] if isinstance(choices[0], dict) else {}
    message = first.get("message") if isinstance(first.get("message"), dict) else {}
    content = message.get("content")
    # content может быть строкой или списком частей (vision/tools)
    if isinstance(content, list):
        text = "".join(
            str(part.get("text") or "")
            for part in content
            if isinstance(part, dict)
        )
    else:
        text = str(content or "")
    finish = str(first.get("finish_reason") or "")
    if not text.strip():
        raise GptApiError(
            "GPT: пустой content в ответе",
            context={"finish_reason": finish, "payload": payload},
        )
    return text, finish


async def _maybe_volume_complete_chat_result(
    result: GptChatResult,
    *,
    prompt: str,
    accompanying: str,
    input_paths: list[Path] | None,
    system: str | None,
    history: list[dict[str, Any]] | None,
    model: str,
    temperature: float | None,
    timeout: float,
    xlsx_write_contract: str,
    volume_complete: bool | None,
) -> GptChatResult:
    """Если apply-ops/db_frames покрыты частично — добор батчами (~80%)."""
    if volume_complete is False:
        return result
    try:
        from app.services.volume_batches import (
            is_db_frames_path,
            volume_complete_apply_ops_reply,
        )
    except Exception:  # noqa: BLE001
        return result
    paths = list(input_paths or [])
    has_db_frames = any(is_db_frames_path(p) for p in paths)
    if volume_complete is None:
        if xlsx_write_contract != "apply_ops" and not has_db_frames:
            return result
    try:
        new_text, did = await volume_complete_apply_ops_reply(
            result.text or "",
            input_paths=paths,
            prompt=prompt,
            accompanying=accompanying,
            system=system,
            history=history,
            model=model,
            temperature=temperature,
            timeout=timeout,
        )
    except Exception as e:  # noqa: BLE001
        logger.warning("gpt_api.chat volume_complete failed: {}", e)
        return result
    if not did:
        return result
    return GptChatResult(
        text=new_text,
        model=result.model or model,
        finish_reason="volume_continued",
        usage=result.usage,
        raw={
            **(result.raw or {}),
            "volume_continued": True,
            "volume_chars": len(new_text or ""),
        },
        response_id=result.response_id,
    )


_LLM_PACK_PARALLEL = 3


def _merge_packed_apply_ops(texts: list[str]) -> str:
    from app.services.db_apply import extract_apply_ops_json

    ops: list[dict[str, Any]] = []
    extra: dict[str, Any] = {}
    for text in texts:
        data = extract_apply_ops_json(text or "")
        if not isinstance(data, dict):
            continue
        for op in data.get("ops") or []:
            if isinstance(op, dict):
                ops.append(op)
        for key in ("characters", "scenes"):
            val = data.get(key)
            if val and key not in extra:
                extra[key] = val
    if not ops:
        return "\n\n---\n\n".join(t for t in texts if (t or "").strip())
    payload = {"ops": ops, **extra}
    return json.dumps(payload, ensure_ascii=False)


async def _chat_packed_parallel(
    *,
    slice_paths: list[Path],
    prompt: str,
    accompanying: str,
    input_paths: list[Path],
    system: str | None,
    history: list[dict[str, Any]] | None,
    model: str | None,
    temperature: float | None,
    timeout: float | None,
    max_retries: int | None,
    xlsx_write_contract: str,
) -> GptChatResult:
    """Несколько независимых chat() по нарезанным db_frames, потом склейка ops."""
    import asyncio

    from app.services.volume_batches import is_db_frames_path

    sem = asyncio.Semaphore(_LLM_PACK_PARALLEL)

    async def _one(slice_path: Path, idx: int) -> GptChatResult:
        async with sem:
            new_paths = [
                slice_path if is_db_frames_path(p) else p for p in input_paths
            ]
            acc = (
                f"{accompanying}\n\n"
                f"# PACK {idx}/{len(slice_paths)} — только uuid из {slice_path.name}. "
                "Верни JSON ops на ВСЕ кадры ЭТОГО файла. Без продолжений.\n"
            )
            logger.info(
                "gpt_api.chat pack {}/{} file={} bytes={}",
                idx,
                len(slice_paths),
                slice_path.name,
                slice_path.stat().st_size,
            )
            return await chat(
                prompt=prompt,
                accompanying=acc,
                input_paths=new_paths,
                system=system,
                history=None,
                model=model,
                temperature=temperature,
                timeout=timeout,
                max_retries=max_retries,
                xlsx_write_contract=xlsx_write_contract,
                volume_complete=False,
                auto_pack=False,
            )

    parts = await asyncio.gather(
        *[_one(p, i) for i, p in enumerate(slice_paths, start=1)]
    )
    merged = _merge_packed_apply_ops([p.text for p in parts])
    first = parts[0]
    return GptChatResult(
        text=merged,
        model=first.model,
        finish_reason="packed",
        usage=first.usage,
        raw={
            **(first.raw or {}),
            "packed_slices": len(parts),
            "packed_chars": len(merged),
        },
        response_id=first.response_id,
    )


async def chat(
    *,
    prompt: str,
    accompanying: str = "",
    input_paths: list[Path] | None = None,
    system: str | None = None,
    history: list[dict[str, Any]] | None = None,
    model: str | None = None,
    temperature: float | None = None,
    timeout: float | None = None,
    max_retries: int | None = None,
    xlsx_write_contract: str = "tsv",
    volume_complete: bool | None = None,
    auto_pack: bool = True,
    pack_kind: str | None = None,
) -> GptChatResult:
    """Вызвать текстовый LLM (kie GPT / TokenRouter Kimi) с ретраями.

    ``volume_complete``: после частичного apply-ops добрать остаток
    (None = авто: apply_ops contract или db_frames*.json во вложениях).
    ``auto_pack``: нарезать db_frames на батчи
    (img_pr: n*4000/228000; иначе len(закадр)/3500) и гнать пачки параллельно.
    ``pack_kind``: явный ``img_pr`` | ``vo`` (img_pr всегда важнее VO-файла).
    """
    if auto_pack:
        from app.services.output_batch_plan import plan_db_frames_slices

        slices = plan_db_frames_slices(
            input_paths,
            pack_kind=pack_kind,
            prompt=prompt,
            accompanying=accompanying,
        )
        if slices:
            logger.info(
                "gpt_api.chat auto_pack kind={} slices={} files={}",
                pack_kind or "auto",
                len(slices),
                [p.name for p in slices],
            )
            return await _chat_packed_parallel(
                slice_paths=slices,
                prompt=prompt,
                accompanying=accompanying,
                input_paths=list(input_paths or []),
                system=system,
                history=history,
                model=model,
                temperature=temperature,
                timeout=timeout,
                max_retries=max_retries,
                xlsx_write_contract=xlsx_write_contract,
            )

    headers = _headers()
    from app.services.llm_override import current_text_model_id

    use_model = (
        model or current_text_model_id() or settings.gpt_model_effective or "gpt-5.5"
    ).strip()
    url = _chat_url(use_model)
    use_timeout = float(timeout if timeout is not None else settings.gpt_timeout_s)
    retries = int(max_retries if max_retries is not None else settings.gpt_max_retries)
    ov_model = current_text_model_id()
    provider_label = (
        f"vibecode.moe ({ov_model})" if _node_vibecode_override() else settings.text_llm_label
    )
    proxy = _gpt_proxy_url()
    logger.info(
        "text_llm.chat → {} model={} url={} proxy={}",
        provider_label,
        use_model,
        url,
        _mask_proxy_url(proxy) if proxy else "direct",
    )

    responses_mode = is_responses_mode()
    if responses_mode:
        # stream=true: иначе Cloudflare рвёт длинный non-stream JSON, а у kie
        # задача уже success (см. kie.ai/logs Task ID = resp_…).
        body: dict[str, Any] = {
            "model": use_model,
            "input": build_input(
                prompt=prompt,
                accompanying=accompanying,
                input_paths=input_paths,
                system=system,
                history=history,
                xlsx_write_contract=xlsx_write_contract,
            ),
            "stream": True,
        }
    else:
        body = {
            "model": use_model,
            "messages": build_messages(
                prompt=prompt,
                accompanying=accompanying,
                input_paths=input_paths,
                system=system,
                history=history,
                xlsx_write_contract=xlsx_write_contract,
            ),
        }
    if temperature is not None:
        body["temperature"] = temperature

    attempt = 0
    last_exc: Exception | None = None
    while attempt <= retries:
        attempt += 1
        try:
            if responses_mode:
                result = await _chat_responses_stream(
                    url=url,
                    headers=headers,
                    body=body,
                    timeout=use_timeout,
                    use_model=use_model,
                )
                # Cloudflare/kie рвёт длинный SSE: дельты уже есть, но JSON
                # незакрыт — добираем хвост коротким continue (без тяжёлых файлов).
                cont_round = 0
                while cont_round < 2 and looks_truncated_llm_text(result.text):
                    cont_round += 1
                    tail = (result.text or "")[-4000:]
                    cont_prompt = (
                        "Предыдущий ответ оборван сетью (прокси/Cloudflare). "
                        "Продолжи СТРОГО с места обрыва: не повторяй уже "
                        "написанное, без пояснений — только продолжение текста.\n\n"
                        f"--- хвост уже полученного ---\n{tail}\n"
                        "--- конец хвоста ---"
                    )
                    cont_body: dict[str, Any] = {
                        "model": use_model,
                        "input": [
                            {
                                "role": "user",
                                "content": [
                                    {"type": "input_text", "text": cont_prompt}
                                ],
                            }
                        ],
                        "stream": True,
                    }
                    if temperature is not None:
                        cont_body["temperature"] = temperature
                    logger.warning(
                        "gpt_api.chat CF-continue round={} chars_so_far={} task_id={}",
                        cont_round,
                        len(result.text or ""),
                        result.response_id or "-",
                    )
                    try:
                        cont = await _chat_responses_stream(
                            url=url,
                            headers=headers,
                            body=cont_body,
                            timeout=min(use_timeout, 300.0),
                            use_model=use_model,
                        )
                    except GptApiError as e:
                        # Continue без исходного файла часто → пустой output.
                        # Уже полученный текст цельного ответа нельзя выбрасывать.
                        if (result.text or "").strip() and (
                            e.context.get("error_kind") == "empty_stream"
                            or "пустой output" in str(e)
                        ):
                            logger.warning(
                                "gpt_api.chat CF-continue empty — keep salvage "
                                "chars={} ({})",
                                len(result.text or ""),
                                e,
                            )
                            break
                        raise
                    merged = stitch_llm_continuation(result.text, cont.text)
                    if len(merged) <= len(result.text or ""):
                        break
                    result = GptChatResult(
                        text=merged,
                        model=use_model,
                        finish_reason="stream_continued",
                        usage=result.usage,
                        raw={
                            **(result.raw or {}),
                            "cf_continued": True,
                            "cf_continue_rounds": cont_round,
                            "continue_task_id": cont.response_id or "",
                        },
                        response_id=result.response_id or cont.response_id,
                    )
                result = await _maybe_volume_complete_chat_result(
                    result,
                    prompt=prompt,
                    accompanying=accompanying,
                    input_paths=input_paths,
                    system=system,
                    history=history,
                    model=use_model,
                    temperature=temperature,
                    timeout=use_timeout,
                    xlsx_write_contract=xlsx_write_contract,
                    volume_complete=volume_complete,
                )
                logger.info(
                    "gpt_api.chat OK provider={} model={} attempt={} "
                    "finish={} chars={} task_id={}",
                    provider_label,
                    use_model,
                    attempt,
                    result.finish_reason,
                    len(result.text),
                    result.response_id or result.raw.get("id") or "-",
                )
                return result

            if settings.text_llm_is_vibecode:
                result = await _chat_completions_stream(
                    url=url,
                    headers=headers,
                    body=body,
                    timeout=use_timeout,
                    use_model=use_model,
                )
                cont_round = 0
                while cont_round < 2 and looks_truncated_llm_text(result.text):
                    cont_round += 1
                    tail = (result.text or "")[-4000:]
                    cont_prompt = (
                        "Предыдущий ответ оборван сетью. Продолжи СТРОГО с места "
                        "обрыва: не повторяй уже написанное, без пояснений — "
                        "только продолжение текста.\n\n"
                        f"--- хвост уже полученного ---\n{tail}\n"
                        "--- конец хвоста ---"
                    )
                    cont_body = {
                        "model": use_model,
                        "messages": [
                            {"role": "user", "content": cont_prompt},
                        ],
                    }
                    try:
                        cont = await _chat_completions_stream(
                            url=url,
                            headers=headers,
                            body=cont_body,
                            timeout=min(use_timeout, 180.0),
                            use_model=use_model,
                        )
                    except GptApiError as e:
                        if (result.text or "").strip() and (
                            e.context.get("error_kind") == "empty_stream"
                            or "пустой output" in str(e)
                        ):
                            logger.warning(
                                "gpt_api.chat vibecode-continue empty — keep salvage "
                                "chars={} ({})",
                                len(result.text or ""),
                                e,
                            )
                            break
                        raise
                    merged = stitch_llm_continuation(result.text, cont.text)
                    if len(merged) <= len(result.text or ""):
                        break
                    result = GptChatResult(
                        text=merged,
                        model=use_model,
                        finish_reason="stream_continued",
                        usage=result.usage,
                        raw={
                            **(result.raw or {}),
                            "cf_continued": True,
                            "cf_continue_rounds": cont_round,
                        },
                        response_id=result.response_id or cont.response_id,
                    )
                result = await _maybe_volume_complete_chat_result(
                    result,
                    prompt=prompt,
                    accompanying=accompanying,
                    input_paths=input_paths,
                    system=system,
                    history=history,
                    model=use_model,
                    temperature=temperature,
                    timeout=use_timeout,
                    xlsx_write_contract=xlsx_write_contract,
                    volume_complete=volume_complete,
                )
                logger.info(
                    "gpt_api.chat OK provider={} model={} attempt={} "
                    "finish={} chars={} task_id={}",
                    provider_label,
                    use_model,
                    attempt,
                    result.finish_reason,
                    len(result.text),
                    result.response_id or result.raw.get("id") or "-",
                )
                return result

            async with _async_client(timeout=_http_timeout(use_timeout)) as client:
                resp = await client.post(url, headers=headers, json=body)
            status = resp.status_code
            _raise_http_status(status, resp.text, use_model=use_model)
            try:
                payload = resp.json()
            except Exception as e:  # noqa: BLE001
                raise GptApiError(
                    f"GPT: ответ не JSON: {resp.text[:300]}",
                    context={"retryable": True, "model": use_model},
                ) from e
            text, finish = _parse_choice(payload)
            usage = payload.get("usage") if isinstance(payload.get("usage"), dict) else {}
            result = GptChatResult(
                text=text, model=use_model, finish_reason=finish, usage=usage, raw=payload
            )
            result = await _maybe_volume_complete_chat_result(
                result,
                prompt=prompt,
                accompanying=accompanying,
                input_paths=input_paths,
                system=system,
                history=history,
                model=use_model,
                temperature=temperature,
                timeout=use_timeout,
                xlsx_write_contract=xlsx_write_contract,
                volume_complete=volume_complete,
            )
            logger.info(
                "gpt_api.chat OK provider={} model={} attempt={} finish={} chars={}",
                provider_label,
                use_model,
                attempt,
                result.finish_reason,
                len(result.text),
            )
            return result
        except httpx.TimeoutException:
            hint = ""
            if settings.text_llm_is_tokenrouter:
                hint = (
                    " — Kimi free на TokenRouter часто тормозит/висит; "
                    "подожди или переключи бейдж модели на GPT (kie)"
                )
            last_exc = GptApiError(
                f"{provider_label} timeout {use_timeout:.0f}s "
                f"(попытка {attempt}/{retries + 1}){hint}",
                context={
                    "error_kind": "timeout",
                    "retryable": True,
                    "model": use_model,
                    "provider": provider_label,
                },
            )
            logger.warning(str(last_exc))
        except httpx.HTTPError as e:
            last_exc = GptApiError(
                f"GPT сетевая ошибка: {type(e).__name__}: {e}",
                context={"error_kind": "network", "retryable": True, "model": use_model},
            )
            logger.warning(str(last_exc))
        except GptApiError as e:
            if not e.retryable:
                raise
            # kie иногда падает на PDF input_file → code=500; оставляем только текст.
            # Не сжигаем попытку: attempt -= 1, чтобы text-only точно ушёл.
            if (
                responses_mode
                and (
                    int(e.context.get("provider_code") or 0) == 500
                    or int(e.context.get("status_code") or 0) == 500
                    or "server exception" in str(e).lower()
                )
                and _drop_input_file_parts(body)
            ):
                logger.warning(
                    "gpt_api.chat: PDF input_file → 500, повтор только с текстом PDF"
                )
                last_exc = e
                attempt = max(0, attempt - 1)
                continue
            last_exc = e
            logger.warning("gpt_api.chat retryable: {}", e)

        if attempt <= retries:
            backoff = min(2.0 * (2 ** (attempt - 1)), 30.0)
            await asyncio.sleep(backoff)

    raise last_exc or GptApiError("GPT: неизвестная ошибка", context={"model": use_model})


def is_pdf_provider_failure(exc: BaseException) -> bool:
    """kie 500 Server exception / timeout на большом PDF — имеет смысл chunked retry."""
    if isinstance(exc, GptApiError):
        kind = str(exc.context.get("error_kind") or "")
        code = int(exc.context.get("provider_code") or exc.context.get("status_code") or 0)
        if kind == "timeout" or code >= 500:
            return True
    msg = str(exc).lower()
    return (
        "server exception" in msg
        or "code=500" in msg
        or "timeout" in msg
        or "http 500" in msg
        or "http 502" in msg
        or "http 503" in msg
    )


async def chat_pdf_in_chunks(
    *,
    prompt: str,
    accompanying: str = "",
    pdf_paths: list[Path],
    other_paths: list[Path] | None = None,
    system: str | None = None,
    history: list[dict[str, Any]] | None = None,
    model: str | None = None,
    temperature: float | None = None,
    timeout: float | None = None,
    max_retries: int | None = None,
    chunk_chars: int = _PDF_PROVIDER_SAFE_CHARS,
    on_chunk: Any | None = None,
) -> GptChatResult:
    """Перевод/обработка PDF по частям — обход kie code=500 / hang на полном тексте.

    Текст извлекается локально (pypdf); в API уходит только текущий кусок без
    повторной отправки всего PDF. История чата — только на первом куске.

    Live: kie часто даёт code=500 на *втором* куске после успешного первого —
    раньше весь чат падал. Теперь: пауза между кусками, ≥3 ретрая, раскол
    куска пополам при стойком 500, иначе пометка и продолжение остальных.
    """
    jobs: list[tuple[str, str]] = []
    for pdf in pdf_paths:
        extracted = pdf_to_text(pdf)
        parts = split_pdf_text_chunks(extracted, max_chars=chunk_chars)
        n = max(1, len(parts))
        for i, part in enumerate(parts, start=1):
            jobs.append((f"{pdf.name} · фрагмент {i}/{n}", part))

    if not jobs:
        raise GptApiError(
            "PDF: не удалось извлечь текст для обработки",
            context={"retryable": False, "error_kind": "empty_pdf"},
        )

    logger.info(
        "gpt_api: PDF chunked mode jobs={} chunk_chars={} files={}",
        len(jobs),
        chunk_chars,
        [p.name for p in pdf_paths],
    )
    user_ask = (prompt or "").strip()
    if accompanying.strip():
        user_ask = f"{user_ask}\n\n{accompanying.strip()}".strip()
    outs: list[str] = []
    last: GptChatResult | None = None
    ok_n = 0
    fail_n = 0
    chunk_timeout = float(timeout if timeout is not None else 120.0)
    # Workspace передаёт max_retries=1 — для PDF этого мало (kie flaky).
    chunk_retries = max(3, int(max_retries if max_retries is not None else 2))
    other = list(other_paths or [])

    async def _one_piece(
        *,
        label: str,
        part: str,
        idx: int,
        depth: int,
        use_system: bool,
        use_history: bool,
    ) -> str:
        nonlocal last, ok_n, fail_n
        piece_prompt = (
            f"{user_ask}\n\n"
            f"[Это фрагмент {idx} из {len(jobs)} документа «{label}». "
            f"Выполни задание полностью для ЭТОГО фрагмента: не сокращай, "
            f"не пиши «продолжение следует», без мета-вступлений про части.]\n\n"
            f"{part}"
        )
        try:
            last = await chat(
                prompt=piece_prompt,
                accompanying="",
                # Не тащить pdf_paths снова — иначе полный текст PDF дублируется.
                input_paths=other,
                system=system if use_system else None,
                history=history if use_history else None,
                model=model,
                temperature=temperature,
                timeout=chunk_timeout,
                max_retries=chunk_retries,
            )
            body = (last.text or "").strip() or "[пустой ответ модели]"
            ok_n += 1
            return f"### {label}\n\n{body}"
        except Exception as e:  # noqa: BLE001
            if is_pdf_provider_failure(e) and depth < 2 and len(part) > 1_200:
                cut = len(part) // 2
                for sep in ("\n\n", "\n", " "):
                    pos = part.rfind(sep, max(0, cut - 400), cut + 400)
                    if pos > 200:
                        cut = pos + len(sep)
                        break
                logger.warning(
                    "gpt_api: PDF piece {} failed ({}) — split depth={} lens={}+{}",
                    label,
                    e,
                    depth + 1,
                    cut,
                    len(part) - cut,
                )
                await asyncio.sleep(1.5)
                left = await _one_piece(
                    label=f"{label}·a",
                    part=part[:cut],
                    idx=idx,
                    depth=depth + 1,
                    use_system=use_system,
                    use_history=use_history,
                )
                await asyncio.sleep(1.5)
                right = await _one_piece(
                    label=f"{label}·b",
                    part=part[cut:],
                    idx=idx,
                    depth=depth + 1,
                    use_system=False,
                    use_history=False,
                )
                return f"{left}\n\n---\n\n{right}"
            fail_n += 1
            logger.warning(
                "gpt_api: PDF piece {} skipped after error: {}",
                label,
                e,
            )
            return (
                f"### {label}\n\n"
                f"[фрагмент не обработан провайдером: {e} — "
                f"остальные части ниже, если есть]"
            )

    for idx, (label, part) in enumerate(jobs):
        if callable(on_chunk):
            try:
                on_chunk(idx + 1, len(jobs), label)
            except Exception:  # noqa: BLE001
                logger.debug("gpt_api: on_chunk callback failed", exc_info=True)
        if idx > 0:
            # Пауза: подряд идущие вызовы после OK часто ловят code=500.
            await asyncio.sleep(1.75)
        outs.append(
            await _one_piece(
                label=label,
                part=part,
                idx=idx + 1,
                depth=0,
                use_system=(idx == 0),
                use_history=(idx == 0),
            )
        )

    if ok_n == 0:
        raise GptApiError(
            "GPT провайдер code=500: Server exception на всех фрагментах PDF "
            "(kie). Попробуй ещё раз или разбей PDF на меньший файл.",
            context={"retryable": True, "error_kind": "pdf_all_chunks_failed"},
        )

    joined = "\n\n---\n\n".join(outs)
    if fail_n:
        joined = (
            f"[PDF: часть фрагментов пропущена из‑за сбоя провайдера "
            f"(ok={ok_n}, fail≈{fail_n}). Повтори запрос для пропущенных.]\n\n"
            f"{joined}"
        )
    return GptChatResult(
        text=joined,
        model=(last.model if last else (model or settings.gpt_model or "")),
        finish_reason="chunked_partial" if fail_n else "chunked",
        usage=(last.usage if last else {}),
        raw={
            "chunked": True,
            "parts": len(jobs),
            "ok_chunks": ok_n,
            "fail_chunks": fail_n,
        },
    )



def _drop_input_file_parts(body: dict[str, Any]) -> bool:
    """Убрать input_file из Responses body. True если что-то сняли."""
    inp = body.get("input")
    if not isinstance(inp, list):
        return False
    changed = False
    new_inp: list[Any] = []
    for item in inp:
        if not isinstance(item, dict):
            new_inp.append(item)
            continue
        content = item.get("content")
        if not isinstance(content, list):
            new_inp.append(item)
            continue
        filtered = [
            p
            for p in content
            if not (isinstance(p, dict) and str(p.get("type") or "") == "input_file")
        ]
        if len(filtered) == len(content):
            new_inp.append(item)
            continue
        changed = True
        if (
            len(filtered) == 1
            and isinstance(filtered[0], dict)
            and filtered[0].get("type") == "input_text"
        ):
            new_inp.append({**item, "content": str(filtered[0].get("text") or "")})
        elif filtered:
            new_inp.append({**item, "content": filtered})
        else:
            new_inp.append({**item, "content": "(файл снят после ошибки провайдера)"})
    if changed:
        body["input"] = new_inp
    return changed


# ─────────────────────────── контент (download/copy) ───────────────────────────

_DATA_URI_RE = re.compile(
    r"data:(image/(?:png|jpeg|jpg|webp|gif|svg\+xml)|application/(?:pdf|zip|octet-stream|"
    r"vnd\.openxmlformats-officedocument\.spreadsheetml\.sheet|"
    r"vnd\.ms-excel))"
    r";base64,([A-Za-z0-9+/=\s]+)",
    re.IGNORECASE,
)


def sniff_file_extension(data: bytes) -> str | None:
    """Определить расширение по magic bytes / сигнатурам (не по имени/.bin)."""
    if not data:
        return None
    # JPEG: достаточно 3 байт; остальным обычно хватает 12
    if len(data) >= 3 and data.startswith(b"\xff\xd8\xff"):
        return ".jpg"
    if len(data) < 8:
        return None
    if data.startswith(b"\x89PNG\r\n\x1a\n"):
        return ".png"
    if data.startswith(b"GIF87a") or data.startswith(b"GIF89a"):
        return ".gif"
    if len(data) >= 12 and data.startswith(b"RIFF") and data[8:12] == b"WEBP":
        return ".webp"
    if data.startswith(b"BM"):
        return ".bmp"
    # AVIF / HEIC (ISO BMFF)
    if len(data) >= 12 and data[4:8] == b"ftyp":
        brand = data[8:12]
        if brand in (b"avif", b"avis"):
            return ".avif"
        if brand in (b"heic", b"heif", b"mif1", b"msf1"):
            return ".heic"
    if data.startswith(b"%PDF"):
        return ".pdf"
    # SVG (часто text/xml без магии)
    head = data[:200].lstrip().lower()
    if head.startswith(b"<svg") or (head.startswith(b"<?xml") and b"<svg" in data[:800].lower()):
        return ".svg"
    if data[:4] in (b"PK\x03\x04", b"PK\x05\x06"):
        head = data[:8000]
        if b"word/" in head:
            return ".docx"
        if b"xl/" in head or b"xl\\" in head:
            return ".xlsx"
        return ".zip"
    # ICO
    if len(data) >= 4 and data[:4] == b"\x00\x00\x01\x00":
        return ".ico"
    # SVG / HTML (текст)
    head = data.lstrip()[:800].lower()
    if head.startswith(b"<svg") or (head.startswith(b"<?xml") and b"<svg" in head):
        return ".svg"
    if head.startswith(b"<!doctype html") or head.startswith(b"<html"):
        return ".html"
    return None


def extension_from_content_type(content_type: str | None) -> str | None:
    if not content_type:
        return None
    ct = content_type.split(";")[0].strip().lower()
    return {
        "image/png": ".png",
        "image/jpeg": ".jpg",
        "image/jpg": ".jpg",
        "image/webp": ".webp",
        "image/gif": ".gif",
        "image/bmp": ".bmp",
        "image/svg+xml": ".svg",
        "image/avif": ".avif",
        "image/heic": ".heic",
        "image/heif": ".heic",
        "image/x-icon": ".ico",
        "image/vnd.microsoft.icon": ".ico",
        "application/pdf": ".pdf",
        "application/zip": ".zip",
        "application/json": ".json",
        "text/plain": ".txt",
        "text/html": ".html",
        "text/csv": ".csv",
        "text/markdown": ".md",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document": ".docx",
    }.get(ct)


def maybe_decode_base64_image(data: bytes) -> tuple[bytes, str | None]:
    """Если тело — base64 PNG/JPEG/… (часто так отдают «файлы»), декодировать."""
    import base64
    import re

    if not data or len(data) < 32:
        return data, None
    sample = data.lstrip()[:32]
    # типичные префиксы base64 картинок
    if not (
        sample.startswith(b"iVBOR")  # PNG
        or sample.startswith(b"/9j/")  # JPEG
        or sample.startswith(b"R0lGOD")  # GIF
        or sample.startswith(b"UklGR")  # WEBP
        or sample.startswith(b"AAAA")  # иногда AVIF/mp4
    ):
        return data, None
    # только printable base64 + whitespace
    if any(b < 9 or (13 < b < 32) for b in data[:4000]):
        return data, None
    raw = re.sub(rb"\s+", b"", data.strip())
    if len(raw) < 32 or len(raw) % 4 not in (0,):  # pad ok
        # try pad
        raw = raw + b"=" * ((4 - len(raw) % 4) % 4)
    try:
        decoded = base64.b64decode(raw, validate=False)
    except Exception:  # noqa: BLE001
        return data, None
    if len(decoded) < 32:
        return data, None
    ext = sniff_file_extension(decoded)
    if ext:
        return decoded, ext
    return data, None


def resolve_bytes_extension(
    data: bytes,
    *,
    content_type: str | None = None,
    url: str | None = None,
    fallback: str = ".dat",
) -> str:
    """Расширение для байтов. НИКОГДА не возвращает .bin."""
    sniffed = sniff_file_extension(data)
    if sniffed:
        return sniffed
    from_ct = extension_from_content_type(content_type)
    if from_ct:
        return from_ct
    if url:
        suf = Path(url.split("?")[0]).suffix.lower()
        if suf in {
            ".png",
            ".jpg",
            ".jpeg",
            ".webp",
            ".gif",
            ".bmp",
            ".svg",
            ".avif",
            ".heic",
            ".ico",
            ".pdf",
            ".zip",
            ".xlsx",
            ".xls",
            ".txt",
            ".md",
            ".csv",
            ".json",
            ".html",
            ".mp4",
            ".webm",
        }:
            return ".jpg" if suf == ".jpeg" else suf
    # text-ish
    if data and all(b in b"\t\n\r\f\b" or 32 <= b < 127 or b >= 160 for b in data[:2000]):
        return ".txt"
    return fallback if fallback != ".bin" else ".dat"


_WEAK_SUFFIXES = frozenset(
    {"", ".bin", ".dat", ".tmp", ".octet-stream", ".unknown", ".download", ".file"}
)
_IMAGE_SUFFIXES = frozenset(
    {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp", ".svg", ".avif", ".heic", ".ico"}
)
_SNIFF_MIME = {
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".gif": "image/gif",
    ".webp": "image/webp",
    ".bmp": "image/bmp",
    ".svg": "image/svg+xml",
    ".avif": "image/avif",
    ".heic": "image/heic",
    ".ico": "image/x-icon",
    ".pdf": "application/pdf",
    ".zip": "application/zip",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ".html": "text/html",
    ".txt": "text/plain",
}


def _read_magic_head(path: Path) -> bytes:
    try:
        size = path.stat().st_size
        if size <= 64_000:
            return path.read_bytes()
        with path.open("rb") as f:
            return f.read(16_384)
    except OSError:
        return b""


def needs_extension_fix(path: Path, sniffed: str | None) -> bool:
    """True если имя слабое (.bin) или не совпадает с magic (картинка)."""
    if not sniffed:
        return False
    cur = path.suffix.lower()
    if cur in _WEAK_SUFFIXES:
        return True
    if sniffed in _IMAGE_SUFFIXES and cur not in _IMAGE_SUFFIXES:
        return True
    return False


def suggested_name_and_mime(path: Path) -> tuple[str, str]:
    """Имя для Content-Disposition + MIME по magic (без обязательного rename на диске)."""
    import mimetypes

    head = _read_magic_head(path)
    sniffed = sniff_file_extension(head)
    if not sniffed and path.suffix.lower() in _WEAK_SUFFIXES:
        sniffed = resolve_bytes_extension(head or b"\x00", fallback=".dat")
    name = path.name
    if sniffed and (
        needs_extension_fix(path, sniffed)
        or path.suffix.lower() in {".bin", ".download", ".octet-stream"}
    ):
        name = f"{path.stem}{sniffed}"
    mime, _ = mimetypes.guess_type(name)
    if not mime and sniffed:
        mime = _SNIFF_MIME.get(sniffed)
    if not mime:
        mime, _ = mimetypes.guess_type(str(path))
    return name, mime or "application/octet-stream"


def ensure_correct_extension(path: Path) -> Path:
    """Переименовать файл по magic/base64; .bin запрещены — уйдут в реальный тип или .dat."""
    if not path.is_file():
        return path
    try:
        raw = path.read_bytes()
    except OSError:
        return path
    decoded, b64_ext = maybe_decode_base64_image(raw)
    if b64_ext and decoded is not raw:
        try:
            path.write_bytes(decoded)
            raw = decoded
        except OSError as e:
            logger.warning("base64 unwrap write failed {}: {}", path, e)
    sniffed = sniff_file_extension(raw) or b64_ext
    if not sniffed and path.suffix.lower() in _WEAK_SUFFIXES:
        # всё равно убрать .bin
        sniffed = resolve_bytes_extension(raw, fallback=".dat")
    if not sniffed:
        return path
    if not needs_extension_fix(path, sniffed) and path.suffix.lower() not in {
        ".bin",
        ".download",
        ".octet-stream",
    }:
        return path
    # всегда чиним .bin даже если sniffed == fallback .dat
    if path.suffix.lower() in {".bin", ".download", ".octet-stream"} or needs_extension_fix(
        path, sniffed
    ):
        target = path.with_suffix(sniffed)
        if target.resolve() == path.resolve():
            return path
        n = 2
        while target.exists():
            target = path.with_name(f"{path.stem}_{n}{sniffed}")
            n += 1
        try:
            path.rename(target)
            logger.info("sniff rename {} → {}", path.name, target.name)
            return target
        except OSError as e:
            logger.warning("sniff rename failed {}: {}", path, e)
            return path
    return path


def finalize_downloaded_file(
    path: Path,
    *,
    content_type: str | None = None,
    url: str | None = None,
) -> Path:
    """После HTTP-download: decode base64 → sniff/CT/url → имя без .bin."""
    if not path.is_file():
        return path
    try:
        raw = path.read_bytes()
    except OSError:
        return path
    decoded, b64_ext = maybe_decode_base64_image(raw)
    if b64_ext and decoded is not raw:
        try:
            path.write_bytes(decoded)
            raw = decoded
        except OSError:
            pass
    ext = resolve_bytes_extension(
        raw, content_type=content_type, url=url, fallback=".dat"
    )
    if b64_ext:
        ext = b64_ext
    cur = path.suffix.lower()
    if cur == ext:
        return path
    # Office-имя + HTML/SVG/текст внутри = типичный «скачали страницу ошибки
    # как .xlsx» (magic 3c21444f = <!DO). Иначе writeback заливает мусор.
    _office = {".xlsx", ".xlsm", ".xls", ".docx"}
    _not_office = {".html", ".htm", ".svg", ".txt", ".json", ".xml", ".dat"}
    office_lied = cur in _office and ext in _not_office
    # слабое имя или несовпадение с картинкой / принудительно убрать .bin
    if (
        cur in _WEAK_SUFFIXES
        or cur == ".bin"
        or office_lied
        or (ext in _IMAGE_SUFFIXES and cur not in _IMAGE_SUFFIXES)
    ):
        target = path.with_suffix(ext)
        n = 2
        while target.exists() and target.resolve() != path.resolve():
            target = path.with_name(f"{path.stem}_{n}{ext}")
            n += 1
        if target.resolve() != path.resolve():
            try:
                path.rename(target)
                logger.info(
                    "download finalize {} → {} (ct={})",
                    path.name,
                    target.name,
                    content_type,
                )
                return target
            except OSError as e:
                logger.warning("download finalize rename failed {}: {}", path, e)
    return ensure_correct_extension(path)


def collect_result_urls(text: str) -> list[str]:
    """Достать http(s)-ссылки; картинки/файлы раньше HTML-страниц."""
    if not text:
        return []
    seen: set[str] = set()
    out: list[str] = []
    for m in _URL_RE.finditer(text):
        u = m.group(0).rstrip(".,;")
        if u not in seen:
            seen.add(u)
            out.append(u)
    out.sort(key=_url_image_priority)
    return out


_IMG_PATH_RE = re.compile(
    r"\.(png|jpe?g|webp|gif|bmp|svg|avif|heic|ico)(?:\?|#|$)",
    re.IGNORECASE,
)


def _url_image_priority(url: str) -> int:
    """Меньше = раньше качаем (прямые картинки важнее wiki/html)."""
    u = (url or "").lower()
    path = u.split("?")[0]
    if _IMG_PATH_RE.search(path):
        return 0
    if any(
        x in u
        for x in (
            "/images/",
            "/image/",
            "/media/",
            "upload.",
            "imgur.",
            "i.ibb.",
            "cdn.",
            "special:filepath",
        )
    ):
        return 1
    if path.endswith((".html", ".htm")) or "/wiki/" in u or "/article/" in u:
        return 5
    return 3


def extract_image_urls_from_html(html: bytes | str, base_url: str) -> list[str]:
    """Достать реальные URL картинок из HTML (og:image, twitter:image, img/srcset)."""
    from urllib.parse import urljoin

    if isinstance(html, bytes):
        text = html.decode("utf-8", errors="replace")
    else:
        text = html
    found: list[str] = []
    patterns = (
        r'<meta[^>]+property=["\']og:image(?::secure_url)?["\'][^>]+content=["\']([^"\']+)["\']',
        r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+property=["\']og:image(?::secure_url)?["\']',
        r'<meta[^>]+name=["\']twitter:image(?::src)?["\'][^>]+content=["\']([^"\']+)["\']',
        r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+name=["\']twitter:image(?::src)?["\']',
        r'<link[^>]+rel=["\']image_src["\'][^>]+href=["\']([^"\']+)["\']',
        r'<meta[^>]+itemprop=["\']image["\'][^>]+content=["\']([^"\']+)["\']',
    )
    for pat in patterns:
        for m in re.finditer(pat, text, re.IGNORECASE):
            found.append(m.group(1).strip())
    for m in re.finditer(r'<img\b[^>]*>', text, re.IGNORECASE):
        tag = m.group(0)
        for attr in ("src", "data-src", "data-original"):
            am = re.search(rf'{attr}=["\']([^"\']+)["\']', tag, re.IGNORECASE)
            if am:
                src = am.group(1).strip()
                if src and not src.startswith("data:"):
                    found.append(src)
        sm = re.search(r'srcset=["\']([^"\']+)["\']', tag, re.IGNORECASE)
        if sm:
            first = sm.group(1).split(",")[0].strip().split()[0]
            if first and not first.startswith("data:"):
                found.append(first)

    out: list[str] = []
    seen: set[str] = set()
    for u in found:
        abs_u = urljoin(base_url, u.replace("&amp;", "&"))
        if not abs_u.startswith("http"):
            continue
        if abs_u in seen:
            continue
        seen.add(abs_u)
        out.append(abs_u)
    out.sort(key=_url_image_priority)
    return out


def extract_data_uri_files(text: str, out_dir: Path, *, prefix: str = "embed") -> list[Path]:
    """Вынуть data:image/...;base64,... из ответа модели в файлы."""
    import base64

    if not text:
        return []
    out_dir.mkdir(parents=True, exist_ok=True)
    saved: list[Path] = []
    for i, m in enumerate(_DATA_URI_RE.finditer(text)):
        mime = (m.group(1) or "application/octet-stream").lower()
        raw_b64 = re.sub(r"\s+", "", m.group(2) or "")
        if len(raw_b64) < 32:
            continue
        ext = {
            "image/png": ".png",
            "image/jpeg": ".jpg",
            "image/jpg": ".jpg",
            "image/webp": ".webp",
            "image/gif": ".gif",
            "image/svg+xml": ".svg",
            "application/pdf": ".pdf",
            "application/zip": ".zip",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
            "application/vnd.ms-excel": ".xls",
        }.get(mime, ".dat")
        try:
            data = base64.b64decode(raw_b64, validate=False)
        except Exception:  # noqa: BLE001
            continue
        # SVG часто короче 32 байт после decode; растр — нет
        min_bytes = 8 if "svg" in mime else 32
        if len(data) < min_bytes:
            continue
        sniffed = sniff_file_extension(data)
        if sniffed:
            ext = sniffed
        path = out_dir / f"{prefix}_{i}{ext}"
        n = 2
        while path.exists():
            path = out_dir / f"{prefix}_{i}_{n}{ext}"
            n += 1
        path.write_bytes(data)
        saved.append(ensure_correct_extension(path))
    return saved


async def materialize_reply_assets(
    text: str,
    out_dir: Path,
    *,
    prefix: str = "file",
    timeout: float = 120.0,
    max_urls: int = 8,
) -> list[Path]:
    """Скачать URL и data-URI из ответа модели в out_dir.

    Если URL отдаёт HTML-страницу — вытаскиваем og:image/img и качаем картинку.
    Сам .html в Результаты не кладём (в браузерном ChatGPT так же видна картинка, не HTML).
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    saved: list[Path] = []
    saved.extend(extract_data_uri_files(text, out_dir, prefix=f"{prefix}_embed"))

    for i, url in enumerate(collect_result_urls(text)[:max_urls]):
        dest = out_dir / f"{prefix}_url_{i}.download"
        n = 2
        while dest.exists():
            dest = out_dir / f"{prefix}_url_{i}_{n}.download"
            n += 1
        try:
            got = await download_content(url, dest, timeout=timeout)
            got = ensure_correct_extension(got)
            if got.suffix.lower() == ".bin":
                got = finalize_downloaded_file(got, url=url)

            # HTML-страница вместо файла → достать картинку
            if got.suffix.lower() in {".html", ".htm"}:
                try:
                    html_bytes = got.read_bytes()
                except OSError:
                    html_bytes = b""
                img_urls = extract_image_urls_from_html(html_bytes, url)
                try:
                    got.unlink(missing_ok=True)
                except OSError:
                    pass
                pulled = False
                for j, iurl in enumerate(img_urls[:4]):
                    idest = out_dir / f"{prefix}_url_{i}_img_{j}.download"
                    k = 2
                    while idest.exists():
                        idest = out_dir / f"{prefix}_url_{i}_img_{j}_{k}.download"
                        k += 1
                    try:
                        igot = await download_content(iurl, idest, timeout=timeout)
                        igot = ensure_correct_extension(igot)
                        if igot.suffix.lower() in {".html", ".htm", ".txt", ".dat"}:
                            # не картинка — выкинуть
                            try:
                                igot.unlink(missing_ok=True)
                            except OSError:
                                pass
                            continue
                        if (
                            igot.suffix.lower() in _IMAGE_SUFFIXES
                            or sniff_file_extension(igot.read_bytes()[:64]) in _IMAGE_SUFFIXES
                        ):
                            saved.append(igot)
                            pulled = True
                            logger.info(
                                "materialize: HTML {} → image {}",
                                url[:100],
                                igot.name,
                            )
                            break
                        # не image — тоже не оставляем мусор
                        try:
                            igot.unlink(missing_ok=True)
                        except OSError:
                            pass
                    except Exception as e:  # noqa: BLE001
                        logger.warning(
                            "materialize: img from html fail {}: {}",
                            iurl[:100],
                            e,
                        )
                if not pulled:
                    logger.warning(
                        "materialize: HTML без картинки, skip: {}",
                        url[:120],
                    )
                continue

            # не тащим html/пустые заглушки в Результаты
            if got.suffix.lower() in {".html", ".htm"}:
                try:
                    got.unlink(missing_ok=True)
                except OSError:
                    pass
                continue
            saved.append(got)
        except Exception as e:  # noqa: BLE001
            logger.warning("materialize_reply_assets: skip {}: {}", url[:120], e)
    return saved


async def search_web_image_urls(query: str, *, limit: int = 4) -> list[str]:
    """Найти URL картинок (Openverse + запасные источники, без API-ключа)."""
    queries = [q.strip() for q in (query or "").split("||") if q.strip()]
    if not queries:
        q = (query or "").strip()
        if q:
            queries = [q]
    if not queries:
        return []
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        ),
        "Accept": "application/json,text/html,*/*",
    }
    found: list[str] = []
    seen: set[str] = set()

    def _add(u: str | None) -> None:
        u = (u or "").strip()
        if not u or not u.startswith("http") or u in seen:
            return
        if u.startswith("data:"):
            return
        seen.add(u)
        found.append(u)

    async with _async_client(
        timeout=25.0, follow_redirects=True, headers=headers
    ) as client:
        for q in queries:
            if len(found) >= limit:
                break
            try:
                r = await client.get(
                    "https://api.openverse.org/v1/images/",
                    params={"q": q, "page_size": str(max(limit, 5))},
                )
                if r.status_code < 400 and r.content:
                    data = r.json()
                    for item in (data.get("results") or []) if isinstance(data, dict) else []:
                        if not isinstance(item, dict):
                            continue
                        _add(item.get("url"))
                        _add(item.get("thumbnail"))
            except Exception as e:  # noqa: BLE001
                logger.warning("search_web_image_urls: openverse {!r}: {}", q, e)

            if len(found) >= limit:
                break
            try:
                r = await client.get(
                    "https://api.duckduckgo.com/",
                    params={
                        "q": q,
                        "format": "json",
                        "no_redirect": "1",
                        "no_html": "1",
                    },
                )
                if r.status_code < 400 and r.content:
                    data = r.json()
                    if isinstance(data, dict):
                        _add(data.get("Image"))
            except Exception as e:  # noqa: BLE001
                logger.warning("search_web_image_urls: ddg {!r}: {}", q, e)

    found.sort(key=_url_image_priority)
    return found[:limit]


async def fetch_web_images(
    query: str,
    out_dir: Path,
    *,
    prefix: str = "web",
    limit: int = 3,
    timeout: float = 60.0,
    query_variants: list[str] | None = None,
) -> list[Path]:
    """Поиск + скачивание картинок по запросу пользователя."""
    if query_variants:
        joined = "||".join(query_variants)
    else:
        joined = query
    urls = await search_web_image_urls(joined, limit=limit)
    if not urls:
        return []
    out_dir.mkdir(parents=True, exist_ok=True)
    saved: list[Path] = []
    for i, url in enumerate(urls):
        dest = out_dir / f"{prefix}_{i}.download"
        n = 2
        while dest.exists():
            dest = out_dir / f"{prefix}_{i}_{n}.download"
            n += 1
        try:
            got = await download_content(url, dest, timeout=timeout)
            got = ensure_correct_extension(got)
            if got.suffix.lower() in {".html", ".htm", ".bin", ".download"}:
                try:
                    got.unlink(missing_ok=True)
                except OSError:
                    pass
                continue
            if got.stat().st_size < 64:
                try:
                    got.unlink(missing_ok=True)
                except OSError:
                    pass
                continue
            saved.append(got)
            if len(saved) >= limit:
                break
        except Exception as e:  # noqa: BLE001
            logger.warning("fetch_web_images: skip {}: {}", url[:120], e)
    return saved


async def download_content(
    url: str, out_path: Path, *, timeout: float = 180.0
) -> Path:
    """Скачать файл по URL; расширение — Content-Type + magic, никогда .bin."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    content_type: str | None = None
    try:
        async with _async_client(timeout=timeout, follow_redirects=True) as client:
            r = await client.get(url)
            if r.status_code >= 400:
                raise GptApiError(
                    f"download HTTP {r.status_code}: {url}",
                    context={"status_code": r.status_code, "url": url},
                )
            content_type = r.headers.get("content-type")
            out_path.write_bytes(r.content)
    except httpx.HTTPError as e:
        raise GptApiError(
            f"download сетевая ошибка: {type(e).__name__}: {e}",
            context={"url": url},
        ) from e
    if not out_path.is_file() or out_path.stat().st_size < 1:
        raise GptApiError("download: пустой файл", context={"url": url, "path": str(out_path)})
    return finalize_downloaded_file(out_path, content_type=content_type, url=url)


def copy_content(src: Path, out_path: Path) -> Path:
    """Скопировать локальный контент (когда файл уже на диске, не по URL)."""
    import shutil

    if not src.is_file():
        raise GptApiError("copy: источник не найден", context={"src": str(src)})
    out_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, out_path)
    return out_path
