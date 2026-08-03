"""Точечная правка ячейки project.xlsx из «Базы» + sync известных полей в DB."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from loguru import logger
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm.attributes import flag_modified

from app.models import Entity, Frame, Project
from app.services.db_apply import ApplyOpsError
from app.services.xlsx_v8_import import (
    ROW_DURATION_V8,
    ROW_IMAGE_PROMPT_2_V8,
    ROW_IMAGE_PROMPT_V8,
    ROW_TIMECODE_V8,
    ROW_VIDEO_PROMPT_2_V8,
    ROW_VIDEO_PROMPT_V8,
    ROW_VOICEOVER_V8,
    SHEET_PLAN_V8,
    _normalize_sheet_name,
    _resolve_plan_sheet,
)

ROW_PERSONS_PRIMARY = 8


def _find_sheet(wb: Any, sheet_name: str) -> Any | None:
    want = _normalize_sheet_name(sheet_name).casefold()
    for name in wb.sheetnames:
        if _normalize_sheet_name(name).casefold() == want:
            return wb[name]
    return None


def _cell_str(ws: Any, row: int, col: int) -> str:
    v = ws.cell(row=row, column=col).value
    if v is None:
        return ""
    return str(v).strip()


async def _sync_persons_column(
    session: AsyncSession,
    project: Project,
    ws: Any,
    col: int,
) -> str | None:
    """Синк колонки листа «Персонажи» → Entity. None если колонка пустая."""
    from app.services.excel_characters import (
        ROW_CHAR,
        ROW_CLOTHES,
        ROW_ID,
        ROW_LOOK,
        ROW_NAME,
        ROW_RULES,
    )

    code = _cell_str(ws, ROW_ID, col)
    if not code:
        return None
    name = _cell_str(ws, ROW_NAME, col)
    look = _cell_str(ws, ROW_LOOK, col)
    clothes = _cell_str(ws, ROW_CLOTHES, col)
    char = _cell_str(ws, ROW_CHAR, col)
    rules = _cell_str(ws, ROW_RULES, col)
    attrs = {
        "look": look,
        "clothes": clothes,
        "char": char,
        "rules": rules,
        "внешность": look,
        "одежда": clothes,
        "характер": char,
        "правила": rules,
    }
    existing = (
        await session.execute(
            select(Entity).where(
                Entity.project_id == project.id,
                Entity.type == "character",
                Entity.code == code,
            )
        )
    ).scalar_one_or_none()
    if existing is None:
        max_key = (
            await session.execute(
                select(Entity.sort_key)
                .where(Entity.project_id == project.id, Entity.type == "character")
                .order_by(Entity.sort_key.desc())
                .limit(1)
            )
        ).scalar_one_or_none()
        sk = float(max_key or 0) + 10.0
        session.add(
            Entity(
                project_id=project.id,
                type="character",
                code=code,
                name=name or None,
                attrs=attrs,
                sort_key=sk,
            )
        )
    else:
        existing.name = name or None
        existing.attrs = attrs
        flag_modified(existing, "attrs")
    return code


async def _reload_excel_hero_meta(
    session: AsyncSession, project: Project, xlsx: Path
) -> None:
    from app.services.excel_characters import parse_persons_sheet

    try:
        chars = parse_persons_sheet(xlsx)
    except Exception as e:  # noqa: BLE001
        logger.warning(
            "[#{}] sheet_cell: excel_hero reload failed: {}", project.id, e
        )
        return
    meta = dict(project.meta or {})
    if chars:
        meta["excel_hero"] = {"characters": [c.to_dict() for c in chars]}
    else:
        meta.pop("excel_hero", None)
    project.meta = meta
    flag_modified(project, "meta")
    await session.flush()


async def _sync_plan_cell(
    session: AsyncSession,
    project: Project,
    *,
    row: int,
    col: int,
    value: str,
) -> str | None:
    """Синк известной строки листа «план» → Frame. Возвращает label синка."""
    from app.services.plan_shot2 import SHOT2_PROMPT_ATTR, SHOT2_VIDEO_PROMPT_ATTR

    frame_number = col - 2
    if frame_number < 1:
        return None
    fr = (
        await session.execute(
            select(Frame).where(
                Frame.project_id == project.id,
                Frame.number == frame_number,
            )
        )
    ).scalar_one_or_none()
    if fr is None:
        return None

    attrs = dict(fr.attrs or {}) if isinstance(fr.attrs, dict) else {}
    synced: str | None = None
    if row == ROW_VOICEOVER_V8:
        fr.voiceover_text = value
        synced = "voiceover_text"
    elif row == ROW_IMAGE_PROMPT_V8:
        fr.image_prompt = value
        synced = "image_prompt"
    elif row == ROW_VIDEO_PROMPT_V8:
        fr.animation_prompt = value
        synced = "animation_prompt"
    elif row == ROW_IMAGE_PROMPT_2_V8:
        attrs[SHOT2_PROMPT_ATTR] = value
        fr.attrs = attrs
        flag_modified(fr, "attrs")
        synced = SHOT2_PROMPT_ATTR
    elif row == ROW_VIDEO_PROMPT_2_V8:
        attrs[SHOT2_VIDEO_PROMPT_ATTR] = value
        fr.attrs = attrs
        flag_modified(fr, "attrs")
        synced = SHOT2_VIDEO_PROMPT_ATTR
    elif row == ROW_DURATION_V8:
        try:
            fr.duration_seconds = float(value) if value.strip() else None
        except ValueError:
            fr.duration_seconds = None
        synced = "duration_seconds"
    elif row == ROW_PERSONS_PRIMARY:
        attrs["characters"] = value
        attrs["persons"] = value
        attrs["персонажи"] = value
        fr.attrs = attrs
        flag_modified(fr, "attrs")
        synced = "characters"
    elif row == ROW_TIMECODE_V8:
        # Таймкод в xlsx; start/end_ts не трогаем без парсера.
        synced = "timecode_xlsx_only"
    return synced


async def write_sheet_cell(
    session: AsyncSession,
    project: Project,
    *,
    sheet: str,
    row: int,
    col: int,
    value: str,
) -> dict[str, Any]:
    """Записать ячейку в project.xlsx и синхронизировать известные поля в DB."""
    from openpyxl import load_workbook

    from app.services.excel_characters import SHEET_PERSONS
    from app.services.xlsx_versioning import backup_to_old

    if row < 1 or col < 1:
        raise ApplyOpsError("row/col должны быть >= 1 (Excel 1-based)")
    sheet_name = (sheet or "").strip()
    if not sheet_name:
        raise ApplyOpsError("sheet обязателен")

    path = Path(project.data_dir) / "project.xlsx"
    if not path.is_file():
        raise ApplyOpsError("project.xlsx не найден — сначала сгенерируй книгу")

    try:
        snap = backup_to_old(path)
    except Exception:  # noqa: BLE001
        snap = None

    text = "" if value is None else str(value)
    wb = load_workbook(path)
    synced: str | None = None
    try:
        ws = _find_sheet(wb, sheet_name)
        if ws is None:
            # Создаём лист только если это известные имена; иначе ошибка.
            raise ApplyOpsError(f"лист {sheet_name!r} не найден в project.xlsx")
        ws.cell(row=row, column=col, value=text)
        wb.save(path)
    finally:
        wb.close()

    sheet_norm = _normalize_sheet_name(sheet_name).casefold()
    persons_norm = _normalize_sheet_name(SHEET_PERSONS).casefold()
    plan_norm = _normalize_sheet_name(SHEET_PLAN_V8).casefold()

    if sheet_norm == persons_norm and col >= 2:
        # Перечитать книгу после save для синка всей колонки.
        wb2 = load_workbook(path, data_only=True)
        try:
            ws2 = _find_sheet(wb2, sheet_name)
            if ws2 is not None:
                code = await _sync_persons_column(session, project, ws2, col)
                if code:
                    synced = f"entity:{code}"
        finally:
            wb2.close()
        await _reload_excel_hero_meta(session, project, path)
    elif sheet_norm == plan_norm or sheet_norm == "кадры":
        # Для плана имя в книге может отличаться регистром — уже нашли ws.
        synced = await _sync_plan_cell(
            session, project, row=row, col=col, value=text
        )

    await session.flush()
    logger.info(
        "[#{}] sheet_cell {}!R{}C{} synced={} backup={}",
        project.id,
        sheet_name,
        row,
        col,
        synced,
        getattr(snap, "name", None),
    )
    return {
        "ok": True,
        "sheet": sheet_name,
        "row": row,
        "col": col,
        "value": text,
        "synced": synced,
        "backup": getattr(snap, "name", None),
    }
