"""Полная логика xlsx-шагов ChatGPT — единый источник для TG-бота и воркера.

Telegram `_run_*_xlsx` и orchestrator steps вызывают одни и те же функции.
GPT-сессия (browser → new_conversation → ask_with_files → download) — в
`xlsx_gpt_flow`; здесь — подготовка файлов, валидация, backup/replace, sync.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from loguru import logger
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Frame, Project, ProjectStatus
from app.services import chatgpt_xlsx as cx
from app.services import xlsx_gpt_flow as xgf
from app.services.xlsx_versioning import (
    backup_to_old,
    normalize_xlsx_to_reference_layout,
    replace_with,
    validate_xlsx,
)
from app.services.voiceover_split_local import (
    parse_dash_separated_blocks,
    split_voiceover_locally,
    write_voiceover_blocks_to_xlsx,
)
from app.storage import for_project as _sheet_for_project

# Должен совпадать со строкой 4 в web/STUDIO_VERSION. Если в логе make_plan
# нет «xlsx_step_runners» — на диске старый make_plan.py (текст 30k в ask).
XLSX_STEP_RUNNERS_ID = "xlsx_step_runners-v82-img-pr-volume-batches"


def _plan_empty_error(xlsx_path: Path, *, plan_len: int) -> RuntimeError:
    """Понятная ошибка: импортёр читает «Общий план», GPT часто пишет в «план»."""
    sheets: list[str] = []
    try:
        from openpyxl import load_workbook

        wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
        sheets = list(wb.sheetnames)
        wb.close()
    except Exception:  # noqa: BLE001
        pass
    sheets_s = ", ".join(sheets) if sheets else "?"
    return RuntimeError(
        "лист «Общий план» пуст/шаблон после GPT "
        f"(прочитано {plan_len} симв., нужно ≥200); "
        "GPT заполнил не тот лист (часто «план» вместо «Общий план»). "
        f"Листы файла: [{sheets_s}]"
    )


def _assert_downloaded_plan_meaningful(xlsx_path: Path) -> None:
    """До replace project.xlsx — отказать, если «Общий план» не заполнен."""
    from openpyxl import load_workbook

    from app.services.plan_validation import is_meaningful_general_plan
    from app.services.xlsx_v8_import import _read_general_plan

    try:
        wb = load_workbook(filename=str(xlsx_path), data_only=True)
    except Exception as e:  # noqa: BLE001
        raise RuntimeError(f"скачанный xlsx не читается: {e}") from e
    try:
        plan_text = (_read_general_plan(wb) or "").strip()
    finally:
        wb.close()
    if not is_meaningful_general_plan(plan_text):
        raise _plan_empty_error(xlsx_path, plan_len=len(plan_text))


def _apply_split_fallback(
    xlsx_path: Path,
    voiceover_path: Path,
    *,
    gpt_reply: str,
) -> int:
    """GPT часто не пишет R49 — пробуем блоки из ответа или voiceover.txt."""
    blocks = parse_dash_separated_blocks(gpt_reply)
    if len(blocks) < 2 and voiceover_path.exists():
        blocks = split_voiceover_locally(
            voiceover_path.read_text(encoding="utf-8")
        )
    if len(blocks) < 2:
        return _count_v8_voiceover_blocks(xlsx_path)
    write_voiceover_blocks_to_xlsx(xlsx_path, blocks)
    return _count_v8_voiceover_blocks(xlsx_path)


def diagnose_split_xlsx(xlsx_path: Path) -> str:
    """Краткая диагностика для ошибок split: листы + блоки R49."""
    from openpyxl import load_workbook

    from app.services.xlsx_v8_import import (
        ROW_VOICEOVER_V8,
        _read_voiceover_blocks,
        has_v8_plan_sheet,
    )

    if not xlsx_path.exists():
        return f"файл не найден: {xlsx_path}"
    try:
        wb = load_workbook(filename=str(xlsx_path), data_only=True)
        try:
            sheets = list(wb.sheetnames)
            if not has_v8_plan_sheet(wb):
                return (
                    f"листы={sheets!r} — нет листа «план» (v8); "
                    f"voiceover должен быть в строке {ROW_VOICEOVER_V8}"
                )
            blocks = len(_read_voiceover_blocks(wb))
            return f"листы={sheets!r}, voiceover-блоков (R{ROW_VOICEOVER_V8})={blocks}"
        finally:
            wb.close()
    except Exception as e:  # noqa: BLE001
        return f"не удалось прочитать {xlsx_path.name}: {e}"


def _count_v8_voiceover_blocks(xlsx_path: Path) -> int:
    """Сколько voiceover-блоков уже записано в v8-xlsx (после разбивки)."""
    from openpyxl import load_workbook

    from app.services.xlsx_v8_import import _read_voiceover_blocks, has_v8_plan_sheet

    if not xlsx_path.exists():
        return 0
    try:
        wb = load_workbook(filename=str(xlsx_path), data_only=True)
        try:
            if not has_v8_plan_sheet(wb):
                return 0
            return len(_read_voiceover_blocks(wb))
        finally:
            wb.close()
    except Exception as e:  # noqa: BLE001
        logger.warning("split_xlsx: cannot count voiceover blocks in {}: {}", xlsx_path, e)
        return 0


def _try_reuse_split_download(
    tmp_dir: Path, proj_xlsx: Path, *, min_blocks: int = 2
) -> XlsxRoundtripResult | None:
    """Если GPT уже отдал xlsx в tmp_gpt, не дергаем ChatGPT повторно.

    Только если скачанный файл новее project.xlsx (иначе это устаревший кэш).
    """
    if not proj_xlsx.exists():
        return None
    proj_mtime = proj_xlsx.stat().st_mtime
    candidates = sorted(
        tmp_dir.glob("split_*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    for candidate in candidates[:5]:
        if candidate.stat().st_size < 1024:
            continue
        if candidate.stat().st_mtime <= proj_mtime:
            continue
        if validate_xlsx(candidate) is not None:
            continue
        blocks = _count_v8_voiceover_blocks(candidate)
        if blocks < min_blocks:
            continue
        from app.storage.plan_sheet_v8 import merge_gpt_voiceover_row_into_project

        backup = backup_to_old(proj_xlsx)
        merged = merge_gpt_voiceover_row_into_project(proj_xlsx, candidate)
        if merged < min_blocks:
            # Нет листа «план» в шаблоне / не v8 — полный replace как раньше.
            replace_with(proj_xlsx, candidate)
            merged = _count_v8_voiceover_blocks(proj_xlsx)
        logger.info(
            "split_xlsx: reuse downloaded {} (R49 merged={}, blocks={}) — skip GPT",
            candidate.name,
            merged,
            blocks,
        )
        return XlsxRoundtripResult(
            reply_text="",
            downloaded_path=candidate,
            project_xlsx=proj_xlsx,
            backup_path=backup,
        )
    return None


@dataclass
class XlsxRoundtripResult:
    """Результат GPT round-trip (DB-first; xlsx только экспорт)."""

    reply_text: str
    downloaded_path: Path
    project_xlsx: Path
    backup_path: Path | None = None
    # DB-first plan: текст общего плана из apply-ops (без скачивания xlsx).
    plan_text: str | None = None
    # split: список {закадр, длительность?}
    frames_spec: list[dict] | None = None
    # img_pr / др.: готовые ops для apply_ops
    apply_ops: list[dict] | None = None
    # True = ops уже записаны в DB по батчам внутри runner (не apply повторно).
    ops_applied_inline: bool = False


def _ts() -> str:
    return datetime.utcnow().strftime("%Y%m%d_%H%M%S")


def _ensure_project_xlsx(project: Project) -> Path:
    proj_xlsx = project.data_dir / "project.xlsx"
    if proj_xlsx.exists():
        return proj_xlsx
    sheet = _sheet_for_project(project)
    proj_xlsx = sheet.ensure_initialized(
        project_id=project.id, slug=project.slug
    )
    if not proj_xlsx.exists():
        raise FileNotFoundError(f"project.xlsx не найден: {proj_xlsx}")
    return proj_xlsx


_PLAN_DB_HINT = (
    "\n\n# ЗАПИСЬ СЦЕНАРИЯ — ИСТОЧНИК ПРАВДЫ БАЗА (НЕ Excel)\n"
    "Верни ТОЛЬКО JSON apply-ops. Без TSV, без `# Лист:`, без `@row=`, "
    "без ссылок на скачивание .xlsx:\n"
    '{"ops":[{"target":"project","fields":{"общий_план":"<полный текст сценария>"}}]}\n'
    "Текст в общем_плане — полный содержательный сценарий/план "
    "(минимум ~200 символов). Никакой прозы вокруг JSON.\n"
)


def extract_general_plan_from_gpt_reply(reply: str) -> str:
    """Достать общий_план из apply-ops JSON ответа модели."""
    from app.services import db_apply

    data = db_apply.extract_apply_ops_json(reply or "")
    if isinstance(data, dict):
        for op in data.get("ops") or []:
            if not isinstance(op, dict):
                continue
            fields = op.get("fields") or {}
            if not isinstance(fields, dict):
                continue
            for key in ("общий_план", "general_plan", "план", "сценарий"):
                val = fields.get(key)
                if isinstance(val, str) and val.strip():
                    return val.strip()
            if str(op.get("target") or "").strip().lower() == "project":
                for val in fields.values():
                    if isinstance(val, str) and len(val.strip()) >= 80:
                        return val.strip()
    return ""


async def run_plan_xlsx(
    project: Project,
    *,
    topic: str | None = None,
    project_id: int | None = None,
) -> XlsxRoundtripResult:
    """Шаг «План»: GPT → apply-ops (общий_план) → DB; Excel только экспорт позже.

    Больше не скачиваем .xlsx / TSV writeback (HTML-страницы ломали шаг).
    """
    from app.services.plan_validation import is_meaningful_general_plan

    proj_xlsx = _ensure_project_xlsx(project)
    actual_topic = topic if topic is not None else (project.topic or "")

    ts = _ts()
    tmp_dir = cx.tmp_gpt_dir(project)
    prompt_file = cx.write_plan_prompt_file(
        project, tmp_dir, topic=actual_topic, ts=ts
    )
    chat_msg = cx.chat_message(
        project, "plan", topic=actual_topic, prompt_file_name=prompt_file.name
    )
    chat_msg = f"{chat_msg}{_PLAN_DB_HINT}"

    logger.info(
        "plan_db: prompt_file={} ({} байт), chat_len={} (без xlsx-download)",
        prompt_file.name,
        prompt_file.stat().st_size,
        len(chat_msg),
    )

    async def _gpt() -> str:
        # Промт + xlsx только как контекст; запись — JSON apply-ops, не файл.
        return await xgf.telegram_style_ask_with_files(
            chat_msg,
            [prompt_file, proj_xlsx],
            project_id=project_id or project.id,
        )

    reply = await xgf.run_under_xlsx_lock(project.id, "plan", _gpt)
    plan_text = extract_general_plan_from_gpt_reply(reply)
    if not is_meaningful_general_plan(plan_text):
        raise RuntimeError(
            "GPT не вернул общий_план в apply-ops JSON "
            f"(досталось {len(plan_text)} символов). Нужен формат: "
            '{"ops":[{"target":"project","fields":{"общий_план":"…"}}]}'
        )

    logger.info("plan_db: общий_план len={}", len(plan_text))
    return XlsxRoundtripResult(
        reply_text=reply,
        downloaded_path=proj_xlsx,
        project_xlsx=proj_xlsx,
        backup_path=None,
        plan_text=plan_text,
    )


_SCRIPT_DB_HINT = (
    "\n\n# ЗАПИСЬ ЗАКАДРА — ИСТОЧНИК ПРАВДЫ БАЗА (НЕ Excel)\n"
    "Верни ТОЛЬКО JSON apply-ops ИЛИ блок <<<VOICEOVER>>>…<<<END>>>:\n"
    '{"ops":[{"target":"project","fields":{"закадровый_текст":"<полный закадр>"}}]}\n'
    "Без TSV, без `# Лист:`, без скачивания файлов.\n"
)


async def run_script_xlsx(
    project: Project,
    *,
    project_id: int | None = None,
) -> tuple[XlsxRoundtripResult, str]:
    """Шаг «Закадровый текст»: GPT → текст/apply-ops → DB+voiceover.txt (без download)."""
    from app.services import db_apply
    from app.services.voiceover_sanitize import (
        ensure_voiceover_format_instruction,
        extract_voiceover_block,
    )

    proj_xlsx = _ensure_project_xlsx(project)
    source_voiceover = cx.ensure_current_voiceover(project)

    ts = _ts()
    tmp_dir = cx.tmp_gpt_dir(project)
    prompt_file = cx.write_script_prompt_file(project, tmp_dir, ts=ts)
    chat_msg = ensure_voiceover_format_instruction(
        cx.chat_message(project, "script", prompt_file_name=prompt_file.name)
    )
    chat_msg = f"{chat_msg}{_SCRIPT_DB_HINT}"
    attach_files: list[Path] = [prompt_file]
    # Контекст: общий план / прошлый закадр — без ожидания скачивания xlsx.
    if (project.general_plan or "").strip():
        gp = tmp_dir / f"general_plan_{ts}.txt"
        gp.write_text(project.general_plan or "", encoding="utf-8")
        attach_files.append(gp)
    if source_voiceover is not None:
        attach_files.append(source_voiceover)

    logger.info(
        "script_db: prompt_file={} chat_len={} (без xlsx-download)",
        prompt_file.name,
        len(chat_msg),
    )

    async def _gpt() -> str:
        return await xgf.telegram_style_ask_with_files(
            chat_msg,
            attach_files,
            timeout=1800.0,
            project_id=project_id or project.id,
        )

    reply = await xgf.run_under_xlsx_lock(project.id, "script", _gpt)

    voiceover_text = ""
    data = db_apply.extract_apply_ops_json(reply or "")
    if isinstance(data, dict):
        for op in data.get("ops") or []:
            if not isinstance(op, dict):
                continue
            fields = op.get("fields") or {}
            if not isinstance(fields, dict):
                continue
            for key in (
                "закадровый_текст",
                "script_text",
                "сценарий",
                "voiceover",
                "общий_план",
            ):
                # общий_план тут не пишем в VO — только script aliases
                if key == "общий_план":
                    continue
                val = fields.get(key)
                if isinstance(val, str) and len(val.strip()) >= 80:
                    voiceover_text = val.strip()
                    break
            if voiceover_text:
                break
    if not voiceover_text:
        voiceover_text = (extract_voiceover_block(reply) or "").strip()
    if not voiceover_text:
        # Последний шанс: весь ответ, если это не голый JSON
        raw = (reply or "").strip()
        if raw and '"ops"' not in raw[:40] and len(raw) >= 200:
            voiceover_text = raw
    if len(voiceover_text) < 200:
        raise RuntimeError(
            "GPT не вернул закадр (apply-ops закадровый_текст или "
            f"<<<VOICEOVER>>>, len={len(voiceover_text)})"
        )

    voiceover_text = cx.save_voiceover_text(
        project, proj_xlsx.parent / "voiceover.txt", voiceover_text
    )

    return (
        XlsxRoundtripResult(
            reply_text=reply,
            downloaded_path=proj_xlsx.parent / "voiceover.txt",
            project_xlsx=proj_xlsx,
            apply_ops=[
                {
                    "target": "project",
                    "fields": {"закадровый_текст": voiceover_text},
                }
            ],
        ),
        voiceover_text,
    )


_SPLIT_DB_HINT = (
    "\n\n# РАЗБИВКА — ИСТОЧНИК ПРАВДЫ БАЗА (НЕ Excel)\n"
    "Верни ТОЛЬКО JSON apply-ops. Без TSV, без `# Лист:`, без `@row=`, "
    "без скачивания .xlsx:\n"
    '{"ops":[{"target":"replace_frames","frames":['
    '{"закадр":"текст кадра 1","длительность":3},'
    '{"закадр":"текст кадра 2"}'
    "]}]}\n"
    "Нужно ≥2 кадра. Каждый кадр — отдельный объект с полем закадр.\n"
)


def extract_frames_spec_from_gpt_reply(reply: str, *, voiceover_path: Path | None) -> list[dict]:
    """Достать кадры из replace_frames JSON или --- блоков / локальной разбивки."""
    from app.services import db_apply

    data = db_apply.extract_apply_ops_json(reply or "")
    if isinstance(data, dict):
        for op in data.get("ops") or []:
            if not isinstance(op, dict):
                continue
            if str(op.get("target") or "") != "replace_frames":
                continue
            raw = op.get("frames") or op.get("кадры") or []
            if isinstance(raw, list) and len(raw) >= 2:
                out: list[dict] = []
                for item in raw:
                    if isinstance(item, str) and item.strip():
                        out.append({"закадр": item.strip()})
                    elif isinstance(item, dict):
                        out.append(item)
                if len(out) >= 2:
                    return out
    blocks = parse_dash_separated_blocks(reply or "")
    if len(blocks) < 2 and voiceover_path is not None and voiceover_path.exists():
        blocks = split_voiceover_locally(
            voiceover_path.read_text(encoding="utf-8", errors="replace")
        )
    if len(blocks) >= 2:
        return [{"закадр": b.strip()} for b in blocks if b.strip()]
    return []


async def run_split_xlsx(
    project: Project,
    *,
    project_id: int | None = None,
) -> XlsxRoundtripResult:
    """Шаг «Разбивка»: GPT → replace_frames (DB); Excel только через export."""
    proj_xlsx = _ensure_project_xlsx(project)
    voiceover = cx.ensure_current_voiceover(project)
    if voiceover is None:
        raise FileNotFoundError(
            "voiceover.txt не найден — сначала пройди шаг «Закадровый текст»"
        )

    ts = _ts()
    tmp_dir = cx.tmp_gpt_dir(project)
    prompt_file = cx.write_split_prompt_file(project, tmp_dir, ts=ts)
    chat_msg = (
        cx.chat_message(project, "split", prompt_file_name=prompt_file.name)
        + _SPLIT_DB_HINT
    )

    logger.info(
        "split_db: prompt={}, voiceover={}, chat_len={} (без xlsx-download)",
        prompt_file.name,
        voiceover.name,
        len(chat_msg),
    )

    async def _gpt() -> str:
        return await xgf.telegram_style_ask_with_files(
            chat_msg,
            [prompt_file, voiceover],
            project_id=project_id or project.id,
        )

    reply = await xgf.run_under_xlsx_lock(project.id, "split", _gpt)
    frames_spec = extract_frames_spec_from_gpt_reply(reply, voiceover_path=voiceover)
    if len(frames_spec) < 2:
        raise RuntimeError(
            "GPT не вернул разбивку в apply-ops replace_frames "
            f"(кадров={len(frames_spec)}). Нужен JSON "
            '{"ops":[{"target":"replace_frames","frames":[...]}]}'
        )

    logger.info("split_db: кадров из GPT/fallback={}", len(frames_spec))
    return XlsxRoundtripResult(
        reply_text=reply,
        downloaded_path=proj_xlsx,
        project_xlsx=proj_xlsx,
        backup_path=None,
        frames_spec=frames_spec,
        apply_ops=[{"target": "replace_frames", "frames": frames_spec}],
    )


_IMG_PR_DB_HINT = (
    "\n\n# ПРОМТЫ КАРТИНОК — ИСТОЧНИК ПРАВДЫ БАЗА (НЕ Excel)\n"
    "Верни ТОЛЬКО JSON apply-ops. Без TSV, без `# Лист:`, без `@row=`, "
    "без скачивания .xlsx:\n"
    '{"ops":[{"frame_uuid":"<uuid>","fields":{"промт_картинки":"…"}}]}\n'
    "Адрес кадра — ТОЛЬКО frame_uuid из db_frames.json ЭТОГО батча.\n"
    "Одна операция = один кадр. Пройди все кадры из текущего db_frames.json.\n"
    "Сборка кадра (порядок): ref(cXX?) → shot01_bg → shot01_action → "
    "lighting/scene_lighting → accent → scene_sense → scene_feature → "
    "shot01_description/props → place/время → STYLE. "
    "accent/scene_sense/scene_feature — отдельные строки. "
    "В промт_картинки пиши ТОЛЬКО сцену — STYLE LOCK НЕ копируй "
    "(пайплайн допишет). characters[] Entity. Не копируй voiceover_text.\n"
)

_PLASTILIN_IMG_PR_HINT = (
    "\n\n# ПЛАСТИЛИН — ЗАПИСЬ В БАЗУ (НЕ Excel)\n"
    "Верни ТОЛЬКО JSON apply-ops. Без TSV, без `# Лист:`, без `@row=`, "
    "без скачивания .xlsx:\n"
    '{"ops":[{"frame_uuid":"<uuid>","fields":{"промт_картинки":"…","персонажи":"c01"}}]}\n'
    "Адрес кадра — ТОЛЬКО frame_uuid из db_frames.json ЭТОГО батча.\n"
    "Одна операция = один кадр. Пройди все кадры из текущего файла.\n"
    "В `промт_картинки` пиши ПОЛНЫЙ промт: стиль пластилина ТРИ раза + сцена "
    "+ Negative. Пайплайн НЕ допишет watercolor/noir — не копируй Archival Noir. "
    "Не копируй voiceover_text. Без ASCII двойных кавычек в тексте промта.\n"
)


async def _load_img_pr_context(
    project: Project,
    *,
    only_uuids: set[str] | None = None,
    skip_uuids: set[str] | None = None,
) -> tuple[list[Frame], list[dict], str]:
    """Кадры + Entity cards + general_plan для img_pr."""
    from app.db import SessionLocal
    from app.models import Entity
    from app.services import db_v2
    from app.services.excel_characters import entity_cards_for_gpt

    async with SessionLocal() as session:
        proj = await session.get(Project, project.id)
        if proj is None:
            raise RuntimeError(f"project #{project.id} not found for img_pr db_frames")
        await db_v2.backfill_project_v2(session, proj)
        frames = list(
            (
                await session.execute(
                    select(Frame)
                    .where(Frame.project_id == proj.id)
                    .order_by(Frame.number)
                )
            ).scalars().all()
        )
        ents = list(
            (
                await session.execute(
                    select(Entity).where(Entity.project_id == proj.id)
                )
            ).scalars().all()
        )
        general_plan = proj.general_plan or ""
        cards = entity_cards_for_gpt(ents)

    selected: list[Frame] = []
    for fr in frames:
        uuid = (fr.uuid or "").strip()
        if not uuid:
            continue
        if only_uuids is not None and uuid not in only_uuids:
            continue
        if skip_uuids and uuid in skip_uuids:
            continue
        if not (fr.voiceover_text or "").strip():
            continue
        # Уже заполненные пропускаем (resume / soft retry).
        if (fr.image_prompt or "").strip():
            continue
        selected.append(fr)
    return selected, cards, general_plan


def _write_img_pr_db_frames_for(
    project: Project,
    tmp_dir: Path,
    frames: list[Frame],
    characters: list[dict],
    general_plan: str,
    *,
    batch_tag: str = "",
    include_characters: bool = True,
) -> Path:
    import json

    from app.services.db_frames_context import build_img_pr_db_context

    ctx = build_img_pr_db_context(
        project_id=project.id,
        slug=project.slug or "",
        frames=frames,
        characters=characters,
        general_plan=general_plan if include_characters else "",
        include_characters=include_characters,
        include_field_map=include_characters,
    )
    name = f"db_frames{('_' + batch_tag) if batch_tag else ''}.json"
    out = tmp_dir / name
    # Компактно: меньше шанс упереться в trim 60k в gpt_api.file_to_context.
    out.write_text(
        json.dumps(ctx, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    logger.info(
        "img_pr_db: {} frames={} chars={} bytes={}",
        out.name,
        len(ctx.get("frames") or []),
        len(ctx.get("characters") or []),
        out.stat().st_size,
    )
    return out


async def _write_img_pr_db_frames(project: Project, tmp_dir: Path) -> Path:
    """Пишет полный db_frames.json (smoke / legacy)."""
    frames, cards, general_plan = await _load_img_pr_context(project)
    return _write_img_pr_db_frames_for(
        project, tmp_dir, frames, cards, general_plan
    )


async def _apply_img_pr_ops_now(
    project: Project,
    ops: list[dict],
    *,
    export_xlsx: bool = False,
    label: str = "",
) -> None:
    """Сразу записать apply-ops батча в DB (Excel — только явный Export)."""
    if not ops:
        return
    from app.db import SessionLocal
    from app.services import db_apply

    async with SessionLocal() as session:
        proj = await session.get(Project, project.id)
        if proj is None:
            raise RuntimeError(f"project #{project.id} gone during img_pr apply")
        await db_apply.apply_ops(
            session, proj, ops, export_xlsx=export_xlsx, node_kind="img_pr"
        )
        await session.commit()
    logger.info(
        "img_pr_db: applied to DB ops={} export_xlsx={} {}",
        len(ops),
        export_xlsx,
        label,
    )


async def run_img_pr_xlsx(
    project: Project,
    *,
    n_frames: int | None = None,
    project_id: int | None = None,
    uuid_map_text: str = "",
) -> XlsxRoundtripResult:
    """Шаг «Промты картинок»: GPT батчами → apply-ops (DB)."""
    from app.services import img_pr_batches as ipb

    proj_xlsx = _ensure_project_xlsx(project)
    tmp_dir = cx.tmp_gpt_dir(project)
    prompt_file = cx.write_img_pr_prompt_file(project, tmp_dir, ts=_ts())
    from app.services.img_pr_style import is_plastilin_master, resolve_project_img_style

    master_head = ""
    try:
        master_head = prompt_file.read_text(encoding="utf-8")[:2000]
    except OSError:
        master_head = ""
    plastilin = is_plastilin_master(prompt_file.name, master_head)
    style_id = resolve_project_img_style(
        project, variant=prompt_file.name, master=master_head
    )
    logger.info("img_pr_db: style_id={!r} plastilin={}", style_id, plastilin)
    img_pr_hint = _PLASTILIN_IMG_PR_HINT if plastilin else _IMG_PR_DB_HINT
    if plastilin:
        logger.info("img_pr_db: plastilin master — keep clay style in prompt, no watercolor wrap")

    ckpt = ipb.load_checkpoint(project.data_dir)
    done_uuids = list(ckpt.get("done_uuids") or [])
    all_ops: list[dict] = list(ckpt.get("ops") or [])
    done_set = set(done_uuids)

    # НЕ пишем в DB по батчам (SQLite lock). Чекпоинт на диске → apply один раз в конце.
    frames, cards, general_plan = await _load_img_pr_context(
        project, skip_uuids=done_set or None
    )
    if not frames:
        if all_ops:
            logger.info(
                "img_pr_db: checkpoint complete ops={} — apply once at end",
                len(all_ops),
            )
            return XlsxRoundtripResult(
                reply_text="(from checkpoint)",
                downloaded_path=proj_xlsx,
                project_xlsx=proj_xlsx,
                backup_path=None,
                apply_ops=all_ops,
                ops_applied_inline=False,
            )
        # Уже всё в DB с прошлого успешного прогона.
        frames_db, _, _ = await _load_img_pr_context(project)
        if not frames_db:
            logger.info("img_pr_db: nothing to do — all frames already have prompts")
            return XlsxRoundtripResult(
                reply_text="(already in DB)",
                downloaded_path=proj_xlsx,
                project_xlsx=proj_xlsx,
                backup_path=None,
                apply_ops=[],
                ops_applied_inline=True,
            )
        raise RuntimeError(
            "нет кадров без image_prompt для img_pr "
            f"(done_checkpoint={len(done_set)})"
        )

    from collections import deque

    from app.services.volume_batches import (
        inferred_batch_size,
        plan_remainder_batches,
        rechunk_tail,
    )

    batch_size = ipb.plan_batch_size(len(frames))
    work: deque[list] = deque(ipb.chunk_frames(frames, size=batch_size))
    logger.info(
        "img_pr_db: ONE session + ~{} batches×~{} (frames={} checkpoint_done={})",
        len(work),
        batch_size,
        len(frames),
        len(done_set),
    )

    vo = cx.ensure_current_voiceover(project)
    replies: list[str] = []
    api_batches = 0

    from app.services.gpt_client import get_gpt_client
    from app.services.step_cancel import raise_if_cancelled

    gpt = get_gpt_client()

    async def _run_batches() -> None:
        nonlocal done_uuids, all_ops, done_set, batch_size, api_batches
        await gpt.new_conversation()
        history: list[dict[str, str]] = []
        bi = 0
        first_attach = True

        while work:
            raise_if_cancelled(project.id)
            batch = work.popleft()
            if not batch:
                continue
            bi += 1
            batch_n = bi + len(work)
            batch_tag = f"b{bi:02d}"
            db_path = _write_img_pr_db_frames_for(
                project,
                tmp_dir,
                batch,
                cards,
                general_plan,
                batch_tag=batch_tag,
                include_characters=True,
            )
            uuid_lines = "\n".join(
                f"кадр {fr.number} = {fr.uuid}" for fr in batch if fr.uuid
            )
            footer = ipb.batch_footer(
                batch_i=bi, batch_n=batch_n, n=len(batch), plastilin=plastilin
            )
            if first_attach:
                chat_msg = cx.chat_message(
                    project,
                    "img_pr",
                    prompt_file_name=prompt_file.name,
                    n_frames=len(batch),
                )
                chat_msg = (
                    f"{chat_msg}{img_pr_hint}\n{footer}\n"
                    f"Адресация:\n{uuid_lines}\n"
                )
                if uuid_map_text.strip():
                    chat_msg = (
                        f"{chat_msg}\n# uuid map (справочно)\n"
                        f"{uuid_map_text.strip()[:2000]}\n"
                    )
                attach = ipb.batch_attach_files(
                    batch_i=1,
                    prompt_file=prompt_file,
                    db_path=db_path,
                    voiceover=vo,
                )
                treat_txt = True
                use_history: list[dict[str, str]] | None = None
            else:
                chat_msg = (
                    f"{ipb.followup_message(batch_i=bi, batch_n=batch_n, n=len(batch), plastilin=plastilin)}\n"
                    f"Адресация:\n{uuid_lines}\n"
                    "Файл db_frames полный — не отказывайся из‑за «обрезки». "
                    "Верни JSON ops на ВСЕ uuid из файла.\n"
                )
                attach = ipb.batch_attach_files(
                    batch_i=bi,
                    prompt_file=prompt_file,
                    db_path=db_path,
                    voiceover=vo,
                )
                treat_txt = True
                use_history = history or None

            batch_ops: list[dict] = []
            last_reply = ""
            for attempt in range(1, ipb._GPT_ATTEMPTS + 1):
                # После фейла follow-up — свежая сессия + короткий master,
                # чтобы JSON не резался в хвосте длинной history.
                if attempt > 1 and not first_attach:
                    await gpt.new_conversation()
                    history.clear()
                    use_history = None
                    treat_txt = True
                    attach = ipb.batch_attach_files(
                        batch_i=bi,
                        prompt_file=prompt_file,
                        db_path=db_path,
                        voiceover=None,
                    )
                    chat_msg = (
                        f"{img_pr_hint}\n{footer}\n"
                        f"Адресация:\n{uuid_lines}\n"
                        "Только JSON apply-ops. "
                        + (
                            "Стиль пластилина оставь в промт_картинки.\n"
                            if plastilin
                            else "STYLE LOCK не пиши.\n"
                        )
                    )
                    logger.info(
                        "img_pr_db: batch {}/{} fresh session retry",
                        bi,
                        batch_n,
                    )
                logger.info(
                    "img_pr_db: batch {}/{} attempt {} frames={} attach={} "
                    "history={} bytes={}",
                    bi,
                    batch_n,
                    attempt,
                    [fr.number for fr in batch],
                    [p.name for p in attach],
                    len(use_history or []),
                    db_path.stat().st_size,
                )
                last_reply = await gpt.ask_with_files(
                    chat_msg,
                    attach,
                    project_id=project_id or project.id,
                    expect_file_download=False,
                    history=use_history,
                    treat_txt_as_prompt=treat_txt,
                )
                replies.append(last_reply or "")
                batch_ops = ipb.parse_img_pr_ops(
                    last_reply or "",
                    wrap_style=not plastilin,
                    style_id=style_id,
                )
                if batch_ops:
                    break
                rej = ipb.write_rejected_reply(
                    tmp_dir,
                    batch_i=bi,
                    attempt=attempt,
                    reply=last_reply or "",
                    reason="no_prompt_ops",
                )
                logger.warning(
                    "img_pr_db: batch {}/{} attempt {} failed reply_len={} {}",
                    bi,
                    batch_n,
                    attempt,
                    len(last_reply or ""),
                    rej.name,
                )

            if not batch_ops:
                if all_ops:
                    logger.error(
                        "img_pr_db: batch {}/{} failed — возвращаю partial "
                        "ops={} (чекпоинт есть, soft retry добьёт)",
                        bi,
                        batch_n,
                        len(all_ops),
                    )
                    return
                raise RuntimeError(
                    f"img_pr batch {bi}/{batch_n}: нет apply-ops "
                    f"(reply_len={len(last_reply or '')}). "
                    f"Смотри tmp_gpt/img_pr_rejected_b{bi}_*.txt"
                )

            first_attach = False
            api_batches += 1

            # В историю — коротко, без гигантского JSON ops.
            history.append({"role": "user", "content": chat_msg[:2000]})
            history.append(
                {
                    "role": "assistant",
                    "content": f"OK batch {bi}/{batch_n}: {len(batch_ops)} ops.",
                }
            )

            expected = {
                (fr.uuid or "").strip() for fr in batch if (fr.uuid or "").strip()
            }
            got_uuids = {
                ipb.uuid_of_op(op) for op in batch_ops if ipb.uuid_of_op(op)
            } & expected
            for u in sorted(got_uuids):
                if u and u not in done_set:
                    done_uuids.append(u)
                    done_set.add(u)
            # Только ops по uuid этого батча (лишние uuid модели отбрасываем).
            kept_ops = [
                op for op in batch_ops if ipb.uuid_of_op(op) in got_uuids
            ]
            all_ops.extend(kept_ops)
            ipb.save_checkpoint(
                project.data_dir, done_uuids=done_uuids, ops=all_ops
            )
            missing_uuids = expected - got_uuids
            logger.info(
                "img_pr_db: batch {}/{} ops=+{} total={} missing={} "
                "(checkpoint only, DB apply once at end)",
                bi,
                batch_n,
                len(kept_ops),
                len(all_ops),
                len(missing_uuids),
            )

            # Частичный ответ = объём превышен → добор / нарезка ~80%.
            if missing_uuids and got_uuids:
                missing_frames = [
                    fr
                    for fr in batch
                    if (fr.uuid or "").strip() in missing_uuids
                ]
                delivered = len(got_uuids)
                extras = plan_remainder_batches(
                    missing_frames, delivered=delivered
                )
                new_size = inferred_batch_size(delivered)
                tail: list = []
                while work:
                    tail.extend(work.popleft())
                requeued = extras + rechunk_tail(tail, size=new_size)
                work.extend(requeued)
                batch_size = new_size
                logger.warning(
                    "img_pr_db: volume partial got={}/{} → remainder_batches={} "
                    "new_batch_size={} (queue_left={})",
                    delivered,
                    len(expected),
                    len(extras),
                    new_size,
                    len(work),
                )
            elif missing_uuids and not got_uuids:
                logger.warning(
                    "img_pr_db: batch {} — ops без uuid батча, missing={}",
                    bi,
                    len(missing_uuids),
                )

    await xgf.run_under_xlsx_lock(project.id, "img_pr", _run_batches)

    if not all_ops:
        raise RuntimeError(
            "GPT не вернул apply-ops с промт_картинки по frame_uuid. "
            'Нужен {"ops":[{"frame_uuid":"…","fields":{"промт_картинки":"…"}}]}'
        )

    logger.info(
        "img_pr_db: complete ops={} api_batches={} — single DB apply next",
        len(all_ops),
        api_batches,
    )
    return XlsxRoundtripResult(
        reply_text="\n\n---\n\n".join(replies) if replies else "",
        downloaded_path=proj_xlsx,
        project_xlsx=proj_xlsx,
        backup_path=None,
        apply_ops=all_ops,
        ops_applied_inline=False,
    )


async def sync_after_plan(
    session: AsyncSession, project: Project, xlsx_path: Path
) -> None:
    await cx.sync_project_xlsx(session, project, xlsx_path, keep_fields=False)
    from app.services.plan_validation import is_meaningful_general_plan

    plan_text = (project.general_plan or "").strip()
    if not is_meaningful_general_plan(plan_text):
        logger.warning(
            "[#{}] sync_after_plan: general_plan len={} path={}",
            project.id,
            len(plan_text),
            xlsx_path,
        )
        raise _plan_empty_error(xlsx_path, plan_len=len(plan_text))


async def sync_after_split(
    session: AsyncSession, project: Project, xlsx_path: Path
) -> dict | None:
    return await cx.sync_project_xlsx(
        session,
        project,
        xlsx_path,
        keep_fields=False,
        update_frames_voiceover=True,
    )


async def sync_after_img_pr(
    session: AsyncSession, project: Project, xlsx_path: Path
) -> None:
    await cx.sync_project_xlsx(session, project, xlsx_path, keep_fields=False)
    from app.services.xlsx_v8_import import apply_v8_image_prompts_from_xlsx

    applied = await apply_v8_image_prompts_from_xlsx(session, project, xlsx_path)
    if applied:
        logger.info(
            "[#{}] sync_after_img_pr: image_prompt из xlsx для кадров {}",
            project.id,
            applied,
        )
    from app.services.plan_shot2 import (
        SHOT2_PROMPT_ATTR,
        SHOT2_STATUS_ATTR,
        read_shot2_columns,
    )

    frames = (
        await session.execute(
            select(Frame)
            .where(Frame.project_id == project.id)
            .order_by(Frame.number)
        )
    ).scalars().all()
    by_num = read_shot2_columns(xlsx_path)
    shot2_n = 0
    for fr in frames:
        info = by_num.get(fr.number)
        if info is None or not info.has_shot2:
            continue
        attrs = dict(fr.attrs or {})
        attrs[SHOT2_PROMPT_ATTR] = info.prompt
        if SHOT2_STATUS_ATTR not in attrs:
            attrs[SHOT2_STATUS_ATTR] = "image_prompt_ready"
        fr.attrs = attrs
        shot2_n += 1
    if shot2_n:
        await session.flush()
        logger.info(
            "[#{}] sync_after_img_pr: shot_02 промты для {} кадров",
            project.id,
            shot2_n,
        )


def set_status_if_behind(
    project: Project, target: ProjectStatus
) -> None:
    """Ставит статус, если текущий «ниже» target (как в bot после xlsx)."""
    from app.telegram.menu import status_order as _ord

    if _ord(project.status) < _ord(target):
        project.status = target
