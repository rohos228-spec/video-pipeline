"""Второй кадр сцены (shot_02) на листе «план» v8.

Строки 16–29 — описание shot_02 (enrich_1).
Строка 46 — промт для картинки 2.
Файл на диске: ``frame_NNN_s2_<uuid>.png``; референс — ``frame_NNN_*.png`` (без ``_s2_``).
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from loguru import logger
from openpyxl import load_workbook

from app.generation_options import is_skippable_empty_prompt
from app.services.xlsx_v8_import import (
    ROW_IMAGE_PROMPT_V8,
    ROW_VOICEOVER_V8,
    _cell_text,
    _resolve_plan_sheet,
)

ROW_IMAGE_PROMPT_2_V8 = 46
ROW_VIDEO_PROMPT_2_V8 = 64  # промт для видео shot_02 (аналог R48 для shot_01)
ROW_SHOT2_ID_SHOT_V8 = 18
ROW_SHOT2_ACTION_V8 = 29

SHOT2_PROMPT_ATTR = "image_prompt_shot2"
SHOT2_STATUS_ATTR = "shot2_status"
SHOT2_VIDEO_PROMPT_ATTR = "animation_prompt_shot2"
SHOT2_VIDEO_STATUS_ATTR = "shot2_video_status"
MIN_SHOT2_VIDEO_PROMPT_LEN = 10


@dataclass(frozen=True)
class Shot2ColumnInfo:
    frame_number: int
    prompt: str
    has_shot2: bool


def plan_column_for_frame(frame_number: int) -> int:
    return frame_number + 2


def _shot2_block_filled(ws, col: int) -> bool:
    """Есть ли содержательное описание shot_02 в блоке строк 16–29."""
    shot_id = (_cell_text(ws, ROW_SHOT2_ID_SHOT_V8, col) or "").strip().lower()
    if shot_id in ("shot_02", "shot02", "02", "2"):
        return True
    action = (_cell_text(ws, ROW_SHOT2_ACTION_V8, col) or "").strip()
    if len(action) >= 15:
        return True
    for row in range(16, 30):
        if row in (ROW_SHOT2_ID_SHOT_V8, ROW_SHOT2_ACTION_V8):
            continue
        if len((_cell_text(ws, row, col) or "").strip()) >= 8:
            return True
    return False


def read_shot2_columns(xlsx_path: Path) -> dict[int, Shot2ColumnInfo]:
    """frame_number → данные shot_02 из xlsx."""
    out: dict[int, Shot2ColumnInfo] = {}
    if not xlsx_path.is_file():
        return out
    try:
        wb = load_workbook(filename=str(xlsx_path), data_only=True)
    except Exception as e:  # noqa: BLE001
        logger.warning("read_shot2_columns: openpyxl {}: {}", xlsx_path, e)
        return out
    try:
        ws = _resolve_plan_sheet(wb)
        if ws is None:
            return out
        max_col = ws.max_column or 0
        if max_col < 3:
            return out
        frame_no = 0
        for col in range(3, max_col + 1):
            voice = _cell_text(ws, ROW_VOICEOVER_V8, col)
            if voice is None:
                continue
            frame_no += 1
            prompt_2 = (_cell_text(ws, ROW_IMAGE_PROMPT_2_V8, col) or "").strip()
            block = _shot2_block_filled(ws, col)
            has = bool(prompt_2) or block
            if has and not prompt_2 and block:
                action = (_cell_text(ws, ROW_SHOT2_ACTION_V8, col) or "").strip()
                prompt_2 = action
            if is_skippable_empty_prompt(prompt_2):
                prompt_2 = ""
            out[frame_no] = Shot2ColumnInfo(
                frame_number=frame_no,
                prompt=prompt_2,
                has_shot2=bool(prompt_2),
            )
    finally:
        wb.close()
    return out


def shot2_file_pattern(frame_number: int) -> str:
    return f"frame_{frame_number:03d}_s2_*.png"


def disk_has_shot2_image(scenes_dir: Path, frame_number: int) -> bool:
    if not scenes_dir.is_dir():
        return False
    return any(scenes_dir.glob(shot2_file_pattern(frame_number)))


def find_shot2_image(scenes_dir: Path, frame_number: int) -> Path | None:
    """Последний PNG shot_02 (``frame_NNN_s2_*.png``)."""
    if not scenes_dir.is_dir():
        return None
    candidates = list(scenes_dir.glob(shot2_file_pattern(frame_number)))
    if not candidates:
        return None
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0]


def shot2_video_file_pattern(frame_number: int) -> str:
    return f"clip_{frame_number:03d}_s2_*.mp4"


def disk_has_shot2_video(videos_dir: Path, frame_number: int) -> bool:
    if not videos_dir.is_dir():
        return False
    return any(videos_dir.glob(shot2_video_file_pattern(frame_number)))


def find_shot1_image(scenes_dir: Path, frame_number: int) -> Path | None:
    """Последний PNG первого кадра (без ``_s2_`` в имени)."""
    if not scenes_dir.is_dir():
        return None
    candidates = [
        p
        for p in scenes_dir.glob(f"frame_{frame_number:03d}_*.png")
        if "_s2_" not in p.name
    ]
    if not candidates:
        return None
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0]


def effective_shot_from_artifact(
    meta: dict | None, path: str | Path
) -> int:
    """shot_02 по ``_s2_`` в имени важнее meta (дефолт meta=1 ломал skip shot_01)."""
    if "_s2_" in Path(path).name:
        return 2
    meta_shot = (meta or {}).get("shot", 1)
    return meta_shot if meta_shot in (1, 2) else 1


def apply_shot2_prompts_to_frames(
    frames: list[Any],
    xlsx_path: Path,
) -> int:
    """Записать ``image_prompt_shot2`` / ``shot2_status`` в ``frame.attrs``."""
    by_num = read_shot2_columns(xlsx_path)
    changed = 0
    for fr in frames:
        info = by_num.get(fr.number)
        if info is None or not info.has_shot2:
            continue
        attrs = dict(fr.attrs or {})
        if attrs.get(SHOT2_PROMPT_ATTR) != info.prompt:
            attrs[SHOT2_PROMPT_ATTR] = info.prompt
            changed += 1
        if SHOT2_STATUS_ATTR not in attrs:
            attrs[SHOT2_STATUS_ATTR] = "image_prompt_ready"
            changed += 1
        fr.attrs = attrs
    return changed
