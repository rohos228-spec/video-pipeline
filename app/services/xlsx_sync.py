"""Синхронизация xlsx → БД: читает project.xlsx, который пользователь правил
руками, и подтягивает изменения в SQLite.

Что переносится:
  - Лист «Общий план ролика»:
      «Общий план (от ChatGPT)»  → project.general_plan
      «Закадровый текст (от ChatGPT)» → project.script_text
      «Описание героя»           → project.hero_description
  - Лист «Кадры», для каждого столбца N=2,3,...:
      R28 (логика кадра)         → frame.meaning
      R29 (промт картинки)       → frame.image_prompt
      R30 (промт видео)          → frame.animation_prompt
      R31 (время видео)          → frame.duration_seconds
      R32 (закадровый текст)     → frame.voiceover_text
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from loguru import logger
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Frame, Project
from app.storage.project_sheet import (
    ROW_FRAME_LOGIC,
    ROW_HEADER,
    ROW_IMAGE_PROMPT,
    ROW_VIDEO_DURATION,
    ROW_VIDEO_PROMPT,
    ROW_VOICEOVER,
    SHEET_FRAMES,
    SHEET_GENERAL,
)

_GENERAL_LABEL_TO_FIELD = {
    "Общий план (от ChatGPT)": "general_plan",
    "Закадровый текст (от ChatGPT)": "script_text",
    # Старая метка (совместимость с ранее созданными xlsx-файлами):
    "Сценарий (от ChatGPT)": "script_text",
    "Описание героя": "hero_description",
}


def _to_str(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _to_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


async def reload_from_xlsx(
    session: AsyncSession, project: Project, xlsx_path: Path
) -> dict[str, Any]:
    """Читает xlsx и обновляет БД. Возвращает summary
    {project_fields_changed: [...], frames_changed: [n, ...]}."""
    from openpyxl import load_workbook

    if not xlsx_path.exists():
        return {"error": f"файл не найден: {xlsx_path}"}

    wb = load_workbook(filename=str(xlsx_path), data_only=True)
    summary: dict[str, Any] = {"project_fields_changed": [], "frames_changed": []}

    # ---------------- лист «Общий план ролика» ------------------------------
    if SHEET_GENERAL in wb.sheetnames:
        ws_g = wb[SHEET_GENERAL]
        for r in range(1, (ws_g.max_row or 0) + 1):
            label = _to_str(ws_g.cell(row=r, column=1).value)
            value = _to_str(ws_g.cell(row=r, column=2).value)
            if not label or value is None:
                continue
            field = _GENERAL_LABEL_TO_FIELD.get(label)
            if field is None:
                continue
            current = getattr(project, field, None)
            if value != (current or None):
                setattr(project, field, value)
                summary["project_fields_changed"].append(field)
                logger.info(
                    "[#{}] xlsx→DB: project.{} обновлено ({} симв)",
                    project.id,
                    field,
                    len(value),
                )

    # ---------------- лист «Кадры» -----------------------------------------
    if SHEET_FRAMES in wb.sheetnames:
        ws_f = wb[SHEET_FRAMES]
        # Сопоставляем колонки → номера кадров (по строке-заголовку).
        col_to_frame: dict[int, int] = {}
        max_col = ws_f.max_column or 0
        for col in range(2, max_col + 1):
            n = ws_f.cell(row=ROW_HEADER, column=col).value
            try:
                col_to_frame[col] = int(n)
            except (TypeError, ValueError):
                continue

        # все фреймы проекта → быстрый доступ по номеру
        rows = (
            await session.execute(
                select(Frame).where(Frame.project_id == project.id)
            )
        ).scalars().all()
        by_number = {f.number: f for f in rows}

        for col, fnum in col_to_frame.items():
            fr = by_number.get(fnum)
            if fr is None:
                continue
            new_voice = _to_str(ws_f.cell(row=ROW_VOICEOVER, column=col).value)
            new_meaning = _to_str(ws_f.cell(row=ROW_FRAME_LOGIC, column=col).value)
            new_imgp = _to_str(ws_f.cell(row=ROW_IMAGE_PROMPT, column=col).value)
            new_vidp = _to_str(ws_f.cell(row=ROW_VIDEO_PROMPT, column=col).value)
            new_dur = _to_float(ws_f.cell(row=ROW_VIDEO_DURATION, column=col).value)

            changed = False
            if new_voice and new_voice != fr.voiceover_text:
                fr.voiceover_text = new_voice
                changed = True
            if new_meaning and new_meaning != fr.meaning:
                fr.meaning = new_meaning
                changed = True
            if new_imgp and new_imgp != fr.image_prompt:
                fr.image_prompt = new_imgp
                changed = True
            if new_vidp and new_vidp != fr.animation_prompt:
                fr.animation_prompt = new_vidp
                changed = True
            if new_dur is not None and abs((fr.duration_seconds or 0.0) - new_dur) > 0.01:
                fr.duration_seconds = new_dur
                changed = True
            if changed:
                summary["frames_changed"].append(fnum)
                logger.info("[#{}] xlsx→DB: кадр {} обновлён", project.id, fnum)

    await session.flush()

    # ROOT FIX: после импорта данных из xlsx — перевычислить project.status,
    # чтобы он отражал реальное содержимое (а не остался в зависшем
    # `hero_ready` при пустых полях, например). Раньше клик «Перечитать
    # xlsx» обновлял поля, но статус не двигал — менюшка продолжала
    # врать ✅.
    try:
        from app.services.project_state import recompute_status
        old, new, changed = await recompute_status(
            session, project, log_prefix="recompute(after xlsx reload)"
        )
        if changed:
            summary["status_recomputed"] = f"{old.value} → {new.value}"
    except Exception as e:  # noqa: BLE001
        logger.warning("[#{}] recompute_status after xlsx reload failed: {}",
                       project.id, e)

    return summary
