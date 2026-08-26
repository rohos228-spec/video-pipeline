"""Единый fail-closed контракт записи в DB v2 (`apply-ops`).

Используется в двух местах:

1. REST endpoint `POST /api/db/projects/{id}/apply-ops` (тонкая обёртка).
2. GPT-ноды оркестратора: модель возвращает JSON ``{"ops": [...]}`` —
   ``extract_apply_ops_json`` достаёт его из ответа, ``apply_ops`` пишет.

Поля можно писать по-человечески («закадр», «промт_картинки») —
``FIELD_ALIASES`` переводит в канонические ключи. Ошибки — ``ApplyOpsError``
(неизвестный uuid/поле/target, пустые ops/fields): вызывающий код решает,
вернуть ли 400 (HTTP) или залогировать и откатить шаг (оркестратор).
"""

from __future__ import annotations

import json
import re
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Frame, Project
from app.services import db_v2
from app.services.xlsx_v8_import import (
    ROW_DURATION_V8,
    ROW_IMAGE_PROMPT_2_V8,
    ROW_IMAGE_PROMPT_V8,
    ROW_TIMECODE_V8,
    ROW_VIDEO_PROMPT_2_V8,
    ROW_VIDEO_PROMPT_V8,
    ROW_VOICEOVER_V8,
    SHEET_GENERAL_V8,
    _normalize_sheet_name,
    _resolve_plan_sheet,
)


def coerce_duration_seconds(value: Any) -> float | None:
    """Число или строка «3» / «3.5». «3 сек» → None (не стопать apply)."""
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


# Аналитика / shot → Excel (лист «план»), хранится в Frame.attrs.
ROW_MAIN_ACTION_V8 = 52
ROW_PLACE_V8 = 54
ROW_ACCENT_V8 = 55
ROW_SCENE_SENSE_V8 = 56
ROW_VISUAL_TYPE_V8 = 57
ROW_SCENE_FEATURE_V8 = 58
ROW_CLUSTER_V8 = 59
# Scene Grammar: структура/стык/переход + словесные границы сцены
ROW_SCENE_STRUCTURE_V8 = 60
ROW_EDIT_TYPE_V8 = 61
ROW_SCENE_TRANSITION_V8 = 62
ROW_SCENE_START_WORDS_V8 = 65
ROW_SCENE_END_WORDS_V8 = 66

# attr_key → номер строки Excel при экспорте
_ATTR_EXCEL_ROWS: dict[str, int] = {
    "main_action": ROW_MAIN_ACTION_V8,
    "place": ROW_PLACE_V8,
    "accent": ROW_ACCENT_V8,
    "scene_sense": ROW_SCENE_SENSE_V8,
    "visual_type": ROW_VISUAL_TYPE_V8,
    "scene_feature": ROW_SCENE_FEATURE_V8,
    "cluster": ROW_CLUSTER_V8,
    "scene_structure": ROW_SCENE_STRUCTURE_V8,
    "edit_type": ROW_EDIT_TYPE_V8,
    "scene_transition": ROW_SCENE_TRANSITION_V8,
    "scene_start_words": ROW_SCENE_START_WORDS_V8,
    "scene_end_words": ROW_SCENE_END_WORDS_V8,
    # shot_01
    "shot01_id_scene": 2,
    "shot01_id_shot": 3,
    "shot01_number": 4,
    "shot01_prev": 5,
    "shot01_next": 6,
    "shot01_bg": 7,
    "shot01_props": 9,
    "shot01_action": 10,
    "shot01_description": 11,
    "shot01_transition": 12,
    "shot01_notes": 13,
    "shot01_status": 14,
    # shot_02 (без 22–24)
    "shot02_id_scene": 16,
    "shot02_id_shot": 17,
    "shot02_number": 18,
    "shot02_prev": 19,
    "shot02_next": 20,
    "shot02_characters": 21,
    "shot02_action": 25,
    "shot02_description": 26,
    "shot02_transition": 27,
    "shot02_notes": 28,
    "shot02_status": 29,
}

# Человеческие имена полей → канонические ключи DB. Ключ нормализуется:
# lower + пробелы→подчёркивания.
FIELD_ALIASES: dict[str, str] = {
    "voiceover_text": "voiceover_text",
    "voiceover": "voiceover_text",
    "закадр": "voiceover_text",
    "закадровый_текст": "voiceover_text",
    "закадровый": "voiceover_text",
    "закадровый_text": "voiceover_text",
    "реплика": "voiceover_text",
    "image_prompt": "image_prompt",
    "img_prompt": "image_prompt",
    "промт_картинки": "image_prompt",
    "промпт_картинки": "image_prompt",
    "промт_картинка": "image_prompt",
    "картинка": "image_prompt",
    "animation_prompt": "animation_prompt",
    "video_prompt": "animation_prompt",
    "промт_видео": "animation_prompt",
    "промпт_видео": "animation_prompt",
    "видео_промт": "animation_prompt",
    "промт_анимации": "animation_prompt",
    "промпт_анимации": "animation_prompt",
    "анимация": "animation_prompt",
    "meaning": "meaning",
    "смысл": "meaning",
    "описание_кадра": "meaning",
    # Биты сценария (смысловые части «было → стало») → Frame.attrs["биты"].
    "биты": "биты",
    "смысловые_части": "биты",
    "beats": "биты",
    "duration_seconds": "duration_seconds",
    "длительность": "duration_seconds",
    "время": "duration_seconds",
    "секунды": "duration_seconds",
    "image_prompt_shot2": "image_prompt_shot2",
    "img_prompt_2": "image_prompt_shot2",
    "промт_картинки_2": "image_prompt_shot2",
    "промпт_картинки_2": "image_prompt_shot2",
    "animation_prompt_shot2": "animation_prompt_shot2",
    "video_prompt_2": "animation_prompt_shot2",
    "промт_видео_2": "animation_prompt_shot2",
    "промпт_видео_2": "animation_prompt_shot2",
    # ID персонажей кадра (c01, c03) → Frame.attrs["characters"] + Excel R8
    "characters": "characters",
    "персонажи": "characters",
    "persons": "characters",
    # Аналитика 54–59 + главное действие 52 → Frame.attrs + Excel
    "main_action": "main_action",
    "главное_действие": "main_action",
    "главное_действие_в_кадре": "main_action",
    "place": "place",
    "место": "place",
    "accent": "accent",
    "акцент": "accent",
    "scene_sense": "scene_sense",
    "смысл_сцены": "scene_sense",
    "visual_type": "visual_type",
    "тип_сцены": "visual_type",
    "scene_feature": "scene_feature",
    "особенность_сцены": "scene_feature",
    "особенность": "scene_feature",
    "cluster": "cluster",
    "номер_кластера": "cluster",
    "кластер": "cluster",
    # Scene Grammar (сцены по словам, не по столбцу)
    "scene_structure": "scene_structure",
    "структура_сцены": "scene_structure",
    "edit_type": "edit_type",
    "тип_стыка": "edit_type",
    "scene_transition": "scene_transition",
    "переход_в_кадр": "scene_transition",
    "переход_в_сцену": "scene_transition",
    "scene_start_words": "scene_start_words",
    "сцена_start_words": "scene_start_words",
    "start_words": "scene_start_words",
    "scene_end_words": "scene_end_words",
    "сцена_end_words": "scene_end_words",
    "end_words": "scene_end_words",
    # shot_01
    "shot01_id_scene": "shot01_id_scene",
    "id_scene": "shot01_id_scene",
    "shot01_id_shot": "shot01_id_shot",
    "id_shot": "shot01_id_shot",
    "shot01_number": "shot01_number",
    "номер_кадра": "shot01_number",
    "shot01_prev": "shot01_prev",
    "предыдущий_кадр": "shot01_prev",
    "shot01_next": "shot01_next",
    "следующий_кадр": "shot01_next",
    "shot01_bg": "shot01_bg",
    "фон": "shot01_bg",
    "lighting": "lighting",
    "освещение": "lighting",
    "scene_lighting": "scene_lighting",
    "освещение_сцены": "scene_lighting",
    "shot01_props": "shot01_props",
    "предметы": "shot01_props",
    "shot01_action": "shot01_action",
    "действие": "shot01_action",
    # Меню съёмки (группа script_frames_qc): крупность / движение / SET.
    "shot_size": "shot_size",
    "крупность": "shot_size",
    "size": "shot_size",
    "shot_move": "shot_move",
    "движение": "shot_move",
    "движение_камеры": "shot_move",
    "move": "shot_move",
    "shot_set": "shot_set",
    "набор": "shot_set",
    "set": "shot_set",
    "shot01_description": "shot01_description",
    "описание_shot01": "shot01_description",
    "shot01_transition": "shot01_transition",
    "логика_перехода": "shot01_transition",
    "shot01_notes": "shot01_notes",
    "заметки_по_консистенции": "shot01_notes",
    "shot01_status": "shot01_status",
    "статус": "shot01_status",
    # shot_02
    "shot02_id_scene": "shot02_id_scene",
    "shot02_id_shot": "shot02_id_shot",
    "shot02_number": "shot02_number",
    "shot02_prev": "shot02_prev",
    "shot02_next": "shot02_next",
    "shot02_characters": "shot02_characters",
    "персонажи_shot02": "shot02_characters",
    "shot02_action": "shot02_action",
    "действие_shot02": "shot02_action",
    "shot02_description": "shot02_description",
    "описание_shot02": "shot02_description",
    "shot02_transition": "shot02_transition",
    "логика_перехода_shot02": "shot02_transition",
    "shot02_notes": "shot02_notes",
    "shot02_status": "shot02_status",
}

# Поля, которые живут в Frame.attrs (не колонки Frame.*).
_ATTR_FIELD_KEYS = frozenset(_ATTR_EXCEL_ROWS) | {
    "characters",
    "image_prompt_shot2",
    "animation_prompt_shot2",
}

PROJECT_FIELD_ALIASES: dict[str, str] = {
    "general_plan": "general_plan",
    "общий_план": "general_plan",
    "план": "general_plan",
    "script_text": "script_text",
    "закадровый_текст": "script_text",
    "сценарий": "script_text",
    "voiceover": "script_text",
}

TARGETS = ("frame", "project", "replace_frames")


class ApplyOpsError(ValueError):
    """Ошибка валидации apply-ops (неизвестный uuid/поле/target, пустые ops)."""


def _canon_key(raw: str, aliases: dict[str, str]) -> str | None:
    key = str(raw).strip().lower().replace(" ", "_")
    return aliases.get(key)


def normalize_fields(raw: dict, aliases: dict[str, str], *, scope: str) -> dict:
    out: dict = {}
    unknown: list[str] = []
    for k, v in (raw or {}).items():
        canon = _canon_key(k, aliases)
        if canon is None:
            unknown.append(str(k))
            continue
        out[canon] = v
    if unknown:
        allowed = sorted(set(aliases))
        raise ApplyOpsError(
            f"{scope}: неизвестные поля {unknown}. "
            f"Разрешённые (можно по-человечески): {allowed}"
        )
    return out


_JSON_FENCE_RE = re.compile(r"```(?:json)?\s*(\{.*?\})\s*```", re.DOTALL)
_HEX_UUID_RE = re.compile(r"^[0-9a-fA-F]+$")


def _hamming_distance(a: str, b: str) -> int:
    if len(a) != len(b):
        return 10**9
    return sum(1 for x, y in zip(a, b, strict=True) if x != y)


def repair_near_miss_frame_uuids(
    ops: list[dict[str, Any]],
    known_uuids: list[str] | set[str],
    *,
    max_distance: int = 1,
) -> list[tuple[str, str]]:
    """Починить опечатки hex в ``frame_uuid`` (…4d… → …4b…).

    Fail-closed: чиним только если ровно один known uuid на расстоянии
    ``max_distance`` (по умолчанию 1 символ). Возвращает список
    ``(было, стало)``; мутирует ``ops`` in-place.
    """
    known = [str(u).strip() for u in known_uuids if str(u).strip()]
    known_set = set(known)
    remaps: list[tuple[str, str]] = []
    for op in ops:
        if not isinstance(op, dict):
            continue
        if op.get("target", "frame") not in (None, "frame"):
            continue
        raw = str(op.get("frame_uuid") or "").strip()
        if not raw or raw in known_set:
            continue
        if not _HEX_UUID_RE.fullmatch(raw):
            continue
        hits = [
            k
            for k in known
            if len(k) == len(raw)
            and _HEX_UUID_RE.fullmatch(k)
            and _hamming_distance(raw.lower(), k.lower()) <= max_distance
        ]
        # уникальный near-miss
        uniq = list(dict.fromkeys(hits))
        if len(uniq) != 1:
            continue
        fixed = uniq[0]
        op["frame_uuid"] = fixed
        remaps.append((raw, fixed))
    return remaps


_FRAME_NUMBER_ID_RE = re.compile(r"^\d{1,6}$")


def remap_frame_number_uuids(
    ops: list[dict[str, Any]],
    number_to_uuid: dict[int, str],
) -> list[tuple[str, str]]:
    """Если GPT прислал номер кадра вместо uuid (``"78"``) — подставить uuid.

    Fail-closed: только чистое целое 1…999999 и ровно один кадр с таким
    ``Frame.number``. Мутирует ``ops`` in-place. Возвращает ``(было, стало)``.
    """
    remaps: list[tuple[str, str]] = []
    if not number_to_uuid:
        return remaps
    for op in ops:
        if not isinstance(op, dict):
            continue
        if op.get("target", "frame") not in (None, "frame"):
            continue
        raw = str(op.get("frame_uuid") or "").strip()
        if not raw or not _FRAME_NUMBER_ID_RE.fullmatch(raw):
            continue
        try:
            num = int(raw)
        except ValueError:
            continue
        fixed = number_to_uuid.get(num)
        if not fixed:
            continue
        op["frame_uuid"] = fixed
        remaps.append((raw, fixed))
    return remaps


def salvage_ops_from_partial_json(text: str) -> list[dict[str, Any]]:
    """Достать целые объекты из обрезанного ``{"ops":[…`` (без закрытия)."""
    if not text:
        return []
    m = re.search(r'"ops"\s*:\s*\[', text)
    if not m:
        m = re.search(r'"actions"\s*:\s*\[', text)
    if not m:
        return []
    i = m.end()
    ops: list[dict[str, Any]] = []
    n = len(text)
    while i < n:
        while i < n and text[i] in " \t\r\n,":
            i += 1
        if i >= n or text[i] == "]":
            break
        if text[i] != "{":
            break
        depth = 0
        start = i
        in_str = False
        esc = False
        ended = False
        for j in range(i, n):
            ch = text[j]
            if in_str:
                if esc:
                    esc = False
                elif ch == "\\":
                    esc = True
                elif ch == '"':
                    in_str = False
                continue
            if ch == '"':
                in_str = True
            elif ch == "{":
                depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0:
                    chunk = text[start : j + 1]
                    try:
                        obj = json.loads(chunk)
                    except Exception:  # noqa: BLE001
                        obj = None
                    if isinstance(obj, dict) and (
                        obj.get("frame_uuid") or obj.get("target")
                    ):
                        ops.append(obj)
                    i = j + 1
                    ended = True
                    break
        if not ended:
            break
    return ops


def extract_apply_ops_json(text: str) -> dict[str, Any] | None:
    """Достать JSON с ``ops`` / ``characters`` / ``scenes`` из ответа модели.

    Если целый объект обрезан Cloudflare/моделью — salvage целых ops из
    хвоста ``"ops":[…``.
    """
    if not text:
        return None
    candidates: list[str] = [m.group(1) for m in _JSON_FENCE_RE.finditer(text)]
    idxs = [
        i
        for i in (
            text.find('"ops"'),
            text.find('"actions"'),
            text.find('"characters"'),
            text.find('"scenes"'),
        )
        if i != -1
    ]
    idx = min(idxs) if idxs else -1
    if idx != -1:
        start = text.rfind("{", 0, idx)
        if start != -1:
            depth = 0
            in_str = False
            esc = False
            for i in range(start, len(text)):
                ch = text[i]
                if in_str:
                    if esc:
                        esc = False
                    elif ch == "\\":
                        esc = True
                    elif ch == '"':
                        in_str = False
                    continue
                if ch == '"':
                    in_str = True
                elif ch == "{":
                    depth += 1
                elif ch == "}":
                    depth -= 1
                    if depth == 0:
                        candidates.append(text[start : i + 1])
                        break
    for cand in candidates:
        try:
            data = json.loads(cand)
        except Exception:  # noqa: BLE001
            continue
        if isinstance(data, dict) and (
            isinstance(data.get("ops"), list)
            or isinstance(data.get("actions"), list)
            or isinstance(data.get("characters"), list)
            or isinstance(data.get("scenes"), list)
        ):
            return data
    # Обрезанный JSON: ops частично целые — лучше, чем JSON-retry с нуля.
    partial_ops = salvage_ops_from_partial_json(text)
    if partial_ops:
        return {"ops": partial_ops, "_salvaged_partial": True}
    return None


def _normalize_scene_card(raw: dict[str, Any]) -> dict[str, Any]:
    """Сцена из ответа агента → реестр (границы по словам, не по столбцу)."""
    if not isinstance(raw, dict):
        raise ApplyOpsError("scenes: каждый элемент — объект")
    sid = str(
        raw.get("id_scene") or raw.get("id") or raw.get("scene_id") or ""
    ).strip()
    if not sid:
        raise ApplyOpsError("scenes: нужен id_scene")
    start = str(
        raw.get("start_words")
        or raw.get("сцена_start_words")
        or raw.get("scene_start_words")
        or ""
    ).strip()
    end = str(
        raw.get("end_words")
        or raw.get("сцена_end_words")
        or raw.get("scene_end_words")
        or ""
    ).strip()
    if not start or not end:
        raise ApplyOpsError(
            f"scenes[{sid}]: нужны start_words и end_words (цитаты из полного закадра)"
        )
    shots = raw.get("shots")
    if shots is not None and not isinstance(shots, list):
        raise ApplyOpsError(f"scenes[{sid}]: shots должен быть списком")
    time_sec: float | None = None
    time_raw = raw.get("время_сек", raw.get("длительность_сек", raw.get("duration_seconds")))
    if time_raw is not None and str(time_raw).strip() != "":
        try:
            time_sec = round(float(time_raw), 1)
        except (TypeError, ValueError):
            time_sec = None
    return {
        "id_scene": sid,
        "start_words": start,
        "end_words": end,
        "время_сек": time_sec,
        "структура_сцены": str(
            raw.get("структура_сцены") or raw.get("scene_structure") or ""
        ).strip(),
        "тип_стыка": str(raw.get("тип_стыка") or raw.get("edit_type") or "").strip(),
        "переход_в_сцену": str(
            raw.get("переход_в_сцену")
            or raw.get("переход_в_кадр")
            or raw.get("scene_transition")
            or ""
        ).strip(),
        "место": str(raw.get("место") or raw.get("place") or "").strip(),
        "освещение": str(
            raw.get("освещение") or raw.get("lighting") or raw.get("scene_lighting") or ""
        ).strip(),
        "акцент": str(raw.get("акцент") or raw.get("accent") or "").strip(),
        "смысл_сцены": str(
            raw.get("смысл_сцены") or raw.get("scene_sense") or ""
        ).strip(),
        "тип_сцены": str(raw.get("тип_сцены") or raw.get("visual_type") or "").strip(),
        "особенность_доминанта": str(
            raw.get("особенность_доминанта")
            or raw.get("особенность_сцены")
            or raw.get("scene_feature")
            or ""
        ).strip(),
        "номер_кластера": str(
            raw.get("номер_кластера") or raw.get("cluster") or ""
        ).strip(),
        "персонажи_сцены": str(
            raw.get("персонажи_сцены") or raw.get("characters") or ""
        ).strip(),
        "shots": list(shots or []),
    }


def upsert_scene_registry(project: Project, scenes: list[Any]) -> int:
    """Сохранить реестр сцен в ``project.meta["scene_registry"]``."""
    if not isinstance(scenes, list):
        raise ApplyOpsError("scenes: нужен список")
    if not scenes:
        return 0
    normalized = [_normalize_scene_card(s) for s in scenes]
    meta = dict(project.meta or {})
    meta["scene_registry"] = normalized
    project.meta = meta
    return len(normalized)


def _join_frame_voiceovers(
    frames: list[Frame],
) -> tuple[str, list[tuple[Frame, int, int]]]:
    """Склейка закадра + span каждого кадра в общей строке."""
    chunks: list[str] = []
    spans: list[tuple[Frame, int, int]] = []
    pos = 0
    for fr in frames:
        text = (getattr(fr, "voiceover_text", None) or "").strip()
        if chunks and text:
            chunks.append(" ")
            pos += 1
        start = pos
        if text:
            chunks.append(text)
            pos += len(text)
        spans.append((fr, start, pos))
    return "".join(chunks), spans


def _scene_span_in_text(full: str, start_words: str, end_words: str) -> tuple[int, int] | None:
    """Найти [start, end) сцены по цитатам; None если не найдено однозначно."""
    s = (start_words or "").strip()
    e = (end_words or "").strip()
    if not s or not e or not full:
        return None
    i0 = full.find(s)
    if i0 < 0:
        return None
    i1 = full.find(e, i0)
    if i1 < 0:
        return None
    return i0, i1 + len(e)


def expand_scene_registry_onto_frames(
    frames: list[Frame],
    registry: list[Any] | None,
) -> int:
    """Сцена → кадры: привязка по словам + scene-level + 1 shot → 1 кадр.

    SoT = ``scenes[]`` (монтаж), не число колонок закадра.
    1) кадры в диапазоне start_words…end_words получают id_scene;
    2) scene-level поля (место/структура/…) на все кадры сцены;
    3) ``shots[i]`` → i-й кадр сцены (действие/ракурс уникально), лишние
       колонки внутри сцены получают только scene-level (не клон шота).
    """
    if not registry or not frames:
        return 0
    scenes: list[dict[str, Any]] = []
    for raw in registry:
        if isinstance(raw, dict) and str(raw.get("id_scene") or "").strip():
            scenes.append(raw)
    if not scenes:
        return 0

    full, spans = _join_frame_voiceovers(frames)
    # sid → frames in order
    scene_frames: dict[str, list[Frame]] = {str(sc["id_scene"]).strip(): [] for sc in scenes}
    for fr, a, b in spans:
        if a >= b and not (getattr(fr, "voiceover_text", None) or "").strip():
            continue
        for sc in scenes:
            sid = str(sc.get("id_scene") or "").strip()
            span = _scene_span_in_text(
                full,
                str(sc.get("start_words") or sc.get("scene_start_words") or ""),
                str(sc.get("end_words") or sc.get("scene_end_words") or ""),
            )
            if span is None:
                continue
            s0, s1 = span
            # пересечение span кадра со сценой
            if b > s0 and a < s1:
                scene_frames[sid].append(fr)
                break

    def _set(attrs: dict[str, Any], key: str, value: Any, *, overwrite: bool = False) -> bool:
        val = str(value or "").strip()
        if not val:
            return False
        if not overwrite and str(attrs.get(key) or "").strip():
            return False
        attrs[key] = val
        return True

    n = 0
    by_id = {str(sc.get("id_scene") or "").strip(): sc for sc in scenes}
    # Также кадры, у которых id_scene уже был из ops, но слова не сматчились.
    for fr in frames:
        attrs = dict(fr.attrs or {})
        sid = str(attrs.get("shot01_id_scene") or "").strip()
        if sid and sid in by_id and fr not in scene_frames.get(sid, []):
            scene_frames.setdefault(sid, []).append(fr)

    for sid, frs in scene_frames.items():
        sc = by_id.get(sid)
        if not sc or not frs:
            continue
        # de-dupe preserving order
        seen: set[int] = set()
        ordered: list[Frame] = []
        for fr in frs:
            fid = id(fr)
            if fid in seen:
                continue
            seen.add(fid)
            ordered.append(fr)
        shots = sc.get("shots") if isinstance(sc.get("shots"), list) else []
        for idx, fr in enumerate(ordered):
            attrs = dict(fr.attrs or {})
            changed = False
            if _set(attrs, "shot01_id_scene", sid, overwrite=True):
                changed = True
            for key, src in (
                ("place", sc.get("место") or sc.get("place")),
                ("scene_sense", sc.get("смысл_сцены") or sc.get("scene_sense")),
                ("visual_type", sc.get("тип_сцены") or sc.get("visual_type")),
                ("cluster", sc.get("номер_кластера") or sc.get("cluster")),
                (
                    "scene_lighting",
                    sc.get("освещение")
                    or sc.get("lighting")
                    or sc.get("scene_lighting"),
                ),
                (
                    "scene_structure",
                    sc.get("структура_сцены") or sc.get("scene_structure"),
                ),
                ("edit_type", sc.get("тип_стыка") or sc.get("edit_type")),
                (
                    "scene_transition",
                    sc.get("переход_в_сцену")
                    or sc.get("переход_в_кадр")
                    or sc.get("scene_transition"),
                ),
                (
                    "scene_start_words",
                    sc.get("start_words") or sc.get("scene_start_words"),
                ),
                ("scene_end_words", sc.get("end_words") or sc.get("scene_end_words")),
                ("scene_time_sec", sc.get("время_сек")),
            ):
                if _set(attrs, key, src, overwrite=True):
                    changed = True
            # Один shot → один кадр (по порядку внутри сцены).
            # accent — с ШОТА, не клон сцены на все колонки.
            if idx < len(shots) and isinstance(shots[idx], dict):
                sh = shots[idx]
                shot_id = str(sh.get("id_shot") or f"shot_{idx + 1:02d}").strip()
                if _set(attrs, "shot01_id_shot", shot_id, overwrite=True):
                    changed = True
                shot_accent = (
                    sh.get("акцент")
                    or sh.get("accent")
                    or sh.get("предметы")
                    or sh.get("действие")
                    or ""
                )
                shot_lighting = (
                    sh.get("освещение")
                    or sh.get("lighting")
                    or sc.get("освещение")
                    or sc.get("lighting")
                    or ""
                )
                for key, src in (
                    ("scene_feature", sh.get("особенность_сцены") or sh.get("ракурс")),
                    ("shot01_action", sh.get("действие") or sh.get("action")),
                    (
                        "shot01_description",
                        sh.get("описание_кадра") or sh.get("description"),
                    ),
                    ("shot01_transition", sh.get("логика_перехода")),
                    ("shot01_bg", sh.get("фон")),
                    ("lighting", shot_lighting),
                    ("shot01_notes", shot_lighting),
                    ("shot01_props", sh.get("предметы")),
                    ("main_action", sh.get("главное_действие")),
                    ("accent", shot_accent),
                ):
                    if _set(attrs, key, src, overwrite=True):
                        changed = True
                if sh.get("персонажи") is not None:
                    if _set(attrs, "characters", sh.get("персонажи"), overwrite=True):
                        changed = True
                elif sc.get("персонажи_сцены") is not None:
                    if _set(
                        attrs, "characters", sc.get("персонажи_сцены"), overwrite=True
                    ):
                        changed = True
            else:
                # Кадр без своего shot — не копируем scene.accent (иначе все
                # колонки сцены с одним акцентом).
                if "accent" in attrs and len(shots) > 0:
                    attrs.pop("accent", None)
                    changed = True
            if changed:
                fr.attrs = attrs
                n += 1
    return n


# Shot-level attrs, которые нельзя массово копировать со сцены на кадры.
# Не включать shot01_bg / lighting / shot01_notes: одинаковый фон/свет
# на соседних кадрах одного места — это правильная continuity, не клон-баг.
_SHOT_DETAIL_ATTR_KEYS = (
    "shot01_action",
    "shot01_description",
    "shot01_props",
    "shot01_transition",
    "main_action",
    "scene_feature",
    "accent",
)


def strip_duplicated_shot_details(frames: list[Frame]) -> int:
    """Снять с кадров shot-детали, если один и тот же текст на ≥3 кадрах.

    Чинит последствия старого expand (один shot → 10+ одинаковых описаний).
    Scene-level (place/structure/id_scene) не трогает.
    """
    from collections import defaultdict

    buckets: dict[tuple[str, str], list[Frame]] = defaultdict(list)
    for fr in frames:
        attrs = fr.attrs or {}
        for key in _SHOT_DETAIL_ATTR_KEYS:
            val = str(attrs.get(key) or "").strip()
            if not val:
                continue
            buckets[(key, val)].append(fr)
    n = 0
    for (key, _val), group in buckets.items():
        if len(group) < 3:
            continue
        for fr in group:
            attrs = dict(fr.attrs or {})
            if key in attrs:
                attrs.pop(key, None)
                fr.attrs = attrs
                n += 1
    return n


def _normalize_character_card(raw: dict[str, Any]) -> dict[str, Any]:
    """Карточка персонажа из ответа агента → code/name/attrs."""
    if not isinstance(raw, dict):
        raise ApplyOpsError("characters: каждый элемент — объект")
    code = str(
        raw.get("id") or raw.get("code") or raw.get("ID") or ""
    ).strip()
    if not code:
        raise ApplyOpsError("characters: нужен id (c01…)")
    name = str(raw.get("имя") or raw.get("name") or "").strip()
    look = str(raw.get("внешность") or raw.get("look") or "").strip()
    clothes = str(raw.get("одежда") or raw.get("clothes") or "").strip()
    char = str(raw.get("характер") or raw.get("char") or "").strip()
    rules = str(raw.get("правила") or raw.get("rules") or "").strip()
    return {
        "code": code,
        "name": name,
        "attrs": {
            "look": look,
            "clothes": clothes,
            "char": char,
            "rules": rules,
            "внешность": look,
            "одежда": clothes,
            "характер": char,
            "правила": rules,
        },
    }


async def upsert_characters(
    session: AsyncSession,
    project: Project,
    characters: list[Any],
) -> int:
    """Записать/обновить сущности type=character по code (c01…)."""
    from app.models import Entity

    if not isinstance(characters, list):
        raise ApplyOpsError("characters: нужен список")
    if not characters:
        return 0
    existing = list(
        (
            await session.execute(
                select(Entity).where(
                    Entity.project_id == project.id,
                    Entity.type == "character",
                )
            )
        ).scalars()
    )
    by_code = {
        str(e.code or "").strip(): e for e in existing if e.code
    }
    max_key = max((float(e.sort_key or 0) for e in existing), default=0.0)
    n = 0
    for raw in characters:
        card = _normalize_character_card(raw)
        code = card["code"]
        en = by_code.get(code)
        if en is None:
            max_key += 10.0
            en = Entity(
                project_id=project.id,
                type="character",
                code=code,
                name=card["name"],
                attrs=card["attrs"],
                sort_key=max_key,
            )
            session.add(en)
            by_code[code] = en
        else:
            en.name = card["name"]
            en.attrs = card["attrs"]
            en.type = "character"
        n += 1
    await session.flush()
    return n


def _write_persons_sheet(wb: Any, entities: list[Any]) -> int:
    """Синк листа «Персонажи» из Entity (колонки B.. как в excel_characters)."""
    from app.services.excel_characters import (
        ROW_CHAR,
        ROW_CLOTHES,
        ROW_ID,
        ROW_LOOK,
        ROW_NAME,
        ROW_RULES,
        SHEET_PERSONS,
    )

    chars = [
        e
        for e in entities
        if getattr(e, "type", None) == "character" and getattr(e, "code", None)
    ]
    chars.sort(key=lambda e: (float(getattr(e, "sort_key", 0) or 0), str(e.code)))
    if SHEET_PERSONS in wb.sheetnames:
        ws = wb[SHEET_PERSONS]
    else:
        ws = wb.create_sheet(SHEET_PERSONS)
        ws["A1"] = "ID персонажа"
        ws["A3"] = "имя"
        ws["A4"] = "внешность"
        ws["A5"] = "одежда"
        ws["A6"] = "характер"
        ws["A7"] = "правила"
    # Чистим старые колонки B.. (до разумного лимита)
    # openpyxl: cell(..., value=None) не пишет — только .value = None
    for col in range(2, max(ws.max_column or 2, 2) + 1):
        for row in (ROW_ID, ROW_NAME, ROW_LOOK, ROW_CLOTHES, ROW_CHAR, ROW_RULES):
            ws.cell(row=row, column=col).value = None
    cells = 0
    for i, en in enumerate(chars):
        col = 2 + i
        attrs = en.attrs if isinstance(en.attrs, dict) else {}
        look = str(attrs.get("look") or attrs.get("внешность") or "")
        clothes = str(attrs.get("clothes") or attrs.get("одежда") or "")
        char = str(attrs.get("char") or attrs.get("характер") or "")
        rules = str(attrs.get("rules") or attrs.get("правила") or "")
        ws.cell(row=ROW_ID, column=col, value=str(en.code))
        ws.cell(row=ROW_NAME, column=col, value=str(en.name or ""))
        ws.cell(row=ROW_LOOK, column=col, value=look)
        ws.cell(row=ROW_CLOTHES, column=col, value=clothes)
        ws.cell(row=ROW_CHAR, column=col, value=char)
        ws.cell(row=ROW_RULES, column=col, value=rules)
        cells += 6
    return cells


def _fmt_timecode(start: float | None, end: float | None) -> str | None:
    if start is None or end is None:
        return None

    def _ts(v: float) -> str:
        m = int(v // 60)
        s = v - m * 60
        return f"{m}:{s:05.2f}"

    return f"{_ts(float(start))}-{_ts(float(end))}"


def export_project_xlsx(
    project: Project,
    frames: list[Frame],
    *,
    entities: list[Any] | None = None,
) -> dict:
    """DB → project.xlsx (план + Общий план + Персонажи + R8).

    Пишет строки DB: R45/R46/R48/R64/R49, R50, R15, плюс R8 (персонажи кадра)
    и лист «Персонажи» из Entity. Перед записью — backup.
    """
    from app.services.plan_shot2 import (
        SHOT2_PROMPT_ATTR,
        SHOT2_STATUS_ATTR,
        SHOT2_VIDEO_PROMPT_ATTR,
    )
    from app.services.xlsx_versioning import backup_to_old

    ROW_PERSONS_PRIMARY = 8  # как в generate_images / baza

    path = project.data_dir / "project.xlsx"
    if not path.is_file():
        raise ApplyOpsError("project.xlsx не найден — сначала сгенерируй лист «план»")
    from openpyxl import load_workbook

    wb = load_workbook(path)
    try:
        ws = _resolve_plan_sheet(wb)
        if ws is None:
            raise ApplyOpsError("в project.xlsx нет листа «план» (v8)")
        snap = None
        try:
            snap = backup_to_old(path)
        except Exception:  # noqa: BLE001
            snap = None
        cells = 0
        for fr in frames:
            col = fr.number + 2
            attrs = fr.attrs or {}
            writes = {
                ROW_IMAGE_PROMPT_V8: fr.image_prompt or "",
                ROW_IMAGE_PROMPT_2_V8: str(attrs.get(SHOT2_PROMPT_ATTR) or ""),
                ROW_VIDEO_PROMPT_V8: fr.animation_prompt or "",
                ROW_VIDEO_PROMPT_2_V8: str(attrs.get(SHOT2_VIDEO_PROMPT_ATTR) or ""),
                ROW_VOICEOVER_V8: fr.voiceover_text or "",
            }
            for row, value in writes.items():
                ws.cell(row=row, column=col, value=value)
                cells += 1
            if fr.duration_seconds is not None:
                ws.cell(row=ROW_DURATION_V8, column=col, value=float(fr.duration_seconds))
                cells += 1
            tc = _fmt_timecode(fr.start_ts, fr.end_ts)
            if tc:
                ws.cell(row=ROW_TIMECODE_V8, column=col, value=tc)
                cells += 1
            persons = str(
                attrs.get("characters")
                or attrs.get("persons")
                or attrs.get("персонажи")
                or ""
            ).strip()
            ws.cell(row=ROW_PERSONS_PRIMARY, column=col, value=persons)
            cells += 1
            # Аналитика 52/54–59 + shot_01/shot_02 из attrs
            for attr_key, row in _ATTR_EXCEL_ROWS.items():
                if attr_key not in attrs:
                    continue
                val = attrs.get(attr_key)
                if val is None or str(val).strip() == "":
                    continue
                ws.cell(row=row, column=col, value=str(val))
                cells += 1
        # Стереть хвост колонок после последнего кадра — иначе sync_project_xlsx
        # снова поднимет удалённые VO/таймкоды (было 182 после truncate до 20).
        max_col = int(ws.max_column or 2)
        clear_from = len(frames) + 3  # number N → col N+2; первая лишняя = N+3
        if clear_from <= max_col:
            rows_clear = {
                ROW_IMAGE_PROMPT_V8,
                ROW_IMAGE_PROMPT_2_V8,
                ROW_VIDEO_PROMPT_V8,
                ROW_VIDEO_PROMPT_2_V8,
                ROW_VOICEOVER_V8,
                ROW_DURATION_V8,
                ROW_TIMECODE_V8,
                ROW_PERSONS_PRIMARY,
                *(_ATTR_EXCEL_ROWS.values()),
            }
            for col in range(clear_from, max_col + 1):
                for row in rows_clear:
                    # openpyxl: cell(..., value=None) НЕ пишет — None значит «не менять»
                    ws.cell(row=row, column=col).value = None
                    cells += 1

        general_plan = (project.meta or {}).get("general_plan")
        if general_plan:
            for name in wb.sheetnames:
                if _normalize_sheet_name(name).casefold() == _normalize_sheet_name(
                    SHEET_GENERAL_V8
                ).casefold():
                    wb[name]["B2"] = str(general_plan)
                    cells += 1
                    break
        if entities is not None:
            cells += _write_persons_sheet(wb, entities)
        wb.save(path)
    finally:
        wb.close()
    return {
        "frames": len(frames),
        "cells": cells,
        "backup": getattr(snap, "name", None),
        "path": str(path),
    }


async def apply_ops(
    session: AsyncSession,
    project: Project,
    ops: list[dict[str, Any]] | None = None,
    *,
    characters: list[Any] | None = None,
    scenes: list[Any] | None = None,
    export_xlsx: bool = False,
    node_kind: str | None = None,
) -> dict:
    """Применить JSON-операции к проекту (fail-closed, одна транзакция).

    ``ops`` — список ``{"target": "frame"|"project"|"replace_frames", ...}``.
    ``characters`` — реестр персонажей ``[{id, имя, внешность, …}]`` → Entity.
    ``scenes`` — реестр сцен с ``start_words``/``end_words`` → meta.scene_registry.
    ``replace_frames``: ``{"target":"replace_frames","frames":[{"закадр":"…"},…]}``.
    ``node_kind`` — опциональный фильтр полей (img_pr / anim_pr / excel_gpt*);
    ``None`` = не трогать ops (поведение остальных вызывающих без изменений).
    Excel не трогаем по умолчанию — только явный Export
    (``export_project_xlsx`` / кнопка). ``export_xlsx=True`` — legacy opt-in.
    """
    from app.models import Entity
    from app.services.plan_shot2 import (
        SHOT2_PROMPT_ATTR,
        SHOT2_STATUS_ATTR,
        SHOT2_VIDEO_PROMPT_ATTR,
    )

    ops = list(ops or [])
    if node_kind:
        from app.services.node_write_contract import filter_ops_for_node

        ops = filter_ops_for_node(ops, node_kind=str(node_kind))
    chars_n = 0
    scenes_n = 0
    if characters:
        chars_n = await upsert_characters(session, project, characters)
    if scenes:
        scenes_n = upsert_scene_registry(project, scenes)
    if not ops and not chars_n and not scenes_n:
        raise ApplyOpsError("пустой ops — нечего применять")
    for op in ops:
        if op.get("target", "frame") not in TARGETS:
            raise ApplyOpsError(
                f"неизвестный target {op.get('target')!r}; разрешены: {list(TARGETS)}"
            )

    replace_ops = [op for op in ops if op.get("target") == "replace_frames"]
    if len(replace_ops) > 1:
        raise ApplyOpsError("replace_frames: только одна операция за раз")
    if replace_ops and len(ops) > 1:
        raise ApplyOpsError(
            "replace_frames нельзя смешивать с другими ops в одном запросе"
        )

    frame_ops = [op for op in ops if op.get("target", "frame") == "frame"]
    project_ops = [op for op in ops if op.get("target", "frame") == "project"]

    if replace_ops:
        specs = replace_ops[0].get("frames") or replace_ops[0].get("кадры") or []
        if not isinstance(specs, list):
            raise ApplyOpsError("replace_frames: нужен список frames")
        try:
            created = await db_v2.replace_all_frames(session, project, specs)
        except ValueError as e:
            raise ApplyOpsError(str(e)) from None
        exported = None
        if export_xlsx:
            exported = export_project_xlsx(project, created)
        return {
            "ok": True,
            "updated": len(created),
            "exported": exported,
            "replace_frames": len(created),
        }

    by_uuid: dict[str, Frame] = {}
    uuid_repairs: list[tuple[str, str]] = []
    number_repairs: list[tuple[str, str]] = []
    if frame_ops:
        uuids = [str(op.get("frame_uuid") or "") for op in frame_ops]
        if any(not u for u in uuids):
            raise ApplyOpsError("для target=frame нужен frame_uuid")
        # Все кадры проекта — иначе near-miss не к чему привязать.
        frames = list(
            (
                await session.execute(
                    select(Frame).where(Frame.project_id == project.id)
                )
            ).scalars()
        )
        by_uuid = {f.uuid: f for f in frames}
        # GPT часто пишет номер кадра ("78") вместо uuid — сначала remap по number.
        by_number: dict[int, list[str]] = {}
        for f in frames:
            n = getattr(f, "number", None)
            if n is None:
                continue
            try:
                ni = int(n)
            except (TypeError, ValueError):
                continue
            by_number.setdefault(ni, []).append(str(f.uuid))
        number_to_uuid = {
            n: ids[0] for n, ids in by_number.items() if len(ids) == 1
        }
        number_repairs = remap_frame_number_uuids(frame_ops, number_to_uuid)
        if number_repairs:
            from loguru import logger

            logger.warning(
                "db_apply: frame_number→uuid remapped n={} samples={}",
                len(number_repairs),
                number_repairs[:5],
            )
        uuid_repairs = repair_near_miss_frame_uuids(frame_ops, list(by_uuid.keys()))
        if uuid_repairs:
            from loguru import logger

            logger.warning(
                "db_apply: near-miss frame_uuid repaired n={} samples={}",
                len(uuid_repairs),
                uuid_repairs[:5],
            )
        uuids = [str(op.get("frame_uuid") or "") for op in frame_ops]
        missing = [u for u in uuids if u not in by_uuid]
        if missing:
            raise ApplyOpsError(f"неизвестные frame_uuid: {missing}")

    updated = 0
    for op in frame_ops:
        uuid = str(op.get("frame_uuid") or "")
        fields = normalize_fields(op.get("fields") or {}, FIELD_ALIASES, scope=f"кадр {uuid}")
        if not fields:
            raise ApplyOpsError(f"кадр {uuid}: пустые fields")
        fr = by_uuid[uuid]
        if "voiceover_text" in fields:
            fr.voiceover_text = str(fields["voiceover_text"] or "")
        if "meaning" in fields:
            fr.meaning = None if fields["meaning"] is None else str(fields["meaning"])
        if "duration_seconds" in fields:
            raw_dur = fields["duration_seconds"]
            if raw_dur is None or str(raw_dur).strip() == "":
                fr.duration_seconds = None
            else:
                parsed_dur = coerce_duration_seconds(raw_dur)
                if parsed_dur is None:
                    from loguru import logger

                    logger.warning(
                        "db_apply: кадр {}: skip non-numeric duration {!r}",
                        uuid,
                        raw_dur,
                    )
                else:
                    fr.duration_seconds = parsed_dur
        if "image_prompt" in fields:
            fr.image_prompt = (
                None if fields["image_prompt"] is None else str(fields["image_prompt"])
            )
            await db_v2.add_prompt_version(
                session,
                project.id,
                fr.id,
                kind="img",
                text=fr.image_prompt or "",
                set_active=True,
            )
            # Чтобы шаг images подхватил кадр после excel_gpt/img_pr apply-ops.
            from app.models import FrameStatus
            from app.generation_options import is_skippable_empty_prompt

            if not is_skippable_empty_prompt(fr.image_prompt or "") and fr.status not in (
                FrameStatus.image_generated,
                FrameStatus.image_approved,
            ):
                fr.status = FrameStatus.image_prompt_ready
        if "animation_prompt" in fields:
            fr.animation_prompt = (
                None
                if fields["animation_prompt"] is None
                else str(fields["animation_prompt"])
            )
            await db_v2.add_prompt_version(
                session,
                project.id,
                fr.id,
                kind="video",
                text=fr.animation_prompt or "",
                set_active=True,
            )
        attrs = dict(fr.attrs or {})
        if "image_prompt_shot2" in fields:
            attrs[SHOT2_PROMPT_ATTR] = str(fields["image_prompt_shot2"] or "")
            from app.generation_options import is_skippable_empty_prompt

            if not is_skippable_empty_prompt(attrs[SHOT2_PROMPT_ATTR]):
                attrs[SHOT2_STATUS_ATTR] = "image_prompt_ready"

        if "animation_prompt_shot2" in fields:
            attrs[SHOT2_VIDEO_PROMPT_ATTR] = str(fields["animation_prompt_shot2"] or "")
        if "characters" in fields:
            persons = str(fields["characters"] or "").strip()
            attrs["characters"] = persons
            attrs["persons"] = persons
            attrs["персонажи"] = persons
        if "биты" in fields:
            # Список битов храним как есть (JSON), не str().
            bits_raw = fields["биты"]
            attrs["биты"] = bits_raw if isinstance(bits_raw, list) else str(
                bits_raw or ""
            )
        for attr_key in _ATTR_EXCEL_ROWS:
            if attr_key in fields:
                attrs[attr_key] = str(fields[attr_key] or "")
        cs_raw = attrs.get("camera_subdivide")
        cs = dict(cs_raw) if isinstance(cs_raw, dict) else {}
        _cs_map = (
            ("shot_size", "крупность"),
            ("shot_move", "движение"),
            ("shot_set", "набор"),
        )
        wrote_cs = False
        for canon, cs_key in _cs_map:
            if canon in fields and str(fields[canon] or "").strip():
                cs[cs_key] = str(fields[canon]).strip()
                wrote_cs = True
        if wrote_cs:
            attrs["camera_subdivide"] = cs
            if cs.get("набор"):
                attrs["place"] = cs["набор"]
        fr.attrs = attrs
        updated += 1

    for op in project_ops:
        fields = normalize_fields(
            op.get("fields") or {}, PROJECT_FIELD_ALIASES, scope="проект"
        )
        if not fields:
            raise ApplyOpsError("проект: пустые fields")
        meta = dict(project.meta or {})
        if "general_plan" in fields:
            plan_val = str(fields["general_plan"] or "")
            meta["general_plan"] = plan_val
            # Колонка HITL/статусов — тот же SoT, что meta (не только Excel B2).
            project.general_plan = plan_val
        if "script_text" in fields:
            project.script_text = str(fields["script_text"] or "")
            # Файл voiceover.txt — артефакт для audio/music (зеркало DB).
            try:
                vo_path = project.data_dir / "voiceover.txt"
                vo_path.parent.mkdir(parents=True, exist_ok=True)
                vo_path.write_text(project.script_text or "", encoding="utf-8")
            except OSError:
                pass
        project.meta = meta
        updated += 1

    await session.flush()

    # Детали из scenes[].shots[] → Frame.attrs только если ЭТОТ apply
    # реально обновил scenes. Иначе старый scene_registry затирает
    # свежие frame ops (место/смысл после повторного ▶ excel_gpt).
    all_frames = list(
        (
            await session.execute(
                select(Frame)
                .where(Frame.project_id == project.id)
                .order_by(Frame.number)
            )
        ).scalars()
    )
    meta_now = project.meta if isinstance(project.meta, dict) else {}
    expanded = 0
    if scenes_n > 0:
        expanded = expand_scene_registry_onto_frames(
            all_frames, meta_now.get("scene_registry")
        )
        if expanded:
            await session.flush()

    exported = None
    if export_xlsx:
        ents = list(
            (
                await session.execute(
                    select(Entity).where(Entity.project_id == project.id)
                )
            ).scalars()
        )
        exported = export_project_xlsx(project, all_frames, entities=ents)
    return {
        "ok": True,
        "updated": updated,
        "characters": chars_n,
        "scenes": scenes_n,
        "expanded_frames": expanded,
        "exported": exported,
        "uuid_repairs": [
            {"from": a, "to": b} for a, b in (number_repairs + uuid_repairs)
        ],
    }
