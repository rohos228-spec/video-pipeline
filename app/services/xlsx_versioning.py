"""Версионирование project.xlsx: перед каждой подменой файла GPT-ответом
старая версия сохраняется в `data/projects/<id>/old/<timestamp>.xlsx`.

Использование:
    from app.services.xlsx_versioning import backup_to_old, replace_with
    backup_to_old(current_path)              # сохранил предыдущую версию
    replace_with(current_path, new_xlsx)     # подменил current на новую

Или одной операцией:
    replace_with_backup(current_path, new_xlsx)
"""

from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from loguru import logger

_XLSX_FORMAT_ERROR_PREFIX = "ошибка формата эксель таблицы"


def _ts() -> str:
    return datetime.utcnow().strftime("%Y%m%d_%H%M%S")


def old_dir_for(project_xlsx: Path) -> Path:
    """Папка для архивных версий рядом с project.xlsx (data/projects/<id>/old/)."""
    return project_xlsx.parent / "old"


def backup_to_old(project_xlsx: Path) -> Path | None:
    """Копирует текущий project.xlsx в old/<timestamp>_<orig>.xlsx.

    Возвращает путь к копии или None если исходного файла не было.
    """
    project_xlsx = Path(project_xlsx)
    if not project_xlsx.exists():
        return None
    old_dir = old_dir_for(project_xlsx)
    old_dir.mkdir(parents=True, exist_ok=True)
    dest = old_dir / f"{_ts()}_{project_xlsx.name}"
    shutil.copy2(project_xlsx, dest)
    logger.info("xlsx_versioning: backup {} -> {}", project_xlsx, dest)
    return dest


def replace_with(project_xlsx: Path, new_file: Path) -> None:
    """Заменяет project.xlsx содержимым `new_file`. БЕЗ бэкапа предыдущей версии."""
    project_xlsx = Path(project_xlsx)
    new_file = Path(new_file)
    if not new_file.exists():
        raise FileNotFoundError(f"replace_with: исходный файл не найден {new_file}")
    project_xlsx.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(new_file, project_xlsx)
    logger.info("xlsx_versioning: replace {} <- {}", project_xlsx, new_file)


def replace_with_backup(project_xlsx: Path, new_file: Path) -> Path | None:
    """Бэкапит текущую версию project.xlsx и заменяет её на `new_file`.

    Возвращает путь к бэкапу (или None если предыдущей версии не было).
    """
    backup = backup_to_old(project_xlsx)
    replace_with(project_xlsx, new_file)
    return backup


def _workbook_sheet_names(path: Path) -> list[str]:
    from openpyxl import load_workbook  # noqa: PLC0415

    wb = load_workbook(path, read_only=True)
    names = list(wb.sheetnames)
    wb.close()
    return names


def _allowed_sheet_layouts() -> list[frozenset[str]]:
    """Допустимые наборы листов: шаблон из settings + известные v7/v8."""
    layouts: list[frozenset[str]] = []
    seen: set[frozenset[str]] = set()

    def _add(names: frozenset[str]) -> None:
        if names and names not in seen:
            seen.add(names)
            layouts.append(names)

    try:
        from app.services.xlsx_v8_import import SHEET_GENERAL_V8, SHEET_PLAN_V8
        from app.storage.project_sheet import (
            SHEET_FRAMES,
            SHEET_GENERAL,
            resolve_default_template_path,
        )

        tpl = resolve_default_template_path()
        if tpl.is_file():
            _add(frozenset(_workbook_sheet_names(tpl)))
        _add(frozenset({SHEET_FRAMES, SHEET_GENERAL}))
        _add(frozenset({SHEET_PLAN_V8, SHEET_GENERAL_V8}))
    except Exception as e:  # noqa: BLE001
        logger.warning("xlsx_versioning: cannot build sheet layouts: {}", e)
    return layouts


def validate_xlsx_sheets(path: Path) -> str | None:
    """Имена и количество листов должны совпадать с одним из шаблонов."""
    try:
        actual = frozenset(_workbook_sheet_names(path))
    except Exception as e:  # noqa: BLE001
        return f"{_XLSX_FORMAT_ERROR_PREFIX}: не удалось прочитать листы: {e}"
    if not actual:
        return f"{_XLSX_FORMAT_ERROR_PREFIX}: в файле нет листов"
    for expected in _allowed_sheet_layouts():
        if actual == expected:
            return None
    actual_str = ", ".join(sorted(actual))
    expected_hint = " | ".join(
        ", ".join(sorted(layout)) for layout in _allowed_sheet_layouts()
    )
    return (
        f"{_XLSX_FORMAT_ERROR_PREFIX}: листы [{actual_str}] "
        f"не совпадают с шаблоном (ожидалось: {expected_hint})"
    )


def _resolve_workbook_sheet(wb: Any, name: str) -> Any | None:
    if name in wb.sheetnames:
        return wb[name]
    low = name.casefold()
    for sheet_name in wb.sheetnames:
        if sheet_name.casefold() == low:
            return wb[sheet_name]
    return None


def normalize_xlsx_to_reference_layout(source: Path, reference: Path) -> bool:
    """Убрать лишние листы GPT: оставить layout reference, значения — из source.

    Возвращает True, если файл перезаписан (были лишние листы или другой порядок).
    """
    source = Path(source)
    reference = Path(reference)
    if not source.is_file() or not reference.is_file():
        return False

    try:
        ref_names = _workbook_sheet_names(reference)
        src_names = _workbook_sheet_names(source)
    except Exception as e:  # noqa: BLE001
        logger.debug("xlsx_versioning: normalize skip (read sheets): {}", e)
        return False

    if not ref_names:
        return False

    expected = frozenset(ref_names)
    actual = frozenset(src_names)
    if not expected.issubset(actual):
        return False
    if src_names == ref_names:
        return False

    from openpyxl import load_workbook  # noqa: PLC0415

    tmp = source.with_name(f"{source.stem}.norm{source.suffix}")
    try:
        shutil.copy2(reference, tmp)
        src_wb = load_workbook(source, data_only=False)
        dst_wb = load_workbook(tmp)
        try:
            for ref_name in ref_names:
                src_ws = _resolve_workbook_sheet(src_wb, ref_name)
                dst_ws = _resolve_workbook_sheet(dst_wb, ref_name)
                if src_ws is None or dst_ws is None:
                    continue
                for row in src_ws.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            dst_ws.cell(
                                row=cell.row,
                                column=cell.column,
                                value=cell.value,
                            )
            dst_wb.save(tmp)
        finally:
            src_wb.close()
            dst_wb.close()

        shutil.move(str(tmp), str(source))
    except Exception as e:  # noqa: BLE001
        logger.warning("xlsx_versioning: normalize failed {}: {}", source, e)
        if tmp.exists():
            try:
                tmp.unlink()
            except OSError:
                pass
        return False

    dropped = sorted(actual - expected)
    logger.info(
        "xlsx_versioning: normalized {} → {} лист(ов) reference, убрано: {}",
        source.name,
        len(ref_names),
        dropped or "reorder",
    )
    return True


def validate_xlsx(path: Path) -> str | None:
    """Проверяет что `path` — валидный xlsx-файл.

    Возвращает None если ок, иначе человекочитаемое сообщение об ошибке.
    Используется после `download_attachment_from_last_reply`, чтобы не
    подменить project.xlsx «пустышкой» (svg-иконкой, html-ошибкой и т.п.).
    Также сверяет набор листов с шаблоном (имена + количество).
    """
    path = Path(path)
    if not path.exists():
        return f"файл не существует: {path}"
    size = path.stat().st_size
    # Минимальный валидный xlsx ≈ 5 КБ. Меньше 1 КБ — почти наверняка
    # это не xlsx, а svg/html/иконка которую ChatGPT отдал по ошибочному
    # клику.
    if size < 1024:
        return f"файл подозрительно мал ({size} байт)"
    with path.open("rb") as f:
        magic = f.read(4)
    # xlsx — это zip, у zip-архивов всегда первые 2 байта 'PK'.
    if magic[:2] != b"PK":
        preview = magic.hex()
        return (
            f"файл не является xlsx (первые 4 байта: {preview}, "
            f"ожидался zip-magic 'PK')"
        )
    # Финальная проверка — действительно ли openpyxl откроет файл.
    try:
        from openpyxl import load_workbook  # noqa: PLC0415

        wb = load_workbook(path, read_only=True)
        wb.close()
    except Exception as e:  # noqa: BLE001
        return f"openpyxl не смог открыть файл: {e}"
    sheet_err = validate_xlsx_sheets(path)
    if sheet_err is not None:
        return sheet_err
    return None


def is_valid_xlsx(path: Path) -> bool:
    return validate_xlsx(path) is None


def _quarantine_corrupt(project_xlsx: Path) -> Path | None:
    """Убрать битый project.xlsx в old/ (не мешает restore)."""
    project_xlsx = Path(project_xlsx)
    if not project_xlsx.exists():
        return None
    old_dir = old_dir_for(project_xlsx)
    old_dir.mkdir(parents=True, exist_ok=True)
    dest = old_dir / f"{_ts()}_CORRUPT_{project_xlsx.name}"
    try:
        shutil.move(str(project_xlsx), str(dest))
        logger.warning("xlsx_versioning: corrupt quarantined {} -> {}", project_xlsx, dest)
        return dest
    except OSError as e:
        logger.warning("xlsx_versioning: cannot quarantine {}: {}", project_xlsx, e)
        return None


def restore_latest_valid_backup(project_xlsx: Path) -> Path | None:
    """Восстановить project.xlsx из последнего валидного файла в old/."""
    project_xlsx = Path(project_xlsx)
    old_dir = old_dir_for(project_xlsx)
    if not old_dir.is_dir():
        return None
    candidates = sorted(
        old_dir.glob("*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    for src in candidates:
        if "CORRUPT" in src.name:
            continue
        if not is_valid_xlsx(src):
            continue
        project_xlsx.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(src, project_xlsx)
        logger.info("xlsx_versioning: restored {} <- {}", project_xlsx, src)
        return src
    return None


def repair_project_xlsx_if_corrupt(
    project_xlsx: Path,
    *,
    template_path: Path | None = None,
) -> bool:
    """Если project.xlsx битый — restore из old/ или шаблона. True если починили."""
    project_xlsx = Path(project_xlsx)
    if project_xlsx.exists() and is_valid_xlsx(project_xlsx):
        return False
    if project_xlsx.exists():
        _quarantine_corrupt(project_xlsx)
    restored = restore_latest_valid_backup(project_xlsx)
    if restored is not None:
        return True
    if template_path is not None and Path(template_path).exists():
        project_xlsx.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(template_path, project_xlsx)
        logger.warning(
            "xlsx_versioning: no backup; copied template -> {}",
            project_xlsx,
        )
        return True
    return False
