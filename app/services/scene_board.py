"""Агрегат данных для Scene Board — вертикальный обзор сцен проекта.

Одна сцена = один Frame (столбец листа «план»). Для каждой сцены:
закадровый текст, shot_01/shot_02 (картинка+видео), персонажи, озвучка,
таймслот на монтажной шкале, музыка проекта.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from loguru import logger
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models import Artifact, ArtifactKind, Frame, Project
from app.services.bgm import find_bgm_file, resolve_bgm
from app.services.frame_audio import frame_audio_path
from app.services.plan_shot2 import (
    SHOT2_PROMPT_ATTR,
    SHOT2_STATUS_ATTR,
    SHOT2_VIDEO_PROMPT_ATTR,
    SHOT2_VIDEO_STATUS_ATTR,
    find_shot1_image,
    find_shot2_image,
    read_shot2_columns,
    shot2_video_file_pattern,
)
from app.services.xlsx_v8_import import (
    ROW_VOICEOVER_V8,
    _cell_text,
    _resolve_plan_sheet,
)

_XLSX_ROWS_PERSONS = (8, 23, 38)
_XLSX_ROWS_ITEMS = (9, 24, 39)
_REF_ID_RE = re.compile(r"^(c\d+|i\d+|predmet\d+)$", re.IGNORECASE)

REGEN_TARGETS = (
    "voiceover_text",
    "image_shot1",
    "image_shot2",
    "video_shot1",
    "video_shot2",
    "characters",
    "audio",
    "music",
)

REGEN_TYPES = ("media", "prompt_and_media", "full_scene")


def _file_preview_url(path: Path | str | None) -> str | None:
    if path is None:
        return None
    p = Path(path)
    if not p.is_file():
        return None
    return f"/api/files?path={p}"


def _artifact_preview(art: Artifact | None) -> dict[str, Any] | None:
    if art is None:
        return None
    preview = None
    if art.uuid:
        preview = f"/api/artifacts/{art.uuid}/file"
    elif art.path:
        preview = _file_preview_url(art.path)
    return {
        "artifact_uuid": art.uuid,
        "path": art.path,
        "preview_url": preview,
        "present": bool(preview or (art.path and Path(art.path).is_file())),
    }


def _disk_media(path: Path | None) -> dict[str, Any]:
    present = path is not None and path.is_file()
    return {
        "artifact_uuid": None,
        "path": str(path) if present else None,
        "preview_url": _file_preview_url(path) if present else None,
        "present": present,
    }


def _normalize_ref_id(token: str) -> str | None:
    t = (token or "").strip().lower().rstrip(":;,.)]}»\"'")
    if not t or not _REF_ID_RE.match(t):
        return None
    return t


def _parse_ref_ids(cell_value: object) -> list[str]:
    if cell_value is None:
        return []
    s = str(cell_value).strip()
    if not s:
        return []
    for ch in (";", "+", "/", "|", " "):
        s = s.replace(ch, ",")
    out: list[str] = []
    for tok in s.split(","):
        norm = _normalize_ref_id(tok)
        if norm:
            out.append(norm)
    return out


def _find_ref_preview(base_dir: Path, ref_id: str) -> Path | None:
    if not base_dir.is_dir():
        return None
    aliases = [ref_id]
    if ref_id.startswith("i") and ref_id[1:].isdigit():
        aliases.append(f"predmet{int(ref_id[1:])}")
    elif ref_id.startswith("predmet") and ref_id[7:].isdigit():
        aliases.append(f"i{int(ref_id[7:]):02d}")
    candidates: list[Path] = []
    for alias in aliases:
        for ext in ("png", "jpg", "jpeg", "webp"):
            candidates.extend(base_dir.glob(f"{alias}_*.{ext}"))
            direct = base_dir / f"{alias}.{ext}"
            if direct.is_file():
                candidates.append(direct)
    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


def _read_plan_refs(xlsx_path: Path) -> dict[int, dict[str, list[str]]]:
    """frame_number → {character_ids, item_ids} из листа «план»."""
    out: dict[int, dict[str, list[str]]] = {}
    if not xlsx_path.is_file():
        return out
    try:
        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    except Exception as e:  # noqa: BLE001
        logger.warning("scene_board: cannot open xlsx {}: {}", xlsx_path, e)
        return out
    try:
        ws = _resolve_plan_sheet(wb)
        if ws is None:
            return out
        max_col = ws.max_column or 0
        frame_no = 0
        for col in range(3, max_col + 1):
            voice = _cell_text(ws, ROW_VOICEOVER_V8, col)
            if voice is None:
                continue
            frame_no += 1

            def _merged(rows: tuple[int, ...]) -> list[str]:
                merged: list[str] = []
                seen: set[str] = set()
                for r in rows:
                    for x in _parse_ref_ids(ws.cell(row=r, column=col).value):
                        if x not in seen:
                            seen.add(x)
                            merged.append(x)
                return merged

            out[frame_no] = {
                "character_ids": _merged(_XLSX_ROWS_PERSONS),
                "item_ids": _merged(_XLSX_ROWS_ITEMS),
            }
    finally:
        wb.close()
    return out


def _newest_shot1_video(videos_dir: Path, frame_number: int) -> Path | None:
    if not videos_dir.is_dir():
        return None
    candidates = [
        p
        for p in videos_dir.glob(f"clip_{frame_number:03d}_*.mp4")
        if "_s2_" not in p.name
    ]
    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


def _newest_shot2_video(videos_dir: Path, frame_number: int) -> Path | None:
    if not videos_dir.is_dir():
        return None
    candidates = list(videos_dir.glob(shot2_video_file_pattern(frame_number)))
    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


def _latest_by_frame(
    arts: list[Artifact], kind: ArtifactKind
) -> dict[int, Artifact]:
    best: dict[int, Artifact] = {}
    for a in arts:
        if a.kind != kind or a.frame_id is None:
            continue
        prev = best.get(a.frame_id)
        if prev is None or a.id > prev.id:
            best[a.frame_id] = a
    return best


def _latest_project(arts: list[Artifact], kind: ArtifactKind) -> Artifact | None:
    matched = [a for a in arts if a.kind == kind and a.frame_id is None]
    if not matched:
        matched = [a for a in arts if a.kind == kind]
    if not matched:
        return None
    return max(matched, key=lambda a: a.id)


def _format_ts(seconds: float | None) -> str | None:
    if seconds is None:
        return None
    s = max(0.0, float(seconds))
    m = int(s // 60)
    rem = s - m * 60
    return f"{m}:{rem:05.2f}"


def _missing_flags(scene: dict[str, Any]) -> list[str]:
    missing: list[str] = []
    if not (scene.get("voiceover_text") or "").strip():
        missing.append("voiceover_text")
    img1 = scene.get("image_shot1") or {}
    if not img1.get("present"):
        missing.append("image_shot1")
    if scene.get("has_shot2"):
        img2 = scene.get("image_shot2") or {}
        if not img2.get("present"):
            missing.append("image_shot2")
        vid2 = scene.get("video_shot2") or {}
        if not vid2.get("present"):
            missing.append("video_shot2")
    vid1 = scene.get("video_shot1") or {}
    if not vid1.get("present"):
        missing.append("video_shot1")
    audio = scene.get("audio") or {}
    if not audio.get("present"):
        missing.append("audio")
    if scene.get("start_ts") is None or scene.get("end_ts") is None:
        missing.append("timeslot")
    if not scene.get("characters"):
        missing.append("characters")
    return missing


async def build_scene_board(
    session: AsyncSession, project: Project
) -> dict[str, Any]:
    """Собрать payload для GET /api/projects/{id}/scene-board."""
    frames = (
        await session.execute(
            select(Frame)
            .where(Frame.project_id == project.id)
            .options(selectinload(Frame.artifacts))
            .order_by(Frame.number.asc())
        )
    ).scalars().all()

    arts = (
        await session.execute(
            select(Artifact).where(Artifact.project_id == project.id)
        )
    ).scalars().all()

    images_by_frame = _latest_by_frame(arts, ArtifactKind.scene_image)
    videos_by_frame = _latest_by_frame(arts, ArtifactKind.scene_video)
    audio_art = _latest_project(arts, ArtifactKind.audio)
    music_art = _latest_project(arts, ArtifactKind.music)
    final_art = _latest_project(arts, ArtifactKind.final_video)

    clip_by_number: dict[int, dict[str, Any]] = {}
    if audio_art and isinstance(audio_art.meta, dict):
        for c in audio_art.meta.get("clips") or []:
            if not isinstance(c, dict):
                continue
            try:
                n = int(c.get("frame_number"))
            except (TypeError, ValueError):
                continue
            clip_by_number[n] = c

    xlsx_path = project.data_dir / "project.xlsx"
    plan_refs = _read_plan_refs(xlsx_path)
    shot2_info = read_shot2_columns(xlsx_path)

    scenes_dir = project.data_dir / "scenes"
    videos_dir = project.data_dir / "videos"
    audio_dir = project.data_dir / "audio"
    chars_dir = project.data_dir / "characters"
    items_dir = project.data_dir / "items"

    excel_chars = []
    meta = project.meta or {}
    eh = meta.get("excel_hero") if isinstance(meta, dict) else None
    if isinstance(eh, dict) and isinstance(eh.get("characters"), list):
        excel_chars = eh["characters"]
    char_by_id = {
        str(c.get("id", "")).lower(): c
        for c in excel_chars
        if isinstance(c, dict) and c.get("id")
    }

    scenes: list[dict[str, Any]] = []
    for fr in frames:
        attrs = dict(fr.attrs or {})
        s2 = shot2_info.get(fr.number)
        has_shot2 = bool(
            (s2 and s2.has_shot2)
            or attrs.get(SHOT2_PROMPT_ATTR)
            or find_shot2_image(scenes_dir, fr.number)
        )

        # Image shot1: prefer artifact if not s2 path, else disk
        img_art = images_by_frame.get(fr.id)
        img1_disk = find_shot1_image(scenes_dir, fr.number)
        if img_art and img_art.path and "_s2_" not in Path(img_art.path).name:
            image_shot1 = _artifact_preview(img_art)
        else:
            image_shot1 = _disk_media(img1_disk)
            if not image_shot1["present"] and img_art:
                image_shot1 = _artifact_preview(img_art)

        img2_disk = find_shot2_image(scenes_dir, fr.number)
        image_shot2 = _disk_media(img2_disk) if has_shot2 else None

        vid_art = videos_by_frame.get(fr.id)
        vid1_disk = _newest_shot1_video(videos_dir, fr.number)
        if vid_art and vid_art.path and "_s2_" not in Path(vid_art.path).name:
            video_shot1 = _artifact_preview(vid_art)
        else:
            video_shot1 = _disk_media(vid1_disk)
            if not video_shot1["present"] and vid_art:
                video_shot1 = _artifact_preview(vid_art)

        vid2_disk = _newest_shot2_video(videos_dir, fr.number)
        video_shot2 = _disk_media(vid2_disk) if has_shot2 else None

        audio_path = frame_audio_path(audio_dir, fr.number)
        audio_media = _disk_media(audio_path if audio_path.is_file() else None)

        clip = clip_by_number.get(fr.number) or {}
        start_ts = fr.start_ts if fr.start_ts is not None else clip.get("start_ts")
        end_ts = fr.end_ts if fr.end_ts is not None else clip.get("end_ts")
        duration = (
            fr.duration_seconds
            if fr.duration_seconds is not None
            else clip.get("duration")
        )
        if duration is None and start_ts is not None and end_ts is not None:
            try:
                duration = float(end_ts) - float(start_ts)
            except (TypeError, ValueError):
                duration = None

        refs = plan_refs.get(fr.number) or {"character_ids": [], "item_ids": []}
        characters: list[dict[str, Any]] = []
        for cid in refs["character_ids"]:
            info = char_by_id.get(cid) or {}
            preview = _find_ref_preview(chars_dir, cid)
            characters.append(
                {
                    "id": cid,
                    "name": info.get("name") or cid,
                    "kind": "character",
                    "preview_url": _file_preview_url(preview),
                    "present": preview is not None,
                }
            )
        items: list[dict[str, Any]] = []
        for iid in refs["item_ids"]:
            preview = _find_ref_preview(items_dir, iid)
            items.append(
                {
                    "id": iid,
                    "name": iid,
                    "kind": "item",
                    "preview_url": _file_preview_url(preview),
                    "present": preview is not None,
                }
            )

        scene: dict[str, Any] = {
            "frame_id": fr.id,
            "number": fr.number,
            "status": fr.status.value if hasattr(fr.status, "value") else str(fr.status),
            "voiceover_text": fr.voiceover_text or "",
            "meaning": fr.meaning,
            "image_prompt": fr.image_prompt,
            "animation_prompt": fr.animation_prompt,
            "image_prompt_shot2": attrs.get(SHOT2_PROMPT_ATTR)
            or (s2.prompt if s2 else None),
            "animation_prompt_shot2": attrs.get(SHOT2_VIDEO_PROMPT_ATTR),
            "shot2_status": attrs.get(SHOT2_STATUS_ATTR),
            "shot2_video_status": attrs.get(SHOT2_VIDEO_STATUS_ATTR),
            "has_shot2": has_shot2,
            "start_ts": float(start_ts) if start_ts is not None else None,
            "end_ts": float(end_ts) if end_ts is not None else None,
            "duration_seconds": float(duration) if duration is not None else None,
            "timeslot_label": (
                f"{_format_ts(float(start_ts) if start_ts is not None else None)}"
                f" – {_format_ts(float(end_ts) if end_ts is not None else None)}"
                if start_ts is not None and end_ts is not None
                else None
            ),
            "image_shot1": image_shot1,
            "image_shot2": image_shot2,
            "video_shot1": video_shot1,
            "video_shot2": video_shot2,
            "audio": audio_media,
            "characters": characters,
            "items": items,
            "attrs": attrs,
        }
        scene["missing"] = _missing_flags(scene)
        scenes.append(scene)

    bgm_cfg = resolve_bgm(project)
    disk_bgm = find_bgm_file(project)
    music_payload: dict[str, Any]
    if bgm_cfg is not None:
        music_payload = {
            "enabled": True,
            "present": True,
            "path": str(bgm_cfg.path),
            "preview_url": _file_preview_url(bgm_cfg.path),
            "level_percent": int(bgm_cfg.level * 100),
            "label": bgm_cfg.path.name,
        }
    elif disk_bgm is not None and disk_bgm.is_file():
        # Файл есть, но BGM выключен флагом meta
        music_payload = {
            "enabled": False,
            "present": True,
            "path": str(disk_bgm),
            "preview_url": _file_preview_url(disk_bgm),
            "level_percent": 0,
            "label": f"{disk_bgm.name} (выкл.)",
        }
    else:
        music_payload = {
            "enabled": False,
            "present": False,
            "path": None,
            "preview_url": None,
            "level_percent": 0,
            "label": "нет файла",
        }

    if music_art and not music_payload["present"]:
        art_prev = _artifact_preview(music_art) or {}
        music_payload = {
            **music_payload,
            **art_prev,
            "enabled": True,
            "label": Path(music_art.path).name if music_art.path else "music",
        }

    regen = {}
    if isinstance(meta, dict) and isinstance(meta.get("scene_board_regen"), dict):
        regen = meta["scene_board_regen"]

    return {
        "project_id": project.id,
        "slug": project.slug,
        "topic": project.topic,
        "status": (
            project.status.value
            if hasattr(project.status, "value")
            else str(project.status)
        ),
        "frame_count": len(scenes),
        "music": music_payload,
        "master_audio": _artifact_preview(audio_art),
        "final_video": _artifact_preview(final_art),
        "scenes": scenes,
        "regen_draft": {
            "note": regen.get("note") or "",
            "selections": regen.get("selections") or [],
        },
        "regen_targets": list(REGEN_TARGETS),
        "regen_types": list(REGEN_TYPES),
    }


def validate_regen_draft(payload: dict[str, Any]) -> dict[str, Any]:
    note = str(payload.get("note") or "").strip()
    raw_sel = payload.get("selections") or []
    if not isinstance(raw_sel, list):
        raise ValueError("selections must be a list")
    selections: list[dict[str, Any]] = []
    for item in raw_sel:
        if not isinstance(item, dict):
            continue
        try:
            frame_id = int(item["frame_id"])
            number = int(item.get("number") or 0)
        except (KeyError, TypeError, ValueError) as e:
            raise ValueError(f"bad selection frame_id: {e}") from e
        targets = [
            t
            for t in (item.get("targets") or [])
            if isinstance(t, str) and t in REGEN_TARGETS
        ]
        regen_type = item.get("regen_type") or "media"
        if regen_type not in REGEN_TYPES:
            regen_type = "media"
        if not targets:
            continue
        selections.append(
            {
                "frame_id": frame_id,
                "number": number,
                "targets": targets,
                "regen_type": regen_type,
            }
        )
    return {"note": note, "selections": selections}
