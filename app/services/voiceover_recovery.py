"""Поиск и восстановление закадрового текста — глубокий скан диска и БД.

Источники (приоритет — от более «исходных» к запасным):
  1. old/*_voiceover*.txt, .trash/*voiceover*
  2. Дочерние проекты (old / voiceover / script_text / tmp_gpt / xlsx)
  3. tmp_gpt/voiceover_*.txt, script_*.txt
  4. old/*.xlsx и project.xlsx (лист «Общий план» / v8 R49)
  5. Кадры в БД (склейка frame.voiceover_text)
  6. script_text / voiceover.txt на диске
"""

from __future__ import annotations

import re
import shutil
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from loguru import logger
from sqlalchemy import Integer, cast, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Frame, Project
from app.services.mass_factory import list_mass_children, mass_parent_id

_BACKUP_RE = re.compile(r"^(\d{8}_\d{6})_.*voiceover.*\.txt$", re.I)
_XLSX_BACKUP_RE = re.compile(r"^(\d{8}_\d{6})_.*\.xlsx$", re.I)
_TMP_VO_RE = re.compile(r"^voiceover_(\d{8}_\d{6})\.txt$", re.I)
_MIN_VO_LEN = 80


@dataclass(frozen=True)
class VoiceoverCandidate:
    text: str
    source: str
    priority: int


def _read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace").strip()


def _ts_sort_key(name: str) -> tuple[str, str]:
    m = re.match(r"^(\d{8}_\d{6})", name)
    if m:
        return m.group(1), name
    return name, name


def backup_sort_key(path: Path) -> tuple[str, str]:
    m = _BACKUP_RE.match(path.name)
    if m:
        return m.group(1), path.name
    return _ts_sort_key(path.name)


def list_voiceover_backups(project: Project) -> list[Path]:
    found: list[Path] = []
    for folder in (project.data_dir / "old", project.data_dir / ".trash"):
        if not folder.is_dir():
            continue
        for pattern in ("*_voiceover.txt", "*_voiceover_deleted.txt", "*voiceover*.txt"):
            for p in folder.glob(pattern):
                if p.is_file() and p.stat().st_size > 0:
                    found.append(p)
    # уникальные пути, старые первыми
    uniq = sorted({p.resolve() for p in found}, key=backup_sort_key)
    return uniq


def oldest_voiceover_backup(project: Project) -> Path | None:
    backups = list_voiceover_backups(project)
    return backups[0] if backups else None


def trash_voiceover_file(project: Project, voice_path: Path) -> Path | None:
    """Переместить voiceover в .trash/ + бэкап в old/ (вместо безвозвратного unlink)."""
    if not voice_path.is_file():
        return None
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    trash_dir = project.data_dir / ".trash"
    trash_dir.mkdir(parents=True, exist_ok=True)
    dest = trash_dir / f"{ts}_voiceover.txt"
    shutil.copy2(voice_path, dest)
    old_dir = project.data_dir / "old"
    old_dir.mkdir(parents=True, exist_ok=True)
    backup = old_dir / f"{ts}_voiceover_deleted.txt"
    shutil.copy2(voice_path, backup)
    voice_path.unlink(missing_ok=True)
    logger.info(
        "[#{}] voiceover trashed → {} and {}",
        project.id,
        dest.name,
        backup.name,
    )
    return dest


def _candidate_from_file(path: Path, source: str, priority: int) -> VoiceoverCandidate | None:
    if not path.is_file() or path.stat().st_size == 0:
        return None
    text = _read_text(path)
    if len(text) < _MIN_VO_LEN:
        return None
    return VoiceoverCandidate(text=text, source=source, priority=priority)


def _candidate_from_text(text: str, source: str, priority: int) -> VoiceoverCandidate | None:
    cleaned = (text or "").strip()
    if len(cleaned) < _MIN_VO_LEN:
        return None
    return VoiceoverCandidate(text=cleaned, source=source, priority=priority)


def _script_text_from_xlsx(path: Path) -> str | None:
    if not path.is_file() or path.stat().st_size < 512:
        return None
    try:
        from openpyxl import load_workbook

        from app.services.xlsx_sync import _GENERAL_LABEL_TO_FIELD, _to_str
        from app.storage.project_sheet import ROW_HEADER, ROW_VOICEOVER, SHEET_FRAMES, SHEET_GENERAL
        from app.services.xlsx_v8_import import _read_voiceover_blocks, has_v8_plan_sheet

        wb = load_workbook(filename=str(path), data_only=True)
        try:
            if SHEET_GENERAL in wb.sheetnames:
                ws_g = wb[SHEET_GENERAL]
                for r in range(1, (ws_g.max_row or 0) + 1):
                    label = _to_str(ws_g.cell(row=r, column=1).value)
                    value = _to_str(ws_g.cell(row=r, column=2).value)
                    if label in _GENERAL_LABEL_TO_FIELD and _GENERAL_LABEL_TO_FIELD[label] == "script_text":
                        if value and len(value) >= _MIN_VO_LEN:
                            return value
            if has_v8_plan_sheet(wb):
                blocks = _read_voiceover_blocks(wb)
                joined = " ".join(b for b in blocks if b.strip())
                if len(joined) >= _MIN_VO_LEN:
                    return joined
            if SHEET_FRAMES in wb.sheetnames:
                ws_f = wb[SHEET_FRAMES]
                parts: list[str] = []
                for col in range(2, (ws_f.max_column or 0) + 1):
                    n = ws_f.cell(row=ROW_HEADER, column=col).value
                    try:
                        int(n)
                    except (TypeError, ValueError):
                        continue
                    voice = _to_str(ws_f.cell(row=ROW_VOICEOVER, column=col).value)
                    if voice:
                        parts.append(voice)
                joined = " ".join(parts)
                if len(joined) >= _MIN_VO_LEN:
                    return joined
        finally:
            wb.close()
    except Exception as exc:  # noqa: BLE001
        logger.debug("xlsx voiceover read {}: {}", path.name, exc)
    return None


def _scan_dir_candidates(data_dir: Path, prefix: str, base_priority: int) -> list[VoiceoverCandidate]:
    out: list[VoiceoverCandidate] = []
    if not data_dir.is_dir():
        return out

    for backup in list_voiceover_backups_from_dir(data_dir):
        c = _candidate_from_file(backup, f"{prefix}old/{backup.name}", base_priority)
        if c:
            out.append(c)

    tmp = data_dir / "tmp_gpt"
    if tmp.is_dir():
        vo_files = sorted(
            [p for p in tmp.glob("voiceover_*.txt") if p.is_file() and p.stat().st_size > 0],
            key=lambda p: _ts_sort_key(p.name),
        )
        for i, p in enumerate(vo_files):
            c = _candidate_from_file(
                p,
                f"{prefix}tmp_gpt/{p.name}",
                base_priority + 30 + i,
            )
            if c:
                out.append(c)
        for p in sorted(tmp.glob("script_*.txt")):
            c = _candidate_from_file(p, f"{prefix}tmp_gpt/{p.name}", base_priority + 50)
            if c:
                out.append(c)

    old_dir = data_dir / "old"
    if old_dir.is_dir():
        xlsx_files = sorted(
            [p for p in old_dir.glob("*.xlsx") if p.is_file()],
            key=lambda p: _ts_sort_key(p.name),
        )
        for i, xlsx in enumerate(xlsx_files):
            text = _script_text_from_xlsx(xlsx)
            c = _candidate_from_text(
                text or "",
                f"{prefix}old/{xlsx.name}",
                base_priority + 20 + i,
            )
            if c:
                out.append(c)

    for name in ("voiceover.txt", "script.txt"):
        c = _candidate_from_file(data_dir / name, f"{prefix}{name}", base_priority + 90)
        if c:
            out.append(c)

    xlsx = data_dir / "project.xlsx"
    text = _script_text_from_xlsx(xlsx)
    c = _candidate_from_text(text or "", f"{prefix}project.xlsx", base_priority + 95)
    if c:
        out.append(c)

    return out


def list_voiceover_backups_from_dir(data_dir: Path) -> list[Path]:
    found: list[Path] = []
    for folder in (data_dir / "old", data_dir / ".trash"):
        if not folder.is_dir():
            continue
        for pattern in ("*_voiceover.txt", "*_voiceover_deleted.txt", "*voiceover*.txt"):
            for p in folder.glob(pattern):
                if p.is_file() and p.stat().st_size > 0:
                    found.append(p)
    return sorted({p.resolve() for p in found}, key=backup_sort_key)


async def _frames_voiceover_candidate(
    session: AsyncSession,
    project_id: int,
    *,
    prefix: str,
    base_priority: int,
) -> VoiceoverCandidate | None:
    frames = (
        await session.execute(
            select(Frame)
            .where(Frame.project_id == project_id)
            .order_by(Frame.number.asc())
        )
    ).scalars().all()
    parts = [(fr.voiceover_text or "").strip() for fr in frames]
    parts = [p for p in parts if p]
    if len(parts) < 2:
        return None
    joined = " ".join(parts)
    return _candidate_from_text(
        joined,
        f"{prefix}frames_db({len(parts)})",
        base_priority + 40,
    )


async def discover_original_candidates(
    session: AsyncSession | None,
    project: Project,
) -> list[VoiceoverCandidate]:
    out: list[VoiceoverCandidate] = []
    out.extend(_scan_dir_candidates(project.data_dir, "", base_priority=0))

    if session is not None:
        c = _candidate_from_text(
            project.script_text or "",
            "script_text",
            priority=100,
        )
        if c:
            out.append(c)

        fc = await _frames_voiceover_candidate(session, project.id, prefix="", base_priority=0)
        if fc:
            out.append(fc)

        children = await list_mass_children(session, project.id)
        children.sort(key=lambda p: p.id)
        for child in children:
            bp = 200 + child.id * 10
            out.extend(
                _scan_dir_candidates(child.data_dir, f"child#{child.id}/", base_priority=bp)
            )
            c2 = _candidate_from_text(
                child.script_text or "",
                f"child#{child.id}/script_text",
                bp + 5,
            )
            if c2:
                out.append(c2)
            fc2 = await _frames_voiceover_candidate(
                session,
                child.id,
                prefix=f"child#{child.id}/",
                base_priority=bp,
            )
            if fc2:
                out.append(fc2)

    # дедуп по тексту — оставляем лучший (минимальный priority)
    best: dict[str, VoiceoverCandidate] = {}
    for c in out:
        prev = best.get(c.text)
        if prev is None or c.priority < prev.priority:
            best[c.text] = c
    return sorted(best.values(), key=lambda x: x.priority)


async def find_original_voiceover(
    session: AsyncSession | None,
    project: Project,
) -> VoiceoverCandidate | None:
    candidates = await discover_original_candidates(session, project)
    if not candidates:
        return None
    return candidates[0]


def is_parent_project(project: Project) -> bool:
    return mass_parent_id(project) is None


def _child_project_result(project: Project) -> dict[str, Any]:
    return {
        "project_id": project.id,
        "slug": project.slug,
        "restored": False,
        "reason": "child_project_skipped",
        "mass_parent_id": mass_parent_id(project),
    }


async def restore_original_voiceover(
    session: AsyncSession,
    project: Project,
    *,
    dry_run: bool = False,
    force: bool = False,
    parents_only: bool = True,
) -> dict[str, Any]:
    if parents_only and not is_parent_project(project):
        return _child_project_result(project)

    from app.services.chatgpt_xlsx import save_voiceover_text

    candidates = await discover_original_candidates(session, project)
    candidate = candidates[0] if candidates else None
    voiceover_path = project.data_dir / "voiceover.txt"
    current = _read_text(voiceover_path) if voiceover_path.is_file() else ""
    current_db = (project.script_text or "").strip()

    if candidate is None:
        return {
            "project_id": project.id,
            "slug": project.slug,
            "restored": False,
            "reason": "no_original_found",
            "current_chars": len(current),
            "current_db_chars": len(current_db),
            "scanned_sources": 0,
        }

    same_as_disk = current == candidate.text
    same_as_db = current_db == candidate.text
    if same_as_disk and same_as_db and not force:
        return {
            "project_id": project.id,
            "slug": project.slug,
            "restored": False,
            "reason": "already_original",
            "source": candidate.source,
            "chars": len(candidate.text),
            "alternatives": len(candidates) - 1,
        }

    if dry_run:
        return {
            "project_id": project.id,
            "slug": project.slug,
            "restored": False,
            "dry_run": True,
            "would_restore": True,
            "source": candidate.source,
            "chars": len(candidate.text),
            "current_chars": len(current),
            "current_db_chars": len(current_db),
            "alternatives": [
                {"source": c.source, "chars": len(c.text)}
                for c in candidates[1:6]
            ],
        }

    save_voiceover_text(project, voiceover_path, candidate.text)
    project.script_text = candidate.text
    await session.flush()
    logger.info(
        "[#{}] restore_original_voiceover: {} симв из {}",
        project.id,
        len(candidate.text),
        candidate.source,
    )
    return {
        "project_id": project.id,
        "slug": project.slug,
        "restored": True,
        "source": candidate.source,
        "chars": len(candidate.text),
        "previous_chars": len(current),
        "alternatives": len(candidates) - 1,
    }


async def restore_all_parent_voiceovers(
    session: AsyncSession,
    *,
    dry_run: bool = False,
    force: bool = False,
) -> dict[str, Any]:
    rows = (await session.execute(select(Project).order_by(Project.id.asc()))).scalars().all()
    parents = [p for p in rows if mass_parent_id(p) is None]
    results: list[dict[str, Any]] = []
    restored = 0
    skipped = 0
    missing = 0
    for p in parents:
        r = await restore_original_voiceover(session, p, dry_run=dry_run, force=force)
        results.append(r)
        if r.get("restored"):
            restored += 1
        elif r.get("reason") == "no_original_found":
            missing += 1
        else:
            skipped += 1
    if not dry_run and restored:
        await session.commit()
    return {
        "parents_total": len(parents),
        "restored": restored,
        "skipped": skipped,
        "missing": missing,
        "dry_run": dry_run,
        "results": results,
    }


async def count_parent_projects(session: AsyncSession) -> int:
    parent_expr = cast(func.json_extract(Project.meta, "$.mass_parent_id"), Integer)
    total = (
        await session.execute(
            select(func.count()).select_from(Project).where(parent_expr.is_(None))
        )
    ).scalar_one()
    return int(total or 0)
