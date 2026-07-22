"""Данные для панели монтажа над нодой assemble (кадры × медиа × Excel)."""

from __future__ import annotations

import asyncio
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any
from urllib.parse import quote

from loguru import logger
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Frame, Project
from app.orchestrator.steps.generate_images import (
    _XLSX_ROWS_PERSONS,
    _find_ref_file_any,
    _parse_ref_ids,
    _resolve_plan_sheet,
)
from app.services.excel_characters import parse_persons_sheet
from app.services.montage_board_cache import (
    get_cached_plan_excel_cells,
    get_cached_shot2_columns,
    probe_video_durations_parallel,
)
from app.services.montage_board_meta import montage_meta, public_board_meta
from app.services.plan_shot2 import (
    MIN_SHOT2_VIDEO_PROMPT_LEN,
    ROW_IMAGE_PROMPT_2_V8,
    ROW_VIDEO_PROMPT_2_V8,
    SHOT2_PROMPT_ATTR,
    SHOT2_VIDEO_PROMPT_ATTR,
    Shot2ColumnInfo,
    find_shot1_image,
    find_shot2_image,
    plan_column_for_frame,
    shot2_video_file_pattern,
)
from app.services.shot2_timeline import split_voiceover_duration
from app.services.xlsx_v8_import import (
    ROW_IMAGE_PROMPT_V8,
    ROW_VIDEO_PROMPT_V8,
    ROW_VOICEOVER_V8,
    _cell_text,
)


def _preview_url(path: Path | None) -> str | None:
    if path is None or not path.is_file():
        return None
    # Кодируем path целиком — пробелы/кириллица иначе ломают <img>/<video>.
    # v=mtime — иначе после replace браузер/прокси может показать старый PNG
    # даже при новом uuid, если path совпал или UI держал кэш по query.
    try:
        mtime_i = int(path.stat().st_mtime)
    except OSError:
        mtime_i = 0
    return f"/api/files?path={quote(str(path), safe='')}&v={mtime_i}"


def _find_shot1_video(videos_dir: Path, frame_number: int) -> Path | None:
    if not videos_dir.is_dir():
        return None
    candidates = [
        p
        for p in videos_dir.glob(f"clip_{frame_number:03d}_*.mp4")
        if "_s2_" not in p.name
    ]
    if not candidates:
        return None
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0]


def _find_shot2_video(videos_dir: Path, frame_number: int) -> Path | None:
    if not videos_dir.is_dir():
        return None
    candidates = list(videos_dir.glob(shot2_video_file_pattern(frame_number)))
    if not candidates:
        return None
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0]


def _character_name_map(xlsx_path: Path) -> dict[str, str]:
    try:
        chars = parse_persons_sheet(xlsx_path)
    except Exception:  # noqa: BLE001
        return {}
    return {c.id.lower(): (c.name or c.id) for c in chars if c.id}


def _merged_plan_ids(ws, col: int, rows: tuple[int, ...]) -> list[str]:
    merged: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for ref_id in _parse_ref_ids(ws.cell(row=row, column=col).value):
            if ref_id not in seen:
                seen.add(ref_id)
                merged.append(ref_id)
    return merged


def _read_plan_excel_cells_uncached(
    xlsx_path: Path,
    *,
    chars_dir: Path,
) -> dict[int, dict[str, Any]]:
    """frame_number → {voiceover_excel, characters, character_refs}."""
    out: dict[int, dict[str, Any]] = {}
    if not xlsx_path.is_file():
        return out
    names = _character_name_map(xlsx_path)
    try:
        wb = load_workbook(filename=str(xlsx_path), data_only=True, read_only=True)
    except Exception as e:  # noqa: BLE001
        logger.warning("montage_board: openpyxl {}: {}", xlsx_path, e)
        return out
    try:
        ws = _resolve_plan_sheet(wb)
        if ws is None:
            return out
        max_col = ws.max_column or 0
        for col in range(3, max_col + 1):
            voice = (_cell_text(ws, ROW_VOICEOVER_V8, col) or "").strip()
            person_ids = _merged_plan_ids(ws, col, _XLSX_ROWS_PERSONS)
            if not voice and not person_ids:
                continue
            frame_num = col - 2
            if frame_num < 1:
                continue
            character_refs: list[dict[str, str | None]] = []
            for ref_id in person_ids:
                image_path = _find_ref_file_any(chars_dir, ref_id)
                character_refs.append(
                    {
                        "id": ref_id,
                        "name": names.get(ref_id.lower(), ref_id),
                        "image_url": _preview_url(image_path),
                    }
                )
            out[frame_num] = {
                "characters": ", ".join(person_ids),
                "voiceover_excel": voice,
                "character_refs": character_refs,
            }
    finally:
        wb.close()
    return out


def _read_plan_excel_cells(
    xlsx_path: Path,
    *,
    chars_dir: Path,
) -> dict[int, dict[str, Any]]:
    return get_cached_plan_excel_cells(
        xlsx_path,
        loader=lambda p: _read_plan_excel_cells_uncached(p, chars_dir=chars_dir),
    )


def _scene_use_durations(
    scene_seconds: float | None,
    *,
    has_shot2: bool,
) -> tuple[float | None, float | None]:
    if scene_seconds is None or scene_seconds <= 0:
        return None, None
    if has_shot2:
        d1, d2 = split_voiceover_duration(scene_seconds)
        return d1, d2
    return round(float(scene_seconds), 3), None


@dataclass(frozen=True, slots=True)
class _FrameBoardSnapshot:
    """Снимок Frame до await/to_thread — без ORM lazy-load / MissingGreenlet."""

    id: int
    number: int
    voiceover_text: str = ""
    start_ts: float | None = None
    end_ts: float | None = None
    duration_seconds: float | None = None
    image_prompt: str = ""
    animation_prompt: str = ""
    attrs: dict[str, Any] = field(default_factory=dict)


def _snapshot_frames(frames: list[Frame]) -> list[_FrameBoardSnapshot]:
    """Читает поля Frame в async greenlet ДО asyncio.to_thread / gather."""
    out: list[_FrameBoardSnapshot] = []
    for fr in frames:
        out.append(
            _FrameBoardSnapshot(
                id=int(fr.id),
                number=int(fr.number),
                voiceover_text=fr.voiceover_text or "",
                start_ts=float(fr.start_ts) if fr.start_ts is not None else None,
                end_ts=float(fr.end_ts) if fr.end_ts is not None else None,
                duration_seconds=(
                    float(fr.duration_seconds)
                    if fr.duration_seconds is not None
                    else None
                ),
                image_prompt=(fr.image_prompt or "").strip(),
                animation_prompt=(fr.animation_prompt or "").strip(),
                attrs=dict(fr.attrs or {}),
            )
        )
    return out


def _empty_prompt_row() -> dict[str, str]:
    return {
        "image_prompt_shot1": "",
        "image_prompt_shot2": "",
        "animation_prompt_shot1": "",
        "animation_prompt_shot2": "",
    }


def _prompts_from_frame_db(frames: list[_FrameBoardSnapshot]) -> dict[int, dict[str, str]]:
    out: dict[int, dict[str, str]] = {}
    for fr in frames:
        attrs = fr.attrs
        out[fr.number] = {
            "image_prompt_shot1": fr.image_prompt,
            "image_prompt_shot2": (attrs.get(SHOT2_PROMPT_ATTR) or "").strip(),
            "animation_prompt_shot1": fr.animation_prompt,
            "animation_prompt_shot2": (attrs.get(SHOT2_VIDEO_PROMPT_ATTR) or "").strip(),
        }
    return out


def _read_source_prompts_once(
    xlsx_path: Path,
    frames: list[_FrameBoardSnapshot],
) -> dict[int, dict[str, str]]:
    """Один openpyxl-проход: R45/R46/R48/R64 для всех кадров.

    frames — plain snapshots, не SQLAlchemy ORM (иначе MissingGreenlet в to_thread).
    """
    out = _prompts_from_frame_db(frames)
    if not frames or not xlsx_path.is_file():
        return out

    excel: dict[int, dict[str, str]] = {
        fr.number: _empty_prompt_row() for fr in frames
    }
    try:
        wb = load_workbook(filename=str(xlsx_path), data_only=True, read_only=True)
    except Exception as e:  # noqa: BLE001
        logger.warning("montage_board: prompts openpyxl {}: {}", xlsx_path, e)
        return out
    try:
        ws = _resolve_plan_sheet(wb)
        if ws is not None:
            for fr in frames:
                col = plan_column_for_frame(fr.number)
                excel[fr.number] = {
                    "image_prompt_shot1": (
                        _cell_text(ws, ROW_IMAGE_PROMPT_V8, col) or ""
                    ).strip(),
                    "image_prompt_shot2": (
                        _cell_text(ws, ROW_IMAGE_PROMPT_2_V8, col) or ""
                    ).strip(),
                    "animation_prompt_shot1": (
                        _cell_text(ws, ROW_VIDEO_PROMPT_V8, col) or ""
                    ).strip(),
                    "animation_prompt_shot2": (
                        _cell_text(ws, ROW_VIDEO_PROMPT_2_V8, col) or ""
                    ).strip(),
                }
    finally:
        wb.close()

    for fr in frames:
        attrs = fr.attrs
        cell = excel.get(fr.number) or {}
        img1 = cell.get("image_prompt_shot1") or fr.image_prompt
        img2 = cell.get("image_prompt_shot2") or (
            attrs.get(SHOT2_PROMPT_ATTR) or ""
        ).strip()
        vid1 = cell.get("animation_prompt_shot1") or fr.animation_prompt
        vid2 = (cell.get("animation_prompt_shot2") or "").strip()
        if len(vid2) < MIN_SHOT2_VIDEO_PROMPT_LEN:
            vid2 = (attrs.get(SHOT2_VIDEO_PROMPT_ATTR) or "").strip()
        out[fr.number] = {
            "image_prompt_shot1": img1,
            "image_prompt_shot2": img2,
            "animation_prompt_shot1": vid1,
            "animation_prompt_shot2": vid2,
        }
    return out


def _load_montage_xlsx_bundle(
    xlsx_path: Path,
    *,
    chars_dir: Path,
    frames: list[_FrameBoardSnapshot],
) -> tuple[
    dict[int, dict[str, Any]],
    dict[int, dict[str, str]],
    dict[int, Shot2ColumnInfo],
]:
    """Последовательно читает Excel в ОДНОМ worker-thread.

    Раньше три параллельных openpyxl на один project.xlsx → на Windows
    PermissionError / lock → API 500 → «Не удалось загрузить данные монтажа».
    """
    excel_by_frame: dict[int, dict[str, Any]] = {}
    prompts_by_frame = _prompts_from_frame_db(frames)
    shot2_by: dict[int, Shot2ColumnInfo] = {}

    if not xlsx_path.is_file():
        return excel_by_frame, prompts_by_frame, shot2_by

    try:
        excel_by_frame = _read_plan_excel_cells(xlsx_path, chars_dir=chars_dir)
    except Exception as e:  # noqa: BLE001
        logger.warning("montage_board: plan excel {}: {}", xlsx_path, e)
        excel_by_frame = {}

    try:
        prompts_by_frame = _read_source_prompts_once(xlsx_path, frames)
    except Exception as e:  # noqa: BLE001
        logger.warning("montage_board: prompts excel {}: {}", xlsx_path, e)

    try:
        shot2_by = get_cached_shot2_columns(xlsx_path)
    except Exception as e:  # noqa: BLE001
        logger.warning("montage_board: shot2 excel {}: {}", xlsx_path, e)
        shot2_by = {}

    return excel_by_frame, prompts_by_frame, shot2_by


def _json_safe_meta(meta: dict[str, Any]) -> dict[str, Any]:
    """Убираем из meta всё, что ломает JSONResponse."""
    import json

    try:
        return json.loads(json.dumps(meta, default=str))
    except Exception:  # noqa: BLE001
        return {
            "video_trims": {},
            "stale_videos": [],
            "highlights": [],
            "corrections": {},
            "pending_ops": [],
            "applied_at": None,
        }


async def build_montage_board(
    session: AsyncSession,
    project: Project,
) -> dict[str, Any]:
    # Project scalars / data_dir — до любого await, пока ORM ещё hot в запросе.
    project_id = int(project.id)
    data_dir = project.data_dir
    try:
        board_meta = _json_safe_meta(public_board_meta(montage_meta(project)))
    except Exception as e:  # noqa: BLE001
        logger.warning("montage_board: meta project {}: {}", project_id, e)
        board_meta = _json_safe_meta({})

    async def _load_frames() -> list[Frame]:
        return list(
            (
                await session.execute(
                    select(Frame)
                    .where(Frame.project_id == project_id)
                    .order_by(Frame.number.asc())
                )
            ).scalars().all()
        )

    frames_orm = await _load_frames()
    # Ручной перенос scenes/videos в новый проект: без Frame в БД доска пустая.
    # project.xlsx → sync; иначе создаём кадры по номерам файлов на диске.
    if not frames_orm:
        xlsx_boot = data_dir / "project.xlsx"
        if xlsx_boot.is_file():
            try:
                from app.services.chatgpt_xlsx import sync_project_xlsx

                await sync_project_xlsx(
                    session,
                    project,
                    xlsx_boot,
                    keep_fields=True,
                    update_frames_voiceover=True,
                )
                frames_orm = await _load_frames()
            except Exception as e:  # noqa: BLE001
                logger.warning(
                    "montage_board: xlsx bootstrap project {}: {}", project_id, e
                )
    try:
        from app.services.ensure_frames_from_disk import ensure_frames_from_disk_media

        created = await ensure_frames_from_disk_media(session, project)
        if created:
            frames_orm = await _load_frames()
    except Exception as e:  # noqa: BLE001
        logger.warning(
            "montage_board: disk frames bootstrap project {}: {}", project_id, e
        )

    # ORM только здесь; дальше — plain snapshots (to_thread не трогает Session).
    frames = _snapshot_frames(frames_orm)

    xlsx_path = data_dir / "project.xlsx"
    chars_dir = data_dir / "characters"
    # Один thread, последовательные openpyxl — без Windows file lock.
    excel_by_frame, prompts_by_frame, shot2_by = await asyncio.to_thread(
        _load_montage_xlsx_bundle,
        xlsx_path,
        chars_dir=chars_dir,
        frames=frames,
    )
    scenes_dir = data_dir / "scenes"
    videos_dir = data_dir / "videos"

    frame_videos: list[
        tuple[_FrameBoardSnapshot, Path | None, Path | None, dict, bool, bool]
    ] = []
    all_vid_paths: list[Path | None] = []
    for fr in frames:
        ex = excel_by_frame.get(fr.number, {})
        vid1 = _find_shot1_video(videos_dir, fr.number)
        vid2 = _find_shot2_video(videos_dir, fr.number)
        shot2_info = shot2_by.get(fr.number)
        has_shot2 = bool(shot2_info is not None and shot2_info.has_shot2)
        has_shot2_video = has_shot2 and vid2 is not None and vid2.is_file()
        frame_videos.append((fr, vid1, vid2, ex, has_shot2, has_shot2_video))
        all_vid_paths.extend([vid1, vid2])

    try:
        durations = await probe_video_durations_parallel(all_vid_paths)
    except Exception as e:  # noqa: BLE001
        logger.warning("montage_board: ffprobe project {}: {}", project_id, e)
        durations = [None] * len(all_vid_paths)
    dur_iter = iter(durations)

    rows: list[dict[str, Any]] = []
    for fr, vid1, vid2, ex, has_shot2, has_shot2_video in frame_videos:
        img1 = find_shot1_image(scenes_dir, fr.number)
        img2 = find_shot2_image(scenes_dir, fr.number)
        scene_seconds = (
            fr.duration_seconds
            if fr.duration_seconds is not None and fr.duration_seconds > 0
            else None
        )
        shot1_use, shot2_use = _scene_use_durations(scene_seconds, has_shot2=has_shot2_video)
        vid1_dur = next(dur_iter)
        vid2_dur = next(dur_iter)
        vo_start = fr.start_ts
        vo_end = fr.end_ts
        shot1_timeline_start = vo_start
        shot1_timeline_end = (
            round(vo_start + shot1_use, 3)
            if vo_start is not None and shot1_use is not None
            else None
        )
        shot2_timeline_start = shot1_timeline_end
        shot2_timeline_end = vo_end if has_shot2 else None
        prompts = prompts_by_frame.get(fr.number) or {}

        rows.append(
            {
                "frame_id": fr.id,
                "number": fr.number,
                "voiceover_text": fr.voiceover_text,
                "voiceover_excel": ex.get("voiceover_excel") or "",
                "characters": ex.get("characters") or "",
                "character_refs": ex.get("character_refs") or [],
                "start_ts": fr.start_ts,
                "end_ts": fr.end_ts,
                "duration_seconds": fr.duration_seconds,
                "has_shot2": has_shot2,
                "shot1_use_seconds": shot1_use,
                "shot2_use_seconds": shot2_use,
                "shot1_timeline_start": shot1_timeline_start,
                "shot1_timeline_end": shot1_timeline_end,
                "shot2_timeline_start": shot2_timeline_start,
                "shot2_timeline_end": shot2_timeline_end,
                "video_shot1_duration": vid1_dur,
                "video_shot2_duration": vid2_dur,
                "image_shot1_url": _preview_url(img1),
                "image_shot2_url": _preview_url(img2),
                "video_shot1_url": _preview_url(vid1),
                "video_shot2_url": _preview_url(vid2),
                "image_prompt_shot1": prompts.get("image_prompt_shot1") or "",
                "image_prompt_shot2": prompts.get("image_prompt_shot2") or "",
                "animation_prompt_shot1": prompts.get("animation_prompt_shot1") or "",
                "animation_prompt_shot2": prompts.get("animation_prompt_shot2") or "",
                "plan_column": plan_column_for_frame(fr.number),
            }
        )

    return {"frames": rows, "frame_count": len(rows), "meta": board_meta}
