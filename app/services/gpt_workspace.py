"""Studio GPT workspace — свободный чат через API (история, вложения, сохранение).

Файлы: data/gpt_workspace/<session_id>/
  meta.json, messages.json, attachments/, outputs/
"""

from __future__ import annotations

import json
import re
import shutil
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import quote

from loguru import logger

from app.settings import settings

_SAFE_NAME = re.compile(r"[^a-zA-Z0-9._\-а-яА-ЯёЁ ]+")

# Vision ≠ исходные байты. Контент файлов — из ответа GPT / скачанных URL.
_WORKSPACE_SYSTEM = (
    "Ты в чате Studio GPT (HTTP API). Вложения пользователя уже приложены.\n"
    "Изображения во вложениях — vision: ты их видишь и анализируешь.\n"
    "ЗАПРЕЩЕНО писать, что не можешь прикрепить файл / нет инструмента вложений.\n"
    "Документ / договор / .docx / «отправь файл» — выдай ПОЛНЫЙ текст файла "
    "без мета-оговорок. Клиент сохранит ответ как скачиваемый файл.\n"
    "Пустой .txt — ответь пустой строкой или одним переводом строки, без пояснений.\n"
    "Картинка / фото / арт «найди / пришли»: НЕ генерируй и НЕ шли data:image/base64. "
    "Верни 1–3 РЕАЛЬНЫХ https URL (прямая .png/.jpg/.webp или страница Wikipedia/wiki "
    "с картинкой). Без выдуманных ссылок и без имён файлов без URL.\n"
    "Excel — TSV с строками «# Лист: …». Word — обычный текст документа.\n"
    "Если просят и Excel, и Word — сначала блок(и) # Лист:, затем текст для .docx.\n"
    "Неизвестные поля — прочерки «—»."
)

# Пользователь хочет файл в «Результаты» (не обязательно исходник).
_ASKS_FILE_RE = re.compile(
    r"(?i)("
    r"верн[иу]|пришл[иу]|отправ[ьи]|скача[йть]|дай\s+файл|выгруз|исходник|"
    r"файл[оа]м|как\s+файл|в\s+виде\s+файл|прилож[иь]|"
    r"\.docx|\.txt|\.md|\bdocx\b|word|ворд|"
    r"download|send\s+(back\s+)?(the\s+)?file|as\s+a\s+file"
    r")"
)

# Нужна работа модели (не просто «положи уже готовое»).
_NEEDS_WORK_RE = re.compile(
    r"(?i)("
    r"анализ|опис|что\s+(на|в)\s+|расскаж|проверь|сравн|"
    r"перепиши|переработ|передел|обработ|отредактир|исправ|"
    r"сгенерир|сформируй|состав[ьи]|подготов[ьи]|напиши|написать|"
    r"сделай|измени|договор|контракт|соглашени|"
    r"describe|analy[sz]e|compare|explain|rewrite|rework"
    r")"
)

_DOCX_RE = re.compile(r"(?i)(\.docx|\bdocx\b|word|ворд)")
_XLSX_RE = re.compile(
    r"(?i)(\.xlsx|\.xls|\bxlsx\b|\bexcel\b|эксел[ьяю]|таблиц[аеуы])"
)
_IMAGE_ASK_RE = re.compile(
    r"(?i)("
    r"картин?к|"  # картинка/картинку + опечатка «картику»
    r"изображен|рисунок|фото|пикч|пиксель|арт\b|скрин|"
    r"\.png|\.jpe?g|\.webp|\.gif|\bpng\b|\bimage\b|\bpicture\b|\bphoto\b"
    r")"
)
_DATA_URI_INLINE_RE = re.compile(
    r"data:(?:image|application)/[^;,\s]+;base64,[A-Za-z0-9+/=\s]+",
    re.IGNORECASE,
)

# «пустой txt» — контент всё равно от GPT (может быть пустой строкой).
_BLANK_FILE_RE = re.compile(
    r"(?i)("
    r"пуст(?:ой|ая|ое|ую|ых)?\s*(?:\.?txt|\.?docx|\.?md|файл|текст|document)?"
    r"|empty\s+(?:\.?txt|\.?docx|file|text|document)"
    r"|blank\s+(?:\.?txt|file)"
    r")"
)

_EXCUSE_RE = re.compile(
    r"(?im)^.*("
    r"недоступен инструмент|не\s+могу\s+прикрепить|не\s+смогу\s+прикрепить|"
    r"генерации файлов|создания новых вложений|прикрепить\s+\.docx|"
    r"cannot\s+attach|no\s+attachment\s+tool"
    r").*(?:\n|$)"
)

# Совместимость со старыми тестами / именами
_RETURN_FILE_RE = _ASKS_FILE_RE
_FILE_DELIVERY_RE = _ASKS_FILE_RE
_ANALYZE_RE = _NEEDS_WORK_RE


def _asks_file(message: str) -> bool:
    return bool(_ASKS_FILE_RE.search(message or ""))


def _is_meta_chat_question(message: str) -> bool:
    """Вопрос — ответить в чате, не слать/паковать файл.

    Правило: если в конце `?` / `？` — всегда вопрос.
    Также жалобы/meta без знака («я вопрос задал», «почему не прислал»).
    """
    t = (message or "").strip()
    if not t:
        return False
    # хвостовая пунктуация: «…файл???» / «файл？»
    if re.search(r"[?？]\s*$", t):
        return True
    if re.search(
        r"(?i)("
        r"почему|зачем|как\s+так| wh?y\b|"
        r"я\s+вопрос|это\s+(?:был\s+)?вопрос|ответь\s+на\s+вопрос|"
        r"объясни|что\s+значит|"
        r"с\s+первого\s+раза|"
        r"не\s+(?:прислал|отправил|дал|положил)"
        r")",
        t,
    ):
        return True
    return False


def _is_short_affirmative(message: str) -> bool:
    return bool(
        re.fullmatch(
            r"(?is)\s*(да|даа+|ага|угу|ок|окей|okay|yes|lf|конечно|сделай|давай)\s*[.!…]?\s*",
            message or "",
        )
    )


def _looks_like_document(text: str) -> bool:
    """Длинный текст договора/документа в пузыре — кандидат на .txt/.docx."""
    t = (text or "").strip()
    if len(t) < 400:
        return False
    if re.match(r"(?i)^(готовые файлы|studio\s+положила|studio\s+вернула)\b", t):
        return False
    score = 0
    if re.search(r"(?i)\bдоговор\b", t):
        score += 2
    if re.search(
        r"(?i)(реквизит|исполнител|заказчик|предмет\s+договора|сторон[ыа])",
        t,
    ):
        score += 1
    if len(re.findall(r"(?m)^\s*\d+(?:\.\d+)+\.?\s", t)) >= 5:
        score += 1
    if len(t) >= 2500:
        score += 1
    return score >= 2


def _needs_work(message: str) -> bool:
    return bool(_NEEDS_WORK_RE.search(message or ""))


def _wants_analysis(message: str) -> bool:
    return _needs_work(message)


def _wants_docx(message: str) -> bool:
    return bool(_DOCX_RE.search(message or ""))


def _wants_xlsx(message: str) -> bool:
    return bool(_XLSX_RE.search(message or ""))


def _wants_image_file(message: str) -> bool:
    return bool(_IMAGE_ASK_RE.search(message or ""))


def _strip_media_payloads(text: str) -> str:
    """Убрать data-URI / прямые URL картинок из текста пузыря."""
    t = _DATA_URI_INLINE_RE.sub("", text or "")
    t = re.sub(
        r"https?://\S+\.(?:png|jpe?g|webp|gif|bmp|svg)(?:\?\S*)?",
        "",
        t,
        flags=re.I,
    )
    return re.sub(r"\n{3,}", "\n\n", t).strip()


def _is_media_heavy_reply(reply: str) -> bool:
    """Ответ почти целиком картинка/data-URI — не паковать в .txt."""
    raw = (reply or "").strip()
    if not raw or not _DATA_URI_INLINE_RE.search(raw):
        return False
    return len(_strip_media_payloads(raw)) < 40


def _explicit_text_doc_ask(message: str) -> bool:
    """Явно просят текстовый/док файл (не «пришли картинку»)."""
    t = message or ""
    return bool(
        re.search(
            r"(?i)("
            r"\.txt|\.md|\btxt\b|\bmd\b|"
            r"пуст(?:ой|ая|ое|ую).*(?:файл|txt|текст)|"
            r"текст(?:ом|овый)?\s+файл|"
            r"как\s+файл|в\s+виде\s+файл|файл[оа]м|"
            r"договор|контракт|соглашени"
            r")",
            t,
        )
    )


def _should_pack_text_document(
    user_text: str, reply: str, *, media_count: int
) -> bool:
    """Паковать .txt/.docx только для явных документных запросов."""
    if _wants_image_file(user_text) and not _explicit_text_doc_ask(user_text):
        return False
    if _wants_blank_file(user_text):
        return True
    if _wants_docx(user_text):
        return True
    if _wants_xlsx(user_text) and not _explicit_text_doc_ask(user_text):
        return False
    if media_count > 0 and _is_media_heavy_reply(reply):
        return False
    if _explicit_text_doc_ask(user_text):
        return True
    # «отправь/верни/скачай файл» без картинки
    if re.search(
        r"(?i)(отправ[ьи]|верн[иу]|скача[йть]|дай)\s+(?:мне\s+)?файл",
        user_text or "",
    ):
        return True
    # НЕ пакуем от голого «пришли …» (иначе любой запрос картинки → .txt)
    return False


def _write_simple_xlsx(path: Path, text: str) -> None:
    """Минимальный .xlsx из ответа GPT (TSV / # Лист / строки)."""
    from openpyxl import Workbook

    from app.services.xlsx_text_writeback import extract_sheet_blocks

    blocks = extract_sheet_blocks(text or "")
    wb = Workbook()
    if blocks:
        default = wb.active
        if default is not None:
            wb.remove(default)
        for name, rows in blocks.items():
            ws = wb.create_sheet(title=(name or "Данные")[:31])
            for r_i, row in enumerate(rows, start=1):
                for c_i, val in enumerate(row, start=1):
                    ws.cell(r_i, c_i, val)
    else:
        ws = wb.active
        ws.title = "Данные"
        lines = [
            ln
            for ln in _strip_media_payloads(text or "").splitlines()
            if ln.strip()
        ]
        if not lines:
            ws.cell(1, 1, "")
        else:
            for r_i, line in enumerate(lines[:800], start=1):
                if "\t" in line:
                    for c_i, val in enumerate(line.split("\t"), start=1):
                        ws.cell(r_i, c_i, val)
                else:
                    ws.cell(r_i, 1, line)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _strip_attachment_excuses(text: str) -> str:
    cleaned = _EXCUSE_RE.sub("", text or "")
    return re.sub(r"\n{3,}", "\n\n", cleaned).strip()


def _last_assistant_document(prior_raw: list[Any]) -> str:
    """Последний документ в чате (не «Готовые файлы» и не отмазка GPT)."""
    fallback = ""
    for m in reversed(prior_raw or []):
        if not isinstance(m, dict) or m.get("role") != "assistant":
            continue
        content = str(m.get("content") or "").strip()
        if not content:
            continue
        content = re.split(
            r"\n—\n(?:Studio|Исходники)|"
            r"\n\nГотовые файлы\b",
            content,
            maxsplit=1,
            flags=re.I,
        )[0].strip()
        content = _strip_attachment_excuses(content)
        content = _strip_media_payloads(content)
        if not content or len(content) < 80:
            continue
        if re.match(r"(?i)^(готовые файлы|studio\s+положила|studio\s+вернула)\b", content):
            continue
        if _looks_like_document(content):
            return content
        if len(content) > len(fallback):
            fallback = content
    return fallback


def _prior_turn_wanted_file(prior_raw: list[Any]) -> bool:
    """Недавний запрос пользователя про файл / уточнение ассистента про выдачу файла."""
    users: list[str] = []
    last_assistant = ""
    for m in reversed(prior_raw or []):
        if not isinstance(m, dict):
            continue
        role = m.get("role")
        content = str(m.get("content") or "").strip()
        if role == "user" and content:
            users.append(content)
            if len(users) >= 3:
                break
        elif role == "assistant" and content and not last_assistant:
            last_assistant = content
    for u in users:
        if _is_meta_chat_question(u):
            continue
        if _asks_file(u) or _explicit_text_doc_ask(u) or _wants_docx(u):
            return True
    if last_assistant and re.search(
        r"(?i)("
        r"\.txt|\.docx|файл|выдать|пришл|отправ|"
        r"готовые\s+файлы|убрать\s+лишн"
        r")",
        last_assistant,
    ):
        return True
    return False


def _image_search_query(message: str) -> str:
    """Текст запроса картинки без «пришли мне картинку»."""
    t = (message or "").strip()
    t = re.sub(
        r"(?i)^(пришл[иу]|отправ[ьи]|найд[ий]|скача[йть]|дай|покаж[иь])\s+"
        r"(мне\s+)?(пожалуйста\s+)?",
        "",
        t,
    ).strip()
    t = re.sub(
        r"(?i)^(картин?к\w*|изображен\w*|фото|рисунок|арт|пикч\w*|image|picture|photo)\s+"
        r"(of\s+|с\s+|про\s+)?",
        "",
        t,
    ).strip()
    t = re.sub(
        r"(?i)\b(из\s+интернета|из\s+сети|в\s+интернете|ссылк\w*|url|пожалуйста)\b",
        "",
        t,
    ).strip(" ,.-")
    return t or (message or "").strip()


def _image_query_variants(message_or_query: str) -> list[str]:
    """Несколько поисковых формулировок (RU/EN), чтобы Openverse находил арт."""
    base = _image_search_query(message_or_query)
    variants: list[str] = []
    seen: set[str] = set()

    def add(s: str) -> None:
        s = re.sub(r"\s+", " ", (s or "").strip())
        if not s:
            return
        key = s.lower()
        if key in seen:
            return
        seen.add(key)
        variants.append(s)

    add(base)
    low = base.lower()
    if re.search(r"[а-яё]", low):
        if "фантом" in low and "ассас" in low:
            add("Phantom Assassin Dota 2")
        if "дот" in low:
            add(re.sub(r"(?i)из\s+доты|в\s+доте|дота\s*2?|доты", "Dota 2", base))
            add(f"{base} Dota 2")
        add(f"{base} art")
        add(f"{base} official")
    else:
        add(f"{base} art")
        if "dota" not in low and re.search(
            r"(?i)assassin|phantom|invoker|pudge", low
        ):
            add(f"{base} Dota 2")
    if "бел" in low or "white" in low:
        add((variants[0] if variants else base) + " white background")
    return variants


def _is_placeholder_image(path: Path) -> bool:
    """1×1 / крошечный data-URI — заглушка модели, не реальная картинка."""
    import struct

    try:
        data = path.read_bytes()
    except OSError:
        return True
    suf = path.suffix.lower()
    if suf not in {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp", ".svg"}:
        return False
    if len(data) < 64:
        return True
    if data.startswith(b"\x89PNG") and len(data) >= 24:
        w, h = struct.unpack(">II", data[16:24])
        if w <= 2 or h <= 2:
            return True
        return False
    if data.startswith(b"\xff\xd8\xff") and len(data) < 400:
        return True
    if suf == ".svg" and len(data) < 80:
        return True
    return False


def _filter_placeholder_images(out_dir: Path, names: list[str]) -> list[str]:
    kept: list[str] = []
    for name in names:
        p = out_dir / Path(name).name
        if p.is_file() and _is_placeholder_image(p):
            logger.warning(
                "gpt_workspace: drop placeholder image {} ({} B)",
                p.name,
                p.stat().st_size,
            )
            try:
                p.unlink()
            except OSError:
                pass
            continue
        kept.append(name)
    return kept


def _wants_blank_file(message: str) -> bool:
    """Просят пустой файл — контент от GPT может быть пустым."""
    text = (message or "").strip()
    return bool(text) and bool(_BLANK_FILE_RE.search(text)) and (
        _asks_file(text)
        or bool(re.search(r"(?i)(\.txt|\.docx|\.md|\btxt\b|файл)", text))
    )


def resolve_file_intent(
    message: str,
    *,
    has_attachments: bool,
    last_doc: str,
    prior_raw: list[Any] | None = None,
) -> str:
    """Одна развилка вместо кучи флагов.

    return_attachments — скопировать вложения пользователя (байты юзера)
    pack_last         — упаковать предыдущий ответ ассистента
    pack_reply        — вызвать GPT и упаковать его ответ
    none              — обычный чат
    """
    text = (message or "").strip()
    if not text:
        return "none"
    # «почему не прислал файлом?» / «я вопрос задал» — ответ текстом, не новый .txt
    if _is_meta_chat_question(text):
        return "none"
    asks = _asks_file(text)
    work = _needs_work(text)
    short = len(text) < 240
    has_doc = len(last_doc or "") >= 80 or _looks_like_document(last_doc or "")

    # «да» после «убрать ** и пришли txt» — продолжить работу и сразу упаковать
    if _is_short_affirmative(text) and _prior_turn_wanted_file(prior_raw or []):
        return "pack_reply"

    # «составь договор и пришли файлом» / «переработай и отправь» / «пустой txt»
    if asks and work:
        return "pack_reply"
    if asks and _wants_blank_file(text):
        return "pack_reply"
    # «отправь файл» после готового текста в чате
    if asks and short and has_doc and not work:
        return "pack_last"
    # «верни файл» — только исходники со скрепки
    if asks and short and has_attachments and not work and not has_doc:
        return "return_attachments"
    # «отправь файл» без контекста — GPT, затем сохранить ответ как файл
    if asks:
        return "pack_reply"
    # договор без слова «файл»
    if work and re.search(r"(?i)(договор|контракт|соглашени|\.docx)", text):
        return "pack_reply"
    return "none"


# --- aliases для старых тестов ---
def _wants_file_return(message: str) -> bool:
    return _asks_file(message) and not _needs_work(message)


def _pure_file_return(message: str) -> bool:
    text = (message or "").strip()
    return bool(text) and _wants_file_return(text) and len(text) < 240


def _wants_deliverable_file(message: str) -> bool:
    return _asks_file(message) or bool(
        re.search(r"(?i)(договор|контракт|соглашени)", message or "")
        and _needs_work(message)
    )


def _write_simple_docx(path: Path, text: str) -> None:
    """Минимальный .docx (OOXML zip) без python-docx — открывается в Word/LibreOffice."""
    import zipfile
    from xml.sax.saxutils import escape

    lines = (text or "").replace("\r\n", "\n").replace("\r", "\n").split("\n")
    body_parts: list[str] = []
    for line in lines:
        body_parts.append(
            "<w:p><w:r><w:t xml:space=\"preserve\">"
            f"{escape(line)}"
            "</w:t></w:r></w:p>"
        )
    if not body_parts:
        body_parts.append("<w:p><w:r><w:t></w:t></w:r></w:p>")
    document_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        f"<w:body>{''.join(body_parts)}<w:sectPr/></w:body></w:document>"
    )
    content_types = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
</Types>
"""
    rels = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>
</Relationships>
"""
    doc_rels = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"/>
"""
    path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", content_types)
        zf.writestr("_rels/.rels", rels)
        zf.writestr("word/document.xml", document_xml)
        zf.writestr("word/_rels/document.xml.rels", doc_rels)


def _deliver_reply_as_file(
    *,
    out_dir: Path,
    reply: str,
    user_text: str,
    attachments: list[Path],
    ts: str,
) -> Path | None:
    """Сохранить текст ответа GPT как .txt/.docx (без data-URI)."""
    allow_empty = _wants_blank_file(user_text)
    body = _strip_media_payloads(_strip_attachment_excuses(reply))
    if len(body) < 40:
        body = _strip_media_payloads((reply or "").strip())
    if len(body) < 20 and not allow_empty:
        return None
    stem = "empty" if allow_empty and len(body) < 20 else "document"
    for p in attachments:
        if p.suffix.lower() in {".txt", ".md", ".csv", ".tsv", ".docx"}:
            stem = p.stem[:48] or stem
            break
    if re.search(r"(?i)договор", user_text):
        stem = "dogovor"
    elif re.search(r"(?i)контракт", user_text):
        stem = "contract"
    if _wants_docx(user_text):
        dest = out_dir / f"{stem}_{ts}.docx"
        _write_simple_docx(dest, body)
    else:
        dest = out_dir / f"{stem}_{ts}.txt"
        if allow_empty and not body:
            dest.write_bytes(b"")
        else:
            dest.write_text(body, encoding="utf-8")
    if not dest.is_file():
        return None
    if allow_empty:
        return dest
    return dest if dest.stat().st_size > 32 else None


def _file_urls(path: Path) -> dict[str, str]:
    """URL просмотра и принудительного скачивания через /api/files."""
    q = quote(str(path), safe="")
    return {
        "url": f"/api/files?path={q}",
        "download_url": f"/api/files?path={q}&download=1",
    }


def _file_entry(path: Path) -> dict[str, Any]:
    """Карточка файла: sniff-rename на диске + mime/kind для превью в GPT-окне."""
    from app.services.gpt_api import ensure_correct_extension, suggested_name_and_mime

    fixed = ensure_correct_extension(path)
    display_name, mime = suggested_name_and_mime(fixed)
    kind = "image" if (mime or "").startswith("image/") else "file"
    if kind != "image" and fixed.suffix.lower() in {
        ".png",
        ".jpg",
        ".jpeg",
        ".webp",
        ".gif",
        ".bmp",
        ".svg",
        ".avif",
        ".heic",
        ".ico",
    }:
        kind = "image"
    return {
        "name": fixed.name,
        "display_name": display_name or fixed.name,
        "size": fixed.stat().st_size if fixed.is_file() else 0,
        "path": str(fixed),
        "mime": mime,
        "kind": kind,
        **_file_urls(fixed),
    }


def _root() -> Path:
    d = Path(settings.data_dir) / "gpt_workspace"
    d.mkdir(parents=True, exist_ok=True)
    return d


def _session_dir(session_id: str) -> Path:
    sid = "".join(c for c in session_id if c.isalnum() or c in "-_")
    if not sid:
        raise ValueError("пустой session_id")
    return _root() / sid


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _read_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:  # noqa: BLE001
        return default


def _write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def list_sessions() -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for d in sorted(_root().iterdir(), key=lambda p: p.stat().st_mtime, reverse=True):
        if not d.is_dir():
            continue
        meta = _read_json(d / "meta.json", {})
        if not isinstance(meta, dict):
            continue
        msgs = _read_json(d / "messages.json", [])
        items.append(
            {
                "id": d.name,
                "title": str(meta.get("title") or d.name),
                "created_at": meta.get("created_at"),
                "updated_at": meta.get("updated_at"),
                "message_count": len(msgs) if isinstance(msgs, list) else 0,
                "status": meta.get("status") or "idle",
            }
        )
    return items


def create_session(*, title: str | None = None) -> dict[str, Any]:
    sid = uuid.uuid4().hex[:12]
    d = _session_dir(sid)
    d.mkdir(parents=True, exist_ok=True)
    (d / "attachments").mkdir(exist_ok=True)
    (d / "outputs").mkdir(exist_ok=True)
    now = _now()
    meta = {
        "id": sid,
        "title": (title or "").strip() or f"Чат {now[:16].replace('T', ' ')}",
        "created_at": now,
        "updated_at": now,
        "status": "idle",
    }
    _write_json(d / "meta.json", meta)
    _write_json(d / "messages.json", [])
    return get_session(sid)


def reset_running_sessions_on_startup() -> dict[str, Any]:
    """После рестарта бэкенда in-flight ask мёртв — UI не должен крутить «думает»."""
    root = _root()
    if not root.is_dir():
        return {"reset": 0}
    reset = 0
    for d in root.iterdir():
        if not d.is_dir():
            continue
        meta_path = d / "meta.json"
        if not meta_path.is_file():
            continue
        meta = _read_json(meta_path, {})
        if not isinstance(meta, dict):
            continue
        if (meta.get("status") or "").lower() != "running":
            continue
        meta["status"] = "error"
        meta["phase"] = "error"
        meta["phase_detail"] = "Сброс после рестарта: запрос GPT прерван"
        meta["last_error"] = "backend restart — in-flight GPT ask cancelled"
        meta["updated_at"] = _now()
        _write_json(meta_path, meta)
        _append_message(
            d.name,
            "system",
            "Ошибка: бэкенд перезапущен во время ответа GPT. Статус сброшен — "
            "отправь сообщение ещё раз.",
        )
        reset += 1
        logger.warning("gpt_workspace: reset orphan running session={}", d.name)
    return {"reset": reset}


# Если Studio крутит «думает» дольше этого — считаем зависанием провайдера/ретраев.
_STALE_RUNNING_SEC = 8 * 60


def _maybe_reset_stale_running(session_id: str, meta: dict[str, Any]) -> dict[str, Any]:
    if (meta.get("status") or "").lower() != "running":
        return meta
    updated = str(meta.get("updated_at") or "").strip()
    if not updated:
        return meta
    try:
        ts = datetime.fromisoformat(updated.replace("Z", "+00:00"))
        if ts.tzinfo is None:
            ts = ts.replace(tzinfo=timezone.utc)
        age = (datetime.now(timezone.utc) - ts).total_seconds()
    except Exception:  # noqa: BLE001
        return meta
    if age < _STALE_RUNNING_SEC:
        return meta
    meta = dict(meta)
    detail = (
        f"GPT не ответил за {int(age // 60)} мин (таймаут/зависание провайдера). "
        "Отправь сообщение ещё раз."
    )
    meta["status"] = "error"
    meta["phase"] = "error"
    meta["phase_detail"] = detail
    meta["last_error"] = detail
    meta["updated_at"] = _now()
    _write_json(_session_dir(session_id) / "meta.json", meta)
    _append_message(session_id, "system", f"Ошибка: {detail}")
    logger.warning(
        "gpt_workspace: stale running session={} age={:.0f}s → error",
        session_id,
        age,
    )
    return meta


def get_session(session_id: str) -> dict[str, Any]:
    d = _session_dir(session_id)
    if not d.is_dir():
        raise FileNotFoundError(f"сессия не найдена: {session_id}")
    # Лениво: .bin на диске → png/jpg по magic (в т.ч. старые сессии)
    renames = _normalize_session_files(d)
    meta = _read_json(d / "meta.json", {})
    if not isinstance(meta, dict):
        meta = {}
    meta = _maybe_reset_stale_running(session_id, meta)
    messages = _read_json(d / "messages.json", [])
    if renames and isinstance(messages, list):
        messages = _rewrite_message_filenames(messages, renames)
        _write_json(d / "messages.json", messages)
    attachments = []
    att_dir = d / "attachments"
    if att_dir.is_dir():
        for p in sorted(att_dir.iterdir()):
            if p.is_file():
                attachments.append(_file_entry(p))
    outputs = []
    out_dir = d / "outputs"
    if out_dir.is_dir():
        for p in sorted(out_dir.iterdir(), key=lambda x: x.stat().st_mtime, reverse=True):
            if p.is_file():
                outputs.append(_file_entry(p))
    return {
        "id": d.name,
        "title": meta.get("title") or d.name,
        "created_at": meta.get("created_at"),
        "updated_at": meta.get("updated_at"),
        "status": meta.get("status") or "idle",
        "phase": meta.get("phase") or "",
        "phase_detail": meta.get("phase_detail") or "",
        "messages": messages if isinstance(messages, list) else [],
        "attachments": attachments,
        "outputs": outputs,
    }


def _normalize_session_files(d: Path) -> dict[str, str]:
    """Переименовать .bin/слабые расширения по magic/CT. {old_name: new_name}."""
    from app.services.gpt_api import ensure_correct_extension, finalize_downloaded_file

    renames: dict[str, str] = {}
    for sub in ("attachments", "outputs"):
        folder = d / sub
        if not folder.is_dir():
            continue
        for p in list(folder.iterdir()):
            if not p.is_file():
                continue
            old = p.name
            if p.suffix.lower() in {".bin", ".download", ".octet-stream", ".dat", ".tmp"}:
                fixed = finalize_downloaded_file(p)
            else:
                fixed = ensure_correct_extension(p)
            if fixed.name != old:
                renames[old] = fixed.name
    return renames


def _rewrite_message_filenames(
    messages: list[Any], renames: dict[str, str]
) -> list[Any]:
    out: list[Any] = []
    for m in messages:
        if not isinstance(m, dict):
            out.append(m)
            continue
        nm = dict(m)
        for key in ("attachment_names", "output_files"):
            names = nm.get(key)
            if isinstance(names, list):
                nm[key] = [renames.get(str(n), n) for n in names]
        content = nm.get("content")
        if isinstance(content, str) and content:
            for old, new in renames.items():
                if old and new and old != new and old in content:
                    content = content.replace(old, new)
            nm["content"] = content
        out.append(nm)
    return out


def delete_session(session_id: str) -> None:
    d = _session_dir(session_id)
    if d.is_dir():
        shutil.rmtree(d, ignore_errors=True)


def rename_session(session_id: str, title: str) -> dict[str, Any]:
    d = _session_dir(session_id)
    meta = _read_json(d / "meta.json", {})
    meta["title"] = (title or "").strip() or meta.get("title") or session_id
    meta["updated_at"] = _now()
    _write_json(d / "meta.json", meta)
    return get_session(session_id)


def _safe_filename(name: str) -> str:
    base = Path(name).name
    cleaned = _SAFE_NAME.sub("_", base).strip(" ._") or "file"
    return cleaned[:120]


def save_attachment(session_id: str, filename: str, data: bytes) -> dict[str, Any]:
    from app.services.gpt_api import resolve_bytes_extension, sniff_file_extension

    d = _session_dir(session_id)
    if not d.is_dir() or not (d / "meta.json").is_file():
        raise FileNotFoundError(f"сессия не найдена: {session_id}")
    if not data:
        raise ValueError("пустой файл")
    att = d / "attachments"
    att.mkdir(exist_ok=True)
    safe = _safe_filename(filename)
    # Никогда не оставляем .bin — magic / fallback .dat
    sniffed = sniff_file_extension(data) or resolve_bytes_extension(data, fallback=".dat")
    suf = Path(safe).suffix.lower()
    if suf in {"", ".bin", ".dat", ".tmp", ".octet-stream", ".download"} or (
        sniffed.startswith(".")
        and sniffed in {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp", ".svg", ".avif"}
        and suf not in {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp", ".svg", ".avif"}
    ):
        safe = f"{Path(safe).stem}{sniffed}"
    path = att / safe
    # уникальность
    if path.exists():
        stem, suf = path.stem, path.suffix
        i = 2
        while True:
            cand = att / f"{stem}_{i}{suf}"
            if not cand.exists():
                path = cand
                break
            i += 1
    path.write_bytes(data)
    meta = _read_json(d / "meta.json", {})
    meta["updated_at"] = _now()
    _write_json(d / "meta.json", meta)
    return _file_entry(path)


def delete_attachment(session_id: str, name: str) -> None:
    d = _session_dir(session_id)
    if not d.is_dir() or not (d / "meta.json").is_file():
        raise FileNotFoundError(f"сессия не найдена: {session_id}")
    path = d / "attachments" / Path(name).name
    if path.exists():
        path.unlink()


def copy_attachment_to_outputs(session_id: str, name: str) -> dict[str, Any]:
    """Скопировать вложение в outputs/ — «вернуть файл» без генерации моделью."""
    from app.services.gpt_api import ensure_correct_extension

    d = _session_dir(session_id)
    if not d.is_dir() or not (d / "meta.json").is_file():
        raise FileNotFoundError(f"сессия не найдена: {session_id}")
    src = d / "attachments" / Path(name).name
    if not src.is_file():
        raise FileNotFoundError(f"нет вложения {name}")
    out_dir = d / "outputs"
    out_dir.mkdir(exist_ok=True)
    # Имя с корректным расширением (если исходник .bin, а внутри PNG)
    fixed_src = ensure_correct_extension(src)
    safe = _safe_filename(fixed_src.name)
    dest = out_dir / safe
    if dest.exists():
        stem, suf = Path(safe).stem, Path(safe).suffix
        i = 2
        while True:
            cand = out_dir / f"{stem}_out{i}{suf}"
            if not cand.exists():
                dest = cand
                break
            i += 1
    shutil.copy2(fixed_src, dest)
    dest = ensure_correct_extension(dest)
    meta = _read_json(d / "meta.json", {})
    meta["updated_at"] = _now()
    _write_json(d / "meta.json", meta)
    return _file_entry(dest)


def build_outputs_zip(session_id: str) -> Path:
    """Собрать zip всех outputs сессии (скачать всё, как из ChatGPT)."""
    import zipfile

    d = _session_dir(session_id)
    if not d.is_dir() or not (d / "meta.json").is_file():
        raise FileNotFoundError(f"сессия не найдена: {session_id}")
    out_dir = d / "outputs"
    files = [p for p in out_dir.iterdir() if p.is_file()] if out_dir.is_dir() else []
    if not files:
        raise FileNotFoundError("нет файлов в Результатах")
    zip_path = d / f"outputs_{session_id}.zip"
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for p in files:
            zf.write(p, arcname=p.name)
    return zip_path


def _append_message(session_id: str, role: str, content: str, **extra: Any) -> None:
    d = _session_dir(session_id)
    msgs = _read_json(d / "messages.json", [])
    if not isinstance(msgs, list):
        msgs = []
    entry = {
        "id": uuid.uuid4().hex[:10],
        "role": role,
        "content": content,
        "at": _now(),
        **extra,
    }
    msgs.append(entry)
    _write_json(d / "messages.json", msgs)
    meta = _read_json(d / "meta.json", {})
    meta["updated_at"] = _now()
    # автозаголовок из первого юзер-сообщения
    if role == "user" and len(msgs) <= 2:
        title = (content or "").strip().replace("\n", " ")[:48]
        if title:
            meta["title"] = title
    _write_json(d / "meta.json", meta)


def _promote_attachments(session_id: str, files: list[Path]) -> list[str]:
    """Скопировать исходные байты вложений в outputs/ (возврат без модели)."""
    names: list[str] = []
    for src in files:
        try:
            dest_info = copy_attachment_to_outputs(session_id, src.name)
            if dest_info["name"] not in names:
                names.append(dest_info["name"])
        except Exception as e:  # noqa: BLE001
            logger.warning("gpt_workspace: promote {}: {}", src.name, e)
    return names


# Типы, которые обещаем корректно отдать (остальное — мусор / не в уведомление).
_READY_SUFFIXES = frozenset(
    {
        ".txt",
        ".md",
        ".csv",
        ".tsv",
        ".docx",
        ".xlsx",
        ".xlsm",
        ".png",
        ".jpg",
        ".jpeg",
        ".webp",
        ".gif",
        ".bmp",
        ".svg",
        ".pdf",
        ".zip",
        ".mp3",
        ".wav",
        ".mp4",
        ".webm",
    }
)


def _verify_ready_files(out_dir: Path, names: list[str], *, min_size: int = 0) -> list[str]:
    """Только реально лежащие на диске файлы (без reply_*.txt и мусора)."""
    junk = {".bin", ".download", ".tmp", ".dat", ".octet-stream", ".html", ".htm"}
    seen: set[str] = set()
    ok: list[str] = []
    for raw in names:
        name = Path(raw).name
        if not name or name in seen:
            continue
        if re.match(r"^reply_\d", name, re.I):
            continue
        p = out_dir / name
        if not p.is_file() or p.stat().st_size < min_size:
            continue
        suf = p.suffix.lower()
        if suf in junk:
            continue
        # генерация: предпочитаем обещанные типы; исходники «верни» — любой не-junk
        if suf and suf not in _READY_SUFFIXES and not re.search(
            r"[a-zA-Z0-9]", suf[1:] or ""
        ):
            continue
        seen.add(name)
        ok.append(name)
    return ok


def _ready_files_notice(names: list[str]) -> str:
    """Уведомление в пузыре: полный список успешно созданных файлов."""
    if not names:
        return "Файлы не созданы — на диск ничего не записано."
    lines = ["Готовые файлы:"]
    for n in names:
        lines.append(f"• {n}")
    return "\n".join(lines)


def _studio_return_reply(returned: list[str]) -> str:
    if not returned:
        return (
            "Нет вложений в этой сессии. Прикрепи файл(ы) скрепкой, "
            "затем снова попроси вернуть."
        )
    return _ready_files_notice(returned)


async def ask(
    session_id: str,
    message: str,
    *,
    with_attachments: bool = True,
) -> dict[str, Any]:
    """Отправить сообщение в GPT API, сохранить ответ и файлы в outputs/.

    Память диалога: прошлые user/assistant реплики сессии уходят в API как history.

    Возврат исходников: Studio копирует attachments/ → outputs/ сама.
    Vision даёт модели только визуальные токены, не байты файла.
    """
    from app.services.gpt_api import normalize_history
    from app.services.gpt_client import get_gpt_client

    d = _session_dir(session_id)
    if not d.is_dir():
        raise FileNotFoundError(f"сессия не найдена: {session_id}")

    text = (message or "").strip()
    if not text:
        raise ValueError("пустое сообщение")

    files: list[Path] = []
    if with_attachments:
        att = d / "attachments"
        if att.is_dir():
            files = [p for p in sorted(att.iterdir()) if p.is_file()]

    # История ДО текущего сообщения (UI хранит всё; в API — только user/assistant)
    prior_raw = _read_json(d / "messages.json", [])
    history = normalize_history(prior_raw if isinstance(prior_raw, list) else [])

    meta = _read_json(d / "meta.json", {})
    meta["status"] = "running"
    meta["phase"] = "accepted"
    meta["phase_detail"] = "Запрос принят"
    meta["last_error"] = ""
    meta["updated_at"] = _now()
    _write_json(d / "meta.json", meta)

    _append_message(
        session_id,
        "user",
        text,
        attachment_names=[p.name for p in files],
    )

    out_dir = d / "outputs"
    out_dir.mkdir(exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    saved_files: list[str] = []
    last_doc = _last_assistant_document(
        prior_raw if isinstance(prior_raw, list) else []
    )
    intent = resolve_file_intent(
        text,
        has_attachments=bool(files),
        last_doc=last_doc,
        prior_raw=prior_raw if isinstance(prior_raw, list) else [],
    )
    want_pack = intent in {"pack_last", "pack_reply"}
    want_return = intent == "return_attachments"

    # 1) Упаковать уже готовый текст из чата
    if intent == "pack_last":
        delivered = _deliver_reply_as_file(
            out_dir=out_dir,
            reply=last_doc,
            user_text=text,
            attachments=files,
            ts=ts,
        )
        if delivered is not None:
            saved_files.append(delivered.name)
            ready = _verify_ready_files(out_dir, saved_files)
            reply = _ready_files_notice(ready)
            (out_dir / f"reply_{ts}.txt").write_text(reply, encoding="utf-8")
            _append_message(
                session_id,
                "assistant",
                reply,
                output_files=list(ready),
            )
            meta = _read_json(d / "meta.json", {})
            meta["status"] = "idle"
            meta["phase"] = "done"
            meta["phase_detail"] = "Готово (файл из ответа)"
            meta["updated_at"] = _now()
            _write_json(d / "meta.json", meta)
            logger.info(
                "gpt_workspace: session={} intent=pack_last file={}",
                session_id,
                delivered.name,
            )
            return get_session(session_id)

    # 2) Вернуть исходные вложения со скрепки
    if want_return:
        saved_files.extend(_promote_attachments(session_id, files))
        ready = _verify_ready_files(out_dir, saved_files)
        reply = _studio_return_reply(ready)
        (out_dir / f"reply_{ts}.txt").write_text(reply, encoding="utf-8")
        _append_message(
            session_id,
            "assistant",
            reply,
            output_files=list(ready),
            studio_returned=True,
        )
        meta = _read_json(d / "meta.json", {})
        meta["status"] = "idle"
        meta["phase"] = "done"
        meta["phase_detail"] = "Готово (возврат файлов)"
        meta["updated_at"] = _now()
        _write_json(d / "meta.json", meta)
        logger.info(
            "gpt_workspace: session={} intent=return_attachments files={}",
            session_id,
            saved_files,
        )
        return get_session(session_id)

    try:
        gpt = get_gpt_client()
        has_xlsx = any(p.suffix.lower() in {".xlsx", ".xlsm"} for p in files)
        ask_text = text

        meta = _read_json(d / "meta.json", {})
        meta["phase"] = "thinking"
        meta["phase_detail"] = "GPT думает / генерирует ответ (vision может занять минуты)…"
        meta["updated_at"] = _now()
        _write_json(d / "meta.json", meta)

        try:
            # Studio chat: 1 повтор максимум — иначе 5×180с = «зависание» на 15 мин.
            ask_timeout = 240.0 if files else 120.0
            reply = await gpt.ask_with_files(
                ask_text,
                files,
                timeout=min(float(settings.gpt_timeout_s or 600), ask_timeout),
                expect_file_download=has_xlsx,
                history=history,
                treat_txt_as_prompt=False,
                system=_WORKSPACE_SYSTEM,
                max_retries=1,
            )
        except Exception as e:  # noqa: BLE001
            # Модель иногда отдаёт пустой output на «пустой txt» — это валидный контент файла.
            err = str(e)
            if want_pack and "пустой output" in err.lower():
                logger.info(
                    "gpt_workspace: session={} empty GPT output → pack as empty file",
                    session_id,
                )
                reply = ""
            else:
                raise
        reply = (reply or "").strip()

        meta = _read_json(d / "meta.json", {})
        meta["phase"] = "saving"
        meta["phase_detail"] = "Сохраняю ответ и файлы…"
        meta["updated_at"] = _now()
        _write_json(d / "meta.json", meta)

        # Текст ответа на диск (для zip), в пузыре не дублируем reply_*.txt
        reply_path = out_dir / f"reply_{ts}.txt"
        reply_path.write_text(reply, encoding="utf-8")

        # 1) Картинки / URL из ответа модели
        try:
            from app.services.gpt_api import ensure_correct_extension, materialize_reply_assets

            assets = await materialize_reply_assets(
                reply,
                out_dir,
                prefix=f"gpt_{ts}",
                timeout=90.0,
            )
            for p in assets:
                p = ensure_correct_extension(p)
                if p.is_file() and p.name not in saved_files:
                    saved_files.append(p.name)
            for p in list(out_dir.iterdir()):
                if p.is_file() and p.suffix.lower() in {
                    ".bin",
                    ".dat",
                    ".download",
                    ".octet-stream",
                }:
                    from app.services.gpt_api import finalize_downloaded_file

                    fixed = finalize_downloaded_file(p)
                    if fixed.name not in saved_files and fixed.suffix.lower() not in {
                        ".bin",
                        ".download",
                    }:
                        saved_files.append(fixed.name)
        except Exception as e:  # noqa: BLE001
            logger.warning("gpt_workspace: materialize assets: {}", e)

        saved_files = _filter_placeholder_images(out_dir, saved_files)

        media_count = sum(
            1
            for n in saved_files
            if Path(n).suffix.lower() in {
                ".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp", ".svg",
            }
        )

        # Картинка из интернета: если GPT не дал рабочих URL — ищем сами.
        if _wants_image_file(text) and media_count == 0:
            q = _image_search_query(text)
            variants = _image_query_variants(text)
            try:
                from app.services.gpt_api import fetch_web_images

                web_imgs = await fetch_web_images(
                    q,
                    out_dir,
                    prefix=f"web_{ts}",
                    limit=3,
                    timeout=60.0,
                    query_variants=variants,
                )
                for p in web_imgs:
                    if p.is_file() and p.name not in saved_files:
                        saved_files.append(p.name)
                saved_files = _filter_placeholder_images(out_dir, saved_files)
                media_count = sum(
                    1
                    for n in saved_files
                    if Path(n).suffix.lower() in {
                        ".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp", ".svg",
                    }
                )
                logger.info(
                    "gpt_workspace: session={} web image search q={!r} variants={} got={}",
                    session_id,
                    q,
                    variants[:4],
                    media_count,
                )
            except Exception as e:  # noqa: BLE001
                logger.warning("gpt_workspace: web image search: {}", e)

        # 2) Excel — если просили (и/или в ответе есть # Лист:)
        if _wants_xlsx(text) or ("# Лист:" in reply) or ("# Лист：" in reply) or has_xlsx:
            try:
                xlsx_path = out_dir / f"table_{ts}.xlsx"
                if has_xlsx or ("# Лист:" in reply) or ("# Лист：" in reply):
                    try:
                        await gpt.download_attachment_from_last_reply(
                            xlsx_path,
                            timeout=120 if has_xlsx else 60,
                            fallback_text=reply,
                            allow_reply_text_fallback=True,
                        )
                    except Exception:  # noqa: BLE001
                        pass
                if not xlsx_path.exists() or xlsx_path.stat().st_size < 64:
                    _write_simple_xlsx(xlsx_path, reply)
                if xlsx_path.exists() and xlsx_path.stat().st_size >= 64:
                    if xlsx_path.name not in saved_files:
                        saved_files.append(xlsx_path.name)
                elif xlsx_path.exists():
                    try:
                        xlsx_path.unlink()
                    except OSError:
                        pass
            except Exception as e:  # noqa: BLE001
                logger.warning("gpt_workspace: xlsx deliver: {}", e)

        # 3) .txt/.docx — документные запросы; длинный договор в ответе — тоже файл
        delivered: Path | None = None
        pack_doc = (
            want_pack
            and _should_pack_text_document(text, reply, media_count=media_count)
        ) or (
            media_count == 0
            and _looks_like_document(reply)
            and not _is_meta_chat_question(text)
        )
        if (
            pack_doc
            and not any(p.suffix.lower() in {".xlsx", ".xlsm"} for p in files)
        ):
            try:
                delivered = _deliver_reply_as_file(
                    out_dir=out_dir,
                    reply=reply,
                    user_text=text,
                    attachments=files,
                    ts=ts,
                )
                if delivered is not None and delivered.name not in saved_files:
                    saved_files.append(delivered.name)
            except Exception as e:  # noqa: BLE001
                logger.warning("gpt_workspace: deliver reply file: {}", e)

        ready = _verify_ready_files(out_dir, saved_files)
        body = _strip_media_payloads(_strip_attachment_excuses(reply)).rstrip()
        body = re.sub(
            r"\n*—\n*(?:Studio положила|Studio вернула|Готовые файлы:)[\s\S]*$",
            "",
            body,
            flags=re.I,
        ).rstrip()
        # Длинный договор уже в файле — в пузыре не дублируем простыню
        if ready and _looks_like_document(body) and any(
            Path(n).suffix.lower() in {".txt", ".docx", ".md"} for n in ready
        ):
            body = ""
        if _wants_image_file(text) and media_count == 0:
            logger.warning(
                "gpt_workspace: session={} image ask but no real media "
                "raw_reply_len={} has_data_uri={}",
                session_id,
                len(reply or ""),
                bool(_DATA_URI_INLINE_RE.search(reply or "")),
            )
            if not body:
                body = (
                    "Не удалось найти картинку в интернете по запросу. "
                    "Уточни название (например: «Phantom Assassin Dota 2») "
                    "или пришли прямую ссылку на изображение."
                )
        if ready:
            notice = _ready_files_notice(ready)
            reply = f"{body}\n\n{notice}" if body else notice
        else:
            reply = body

        _append_message(
            session_id,
            "assistant",
            reply,
            output_files=list(ready),
        )
        meta = _read_json(d / "meta.json", {})
        meta["status"] = "idle"
        meta["phase"] = "done"
        meta["phase_detail"] = "Готово"
        meta["updated_at"] = _now()
        _write_json(d / "meta.json", meta)
        logger.info(
            "gpt_workspace: session={} intent={} reply_len={} files={} history={}",
            session_id,
            intent,
            len(reply),
            saved_files,
            len(history),
        )
        return get_session(session_id)
    except Exception as e:
        meta = _read_json(d / "meta.json", {})
        meta["status"] = "error"
        meta["phase"] = "error"
        meta["phase_detail"] = str(e)[:200]
        meta["last_error"] = str(e)[:500]
        meta["updated_at"] = _now()
        _write_json(d / "meta.json", meta)
        _append_message(session_id, "system", f"Ошибка: {e}")
        raise


def save_output_to_project(
    session_id: str,
    *,
    output_name: str,
    project_data_dir: Path,
    as_name: str | None = None,
) -> dict[str, Any]:
    """Скопировать output в папку проекта (сохранение «как в оригинале»)."""
    src = _session_dir(session_id) / "outputs" / Path(output_name).name
    if not src.is_file():
        raise FileNotFoundError(f"нет файла {output_name}")
    project_data_dir.mkdir(parents=True, exist_ok=True)
    dest_name = _safe_filename(as_name or src.name)
    dest = project_data_dir / dest_name
    # project.xlsx — особый случай: бэкап
    if dest_name.lower() == "project.xlsx" and dest.exists():
        bak = project_data_dir / "old" / f"project_before_gpt_{uuid.uuid4().hex[:8]}.xlsx"
        bak.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(dest, bak)
    shutil.copy2(src, dest)
    return {"saved_as": str(dest), "name": dest.name, "size": dest.stat().st_size}


def save_reply_as_voiceover(
    session_id: str,
    *,
    project_data_dir: Path,
    message_id: str | None = None,
) -> dict[str, Any]:
    """Сохранить текст ответа ассистента как voiceover.txt."""
    session = get_session(session_id)
    msgs = list(session.get("messages") or [])
    text = ""
    if message_id:
        for m in msgs:
            if m.get("id") == message_id and m.get("role") == "assistant":
                text = str(m.get("content") or "")
                break
    if not text:
        for m in reversed(msgs):
            if m.get("role") == "assistant":
                text = str(m.get("content") or "")
                break
    if not text.strip():
        raise ValueError("нет текста ответа для сохранения")
    project_data_dir.mkdir(parents=True, exist_ok=True)
    dest = project_data_dir / "voiceover.txt"
    dest.write_text(text.strip() + "\n", encoding="utf-8")
    return {"saved_as": str(dest), "name": dest.name, "chars": len(text.strip())}
