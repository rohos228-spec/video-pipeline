"""Чтение листа «Персонажи» из project.xlsx (v8-шаблон).

Формат листа:
    R1   ID персонажа       — c01, c02, c03, ... (по 1 столбцу на персонажа)
    R3   имя
    R4   внешность
    R5   одежда
    R6   характер
    R7   правила            — если содержит ID других персонажей, это
                              делает персонажа реф-вариацией тех

Каждый персонаж — это один столбец (B..N). Колонка A — подписи строк
(служебные).

Если в R7 встречаются ID других персонажей (например `c01`, `c02`) — все
они становятся `ref_ids`: их картинки будут использованы как референсы
при генерации этого персонажа (без вызова ChatGPT).
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

# Какие строки на листе «Персонажи» какому полю соответствуют.
# (1-based, как в openpyxl).
ROW_ID = 1
ROW_NAME = 3
ROW_LOOK = 4
ROW_CLOTHES = 5
ROW_CHAR = 6
ROW_RULES = 7

SHEET_PERSONS = "Персонажи"

# Шаблон допустимого ID персонажа в R1 (для отсеивания «мусорных» столбцов).
# Поддерживаем латиницу, кириллицу, цифры, `_`, `-`. Должен начинаться
# с буквы (любой). Без пробелов и спецсимволов — чтобы безопасно класть
# в имя файла (`<id>.png`) и в callback_data Telegram-инлайн-кнопок.
_ID_RE = re.compile(r"^[^\W\d_][\w-]*$", re.UNICODE)


@dataclass
class ExcelCharacter:
    """Одна запись из листа «Персонажи»."""

    id: str
    name: str = ""
    look: str = ""
    clothes: str = ""
    char: str = ""
    rules: str = ""
    # ID других персонажей, упомянутых в `rules` (в порядке появления).
    ref_ids: list[str] = field(default_factory=list)
    # Выбранный пользователем в TG промт (имя файла в prompts/04_hero_style/
    # или null если ещё не выбран). Для реф-вариаций промт всё равно
    # запрашивается, но при генерации не используется (выбор только для
    # консистентного UI).
    prompt_name: str | None = None

    @property
    def has_refs(self) -> bool:
        return bool(self.ref_ids)

    def changes_text(self) -> str:
        """Текст «изменений» для outsee-вариации: 4 поля с подписями,
        без «правил» (по требованию пользователя)."""
        parts: list[str] = []
        if self.name:
            parts.append(f"Имя: {self.name}")
        if self.look:
            parts.append(f"Внешность: {self.look}")
        if self.clothes:
            parts.append(f"Одежда: {self.clothes}")
        if self.char:
            parts.append(f"Характер: {self.char}")
        return "\n".join(parts)

    def brief_for_gpt(self) -> str:
        """Объединённый brief для GPT (не-реф персонажи): все поля
        с подписями, включая правила — это полное описание персонажа."""
        parts: list[str] = []
        if self.name:
            parts.append(f"Имя: {self.name}")
        if self.look:
            parts.append(f"Внешность: {self.look}")
        if self.clothes:
            parts.append(f"Одежда: {self.clothes}")
        if self.char:
            parts.append(f"Характер: {self.char}")
        if self.rules:
            parts.append(f"Правила: {self.rules}")
        return "\n".join(parts)

    def to_dict(self) -> dict:
        return {
            "id": self.id,
            "name": self.name,
            "look": self.look,
            "clothes": self.clothes,
            "char": self.char,
            "rules": self.rules,
            "ref_ids": list(self.ref_ids),
            "prompt_name": self.prompt_name,
        }

    @classmethod
    def from_dict(cls, d: dict) -> ExcelCharacter:
        return cls(
            id=str(d.get("id") or "").strip(),
            name=str(d.get("name") or "").strip(),
            look=str(d.get("look") or "").strip(),
            clothes=str(d.get("clothes") or "").strip(),
            char=str(d.get("char") or "").strip(),
            rules=str(d.get("rules") or "").strip(),
            ref_ids=[str(x).strip() for x in (d.get("ref_ids") or []) if str(x).strip()],
            prompt_name=(d.get("prompt_name") or None),
        )


def _cell_text(ws, row: int, col: int) -> str:
    """Безопасное извлечение текста из ячейки (строка, без переводов
    нач/конца)."""
    v = ws.cell(row=row, column=col).value
    if v is None:
        return ""
    return str(v).strip()


def parse_persons_sheet(xlsx_path: Path) -> list[ExcelCharacter]:
    """Парсит лист «Персонажи» project.xlsx и возвращает список персонажей
    (только тех, у кого заполнен ID в R1 и хотя бы одно из имя/внешность).

    Бросает `FileNotFoundError` если файла нет, `RuntimeError` если в
    книге нет листа «Персонажи».
    """
    from openpyxl import load_workbook

    if not xlsx_path.exists():
        raise FileNotFoundError(f"project.xlsx не найден: {xlsx_path}")

    wb = load_workbook(filename=str(xlsx_path), data_only=True)
    if SHEET_PERSONS not in wb.sheetnames:
        raise RuntimeError(
            f"в xlsx нет листа «{SHEET_PERSONS}» — это не v8-шаблон?"
        )

    ws = wb[SHEET_PERSONS]
    # Сначала собираем IDs из R1 (столбцы B.. = col 2..max_col).
    # Это даёт нам список ID, по которому потом будем матчить ref_ids в R7.
    max_col = max(ws.max_column or 0, 2)
    raw: list[tuple[int, str]] = []  # (column, id)
    for c in range(2, max_col + 1):
        cid = _cell_text(ws, ROW_ID, c)
        if cid and _ID_RE.match(cid):
            raw.append((c, cid))

    known_ids = {cid for _, cid in raw}
    out: list[ExcelCharacter] = []
    for col, cid in raw:
        name = _cell_text(ws, ROW_NAME, col)
        look = _cell_text(ws, ROW_LOOK, col)
        clothes = _cell_text(ws, ROW_CLOTHES, col)
        char = _cell_text(ws, ROW_CHAR, col)
        rules = _cell_text(ws, ROW_RULES, col)

        # Требуем хотя бы одно непустое описательное поле — иначе колонка
        # реально пустая, ID = «c01» это просто заглушка шаблона.
        if not any([name, look, clothes, char, rules]):
            continue

        ref_ids = _extract_refs(rules, known_ids, exclude=cid)
        out.append(
            ExcelCharacter(
                id=cid,
                name=name,
                look=look,
                clothes=clothes,
                char=char,
                rules=rules,
                ref_ids=ref_ids,
            )
        )
    return out


def _extract_refs(rules_text: str, known_ids: set[str], *, exclude: str) -> list[str]:
    """Находит в `rules_text` упоминания других ID персонажей.
    Возвращает их в порядке первого появления, без дубликатов.
    `exclude` — собственный ID персонажа (на себя ссылаться не считаем)."""
    if not rules_text or not known_ids:
        return []
    # Ищем «слова» из букв/цифр/подчёркивания/дефиса (UNICODE — кириллица
    # тоже допускается). Это покрывает форматы вида c01, p_02, hero-3,
    # пав1, нико-2 и т.п.
    found: list[str] = []
    for tok in re.findall(r"[^\W\d_][\w-]*", rules_text, flags=re.UNICODE):
        if tok == exclude:
            continue
        if tok in known_ids and tok not in found:
            found.append(tok)
    return found
