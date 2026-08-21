"""Запись project.xlsx из текстового ответа GPT API (без браузерного download).

Модель не отдаёт бинарный .xlsx — поэтому:
  1) если в ответе/скачанных файлах есть .xlsx — накладываем НЕПУСТЫЕ ячейки
     на существующий шаблон (full replace только если шаблона ещё нет);
  2) иначе TSV `# Лист:` + `@row=N` накладываем на шаблон БЕЗ wipe строк/merge;
  3) если TSV нет, но ответ длинный — пишем в B2 «Общий план» только
     для простых книг без листа кадров «план» (иначе книга не трогается).
"""

from __future__ import annotations

import re
import shutil
from pathlib import Path

from loguru import logger

_SHEET_HEADER_RE = re.compile(
    r"^\s*#\s*Лист\s*:\s*(.+?)\s*$",
    re.IGNORECASE | re.MULTILINE,
)
_FENCE_RE = re.compile(
    r"```(?:tsv|csv|text|xlsx)?\s*\n(.*?)```",
    re.IGNORECASE | re.DOTALL,
)
_ROW_MARK_RE = re.compile(r"^@row=(\d+)$", re.IGNORECASE)

# Совпадает с app.services.xlsx_v8_import.SHEET_GENERAL_V8 / plan_validation.
_GENERAL_PLAN_SHEET = "Общий план"
_MIN_PROSE_PLAN_CHARS = 200


_CONTINUE_LINE_RE = re.compile(r"^\s*CONTINUE_XLSX\b", re.IGNORECASE)
_CONTROL_JUNK_RE = re.compile(
    r"^\s*(CONTINUE_XLSX\b|#\s*КОНТРАКТ|===VP_|===BODY===|===END===)",
    re.IGNORECASE,
)
# CONTINUE / контракт в любой ячейке (в т.ч. `@row=1\tCONTINUE…\t1\t2`).
_CONTINUE_ANYWHERE_RE = re.compile(r"CONTINUE\s*[_ ]?\s*XLSX", re.IGNORECASE)
_PLAN_LABEL_ROWS_CHECK = 60


def _cell_is_control_junk(value: object) -> bool:
    """True если значение — служебный маркер, а не данные таблицы."""
    s = str(value or "").strip()
    if not s:
        return False
    if _CONTINUE_LINE_RE.search(s) or _CONTINUE_ANYWHERE_RE.search(s):
        return True
    if _CONTROL_JUNK_RE.search(s):
        return True
    if "CONTINUE_XLSX" in s.upper().replace(" ", ""):
        return True
    return False


def _row_has_control_junk(cells: list[str]) -> bool:
    return any(_cell_is_control_junk(c) for c in cells)


def extract_sheet_blocks(
    text: str,
    *,
    orphan_sheet: str = "Данные",
) -> dict[str, list[list[str]]]:
    """Разобрать `# Лист: Name` + TSV-строки → {sheet_name: rows}.

    Строки могут начинаться с `@row=N` (абсолютный номер Excel).
    Строки с табами без `# Лист:` попадают на ``orphan_sheet`` (по умолчанию
    «Данные»; при CONTINUE лучше передавать имя целевого листа).
    """
    if not (text or "").strip():
        return {}
    # Сначала пробуем fenced-блоки; если пусто — весь текст.
    chunks: list[str] = []
    for m in _FENCE_RE.finditer(text):
        body = (m.group(1) or "").strip()
        if body:
            chunks.append(body)
    if not chunks:
        chunks = [text]

    out: dict[str, list[list[str]]] = {}
    for chunk in chunks:
        current: str | None = None
        for line in chunk.splitlines():
            hm = _SHEET_HEADER_RE.match(line)
            if hm:
                current = hm.group(1).strip()
                out.setdefault(current, [])
                continue
            # CONTINUE_XLSX / служебные строки — НЕ ячейки (иначе A1=CONTINUE…).
            if _CONTINUE_LINE_RE.search(line) or _CONTROL_JUNK_RE.search(line):
                continue
            if _CONTINUE_ANYWHERE_RE.search(line):
                continue
            if current is None:
                # Без `# Лист:` — только явный TSV (табы). Запятые в русской
                # прозе (. «армия, право, дороги») нельзя считать CSV.
                if "\t" in line:
                    current = orphan_sheet or "Данные"
                    out.setdefault(current, [])
                else:
                    continue
            if not line.strip() or line.strip().startswith("…"):
                continue
            if "\t" in line:
                cells = line.split("\t")
            elif "|" in line and not line.strip().startswith("|--"):
                # markdown table row
                raw = [c.strip() for c in line.strip().strip("|").split("|")]
                if raw and all(re.fullmatch(r":?-{2,}:?", c or "") for c in raw):
                    continue
                cells = raw
            else:
                # Без таба внутри `# Лист:` — не данные (CONTINUE / проза / мусор).
                # Раньше уезжало в A1..A3 как sequential.
                continue
            # Любая ячейка со служебным CONTINUE/контрактом — вся строка мусор.
            if cells and _row_has_control_junk(cells):
                continue
            out[current].append(cells)
    # Убрать пустые листы
    return {k: v for k, v in out.items() if v}


def _parse_row_mark(row: list[str]) -> tuple[int | None, list[str]]:
    """Если первая ячейка `@row=N` — вернуть (N, остальные ячейки)."""
    if not row:
        return None, []
    first = str(row[0] or "").strip()
    m = _ROW_MARK_RE.match(first)
    if not m:
        return None, list(row)
    return int(m.group(1)), list(row[1:])


def _label_row_map(ws) -> dict[str, int]:
    """Подпись колонки A → номер строки (для legacy TSV без @row=)."""
    out: dict[str, int] = {}
    max_r = int(ws.max_row or 0)
    for r in range(1, max_r + 1):
        from app.services.xlsx_v8_import import _resolve_plan_cell

        cell = _resolve_plan_cell(ws, r, 1)
        if type(cell).__name__ == "MergedCell":
            continue
        lab = str(cell.value or "").strip()
        if lab:
            out.setdefault(lab.casefold(), r)
    return out


def _set_cell_value(ws, row: int, col: int, value: str) -> None:
    """Записать значение с учётом merge (как в чате GPT: структура листа не ломается)."""
    from app.services.xlsx_v8_import import _resolve_plan_cell

    cell = _resolve_plan_cell(ws, row, col)
    # MergedCell без top-left — пропускаем
    if type(cell).__name__ == "MergedCell":
        return
    cell.value = value


def assert_frame_plan_labels_intact(
    written_xlsx: Path,
    *,
    reference_xlsx: Path,
) -> None:
    """Fail-closed: подписи колонки A листа «план» не должны меняться / содержать CONTINUE.

    Если GPT-TSV всё же пролез — откатываем запись (вызывающий не делает move).
    """
    from openpyxl import load_workbook

    if not written_xlsx.exists() or not reference_xlsx.exists():
        return
    wb_w = load_workbook(filename=str(written_xlsx))
    wb_r = load_workbook(filename=str(reference_xlsx), read_only=True, data_only=True)
    try:
        names_w = {n.lower(): n for n in wb_w.sheetnames}
        names_r = {n.lower(): n for n in wb_r.sheetnames}
        if "план" not in names_w or "план" not in names_r:
            return
        ws_w = wb_w[names_w["план"]]
        ws_r = wb_r[names_r["план"]]
        # CONTINUE / служебное в любой ячейке «план» — брак.
        max_r = min(int(ws_w.max_row or 0), 120)
        max_c = min(int(ws_w.max_column or 0), 40)
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                val = ws_w.cell(r, c).value
                if _cell_is_control_junk(val):
                    raise ValueError(
                        f"xlsx_writeback: отказ — в «план»!{ws_w.cell(r, c).coordinate} "
                        f"служебный маркер {str(val)[:60]!r}"
                    )
        # Колонка A (подписи строк) должна совпасть с шаблоном до записи.
        for r in range(1, _PLAN_LABEL_ROWS_CHECK + 1):
            ref = str(ws_r.cell(r, 1).value or "").strip()
            got = str(ws_w.cell(r, 1).value or "").strip()
            if not ref:
                # Пустая A в шаблоне не должна получить мусор/сдвиг подписей.
                if got and _cell_is_control_junk(got):
                    raise ValueError(
                        f"xlsx_writeback: отказ — «план»!A{r}={got[:60]!r}"
                    )
                continue
            if got.casefold() != ref.casefold():
                raise ValueError(
                    f"xlsx_writeback: отказ — сдвиг подписей «план»!A{r}: "
                    f"было {ref!r}, стало {got!r}"
                )
    finally:
        wb_w.close()
        wb_r.close()


def _preferred_content_sheet(wb) -> str | None:
    """Целевой лист для orphan «Данные» на v8-книге."""
    names = list(wb.sheetnames)
    lower = {n.lower(): n for n in names}
    for key in ("план", "общий план", "общий план ролика"):
        if key in lower:
            return lower[key]
    return None


def _resolve_block_sheet_name(
    sheet_name: str,
    *,
    name_map: dict[str, str],
    preferred: str | None,
) -> str | None:
    """Имя листа в книге или None (пропустить). «Данные» → preferred на strict layout."""
    key = sheet_name.lower()
    if key in name_map:
        return name_map[key]
    if key == "данные" and preferred is not None:
        logger.info(
            "xlsx_writeback: remap sheet {!r} → {!r}",
            sheet_name,
            preferred,
        )
        return preferred
    return None


def apply_sheet_blocks_to_xlsx(
    src_xlsx: Path,
    blocks: dict[str, list[list[str]]],
    dest_xlsx: Path,
) -> Path:
    """Наложить TSV на существующий workbook БЕЗ wipe строк/merge.

    Как скачанный .xlsx из чата: шаблон (merge, подписи, другие листы) сохраняем,
    пишем только пришедшие значения ячеек.

    Абсолютные строки: префикс `@row=N` в первой колонке TSV.
    Без префикса: sequential (legacy); для листа «план» / «Общий план»
    пробуем сопоставить подпись A с шаблоном, чтобы не сдвинуть R45/R49.
    """
    from openpyxl import Workbook, load_workbook

    if not blocks:
        raise ValueError("нет блоков листов для записи")

    if src_xlsx.exists() and src_xlsx.stat().st_size > 0:
        wb = load_workbook(filename=str(src_xlsx))
    else:
        wb = Workbook()
        if wb.sheetnames == ["Sheet"] and "Данные" in blocks:
            std = wb.active
            if std is not None:
                wb.remove(std)

    name_map = {n.lower(): n for n in wb.sheetnames}
    # Полный project.xlsx (v8) — нельзя плодить лишние листы (ломают validate).
    strict_layout = (
        _GENERAL_PLAN_SHEET in wb.sheetnames or "план" in name_map
    )
    preferred = _preferred_content_sheet(wb) if strict_layout else None

    for sheet_name, rows in blocks.items():
        key = sheet_name.lower()
        resolved = _resolve_block_sheet_name(
            sheet_name, name_map=name_map, preferred=preferred
        )
        if resolved is not None:
            ws = wb[resolved]
            key = resolved.lower()
        elif strict_layout:
            logger.warning(
                "xlsx_writeback: skip unknown sheet {!r} (сохраняем layout)",
                sheet_name,
            )
            continue
        else:
            ws = wb.create_sheet(title=sheet_name[:31] or "Данные")
            name_map[key] = ws.title

        has_row_marks = any(_parse_row_mark(r)[0] is not None for r in rows)
        label_map: dict[str, int] = {}
        # Fail-closed только для листа кадров «план» (R45/R48/R49).
        # «Общий план» часто пишут без @row= (эпизоды / prose) — sequential OK
        # ТОЛЬКО если во всём блоке нет ни одного @row= (иначе мусор/CONTINUE
        # уезжал в A1..A3 sequential и ломал шаблон).
        frame_plan = key == "план"
        plan_like = key in {"план", "общий план", "общий план ролика"}
        use_label_fallback = not has_row_marks and frame_plan
        # «план»: колонку A (подписи R*) никогда не трогаем — только B+.
        # «Общий план»: бережём A если уже заполнена.
        protect_label_col = plan_like
        freeze_label_col = frame_plan
        if use_label_fallback:
            label_map = _label_row_map(ws)
            if not label_map:
                raise ValueError(
                    f"xlsx_writeback: лист «{ws.title}» без @row= и без подписей "
                    "в колонке A — отказ (риск сдвига R45/R48/R49)"
                )
        elif plan_like and not has_row_marks and strict_layout:
            # v8 + «Общий план» без @row=: только по подписи A, не sequential 1..N
            label_map = _label_row_map(ws)

        sequential_idx = 0
        for row in rows:
            abs_row, cells = _parse_row_mark(row)
            if abs_row is None:
                # В блоке уже есть @row= — строки без маркера = мусор/CONTINUE-хвост.
                if has_row_marks and plan_like:
                    logger.warning(
                        "xlsx_writeback: skip unmarked row on «{}» ({!r})",
                        ws.title,
                        (cells[0] if cells else "")[:80],
                    )
                    continue
                if label_map and cells:
                    lab = str(cells[0] or "").strip().casefold()
                    if lab and lab in label_map:
                        abs_row = label_map[lab]
                    elif plan_like and strict_layout:
                        logger.warning(
                            "xlsx_writeback: skip unknown label on «{}» ({!r})",
                            ws.title,
                            lab,
                        )
                        continue
                    else:
                        raise ValueError(
                            f"xlsx_writeback: лист «{ws.title}» строка без @row= "
                            f"и без известной подписи {lab!r} — отказ"
                        )
                elif frame_plan and not has_row_marks:
                    raise ValueError(
                        f"xlsx_writeback: лист «{ws.title}» без @row= — отказ"
                    )
                elif plan_like and strict_layout:
                    logger.warning(
                        "xlsx_writeback: skip sequential junk on v8 «{}»",
                        ws.title,
                    )
                    continue
                else:
                    sequential_idx += 1
                    abs_row = sequential_idx
            if abs_row < 1:
                continue
            for c_idx, val in enumerate(cells, start=1):
                if val is None:
                    continue
                s = str(val)
                # Пустая ячейка в TSV — не затираем подпись/merge шаблона
                if not s.strip():
                    continue
                if _cell_is_control_junk(s):
                    continue
                # Лист кадров: A1..An — священные подписи шаблона.
                if freeze_label_col and c_idx == 1:
                    continue
                # «Общий план»: не затираем уже существующую подпись в A.
                if protect_label_col and c_idx == 1:
                    from app.services.xlsx_v8_import import _resolve_plan_cell

                    existing = _resolve_plan_cell(ws, abs_row, 1)
                    if type(existing).__name__ != "MergedCell":
                        old = str(existing.value or "").strip()
                        if old and old.casefold() != s.strip().casefold():
                            continue
                _set_cell_value(ws, abs_row, c_idx, s)

    dest_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(dest_xlsx))
    wb.close()
    return dest_xlsx


def _is_v8_project_workbook(path: Path) -> bool:
    if not path.exists() or path.stat().st_size < 64:
        return False
    try:
        from openpyxl import load_workbook

        wb = load_workbook(filename=str(path), read_only=True)
        try:
            names = {n.lower() for n in wb.sheetnames}
            return "план" in names or "общий план" in names
        finally:
            wb.close()
    except Exception:  # noqa: BLE001
        return False


def _workbook_has_frame_plan_sheet(path: Path) -> bool:
    """True если в книге есть лист кадров «план» (полный project.xlsx)."""
    if not path.exists() or path.stat().st_size < 64:
        return False
    try:
        from openpyxl import load_workbook

        wb = load_workbook(filename=str(path), read_only=True)
        try:
            return any(n.lower() == "план" for n in wb.sheetnames)
        finally:
            wb.close()
    except Exception:  # noqa: BLE001
        return False


def merge_xlsx_nonempty_overlay(
    project_xlsx: Path,
    gpt_xlsx: Path,
    dest_xlsx: Path | None = None,
) -> Path:
    """Скопировать непустые ячейки из gpt_xlsx поверх project_xlsx.

    Пустые/отсутствующие в GPT ячейки НЕ затирают шаблон (enrich R2–44 и т.п.).
    Листы, которых нет в шаблоне, на v8-книге пропускаем.
    """
    from openpyxl import load_workbook

    dest = dest_xlsx or project_xlsx
    wb_proj = load_workbook(filename=str(project_xlsx))
    wb_gpt = load_workbook(filename=str(gpt_xlsx), data_only=True)
    try:
        name_map = {n.lower(): n for n in wb_proj.sheetnames}
        strict = "план" in name_map or "общий план" in name_map
        written = 0
        for gpt_name in wb_gpt.sheetnames:
            key = gpt_name.lower()
            if key not in name_map:
                if strict:
                    logger.warning(
                        "xlsx_overlay: skip unknown sheet {!r}", gpt_name
                    )
                    continue
                ws_p = wb_proj.create_sheet(title=gpt_name[:31])
                name_map[key] = ws_p.title
            else:
                ws_p = wb_proj[name_map[key]]
            ws_g = wb_gpt[gpt_name]
            max_r = int(ws_g.max_row or 0)
            # Раньше 80: обрезка широких строк «план» (100–150 кадров).
            max_c = min(int(ws_g.max_column or 0), 256)
            for r in range(1, max_r + 1):
                for c in range(1, max_c + 1):
                    val = ws_g.cell(row=r, column=c).value
                    if val is None:
                        continue
                    s = str(val)
                    if not s.strip():
                        continue
                    if _cell_is_control_junk(s):
                        continue
                    # Не трогаем подписи строк на листе кадров.
                    if key == "план" and c == 1:
                        continue
                    _set_cell_value(ws_p, r, c, s)
                    written += 1
        dest.parent.mkdir(parents=True, exist_ok=True)
        wb_proj.save(str(dest))
        logger.info(
            "xlsx_overlay: {} ← {} cells={} → {}",
            project_xlsx.name,
            gpt_xlsx.name,
            written,
            dest.name,
        )
    finally:
        wb_gpt.close()
        wb_proj.close()
    return dest


def _strip_reply_noise(text: str) -> str:
    t = (text or "").strip()
    if not t:
        return ""
    t = re.sub(r"```[\w]*\n?", "", t)
    t = t.replace("```", "").strip()
    return t


def write_prose_into_general_plan(
    src_xlsx: Path,
    dest_xlsx: Path,
    prose: str,
) -> Path:
    """Записать длинный текст плана в value-ячейку шаблона «Общий план».

    Не добавляем новые строки в конец — пишем в B2 (типичный content merge),
    подписи колонки A и merge ranges не трогаем.
    """
    from openpyxl import load_workbook
    from app.services.xlsx_v8_import import _resolve_plan_cell

    cleaned = _strip_reply_noise(prose)
    if len(cleaned) < _MIN_PROSE_PLAN_CHARS:
        raise ValueError(
            f"проза слишком короткая для плана ({len(cleaned)} < {_MIN_PROSE_PLAN_CHARS})"
        )
    if not src_xlsx.exists():
        raise FileNotFoundError(str(src_xlsx))

    wb = load_workbook(filename=str(src_xlsx))
    try:
        if _GENERAL_PLAN_SHEET not in wb.sheetnames:
            raise ValueError(f"нет листа «{_GENERAL_PLAN_SHEET}»")
        ws = wb[_GENERAL_PLAN_SHEET]
        # Пишем пару A2/B2: импортёр _read_general_plan берёт только a+b.
        # Подписи A1 и merge не ломаем; новые строки в конец не добавляем.
        label_cell = _resolve_plan_cell(ws, 2, 1)
        value_cell = _resolve_plan_cell(ws, 2, 2)
        if type(value_cell).__name__ == "MergedCell":
            value_cell = _resolve_plan_cell(ws, 3, 2)
        if type(value_cell).__name__ == "MergedCell":
            raise ValueError("не удалось записать в ячейки «Общий план» (merge)")
        if type(label_cell).__name__ != "MergedCell":
            if not (label_cell.value and str(label_cell.value).strip()):
                label_cell.value = "План (GPT)"
        value_cell.value = cleaned[:32000]
        dest_xlsx.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(dest_xlsx))
    finally:
        wb.close()
    return dest_xlsx


def writeback_project_xlsx(
    *,
    project_xlsx: Path,
    reply_text: str,
    downloaded_paths: list[Path] | None = None,
) -> Path | None:
    """Запрещено: Excel меняется только через excel_io Export/Import.

    Раньше обновлял project.xlsx из скачанного .xlsx или TSV.
    """
    raise RuntimeError(
        "Excel writeback отключён: правки только через apply-ops (БД) "
        "или кнопки Import/Export (excel_io). "
        f"Игнорирую writeback для {project_xlsx}"
    )
    # unreachable — legacy body removed from runtime path
    downloaded_paths = list(downloaded_paths or [])
    reply = reply_text or ""
    del reply, downloaded_paths
    return None


def _writeback_project_xlsx_legacy_disabled(
    *,
    project_xlsx: Path,
    reply_text: str,
    downloaded_paths: list[Path] | None = None,
) -> Path | None:
    """Legacy implementation kept for reference/tests that patch this — unused."""
    downloaded_paths = list(downloaded_paths or [])
    reply = reply_text or ""
    backed_up = False

    def _ensure_pre_write_backup() -> None:
        nonlocal backed_up
        if backed_up or not project_xlsx.exists() or project_xlsx.stat().st_size < 64:
            return
        try:
            from app.services.xlsx_versioning import backup_to_old

            snap = backup_to_old(project_xlsx)
            if snap is not None:
                logger.info(
                    "xlsx_writeback: pre-write backup → {}",
                    snap.name,
                )
        except Exception as e:  # noqa: BLE001
            logger.warning("xlsx_writeback: pre-write backup failed: {}", e)
        backed_up = True

    # Отчёт проверки (JSON/TXT) нельзя заливать в «Общий план» через prose/TSV.
    try:
        from app.services.check_analysis import looks_like_check_payload
    except Exception:  # noqa: BLE001
        looks_like_check_payload = None  # type: ignore[assignment]

    if callable(looks_like_check_payload) and looks_like_check_payload(reply):
        has_xlsx = any(
            p.suffix.lower() in {".xlsx", ".xlsm", ".xls"}
            and p.exists()
            and p.stat().st_size > 64
            for p in downloaded_paths
        )
        logger.warning(
            "xlsx_writeback: skip text — reply looks like check report "
            "(binary_xlsx={})",
            has_xlsx,
        )
        if not has_xlsx:
            return None
        reply = ""

    for p in downloaded_paths:
        if p.suffix.lower() in {".xlsx", ".xlsm", ".xls"} and p.exists() and p.stat().st_size > 64:
            try:
                magic = p.read_bytes()[:2]
            except OSError:
                magic = b""
            if magic != b"PK":
                # HTML/ошибка авторизации с именем .xlsx — дальше пробуем TSV.
                logger.warning(
                    "xlsx_writeback: skip не-zip download {} (magic={})",
                    p.name,
                    magic.hex() if magic else "?",
                )
                continue
            project_xlsx.parent.mkdir(parents=True, exist_ok=True)
            if p.resolve() == project_xlsx.resolve():
                return project_xlsx
            _ensure_pre_write_backup()
            if _is_v8_project_workbook(project_xlsx):
                # Не подменяем весь файл — GPT часто отдаёт «урезанную» книгу
                # и затирает enrich / чужие строки (как старый баг img_pr).
                try:
                    merge_xlsx_nonempty_overlay(project_xlsx, p, project_xlsx)
                    return project_xlsx
                except Exception as e:  # noqa: BLE001
                    logger.warning(
                        "xlsx_writeback: overlay failed ({}), fallback copy: {}",
                        p.name,
                        e,
                    )
            shutil.copy2(p, project_xlsx)
            logger.info(
                "xlsx_writeback: скопирован {} → {} ({} байт)",
                p.name,
                project_xlsx.name,
                project_xlsx.stat().st_size,
            )
            return project_xlsx

    blocks = extract_sheet_blocks(reply)
    if blocks:
        # Диагностика: первые строки TSV (видно CONTINUE/@row= в логе tutor).
        for sname, rows in list(blocks.items())[:3]:
            preview = [" | ".join(r[:6]) for r in rows[:6]]
            logger.info(
                "xlsx_writeback: TSV preview «{}» rows={} :: {}",
                sname,
                len(rows),
                " || ".join(preview)[:500],
            )
        tmp = project_xlsx.with_suffix(".writeback.tmp.xlsx")
        try:
            _ensure_pre_write_backup()
            apply_sheet_blocks_to_xlsx(project_xlsx, blocks, tmp)
            assert_frame_plan_labels_intact(tmp, reference_xlsx=project_xlsx)
            shutil.move(str(tmp), str(project_xlsx))
            logger.info(
                "xlsx_writeback: TSV→{} листов={}, size={}",
                project_xlsx.name,
                list(blocks.keys()),
                project_xlsx.stat().st_size,
            )
            return project_xlsx
        except Exception as e:  # noqa: BLE001
            logger.warning("xlsx_writeback failed: {}", e)
            if tmp.exists():
                tmp.unlink(missing_ok=True)
            # Порча подписей «план» — не fallback в prose, а жёсткий отказ.
            if "отказ" in str(e).lower() or "сдвиг подписей" in str(e).lower():
                raise
            # fall through to prose fallback

    # Fallback: GPT отдал прозу без `# Лист:` / TSV.
    # На полном project.xlsx (есть лист кадров «план») это ЛОМАЕТ книгу —
    # заливает B2 «Общий план» болтовнёй вместо `@row=` TSV. Разрешено
    # только для простых книг без листа «план».
    prose = _strip_reply_noise(reply)
    has_frame_sheet = _workbook_has_frame_plan_sheet(project_xlsx)
    if (
        len(prose) >= _MIN_PROSE_PLAN_CHARS
        and project_xlsx.exists()
        and not has_frame_sheet
    ):
        tmp = project_xlsx.with_suffix(".writeback.prose.tmp.xlsx")
        try:
            _ensure_pre_write_backup()
            write_prose_into_general_plan(project_xlsx, tmp, prose)
            shutil.move(str(tmp), str(project_xlsx))
            logger.info(
                "xlsx_writeback: prose→«{}» ({} симв) → {}",
                _GENERAL_PLAN_SHEET,
                len(prose),
                project_xlsx.name,
            )
            return project_xlsx
        except Exception as e:  # noqa: BLE001
            logger.warning("xlsx_writeback prose fallback failed: {}", e)
            if tmp.exists():
                tmp.unlink(missing_ok=True)
    elif len(prose) >= _MIN_PROSE_PLAN_CHARS and has_frame_sheet:
        logger.warning(
            "xlsx_writeback: отказ prose→«Общий план» на книге с листом «план» "
            "({} симв без `# Лист:`/`@row=`) — книга не тронута",
            len(prose),
        )

    logger.debug(
        "xlsx_writeback: в ответе нет TSV/листов и prose-fallback не сработал"
    )
    return None


WRITEBACK_HINT = (
    "Верни обновлённый Excel в текстовом формате: для каждого листа строка "
    "`# Лист: <имя>` и далее строки TSV (ячейки через табуляцию). "
    "Каждая строка ОБЯЗАТЕЛЬНО начинается с `@row=<номер Excel>` как во входе "
    "(пример: `@row=49\\tзакадровый текст\\t...`) — не уплотняй пустые ряды "
    "и не сдвигай номера строк. "
    "Для шага план обязателен лист «Общий план» (`# Лист: Общий план`). "
    "Сохрани имена листов как во входе. Либо приложи прямую ссылку на .xlsx.\n"
    "ПОЛНОЕ ЗАПОЛНЕНИЕ (критично): верни ВСЕ целевые строки листа из входа "
    "(`@row=…`), не сокращай до «примера» / первых 10–20%. Пустые обязательные "
    "ячейки — заполни. Если ответ не вмещается в один раз — в конце напиши "
    "`CONTINUE_XLSX: <имя_листа> @row=<следующий>` и жди продолжения; "
    "не обрывай молча на середине."
)

# Сколько дозапросов «продолжи TSV», если модель отдала кусок книги.
MAX_WRITEBACK_CONTINUES = 3


def count_sheet_rows_in_tsv(text: str) -> dict[str, int]:
    """Сколько TSV-строк на лист в ответе (по extract_sheet_blocks)."""
    blocks = extract_sheet_blocks(text or "")
    return {name: len(rows) for name, rows in blocks.items() if rows}


def last_row_mark_in_tsv(text: str, sheet: str) -> int | None:
    """Максимальный @row=N на листе в TSV-ответе."""
    blocks = extract_sheet_blocks(text or "")
    rows = blocks.get(sheet) or []
    last: int | None = None
    for cells in rows:
        if not cells:
            continue
        n, _rest = _parse_row_mark(cells)
        if n is None:
            continue
        last = n if last is None else max(last, n)
    return last


def template_nonempty_row_counts(xlsx_path: Path) -> dict[str, int]:
    """Непустые строки по листам шаблона (для оценки полноты writeback)."""
    try:
        from openpyxl import load_workbook
    except Exception:  # noqa: BLE001
        return {}
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        return {}
    out: dict[str, int] = {}
    try:
        for ws in wb.worksheets:
            n = 0
            for row in ws.iter_rows(values_only=True):
                if any(c is not None and str(c).strip() for c in row):
                    n += 1
            out[ws.title] = n
    finally:
        import contextlib

        with contextlib.suppress(Exception):
            wb.close()
    return out


def writeback_looks_incomplete(
    *,
    reply_text: str,
    template_xlsx: Path | None,
    min_ratio: float = 0.45,
    min_template_rows: int = 20,
) -> tuple[bool, str, str | None, int | None]:
    """(incomplete, reason, sheet, next_row)."""
    reply = reply_text or ""
    m = re.search(
        r"CONTINUE_XLSX\s*:\s*(.+?)\s*@row\s*=\s*(\d+)",
        reply,
        re.IGNORECASE,
    )
    template_names: list[str] = []
    if template_xlsx is not None and template_xlsx.is_file():
        try:
            from openpyxl import load_workbook

            twb = load_workbook(template_xlsx, read_only=True)
            try:
                template_names = list(twb.sheetnames)
            finally:
                twb.close()
        except Exception:  # noqa: BLE001
            template_names = []
    preferred = None
    if template_names:
        lower = {n.lower(): n for n in template_names}
        for key in ("план", "общий план", "общий план ролика"):
            if key in lower:
                preferred = lower[key]
                break

    if m:
        sheet_raw = m.group(1).strip()
        sheet = sheet_raw
        if sheet_raw.lower() == "данные" and preferred:
            sheet = preferred
        nxt = int(m.group(2))
        # Модель часто пишет CONTINUE_XLSX «по привычке», даже когда строк уже
        # достаточно / следующий @row за пределами шаблона — не считаем incomplete.
        got = count_sheet_rows_in_tsv(reply)
        want = (
            template_nonempty_row_counts(template_xlsx)
            if template_xlsx is not None and template_xlsx.is_file()
            else {}
        )
        n_got = got.get(sheet) or 0
        if n_got == 0:
            for gn, gv in got.items():
                if gn.lower() == sheet.lower() or (
                    gn.lower() == "данные" and preferred and sheet == preferred
                ):
                    n_got += gv
        n_want = want.get(sheet)
        if n_want is None:
            for tn, tv in want.items():
                if tn.lower() == sheet.lower():
                    n_want = tv
                    break
        max_template_row = n_want  # approx: nonempty count ≈ useful span
        if n_want and n_got >= max(8, int(n_want * min_ratio)):
            logger.info(
                "xlsx_writeback: ignore CONTINUE_XLSX (coverage {}/{} rows on {!r})",
                n_got,
                n_want,
                sheet,
            )
        elif max_template_row and nxt > max_template_row + 5:
            logger.info(
                "xlsx_writeback: ignore CONTINUE_XLSX (@row={} past template ~{})",
                nxt,
                max_template_row,
            )
        else:
            return True, "model marked CONTINUE_XLSX", sheet, nxt
    got = count_sheet_rows_in_tsv(reply)
    if not got:
        return False, "no tsv blocks", None, None
    if template_xlsx is None or not template_xlsx.is_file():
        return False, "no template", None, None
    want = template_nonempty_row_counts(template_xlsx)
    if not want:
        return False, "empty template", None, None
    # Сравниваем только листы, которые модель уже трогала.
    for name, n_got in got.items():
        resolved = name
        n_want = want.get(name)
        if n_want is None and name.lower() == "данные" and preferred:
            resolved = preferred
            n_want = want.get(preferred)
            n_got = got.get(preferred, 0) + n_got
        if n_want is None:
            for tn, tv in want.items():
                if tn.lower() == name.lower():
                    n_want = tv
                    resolved = tn
                    break
        if n_want is None:
            continue
        if n_want < min_template_rows:
            continue
        if n_got < max(8, int(n_want * min_ratio)):
            last = last_row_mark_in_tsv(reply, name)
            if last is None:
                last = last_row_mark_in_tsv(reply, resolved)
            nxt = (last + 1) if last else 2
            return (
                True,
                f"sheet {resolved!r}: writeback {n_got}/{n_want} rows "
                f"(<{min_ratio:.0%})",
                resolved,
                nxt,
            )
    return False, "ok", None, None


def build_writeback_continue_prompt(
    *,
    sheet_hint: str,
    next_row: int,
    previous_reply: str,
) -> str:
    """Промт дозапроса: продолжить TSV с указанного @row."""
    prev_tail = (previous_reply or "")[-2500:]
    return (
        "Предыдущий ответ обрезал Excel — заполнение НЕПОЛНОЕ.\n"
        f"Продолжи лист «{sheet_hint}» начиная с `@row={next_row}` "
        "(и дальше по шаблону). Верни ТОЛЬКО:\n"
        f"`# Лист: {sheet_hint}`\n"
        "и строки TSV с `@row=…` — без болтовни, без повтора уже отданных строк.\n"
        "Если снова не влезет — в конце `CONTINUE_XLSX: "
        f"{sheet_hint} @row=<следующий>`.\n\n"
        f"Хвост прошлого ответа:\n{prev_tail}"
    )


def merge_writeback_texts(
    *parts: str,
    default_sheet: str | None = None,
) -> str:
    """Склеить несколько TSV-ответов: блоки по листам.

    При дубле `@row=N` побеждает ПОСЛЕДНИЙ кусок (CONTINUE часто чинит
    кривые ранние строки). Строки без `# Лист:` наследуют лист предыдущего
    куска / ``default_sheet``, а не выдуманный «Данные».
    """
    merged: dict[str, list[list[str]]] = {}
    row_index: dict[str, dict[int, int]] = {}
    orphan = default_sheet or "Данные"
    for part in parts:
        blocks = extract_sheet_blocks(part or "", orphan_sheet=orphan)
        if blocks:
            # Следующий кусок без заголовка — на последний лист этого куска.
            orphan = next(reversed(blocks))
        for name, rows in blocks.items():
            bucket = merged.setdefault(name, [])
            idx_map = row_index.setdefault(name, {})
            for cells in rows:
                n, rest = _parse_row_mark(cells)
                if n is not None:
                    payload = [f"@row={n}", *rest]
                    if n in idx_map:
                        bucket[idx_map[n]] = payload
                    else:
                        idx_map[n] = len(bucket)
                        bucket.append(payload)
                else:
                    bucket.append(list(cells))
    chunks: list[str] = []
    for name, rows in merged.items():
        chunks.append(f"# Лист: {name}")
        for cells in rows:
            chunks.append("\t".join(cells))
        chunks.append("")
    return "\n".join(chunks).strip()


async def maybe_continue_partial_xlsx_reply(
    *,
    reply_text: str,
    template_xlsx: Path | None,
    input_paths: list[Path],
    accompanying: str = "",
    log_label: str = "xlsx",
    raise_if_still_incomplete: bool = False,
) -> str:
    """Если TSV-ответ явно кусок книги — дозапросить продолжение и склеить.

    Если после MAX_WRITEBACK_CONTINUES всё ещё CONTINUE/неполно и
    ``raise_if_still_incomplete`` — RuntimeError (книгу не писать).
    """
    from app.services.gpt_api import chat

    parts = [reply_text or ""]
    last_sheet: str | None = None
    for cont_i in range(MAX_WRITEBACK_CONTINUES):
        merged_so_far = merge_writeback_texts(
            *parts, default_sheet=last_sheet
        )
        incomplete, reason, sheet, nxt = writeback_looks_incomplete(
            reply_text=merged_so_far,
            template_xlsx=template_xlsx,
        )
        if not incomplete or not sheet or nxt is None:
            break
        last_sheet = sheet
        logger.warning(
            "{}: partial xlsx writeback ({}) — continue {}/{}",
            log_label,
            reason,
            cont_i + 1,
            MAX_WRITEBACK_CONTINUES,
        )
        cont = await chat(
            prompt=build_writeback_continue_prompt(
                sheet_hint=sheet,
                next_row=nxt,
                previous_reply=parts[-1],
            ),
            accompanying=accompanying,
            input_paths=list(input_paths),
        )
        parts.append(cont.text or "")
    merged = (
        parts[0]
        if len(parts) == 1
        else merge_writeback_texts(*parts, default_sheet=last_sheet)
    )
    if raise_if_still_incomplete:
        still, reason, sheet, nxt = writeback_looks_incomplete(
            reply_text=merged,
            template_xlsx=template_xlsx,
        )
        if still:
            raise RuntimeError(
                f"xlsx writeback всё ещё неполный после "
                f"{MAX_WRITEBACK_CONTINUES} continue ({reason}; "
                f"sheet={sheet!r} @row={nxt}) — project.xlsx не изменён"
            )
    return merged
