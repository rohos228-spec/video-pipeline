"""Лист «план» v8: закадровый текст (R49), номера кадров (R1), таймкоды Whisper."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from loguru import logger
from openpyxl import load_workbook

from app.services.xlsx_v8_import import (
    ROW_DURATION_V8,
    ROW_VOICEOVER_V8,
    SHEET_PLAN_V8,
)

# Строка 1 — номера кадров в шапке колонок (1, 2, 3…).
ROW_FRAME_NUMBER_V8 = 1
# Строка 2 — «id scene»; если в ячейке число — используем как номер кадра.
ROW_FRAME_NUMBER_ALT_V8 = 2
# Пустая строка под таймкод конца последнего слова (сек).
ROW_VOICE_END_V8 = 51
ROW_VOICE_END_LABEL = "таймкод конца (сек)"


@dataclass(frozen=True)
class PlanColumn:
    """Одна колонка кадра на листе «план»."""

    column: int
    frame_number: int
    voiceover_text: str


def _cell_text(ws, row: int, col: int) -> str | None:
    v = ws.cell(row=row, column=col).value
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    return " ".join(s.split())


def _parse_frame_number(ws, col: int, fallback: int) -> int:
    raw = _cell_text(ws, ROW_FRAME_NUMBER_V8, col)
    if raw is not None:
        try:
            return int(float(raw))
        except (TypeError, ValueError):
            pass
    return fallback


def read_plan_columns(xlsx_path: Path) -> list[PlanColumn]:
    """Читает блоки закадрового текста (R49) и номера кадров из R1/R2."""
    if not xlsx_path.exists():
        raise FileNotFoundError(xlsx_path)
    wb = load_workbook(filename=str(xlsx_path), data_only=True)
    if SHEET_PLAN_V8 not in wb.sheetnames:
        raise ValueError(f"лист «{SHEET_PLAN_V8}» не найден в {xlsx_path}")
    ws = wb[SHEET_PLAN_V8]
    out: list[PlanColumn] = []
    seq = 0
    for col in range(3, ws.max_column + 1):
        voice = _cell_text(ws, ROW_VOICEOVER_V8, col)
        if voice is None:
            continue
        seq += 1
        out.append(
            PlanColumn(
                column=col,
                frame_number=_parse_frame_number(ws, col, seq),
                voiceover_text=voice,
            )
        )
    return out


def write_whisper_timecodes(
    xlsx_path: Path,
    timings: list[tuple[int, float, float, float]],
    *,
    column_by_frame: dict[int, int] | None = None,
) -> None:
    """Пишет end_ts в R51 и длительность в R50.

    timings: [(frame_number, start_ts, end_ts, duration), ...]
    column_by_frame: опционально frame_number → excel column (из read_plan_columns).
    """
    if not xlsx_path.exists():
        raise FileNotFoundError(xlsx_path)
    wb = load_workbook(filename=str(xlsx_path))
    if SHEET_PLAN_V8 not in wb.sheetnames:
        raise ValueError(f"лист «{SHEET_PLAN_V8}» не найден")
    ws = wb[SHEET_PLAN_V8]
    if not _cell_text(ws, ROW_VOICE_END_V8, 1):
        ws.cell(row=ROW_VOICE_END_V8, column=1, value=ROW_VOICE_END_LABEL)

    if column_by_frame is None:
        cols = read_plan_columns(xlsx_path)
        column_by_frame = {c.frame_number: c.column for c in cols}

    for frame_number, start_ts, end_ts, duration in timings:
        col = column_by_frame.get(frame_number)
        if col is None:
            logger.warning(
                "xlsx: нет колонки для кадра {}, пропускаем таймкоды", frame_number
            )
            continue
        ws.cell(row=ROW_VOICE_END_V8, column=col, value=round(end_ts, 3))
        ws.cell(row=ROW_DURATION_V8, column=col, value=round(duration, 3))

    wb.save(xlsx_path)
    logger.info("xlsx: записаны таймкоды Whisper для {} кадров → {}", len(timings), xlsx_path)


def plan_columns_to_cells(columns: list[PlanColumn]) -> list[tuple[int, str]]:
    """[(frame_number, voiceover_text), ...] для map_frames."""
    return [(c.frame_number, c.voiceover_text) for c in columns]


def column_map(columns: list[PlanColumn]) -> dict[int, int]:
    return {c.frame_number: c.column for c in columns}
