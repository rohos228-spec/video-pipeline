"""Локальная разбивка voiceover → xlsx (fallback если GPT не записал R49).

GPT часто кладёт блоки в текст ответа (через «-»), а xlsx возвращает
без изменений — особенно со старым default.md. Тогда пишем блоки сами.
"""

from __future__ import annotations

import re
from pathlib import Path

from loguru import logger
from openpyxl import load_workbook

from app.services.xlsx_v8_import import (
    ROW_VOICEOVER_V8,
    SHEET_PLAN_V8,
    _resolve_plan_sheet,
)

_MIN_BLOCK_LEN = 8


def frames_spec_cell_stats(frames_spec: list[dict]) -> dict[str, float | int]:
    lens: list[int] = []
    for item in frames_spec:
        if not isinstance(item, dict):
            continue
        text = str(
            item.get("закадр")
            or item.get("voiceover_text")
            or item.get("voiceover")
            or ""
        ).strip()
        if text:
            lens.append(len(text))
    if not lens:
        return {"n": 0, "min": 0, "max": 0, "avg": 0.0}
    return {
        "n": len(lens),
        "min": min(lens),
        "max": max(lens),
        "avg": round(sum(lens) / len(lens), 1),
    }


def _split_oversized(text: str, *, max_chars: int, target: int) -> list[str]:
    raw = " ".join((text or "").split())
    if not raw:
        return []
    if len(raw) <= max_chars:
        return [raw]
    sentences = re.split(r"(?<=[.!?…])\s+", raw)
    if len(sentences) <= 1:
        words = raw.split()
        out: list[str] = []
        buf: list[str] = []
        for w in words:
            # одно «слово» длиннее max — режем посимвольно
            if len(w) > max_chars:
                if buf:
                    out.append(" ".join(buf).strip())
                    buf = []
                step = max(1, min(max_chars, target))
                for i in range(0, len(w), step):
                    chunk = w[i : i + step]
                    if chunk:
                        out.append(chunk)
                continue
            cand = (" ".join(buf + [w])).strip()
            if buf and len(cand) > max_chars:
                chunk = " ".join(buf).strip()
                if chunk:
                    out.append(chunk)
                buf = [w]
            else:
                buf.append(w)
        if buf:
            out.append(" ".join(buf).strip())
        return [b for b in out if b]
    out: list[str] = []
    buf = ""
    for sent in sentences:
        sent = sent.strip()
        if not sent:
            continue
        if not buf:
            buf = sent
            continue
        if len(buf) + 1 + len(sent) <= max_chars:
            buf = f"{buf} {sent}"
        else:
            out.append(buf)
            buf = sent
    if buf:
        out.append(buf)
    fixed: list[str] = []
    for b in out:
        if len(b) <= max_chars:
            fixed.append(b)
        else:
            fixed.extend(_split_oversized(b, max_chars=max_chars, target=target))
    return fixed


def _pack_words_to_limits(
    words: list[str],
    *,
    lo: int,
    hi: int,
    target: int,
) -> list[str]:
    """Упаковать слова в ячейки около target, не превышая hi (кроме слова > hi)."""
    if not words:
        return []
    chunks: list[str] = []
    buf: list[str] = []

    def flush() -> None:
        nonlocal buf
        if buf:
            chunks.append(" ".join(buf))
            buf = []

    for w in words:
        if len(w) > hi:
            flush()
            step = max(1, min(hi, target))
            for i in range(0, len(w), step):
                piece = w[i : i + step]
                if piece:
                    chunks.append(piece)
            continue
        if not buf:
            buf = [w]
            continue
        cand = f"{' '.join(buf)} {w}"
        cur_len = len(" ".join(buf))
        if len(cand) <= hi and not (
            cur_len >= target and cur_len >= lo and len(cand) > target
        ):
            buf.append(w)
            continue
        if cur_len >= lo:
            flush()
            buf = [w]
            continue
        # buf ещё короче min — добираем даже если уйдём чуть выше target,
        # пока не упрёмся в hi.
        if len(cand) <= hi:
            buf.append(w)
        else:
            flush()
            buf = [w]
    flush()
    return chunks


def _rebalance_short_chunks(chunks: list[str], *, lo: int, hi: int) -> list[str]:
    """Убрать ячейки < lo: украсть слова у соседей, иначе склеить (редко > hi)."""
    if not chunks:
        return []
    out = [c for c in chunks if c.strip()]
    if len(out) == 1:
        return out

    guard = 0
    while guard < 10_000:
        guard += 1
        short_i = next((i for i, c in enumerate(out) if len(c) < lo), None)
        if short_i is None:
            break

        i = short_i
        cur_words = out[i].split()
        moved = False

        # 1) тянем слова из следующей ячейки
        if i + 1 < len(out):
            nxt_words = out[i + 1].split()
            while nxt_words and len(" ".join(cur_words)) < lo:
                trial = cur_words + [nxt_words[0]]
                trial_s = " ".join(trial)
                rest_s = " ".join(nxt_words[1:])
                if len(trial_s) > hi:
                    break
                # не оставляем соседа пустым «в никуда» — ok; если rest < lo,
                # починим на следующем проходе
                if rest_s and len(rest_s) < lo and len(trial_s) + 1 + len(rest_s) <= hi:
                    # выгоднее сразу склеить целиком
                    break
                cur_words.append(nxt_words.pop(0))
                moved = True
            out[i] = " ".join(cur_words)
            if nxt_words:
                out[i + 1] = " ".join(nxt_words)
            else:
                del out[i + 1]
            if len(out[i]) >= lo:
                continue

        # 2) тянем слова с конца предыдущей
        if i > 0 and len(out[i]) < lo:
            prev_words = out[i - 1].split()
            cur_words = out[i].split()
            while prev_words and len(" ".join(cur_words)) < lo:
                trial = [prev_words[-1]] + cur_words
                trial_s = " ".join(trial)
                rest_s = " ".join(prev_words[:-1])
                if len(trial_s) > hi:
                    break
                if rest_s and len(rest_s) < lo:
                    break
                cur_words.insert(0, prev_words.pop())
                moved = True
            out[i] = " ".join(cur_words)
            if prev_words:
                out[i - 1] = " ".join(prev_words)
            else:
                del out[i - 1]
                i = max(0, i - 1)
            if len(out[i]) >= lo:
                continue

        # 3) принудительный merge — лучше чуть > hi, чем ячейка < lo
        if len(out[i]) < lo:
            if i + 1 < len(out):
                out[i] = f"{out[i]} {out[i + 1]}"
                del out[i + 1]
                moved = True
            elif i > 0:
                out[i - 1] = f"{out[i - 1]} {out[i]}"
                del out[i]
                moved = True
            else:
                break
        if not moved:
            break

    return [c for c in out if c.strip()]


def enforce_split_cell_limits(
    blocks: list[str],
    *,
    min_chars: int | None,
    max_chars: int | None,
    avg_min: int | None = None,
    avg_max: int | None = None,
) -> list[str]:
    """Склеить короткие / разрезать длинные ячейки под лимиты из настроек.

    Алгоритм: склеить весь текст → упаковать слова около target ≤ max →
    добить короткие (< min) кражей слов у соседей / merge.
    Старый split-then-merge оставлял хвосты 8–20 между «полными» соседями.
    """
    cleaned = [" ".join(str(b).split()) for b in blocks if str(b).strip()]
    if not cleaned:
        return []
    if min_chars is None and max_chars is None:
        return cleaned

    lo = int(min_chars) if min_chars is not None else 1
    hi = int(max_chars) if max_chars is not None else 10_000
    if hi < lo:
        hi = lo
    if avg_min is not None and avg_max is not None and avg_max >= avg_min:
        target = int(round((avg_min + avg_max) / 2))
    elif avg_min is not None:
        target = int(avg_min)
    elif avg_max is not None:
        target = int(avg_max)
    else:
        target = int(round((lo + hi) / 2))
    target = max(lo, min(hi, target))

    full = " ".join(cleaned)
    if len(full) < lo:
        return [full]

    words = full.split()
    packed = _pack_words_to_limits(words, lo=lo, hi=hi, target=target)
    return _rebalance_short_chunks(packed, lo=lo, hi=hi)


def parse_dash_separated_blocks(text: str) -> list[str]:
    """Блоки из ответа GPT: разделитель «-» на отдельной строке или inline."""
    raw = (text or "").strip()
    if not raw:
        return []
    parts = re.split(r"(?:\n\s*-\s*\n|\n\s*-\s+|\r\n\s*-\s+)", raw)
    if len(parts) <= 1 and " - " in raw:
        parts = raw.split(" - ")
    out: list[str] = []
    for p in parts:
        s = " ".join(p.split())
        if len(s) >= _MIN_BLOCK_LEN:
            out.append(s)
    return out


def split_voiceover_locally(text: str, *, target_len: int = 70) -> list[str]:
    """Грубая разбивка voiceover.txt по предложениям (~45–100 симв)."""
    raw = " ".join((text or "").split())
    if len(raw) < _MIN_BLOCK_LEN:
        return []
    sentences = re.split(r"(?<=[.!?…])\s+", raw)
    blocks: list[str] = []
    buf = ""
    for sent in sentences:
        sent = sent.strip()
        if not sent:
            continue
        if not buf:
            buf = sent
        elif len(buf) + 1 + len(sent) <= target_len + 30:
            buf = f"{buf} {sent}"
        else:
            if len(buf) >= _MIN_BLOCK_LEN:
                blocks.append(buf)
            buf = sent
    if buf and len(buf) >= _MIN_BLOCK_LEN:
        blocks.append(buf)
    return blocks


def write_voiceover_blocks_to_xlsx(xlsx_path: Path, blocks: list[str]) -> int:
    """Пишет блоки в строку 49 листа «план», колонки C..N."""
    if not blocks:
        return 0
    path = Path(xlsx_path)
    wb = load_workbook(filename=str(path))
    ws = _resolve_plan_sheet(wb)
    if ws is None:
        if SHEET_PLAN_V8 in wb.sheetnames:
            ws = wb[SHEET_PLAN_V8]
        else:
            ws = wb.active
            ws.title = SHEET_PLAN_V8
    for i, block in enumerate(blocks):
        ws.cell(row=ROW_VOICEOVER_V8, column=3 + i, value=block)
    wb.save(path)
    logger.info(
        "voiceover_split_local: wrote {} blocks to {} R{}",
        len(blocks),
        path.name,
        ROW_VOICEOVER_V8,
    )
    return len(blocks)
