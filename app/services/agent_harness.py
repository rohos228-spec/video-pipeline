"""Автономный харнес: verify проект → ops journal → soft repair (без wipe/audio/music)."""

from __future__ import annotations

import json
import sqlite3
import urllib.error
import urllib.request
import uuid
from dataclasses import asdict, dataclass, field
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from loguru import logger

from app.services.xlsx_v8_import import (
    ROW_IMAGE_PROMPT_V8,
    ROW_VIDEO_PROMPT_V8,
    ROW_VOICEOVER_V8,
)

HARNESS_HTTP_BASE = "http://127.0.0.1:8765"

# Шаги, которые харнес НИКОГДА не запускает.
HARNESS_FORBIDDEN_STEPS = frozenset({"audio", "music"})
# Soft repair кандидаты (если verify скажет missing/failed).
HARNESS_REPAIRABLE_STEPS = frozenset(
    {
        "plan",
        "script",
        "split",
        "hero",
        "items",
        "img_pr",
        "image_prompts",
        "img",
        "images",
        "anim_pr",
        "animation_prompts",
        "video",
        "videos",
        "assemble",
        "excel_gpt",
    }
)

# Шаговые N/N-гейты (harness_gate_or_raise step=), не путать со status assembled.
_IMG_PR_STEPS = frozenset({"img_pr", "image_prompts"})
_ANIM_PR_STEPS = frozenset({"anim_pr", "animation_prompts"})
_EXCEL_GPT_STEPS = frozenset(
    {"excel_gpt", "enrich_1", "enrich_2", "enrich_3", "enrich_4", "enrich_5"}
)

# PNG сцены — после шага img. anim_pr пишет текст из image_prompt, картинок ещё нет.
_SCENES_REQUIRED_STATUSES = {
    "images_ready",
    "generating_videos",
    "videos_ready",
    "generating_music",
    "music_ready",
    "generating_audio",
    "audio_ready",
    "assembling",
    "assembled",
    "publishing",
    "published",
}
_VIDEOS_REQUIRED_STATUSES = {
    "videos_ready",
    "generating_music",
    "music_ready",
    "generating_audio",
    "audio_ready",
    "assembling",
    "assembled",
    "publishing",
    "published",
}


@dataclass
class HarnessCheck:
    name: str
    ok: bool
    detail: str = ""


@dataclass
class HarnessReport:
    project_id: int
    run_id: str
    ok: bool
    status: str
    checks: list[HarnessCheck] = field(default_factory=list)
    next_action: str = "none"
    repair_steps: list[str] = field(default_factory=list)
    updated_at: str = ""

    def to_dict(self) -> dict[str, Any]:
        return {
            "project_id": self.project_id,
            "run_id": self.run_id,
            "ok": self.ok,
            "status": self.status,
            "checks": [asdict(c) for c in self.checks],
            "next_action": self.next_action,
            "repair_steps": list(self.repair_steps),
            "updated_at": self.updated_at,
        }


def _now() -> str:
    return datetime.now(UTC).isoformat()


def _repo_root() -> Path:
    return Path(__file__).resolve().parents[2]


def _db_path() -> Path:
    """Путь к SQLite для raw-проверок harness.

    Источник правды — ``settings.sqlite_path`` (тот же файл, что у приложения).
    Опциональный ``settings.db_path`` оставлен для monkeypatch в тестах.
    """
    from app.settings import settings

    legacy = getattr(settings, "db_path", None)
    if legacy:
        p = Path(legacy)
        if str(p):
            return p
    sp = getattr(settings, "sqlite_path", None)
    if sp:
        return Path(sp)
    return _repo_root() / "data" / "state.db"


def ops_dir_for_project(data_dir: Path) -> Path:
    d = data_dir / "ops"
    d.mkdir(parents=True, exist_ok=True)
    return d


def append_ops_event(data_dir: Path, event: dict[str, Any]) -> Path:
    path = ops_dir_for_project(data_dir) / "events.jsonl"
    row = dict(event)
    row.setdefault("at", _now())
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(row, ensure_ascii=False) + "\n")
    return path


def read_ops_telemetry(meta: dict[str, Any] | None) -> dict[str, Any]:
    m = meta if isinstance(meta, dict) else {}
    raw = m.get("ops_telemetry")
    return dict(raw) if isinstance(raw, dict) else {}


def write_ops_telemetry(meta: dict[str, Any], report: HarnessReport) -> dict[str, Any]:
    out = dict(meta) if isinstance(meta, dict) else {}
    out["ops_telemetry"] = {
        "harness_run_id": report.run_id,
        "updated_at": report.updated_at or _now(),
        "phase": "verify",
        "last_step": report.repair_steps[0] if report.repair_steps else None,
        "last_outcome": "verify_pass" if report.ok else "verify_fail",
        "artifact_ok": report.ok,
        "next_action": report.next_action,
        "checks": [asdict(c) for c in report.checks],
        "pointers": {
            "step_failure": (out.get("step_failure") or {}),
            "chrome_recovery": (out.get("chrome_recovery") or {}),
        },
    }
    return out


def _count_plan_rows_filled(xlsx: Path, rows: tuple[int, ...]) -> dict[int, int]:
    if not xlsx.is_file():
        return {r: 0 for r in rows}
    try:
        from openpyxl import load_workbook

        wb = load_workbook(xlsx, read_only=True, data_only=True)
        ws = None
        for name in wb.sheetnames:
            if name.strip().casefold() == "план":
                ws = wb[name]
                break
        if ws is None:
            wb.close()
            return {r: 0 for r in rows}
        out = {r: 0 for r in rows}
        max_col = max(int(ws.max_column or 2), 2)
        for col in range(3, max_col + 1):
            for r in rows:
                v = ws.cell(r, col).value
                if v and str(v).strip():
                    out[r] += 1
        wb.close()
        return out
    except Exception as e:  # noqa: BLE001
        logger.warning("harness plan rows read failed: {}", e)
        return {r: 0 for r in rows}


def _count_r48_filled(xlsx: Path) -> int:
    return _count_plan_rows_filled(xlsx, (ROW_VIDEO_PROMPT_V8,))[ROW_VIDEO_PROMPT_V8]


def _count_project_log_errors(project_id: int, slug: str) -> tuple[int, str]:
    data = _repo_root() / "data"
    logs = sorted(data.glob("backend-*.log"), key=lambda p: p.stat().st_mtime)
    log = logs[-1] if logs else data / "backend.log"
    if not log.is_file():
        return 0, "no-log"
    needles = (f"#{project_id}", slug)
    hits: list[str] = []
    try:
        lines = log.read_text(encoding="utf-8", errors="replace").splitlines()[-400:]
    except Exception as e:  # noqa: BLE001
        return 0, f"read-fail:{e}"
    for ln in lines:
        if any(n and n in ln for n in needles) and (
            "ERROR" in ln or "WARNING" in ln or "Traceback" in ln
        ):
            hits.append(ln)
    return len(hits), log.name


def _load_frame_prompt_rows(
    project_id: int,
) -> tuple[list[tuple[int, str, str, str]], str]:
    """Кадры: (number, voiceover, image_prompt, animation_prompt) + ошибка или ''."""
    db_file = _db_path()
    if not db_file.is_file():
        return [], ""
    try:
        db = sqlite3.connect(str(db_file))
        raw = db.execute(
            "SELECT number, "
            "COALESCE(voiceover_text, ''), "
            "COALESCE(image_prompt, ''), "
            "COALESCE(animation_prompt, '') "
            "FROM frames WHERE project_id=? ORDER BY number",
            (project_id,),
        ).fetchall()
        db.close()
    except Exception as e:  # noqa: BLE001
        return [], str(e)
    rows: list[tuple[int, str, str, str]] = []
    for r in raw:
        rows.append(
            (int(r[0] or 0), str(r[1] or ""), str(r[2] or ""), str(r[3] or ""))
        )
    return rows, ""


def _append_prompt_nn_checks(
    checks: list[HarnessCheck],
    repair: list[str],
    frame_rows: list[tuple[int, str, str, str]],
    *,
    status: str,
    step: str | None,
) -> None:
    """N/N по БД: img_pr — все VO с image_prompt; anim_pr — usable img → anim.

    Срабатывает по ``step`` (гейт до *_ready) и по уже выставленному status
    (второй рубеж auto_advance). excel_gpt не требует image_prompt.
    """
    step_key = (step or "").strip().lower()
    want_img = step_key in _IMG_PR_STEPS or status == "image_prompts_ready"
    want_anim = step_key in _ANIM_PR_STEPS or status == "animation_prompts_ready"
    if want_img:
        missing = [n for n, vo, img, _ in frame_rows if vo.strip() and not img.strip()]
        vo_n = sum(1 for _, vo, _, _ in frame_rows if vo.strip())
        img_n = vo_n - len(missing)
        ok = not missing
        checks.append(
            HarnessCheck(
                "img_pr_vo_coverage",
                ok,
                f"img_pr={img_n}/{vo_n} missing={missing[:20]}",
            )
        )
        if not ok:
            repair.append("img_pr")
    if want_anim:
        from app.generation_options import is_skippable_empty_prompt

        missing = [
            n
            for n, _, img, anim in frame_rows
            if (not is_skippable_empty_prompt(img)) and is_skippable_empty_prompt(anim)
        ]
        usable = sum(
            1 for _, _, img, _ in frame_rows if not is_skippable_empty_prompt(img)
        )
        ok = not missing
        checks.append(
            HarnessCheck(
                "anim_pr_coverage",
                ok,
                f"anim={usable - len(missing)}/{usable} missing={missing[:20]}",
            )
        )
        if not ok:
            repair.append("anim_pr")
    if step_key in _EXCEL_GPT_STEPS:
        checks.append(
            HarnessCheck(
                "excel_gpt_gate",
                True,
                "disk checks N/A; node coverage already applied",
            )
        )


def verify_project_disk(
    project_id: int, data_dir: Path, status: str, *, step: str | None = None
) -> HarnessReport:
    """Синхронная проверка диска/БД без HTTP (для CLI и сервиса)."""
    run_id = uuid.uuid4().hex[:12]
    checks: list[HarnessCheck] = []
    repair: list[str] = []

    xlsx = data_dir / "project.xlsx"
    checks.append(
        HarnessCheck("project_xlsx", xlsx.is_file(), str(xlsx) if xlsx.is_file() else "missing")
    )
    if not xlsx.is_file():
        repair.append("plan")

    scenes = list((data_dir / "scenes").glob("*.png")) if (data_dir / "scenes").is_dir() else []
    videos = list((data_dir / "videos").glob("*.mp4")) if (data_dir / "videos").is_dir() else []
    heroes = (
        list((data_dir / "characters").glob("*.png"))
        if (data_dir / "characters").is_dir()
        else []
    )
    final = list((data_dir / "final").glob("*.mp4")) if (data_dir / "final").is_dir() else []

    scenes_required = status in _SCENES_REQUIRED_STATUSES
    videos_required = status in _VIDEOS_REQUIRED_STATUSES
    checks.append(
        HarnessCheck(
            "scenes_png",
            (not scenes_required) or len(scenes) >= 1,
            f"n={len(scenes)} required={scenes_required} status={status}",
        )
    )
    checks.append(
        HarnessCheck(
            "videos_mp4",
            (not videos_required) or len(videos) >= 1,
            f"n={len(videos)} required={videos_required} status={status}",
        )
    )
    # Персонажи: после hero_ready png обязательны, если в xlsx есть записи.
    excel_persons: list[Any] = []
    polluted: list[str] = []
    if xlsx.is_file():
        try:
            from app.services.excel_characters import (
                is_polluted_character_field,
                parse_persons_sheet,
            )

            excel_persons = parse_persons_sheet(xlsx)
            for ch in excel_persons:
                if is_polluted_character_field(ch.name) or is_polluted_character_field(
                    ch.look
                ):
                    polluted.append(ch.id)
        except Exception as e:  # noqa: BLE001
            logger.debug("harness persons parse skip: {}", e)

    # Полный комплект characters/*.png — только ПОСЛЕ фазы hero.
    # На hero_ready в HITL/ai_control нормально 1 из N (сначала c01, аппрув,
    # потом c02-реф). Требовать всех сразу → harness PAUSE, c02 не стартует.
    _HEROES_FULL_STATUSES = {
        "items_ready",
        "generating_items",
        "enrich_1_ready",
        "enrich_2_ready",
        "enrich_3_ready",
        "enrich_4_ready",
        "enrich_5_ready",
        "generating_image_prompts",
        "image_prompts_ready",
        "generating_images",
        "images_ready",
        *_SCENES_REQUIRED_STATUSES,
    }
    heroes_full = bool(excel_persons) and status in _HEROES_FULL_STATUSES
    heroes_ok = (not heroes_full) or len(heroes) >= len(excel_persons)
    checks.append(
        HarnessCheck(
            "heroes_png",
            heroes_ok,
            f"n={len(heroes)} excel={len(excel_persons)} "
            f"full_required={heroes_full} status={status}",
        )
    )
    if heroes_full and not heroes_ok:
        repair.append("hero")

    if polluted:
        checks.append(
            HarnessCheck(
                "excel_hero_fields_sane",
                False,
                f"polluted character fields: {', '.join(polluted)}",
            )
        )
        repair.append("hero")

    # Sheet vs cinematic scene (c02-баг: png есть, но это не turnaround).
    try:
        from app.services.hero_quality import collect_hero_quality_checks

        for hc in collect_hero_quality_checks(project_id, data_dir):
            # не дублировать excel_hero_fields_sane если уже добавили polluted
            if hc.name == "excel_hero_fields_sane" and polluted:
                continue
            checks.append(hc)
            if not hc.ok:
                repair.append("hero")
    except Exception as e:  # noqa: BLE001
        checks.append(HarnessCheck("hero_quality", False, str(e)[:200]))
        repair.append("hero")

    checks.append(
        HarnessCheck(
            "final_mp4",
            status != "assembled" or len(final) >= 1,
            f"n={len(final)} status={status}",
        )
    )
    if status == "assembled" and not final:
        repair.append("assemble")
    if scenes_required and not scenes:
        repair.append("img")
    if videos_required and not videos:
        repair.append("video")

    r48 = _count_r48_filled(xlsx) if xlsx.is_file() else 0
    # R48 обязателен только на/после anim_pr. На hero_ready старые scenes/
    # пустой R48 НЕ должны PAUSE'ить генерацию c02.
    _R48_REQUIRED_STATUSES = {
        "animation_prompts_ready",
        "generating_videos",
        "videos_ready",
        "generating_audio",
        "audio_ready",
        "generating_music",
        "music_ready",
        "assembling",
        "assembled",
        "published",
    }
    if scenes and status in _R48_REQUIRED_STATUSES:
        r48_ok = r48 >= len(scenes)
        if not r48_ok:
            repair.append("anim_pr")
    else:
        r48_ok = True
    checks.append(
        HarnessCheck(
            "r48_anim",
            r48_ok,
            f"filled={r48} scenes={len(scenes)} status={status}",
        )
    )

    # DB frames ↔ xlsx plan parity (строго для assembled, иначе диагностика).
    plan_rows = _count_plan_rows_filled(
        xlsx,
        (ROW_IMAGE_PROMPT_V8, ROW_VIDEO_PROMPT_V8, ROW_VOICEOVER_V8),
    )
    r45 = plan_rows[ROW_IMAGE_PROMPT_V8]
    vo_xlsx = plan_rows[ROW_VOICEOVER_V8]
    frames_total = img_pr_db = anim_pr_db = vo_db = 0
    parity_err = ""
    try:
        db_file = _db_path()
        if not db_file.is_file():
            # БД ещё нет — для ранних статусов это не блок (0 кадров).
            frames_total = img_pr_db = anim_pr_db = vo_db = 0
        else:
            db = sqlite3.connect(str(db_file))
            row = db.execute(
                "SELECT COUNT(*), "
                "SUM(CASE WHEN image_prompt IS NOT NULL AND trim(image_prompt)<>'' THEN 1 ELSE 0 END), "
                "SUM(CASE WHEN animation_prompt IS NOT NULL AND trim(animation_prompt)<>'' THEN 1 ELSE 0 END), "
                "SUM(CASE WHEN voiceover_text IS NOT NULL AND trim(voiceover_text)<>'' THEN 1 ELSE 0 END) "
                "FROM frames WHERE project_id=?",
                (project_id,),
            ).fetchone()
            db.close()
            if row:
                frames_total = int(row[0] or 0)
                img_pr_db = int(row[1] or 0)
                anim_pr_db = int(row[2] or 0)
                vo_db = int(row[3] or 0)
    except Exception as e:  # noqa: BLE001
        parity_err = str(e)
    if parity_err:
        checks.append(HarnessCheck("frames_xlsx_parity", False, parity_err))
    else:
        parity_ok = True
        if status == "assembled" and scenes:
            n = len(scenes)
            parity_ok = (
                frames_total >= n
                and img_pr_db >= n
                and anim_pr_db >= n
                and vo_db >= n
                and r45 >= n
                and r48 >= n
                and vo_xlsx >= n
            )
            if not parity_ok:
                if img_pr_db < n or r45 < n:
                    repair.append("img_pr")
                if anim_pr_db < n or r48 < n:
                    repair.append("anim_pr")
        checks.append(
            HarnessCheck(
                "frames_xlsx_parity",
                parity_ok,
                f"frames={frames_total} img_db={img_pr_db} anim_db={anim_pr_db} vo_db={vo_db} "
                f"r45={r45} r48={r48} r49={vo_xlsx} scenes={len(scenes)}",
            )
        )

    nn_rows, nn_err = _load_frame_prompt_rows(project_id)
    if nn_err:
        step_key = (step or "").strip().lower()
        if step_key in _IMG_PR_STEPS or status == "image_prompts_ready":
            checks.append(HarnessCheck("img_pr_vo_coverage", False, nn_err))
            repair.append("img_pr")
        elif step_key in _ANIM_PR_STEPS or status == "animation_prompts_ready":
            checks.append(HarnessCheck("anim_pr_coverage", False, nn_err))
            repair.append("anim_pr")
        if step_key in _EXCEL_GPT_STEPS:
            checks.append(
                HarnessCheck(
                    "excel_gpt_gate",
                    True,
                    "disk checks N/A; node coverage already applied",
                )
            )
    else:
        _append_prompt_nn_checks(
            checks, repair, nn_rows, status=status, step=step
        )

    # node_runs failed
    failed_n = 0
    try:
        db_file = _db_path()
        if not db_file.is_file():
            checks.append(HarnessCheck("node_runs_failed", True, "failed=0 db_missing"))
        else:
            db = sqlite3.connect(str(db_file))
            row = db.execute(
                "SELECT id FROM workflow_runs WHERE project_id=? ORDER BY id DESC LIMIT 1",
                (project_id,),
            ).fetchone()
            if row:
                failed_n = db.execute(
                    "SELECT COUNT(*) FROM node_runs WHERE workflow_run_id=? AND status='failed'",
                    (row[0],),
                ).fetchone()[0]
            db.close()
            checks.append(
                HarnessCheck("node_runs_failed", failed_n == 0, f"failed={failed_n}")
            )
    except Exception as e:  # noqa: BLE001
        checks.append(HarnessCheck("node_runs", False, str(e)))

    log_hits, log_name = _count_project_log_errors(project_id, data_dir.name)
    checks.append(
        HarnessCheck(
            "project_log_clean",
            log_hits == 0,
            f"hits={log_hits} log={log_name}",
        )
    )

    # voiceover hygiene (optional)
    vo = data_dir / "voiceover.txt"
    if vo.is_file():
        text = vo.read_text(encoding="utf-8", errors="replace")
        polluted = "# Лист:" in text or "vp.check.v1" in text or "ОТЧЁТ ПРОВЕРКИ" in text
        checks.append(HarnessCheck("voiceover_clean", not polluted, "polluted" if polluted else "ok"))
        if polluted:
            repair.append("script")

    ok = all(c.ok for c in checks)
    # Sheet/quality checks may append "hero" несколько раз — схлопнуть.
    repair = list(dict.fromkeys(s for s in repair if s not in HARNESS_FORBIDDEN_STEPS))
    next_action = "none" if ok else ("repair:" + ",".join(repair) if repair else "investigate")
    return HarnessReport(
        project_id=project_id,
        run_id=run_id,
        ok=ok,
        status=status,
        checks=checks,
        next_action=next_action,
        repair_steps=repair,
        updated_at=_now(),
    )


def _http_get(url: str, *, timeout: float = 20.0, max_bytes: int = 65_536) -> tuple[int, int, bytes]:
    try:
        with urllib.request.urlopen(url, timeout=timeout) as r:
            data = r.read(max_bytes)
            cl = r.headers.get("Content-Length")
            n = int(cl) if cl and cl.isdigit() else len(data)
            return r.status, n, data
    except urllib.error.HTTPError as e:
        return e.code, 0, b""
    except Exception:
        return 0, 0, b""


def _http_json(url: str, *, timeout: float = 20.0) -> tuple[int, Any]:
    status, _, data = _http_get(url, timeout=timeout)
    if status != 200:
        return status, None
    try:
        return status, json.loads(data.decode("utf-8"))
    except Exception:
        return status, None


def verify_project_http(
    project_id: int,
    data_dir: Path,
    *,
    base_url: str = HARNESS_HTTP_BASE,
) -> list[HarnessCheck]:
    """HTTP-слой проверки: Studio API, artifacts/assets/xlsx доступны по факту."""
    base = base_url.rstrip("/")
    checks: list[HarnessCheck] = []

    st, ver = _http_json(f"{base}/api/studio-version")
    ver_ok = bool(isinstance(ver, dict) and ver.get("backend_ok") and ver.get("pipeline_ok"))
    checks.append(
        HarnessCheck(
            "studio_http",
            st == 200 and ver_ok,
            f"status={st} build={ver.get('build') if isinstance(ver, dict) else '?'} "
            f"backend_git={ver.get('backend_git') if isinstance(ver, dict) else '?'}",
        )
    )

    st, proj = _http_json(f"{base}/api/projects/{project_id}")
    checks.append(
        HarnessCheck(
            "project_http",
            st == 200 and isinstance(proj, dict),
            f"status={st} project_status={proj.get('status') if isinstance(proj, dict) else '?'} "
            f"slug={proj.get('slug') if isinstance(proj, dict) else '?'}",
        )
    )

    st, frames = _http_json(f"{base}/api/projects/{project_id}/frames")
    if st == 200 and isinstance(frames, list):
        img_pr = sum(1 for f in frames if (f.get("image_prompt") or "").strip())
        anim_pr = sum(1 for f in frames if (f.get("animation_prompt") or "").strip())
        vo = sum(1 for f in frames if (f.get("voiceover_text") or "").strip())
        checks.append(
            HarnessCheck(
                "frames_http",
                True,
                f"frames={len(frames)} img_pr={img_pr} anim_pr={anim_pr} voiceover={vo}",
            )
        )
    else:
        checks.append(HarnessCheck("frames_http", False, f"status={st}"))

    artifacts: list[sqlite3.Row] = []
    try:
        db = sqlite3.connect(str(_db_path()))
        db.row_factory = sqlite3.Row
        artifacts = db.execute(
            "SELECT kind, uuid, path FROM artifacts WHERE project_id=? ORDER BY kind, id",
            (project_id,),
        ).fetchall()
        db.close()
    except Exception as e:  # noqa: BLE001
        checks.append(HarnessCheck("artifacts_http", False, f"db: {e}"))
    else:
        fail = 0
        checked = 0
        for a in artifacts:
            checked += 1
            status, nbytes, _ = _http_get(
                f"{base}/api/artifacts/{a['uuid']}/file", timeout=30.0, max_bytes=1
            )
            if status != 200 or nbytes <= 0:
                fail += 1
        checks.append(
            HarnessCheck(
                "artifacts_http",
                fail == 0,
                f"checked={checked} fail={fail}",
            )
        )

    st, assets = _http_json(f"{base}/api/projects/{project_id}/assets")
    if st == 200 and isinstance(assets, list):
        fail = 0
        checked = 0
        for a in assets:
            prev = a.get("preview_url") or ""
            if not prev:
                continue
            url = prev if prev.startswith("http") else base + prev
            status, nbytes, _ = _http_get(url, timeout=30.0, max_bytes=1)
            checked += 1
            if status != 200 or nbytes <= 0:
                fail += 1
            if checked >= 40:
                break
        checks.append(
            HarnessCheck("assets_http", fail == 0, f"checked={checked} fail={fail}")
        )
    else:
        checks.append(HarnessCheck("assets_http", False, f"status={st}"))

    st, nbytes, _ = _http_get(f"{base}/api/projects/{project_id}/xlsx", timeout=30.0)
    checks.append(HarnessCheck("xlsx_http", st == 200 and nbytes > 0, f"status={st} bytes={nbytes}"))

    scenes = list((data_dir / "scenes").glob("*.png")) if (data_dir / "scenes").is_dir() else []
    checks.append(HarnessCheck("http_scene_parity", True, f"scenes_disk={len(scenes)}"))
    return checks


_GATE_OBSERVABILITY = frozenset({"project_log_clean", "node_runs_failed"})


async def harness_gate_or_raise(
    session: Any, project: Any, *, step: str
) -> HarnessReport:
    """Harness-гейт после шага: плохие данные → RuntimeError (soft retry шага).

    Наблюдаемость (лог/старые failed-записи) в гейт не входит — иначе ошибка
    шага в логе блокировала бы повтор навсегда. Используется нодами после
    записи (anim_pr/img_pr/split/plan…); центральный гейт — в auto_advance.
    Вызывать ДО смены статуса на *_ready: N/N смотрит ``step``, не status.
    """
    report = await run_harness_verify(
        session, project, allow_repair=False, include_http=False, step=step
    )
    bad = [
        f"{c.name}({c.detail})"
        for c in report.checks
        if not c.ok and c.name not in _GATE_OBSERVABILITY
    ]
    if bad:
        raise RuntimeError(f"{step} harness gate failed: {bad}")
    logger.info(
        "[#{}] {}: harness gate ok ({} checks)", project.id, step, len(report.checks)
    )
    return report


async def run_harness_verify(
    session: Any,
    project: Any,
    *,
    allow_repair: bool = False,
    include_http: bool = False,
    http_base_url: str = HARNESS_HTTP_BASE,
    step: str | None = None,
) -> HarnessReport:
    """Verify + optional soft repair (POST step run via project_steps)."""
    data_dir = Path(project.data_dir)
    status = str(getattr(project.status, "value", project.status) or "")
    report = verify_project_disk(int(project.id), data_dir, status, step=step)
    if include_http:
        try:
            import anyio

            http_checks = await anyio.to_thread.run_sync(
                lambda: verify_project_http(
                    int(project.id), data_dir, base_url=http_base_url
                )
            )
        except Exception as e:  # noqa: BLE001
            http_checks = [HarnessCheck("http_checks", False, str(e))]
        report.checks.extend(http_checks)
        report.ok = all(c.ok for c in report.checks)
        if not report.ok and report.next_action == "none":
            report.next_action = "investigate:http"

    meta = dict(project.meta) if isinstance(project.meta, dict) else {}
    project.meta = write_ops_telemetry(meta, report)
    append_ops_event(
        data_dir,
        {
            "type": "harness_verify",
            "run_id": report.run_id,
            "ok": report.ok,
            "status": status,
            "checks": [asdict(c) for c in report.checks],
            "next_action": report.next_action,
        },
    )

    if allow_repair and report.repair_steps:
        step = report.repair_steps[0]
        if step in HARNESS_FORBIDDEN_STEPS:
            logger.warning("harness: refuse forbidden step {}", step)
        else:
            try:
                from app.services.project_steps import start_step

                code = {
                    "image_prompts": "img_pr",
                    "images": "img",
                    "animation_prompts": "anim_pr",
                    "videos": "video",
                }.get(step, step)
                if code in HARNESS_FORBIDDEN_STEPS:
                    raise RuntimeError(f"forbidden step {code}")
                logger.info("harness: soft repair step={} project={}", code, project.id)
                append_ops_event(
                    data_dir,
                    {"type": "harness_repair", "run_id": report.run_id, "step": code},
                )
                await start_step(
                    session,
                    project,
                    code,
                    require_node_fsm=False,
                    explicit_ui_start=True,
                )
                report.next_action = f"repaired:{code}"
            except Exception as e:  # noqa: BLE001
                logger.warning("harness repair failed: {}", e)
                report.next_action = f"repair_failed:{step}:{e}"
                append_ops_event(
                    data_dir,
                    {
                        "type": "harness_repair_failed",
                        "run_id": report.run_id,
                        "step": step,
                        "error": str(e),
                    },
                )

    await session.flush()
    return report


def harness_status_payload(project: Any) -> dict[str, Any]:
    meta = project.meta if isinstance(project.meta, dict) else {}
    tel = read_ops_telemetry(meta)
    data_dir = Path(project.data_dir)
    events_path = data_dir / "ops" / "events.jsonl"
    last_events: list[dict[str, Any]] = []
    if events_path.is_file():
        lines = events_path.read_text(encoding="utf-8", errors="replace").splitlines()
        for ln in lines[-20:]:
            try:
                last_events.append(json.loads(ln))
            except json.JSONDecodeError:
                continue
    return {
        "project_id": project.id,
        "status": str(getattr(project.status, "value", project.status)),
        "ops_telemetry": tel,
        "events_tail": last_events,
        "forbidden_steps": sorted(HARNESS_FORBIDDEN_STEPS),
    }
