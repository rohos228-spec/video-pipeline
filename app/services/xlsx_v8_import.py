"""Импорт v8-xlsx → БД. Используется в двух местах:

1. _backfill_from_disk на старте бота — подтягиваем xlsx/voiceover.txt
   в БД для всех проектов, чтобы recompute_status не откатил статус
   из-за пустых полей.
2. После xlsx-flow шагов 1 («План») и 3 («Разбивка») в TG-боте —
   синхронизируем свежий xlsx, который GPT прислал, в БД.

Логика вытащена из standalone-скрипта `import_from_xlsx.py` (см. там
оригинал и описание формата v8).

v8-шаблон отличается от старого (v7):
  - лист «Общий план» (без «ролика» в имени)
  - лист «план» (вместо «Кадры»), кадры стоят колонками 3..N,
    voiceover лежит в строке 49.

Идемпотентный: повторный запуск ничего не ломает, обновляет только то,
что изменилось.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from loguru import logger
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.generation_options import is_skippable_empty_prompt
from app.models import Frame, FrameStatus, Project
from app.services.content_locks import is_ui_locked
from app.services.plan_validation import is_meaningful_general_plan

# --- константы под v8-шаблон ---------------------------------------------
SHEET_GENERAL_V8 = "Общий план"
SHEET_PLAN_V8 = "план"
# В v8-шаблоне «план» каждый кадр — это колонка (col=3..N). Эти строки —
# поля одного кадра. См. templates/project_template_v8.xlsx (column A):
ROW_IMAGE_PROMPT_V8 = 45  # «промт для картинки 1»
ROW_IMAGE_PROMPT_2_V8 = 46  # «промт для картинки 2» (shot_02)
# R46/R47 — резервные «картинка 2/3» (модель пока одну хранит, см. Frame.image_prompt)
ROW_VIDEO_PROMPT_V8 = 48  # «промт для видео» (shot_01)
ROW_VOICEOVER_V8 = 49     # «закадровый текст»
ROW_VIDEO_PROMPT_2_V8 = 64  # «промт для видео 2» (shot_02)
ROW_DURATION_V8 = 50      # «Время на кадр»

# Подписи строки промта shot_01 (колонка A/B) — fallback если R45 пустая.
_PLAN_IMAGE_PROMPT_LABELS: tuple[str, ...] = (
    "промт для картинки 1",
    "промт для картинки",
    "промт картинки",
    "image prompt 1",
    "image prompt",
)

# Длительность кадра — проп. длине voiceover-блока (русская речь ~14 симв/сек).
CHARS_PER_SEC = 14.0
MIN_FRAME = 1.5
MAX_FRAME = 6.0


def _normalize_sheet_name(name: str) -> str:
    """Убрать NBSP/пробелы по краям — Excel часто даёт «план » или «план\\xa0»."""
    return " ".join(str(name).replace("\xa0", " ").split())


def _resolve_plan_sheet(wb):
    """Лист «план» (v8): без учёта регистра, с trim/NBSP в имени."""
    target = _normalize_sheet_name(SHEET_PLAN_V8).casefold()
    for name in wb.sheetnames:
        if _normalize_sheet_name(name).casefold() == target:
            return wb[name]
    return None


def has_v8_plan_sheet(wb) -> bool:
    return _resolve_plan_sheet(wb) is not None


def _distribute_durations(cells: list[str]) -> list[float]:
    if not cells:
        return []
    return [
        round(min(max(len(c) / CHARS_PER_SEC, MIN_FRAME), MAX_FRAME), 2)
        for c in cells
    ]


def _read_general_plan(wb) -> str | None:
    if SHEET_GENERAL_V8 not in wb.sheetnames:
        return None
    ws = wb[SHEET_GENERAL_V8]
    lines: list[str] = []
    block_header_row: int | None = None

    for r in range(1, min(ws.max_row, 200) + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        a_s = str(a).strip() if a is not None else ""
        b_s = str(b).strip() if b is not None else ""

        if a_s and not b_s:
            lines.append(f"\n## {a_s}\n")
            block_header_row = r + 1
            continue

        if block_header_row == r:
            headers = [
                ws.cell(row=r, column=c).value for c in range(1, 6)
            ]
            if all(h for h in headers):
                block_header_row = -1
                continue
            block_header_row = None

        if a_s and b_s:
            lines.append(f"**{a_s}:** {b_s}")
            continue

        if block_header_row == -1 and any(
            ws.cell(row=r, column=c).value for c in range(1, 6)
        ):
            row_cells = [
                str(ws.cell(row=r, column=c).value or "").strip()
                for c in range(1, 6)
            ]
            if row_cells[0]:
                lines.append(f"\n### {row_cells[0]}")
            for label, idx in [
                ("Основная мысль", 1),
                ("Подтемы", 2),
                ("Функция", 3),
                ("Как подводит к следующему", 4),
            ]:
                if row_cells[idx]:
                    lines.append(f"- **{label}:** {row_cells[idx]}")

    text = "\n".join(line for line in lines if line.strip()).strip()
    return text if text else None


def _resolve_plan_cell(ws, row: int, col: int):
    """Ячейка листа «план»; для MergedCell — top-left merge (иначе value=None)."""
    cell = ws.cell(row=row, column=col)
    if type(cell).__name__ != "MergedCell":
        return cell
    merged = getattr(ws, "merged_cells", None)
    if merged is None:
        return cell
    for cr in merged.ranges:
        if cr.min_row <= row <= cr.max_row and cr.min_col <= col <= cr.max_col:
            return ws.cell(row=cr.min_row, column=cr.min_col)
    return cell


def _cell_text(ws, row: int, col: int) -> str | None:
    """Прочитать ячейку как trim-string, схлопнув whitespace. None если пусто."""
    v = _resolve_plan_cell(ws, row, col).value
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    return " ".join(s.split())


def _is_formula_cell_text(text: str) -> bool:
    """Формула без закэшированного значения (data_only=False) — не промт."""
    return text.lstrip().startswith("=")


def _discover_labeled_row(ws, labels: tuple[str, ...]) -> int | None:
    """Номер строки по подписи в колонках A/B (без учёта регистра)."""
    max_row = min(ws.max_row or 0, 120)
    for row in range(1, max_row + 1):
        for col in (1, 2):
            raw = _cell_text(ws, row, col)
            if not raw:
                continue
            low = raw.casefold()
            for label in labels:
                if label.casefold() in low:
                    return row
    return None


def _frame_number_for_plan_column(ws, col: int) -> int:
    """Номер кадра для колонки: id в R4/R3/R2/R1 → иначе col-2 (C=1)."""
    for header_row in (4, 3, 2, 1):
        val = _cell_text(ws, header_row, col)
        if val and val.isdigit():
            return int(val)
    if col >= 3:
        return col - 2
    if col >= 2:
        return col - 1
    return col


def _plan_prompt_row(ws) -> int:
    """Строка image_prompt shot_01: R45, либо по подписи в A/B."""
    max_col = ws.max_column or 0
    if any(
        _cell_text(ws, ROW_IMAGE_PROMPT_V8, col)
        for col in range(2, max_col + 1)
    ):
        return ROW_IMAGE_PROMPT_V8
    discovered = _discover_labeled_row(ws, _PLAN_IMAGE_PROMPT_LABELS)
    return discovered or ROW_IMAGE_PROMPT_V8


def _plan_scene_column_range(ws, row: int) -> range:
    """Колонки B..N для строки промта — вся ширина листа (ws.max_column)."""
    last = ws.max_column or 0
    for col in range(last, 1, -1):
        if _cell_text(ws, row, col):
            last = max(last, col)
            break
    end = max(last, 3)
    return range(2, end + 1)


def _plan_prompt_columns(ws, row: int) -> list[int]:
    """Колонки с промтом в строке row — одна на merge-range (не дублировать slaves)."""
    cols: list[int] = []
    seen_merges: set[tuple[int, int, int, int]] = set()
    for col in _plan_scene_column_range(ws, row):
        prompt = _cell_text(ws, row, col)
        if not prompt or _is_formula_cell_text(prompt):
            continue
        merge_key: tuple[int, int, int, int] | None = None
        merged = getattr(ws, "merged_cells", None)
        if merged is not None:
            for cr in merged.ranges:
                if cr.min_row <= row <= cr.max_row and cr.min_col <= col <= cr.max_col:
                    merge_key = (cr.min_row, cr.min_col, cr.max_row, cr.max_col)
                    break
        if merge_key is not None:
            if merge_key in seen_merges:
                continue
            seen_merges.add(merge_key)
        cols.append(col)
    return cols


def _plan_scene_columns_ordered(ws) -> list[tuple[int, int]]:
    """(номер кадра 1..N, колонка Excel) — R49 или уникальные колонки R45/R46."""
    max_col = ws.max_column or 0
    vo_cols = [
        c
        for c in range(2, max_col + 1)
        if _cell_text(ws, ROW_VOICEOVER_V8, c)
    ]
    if vo_cols:
        return list(enumerate(vo_cols, start=1))
    seen: list[int] = []
    for row in (ROW_IMAGE_PROMPT_V8, ROW_IMAGE_PROMPT_2_V8):
        for col in _plan_prompt_columns(ws, row):
            if col not in seen:
                seen.append(col)
    return list(enumerate(seen, start=1))


def _read_plan_image_prompts_ws(ws, *, prompt_row: int | None = None) -> dict[int, str]:
    row = prompt_row or _plan_prompt_row(ws)
    out: dict[int, str] = {}
    for frame_num, col in _plan_scene_columns_ordered(ws):
        prompt = _cell_text(ws, row, col)
        if not prompt or _is_formula_cell_text(prompt):
            continue
        out[frame_num] = prompt
    if out:
        return out
    for col in _plan_scene_column_range(ws, row):
        prompt = _cell_text(ws, row, col)
        if not prompt or _is_formula_cell_text(prompt):
            continue
        out[_frame_number_for_plan_column(ws, col)] = prompt
    return out


def _read_v7_frames_image_prompts_ws(wb) -> dict[int, str]:
    """Лист «Кадры» (legacy v7): R29, номера кадров в строке заголовка."""
    from app.storage.project_sheet import (
        ROW_HEADER,
        ROW_IMAGE_PROMPT,
        SHEET_FRAMES,
    )

    if SHEET_FRAMES not in wb.sheetnames:
        return {}
    ws = wb[SHEET_FRAMES]
    out: dict[int, str] = {}
    max_col = ws.max_column or 0
    col_to_frame: dict[int, int] = {}
    for col in range(2, max_col + 1):
        n = ws.cell(row=ROW_HEADER, column=col).value
        try:
            col_to_frame[col] = int(n)
        except (TypeError, ValueError):
            continue
    if not col_to_frame:
        for col in range(2, max_col + 1):
            col_to_frame[col] = col - 1
    for col, fnum in col_to_frame.items():
        prompt = _cell_text(ws, ROW_IMAGE_PROMPT, col)
        if prompt and not _is_formula_cell_text(prompt):
            out[fnum] = prompt
    return out


def _read_image_prompts_workbook(wb) -> dict[int, str]:
    merged: dict[int, str] = {}
    ws = _resolve_plan_sheet(wb)
    if ws is not None:
        for num, text in _read_plan_image_prompts_ws(ws).items():
            merged[num] = text
    for num, text in _read_v7_frames_image_prompts_ws(wb).items():
        merged.setdefault(num, text)
    return merged


def read_image_prompts_from_project_xlsx(xlsx_path: Path) -> dict[int, str]:
    """image_prompt из project.xlsx: лист «план» R45 (+подпись строки) и «Кадры» v7.

    data_only=True и False — если Excel не пересохраняли, кэш формул может быть пуст.
    """
    from openpyxl import load_workbook

    if not xlsx_path.exists():
        return {}
    merged: dict[int, str] = {}
    for data_only in (True, False):
        try:
            wb = load_workbook(filename=str(xlsx_path), data_only=data_only)
        except Exception as e:  # noqa: BLE001
            logger.warning(
                "read_image_prompts: open {} data_only={}: {}",
                xlsx_path,
                data_only,
                e,
            )
            continue
        try:
            part = _read_image_prompts_workbook(wb)
        finally:
            wb.close()
        for num, text in part.items():
            if num not in merged and text:
                merged[num] = text
    return merged


def describe_image_prompts_xlsx_scan(xlsx_path: Path) -> str:
    """Краткая диагностика для логов/ошибок generate_images."""
    from openpyxl import load_workbook

    if not xlsx_path.exists():
        return "project.xlsx не найден"
    try:
        st = xlsx_path.stat()
        size_m = f"{st.st_size} B, mtime={st.st_mtime:.0f}"
    except OSError:
        size_m = "stat failed"
    try:
        wb = load_workbook(filename=str(xlsx_path), data_only=True)
    except Exception as e:  # noqa: BLE001
        return f"не открыть xlsx ({size_m}): {e}"
    try:
        sheets = list(wb.sheetnames)
        ws = _resolve_plan_sheet(wb)
        if ws is None:
            v7 = "Кадры" in sheets
            norm = [_normalize_sheet_name(s) for s in sheets]
            return (
                f"path={xlsx_path} ({size_m}); лист «план» не найден "
                f"(имена={sheets!r}, norm={norm!r}); v7 Кадры={v7}"
            )
        row = _plan_prompt_row(ws)
        n = len(_read_plan_image_prompts_ws(ws, prompt_row=row))
        return (
            f"path={xlsx_path} ({size_m}); лист «{ws.title}», строка R{row}, "
            f"найдено {n} промтов; листы={sheets!r}"
        )
    finally:
        wb.close()


def _map_prompts_to_frame_numbers(
    frames: list[Any], prompts: dict[int, str]
) -> dict[int, str]:
    """Сопоставить промты xlsx с Frame.number; fallback — по порядку 1:1."""
    if not prompts or not frames:
        return {}
    by_num = {f.number: f for f in frames}
    direct = {n: t for n, t in prompts.items() if n in by_num and (t or "").strip()}
    if direct:
        return direct
    ordered_frames = sorted(by_num.keys())
    ordered_texts = [prompts[k] for k in sorted(prompts.keys()) if prompts[k]]
    return {
        ordered_frames[i]: ordered_texts[i]
        for i in range(min(len(ordered_frames), len(ordered_texts)))
    }


def apply_image_prompts_from_xlsx_to_frames(
    frames: list[Any],
    xlsx_path: Path,
    *,
    reset_failed: bool = True,
) -> int:
    """Перезаписать frame.image_prompt из project.xlsx (ручная замена файла).

    xlsx — источник истины: без фильтра is_skippable на этапе записи в БД.
    """
    from app.models import FrameStatus

    raw = read_image_prompts_from_project_xlsx(xlsx_path)
    mapped = _map_prompts_to_frame_numbers(frames, raw)
    if not mapped:
        return 0
    by_num = {f.number: f for f in frames}
    changed = 0
    for num, text in mapped.items():
        fr = by_num.get(num)
        if fr is None:
            continue
        if fr.image_prompt != text:
            fr.image_prompt = text
            changed += 1
        if reset_failed and fr.status is FrameStatus.failed:
            attrs = dict(fr.attrs or {})
            if attrs.pop("fail_reason", None) is not None:
                fr.attrs = attrs
            fr.status = FrameStatus.image_prompt_ready
            changed += 1
    return changed


def _cell_float(ws, row: int, col: int) -> float | None:
    v = ws.cell(row=row, column=col).value
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _read_voiceover_blocks(wb) -> list[str]:
    ws = _resolve_plan_sheet(wb)
    if ws is None:
        return []
    out: list[str] = []
    for col in range(3, ws.max_column + 1):
        s = _cell_text(ws, ROW_VOICEOVER_V8, col)
        if s is None:
            continue
        out.append(s)
    return out


def read_v8_active_frame_count(xlsx_path: Path) -> int:
    """Сколько кадров в v8-xlsx (колонки с непустым voiceover на листе «план»)."""
    from openpyxl import load_workbook

    if not xlsx_path.exists():
        return 0
    wb = load_workbook(filename=str(xlsx_path), data_only=True)
    try:
        return len(_read_voiceover_blocks(wb))
    finally:
        wb.close()


def read_v8_image_prompts_from_path(xlsx_path: Path) -> dict[int, str]:
    """Номер кадра (1..N) → image_prompt с листа «план», строка R45, колонки C..N."""
    return read_image_prompts_from_project_xlsx(xlsx_path)


def read_v8_voiceovers_from_path(xlsx_path: Path) -> dict[int, str]:
    """Номер кадра → voiceover из R49 (если заполнен)."""
    from openpyxl import load_workbook

    if not xlsx_path.exists():
        return {}
    wb = load_workbook(filename=str(xlsx_path), data_only=True)
    try:
        ws = _resolve_plan_sheet(wb)
        if ws is None:
            return {}
        out: dict[int, str] = {}
        for col in range(3, ws.max_column + 1):
            voice = _cell_text(ws, ROW_VOICEOVER_V8, col)
            if voice:
                out[col - 2] = voice
        return out
    finally:
        wb.close()


@dataclass
class ImageStepBootstrapResult:
    prompts_in_xlsx: int = 0
    shot2_in_xlsx: int = 0
    frames_created: list[int] = field(default_factory=list)
    frames_prompt_updated: list[int] = field(default_factory=list)
    frames_shot2_updated: list[int] = field(default_factory=list)
    frames_status_reset: list[int] = field(default_factory=list)

    @property
    def touched(self) -> list[int]:
        return sorted(
            set(self.frames_created)
            | set(self.frames_prompt_updated)
            | set(self.frames_shot2_updated)
            | set(self.frames_status_reset)
        )


async def bootstrap_frames_for_image_step(
    session: AsyncSession,
    project: Project,
    xlsx_path: Path | None = None,
    *,
    force_prompts_from_xlsx: bool = True,
) -> ImageStepBootstrapResult:
    """Ручной импорт / шаг «Картинки»: xlsx R45/R46 + диск scenes/ — источник истины.

    - создаёт Frame в БД, если в xlsx есть промт (R45 и/или R46), а записи нет;
    - перезаписывает image_prompt из R45 и shot_02 из R46 (БД не важна);
    - сбрасывает failed/planned → image_prompt_ready, если PNG на диске нет.
    """
    from app.services.plan_shot2 import (
        SHOT2_PROMPT_ATTR,
        SHOT2_STATUS_ATTR,
        disk_has_shot2_image,
        read_shot2_columns,
    )
    from app.services.scan_frames import disk_has_valid_frame_image

    path = xlsx_path or (project.data_dir / "project.xlsx")
    result = ImageStepBootstrapResult()
    if not path.is_file():
        logger.warning(
            "[#{}] bootstrap_frames_for_image_step: нет файла {}",
            project.id,
            path,
        )
        return result

    scan = describe_image_prompts_xlsx_scan(path)
    logger.info("[#{}] bootstrap_frames_for_image_step: {}", project.id, scan)

    prompts_shot1 = read_image_prompts_from_project_xlsx(path)
    shot2_map = read_shot2_columns(path)
    shot2_nums = {n for n, info in shot2_map.items() if info.has_shot2}
    frame_nums = sorted(set(prompts_shot1.keys()) | shot2_nums)
    result.prompts_in_xlsx = len(prompts_shot1)
    result.shot2_in_xlsx = len(shot2_nums)
    if not frame_nums:
        logger.warning(
            "[#{}] bootstrap_frames_for_image_step: в листе «план» нет промтов "
            "R45/R46 — {}",
            project.id,
            scan,
        )
        return result

    voiceovers = read_v8_voiceovers_from_path(path)
    scenes_dir = project.data_dir / "scenes"
    rows = (
        await session.execute(select(Frame).where(Frame.project_id == project.id))
    ).scalars().all()
    by_number = {f.number: f for f in rows}

    for num in frame_nums:
        prompt = prompts_shot1.get(num, "")
        fr = by_number.get(num)
        if fr is None:
            fr = Frame(
                project_id=project.id,
                number=num,
                voiceover_text=voiceovers.get(num) or f"Кадр {num}",
                image_prompt=prompt or None,
                status=FrameStatus.image_prompt_ready,
            )
            session.add(fr)
            by_number[num] = fr
            result.frames_created.append(num)
            if prompt:
                result.frames_prompt_updated.append(num)
            continue

        if force_prompts_from_xlsx and prompt:
            fr.image_prompt = prompt
            if num not in result.frames_prompt_updated:
                result.frames_prompt_updated.append(num)
        elif prompt and not (fr.image_prompt or "").strip():
            fr.image_prompt = prompt
            result.frames_prompt_updated.append(num)
        vo = voiceovers.get(num)
        if vo and not (fr.voiceover_text or "").strip():
            fr.voiceover_text = vo

        if not disk_has_valid_frame_image(scenes_dir, num):
            attrs = dict(fr.attrs or {})
            if attrs.pop("fail_reason", None) is not None:
                fr.attrs = attrs
            if fr.status is not FrameStatus.image_approved:
                if fr.status is not FrameStatus.image_prompt_ready:
                    fr.status = FrameStatus.image_prompt_ready
                    result.frames_status_reset.append(num)

    for num, info in shot2_map.items():
        if not info.has_shot2:
            continue
        fr = by_number.get(num)
        if fr is None:
            continue
        attrs = dict(fr.attrs or {})
        attrs[SHOT2_PROMPT_ATTR] = info.prompt
        if disk_has_shot2_image(scenes_dir, num):
            attrs[SHOT2_STATUS_ATTR] = "image_generated"
        else:
            attrs[SHOT2_STATUS_ATTR] = "image_prompt_ready"
        fr.attrs = attrs
        if num not in result.frames_shot2_updated:
            result.frames_shot2_updated.append(num)

    if result.touched:
        await session.flush()
        logger.info(
            "[#{}] bootstrap_frames_for_image_step: R45={} R46={} created={} "
            "shot1={} shot2={} reset_status={}",
            project.id,
            result.prompts_in_xlsx,
            result.shot2_in_xlsx,
            result.frames_created,
            result.frames_prompt_updated,
            result.frames_shot2_updated,
            result.frames_status_reset,
        )
    return result


async def apply_v8_image_prompts_from_xlsx(
    session: AsyncSession,
    project: Project,
    xlsx_path: Path,
) -> list[int]:
    """Подтянуть image_prompt из v8-xlsx в Frame по номеру кадра."""
    boot = await bootstrap_frames_for_image_step(
        session, project, xlsx_path, force_prompts_from_xlsx=True
    )
    return boot.touched


def _read_frame_fields(wb) -> list[dict[str, Any]]:
    """Читает поля кадров с листа «план» v8: image_prompt (R45),
    animation_prompt (R48), voiceover (R49), duration (R50). Возвращает
    список по фреймам (порядок = порядок колонок 3..N, в которых есть
    voiceover). Каждый элемент — dict с ключами
    {image_prompt, animation_prompt, voiceover_text, duration_seconds}.

    Кадр считается «существующим», если в колонке непустой voiceover —
    остальное опционально. Это согласовано с _read_voiceover_blocks.
    """
    ws = _resolve_plan_sheet(wb)
    if ws is None:
        return []
    out: list[dict[str, Any]] = []
    for col in range(3, ws.max_column + 1):
        voice = _cell_text(ws, ROW_VOICEOVER_V8, col)
        if voice is None:
            continue
        out.append({
            "voiceover_text": voice,
            "image_prompt": _cell_text(ws, ROW_IMAGE_PROMPT_V8, col),
            "animation_prompt": _cell_text(ws, ROW_VIDEO_PROMPT_V8, col),
            "duration_seconds": _cell_float(ws, ROW_DURATION_V8, col),
        })
    return out


async def import_v8_xlsx(
    session: AsyncSession,
    project: Project,
    xlsx_path: Path,
    *,
    keep_fields: bool = True,
    update_frames_voiceover: bool = False,
) -> dict[str, Any]:
    """Подтягиваем v8-xlsx в БД для проекта.

    `keep_fields=True` (дефолт) — НЕ перезаписываем непустые
    general_plan/script_text, только заполняем пустые. Это безопасный
    режим для бэкфилла на старте.

    `keep_fields=False` — перезаписываем (используется после xlsx-flow
    шагов плана/разбивки, когда юзер только что прислал свежий xlsx).

    `update_frames_voiceover` — если True, обновляем voiceover_text у
    существующих Frame'ов (для xlsx-flow шага 3). Иначе только создаём
    недостающие.
    """
    from openpyxl import load_workbook

    summary: dict[str, Any] = {
        "project_fields_changed": [],
        "frames_created": [],
        "frames_updated": [],
    }

    if not xlsx_path.exists():
        return {"error": f"файл не найден: {xlsx_path}"}

    try:
        wb = load_workbook(filename=str(xlsx_path), data_only=True)
    except Exception as e:  # noqa: BLE001
        return {"error": f"openpyxl: {e}"}

    # --- general_plan ---
    new_plan = _read_general_plan(wb)
    if new_plan and is_meaningful_general_plan(new_plan):
        if keep_fields:
            if not project.general_plan:
                project.general_plan = new_plan
                summary["project_fields_changed"].append("general_plan")
                logger.info(
                    "[#{}] xlsx-v8→DB: general_plan заполнен ({} симв)",
                    project.id, len(new_plan),
                )
        else:
            if project.general_plan != new_plan:
                project.general_plan = new_plan
                summary["project_fields_changed"].append("general_plan")
                logger.info(
                    "[#{}] xlsx-v8→DB: general_plan обновлён ({} симв)",
                    project.id, len(new_plan),
                )

    # --- script_text + frames из voiceover-блоков ---
    blocks = _read_voiceover_blocks(wb)
    if blocks:
        new_script = " ".join(blocks)
        if not is_ui_locked(project, "script_text"):
            if keep_fields:
                if not project.script_text:
                    project.script_text = new_script
                    summary["project_fields_changed"].append("script_text")
                    logger.info(
                        "[#{}] xlsx-v8→DB: script_text заполнен из блоков "
                        "({} симв, {} блоков)",
                        project.id, len(new_script), len(blocks),
                    )
            else:
                if project.script_text != new_script:
                    project.script_text = new_script
                    summary["project_fields_changed"].append("script_text")
                    logger.info(
                        "[#{}] xlsx-v8→DB: script_text обновлён ({} симв, "
                        "{} блоков)",
                        project.id, len(new_script), len(blocks),
                    )
        else:
            logger.debug(
                "[#{}] xlsx-v8→DB: script_text пропущен (заблокирован UI)",
                project.id,
            )

        # Frame'ы — создаём недостающие, прицепляем поля из v8 (image_prompt,
        # animation_prompt, duration). Это единственный путь для v8-проектов
        # подтянуть промты, заполненные ChatGPT-ом через enrich-слоты, в БД —
        # старый xlsx_sync (лист «Кадры», R29) на v8-файле молча no-op.
        existing = (
            await session.execute(
                select(Frame)
                .where(Frame.project_id == project.id)
                .order_by(Frame.number)
            )
        ).scalars().all()
        by_number = {f.number: f for f in existing}

        # Поля по кадрам — image_prompt, animation_prompt, voiceover, duration.
        # Длины списков совпадают: оба фильтруют по непустому voiceover в
        # колонке (см. _read_frame_fields / _read_voiceover_blocks).
        frame_fields = _read_frame_fields(wb)
        fallback_durations = _distribute_durations(blocks)
        t = 0.0
        prompts_synced: list[int] = []
        for i, (cell, fields) in enumerate(
            zip(blocks, frame_fields, strict=True), start=1
        ):
            dur = fields.get("duration_seconds") or fallback_durations[i - 1]
            start_ts = t
            end_ts = t + dur
            fr = by_number.get(i)
            if fr is None:
                session.add(
                    Frame(
                        project_id=project.id,
                        number=i,
                        voiceover_text=cell,
                        image_prompt=fields.get("image_prompt"),
                        animation_prompt=fields.get("animation_prompt"),
                        start_ts=start_ts,
                        end_ts=end_ts,
                        duration_seconds=dur,
                    )
                )
                summary["frames_created"].append(i)
                if fields.get("image_prompt") or fields.get("animation_prompt"):
                    prompts_synced.append(i)
            else:
                changed = False
                if update_frames_voiceover and fr.voiceover_text != cell:
                    fr.voiceover_text = cell
                    changed = True
                # ROOT FIX: подтягиваем image_prompt / animation_prompt из v8.
                # Перезаписываем только когда в xlsx есть непустое значение и
                # оно отличается от текущего — чтобы случайно очищенная ячейка
                # не стёрла GPT-промт в БД.
                new_imgp = fields.get("image_prompt")
                if new_imgp and new_imgp != fr.image_prompt:
                    fr.image_prompt = new_imgp
                    changed = True
                    prompts_synced.append(i)
                new_animp = fields.get("animation_prompt")
                if new_animp and new_animp != fr.animation_prompt:
                    fr.animation_prompt = new_animp
                    changed = True
                new_dur = fields.get("duration_seconds")
                if new_dur is not None and abs((fr.duration_seconds or 0.0) - new_dur) > 0.01:
                    fr.duration_seconds = new_dur
                    changed = True
                if changed and i not in summary["frames_updated"]:
                    summary["frames_updated"].append(i)
            t = end_ts

        if prompts_synced:
            summary["prompts_synced"] = prompts_synced
            logger.info(
                "[#{}] xlsx-v8→DB: подтянуты image/anim prompts для кадров {}",
                project.id, prompts_synced,
            )
        if summary["frames_created"]:
            logger.info(
                "[#{}] xlsx-v8→DB: создано {} Frame'ов",
                project.id, len(summary["frames_created"]),
            )
        if summary["frames_updated"]:
            logger.info(
                "[#{}] xlsx-v8→DB: обновлено {} Frame'ов (v8-поля)",
                project.id, len(summary["frames_updated"]),
            )

    await session.flush()
    return summary
