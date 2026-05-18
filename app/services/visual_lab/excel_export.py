"""Cumulative ``scores.xlsx`` exporter for the visual lab.

Every iteration the runner calls ``rebuild_excel(storage)`` and we
overwrite ``data/test_prompts/<slug>/scores.xlsx`` with the current full
history. The file is then re-uploaded to ChatGPT alongside each new
phase so GPT has a single, machine-readable view of all iterations.

Columns (in order):
    iter, timestamp, weighted_score, verdict, phase,
    notes, prompt, prompt_len,
    <criterion_id_1>, <criterion_id_2>, ..., <criterion_id_20>,
    delta_weighted_from_prev

Plus a second sheet ``Reference`` with reference image scores, and a
third sheet ``KnowledgeBase`` with word_effects flattened.
"""

from __future__ import annotations

from typing import TYPE_CHECKING

from loguru import logger

from app.services.visual_lab.criteria import CRITERION_IDS

if TYPE_CHECKING:
    from app.services.visual_lab.storage import LabStorage


_BASE_COLUMNS = (
    "iter",
    "timestamp",
    "weighted_score",
    "verdict",
    "phase",
    "notes",
    "prompt_len",
    "prompt",
)
_DELTA_COL = "delta_weighted_from_prev"

ALL_COLUMNS: tuple[str, ...] = (
    *_BASE_COLUMNS,
    *CRITERION_IDS,
    _DELTA_COL,
)


def rebuild_excel(storage: LabStorage) -> None:
    """Rewrite ``scores.xlsx`` from the on-disk iter docs / references / KB.

    No-op if ``openpyxl`` is unavailable (logs a warning) so tests on a
    minimal env don't crash.
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        logger.warning(
            "visual_lab.excel_export: openpyxl not installed, skipping xlsx"
        )
        return

    wb = Workbook()

    # Sheet 1: iterations.
    ws = wb.active
    ws.title = "Iterations"
    ws.append(list(ALL_COLUMNS))

    prev_weighted: float | None = None
    for iter_doc in storage.load_all_iters():
        scores = iter_doc.analysis.scores if iter_doc.analysis else {}
        row = [
            iter_doc.iter,
            iter_doc.timestamp,
            round(iter_doc.weighted_score, 3),
            iter_doc.verdict,
            iter_doc.phase,
            iter_doc.notes,
            len(iter_doc.prompt.text or ""),
            (iter_doc.prompt.text or "")[:32760],  # Excel cell soft cap
        ]
        for cid in CRITERION_IDS:
            row.append(scores.get(cid, ""))
        if prev_weighted is None or iter_doc.weighted_score == 0:
            row.append("")
        else:
            row.append(round(iter_doc.weighted_score - prev_weighted, 3))
        if iter_doc.weighted_score > 0:
            prev_weighted = iter_doc.weighted_score
        ws.append(row)

    # Header style + freeze.
    ws.freeze_panes = "A2"
    for col_idx in range(1, len(ALL_COLUMNS) + 1):
        ws.cell(row=1, column=col_idx).style = "Pandas"  # bold-ish default

    # Sheet 2: reference scores.
    ws_ref = wb.create_sheet("Reference")
    ws_ref.append(["file", "prompt_len", "weighted_score", *CRITERION_IDS])
    project = storage.load_project()
    if project:
        for fname in project.references:
            ref_json = storage.reference_dir / f"{fname.rsplit('.', 1)[0]}.json"
            if not ref_json.exists():
                continue
            try:
                import json

                from app.services.visual_lab.models import ReferenceImage

                ref = ReferenceImage.model_validate(
                    json.loads(ref_json.read_text(encoding="utf-8"))
                )
            except Exception as e:  # noqa: BLE001
                logger.warning(
                    "visual_lab.excel_export: bad reference {}: {}",
                    ref_json,
                    e,
                )
                continue
            row = [
                ref.file,
                len(ref.prompt or ""),
                round(ref.weighted_score, 3),
            ]
            for cid in CRITERION_IDS:
                row.append(ref.scores.get(cid, ""))
            ws_ref.append(row)

    # Sheet 3: knowledge base flattened.
    ws_kb = wb.create_sheet("KnowledgeBase")
    ws_kb.append(
        [
            "word",
            "tested",
            "stability",
            "avg_weighted_delta",
            "conflicts_with",
            "synergizes_with",
            "top_criteria_delta",
        ]
    )
    kb = storage.load_knowledge()
    for word, effect in kb.word_effects.items():
        top = sorted(
            effect.avg_delta.items(), key=lambda x: abs(x[1]), reverse=True
        )[:5]
        top_str = ", ".join(f"{k}:{v:+.2f}" for k, v in top)
        ws_kb.append(
            [
                word,
                effect.tested,
                effect.stability,
                round(effect.avg_weighted_delta, 3),
                ", ".join(effect.conflicts_with),
                ", ".join(effect.synergizes_with),
                top_str,
            ]
        )

    wb.save(storage.excel_path)
    logger.info(
        "visual_lab.excel_export: wrote {} with {} iterations",
        storage.excel_path,
        ws.max_row - 1,
    )
