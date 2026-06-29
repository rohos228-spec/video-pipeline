"""Запись полей кадра на лист «план» (v8-xlsx).

Колонки кадров: 3..N (кадр 1 → col 3). Строки см. `xlsx_v8_import`.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import TYPE_CHECKING

from loguru import logger
from openpyxl import load_workbook
from sqlalchemy import select

from app.models import Frame, Project
from app.services.xlsx_v8_import import (
    ROW_IMAGE_PROMPT_V8,
    ROW_TIMECODE_V8,
    ROW_VIDEO_PROMPT_V8,
    ROW_VOICEOVER_V8,
    _cell_text,
    _resolve_plan_sheet,
)

SHEET_PLAN_V8 = "план"
from app.storage.project_sheet import _file_lock

if TYPE_CHECKING:
    from sqlalchemy.ext.asyncio import AsyncSession


def _normalize_timestamp_label(text: str) -> str:
    raw = (text or "").strip().translate(str.maketrans("–—−", "---"))
    if not raw:
        return ""
    return re.sub(r"\s*-\s*", "-", raw)


def _cell_timecode_text(ws, row: int, col: int) -> str:
    """Ячейка таймкода: строка, datetime/time из Excel, unicode-тире."""
    v = ws.cell(row=row, column=col).value
    if v is None:
        return ""
    if isinstance(v, str):
        return _normalize_timestamp_label(v)
    from datetime import datetime, time, timedelta

    if isinstance(v, datetime):
        sec = v.hour * 3600 + v.minute * 60 + v.second + v.microsecond / 1_000_000
        minutes = int(sec // 60)
        return f"{minutes}:{sec - minutes * 60:05.2f}"
    if isinstance(v, time):
        sec = v.hour * 3600 + v.minute * 60 + v.second + v.microsecond / 1_000_000
        minutes = int(sec // 60)
        return f"{minutes}:{sec - minutes * 60:05.2f}"
    if isinstance(v, timedelta):
        sec = max(0.0, v.total_seconds())
        minutes = int(sec // 60)
        return f"{minutes}:{sec - minutes * 60:05.2f}"
    if isinstance(v, (int, float)):
        # Excel time serial (fraction of day) or plain seconds.
        if 0 < float(v) < 1:
            sec = float(v) * 86400.0
        else:
            sec = float(v)
        minutes = int(sec // 60)
        return f"{minutes}:{sec - minutes * 60:05.2f}"
    return _normalize_timestamp_label(str(v))


def _read_r15_label(ws_raw, ws_values, row: int, col: int) -> str:
    """Строка R15: calculated value, literal, time serial."""
    if ws_values is not None:
        text = _cell_timecode_text(ws_values, row, col)
        if text:
            return text
    if ws_raw is not None:
        raw = ws_raw.cell(row=row, column=col).value
        if isinstance(raw, str):
            s = raw.strip()
            if s and not s.startswith("="):
                return _normalize_timestamp_label(s)
        text = _cell_timecode_text(ws_raw, row, col)
        if text:
            return text
    return ""


def read_plan_timestamps_cells(
    project: Project,
    frame_numbers: list[int],
) -> tuple[list[tuple[int, str]], int]:
    """Таймкод кадра — строка 15, колонка N → col N+2 (шаблон v8)."""
    ts_row = ROW_TIMECODE_V8
    if not frame_numbers:
        return [], ts_row
    path = project.data_dir / "project.xlsx"
    if not path.exists():
        return [(n, "") for n in frame_numbers], ts_row

    out: dict[int, str] = dict.fromkeys(frame_numbers, "")
    try:
        with _file_lock(path):
            wb_values = _load_plan_workbook(path, data_only=True)
            wb_raw = _load_plan_workbook(path, data_only=False)
            try:
                ws_values = _resolve_plan_sheet(wb_values)
                ws_raw = _resolve_plan_sheet(wb_raw)
                ws_primary = ws_values or ws_raw
                if ws_primary is not None:
                    col_map = voiceover_frame_columns(ws_primary)
                    for frame_number in frame_numbers:
                        col = _timestamp_column(frame_number, col_map)
                        out[frame_number] = _read_r15_label(
                            ws_raw, ws_values, ts_row, col
                        )
            finally:
                wb_values.close()
                wb_raw.close()
    except Exception as e:  # noqa: BLE001
        logger.warning(
            "[#{}] read_plan_timestamps_cells failed ({}): {}",
            project.id,
            path,
            e,
        )
    return [(n, out[n]) for n in frame_numbers], ts_row


def scan_r15_frame_numbers(project: Project) -> list[int]:
    """Все кадры с непустой меткой в R15 (col 3 = кадр 1, col = frame+2)."""
    path = project.data_dir / "project.xlsx"
    if not path.is_file():
        return []
    nums: list[int] = []
    try:
        with _file_lock(path):
            wb_values = _load_plan_workbook(path, data_only=True)
            wb_raw = _load_plan_workbook(path, data_only=False)
            try:
                ws_values = _resolve_plan_sheet(wb_values)
                ws_raw = _resolve_plan_sheet(wb_raw)
                ws_primary = ws_values or ws_raw
                if ws_primary is None:
                    return []
                col_map = voiceover_frame_columns(ws_primary)
                if col_map:
                    for frame_num in sorted(col_map):
                        col = col_map[frame_num]
                        label = _read_r15_label(ws_raw, ws_values, ROW_TIMECODE_V8, col)
                        if (label or "").strip():
                            nums.append(frame_num)
                    if nums:
                        return nums
                max_col = _scan_row_content_columns(ws_primary, ROW_TIMECODE_V8)
                for col in range(3, max_col + 1):
                    label = _read_r15_label(ws_raw, ws_values, ROW_TIMECODE_V8, col)
                    if (label or "").strip():
                        nums.append(col - 2)
            finally:
                wb_values.close()
                wb_raw.close()
    except Exception as exc:  # noqa: BLE001
        logger.warning("[#{}] scan_r15_frame_numbers: {}", project.id, exc)
    return nums


def plan_frame_column(frame_number: int) -> int:
    """Кадр N (1-based) → колонка на листе «план» (если без пропусков в R49)."""
    return frame_number + 2


def _load_plan_workbook(path: Path, *, data_only: bool):
    """Обычный режим openpyxl — read_only ломает max_column и чтение R49."""
    return load_workbook(path, data_only=data_only, read_only=False)


def _scan_row_content_columns(ws, row: int, *, start_col: int = 3, limit: int = 400) -> int:
    """Последняя колонка с данными в строке (fallback если max_column маленький)."""
    reported = int(getattr(ws, "max_column", None) or start_col)
    last = start_col
    empty_streak = 0
    for col in range(start_col, start_col + limit):
        if _cell_text(ws, row, col):
            last = col
            empty_streak = 0
        else:
            empty_streak += 1
            if empty_streak >= 24 and last > start_col:
                break
    return max(reported, last)


def voiceover_frame_columns(ws) -> dict[int, int]:
    """Номер кадра → колонка xlsx (порядок непустых ячеек R49, как в import_v8)."""
    mapping: dict[int, int] = {}
    num = 1
    max_col = _scan_row_content_columns(ws, ROW_VOICEOVER_V8)
    for col in range(3, max_col + 1):
        if _cell_text(ws, ROW_VOICEOVER_V8, col):
            mapping[num] = col
            num += 1
    return mapping


def _timestamp_column(frame_number: int, col_map: dict[int, int]) -> int:
    return col_map.get(frame_number, plan_frame_column(frame_number))


def read_plan_voiceover(project: Project, frame_number: int) -> str | None:
    """Закадровый текст кадра — строка 49 листа «план» (v8)."""
    cells = read_plan_voiceover_cells(project, [frame_number])
    text = cells[0][1] if cells else ""
    return text or None


def read_plan_animation_prompt_cells(
    project: Project,
    frame_numbers: list[int],
) -> list[tuple[int, str]]:
    """Промт анимации кадра — строка 48 листа «план» (v8)."""
    if not frame_numbers:
        return []
    path = project.data_dir / "project.xlsx"
    if not path.exists():
        return [(n, "") for n in frame_numbers]

    out: dict[int, str] = dict.fromkeys(frame_numbers, "")
    try:
        with _file_lock(path):
            wb = _load_plan_workbook(path, data_only=True)
            ws = _resolve_plan_sheet(wb)
            if ws is not None:
                for frame_number in frame_numbers:
                    col = plan_frame_column(frame_number)
                    text = _cell_text(ws, ROW_VIDEO_PROMPT_V8, col)
                    out[frame_number] = (text or "").strip()
            wb.close()
    except Exception as e:  # noqa: BLE001
        logger.warning(
            "[#{}] read_plan_animation_prompt_cells failed: {}",
            project.id,
            e,
        )
    return [(n, out[n]) for n in frame_numbers]


def write_plan_timestamps(
    project: Project,
    ranges: list[tuple[int, str]],
) -> int:
    """Записать таймкоды в строку 15 листа «план». Возвращает число ячеек."""
    path = project.data_dir / "project.xlsx"
    if not path.exists() or not ranges:
        return 0
    written = 0
    try:
        with _file_lock(path):
            wb = load_workbook(path)
            ws = _resolve_plan_sheet(wb)
            if ws is None:
                wb.close()
                return 0
            if not (_cell_text(ws, ROW_TIMECODE_V8, 1) or "").strip():
                ws.cell(row=ROW_TIMECODE_V8, column=1, value="таймкод M:SS.ss")
            col_map = voiceover_frame_columns(ws)
            for frame_number, label in ranges:
                label = (label or "").strip()
                if not label:
                    continue
                col = _timestamp_column(frame_number, col_map)
                ws.cell(row=ROW_TIMECODE_V8, column=col, value=label)
                written += 1
            wb.save(path)
            wb.close()
    except Exception as e:  # noqa: BLE001
        logger.warning("[#{}] write_plan_timestamps failed: {}", project.id, e)
        return 0
    return written


def read_plan_voiceover_cells(
    project: Project,
    frame_numbers: list[int],
) -> list[tuple[int, str]]:
    """Текст каждого кадра: строка 49 листа «план», одна колонка = одно видео."""
    if not frame_numbers:
        return []
    path = project.data_dir / "project.xlsx"
    if not path.exists():
        return [(n, "") for n in frame_numbers]

    out: dict[int, str] = dict.fromkeys(frame_numbers, "")
    try:
        with _file_lock(path):
            wb = _load_plan_workbook(path, data_only=True)
            ws = _resolve_plan_sheet(wb)
            if ws is not None:
                col_map = voiceover_frame_columns(ws)
                for frame_number in frame_numbers:
                    col = _timestamp_column(frame_number, col_map)
                    text = _cell_text(ws, ROW_VOICEOVER_V8, col)
                    out[frame_number] = (text or "").strip()
            wb.close()
    except Exception as e:  # noqa: BLE001
        logger.warning(
            "[#{}] read_plan_voiceover_cells failed: {}",
            project.id,
            e,
        )
    return [(n, out[n]) for n in frame_numbers]


def _read_plan_voiceover_cells_raw(project: Project, frame_numbers: list[int]) -> list[tuple[int, str]]:
    """R49 без data_only — если формулы не сохранили cached values."""
    if not frame_numbers:
        return []
    path = project.data_dir / "project.xlsx"
    if not path.is_file():
        return [(n, "") for n in frame_numbers]

    out: dict[int, str] = dict.fromkeys(frame_numbers, "")
    try:
        with _file_lock(path):
            wb = _load_plan_workbook(path, data_only=False)
            ws = _resolve_plan_sheet(wb)
            if ws is not None:
                col_map = voiceover_frame_columns(ws)
                for frame_number in frame_numbers:
                    col = _timestamp_column(frame_number, col_map)
                    text = _cell_text(ws, ROW_VOICEOVER_V8, col)
                    out[frame_number] = (text or "").strip()
            wb.close()
    except Exception as e:  # noqa: BLE001
        logger.warning(
            "[#{}] _read_plan_voiceover_cells_raw failed: {}",
            project.id,
            e,
        )
    return [(n, out[n]) for n in frame_numbers]


async def resolve_plan_voiceover_cells(
    session: AsyncSession,
    project: Project,
    frame_numbers: list[int],
) -> tuple[list[tuple[int, str]], str]:
    """Текст кадров для монтажа: xlsx R49 → raw xlsx → Frame.voiceover_text в БД."""
    cells = read_plan_voiceover_cells(project, frame_numbers)
    if any(text.strip() for _, text in cells):
        return cells, "xlsx-r49"

    raw_cells = _read_plan_voiceover_cells_raw(project, frame_numbers)
    if any(text.strip() for _, text in raw_cells):
        logger.warning(
            "[#{}] montage: R49 пуст в data_only — взято из literal/formula xlsx",
            project.id,
        )
        return raw_cells, "xlsx-r49-raw"

    rows = (
        await session.execute(
            select(Frame).where(
                Frame.project_id == project.id,
                Frame.number.in_(frame_numbers),
            )
        )
    ).scalars().all()
    by_number = {fr.number: (fr.voiceover_text or "").strip() for fr in rows}
    db_cells = [(n, by_number.get(n, "")) for n in frame_numbers]
    filled = sum(1 for _, text in db_cells if text.strip())
    if filled:
        logger.warning(
            "[#{}] montage: лист «план» R49 пуст — текст из БД ({} кадров)",
            project.id,
            filled,
        )
        return db_cells, "db-frames"

    return cells, "empty"


def write_plan_image_prompt(
    project: Project,
    frame_number: int,
    image_prompt: str,
) -> bool:
    """Промт картинки — строка 45 листа «план» (v8)."""
    path = project.data_dir / "project.xlsx"
    if not path.exists():
        return False
    col = plan_frame_column(frame_number)
    text = (image_prompt or "").strip()
    if not text:
        return False
    try:
        with _file_lock(path):
            wb = load_workbook(path)
            ws = _resolve_plan_sheet(wb)
            if ws is None:
                wb.close()
                return False
            ws.cell(row=ROW_IMAGE_PROMPT_V8, column=col, value=text)
            wb.save(path)
            wb.close()
        return True
    except Exception as e:  # noqa: BLE001
        logger.warning(
            "[#{}] write_plan_image_prompt frame {} failed: {}",
            project.id,
            frame_number,
            e,
        )
        return False


def merge_gpt_image_prompt_rows_into_project(
    project_xlsx: Path,
    gpt_xlsx: Path,
) -> tuple[int, int]:
    """Скопировать R45/R46 из ответа GPT в ``project.xlsx``, enrich не трогать.

    img_pr раньше подменял весь файл — ChatGPT часто возвращал только промты,
    и строки 2–44 (enrich) обнулялись.
    """
    from app.services.plan_shot2 import ROW_IMAGE_PROMPT_2_V8

    project_xlsx = Path(project_xlsx)
    gpt_xlsx = Path(gpt_xlsx)
    if not project_xlsx.is_file() or not gpt_xlsx.is_file():
        return 0, 0
    n45 = n46 = 0
    try:
        with _file_lock(project_xlsx):
            wb_gpt = load_workbook(filename=str(gpt_xlsx), data_only=True)
            wb_proj = load_workbook(filename=str(project_xlsx))
            try:
                ws_g = _resolve_plan_sheet(wb_gpt)
                ws_p = _resolve_plan_sheet(wb_proj)
                if ws_g is None or ws_p is None:
                    return 0, 0
                max_col = max(ws_g.max_column or 0, ws_p.max_column or 0)
                if max_col < 3:
                    return 0, 0
                for col in range(3, max_col + 1):
                    p45 = (_cell_text(ws_g, ROW_IMAGE_PROMPT_V8, col) or "").strip()
                    p46 = (
                        _cell_text(ws_g, ROW_IMAGE_PROMPT_2_V8, col) or ""
                    ).strip()
                    if p45:
                        ws_p.cell(row=ROW_IMAGE_PROMPT_V8, column=col, value=p45)
                        n45 += 1
                    if p46:
                        ws_p.cell(
                            row=ROW_IMAGE_PROMPT_2_V8, column=col, value=p46
                        )
                        n46 += 1
                if n45 or n46:
                    wb_proj.save(project_xlsx)
            finally:
                wb_gpt.close()
                wb_proj.close()
    except Exception as e:  # noqa: BLE001
        logger.warning(
            "merge_gpt_image_prompt_rows_into_project {} <- {}: {}",
            project_xlsx,
            gpt_xlsx,
            e,
        )
        return 0, 0
    return n45, n46


def write_plan_image_prompts_bulk(
    project: Project,
    prompts_by_frame: dict[int, str],
) -> int:
    """Массовая запись image_prompt в R45. Возвращает число записанных кадров."""
    if not prompts_by_frame:
        return 0
    path = project.data_dir / "project.xlsx"
    if not path.exists():
        return 0
    written = 0
    try:
        with _file_lock(path):
            wb = load_workbook(path)
            ws = _resolve_plan_sheet(wb)
            if ws is None:
                wb.close()
                return 0
            for frame_number, text in prompts_by_frame.items():
                t = (text or "").strip()
                if not t:
                    continue
                col = plan_frame_column(frame_number)
                ws.cell(row=ROW_IMAGE_PROMPT_V8, column=col, value=t)
                written += 1
            if written:
                wb.save(path)
            wb.close()
    except Exception as e:  # noqa: BLE001
        logger.warning(
            "[#{}] write_plan_image_prompts_bulk failed: {}",
            project.id,
            e,
        )
        return 0
    if written:
        logger.info(
            "[#{}] plan R{}: записано image_prompt для {} кадров",
            project.id,
            ROW_IMAGE_PROMPT_V8,
            written,
        )
    return written


def write_plan_animation_prompt(
    project: Project,
    frame_number: int,
    animation_prompt: str,
) -> bool:
    """Пишет промт анимации в строку 48 листа «план» для кадра `frame_number`."""
    path = project.data_dir / "project.xlsx"
    if not path.exists():
        logger.warning(
            "[#{}] write_plan_animation_prompt: нет {}",
            project.id,
            path,
        )
        return False
    col = plan_frame_column(frame_number)
    text = (animation_prompt or "").strip()
    if not text:
        return False
    try:
        with _file_lock(path):
            wb = load_workbook(path)
            ws = _resolve_plan_sheet(wb)
            if ws is None:
                wb.close()
                logger.warning(
                    "[#{}] write_plan_animation_prompt: лист «{}» не найден в {}",
                    project.id,
                    SHEET_PLAN_V8,
                    path.name,
                )
                return False
            ws.cell(row=ROW_VIDEO_PROMPT_V8, column=col, value=text)
            wb.save(path)
            wb.close()
        logger.info(
            "[#{}] plan R{} col {} ← animation_prompt ({} симв.)",
            project.id,
            ROW_VIDEO_PROMPT_V8,
            col,
            len(text),
        )
        return True
    except Exception as e:  # noqa: BLE001
        logger.warning(
            "[#{}] write_plan_animation_prompt frame {} failed: {}",
            project.id,
            frame_number,
            e,
        )
        return False
