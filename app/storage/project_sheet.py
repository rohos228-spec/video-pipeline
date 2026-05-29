"""xlsx-хранилище одного проекта.

Структура файла повторяет `templates/project_template.xlsx`:

  Лист «Кадры»
    R1  (заголовок: «Инфа», 1, 2, 3, ...) — номера кадров
    R2  id ролика
    R3  id изображения      — gen_id картинки (uuid v4) на текущую попытку
    R4  id видео            — gen_id видео
    R5  id сцены            — slug проекта
    R6  id типа генерации   — "image" / "video"
    R7..R14   id персонажа 1..8
    R15 id формата          — "9:16"
    R16 id разрешения фото
    R17 id разрешения видео
    R18 id нейронки фото    — "nano-banana-2"
    R19 id нейронки видео   — "veo-3-fast"
    R20 id стиля
    R21 id проверки фото
    R22 id проверки видео
    R28 логика кадра с учётом видео
    R29 промт картинки
    R30 промт видео
    R31 время видео         — duration_seconds
    R32 закадровый текст    — voiceover_text
    R33..R40  речь персонажа 1..8
    R41 количество символов

    Служебные строки добавляем после R41 (их в шаблоне нет, появятся при
    первой записи):
    R42 image_path     — путь к локальному файлу картинки
    R43 video_path     — путь к локальному файлу видео
    R44 image_url      — URL картинки на CDN outsee (raw_url)
    R45 video_url      — URL видео
    R46 frame_status   — текущий статус кадра (как в БД)
    R47 last_error     — текст последней ошибки на этом кадре (если был)
    R48 attempt        — номер последней попытки
    R49 updated_at     — UTC ISO-таймштамп последней записи

  Лист «Общий план ролика»
    Свободный текст: пара (label, value) по строкам.
"""

from __future__ import annotations

import shutil
import threading
from datetime import datetime
from pathlib import Path
from typing import Any

from loguru import logger

# Путь к шаблону относительно корня репо. CWD на проде — корень.
# Если есть v8-шаблон (новая логика, заполняется самим GPT) — берём его,
# иначе старый. Старый используется существующим кодом в orchestrator/steps,
# который пишет промты/URL в специально размеченные ячейки. v8-шаблон
# структурно другой и заполняется ChatGPT-ом, поэтому при v8 мы пропускаем
# enforce-разметку (см. _ensure_layout).
_V8_TEMPLATE = Path("templates/project_template_v8.xlsx")
_OLD_TEMPLATE = Path("templates/project_template.xlsx")
DEFAULT_TEMPLATE_PATH = _V8_TEMPLATE if _V8_TEMPLATE.exists() else _OLD_TEMPLATE

SHEET_FRAMES = "Кадры"
SHEET_GENERAL = "Общий план ролика"

# ---------- адресация строк листа «Кадры» -----------------------------------
ROW_HEADER = 1

ROW_PROJECT_ID = 2
ROW_IMAGE_GEN_ID = 3
ROW_VIDEO_GEN_ID = 4
ROW_SCENE_ID = 5
ROW_GEN_TYPE_ID = 6

ROW_CHAR_BASE = 7  # персонаж 1..8 → 7..14

ROW_FORMAT_ID = 15
ROW_PHOTO_RES_ID = 16
ROW_VIDEO_RES_ID = 17
ROW_PHOTO_NN_ID = 18
ROW_VIDEO_NN_ID = 19
ROW_STYLE_ID = 20
ROW_PHOTO_CHECK_ID = 21
ROW_VIDEO_CHECK_ID = 22

ROW_FRAME_LOGIC = 28
ROW_IMAGE_PROMPT = 29
ROW_VIDEO_PROMPT = 30
ROW_VIDEO_DURATION = 31
ROW_VOICEOVER = 32
ROW_SPEECH_BASE = 33  # речь 1..8 → 33..40
ROW_CHAR_COUNT = 41

# Служебные (наши, дописываются если в шаблоне нет)
ROW_IMAGE_PATH = 42
ROW_VIDEO_PATH = 43
ROW_IMAGE_URL = 44
ROW_VIDEO_URL = 45
ROW_FRAME_STATUS = 46
ROW_LAST_ERROR = 47
ROW_ATTEMPT = 48
ROW_UPDATED_AT = 49

# Подписи служебных строк (column A)
_SERVICE_LABELS = {
    ROW_IMAGE_PATH: "путь к картинке (файл)",
    ROW_VIDEO_PATH: "путь к видео (файл)",
    ROW_IMAGE_URL: "URL картинки",
    ROW_VIDEO_URL: "URL видео",
    ROW_FRAME_STATUS: "статус кадра",
    ROW_LAST_ERROR: "последняя ошибка",
    ROW_ATTEMPT: "номер попытки",
    ROW_UPDATED_AT: "обновлено (UTC)",
}


# Глобальная блокировка на файл — на одном проекте записывается несколько шагов
# параллельно (например, при per-frame HITL), а openpyxl не толерантен к
# конкурентной правке одного workbook.
_locks: dict[str, threading.Lock] = {}
_locks_guard = threading.Lock()


def _file_lock(path: Path) -> threading.Lock:
    key = str(path.resolve())
    with _locks_guard:
        lock = _locks.get(key)
        if lock is None:
            lock = threading.Lock()
            _locks[key] = lock
        return lock


def _now_iso() -> str:
    return datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")


def _frame_col(n: int) -> int:
    """Кадр N (1-based) сидит в колонке N+1 (column 1 — подписи)."""
    return n + 1


def _ensure_template_exists(template_path: Path) -> Path:
    """Если шаблона нет в репо — собираем минимальный совместимый шаблон сами.
    Это нужно, чтобы пайплайн не падал при свежем checkout без шаблона.
    """
    if template_path.exists():
        return template_path
    logger.warning(
        "project_sheet: шаблон {} не найден, создаю минимальный",
        template_path,
    )
    template_path.parent.mkdir(parents=True, exist_ok=True)
    from openpyxl import Workbook

    wb = Workbook()
    ws_frames = wb.active
    ws_frames.title = SHEET_FRAMES
    ws_frames.cell(row=ROW_HEADER, column=1, value="Инфа")
    labels = {
        ROW_PROJECT_ID: "id ролика",
        ROW_IMAGE_GEN_ID: "id изображения",
        ROW_VIDEO_GEN_ID: "id видео",
        ROW_SCENE_ID: "id сцены",
        ROW_GEN_TYPE_ID: "id типа генерации",
        ROW_FORMAT_ID: "id формата",
        ROW_PHOTO_RES_ID: "id разрешения фото",
        ROW_VIDEO_RES_ID: "id разрешения видео",
        ROW_PHOTO_NN_ID: "id нейронки фото",
        ROW_VIDEO_NN_ID: "id нейронки видео",
        ROW_STYLE_ID: "id стиля",
        ROW_PHOTO_CHECK_ID: "id проверки фото",
        ROW_VIDEO_CHECK_ID: "id проверки видео",
        ROW_FRAME_LOGIC: "логика кадра с учетом видео",
        ROW_IMAGE_PROMPT: "промт картинки",
        ROW_VIDEO_PROMPT: "промт видео",
        ROW_VIDEO_DURATION: "время видео",
        ROW_VOICEOVER: "закадровый текст",
        ROW_CHAR_COUNT: "количество символов",
    }
    for i in range(1, 9):
        labels[ROW_CHAR_BASE + i - 1] = f"id персонажа {i}"
        labels[ROW_SPEECH_BASE + i - 1] = f"речь персонажа {i}"
    for row, label in labels.items():
        ws_frames.cell(row=row, column=1, value=label)
    wb.create_sheet(SHEET_GENERAL)
    wb.save(template_path)
    return template_path


# ---------------------------------------------------------------------------


class ProjectSheet:
    """Обёртка над xlsx-файлом конкретного проекта.

    Все методы синхронные (openpyxl читает файл целиком) и идемпотентные —
    вызвать `write_frame` дважды с одними данными безопасно. Перед каждой
    записью файл открывается под блокировкой, изменяется и сохраняется.
    """

    def __init__(self, file_path: Path, *, template_path: Path | None = None) -> None:
        self.file_path = Path(file_path)
        self.template_path = (
            Path(template_path) if template_path is not None else DEFAULT_TEMPLATE_PATH
        )

    # ---- инициализация --------------------------------------------------

    def ensure_initialized(self, *, project_id: int, slug: str) -> Path:
        """Если файла ещё нет — копируем шаблон. Возвращает путь к файлу.
        Гарантирует, что в файле есть оба листа и нужная разметка (для
        старого шаблона). Для v8-шаблона разметку enforce-ить не надо —
        её формирует GPT."""
        if not self.file_path.exists():
            self.file_path.parent.mkdir(parents=True, exist_ok=True)
            tpl = _ensure_template_exists(self.template_path)
            shutil.copy(tpl, self.file_path)
            logger.info(
                "project_sheet: создан {} (копия шаблона {})",
                self.file_path,
                tpl,
            )

        with _file_lock(self.file_path):
            wb = self._open()
            if SHEET_FRAMES not in wb.sheetnames:
                # v8-шаблон: листа "Кадры" нет, пропускаем enforce-логику.
                return self.file_path
            self._ensure_layout(wb)
            ws = wb[SHEET_FRAMES]
            ws.cell(row=ROW_PROJECT_ID, column=2, value=project_id)
            ws.cell(row=ROW_SCENE_ID, column=2, value=slug)
            ws.cell(row=ROW_PHOTO_NN_ID, column=2, value="nano-banana-2")
            ws.cell(row=ROW_VIDEO_NN_ID, column=2, value="veo-3-fast")
            ws.cell(row=ROW_FORMAT_ID, column=2, value="9:16")
            ws.cell(row=ROW_UPDATED_AT, column=2, value=_now_iso())
            self._save(wb)
        return self.file_path

    def reset_from_template(self, *, project_id: int, slug: str) -> Path:
        """Копирует шаблон заново. Старый project.xlsx уходит в old/."""
        self.file_path.parent.mkdir(parents=True, exist_ok=True)
        if self.file_path.exists():
            old_dir = self.file_path.parent / "old"
            old_dir.mkdir(parents=True, exist_ok=True)
            ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            dest = old_dir / f"{ts}_{self.file_path.name}"
            shutil.move(str(self.file_path), str(dest))
            logger.info("project_sheet: backup {} -> {}", self.file_path.name, dest)
        tpl = _ensure_template_exists(self.template_path)
        shutil.copy(tpl, self.file_path)
        logger.info(
            "project_sheet: reset {} (копия шаблона {})",
            self.file_path,
            tpl,
        )
        with _file_lock(self.file_path):
            wb = self._open()
            if SHEET_FRAMES not in wb.sheetnames:
                return self.file_path
            self._ensure_layout(wb)
            ws = wb[SHEET_FRAMES]
            ws.cell(row=ROW_PROJECT_ID, column=2, value=project_id)
            ws.cell(row=ROW_SCENE_ID, column=2, value=slug)
            ws.cell(row=ROW_PHOTO_NN_ID, column=2, value="nano-banana-2")
            ws.cell(row=ROW_VIDEO_NN_ID, column=2, value="veo-3-fast")
            ws.cell(row=ROW_FORMAT_ID, column=2, value="9:16")
            ws.cell(row=ROW_UPDATED_AT, column=2, value=_now_iso())
            self._save(wb)
        return self.file_path

    # ---- общий план / сценарий -----------------------------------------

    def write_general(self, **fields: Any) -> None:
        """Пишет произвольные пары (label, value) на лист «Общий план ролика».

        Известные ключи:
          topic, slug, hero_mode, status, general_plan, script_text,
          final_video_path
        Любые другие ключи будут добавлены в конец листа как-есть.
        """
        labels_order = [
            ("topic", "Тема ролика"),
            ("slug", "Slug"),
            ("hero_mode", "Режим героя"),
            ("status", "Статус (служебный)"),
            ("image_generator", "Генератор картинок"),
            ("aspect_ratio", "Соотношение сторон"),
            ("image_resolution", "Разрешение картинки"),
            ("video_generator", "Видео-генератор"),
            ("video_resolution", "Разрешение видео"),
            ("general_plan", "Общий план (от ChatGPT)"),
            ("hero_description", "Описание героя"),
            ("script_text", "Закадровый текст (от ChatGPT)"),
            ("final_video_path", "Финальное видео (файл)"),
        ]
        with _file_lock(self.file_path):
            wb = self._open()
            if SHEET_GENERAL not in wb.sheetnames:
                # v8-шаблон: листа "Общий план ролика" нет — данные общего
                # уровня сидят в листе "Общий план" и заполняются GPT-ом,
                # поэтому write_general для v8 — no-op.
                return
            self._ensure_layout(wb)
            ws = wb[SHEET_GENERAL]
            existing: dict[str, int] = {}
            for r in range(1, ws.max_row + 1):
                v = ws.cell(row=r, column=1).value
                if isinstance(v, str) and v:
                    existing[v] = r

            def _put(label: str, value: Any) -> None:
                row = existing.get(label)
                if row is None:
                    row = (ws.max_row or 0) + 1
                    ws.cell(row=row, column=1, value=label)
                    existing[label] = row
                ws.cell(row=row, column=2, value=_stringify(value))

            for key, label in labels_order:
                if key in fields and fields[key] is not None:
                    _put(label, fields[key])
            for key, value in fields.items():
                if value is None:
                    continue
                if any(k == key for k, _ in labels_order):
                    continue
                _put(key, value)
            self._save(wb)

    # ---- per-frame ------------------------------------------------------

    def ensure_frame_columns(self, count: int) -> None:
        """Гарантирует, что в шапке листа `Кадры` есть столбцы для кадров 1..count."""
        with _file_lock(self.file_path):
            wb = self._open()
            if SHEET_FRAMES not in wb.sheetnames:
                return  # v8 — лист Кадры отсутствует, ничего не делаем.
            self._ensure_layout(wb)
            ws = wb[SHEET_FRAMES]
            for n in range(1, count + 1):
                col = _frame_col(n)
                cur = ws.cell(row=ROW_HEADER, column=col).value
                if cur in (None, ""):
                    ws.cell(row=ROW_HEADER, column=col, value=n)
            self._save(wb)

    def write_frame(
        self,
        n: int,
        *,
        voiceover_text: str | None = None,
        meaning: str | None = None,
        duration_seconds: float | None = None,
        char_count: int | None = None,
        image_prompt: str | None = None,
        animation_prompt: str | None = None,
        image_gen_id: str | None = None,
        video_gen_id: str | None = None,
        gen_type: str | None = None,
        image_path: str | None = None,
        video_path: str | None = None,
        image_url: str | None = None,
        video_url: str | None = None,
        frame_status: str | None = None,
        last_error: str | None = None,
        attempt: int | None = None,
    ) -> None:
        """Записывает в столбец кадра `n` любые непустые поля.
        None-значения пропускаются — это позволяет вызывать метод многократно
        с разными подмножествами полей по мере прохождения шагов пайплайна."""
        col = _frame_col(n)
        with _file_lock(self.file_path):
            wb = self._open()
            if SHEET_FRAMES not in wb.sheetnames:
                return  # v8 — лист Кадры отсутствует, пропускаем запись кадра.
            self._ensure_layout(wb)
            ws = wb[SHEET_FRAMES]
            if ws.cell(row=ROW_HEADER, column=col).value in (None, ""):
                ws.cell(row=ROW_HEADER, column=col, value=n)

            mapping: list[tuple[int, Any]] = [
                (ROW_VOICEOVER, voiceover_text),
                (ROW_FRAME_LOGIC, meaning),
                (ROW_VIDEO_DURATION, duration_seconds),
                (ROW_CHAR_COUNT, char_count),
                (ROW_IMAGE_PROMPT, image_prompt),
                (ROW_VIDEO_PROMPT, animation_prompt),
                (ROW_IMAGE_GEN_ID, image_gen_id),
                (ROW_VIDEO_GEN_ID, video_gen_id),
                (ROW_GEN_TYPE_ID, gen_type),
                (ROW_IMAGE_PATH, image_path),
                (ROW_VIDEO_PATH, video_path),
                (ROW_IMAGE_URL, image_url),
                (ROW_VIDEO_URL, video_url),
                (ROW_FRAME_STATUS, frame_status),
                (ROW_LAST_ERROR, last_error),
                (ROW_ATTEMPT, attempt),
            ]
            wrote = False
            for row, value in mapping:
                if value is None:
                    continue
                ws.cell(row=row, column=col, value=_stringify(value))
                wrote = True
            if wrote:
                ws.cell(row=ROW_UPDATED_AT, column=col, value=_now_iso())
            self._save(wb)

    # ---- internal --------------------------------------------------------

    def _open(self):  # type: ignore[no-untyped-def]
        from openpyxl import load_workbook

        # Если Excel открыл файл — он держит его на чтение, и load_workbook
        # обычно проходит. Но если открыто в режиме редактирования и lock
        # держится жёстко, делаем 3 ретрая по 0.2 сек.
        import time as _t

        last: Exception | None = None
        for _ in range(3):
            try:
                return load_workbook(self.file_path)
            except PermissionError as e:
                last = e
                _t.sleep(0.2)
        assert last is not None
        raise last

    def _save(self, wb: Any) -> None:
        """Сохраняем атомарно через временный файл, с ретраями на
        PermissionError (Windows + открытый Excel)."""
        import os as _os
        import time as _t

        tmp = self.file_path.with_suffix(self.file_path.suffix + ".tmp")
        last: Exception | None = None
        for _ in range(5):
            try:
                wb.save(tmp)
                _os.replace(tmp, self.file_path)
                return
            except PermissionError as e:
                last = e
                _t.sleep(0.3)
            except Exception as e:  # noqa: BLE001
                last = e
                break
        # деградация: если файл реально занят (Excel открыт у пользователя) —
        # сохраняем рядом «pending»-копию, чтобы данные не терялись. В этом
        # случае не поднимаем исключение: запись формально удалась (в pending),
        # caller узнает по предупреждению в логе.
        try:
            pending = self.file_path.with_suffix(self.file_path.suffix + ".pending")
            wb.save(pending)
            logger.warning(
                "project_sheet: {} занят (PermissionError, скорее всего открыт в "
                "Excel) — данные ушли в {}; закрой файл, при следующем шаге "
                "пайплайна будет полная запись",
                self.file_path.name,
                pending.name,
            )
            return
        except Exception:  # noqa: BLE001
            pass
        if last is not None:
            raise last

    def _ensure_layout(self, wb: Any) -> None:
        """Гарантирует, что в файле есть оба листа и наши служебные строки."""
        if SHEET_FRAMES not in wb.sheetnames:
            wb.create_sheet(SHEET_FRAMES, 0)
            wb[SHEET_FRAMES].cell(row=ROW_HEADER, column=1, value="Инфа")
        if SHEET_GENERAL not in wb.sheetnames:
            wb.create_sheet(SHEET_GENERAL)

        ws = wb[SHEET_FRAMES]
        for row, label in _SERVICE_LABELS.items():
            cur = ws.cell(row=row, column=1).value
            if cur in (None, ""):
                ws.cell(row=row, column=1, value=label)


def _stringify(value: Any) -> Any:
    """openpyxl плохо переваривает типы кроме str/int/float/bool/None — приводим."""
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


# ---------------------------------------------------------------------------


def sheet_for_slug(slug: str, *, data_dir: Path) -> ProjectSheet:
    """Конструирует ProjectSheet по slug-у проекта.

    Внимание: эта функция знает только slug → собирает старый путь
    `data/videos/<slug>/project.xlsx`. Для батч-подпроектов нужно
    использовать `for_project(project)` — он берёт `project.data_dir`,
    который корректно резолвится в `data/batches/<batch>/sub/<slug>/`.
    """
    file_path = Path(data_dir) / "videos" / slug / "project.xlsx"
    return ProjectSheet(file_path)


def for_project(project: Any) -> ProjectSheet:
    """Сахар: ProjectSheet для текущего Project.

    Использует `project.data_dir` — это автоматически даёт правильный путь
    как для одиночных (`data/videos/<slug>/`), так и для батч-подпроектов
    (`data/batches/<batch_slug>/sub/<slug>/`).

    Файл создаётся при первом обращении, если его ещё нет — это покрывает
    случай миграции старых проектов, заведённых до появления xlsx-хранилища.
    """
    file_path = project.data_dir / "project.xlsx"
    sheet = ProjectSheet(file_path)
    if not sheet.file_path.exists():
        sheet.ensure_initialized(project_id=project.id, slug=project.slug)
    return sheet
