"""xlsx-хранилище списка тем массового проекта (схема v2 — 26 колонок).

Один файл `data/batches/<slug>/topics.xlsx`. Лист «Темы» — каждая строка
карточка одного ролика с полным набором настроек, как при индивидуальной
генерации (плюс выпадающие списки промтов из `prompts/`).

Маппинг колонок (A..Z):

  A: Сценарий                    dropdown · шаг 1  (`prompts/01_plan/*.md`)
  B: Название ролика             manual   · шаг 1
  C: Стиль текста                dropdown · шаг 2  (`prompts/02_script/*.md`)
  D: Стиль анимации              dropdown · шаги 4, 6 (`prompts/05_image_prompts/*.md`)
  E: Тип хука                    зарезервировано (пока не используется)
  F: Научпоп ядро / факт         manual   · шаг 1 (краткое описание факта)
  G: Интеграция продукта         зарезервировано
  H: Генерация видео промтов     dropdown · шаг 8  (`prompts/07_animation/*.md`)
  I: hero_mode                   dropdown ("0и1".."4и5") · шаг 4
  J: hero описание               manual   · шаг 4
  K: Время ролика (сек)          manual   · шаг 1
  L: Закадр. текст (символов)    формула =K×13.5
  M: Генератор картинок          dropdown · шаг 7 (default: Nano Banana Pro)
  N: Качество картинок           dropdown 2K/4K · шаг 7
  O: Соотношение картинок        dropdown 16:9/9:16 · шаг 7
  P: Релакс картинок             dropdown ДА/НЕТ · шаг 7
  Q: Генератор видео             dropdown · шаг 9 (default: Veo 3.1 Lite)
  R: Качество видео              dropdown 720/1080 · шаг 9
  S: Соотношение видео           dropdown 16:9/9:16 · шаг 9
  T: Релакс видео                dropdown ДА/НЕТ · шаг 9
  U: Голос                       зарезервировано
  V: Музыка                      зарезервировано
  W: СЛУЖ. — slug                сервис (бот пишет)
  X: СЛУЖ. — статус              сервис (бот пишет)
  Y: СЛУЖ. — прогресс            сервис (бот пишет)
  Z: СЛУЖ. — обновлён            сервис (бот пишет)

Обязательная колонка только B (Название). Остальные карточные поля —
необязательные. В пустые ячейки бот при сохранении подставляет дефолты
(см. ROW_DEFAULTS) — поэтому юзеру достаточно вписать только B и при
необходимости менять одну-две колонки на ролик.

Источники dropdown'ов:
  * Промт-папки (A/C/D/H) — сканируются динамически через
    `app.services.prompt_library.list_prompts(step_code)` в момент
    создания/перевыпуска xlsx. Если потом в `prompts/<...>/` добавили новый
    `.md` — нужно перевыложить файл («📥 Скачать topics.xlsx» → загрузить
    обратно), чтобы Excel показал свежий список.
  * Картинки/видео-генераторы — `IMAGE_GENERATORS.label` /
    `VIDEO_GENERATORS.label` из `app.generation_options`.

Сервисные колонки W..Z окрашены в серый и заблокированы (только-чтение
для пользователя). Меняет их только бот при выгрузке.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from loguru import logger

from app.generation_options import IMAGE_GENERATORS, VIDEO_GENERATORS
from app.services.prompt_library import list_prompts

SHEET_NAME = "Темы"

# Расширенный «карточный» набор полей. Попадают в Project.meta["topic_card"]
# и в настройки проекта (см. app/services/batches.py::add_topics).
CARD_FIELDS = [
    "scenario",          # A: dropdown → prompt_overrides["plan"]
    "title",             # B: manual  → project.topic
    "script_style",      # C: dropdown → prompt_overrides["script"]
    "anim_style",        # D: dropdown → prompt_overrides["img_pr"]
    "hook_type",         # E: пока не используется
    "fact",              # F: manual  → topic_card.fact + plan-prompt
    "integration",       # G: пока не используется
    "video_prompts_gen", # H: dropdown → prompt_overrides["anim_pr"]
    "hero_combo",        # I: "NиM"  → hero_count + hero_variations
    "hero_description",  # J: manual  → hero_descriptions[0..N-1]
    "duration_sec",      # K: manual  → meta.duration_target_sec
    "voiceover_chars",   # L: формула — только для пользователя
    "image_generator",   # M: dropdown → project.image_generator
    "image_quality",     # N: dropdown → project.image_resolution
    "image_aspect",      # O: dropdown → project.aspect_ratio
    "image_relax",       # P: dropdown → project.image_relax
    "video_generator",   # Q: dropdown → project.video_generator
    "video_quality",     # R: dropdown → project.video_resolution
    "video_aspect",      # S: dropdown (отдельно от картинок не сохраняется в БД)
    "video_relax",       # T: dropdown → project.video_relax
    "voice",             # U: пока не используется
    "music",             # V: пока не используется
]

HEADERS = [
    "Сценарий",                          # A
    "Название ролика",                   # B
    "Стиль текста",                      # C
    "Стиль анимации",                    # D
    "Тип хука",                          # E
    "Научпоп ядро / факт",               # F
    "Интеграция продукта",               # G
    "Генерация видео промтов",           # H
    "hero_mode",                         # I
    "hero описание",                     # J
    "Время ролика (сек)",                # K
    "Закадр. текст (символов = K×13,5)", # L
    "Генератор картинок",                # M
    "Качество (картинки)",               # N
    "Соотношение (картинки)",            # O
    "Релакс (картинки)",                 # P
    "Генератор видео",                   # Q
    "Качество (видео)",                  # R
    "Соотношение (видео)",               # S
    "Релакс (видео)",                    # T
    "Голос",                             # U
    "Музыка",                            # V
    "⛔ СЛУЖ. — slug",                    # W
    "⛔ СЛУЖ. — статус",                  # X
    "⛔ СЛУЖ. — прогресс",                # Y
    "⛔ СЛУЖ. — обновлён",                # Z
]

# 1-based индексы сервисных колонок (W..Z = 23..26).
SERVICE_COL_INDICES = (23, 24, 25, 26)
N_COLS = len(HEADERS)
N_ROWS = 100  # сколько строк ниже шапки покрываем validation'ами

# Ширины колонок (в условных Excel-единицах).
COL_WIDTHS = [
    22, 28, 24, 22, 16, 32, 22, 28, 12, 32, 14, 22,
    22, 12, 16, 12, 22, 12, 16, 12, 14, 14, 24, 18, 16, 20,
]

# Значения, которые подставляются в новую строку «по умолчанию», если в
# xlsx ячейка пустая (или если бот выгружает строку для существующего
# подпроекта, у которого настройка не задана).
ROW_DEFAULTS: dict[str, object] = {
    "scenario":          "default",
    "script_style":      "default",
    "anim_style":        "default",
    "video_prompts_gen": "default",
    "hero_combo":        "0и1",          # без героев (hero_count=0)
    "duration_sec":      30,
    "image_generator":   "Nano Banana Pro",
    "image_quality":     "2K",
    "image_aspect":      "9:16",
    "image_relax":       "НЕТ",
    "video_generator":   "Veo 3.1 Lite",
    "video_quality":     "1080",
    "video_aspect":      "9:16",
    "video_relax":       "НЕТ",
}

# Фиксированные dropdown-варианты для колонки I (hero_mode).
HERO_COMBOS = [
    f"{heroes}и{vars_}"
    for heroes in range(5)         # 0..4 героев
    for vars_ in range(1, 6)       # 1..5 вариаций (для 0и* — формальные пары)
]

YES_NO = ["ДА", "НЕТ"]
ASPECTS = ["16:9", "9:16"]
IMG_QUALITY = ["2K", "4K"]
VIDEO_QUALITY = ["720", "1080"]


def _now_iso() -> str:
    return datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")


def _dropdown_formula(values: list[str]) -> str:
    """Сериализует список значений в формулу `"v1,v2,..."` для openpyxl
    DataValidation. Имена не должны содержать запятых и двойных кавычек.
    """
    # На всякий случай экранируем кавычки удвоением (Excel-формат).
    safe = [v.replace('"', '""') for v in values if v]
    formula = ",".join(safe)
    return f'"{formula}"'


def _prompt_dropdown(step_code: str) -> list[str]:
    """Список вариантов промтов из `prompts/<folder>/*.md` (без расширения).

    `default` всегда идёт первым (если файл существует), остальные — в
    алфавитном порядке.
    """
    names = list_prompts(step_code)
    if "default" in names:
        names = ["default"] + [n for n in names if n != "default"]
    return names


def _image_generator_labels() -> list[str]:
    return [c.label for c in IMAGE_GENERATORS]


def _video_generator_labels() -> list[str]:
    return [c.label for c in VIDEO_GENERATORS]


def _apply_data_validations(ws) -> None:
    """Добавляет dropdown'ы во все строки данных (3..N_ROWS+2)."""
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    data_first = 3
    data_last = 2 + N_ROWS  # шапка занимает 1-2, данные с 3

    def _add_dropdown(col_idx: int, values: list[str]) -> None:
        if not values:
            return
        col_letter = get_column_letter(col_idx)
        dv = DataValidation(
            type="list",
            formula1=_dropdown_formula(values),
            allow_blank=True,
            showErrorMessage=False,  # юзер вписать кастомное значение — не лагать
        )
        dv.add(f"{col_letter}{data_first}:{col_letter}{data_last}")
        ws.add_data_validation(dv)

    # A — Сценарий (шаг 1)
    _add_dropdown(1,  _prompt_dropdown("plan"))
    # C — Стиль текста (шаг 2)
    _add_dropdown(3,  _prompt_dropdown("script"))
    # D — Стиль анимации (шаг 6)
    _add_dropdown(4,  _prompt_dropdown("img_pr"))
    # H — Генерация видео промтов (шаг 8)
    _add_dropdown(8,  _prompt_dropdown("anim_pr"))
    # I — hero_mode
    _add_dropdown(9,  HERO_COMBOS)
    # M — Генератор картинок
    _add_dropdown(13, _image_generator_labels())
    # N — Качество (картинки)
    _add_dropdown(14, IMG_QUALITY)
    # O — Соотношение (картинки)
    _add_dropdown(15, ASPECTS)
    # P — Релакс (картинки)
    _add_dropdown(16, YES_NO)
    # Q — Генератор видео
    _add_dropdown(17, _video_generator_labels())
    # R — Качество (видео)
    _add_dropdown(18, VIDEO_QUALITY)
    # S — Соотношение (видео)
    _add_dropdown(19, ASPECTS)
    # T — Релакс (видео)
    _add_dropdown(20, YES_NO)


def _apply_voiceover_formulas(ws, last_data_row: int | None = None) -> None:
    """Пишет =K{row}*13.5 в колонку L для всех data-строк."""
    last = max(last_data_row or 0, 2 + N_ROWS)
    for r in range(3, last + 1):
        ws.cell(row=r, column=12, value=f"=K{r}*13.5")


def _apply_styling(ws, last_data_row: int) -> None:
    """Покраска заголовков, сервисных колонок, ширины."""
    from openpyxl.styles import Alignment, Font, PatternFill, Protection
    from openpyxl.utils import get_column_letter

    # Ширины
    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Заголовок (строка 2) — bold + центр + перенос
    header_font = Font(bold=True, size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, N_COLS + 1):
        cell = ws.cell(row=2, column=c)
        cell.font = header_font
        cell.alignment = header_align

    ws.row_dimensions[2].height = 36

    # Сервисные колонки W..Z — серый + bold red заголовок + lock
    gray = PatternFill(start_color="FFE0E0E0", end_color="FFE0E0E0", fill_type="solid")
    head_red = Font(color="FF990000", bold=True, size=10)
    for col_idx in SERVICE_COL_INDICES:
        h = ws.cell(row=2, column=col_idx)
        h.fill = gray
        h.font = head_red
        h.alignment = header_align
        for r in range(3, max(last_data_row, 2) + 1):
            c = ws.cell(row=r, column=col_idx)
            c.fill = gray
            c.protection = Protection(locked=True)


def _write_header(ws, batch_name: str) -> None:
    """Шапка: A1 — название батча (merged по всем колонкам), 2-я строка — заголовки."""
    from openpyxl.utils import get_column_letter
    last_col = get_column_letter(N_COLS)
    ws["A1"] = f"🎬 Массовый проект: {batch_name} — шаблон тем"
    ws.merge_cells(f"A1:{last_col}1")
    for col_idx, header in enumerate(HEADERS, start=1):
        ws.cell(row=2, column=col_idx, value=header)


def init_topics_xlsx(path: Path, batch_name: str) -> None:
    """Создаёт пустой xlsx с шапкой и dropdown'ами. Идемпотентно — если
    файл уже есть, не трогает (включая случай, когда у него старая
    схема). Чтобы получить файл новой схемы для старого батча — удалить
    его и перевыпустить.
    """
    if path.exists():
        return
    from openpyxl import Workbook

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    _write_header(ws, batch_name)
    _apply_data_validations(ws)
    _apply_voiceover_formulas(ws)
    _apply_styling(ws, last_data_row=2)

    wb.save(path)
    logger.info("batch topics.xlsx initialized (schema v2): {}", path)


def _row_with_defaults(row: dict) -> dict:
    """Возвращает копию `row` с подставленными дефолтами для незаданных
    полей. Только для непустых строк (есть title)."""
    out = dict(row)
    for key, default in ROW_DEFAULTS.items():
        if out.get(key) in (None, ""):
            out[key] = default
    return out


def write_subprojects_table(
    path: Path,
    rows: list[dict],
    batch_name: str,
) -> None:
    """Полностью переписывает таблицу подпроектов в xlsx.

    Каждый `rows[i]` — dict с ключами из CARD_FIELDS + service-поля
    (`slug`, `status`, `progress`).

    Для незаданных полей подставляются дефолты из ROW_DEFAULTS, чтобы
    у юзера в Excel ячейки выглядели «как из шаблона», а не пустыми.
    """
    from openpyxl import Workbook, load_workbook

    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        wb = load_workbook(path)
        if SHEET_NAME in wb.sheetnames:
            del wb[SHEET_NAME]
        ws = wb.create_sheet(SHEET_NAME, 0)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME

    _write_header(ws, batch_name)

    last_data_row = max(2 + N_ROWS, 2 + len(rows))

    for r, raw in enumerate(rows, start=3):
        row = _row_with_defaults(raw)
        ws.cell(row=r, column=1,  value=row.get("scenario"))           # A
        ws.cell(row=r, column=2,  value=row.get("title") or row.get("topic"))  # B
        ws.cell(row=r, column=3,  value=row.get("script_style"))       # C
        ws.cell(row=r, column=4,  value=row.get("anim_style"))         # D
        ws.cell(row=r, column=5,  value=row.get("hook_type"))          # E
        ws.cell(row=r, column=6,  value=row.get("fact"))               # F
        ws.cell(row=r, column=7,  value=row.get("integration"))        # G
        ws.cell(row=r, column=8,  value=row.get("video_prompts_gen"))  # H
        ws.cell(row=r, column=9,  value=row.get("hero_combo"))         # I
        ws.cell(row=r, column=10, value=row.get("hero_description"))   # J
        ws.cell(row=r, column=11, value=row.get("duration_sec"))       # K
        # L — формула, ставится отдельно (см. ниже)
        ws.cell(row=r, column=13, value=row.get("image_generator"))    # M
        ws.cell(row=r, column=14, value=row.get("image_quality"))      # N
        ws.cell(row=r, column=15, value=row.get("image_aspect"))       # O
        ws.cell(row=r, column=16, value=row.get("image_relax"))        # P
        ws.cell(row=r, column=17, value=row.get("video_generator"))    # Q
        ws.cell(row=r, column=18, value=row.get("video_quality"))      # R
        ws.cell(row=r, column=19, value=row.get("video_aspect"))       # S
        ws.cell(row=r, column=20, value=row.get("video_relax"))        # T
        ws.cell(row=r, column=21, value=row.get("voice"))              # U
        ws.cell(row=r, column=22, value=row.get("music"))              # V
        ws.cell(row=r, column=23, value=row.get("slug"))               # W
        ws.cell(row=r, column=24, value=row.get("status"))             # X
        ws.cell(row=r, column=25, value=row.get("progress"))           # Y
        ws.cell(row=r, column=26, value=_now_iso())                    # Z

    _apply_data_validations(ws)
    _apply_voiceover_formulas(ws, last_data_row=last_data_row)
    _apply_styling(ws, last_data_row=last_data_row)

    wb.save(path)
    logger.info("batch topics.xlsx written: {} rows={}", path, len(rows))


def _s(v) -> str | None:
    """Чистая строка или None."""
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _yes_no(v) -> bool | None:
    """Парсит «ДА»/«НЕТ» (а также True/False/1/0/y/n) → bool. None если пусто."""
    s = _s(v)
    if s is None:
        return None
    low = s.lower()
    if low in ("да", "yes", "true", "1", "y", "+"):
        return True
    if low in ("нет", "no", "false", "0", "n", "-"):
        return False
    return None


def _int(v) -> int | None:
    s = _s(v)
    if s is None:
        return None
    try:
        return int(float(s))
    except (TypeError, ValueError):
        return None


def read_topics(path: Path) -> list[dict]:
    """Читает темы из xlsx. Возвращает список словарей со всеми полями
    из CARD_FIELDS + сервисными.

    Пустые строки (без title) пропускаются.
    Поле `topic` дублирует `title` — для обратной совместимости с местами
    в коде, которые ещё ожидают этот ключ.
    """
    if not path.exists():
        return []
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        return []
    ws = wb[SHEET_NAME]

    out: list[dict] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row:
            continue
        padded = (tuple(row) + (None,) * N_COLS)[:N_COLS]
        (
            scenario, title, script_style, anim_style, hook_type, fact,
            integration, video_prompts_gen, hero_combo, hero_description,
            duration_sec, voiceover_chars,
            image_generator, image_quality, image_aspect, image_relax,
            video_generator, video_quality, video_aspect, video_relax,
            voice, music,
            slug, status, progress, updated_at,
        ) = padded

        title_clean = _s(title) or ""
        if not title_clean:
            continue

        out.append({
            # карточные поля
            "scenario":          _s(scenario),
            "title":             title_clean,
            "topic":             title_clean,  # legacy
            "script_style":      _s(script_style),
            "anim_style":        _s(anim_style),
            "hook_type":         _s(hook_type),
            "fact":              _s(fact),
            "integration":       _s(integration),
            "video_prompts_gen": _s(video_prompts_gen),
            "hero_combo":        _s(hero_combo),
            "hero_description":  _s(hero_description),
            "duration_sec":      _int(duration_sec),
            "voiceover_chars":   _int(voiceover_chars),
            "image_generator":   _s(image_generator),
            "image_quality":     _s(image_quality),
            "image_aspect":      _s(image_aspect),
            "image_relax":       _yes_no(image_relax),
            "video_generator":   _s(video_generator),
            "video_quality":     _s(video_quality),
            "video_aspect":      _s(video_aspect),
            "video_relax":       _yes_no(video_relax),
            "voice":             _s(voice),
            "music":             _s(music),
            # сервисные
            "slug":     _s(slug),
            "status":   _s(status),
            "progress": _s(progress),
            "updated_at": updated_at,
        })
    return out


def collect_new_topics(
    path: Path,
    known_slugs: set[str] | None = None,
) -> list[dict]:
    """Возвращает только новые темы (ещё не созданы подпроекты).

    `known_slugs` — множество реально существующих slug'ов подпроектов
    данного массового (из БД). Строка считается «уже привязанной»
    ТОЛЬКО если её slug-колонка содержит ровно один из этих slug'ов.
    Любая каша / случайный текст / название продукта в колонке W → строка
    идёт в новые темы.

    Если `known_slugs is None` — поведение как раньше: строка считается
    новой, если slug пустой. Используется в тестах.
    """
    rows = read_topics(path)
    if known_slugs is None:
        return [r for r in rows if not r.get("slug")]
    return [
        r for r in rows
        if not r.get("slug") or r["slug"] not in known_slugs
    ]
