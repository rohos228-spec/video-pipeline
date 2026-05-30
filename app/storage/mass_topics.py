"""Парсинг Excel со списком тем (любой xlsx, построчно — название видео)."""

from __future__ import annotations

from pathlib import Path

from loguru import logger

_TOPIC_HEADER_KEYS = frozenset(
    {
        "название",
        "название ролика",
        "тема",
        "topic",
        "title",
        "видео",
        "name",
        "№",
        "#",
        "id",
        "номер",
    }
)

_HEADER_SUBSTRINGS = ("тема", "topic", "title", "назван")


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_header_cell(value: str) -> bool:
    h = value.lower()
    if not h or len(h) > 40:
        return False
    if h in _TOPIC_HEADER_KEYS:
        return True
    if h in ("тема", "topic", "title", "name", "видео", "note", "примечание"):
        return True
    if any(k in h for k in ("назван",)) and len(h) <= 25:
        return True
    return False


def _is_likely_header_row(row: tuple) -> bool:
    """Первая строка — заголовок колонок, а не первая тема."""
    non_empty = [_cell_str(c) for c in row if _cell_str(c)]
    if not non_empty:
        return False
    if any(len(h) > 60 for h in non_empty):
        return False
    hits = sum(1 for h in non_empty if _is_header_cell(h))
    if hits == 0:
        return False
    # Строка «1 | Тема A» — это данные, не шапка.
    if len(non_empty) <= 2 and hits < len(non_empty):
        return False
    return hits >= max(1, len(non_empty) // 2)


def _looks_like_row_number(value: str) -> bool:
    return value.isdigit() and len(value) <= 4


def _detect_topic_col(headers: list[str]) -> int:
    preferred = (
        "название",
        "название ролика",
        "тема",
        "topic",
        "title",
        "видео",
        "name",
    )
    for idx, h in enumerate(headers):
        if h in preferred:
            return idx
    for idx, h in enumerate(headers):
        if any(k in h for k in _HEADER_SUBSTRINGS):
            return idx
    return 0


def _collect_topics_from_rows(
    rows: list[tuple],
    topic_col: int,
    *,
    skip_header: bool,
) -> list[str]:
    data_rows = rows[1:] if skip_header and len(rows) > 1 else rows
    topics: list[str] = []
    seen: set[str] = set()
    for row in data_rows:
        if not row or topic_col >= len(row):
            continue
        title = _cell_str(row[topic_col])
        if not title or title in seen:
            continue
        seen.add(title)
        topics.append(title)
    return topics


def _score_topic_column(rows: list[tuple], col: int, *, skip_header: bool) -> int:
    data_rows = rows[1:] if skip_header and len(rows) > 1 else rows
    score = 0
    for row in data_rows:
        if not row or col >= len(row):
            continue
        title = _cell_str(row[col])
        if not title or _looks_like_row_number(title):
            continue
        score += min(len(title), 80)
    return score


def _filter_row_number_column(topics: list[str]) -> list[str]:
    if len(topics) <= 1:
        return topics
    if all(_looks_like_row_number(t) for t in topics):
        return topics
    filtered = [t for t in topics if not _looks_like_row_number(t)]
    return filtered if filtered else topics


def _topics_from_sheet_rows(rows: list[tuple]) -> list[str]:
    if not rows:
        return []

    lengths = [len(r) for r in rows if r]
    ncol = max(lengths) if lengths else 0
    if ncol <= 1:
        return _collect_topics_from_rows(rows, 0, skip_header=False)

    skip_header = _is_likely_header_row(rows[0])
    headers = [_cell_str(c).lower() for c in rows[0]]
    topic_col = _detect_topic_col(headers)

    topics = _filter_row_number_column(
        _collect_topics_from_rows(rows, topic_col, skip_header=skip_header)
    )
    if topics:
        return topics

    # Fallback: ищем колонку с самым «текстовым» содержимым (часто темы в B,
    # а в A — номера строк).
    best_col = 0
    best_score = -1
    best_skip = skip_header
    for col in range(ncol):
        for skip in (True, False):
            score = _score_topic_column(rows, col, skip_header=skip)
            if score > best_score:
                best_score = score
                best_col = col
                best_skip = skip

    if best_score <= 0:
        return []

    return _filter_row_number_column(
        _collect_topics_from_rows(rows, best_col, skip_header=best_skip)
    )


def _topics_from_batch_sheet(path: Path) -> list[str] | None:
    from app.storage import batch_sheet

    try:
        rows = batch_sheet.read_topics(path)
    except Exception as exc:  # noqa: BLE001
        logger.debug("mass_topics: batch_sheet failed {} — {}", path.name, exc)
        return None
    topics: list[str] = []
    for row in rows:
        title = (row.get("title") or row.get("topic") or "").strip()
        if title:
            topics.append(title)
    return topics if topics else None


def _topics_from_workbook(path: Path) -> list[str]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_names: list[str] = []
        if wb.active is not None and wb.active.title:
            sheet_names.append(wb.active.title)
        if "Темы" in wb.sheetnames and "Темы" not in sheet_names:
            sheet_names.append("Темы")

        for name in sheet_names:
            ws = wb[name]
            rows = list(ws.iter_rows(values_only=True))
            topics = _topics_from_sheet_rows(rows)
            if topics:
                return topics
        return []
    finally:
        wb.close()


def _topics_from_first_sheet(path: Path) -> list[str]:
    return _topics_from_workbook(path)


def parse_topics_xlsx(path: Path) -> list[str]:
    """Вернёт список названий видео из xlsx (только текст темы)."""
    flex = _topics_from_first_sheet(path)
    if flex:
        return flex
    batch = _topics_from_batch_sheet(path)
    if batch:
        return batch
    return []
