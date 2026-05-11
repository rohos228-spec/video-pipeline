"""Восстанавливает данные шагов 1/2/3 проекта из файлов на диске после
потери в БД (general_plan / script_text / frames).

Что делает:
  1. Читает project.xlsx ('Общий план ролика' лист) →
       - 'Общий план (от ChatGPT)'    → project.general_plan
       - 'Закадровый текст (от ChatGPT)' / 'Сценарий (от ChatGPT)'
                                       → project.script_text
       - 'Описание героя'              → project.hero_description (если пусто)
     Перезаписывает только пустые поля. Уже заполненное не трогает.
  2. Если project.script_text всё ещё пуст — берёт текст из voiceover.txt.
  3. Если в БД нет ни одного frame этого проекта — пересоздаёт их из
     project.xlsx (лист 'Кадры'): voiceover_text / meaning / image_prompt /
     animation_prompt / duration_seconds + start_ts/end_ts.
     Если кадры уже есть — НЕ ТРОГАЕТ.
  4. Если статус проекта == failed — переводит в frames_ready, чтобы
     можно было сразу нажать '4. Hero' в /menu.
  5. hero_descriptions / hero_variations / prompt_overrides /
     gpt_text_overrides / hero-артефакты — НЕ ТРОГАЕТ.

Запуск:
    python -m recover_from_disk <project_id>

По умолчанию pid=1.
"""

from __future__ import annotations

import asyncio
import sys

from openpyxl import load_workbook
from sqlalchemy import select

from app.db import session_scope
from app.models import Frame, Project, ProjectStatus
from app.services.xlsx_sync import (
    _GENERAL_LABEL_TO_FIELD,
    _to_float,
    _to_str,
)
from app.settings import settings
from app.storage.project_sheet import (
    ROW_FRAME_LOGIC,
    ROW_HEADER,
    ROW_IMAGE_PROMPT,
    ROW_VIDEO_DURATION,
    ROW_VIDEO_PROMPT,
    ROW_VOICEOVER,
    SHEET_FRAMES,
    SHEET_GENERAL,
)


async def recover(pid: int) -> None:
    async with session_scope() as s:
        project = (
            await s.execute(select(Project).where(Project.id == pid))
        ).scalar_one_or_none()
        if project is None:
            print(f"проект #{pid} не найден")
            return

        slug = project.slug
        proj_dir = settings.data_dir / "videos" / slug
        xlsx_path = proj_dir / "project.xlsx"
        vo_path = proj_dir / "voiceover.txt"

        print(f"проект #{pid} '{project.topic}' slug={slug} status={project.status.value}")
        print(f"  xlsx: {xlsx_path} (exists={xlsx_path.exists()})")
        print(f"  voiceover.txt: {vo_path} (exists={vo_path.exists()})")

        if not xlsx_path.exists():
            print("ошибка: project.xlsx не найден — нечего восстанавливать")
            return

        wb = load_workbook(filename=str(xlsx_path), data_only=True)

        # 1) Общий план / закадровый / описание героя
        if SHEET_GENERAL in wb.sheetnames:
            ws_g = wb[SHEET_GENERAL]
            for r in range(1, (ws_g.max_row or 0) + 1):
                label = _to_str(ws_g.cell(row=r, column=1).value)
                value = _to_str(ws_g.cell(row=r, column=2).value)
                if not label or value is None:
                    continue
                field = _GENERAL_LABEL_TO_FIELD.get(label)
                if field is None:
                    continue
                cur = (getattr(project, field, None) or "").strip()
                if not cur and value:
                    setattr(project, field, value)
                    print(f"  ✓ project.{field} ← xlsx ({len(value)} симв)")

        # 2) Если script_text всё ещё пуст — берём из voiceover.txt
        if not (project.script_text or "").strip() and vo_path.exists():
            vo = vo_path.read_text(encoding="utf-8", errors="replace").strip()
            if vo:
                project.script_text = vo
                print(f"  ✓ project.script_text ← voiceover.txt ({len(vo)} симв)")

        # 3) Кадры (только если их нет в БД)
        existing = (
            await s.execute(select(Frame).where(Frame.project_id == project.id))
        ).scalars().all()
        if existing:
            print(f"  кадры уже есть в БД ({len(existing)} шт), не трогаю")
        elif SHEET_FRAMES in wb.sheetnames:
            ws_f = wb[SHEET_FRAMES]
            max_col = ws_f.max_column or 0
            t = 0.0
            created = 0
            for col in range(2, max_col + 1):
                n = ws_f.cell(row=ROW_HEADER, column=col).value
                try:
                    fnum = int(n)
                except (TypeError, ValueError):
                    continue
                voice = _to_str(ws_f.cell(row=ROW_VOICEOVER, column=col).value)
                meaning = _to_str(ws_f.cell(row=ROW_FRAME_LOGIC, column=col).value)
                imgp = _to_str(ws_f.cell(row=ROW_IMAGE_PROMPT, column=col).value)
                vidp = _to_str(ws_f.cell(row=ROW_VIDEO_PROMPT, column=col).value)
                dur = _to_float(ws_f.cell(row=ROW_VIDEO_DURATION, column=col).value)
                if not voice:
                    continue
                start_ts = t
                end_ts = t + (dur or 0.0)
                t = end_ts
                fr = Frame(
                    project_id=project.id,
                    number=fnum,
                    voiceover_text=voice,
                    meaning=meaning,
                    image_prompt=imgp,
                    animation_prompt=vidp,
                    duration_seconds=dur,
                    start_ts=start_ts,
                    end_ts=end_ts,
                )
                s.add(fr)
                created += 1
            print(f"  ✓ создано frames: {created}")
        else:
            print("  лист 'Кадры' в xlsx не найден — кадры не восстановлены")

        # 4) Статус — в frames_ready (можно сразу нажимать '4. Hero')
        if project.status == ProjectStatus.failed:
            project.status = ProjectStatus.frames_ready
            print("  ✓ status: failed → frames_ready")

        await s.flush()
        print("готово.")


if __name__ == "__main__":
    pid = int(sys.argv[1]) if len(sys.argv) > 1 else 1
    asyncio.run(recover(pid))
