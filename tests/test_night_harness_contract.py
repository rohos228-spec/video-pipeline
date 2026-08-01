"""Tests: llm_contract, excel priority pack, writeback fail-closed, harness guards."""

from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import Workbook, load_workbook

from app.services.agent_harness import HARNESS_FORBIDDEN_STEPS, verify_project_disk
from app.services.gpt_api import xlsx_to_text
from app.services.llm_contract import (
    build_api_accompany,
    validate_model_reply,
)
from app.services.xlsx_text_writeback import apply_sheet_blocks_to_xlsx


def test_llm_contract_voiceover_and_writeback() -> None:
    bad = validate_model_reply("просто текст", kind="voiceover")
    assert bad.ok is False
    good = validate_model_reply(
        "<<<VOICEOVER>>>\nhello\n<<<END>>>", kind="voiceover"
    )
    assert good.ok is True
    wb = validate_model_reply(
        "# Лист: план\n@row=49\ta\tb\n", kind="xlsx_writeback"
    )
    assert wb.ok is True
    accomp = build_api_accompany("hi", expect_xlsx_writeback=True)
    assert "# Лист:" in accomp or "TSV" in accomp or "@row=" in accomp


def test_xlsx_priority_pack_keeps_r48_under_budget(tmp_path: Path) -> None:
    p = tmp_path / "project.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "план"
    # Noise rows that would eat budget if not pinned after
    for r in range(1, 40):
        ws.cell(r, 3, value=("noise-" + ("x" * 200)) if r not in {15, 45, 48, 49} else "")
    ws.cell(15, 3, value="0:00-0:03")
    ws.cell(45, 3, value="img-prompt-keep")
    ws.cell(48, 3, value="anim-prompt-KEEP-R48")
    ws.cell(49, 3, value="voice-keep")
    # Huge filler rows after pin zone
    for r in range(100, 180):
        ws.cell(r, 3, value="FILLER-" + ("Z" * 800))
    wb.create_sheet("Общий план")
    wb["Общий план"]["B2"] = "short plan"
    wb.save(p)
    wb.close()

    text = xlsx_to_text(p, max_chars=12_000)
    assert "anim-prompt-KEEP-R48" in text
    assert "@row=48" in text
    assert "@row=49" in text
    assert "нет project.xlsx" not in text.lower() or "НЕ «нет" in text


def test_writeback_skips_dense_plan_unknown_labels_without_write(tmp_path: Path) -> None:
    """Fail-closed: строки «план» без @row= и без известной подписи — пропуск,
    книга не мутирует (раньше ждали ValueError; контракт изменён на skip)."""
    src = tmp_path / "project.xlsx"
    dest = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "план"
    ws["A45"] = "промт для картинки 1"
    ws["C45"] = "old"
    wb.create_sheet("Общий план")
    wb.save(src)
    wb.close()

    apply_sheet_blocks_to_xlsx(
        src,
        {"план": [["random", "", "bad"], ["other", "", "bad2"]]},
        dest,
    )

    wb2 = load_workbook(dest)
    try:
        ws2 = wb2["план"]
        assert ws2["C45"].value == "old"  # существующие данные не тронуты
        assert ws2["A1"].value in (None, "")  # мусор не уехал в A1
    finally:
        wb2.close()


def test_harness_forbidden_steps() -> None:
    assert "audio" in HARNESS_FORBIDDEN_STEPS
    assert "music" in HARNESS_FORBIDDEN_STEPS


@pytest.mark.asyncio
async def test_harness_gate_or_raise_ok_and_fail(tmp_path: Path) -> None:
    """Общий гейт шага: ок-данные → проходит; нет project.xlsx → RuntimeError."""
    from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

    from app.models import Base
    from app.services.agent_harness import harness_gate_or_raise

    engine = create_async_engine("sqlite+aiosqlite:///:memory:")
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    factory = async_sessionmaker(engine, expire_on_commit=False)

    wb = Workbook()
    wb.active.title = "план"
    wb.save(tmp_path / "project.xlsx")
    wb.close()
    async with factory() as session:
        p = SimpleNamespace(id=999, data_dir=tmp_path, status="plan_ready", meta={})
        rep = await harness_gate_or_raise(session, p, step="plan")  # type: ignore[arg-type]
        assert rep.ok

        (tmp_path / "project.xlsx").unlink()
        with pytest.raises(RuntimeError, match="harness gate failed"):
            await harness_gate_or_raise(session, p, step="plan")  # type: ignore[arg-type]
    await engine.dispose()


def test_harness_verify_disk_assembled(tmp_path: Path) -> None:
    (tmp_path / "scenes").mkdir()
    (tmp_path / "scenes" / "frame_001.png").write_bytes(b"x")
    (tmp_path / "videos").mkdir()
    (tmp_path / "videos" / "clip_001.mp4").write_bytes(b"x")
    (tmp_path / "final").mkdir()
    (tmp_path / "final" / "out.mp4").write_bytes(b"x")
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "план"
    ws.cell(48, 3, value="anim")
    wb.save(tmp_path / "project.xlsx")
    wb.close()
    report = verify_project_disk(999001, tmp_path, "assembled")
    assert report.ok is True or any(c.name == "r48_anim" for c in report.checks)
    assert "audio" not in report.repair_steps
    assert "music" not in report.repair_steps


def test_harness_plan_ready_does_not_require_media(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "план"
    wb.save(tmp_path / "project.xlsx")
    wb.close()

    report = verify_project_disk(999002, tmp_path, "plan_ready")
    by_name = {c.name: c for c in report.checks}
    assert by_name["scenes_png"].ok is True
    assert by_name["videos_mp4"].ok is True
    assert "img" not in report.repair_steps
    assert "video" not in report.repair_steps


def test_harness_r48_counts_beyond_17_columns(tmp_path: Path) -> None:
    (tmp_path / "scenes").mkdir()
    for i in range(18):
        (tmp_path / "scenes" / f"frame_{i:03d}.png").write_bytes(b"x")
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "план"
    for col in range(3, 21):
        ws.cell(48, col, value=f"anim-{col}")
    wb.save(tmp_path / "project.xlsx")
    wb.close()

    report = verify_project_disk(999003, tmp_path, "assembled")
    by_name = {c.name: c for c in report.checks}
    assert by_name["r48_anim"].ok is True
    assert "filled=18" in by_name["r48_anim"].detail
