"""Пустые и заглушечные промты — не отправлять в outsee."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from app.generation_options import is_skippable_empty_prompt
from app.services.plan_shot2 import (
    ROW_IMAGE_PROMPT_2_V8,
    ROW_VOICEOVER_V8,
    read_shot2_columns,
)


def test_empty_and_whitespace() -> None:
    assert is_skippable_empty_prompt("") is True
    assert is_skippable_empty_prompt("   \n  ") is True


def test_placeholder_no_source_data() -> None:
    assert is_skippable_empty_prompt("нет исходных данных для заполнения") is True
    assert (
        is_skippable_empty_prompt(
            "КАДР 2 / PROMPT_2:\nнет исходных данных для заполнения"
        )
        is True
    )


def test_shot2_prefix_only() -> None:
    prefix = (
        "на основе референса, запрещено делать идентичную иллюстрацию "
        "без смены положения камеры"
    )
    assert is_skippable_empty_prompt(prefix) is True
    assert (
        is_skippable_empty_prompt(
            "на основе референса, запрещено делать идентичную иллюстрацию "
            "без смены положения"
        )
        is True
    )


def test_real_prompt_not_skipped() -> None:
    assert (
        is_skippable_empty_prompt(
            "на основе референса, запрещено делать идентичную иллюстрацию "
            "без смены положения камеры\n\nExtreme close-up: trembling hands"
        )
        is False
    )
    assert is_skippable_empty_prompt("wide cinematic shot of a dark corridor") is False
    # Длинный R45 (sekty): «нет исходных данных» внутри инструкции — не заглушка
    assert (
        is_skippable_empty_prompt(
            "Create one unified scene. character id: нет исходных данных для "
            "ID персонажа; do not invent character IDs. Wide shot of corridor."
        )
        is False
    )


def test_read_shot2_skips_placeholder_in_row46(tmp_path: Path) -> None:
    p = tmp_path / "plan.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "план"
    col = 3
    ws.cell(row=ROW_VOICEOVER_V8, column=col, value="voiceover")
    ws.cell(
        row=ROW_IMAGE_PROMPT_2_V8,
        column=col,
        value="нет исходных данных для заполнения",
    )
    wb.save(p)

    info = read_shot2_columns(p)[1]
    assert info.has_shot2 is False
    assert info.prompt == ""
