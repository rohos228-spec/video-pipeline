"""T12: Excel — зеркало, DB — SoT (инциденты 13.08, проект #14).

- wipe img_pr/anim_pr падает целиком, если Excel-зеркало (R45/R48) не
  удалось очистить — иначе stale-ячейки воскрешали промты после wipe.
- import_v8_xlsx: непустая ячейка зеркала НЕ перетирает непустую БД
  (fill-only-empty); конфликт попадает в summary.prompt_conflicts.
"""

from __future__ import annotations

from pathlib import Path

import pytest
import pytest_asyncio
from openpyxl import Workbook
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from app.models import Base, Frame, FrameStatus, Project, ProjectStatus
from app.services.xlsx_v8_import import (
    ROW_IMAGE_PROMPT_V8,
    ROW_VIDEO_PROMPT_V8,
    ROW_VOICEOVER_V8,
    import_v8_xlsx,
)


@pytest_asyncio.fixture
async def session(tmp_path: Path):
    db_url = f"sqlite+aiosqlite:///{tmp_path / 't.db'}"
    engine = create_async_engine(db_url, echo=False)
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    SessionLocal = async_sessionmaker(engine, expire_on_commit=False)
    async with SessionLocal() as s:
        yield s
    await engine.dispose()


async def _mkproject(session, slug: str = "mirror") -> Project:
    p = Project(
        slug=slug,
        topic="t",
        hero_mode="full_auto",
        status=ProjectStatus.new,
    )
    session.add(p)
    await session.flush()
    return p


async def _mkframe(
    session, project: Project, n: int, *, vo: str, img: str = "", anim: str = ""
) -> Frame:
    fr = Frame(
        project_id=project.id,
        number=n,
        uuid=f"{n:024x}",
        voiceover_text=vo,
        image_prompt=img or None,
        animation_prompt=anim or None,
        status=FrameStatus.planned,
    )
    session.add(fr)
    await session.flush()
    return fr


def _write_xlsx(path: Path, *, vo: str, img: str = "", anim: str = "") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "план"
    ws.cell(row=ROW_VOICEOVER_V8, column=3, value=vo)
    if img:
        ws.cell(row=ROW_IMAGE_PROMPT_V8, column=3, value=img)
    if anim:
        ws.cell(row=ROW_VIDEO_PROMPT_V8, column=3, value=anim)
    wb.save(path)
    wb.close()


# ── wipe: сбой очистки зеркала валит wipe ────────────────────────────


@pytest.mark.asyncio
async def test_wipe_img_pr_fails_loud_when_xlsx_mirror_broken(session) -> None:
    from app.services import reset_step

    p = await _mkproject(session)
    fr = await _mkframe(session, p, 1, vo="vo", img="stale prompt")
    await session.commit()
    p.data_dir.mkdir(parents=True, exist_ok=True)
    # Битый файл вместо валидного xlsx — clear_plan_image_prompts упадёт.
    (p.data_dir / "project.xlsx").write_bytes(b"not an xlsx")

    with pytest.raises(RuntimeError, match="R45/R46"):
        await reset_step._wipe_img_pr(session, p)

    # Wipe отменён: откатываем сессию — промт обязан остаться нетронутым.
    fr_id = fr.id
    await session.rollback()
    row = await session.get(Frame, fr_id)
    assert (row.image_prompt or "") == "stale prompt"


@pytest.mark.asyncio
async def test_wipe_anim_pr_fails_loud_when_xlsx_mirror_broken(session) -> None:
    from app.services import reset_step

    p = await _mkproject(session)
    fr = await _mkframe(session, p, 1, vo="vo", anim="stale anim")
    await session.commit()
    p.data_dir.mkdir(parents=True, exist_ok=True)
    (p.data_dir / "project.xlsx").write_bytes(b"not an xlsx")

    with pytest.raises(RuntimeError, match="R48/R64"):
        await reset_step._wipe_anim_pr(session, p)

    fr_id = fr.id
    await session.rollback()
    row = await session.get(Frame, fr_id)
    assert (row.animation_prompt or "") == "stale anim"


@pytest.mark.asyncio
async def test_wipe_img_pr_ok_without_xlsx(session) -> None:
    """Нет файла-зеркала — нечему воскресать: wipe проходит."""
    from app.services import reset_step

    p = await _mkproject(session)
    await _mkframe(session, p, 1, vo="vo", img="prompt")
    p.data_dir.mkdir(parents=True, exist_ok=True)

    out = await reset_step._wipe_img_pr(session, p)
    assert out["frames_cleared"] == 1
    assert out["xlsx_r45_cleared"] == 0


# ── import_v8: fill-only-empty для промтов ───────────────────────────


@pytest.mark.asyncio
async def test_import_v8_does_not_clobber_non_empty_db_prompt(session) -> None:
    p = await _mkproject(session)
    fr = await _mkframe(session, p, 1, vo="vo one", img="fresh db prompt")
    p.data_dir.mkdir(parents=True, exist_ok=True)
    xlsx = p.data_dir / "project.xlsx"
    _write_xlsx(xlsx, vo="vo one", img="stale mirror prompt")

    summary = await import_v8_xlsx(session, p, xlsx)

    await session.refresh(fr)
    assert fr.image_prompt == "fresh db prompt"
    assert 1 in (summary.get("prompt_conflicts") or [])


@pytest.mark.asyncio
async def test_import_v8_fills_empty_db_prompt_from_mirror(session) -> None:
    p = await _mkproject(session)
    fr = await _mkframe(session, p, 1, vo="vo one")
    p.data_dir.mkdir(parents=True, exist_ok=True)
    xlsx = p.data_dir / "project.xlsx"
    _write_xlsx(xlsx, vo="vo one", img="mirror prompt")

    summary = await import_v8_xlsx(session, p, xlsx)

    await session.refresh(fr)
    assert fr.image_prompt == "mirror prompt"
    assert not (summary.get("prompt_conflicts") or [])
