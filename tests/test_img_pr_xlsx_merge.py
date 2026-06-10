"""img_pr: слияние R45/R46 без затирания enrich."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from app.services.plan_shot2 import ROW_IMAGE_PROMPT_2_V8
from app.services.xlsx_v8_import import ROW_IMAGE_PROMPT_V8, ROW_VOICEOVER_V8
from app.storage.plan_sheet_v8 import merge_gpt_image_prompt_rows_into_project


def _make_pair(tmp_path: Path) -> tuple[Path, Path]:
    proj = tmp_path / "project.xlsx"
    gpt = tmp_path / "gpt.xlsx"
    for path in (proj, gpt):
        wb = Workbook()
        ws = wb.active
        ws.title = "план"
        ws.cell(row=ROW_VOICEOVER_V8, column=3, value="voice block one")
        if path == proj:
            ws.cell(row=10, column=3, value="enrich must survive")
        if path == gpt:
            ws.cell(row=ROW_IMAGE_PROMPT_V8, column=3, value="prompt shot one")
            ws.cell(row=ROW_IMAGE_PROMPT_2_V8, column=3, value="prompt shot two")
        wb.save(path)
    return proj, gpt


def test_merge_preserves_enrich_and_writes_prompts(tmp_path: Path) -> None:
    proj, gpt = _make_pair(tmp_path)
    n45, n46 = merge_gpt_image_prompt_rows_into_project(proj, gpt)
    assert n45 == 1
    assert n46 == 1

    from openpyxl import load_workbook

    wb = load_workbook(proj, data_only=True)
    ws = wb["план"]
    assert ws.cell(row=10, column=3).value == "enrich must survive"
    assert "prompt shot one" in str(ws.cell(row=ROW_IMAGE_PROMPT_V8, column=3).value)
    assert "prompt shot two" in str(ws.cell(row=ROW_IMAGE_PROMPT_2_V8, column=3).value)
    wb.close()
