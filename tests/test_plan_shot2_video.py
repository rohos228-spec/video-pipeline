"""plan R64 — промты видео shot_02."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from app.services.plan_shot2 import ROW_IMAGE_PROMPT_2_V8, ROW_SHOT2_ID_SHOT_V8
from app.services.xlsx_v8_import import ROW_IMAGE_PROMPT_V8, ROW_VOICEOVER_V8
from app.services.shot2_timeline import build_assembly_clip_specs, split_voiceover_duration
from app.services.assembly import ClipSpec
from app.storage.plan_sheet_v8 import (
    read_plan_animation_prompt_shot2_cells,
    write_plan_animation_prompt_shot2,
)
from app.models import Frame, Project


def _write_v8_with_shot2(path: Path, *, prompt2_img: str, voiceover: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "план"
    col = 3
    ws.cell(row=ROW_VOICEOVER_V8, column=col, value=voiceover)
    ws.cell(row=ROW_IMAGE_PROMPT_V8, column=col, value="wide shot")
    ws.cell(row=ROW_IMAGE_PROMPT_2_V8, column=col, value=prompt2_img)
    ws.cell(row=ROW_SHOT2_ID_SHOT_V8, column=col, value="shot_02")
    wb.save(path)


def test_write_and_read_plan_r64(tmp_path: Path, monkeypatch) -> None:
    from app.settings import settings

    monkeypatch.setattr(settings, "data_dir", tmp_path)
    data_dir = tmp_path / "videos" / "slug"
    data_dir.mkdir(parents=True)
    xlsx = data_dir / "project.xlsx"
    _write_v8_with_shot2(xlsx, prompt2_img="close-up hands", voiceover="Voice line here.")
    project = Project(topic="t", slug="slug")

    ok = write_plan_animation_prompt_shot2(
        project, 1, "Slow dolly in on trembling hands."
    )
    assert ok
    cells = read_plan_animation_prompt_shot2_cells(project, [1])
    assert cells[0][1] == "Slow dolly in on trembling hands."


def test_assembly_split_shot2_duration(tmp_path: Path) -> None:
    d1, d2 = split_voiceover_duration(4.2)
    assert abs(d1 + d2 - 4.2) < 0.001
    assert d1 == 2.1

    frames = [Frame(project_id=1, number=1)]
    p1 = tmp_path / "a.mp4"
    p2 = tmp_path / "b.mp4"
    p1.write_bytes(b"x")
    p2.write_bytes(b"x")
    clips = build_assembly_clip_specs(
        frames,
        {1: p1},
        {1: p2},
        {1: 4.0},
    )
    assert clips == [ClipSpec(p1, 2.0), ClipSpec(p2, 2.0)]


def test_assembly_fallback_to_shot2_when_shot1_missing(tmp_path: Path) -> None:
    """Нет shot_01 — нормально: весь слот закрывает shot_02."""
    frames = [Frame(project_id=1, number=9)]
    p2 = tmp_path / "clip_009_s2_x.mp4"
    p2.write_bytes(b"x")
    clips = build_assembly_clip_specs(
        frames,
        {},  # shot_01 отсутствует
        {9: p2},
        {9: 3.0},
    )
    assert len(clips) == 1
    assert clips[0].src == p2
    assert clips[0].duration == 3.0
