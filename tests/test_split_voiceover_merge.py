"""split: merge R49 без wipe enrich; voiceover overlay."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.services.xlsx_v8_import import ROW_VOICEOVER_V8, SHEET_PLAN_V8
from app.storage.plan_sheet_v8 import merge_gpt_voiceover_row_into_project


def _proj_with_enrich(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = SHEET_PLAN_V8
    ws["A2"] = "enrich label"
    ws["C2"] = "enrich keep"
    ws["A45"] = "промт"
    ws["C45"] = "img keep"
    ws.cell(row=ROW_VOICEOVER_V8, column=1, value="закадровый текст")
    ws.cell(row=ROW_VOICEOVER_V8, column=3, value="old vo")
    wb.create_sheet("Общий план")
    wb.save(path)
    wb.close()


def test_merge_gpt_voiceover_preserves_enrich(tmp_path: Path) -> None:
    proj = tmp_path / "project.xlsx"
    gpt = tmp_path / "gpt.xlsx"
    _proj_with_enrich(proj)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = SHEET_PLAN_V8
    ws.cell(row=ROW_VOICEOVER_V8, column=3, value="block 1")
    ws.cell(row=ROW_VOICEOVER_V8, column=4, value="block 2")
    # GPT «забыл» enrich — не должен затереть
    wb.save(gpt)
    wb.close()

    n = merge_gpt_voiceover_row_into_project(proj, gpt)
    assert n == 2
    got = load_workbook(proj)
    assert got[SHEET_PLAN_V8]["C2"].value == "enrich keep"
    assert got[SHEET_PLAN_V8]["C45"].value == "img keep"
    assert got[SHEET_PLAN_V8].cell(row=ROW_VOICEOVER_V8, column=3).value == "block 1"
    assert got[SHEET_PLAN_V8].cell(row=ROW_VOICEOVER_V8, column=4).value == "block 2"
    got.close()
