"""План не считается готовым, пока general_plan — не шаблон/заглушка."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from app.models import Project, ProjectStatus
from app.services.plan_validation import MIN_GENERAL_PLAN_CHARS, is_meaningful_general_plan
from app.services.project_state import compute_actual_status
from app.services.xlsx_step_runners import (
    _assert_downloaded_plan_meaningful,
    _plan_empty_error,
)
from app.services.xlsx_v8_import import SHEET_GENERAL_V8, _read_general_plan
from app.storage.project_sheet import resolve_default_template_path

REPO = Path(__file__).resolve().parents[1]


def test_v8_template_general_plan_is_not_meaningful() -> None:
    tpl = resolve_default_template_path()
    wb = load_workbook(tpl, read_only=True, data_only=True)
    text = _read_general_plan(wb)
    assert text
    assert len(text.strip()) < MIN_GENERAL_PLAN_CHARS
    assert not is_meaningful_general_plan(text)


def test_long_plan_is_meaningful() -> None:
    text = "А" * MIN_GENERAL_PLAN_CHARS
    assert is_meaningful_general_plan(text)


def test_plan_output_contract_targets_general_plan_sheet() -> None:
    contract = (
        REPO / "prompts" / "blocks" / "plan_output_contract" / "xlsx_plan_timing.md"
    ).read_text(encoding="utf-8")
    template = (REPO / "prompts" / "steps" / "01_plan" / "template.md").read_text(
        encoding="utf-8"
    )
    assert "Общий план" in contract
    assert "Общий план" in template
    # Не требовать заполнять лист «план» на шаге 1
    assert "Заполняй лист «план»" not in contract
    assert "с заполненным листом «план»" not in template


def test_assert_downloaded_plan_rejects_empty_general_plan(tmp_path: Path) -> None:
    path = tmp_path / "empty_plan.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_GENERAL_V8
    ws["A1"] = "Короткий заголовок"
    wb.create_sheet("план")
    wb.save(path)
    wb.close()

    with pytest.raises(RuntimeError, match="Общий план") as ei:
        _assert_downloaded_plan_meaningful(path)
    assert "не тот лист" in str(ei.value)


def test_assert_downloaded_plan_accepts_filled_general_plan(tmp_path: Path) -> None:
    path = tmp_path / "filled_plan.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_GENERAL_V8
    ws["A1"] = "Хук"
    ws["B1"] = "А" * MIN_GENERAL_PLAN_CHARS
    wb.save(path)
    wb.close()
    _assert_downloaded_plan_meaningful(path)


def test_plan_empty_error_lists_sheets(tmp_path: Path) -> None:
    path = tmp_path / "p.xlsx"
    wb = Workbook()
    wb.active.title = SHEET_GENERAL_V8
    wb.create_sheet("план")
    wb.save(path)
    wb.close()
    err = _plan_empty_error(path, plan_len=12)
    assert "план" in str(err)
    assert "Общий план" in str(err)


@pytest.mark.asyncio
async def test_compute_actual_status_ignores_short_general_plan() -> None:
    from unittest.mock import AsyncMock, MagicMock

    p = Project(
        id=1,
        topic="тест",
        slug="test",
        status=ProjectStatus.plan_ready,
        general_plan="короткий план из шаблона",
    )
    session = AsyncMock()
    call_count = {"n": 0}

    async def mock_execute(stmt):
        call_count["n"] += 1
        m = MagicMock()
        m.scalar_one.return_value = 0
        return m

    session.execute = mock_execute
    st = await compute_actual_status(session, p)
    assert st == ProjectStatus.new
