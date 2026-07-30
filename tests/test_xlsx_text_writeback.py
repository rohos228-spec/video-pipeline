"""TSV/xlsx write-back из текстового ответа GPT API."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.services.xlsx_text_writeback import (
    apply_sheet_blocks_to_xlsx,
    extract_sheet_blocks,
    merge_xlsx_nonempty_overlay,
    writeback_project_xlsx,
)


def test_extract_sheet_blocks_tsv() -> None:
    text = (
        "# Лист: план\n"
        "A\tB\tC\n"
        "1\t2\t3\n"
        "# Лист: Кадры\n"
        "n\ttext\n"
        "1\thello\n"
    )
    blocks = extract_sheet_blocks(text)
    assert list(blocks.keys()) == ["план", "Кадры"]
    assert blocks["план"][0] == ["A", "B", "C"]
    assert blocks["Кадры"][1] == ["1", "hello"]


def test_writeback_incomplete_and_merge(tmp_path: Path) -> None:
    from app.services.xlsx_text_writeback import (
        WRITEBACK_HINT,
        merge_writeback_texts,
        writeback_looks_incomplete,
    )

    assert "ПОЛНОЕ ЗАПОЛНЕНИЕ" in WRITEBACK_HINT
    src = tmp_path / "project.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Общий план"
    for i in range(1, 41):
        ws.cell(row=i, column=1, value=f"row{i}")
        ws.cell(row=i, column=2, value="")
    wb.save(src)
    wb.close()

    partial = "\n".join(
        ["# Лист: Общий план"]
        + [f"@row={i}\trow{i}\tonly-part" for i in range(1, 9)]
    )
    incomplete, reason, sheet, nxt = writeback_looks_incomplete(
        reply_text=partial, template_xlsx=src
    )
    assert incomplete is True
    assert sheet == "Общий план"
    assert nxt == 9

    more = "\n".join(
        ["# Лист: Общий план"]
        + [f"@row={i}\trow{i}\trest" for i in range(9, 41)]
    )
    merged = merge_writeback_texts(partial, more)
    incomplete2, _, _, _ = writeback_looks_incomplete(
        reply_text=merged, template_xlsx=src
    )
    assert incomplete2 is False
    assert merged.count("@row=") == 40


def test_merge_writeback_last_wins_same_row() -> None:
    from app.services.xlsx_text_writeback import merge_writeback_texts

    a = "# Лист: план\n@row=5\tlabel\tOLD\n"
    b = "# Лист: план\n@row=5\tlabel\tNEW\n@row=6\tlabel2\tx\n"
    merged = merge_writeback_texts(a, b)
    assert "OLD" not in merged
    assert "@row=5\tlabel\tNEW" in merged
    assert "@row=6\tlabel2\tx" in merged


def test_merge_orphan_tabs_inherit_sheet() -> None:
    from app.services.xlsx_text_writeback import extract_sheet_blocks, merge_writeback_texts

    first = "# Лист: план\n@row=1\ta\tb\n"
    # CONTINUE без `# Лист:` — раньше уезжало в «Данные» и skip на v8.
    cont = "@row=2\tc\td\n"
    merged = merge_writeback_texts(first, cont, default_sheet="план")
    blocks = extract_sheet_blocks(merged)
    assert "Данные" not in blocks
    assert len(blocks["план"]) == 2


def test_writeback_remaps_dannye_and_keeps_plan_labels(tmp_path: Path) -> None:
    src = tmp_path / "project.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "план"
    ws["A5"] = "закадровый текст"
    ws["B5"] = "старое"
    wb.create_sheet("Общий план")
    wb.save(src)
    wb.close()

    reply = (
        "# Лист: Данные\n"
        "@row=5\tЛОМАЙ ПОДПИСЬ\tновое значение\n"
    )
    out = writeback_project_xlsx(
        project_xlsx=src,
        reply_text=reply,
        downloaded_paths=[],
    )
    assert out == src
    wb2 = load_workbook(src, data_only=True)
    assert wb2["план"]["A5"].value == "закадровый текст"
    assert wb2["план"]["B5"].value == "новое значение"
    wb2.close()


def test_extract_from_fenced_block() -> None:
    text = "вот результат:\n```tsv\n# Лист: Данные\nx\ty\n10\t20\n```\nконец"
    blocks = extract_sheet_blocks(text)
    assert "Данные" in blocks
    assert blocks["Данные"] == [["x", "y"], ["10", "20"]]


def test_apply_and_writeback(tmp_path: Path) -> None:
    src = tmp_path / "project.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "план"
    ws["A1"] = "old"
    ws["A5"] = "keep-me"
    wb.create_sheet("Кадры")
    wb.save(src)
    wb.close()

    reply = "# Лист: план\nname\tval\nfoo\tbar\n# Лист: Кадры\nid\n1\n"
    out = writeback_project_xlsx(
        project_xlsx=src,
        reply_text=reply,
        downloaded_paths=[],
    )
    assert out == src
    wb2 = load_workbook(src)
    assert wb2["план"]["A1"].value == "name"
    assert wb2["план"]["B2"].value == "bar"
    # Overlay: строки вне TSV не стираем
    assert wb2["план"]["A5"].value == "keep-me"
    assert wb2["Кадры"]["A2"].value == "1"
    wb2.close()


def test_apply_row_marks_preserve_sparse_plan_rows(tmp_path: Path) -> None:
    """@row=N пишет в абсолютные строки — R45/R49 не съезжают вниз."""
    src = tmp_path / "project.xlsx"
    dest = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "план"
    ws["A1"] = "заголовок"
    ws["A15"] = "таймкод"
    ws["C15"] = "0:00-0:05"
    ws["A45"] = "промт картинки"
    ws["C45"] = "old prompt"
    ws["A49"] = "закадровый текст"
    ws["C49"] = "old vo"
    ws["A2"] = "enrich keep"
    ws["C2"] = "enrich value"
    wb.create_sheet("Общий план")
    wb.save(src)
    wb.close()

    blocks = extract_sheet_blocks(
        "# Лист: план\n"
        "@row=45\tпромт картинки\t\tnew prompt\n"
        "@row=49\tзакадровый текст\t\tnew vo\n"
    )
    apply_sheet_blocks_to_xlsx(src, blocks, dest)
    wb2 = load_workbook(dest)
    assert wb2["план"]["C45"].value == "new prompt"
    assert wb2["план"]["C49"].value == "new vo"
    assert wb2["план"]["C15"].value == "0:00-0:05"
    assert wb2["план"]["C2"].value == "enrich value"
    assert wb2["план"]["A2"].value == "enrich keep"
    wb2.close()


def test_apply_label_fallback_without_row_marks(tmp_path: Path) -> None:
    """Legacy TSV без @row=: сопоставляем по подписи A на листе «план»."""
    src = tmp_path / "project.xlsx"
    dest = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "план"
    ws["A2"] = "enrich"
    ws["C2"] = "keep"
    ws["A45"] = "промт для картинки 1"
    ws["C45"] = "old"
    ws["A49"] = "закадровый текст"
    ws["C49"] = "old vo"
    wb.create_sheet("Общий план")
    wb.save(src)
    wb.close()

    # Уплотнённый ответ без пустых рядов — раньше писал в R1/R2.
    apply_sheet_blocks_to_xlsx(
        src,
        {
            "план": [
                ["промт для картинки 1", "", "fresh prompt"],
                ["закадровый текст", "", "fresh vo"],
            ]
        },
        dest,
    )
    wb2 = load_workbook(dest)
    assert wb2["план"]["C45"].value == "fresh prompt"
    assert wb2["план"]["C49"].value == "fresh vo"
    assert wb2["план"]["C2"].value == "keep"
    wb2.close()


def test_binary_writeback_overlays_not_wipes(tmp_path: Path) -> None:
    project = tmp_path / "project.xlsx"
    dl = tmp_path / "gpt.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "план"
    ws["A2"] = "enrich"
    ws["C2"] = "must-keep"
    ws["A49"] = "закадровый текст"
    ws["C49"] = "old"
    wb.create_sheet("Общий план")
    wb.save(project)
    wb.close()

    wb2 = Workbook()
    ws2 = wb2.active
    assert ws2 is not None
    ws2.title = "план"
    ws2["C49"] = "from_download"
    # enrich отсутствует в ответе GPT — не должен затереться
    wb2.save(dl)
    wb2.close()

    out = writeback_project_xlsx(
        project_xlsx=project,
        reply_text="",
        downloaded_paths=[dl],
    )
    assert out == project
    got = load_workbook(project)
    assert got["план"]["C49"].value == "from_download"
    assert got["план"]["C2"].value == "must-keep"
    got.close()


def test_merge_xlsx_nonempty_overlay(tmp_path: Path) -> None:
    proj = tmp_path / "p.xlsx"
    gpt = tmp_path / "g.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "план"
    ws["A1"] = "label"
    ws["B1"] = "keep"
    wb.create_sheet("Общий план")
    wb.save(proj)
    wb.close()

    wb2 = Workbook()
    ws2 = wb2.active
    assert ws2 is not None
    ws2.title = "план"
    ws2["B1"] = "updated"
    ws2["C1"] = "new"
    wb2.save(gpt)
    wb2.close()

    merge_xlsx_nonempty_overlay(proj, gpt, proj)
    got = load_workbook(proj)
    assert got["план"]["A1"].value == "label"
    assert got["план"]["B1"].value == "updated"
    assert got["план"]["C1"].value == "new"
    got.close()


def test_apply_does_not_add_foreign_sheets_on_v8_workbook(tmp_path: Path) -> None:
    src = tmp_path / "project.xlsx"
    dest = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Общий план"
    ws["A1"] = "label"
    wb.create_sheet("план")
    wb.save(src)
    wb.close()

    apply_sheet_blocks_to_xlsx(
        src,
        {"Общий план": [["Хук", "длинный текст"]], "ЧужойЛист": [["x"]]},
        dest,
    )
    wb2 = load_workbook(dest)
    assert "ЧужойЛист" not in wb2.sheetnames
    assert wb2["Общий план"]["A1"].value == "Хук"
    assert wb2["Общий план"]["B1"].value == "длинный текст"
    assert "план" in wb2.sheetnames
    wb2.close()


def test_writeback_prefers_downloaded_xlsx(tmp_path: Path) -> None:
    project = tmp_path / "project.xlsx"
    dl = tmp_path / "content_1.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws["A1"] = "from_download"
    wb.save(dl)
    wb.close()
    # Нет v8-листов → полный copy как раньше
    Workbook().save(project)

    out = writeback_project_xlsx(
        project_xlsx=project,
        reply_text="# Лист: X\na\tb\n",
        downloaded_paths=[dl],
    )
    assert out == project
    wb2 = load_workbook(project)
    assert wb2.active["A1"].value == "from_download"
    wb2.close()


def test_writeback_prose_fallback_into_general_plan(tmp_path: Path) -> None:
    from app.services.plan_validation import is_meaningful_general_plan
    from app.services.xlsx_v8_import import _read_general_plan

    src = tmp_path / "project.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Общий план"
    ws["A1"] = "Хук"
    ws["A2"] = "Основная тема"
    ws.merge_cells("B2:D2")
    # Без листа «план» — prose fallback разрешён (простая книга плана).
    wb.save(src)
    wb.close()

    prose = "А" * 250 + "\n\nРим был велик: армия, право, дороги."
    out = writeback_project_xlsx(
        project_xlsx=src,
        reply_text=prose,
        downloaded_paths=[],
    )
    assert out == src
    wb2 = load_workbook(src, data_only=True)
    # Структура строк не сломана: подпись A2 на месте, текст в B2
    assert wb2["Общий план"]["A1"].value == "Хук"
    assert wb2["Общий план"]["A2"].value == "Основная тема"
    assert "Рим был велик" in str(wb2["Общий план"]["B2"].value or "")
    text = _read_general_plan(wb2) or ""
    wb2.close()
    assert is_meaningful_general_plan(text)
    assert "Рим был велик" in text


def test_writeback_refuses_prose_on_frame_plan_workbook(tmp_path: Path) -> None:
    """Полный project.xlsx с листом «план»: проза не должна заливать B2."""
    src = tmp_path / "project.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Общий план"
    ws["A2"] = "Основная тема"
    ws["B2"] = "оригинал плана не трогать"
    plan = wb.create_sheet("план")
    plan["A1"] = "кадр"
    plan["B1"] = "текст"
    wb.save(src)
    wb.close()

    prose = "А" * 250 + "\n\nМодель ответила болтовнёй без TSV."
    out = writeback_project_xlsx(
        project_xlsx=src,
        reply_text=prose,
        downloaded_paths=[],
    )
    assert out is None
    wb2 = load_workbook(src, data_only=True)
    assert wb2["Общий план"]["B2"].value == "оригинал плана не трогать"
    assert wb2["план"]["B1"].value == "текст"
    wb2.close()


def test_writeback_skips_check_report_json(tmp_path: Path) -> None:
    src = tmp_path / "project.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Общий план"
    ws["A2"] = "Тема"
    ws["B2"] = "оригинал плана"
    wb.save(src)
    wb.close()

    report = (
        '{"decision":"regen","confidence":0.99,'
        '"criteria":{"flow_continuity":{"verdict":"pass"}},'
        '"fix_hints":["убери мета"],"issues":["x"]}'
    )
    out = writeback_project_xlsx(
        project_xlsx=src,
        reply_text=report,
        downloaded_paths=[],
    )
    assert out is None
    wb2 = load_workbook(src, data_only=True)
    assert wb2["Общий план"]["B2"].value == "оригинал плана"
    wb2.close()


def test_apply_creates_missing_sheet(tmp_path: Path) -> None:
    src = tmp_path / "a.xlsx"
    dest = tmp_path / "b.xlsx"
    Workbook().save(src)
    apply_sheet_blocks_to_xlsx(src, {"Новый": [["a", "b"]]}, dest)
    wb = load_workbook(dest)
    assert "Новый" in wb.sheetnames
    assert wb["Новый"]["B1"].value == "b"
    wb.close()
