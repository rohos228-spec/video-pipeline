"""plan_sheet_v8: запись/чтение R15 таймкодов."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from app.models import Project
from app.services.xlsx_v8_import import (
    ROW_DURATION_V8,
    ROW_TIMECODE_V8,
    ROW_VOICEOVER_V8,
    SHEET_PLAN_V8,
)
from app.storage.plan_sheet_v8 import (
    read_plan_timestamps_cells,
    write_plan_durations,
    write_plan_timestamps,
)


def test_write_and_read_plan_r15(tmp_path: Path, monkeypatch) -> None:
    from app.services.plan_timestamps import normalize_timestamp_label

    root = tmp_path / "data"
    slug_dir = root / "videos" / "t1"
    slug_dir.mkdir(parents=True)
    xlsx = slug_dir / "project.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_PLAN_V8
    ws.cell(row=ROW_VOICEOVER_V8, column=3, value="кадр один")
    ws.cell(row=ROW_VOICEOVER_V8, column=4, value="кадр два")
    wb.save(xlsx)

    monkeypatch.setattr("app.models.settings.data_dir", root)
    project = Project(topic="t", slug="t1")

    written = write_plan_timestamps(
        project,
        [(1, "0:00.00-0:03.28"), (2, "0:03.28-0:07.50")],
    )
    assert written == 2
    cells, row = read_plan_timestamps_cells(project, [1, 2])
    assert row == ROW_TIMECODE_V8
    norm = [(n, normalize_timestamp_label(t)) for n, t in cells]
    assert norm == [(1, "0:00.00-0:03.28"), (2, "0:03.28-0:07.50")]


def test_write_plan_durations_follows_voiceover_columns(
    tmp_path: Path, monkeypatch
) -> None:
    """R50 «Время на кадр» пишется в те же колонки, что и закадр кадра (R49)."""
    root = tmp_path / "data"
    slug_dir = root / "videos" / "t2"
    slug_dir.mkdir(parents=True)
    xlsx = slug_dir / "project.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_PLAN_V8
    # Колонка 4 намеренно пустая: кадр 2 живёт в колонке 5.
    ws.cell(row=ROW_VOICEOVER_V8, column=3, value="кадр один")
    ws.cell(row=ROW_VOICEOVER_V8, column=5, value="кадр два")
    wb.save(xlsx)

    monkeypatch.setattr("app.models.settings.data_dir", root)
    project = Project(topic="t", slug="t2")

    written = write_plan_durations(project, [(1, 3.28), (2, 4.24)])
    assert written == 2

    from openpyxl import load_workbook

    wb2 = load_workbook(xlsx)
    ws2 = wb2[SHEET_PLAN_V8]
    assert ws2.cell(row=ROW_DURATION_V8, column=3).value == 3.28
    assert ws2.cell(row=ROW_DURATION_V8, column=4).value is None
    assert ws2.cell(row=ROW_DURATION_V8, column=5).value == 4.24
    assert ws2.cell(row=ROW_DURATION_V8, column=1).value
    wb2.close()
