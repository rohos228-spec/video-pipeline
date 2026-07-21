"""Тесты панели монтажа (montage-board)."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from app.models import Base, Frame, Project
from app.orchestrator.steps.generate_images import _XLSX_ROWS_PERSONS
from app.services.montage_board import build_montage_board
from app.services.plan_shot2 import ROW_IMAGE_PROMPT_2_V8, ROW_VIDEO_PROMPT_2_V8
from app.services.xlsx_v8_import import (
    ROW_IMAGE_PROMPT_V8,
    ROW_VIDEO_PROMPT_V8,
    ROW_VOICEOVER_V8,
    SHEET_PLAN_V8,
)


@pytest.fixture
async def session(tmp_path: Path) -> AsyncSession:
    db_path = tmp_path / "montage.db"
    engine = create_async_engine(f"sqlite+aiosqlite:///{db_path}")
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    factory = async_sessionmaker(engine, expire_on_commit=False)
    async with factory() as s:
        yield s
    await engine.dispose()


@pytest.fixture
def montage_project(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Project:
    data_root = tmp_path / "data"
    data_root.mkdir(parents=True, exist_ok=True)
    monkeypatch.setattr("app.settings.settings.data_dir", str(data_root))
    p = Project(id=99, slug="montage-test", topic="Тест", hero_mode="auto")
    p.data_dir.mkdir(parents=True, exist_ok=True)
    return p


@pytest.mark.asyncio
async def test_montage_board_reads_excel_voiceover_and_characters(
    montage_project: Project,
    session: AsyncSession,
) -> None:
    xlsx = montage_project.data_dir / "project.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_PLAN_V8
    persons_row = _XLSX_ROWS_PERSONS[0]
    ws.cell(row=persons_row, column=3, value="c01, c02")
    ws.cell(row=ROW_VOICEOVER_V8, column=3, value="Текст закадровки кадра 1")
    wb.create_sheet("Персонажи")
    persons = wb["Персонажи"]
    persons.cell(row=1, column=2, value="c01")
    persons.cell(row=3, column=2, value="Кот")
    persons.cell(row=1, column=3, value="c02")
    persons.cell(row=3, column=3, value="Мышь")
    wb.save(xlsx)

    chars_dir = montage_project.data_dir / "characters"
    chars_dir.mkdir(parents=True, exist_ok=True)
    (chars_dir / "c01.png").write_bytes(b"png1")
    (chars_dir / "c02.png").write_bytes(b"png2")

    fr = Frame(
        project_id=montage_project.id,
        number=1,
        voiceover_text="из БД",
        status="planned",
    )
    session.add(montage_project)
    session.add(fr)
    await session.flush()

    board = await build_montage_board(session, montage_project)
    assert board["frame_count"] == 1
    assert "meta" in board
    assert board["meta"]["video_trims"] == {}
    row = board["frames"][0]
    assert row["voiceover_excel"] == "Текст закадровки кадра 1"
    assert row["characters"] == "c01, c02"
    assert row["number"] == 1
    assert len(row["character_refs"]) == 2
    assert row["character_refs"][0]["id"] == "c01"
    assert row["character_refs"][0]["name"] == "Кот"
    assert row["character_refs"][0]["image_url"] is not None
    assert row["character_refs"][1]["id"] == "c02"


@pytest.mark.asyncio
async def test_montage_board_exposes_source_prompts_from_excel(
    montage_project: Project,
    session: AsyncSession,
) -> None:
    """«Редактировать промт» должен получать промт исходника (Excel → board DTO)."""
    xlsx = montage_project.data_dir / "project.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_PLAN_V8
    col = 3  # кадр #1
    ws.cell(row=ROW_IMAGE_PROMPT_V8, column=col, value="IMG SHOT1 PROMPT")
    ws.cell(row=ROW_IMAGE_PROMPT_2_V8, column=col, value="IMG SHOT2 PROMPT")
    ws.cell(row=ROW_VIDEO_PROMPT_V8, column=col, value="VID SHOT1 PROMPT")
    ws.cell(row=ROW_VIDEO_PROMPT_2_V8, column=col, value="VID SHOT2 PROMPT long enough")
    wb.save(xlsx)

    fr = Frame(
        project_id=montage_project.id,
        number=1,
        voiceover_text="vo",
        image_prompt="stale-db-image",
        animation_prompt="stale-db-video",
        status="planned",
        attrs={
            "image_prompt_shot2": "stale-db-img2",
            "animation_prompt_shot2": "stale-db-vid2",
        },
    )
    session.add(montage_project)
    session.add(fr)
    await session.flush()

    board = await build_montage_board(session, montage_project)
    row = board["frames"][0]
    assert row["image_prompt_shot1"] == "IMG SHOT1 PROMPT"
    assert row["image_prompt_shot2"] == "IMG SHOT2 PROMPT"
    assert row["animation_prompt_shot1"] == "VID SHOT1 PROMPT"
    assert row["animation_prompt_shot2"] == "VID SHOT2 PROMPT long enough"


@pytest.mark.asyncio
async def test_montage_board_prompt_falls_back_to_frame_db(
    montage_project: Project,
    session: AsyncSession,
) -> None:
    xlsx = montage_project.data_dir / "project.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_PLAN_V8
    wb.save(xlsx)

    fr = Frame(
        project_id=montage_project.id,
        number=1,
        voiceover_text="vo",
        image_prompt="DB IMAGE PROMPT",
        animation_prompt="DB VIDEO PROMPT",
        status="planned",
        attrs={"image_prompt_shot2": "DB IMAGE SHOT2"},
    )
    session.add(montage_project)
    session.add(fr)
    await session.flush()

    board = await build_montage_board(session, montage_project)
    row = board["frames"][0]
    assert row["image_prompt_shot1"] == "DB IMAGE PROMPT"
    assert row["image_prompt_shot2"] == "DB IMAGE SHOT2"
    assert row["animation_prompt_shot1"] == "DB VIDEO PROMPT"


def test_preview_url_encodes_spaces() -> None:
    from app.services.montage_board import _preview_url

    p = Path("/tmp/has space/frame_001.png")
    # Файл может не существовать — тогда None; проверяем encode на существующем.
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_bytes(b"x")
    url = _preview_url(p)
    assert url is not None
    assert " " not in url
    assert "%20" in url or "%2520" in url or "has%20space" in url
    p.unlink(missing_ok=True)


@pytest.mark.asyncio
async def test_build_montage_board_survives_expired_frames(
    montage_project: Project,
    session: AsyncSession,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """ORM Frame в to_thread давал MissingGreenlet → UI «не удалось загрузить»."""
    xlsx = montage_project.data_dir / "project.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_PLAN_V8
    ws.cell(row=ROW_IMAGE_PROMPT_V8, column=3, value="EXPIRED SAFE PROMPT")
    wb.save(xlsx)

    fr = Frame(
        project_id=montage_project.id,
        number=1,
        voiceover_text="vo",
        image_prompt="db-fallback",
        animation_prompt="db-anim",
        status="planned",
        attrs={"image_prompt_shot2": "db-s2"},
    )
    session.add(montage_project)
    session.add(fr)
    await session.flush()

    # Имитация: кадры expired после select, пока Excel читается в worker-thread.
    # Project остаётся hot (как в HTTP-запросе после get_project).
    from app.services import montage_board as mb

    real_snapshot = mb._snapshot_frames

    def _snapshot_then_expire(frames: list[Frame]) -> list:
        snaps = real_snapshot(frames)
        for obj in frames:
            session.expire(obj)
        return snaps

    monkeypatch.setattr(mb, "_snapshot_frames", _snapshot_then_expire)
    board = await build_montage_board(session, montage_project)

    assert board["frame_count"] == 1
    assert board["frames"][0]["image_prompt_shot1"] == "EXPIRED SAFE PROMPT"


@pytest.mark.asyncio
async def test_read_source_prompts_once_accepts_plain_snapshots(
    montage_project: Project,
    tmp_path: Path,
) -> None:
    """Worker-thread не должен получать SQLAlchemy Frame."""
    import asyncio

    from app.services.montage_board import (
        _FrameBoardSnapshot,
        _read_source_prompts_once,
    )

    xlsx = montage_project.data_dir / "project.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_PLAN_V8
    ws.cell(row=ROW_IMAGE_PROMPT_V8, column=3, value="THREAD SAFE")
    wb.save(xlsx)

    snaps = [
        _FrameBoardSnapshot(
            id=1,
            number=1,
            image_prompt="db",
            animation_prompt="anim",
            attrs={},
        )
    ]
    out = await asyncio.to_thread(_read_source_prompts_once, xlsx, snaps)
    assert out[1]["image_prompt_shot1"] == "THREAD SAFE"


@pytest.mark.asyncio
async def test_load_xlsx_bundle_is_sequential_single_thread(
    montage_project: Project,
    session: AsyncSession,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Параллельные openpyxl на Windows дают lock — bundle должен быть 1× to_thread."""
    import asyncio

    from app.services import montage_board as mb

    xlsx = montage_project.data_dir / "project.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_PLAN_V8
    ws.cell(row=ROW_VOICEOVER_V8, column=3, value="vo")
    ws.cell(row=ROW_IMAGE_PROMPT_V8, column=3, value="IMG")
    wb.save(xlsx)

    fr = Frame(
        project_id=montage_project.id,
        number=1,
        voiceover_text="vo",
        status="planned",
    )
    session.add(montage_project)
    session.add(fr)
    await session.flush()

    calls: list[object] = []
    real_to_thread = asyncio.to_thread

    async def tracking_to_thread(fn, /, *args, **kwargs):  # type: ignore[no-untyped-def]
        calls.append(fn)
        return await real_to_thread(fn, *args, **kwargs)

    monkeypatch.setattr(mb.asyncio, "to_thread", tracking_to_thread)
    board = await build_montage_board(session, montage_project)
    assert board["frame_count"] == 1
    assert board["frames"][0]["image_prompt_shot1"] == "IMG"
    # Ровно один to_thread на Excel-бандл (не 3 параллельных openpyxl).
    assert calls == [mb._load_montage_xlsx_bundle]


def test_load_montage_xlsx_bundle_survives_plan_open_failure(
    montage_project: Project,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    from app.services.montage_board import (
        _FrameBoardSnapshot,
        _load_montage_xlsx_bundle,
    )

    xlsx = montage_project.data_dir / "project.xlsx"
    xlsx.write_bytes(b"not-xlsx")
    snaps = [
        _FrameBoardSnapshot(
            id=1,
            number=1,
            image_prompt="DB IMG",
            animation_prompt="DB VID",
            attrs={},
        )
    ]
    excel, prompts, shot2 = _load_montage_xlsx_bundle(
        xlsx,
        chars_dir=montage_project.data_dir / "characters",
        frames=snaps,
    )
    assert excel == {}
    assert prompts[1]["image_prompt_shot1"] == "DB IMG"
    assert shot2 == {}
