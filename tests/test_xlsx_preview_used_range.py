"""xlsx preview must show real used range — not padded 500×200 empty grid."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from app.web.routers.project_ops import (
    _col_letter,
    _trim_trailing_empty_cols,
    _trim_trailing_empty_rows,
)


def test_trim_trailing_empty_cols_does_not_pad() -> None:
    rows = [
        ["idea", "", "", ""],
        ["more", "x", "", ""],
    ]
    out = _trim_trailing_empty_cols(rows)
    assert out == [["idea", ""], ["more", "x"]]


def test_trim_trailing_empty_rows() -> None:
    rows = [["a"], [""], ["b"], ["", ""], [""]]
    assert _trim_trailing_empty_rows(rows) == [["a"], [""], ["b"]]


def test_col_letter() -> None:
    assert _col_letter(0) == "A"
    assert _col_letter(25) == "Z"
    assert _col_letter(26) == "AA"


@pytest.mark.asyncio
async def test_preview_xlsx_no_pad_and_used_range(tmp_path: Path, monkeypatch) -> None:
    from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

    from app.models import Base, Project, ProjectStatus, Workflow
    from app.settings import settings
    from app.web.routers import project_ops

    monkeypatch.setattr(settings, "data_dir", tmp_path)
    data_dir = tmp_path / "videos" / "slug-xlsx"
    data_dir.mkdir(parents=True)
    xlsx = data_dir / "project.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Общий план"
    ws["A1"] = "Общая драматургическая идея"
    ws["A2"] = "Длинный текст плана на второй строке"
    ws["B2"] = "side"
    wb.create_sheet("план")
    wb["план"]["C45"] = "prompt frame 1"
    wb.save(xlsx)

    engine = create_async_engine(f"sqlite+aiosqlite:///{tmp_path / 't.db'}")
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    factory = async_sessionmaker(engine, expire_on_commit=False)
    async with factory() as session:
        session.add(Workflow(name="default", is_default=True, nodes=[], edges=[]))
        p = Project(
            topic="t",
            slug="slug-xlsx",
            status=ProjectStatus.new,
            hero_mode="no_hero",
        )
        session.add(p)
        await session.flush()

        result = await project_ops.preview_xlsx(
            project_id=p.id,
            sheet="Общий план",
            max_rows=500,
            max_cols=200,
            start_row=1,
            row=None,
            raw=True,
            node_key=None,
            session=session,
        )

    assert result["active_sheet"] == "Общий план"
    assert len(result["rows"]) == 2
    assert result["rows"][0][0] == "Общая драматургическая идея"
    assert result["rows"][1][0] == "Длинный текст плана на второй строке"
    # Must NOT be padded to 200 columns
    assert all(len(r) <= 2 for r in result["rows"])
    assert result["col_letters"][0] == "A"
    assert result.get("truncated_cols") is False
