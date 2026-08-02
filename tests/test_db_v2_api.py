"""DB v2 API: fail-closed apply-ops (GPT) и экспорт DB → project.xlsx."""

from __future__ import annotations

import pytest
import pytest_asyncio
from httpx import ASGITransport, AsyncClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from app.models import Base, Frame, Project, ProjectStatus, PromptVersion
from app.services import db_apply, db_v2
from app.services.xlsx_v8_import import (
    ROW_IMAGE_PROMPT_V8,
    ROW_VIDEO_PROMPT_V8,
    ROW_VOICEOVER_V8,
)
from app.settings import settings
from app.web.api import create_app
from app.web.deps import get_session

app = create_app()


@pytest_asyncio.fixture
async def api_client(tmp_path, monkeypatch):
    db_file = tmp_path / "dbv2api.db"
    monkeypatch.setattr(settings, "data_dir", tmp_path)
    # harness читает raw sqlite через settings.sqlite_path — тот же файл, что и API
    monkeypatch.setattr(settings, "sqlite_path", db_file)
    engine = create_async_engine(f"sqlite+aiosqlite:///{db_file}")
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
        await db_v2.migrate_db_v2_schema(conn)
    factory = async_sessionmaker(engine, expire_on_commit=False)

    async def _gen():
        async with factory() as s:
            yield s

    app.dependency_overrides[get_session] = _gen

    async with factory() as session:
        p = Project(topic="dbv2", slug="dbv2-api", status=ProjectStatus.new)
        session.add(p)
        await session.flush()
        for i in (1, 2):
            session.add(
                Frame(
                    project_id=p.id,
                    number=i,
                    voiceover_text=f"реплика {i}",
                    image_prompt=f"img {i}",
                    animation_prompt=f"vid {i}",
                    attrs={},
                )
            )
        await session.commit()
        await session.refresh(p)
        await db_v2.backfill_project_v2(session, p)
        await session.commit()
        project_id = p.id

        # Минимальный v8-файл: лист «план» должен существовать для экспорта.
        data_dir = p.data_dir
        data_dir.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "план"
        wb.create_sheet("Общий план")
        wb.save(data_dir / "project.xlsx")
        wb.close()

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as client:
        yield client, project_id, factory

    app.dependency_overrides.clear()
    await engine.dispose()


@pytest.mark.asyncio
async def test_apply_ops_unknown_uuid_fail_closed(api_client) -> None:
    client, project_id, factory = api_client
    r = await client.post(
        f"/api/db/projects/{project_id}/apply-ops",
        json={"ops": [{"frame_uuid": "deadbeef", "fields": {"voiceover_text": "x"}}]},
    )
    assert r.status_code == 400
    assert "неизвестные frame_uuid" in r.json()["detail"]
    async with factory() as session:
        fr = (
            await session.execute(
                select(Frame).where(Frame.project_id == project_id, Frame.number == 1)
            )
        ).scalar_one()
        assert fr.voiceover_text == "реплика 1"  # не тронут


@pytest.mark.asyncio
async def test_apply_ops_empty_and_empty_fields_rejected(api_client) -> None:
    client, project_id, factory = api_client
    async with factory() as session:
        fr = (
            await session.execute(
                select(Frame).where(Frame.project_id == project_id, Frame.number == 1)
            )
        ).scalar_one()
        uuid = fr.uuid
    r_empty = await client.post(f"/api/db/projects/{project_id}/apply-ops", json={"ops": []})
    assert r_empty.status_code == 400
    r_fields = await client.post(
        f"/api/db/projects/{project_id}/apply-ops",
        json={"ops": [{"frame_uuid": uuid, "fields": {}}]},
    )
    assert r_fields.status_code == 400


@pytest.mark.asyncio
async def test_apply_ops_updates_db_versions_and_exports_xlsx(api_client, tmp_path) -> None:
    client, project_id, factory = api_client
    async with factory() as session:
        fr = (
            await session.execute(
                select(Frame).where(Frame.project_id == project_id, Frame.number == 1)
            )
        ).scalar_one()
        uuid = fr.uuid

    r = await client.post(
        f"/api/db/projects/{project_id}/apply-ops",
        json={
            "ops": [
                {
                    "frame_uuid": uuid,
                    "fields": {
                        "voiceover_text": "новый закадр",
                        "image_prompt": "новый img промт",
                        "animation_prompt": "новый vid промт",
                        "duration_seconds": 3.5,
                    },
                }
            ]
        },
    )
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["updated"] == 1
    assert data["exported"] and data["exported"]["cells"] > 0

    async with factory() as session:
        fr = (
            await session.execute(
                select(Frame).where(Frame.project_id == project_id, Frame.number == 1)
            )
        ).scalar_one()
        assert fr.voiceover_text == "новый закадр"
        assert fr.image_prompt == "новый img промт"
        assert fr.animation_prompt == "новый vid промт"
        assert fr.duration_seconds == 3.5
        active_img = (
            await session.execute(
                select(PromptVersion).where(
                    PromptVersion.frame_id == fr.id,
                    PromptVersion.kind == "img",
                    PromptVersion.is_active.is_(True),
                )
            )
        ).scalar_one()
        # активная версия совпадает с legacy-полем — нет двух правд
        assert active_img.text == fr.image_prompt

    # Excel — производный view: те же значения лежат в R45/R48/R49 колонки кадра.
    wb = load_workbook(tmp_path / "videos" / "dbv2-api" / "project.xlsx")
    try:
        ws = wb["план"]
        col = 1 + 2
        assert ws.cell(row=ROW_IMAGE_PROMPT_V8, column=col).value == "новый img промт"
        assert ws.cell(row=ROW_VIDEO_PROMPT_V8, column=col).value == "новый vid промт"
        assert ws.cell(row=ROW_VOICEOVER_V8, column=col).value == "новый закадр"
    finally:
        wb.close()


@pytest.mark.asyncio
async def test_apply_ops_human_aliases_and_unknown_field(api_client) -> None:
    """Агент пишет по-человечески — сервер переводит в канонические ключи."""
    client, project_id, factory = api_client
    async with factory() as session:
        fr = (
            await session.execute(
                select(Frame).where(Frame.project_id == project_id, Frame.number == 1)
            )
        ).scalar_one()
        uuid = fr.uuid

    r = await client.post(
        f"/api/db/projects/{project_id}/apply-ops",
        json={
            "ops": [
                {
                    "frame_uuid": uuid,
                    "fields": {
                        "закадр": "закадр по-человечески",
                        "промт_картинки": "img по-человечески",
                        "Промт Видео": "vid по-человечески",
                        "длительность": 4,
                    },
                }
            ]
        },
    )
    assert r.status_code == 200, r.text
    async with factory() as session:
        fr = (
            await session.execute(
                select(Frame).where(Frame.project_id == project_id, Frame.number == 1)
            )
        ).scalar_one()
        assert fr.voiceover_text == "закадр по-человечески"
        assert fr.image_prompt == "img по-человечески"
        assert fr.animation_prompt == "vid по-человечески"
        assert fr.duration_seconds == 4.0

    r_bad = await client.post(
        f"/api/db/projects/{project_id}/apply-ops",
        json={"ops": [{"frame_uuid": uuid, "fields": {"неизвестное_поле": "x"}}]},
    )
    assert r_bad.status_code == 400
    assert "неизвестные поля" in r_bad.json()["detail"]


@pytest.mark.asyncio
async def test_apply_ops_project_target_general_plan_exports(api_client, tmp_path) -> None:
    """target=project — поля уровня проекта (нода плана), экспорт в «Общий план»."""
    client, project_id, factory = api_client
    r = await client.post(
        f"/api/db/projects/{project_id}/apply-ops",
        json={"ops": [{"target": "project", "fields": {"общий_план": "Эпизод 1 …"}}]},
    )
    assert r.status_code == 200, r.text
    async with factory() as session:
        p = await session.get(Project, project_id)
        assert (p.meta or {}).get("general_plan") == "Эпизод 1 …"

    wb = load_workbook(tmp_path / "videos" / "dbv2-api" / "project.xlsx")
    try:
        assert wb["Общий план"]["B2"].value == "Эпизод 1 …"
    finally:
        wb.close()

    r_target = await client.post(
        f"/api/db/projects/{project_id}/apply-ops",
        json={"ops": [{"target": "галактика", "fields": {"план": "x"}}]},
    )
    assert r_target.status_code == 400


class _FakeGpt:
    last_prompt: str = ""

    def __init__(self, reply: str) -> None:
        self._reply = reply

    async def ask_fresh(self, text: str, **kw) -> str:
        type(self).last_prompt = text
        return self._reply


@pytest.mark.asyncio
async def test_orchestrator_chat_applies_ops(api_client, monkeypatch) -> None:
    """Диалог с оркестратором: ответ с {ops} применяется через db_apply."""
    client, project_id, factory = api_client
    async with factory() as session:
        fr = (
            await session.execute(
                select(Frame).where(Frame.project_id == project_id, Frame.number == 1)
            )
        ).scalar_one()
        uuid = fr.uuid

    reply = f'{{"ops":[{{"frame_uuid":"{uuid}","fields":{{"закадр":"закадр из чата"}}}}]}}'
    monkeypatch.setattr(
        "app.services.gpt_client.get_gpt_client", lambda: _FakeGpt(reply)
    )
    r = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/chat",
        json={"message": "перепиши закадр в 1 кадре", "history": []},
    )
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["error"] is None
    assert data["applied"] and data["applied"]["updated"] == 1
    async with factory() as session:
        fr = (
            await session.execute(
                select(Frame).where(Frame.project_id == project_id, Frame.number == 1)
            )
        ).scalar_one()
        assert fr.voiceover_text == "закадр из чата"


@pytest.mark.asyncio
async def test_orchestrator_chat_plain_text_no_write(api_client, monkeypatch) -> None:
    """Обычный текстовый ответ (вопрос/обсуждение) — без записи в DB."""
    client, project_id, factory = api_client
    monkeypatch.setattr(
        "app.services.gpt_client.get_gpt_client",
        lambda: _FakeGpt("В проекте 2 кадра, всё заполнено."),
    )
    r = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/chat",
        json={"message": "сколько кадров?", "history": []},
    )
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["applied"] is None and data["error"] is None
    async with factory() as session:
        fr = (
            await session.execute(
                select(Frame).where(Frame.project_id == project_id, Frame.number == 1)
            )
        ).scalar_one()
        assert fr.voiceover_text == "реплика 1"  # не тронут


@pytest.mark.asyncio
async def test_orchestrator_chat_runs_step_action(api_client, monkeypatch) -> None:
    """Оркестратор может запустить шаг пайплайна: {"actions":[{"run_step":...}]}."""
    client, project_id, factory = api_client
    monkeypatch.setattr(
        "app.services.gpt_client.get_gpt_client",
        lambda: _FakeGpt('{"actions":[{"run_step":"anim_pr"}]}'),
    )
    r = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/chat",
        json={"message": "запусти промты видео", "history": []},
    )
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["error"] is None
    # Шаг запущен по пути UI-кнопки (start_step); гард anim_pr честно ответил
    # «нечего генерировать» (нет PNG на диске) — статус не сменился, и это ок.
    assert data["actions_run"] and data["actions_run"][0]["run_step"] == "anim_pr"
    async with factory() as session:
        p = await session.get(Project, project_id)
        assert p.status is ProjectStatus.new

    monkeypatch.setattr(
        "app.services.gpt_client.get_gpt_client",
        lambda: _FakeGpt('{"actions":[{"run_step":"hack"}]}'),
    )
    r_bad = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/chat",
        json={"message": "запусти hack", "history": []},
    )
    assert r_bad.status_code == 200
    assert "неизвестный шаг" in (r_bad.json()["error"] or "")


@pytest.mark.asyncio
async def test_orchestrator_chat_settings_actions(api_client, monkeypatch, tmp_path) -> None:
    """Оркестратор управляет настройками: опции генерации, вариант промта, LLM, стоп."""
    import app.services.prompt_library as pl

    (tmp_path / "prompts" / "01_plan").mkdir(parents=True)
    (tmp_path / "prompts" / "01_plan" / "horror.md").write_text("ужас", encoding="utf-8")
    monkeypatch.setattr(pl, "PROMPTS_ROOT", tmp_path / "prompts")

    client, project_id, factory = api_client
    reply = (
        '{"actions":['
        '{"set_option":{"key":"image_generator","value":"gpt_image_2_vip"}},'
        '{"set_option":{"key":"image_resolution","value":"4k"}},'
        '{"set_option":{"key":"hero_mode","value":"no_hero"}},'
        '{"set_option":{"key":"auto_mode","value":"вкл"}},'
        '{"set_prompt":{"step":"plan","variant":"horror"}},'
        '{"set_text_llm":{"provider":"tokenrouter"}},'
        '{"stop_step":true}'
        "]}"
    )
    monkeypatch.setattr(
        "app.services.gpt_client.get_gpt_client", lambda: _FakeGpt(reply)
    )
    r = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/chat",
        json={"message": "настрой всё и останови", "history": []},
    )
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["error"] is None
    kinds = [next(iter(a)) for a in data["actions_run"]]
    assert kinds == [
        "set_option",
        "set_option",
        "set_option",
        "set_option",
        "set_prompt",
        "set_text_llm",
        "stop_step",
    ]
    async with factory() as session:
        p = await session.get(Project, project_id)
        assert p.image_generator == "gpt_image_2_vip"
        assert p.image_resolution == "4k"
        assert p.hero_mode == "no_hero"
        assert p.auto_mode is True
        assert (p.prompt_overrides or {}).get("plan") == "horror"
        assert (p.meta or {}).get("user_stop") is True


@pytest.mark.asyncio
async def test_orchestrator_chat_invalid_setting_rejected(api_client, monkeypatch) -> None:
    client, project_id, factory = api_client
    monkeypatch.setattr(
        "app.services.gpt_client.get_gpt_client",
        lambda: _FakeGpt('{"actions":[{"set_option":{"key":"image_resolution","value":"16k"}}]}'),
    )
    r = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/chat",
        json={"message": "поставь 16k", "history": []},
    )
    assert r.status_code == 200
    assert "неизвестное значение" in (r.json()["error"] or "")
    async with factory() as session:
        p = await session.get(Project, project_id)
        assert p.image_resolution is None  # не тронут


@pytest.mark.asyncio
async def test_checks_context_includes_node_runs_and_harness(api_client) -> None:
    """Контекст оркестратора: итоги harness-проверок + статусы нод (hitl тоже)."""
    from app.models import NodeRun, NodeRunStatus, WorkflowRun, WorkflowRunStatus
    from app.web.routers.db_browser import _checks_context

    client, project_id, factory = api_client
    async with factory() as session:
        p = await session.get(Project, project_id)
        wr = WorkflowRun(
            workflow_id=1, project_id=project_id, status=WorkflowRunStatus.running
        )
        session.add(wr)
        await session.flush()
        session.add(
            NodeRun(workflow_run_id=wr.id, node_key="n1", node_type="plan", status=NodeRunStatus.done)
        )
        session.add(
            NodeRun(
                workflow_run_id=wr.id,
                node_key="n2",
                node_type="img",
                status=NodeRunStatus.failed,
                error="boom",
            )
        )
        session.add(
            NodeRun(
                workflow_run_id=wr.id,
                node_key="n3",
                node_type="hitl_videos",
                status=NodeRunStatus.waiting_hitl,
            )
        )
        meta = dict(p.meta or {})
        meta["ops_telemetry"] = {
            "updated_at": "t",
            "last_outcome": "verify_fail",
            "checks": [{"name": "scenes_png", "ok": False, "detail": "n=0"}],
            "next_action": "repair:img",
        }
        p.meta = meta
        await session.commit()
        lines = await _checks_context(session, p)
    text = "\n".join(lines)
    assert "ПРОВЕРКИ HARNESS" in text
    assert "scenes_png: НЕ ОК" in text
    assert "next_action: repair:img" in text
    assert "СТАТУСЫ НОД" in text
    assert "plan: готово" in text
    assert "img: ОШИБКА — boom" in text
    assert "hitl_videos: ждёт аппрува человека" in text


@pytest.mark.asyncio
async def test_orchestrator_chat_open_ui_action(api_client, monkeypatch) -> None:
    """open_ui step_prompts → ui_actions с node_type для фронта; мусор отклонён."""
    client, project_id, factory = api_client
    monkeypatch.setattr(
        "app.services.gpt_client.get_gpt_client",
        lambda: _FakeGpt('{"actions":[{"open_ui":{"kind":"step_prompts","step":"anim_pr"}}]}'),
    )
    r = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/chat",
        json={"message": "покажи варианты промтов видео", "history": []},
    )
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["error"] is None
    assert data["ui_actions"] == [
        {"kind": "step_prompts", "step": "anim_pr", "node_type": "animation_prompts"}
    ]

    monkeypatch.setattr(
        "app.services.gpt_client.get_gpt_client",
        lambda: _FakeGpt('{"actions":[{"open_ui":{"kind":"hack","step":"anim_pr"}}]}'),
    )
    r_bad = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/chat",
        json={"message": "x", "history": []},
    )
    assert "неизвестный kind" in (r_bad.json()["error"] or "")


@pytest.mark.asyncio
async def test_orchestrator_chat_open_ui_all_kinds(api_client, monkeypatch) -> None:
    """open_ui: node_studio/topic/baza — валидны; hitl без pending — ошибка."""
    client, project_id, factory = api_client
    monkeypatch.setattr(
        "app.services.gpt_client.get_gpt_client",
        lambda: _FakeGpt(
            '{"actions":['
            '{"open_ui":{"kind":"node_studio","node_type":"animation_prompts"}},'
            '{"open_ui":{"kind":"topic"}},'
            '{"open_ui":{"kind":"baza"}}'
            "]}"
        ),
    )
    r = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/chat",
        json={"message": "открой окна", "history": []},
    )
    assert r.status_code == 200, r.text
    kinds = [u["kind"] for u in r.json()["ui_actions"]]
    assert kinds == ["node_studio", "topic", "baza"]
    assert r.json()["ui_actions"][0]["node_type"] == "animation_prompts"

    monkeypatch.setattr(
        "app.services.gpt_client.get_gpt_client",
        lambda: _FakeGpt('{"actions":[{"open_ui":{"kind":"hitl","node_type":"hitl_videos"}}]}'),
    )
    r_bad = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/chat",
        json={"message": "открой hitl", "history": []},
    )
    assert "нет ожидающих HITL" in (r_bad.json()["error"] or "")

    # Конфиг-ноды (topic/storage) не имеют студии промтов — вежливый отказ.
    monkeypatch.setattr(
        "app.services.gpt_client.get_gpt_client",
        lambda: _FakeGpt('{"actions":[{"open_ui":{"kind":"node_studio","node_type":"topic"}}]}'),
    )
    r_cfg = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/chat",
        json={"message": "открой студию темы", "history": []},
    )
    assert "нет промтов/студии" in (r_cfg.json()["error"] or "")


@pytest.mark.asyncio
async def test_orchestrator_chat_hitl_decision_and_topic(api_client, monkeypatch) -> None:
    """hitl_decision approve решает pending-запрос; set_topic меняет тему."""
    from app.models import HITLDecision, HITLKind, HITLRequest

    client, project_id, factory = api_client
    async with factory() as session:
        session.add(
            HITLRequest(project_id=project_id, kind=HITLKind.approve_images, payload={})
        )
        await session.commit()

    monkeypatch.setattr(
        "app.services.gpt_client.get_gpt_client",
        lambda: _FakeGpt(
            '{"actions":[{"hitl_decision":"approve"},{"set_topic":"Новая тема"}]}'
        ),
    )
    r = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/chat",
        json={"message": "аппрувь и поменяй тему", "history": []},
    )
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["error"] is None
    kinds = [next(iter(a)) for a in data["actions_run"]]
    assert kinds == ["hitl_decision", "set_topic"]
    async with factory() as session:
        p = await session.get(Project, project_id)
        assert p.topic == "Новая тема"
        req = (
            await session.execute(
                select(HITLRequest).where(HITLRequest.project_id == project_id)
            )
        ).scalar_one()
        assert req.decision is HITLDecision.approved


@pytest.mark.asyncio
async def test_orchestrator_chat_canvas_nodes_and_node_key(api_client, monkeypatch) -> None:
    """Контекст содержит ноды канваса с названиями; open_ui принимает node_key."""
    from app.models import Workflow, WorkflowRun

    client, project_id, factory = api_client
    async with factory() as session:
        wf = Workflow(
            name="default",
            is_default=True,
            nodes=[
                {"id": "n_plan", "type": "plan", "position": {}, "data": {"label": "Сценарий"}},
                {"id": "n_script", "type": "script", "position": {}, "data": {"label": "Закадровый текст"}},
            ],
            edges=[],
        )
        session.add(wf)
        await session.flush()
        session.add(
            WorkflowRun(
                workflow_id=wf.id,
                project_id=project_id,
                nodes_snapshot=wf.nodes,
                edges_snapshot=[],
            )
        )
        await session.commit()

    monkeypatch.setattr(
        "app.services.gpt_client.get_gpt_client",
        lambda: _FakeGpt('{"actions":[{"open_ui":{"kind":"node_studio","node_key":"n_plan"}}]}'),
    )
    r = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/chat",
        json={"message": "открой сценарий", "history": []},
    )
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["error"] is None
    assert data["ui_actions"][0]["node_key"] == "n_plan"
    assert data["ui_actions"][0]["node_type"] == "plan"
    # контекст видит названия нод канваса
    assert "«Сценарий»" in _FakeGpt.last_prompt
    assert "n_plan" in _FakeGpt.last_prompt

    monkeypatch.setattr(
        "app.services.gpt_client.get_gpt_client",
        lambda: _FakeGpt('{"actions":[{"open_ui":{"kind":"node_studio","node_key":"n_hack"}}]}'),
    )
    r_bad = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/chat",
        json={"message": "x", "history": []},
    )
    assert "неизвестный node_key" in (r_bad.json()["error"] or "")


@pytest.mark.asyncio
async def test_orchestrator_chat_create_project(api_client, monkeypatch) -> None:
    """create_project: оркестратор сам создаёт проект (без окна генерации)."""
    client, project_id, factory = api_client
    monkeypatch.setattr(
        "app.services.gpt_client.get_gpt_client",
        lambda: _FakeGpt('{"actions":[{"create_project":{"title":"Тест оркестра"}}]}'),
    )
    r = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/chat",
        json={"message": "создай новый проект тест оркестра", "history": []},
    )
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["error"] is None
    assert data["actions_run"][0]["create_project"].startswith("#")
    assert data["ui_actions"][0]["kind"] == "open_project"
    new_id = data["ui_actions"][0]["project_id"]
    assert new_id != project_id
    async with factory() as session:
        p = await session.get(Project, new_id)
        assert p is not None and p.title == "Тест оркестра"
        assert (p.data_dir / "project.xlsx").is_file()


@pytest.mark.asyncio
async def test_orchestrator_chat_create_child_and_catalog(api_client, monkeypatch) -> None:
    """create_child + каталог проектов в промпте оркестратора."""
    client, project_id, factory = api_client
    async with factory() as session:
        parent = await session.get(Project, project_id)
        assert parent is not None
        parent.title = "История ведьм"
        parent.slug = "istoriya-vedm-test"
        parent.hero_mode = "hero"
        parent.image_generator = "gpt_image_2"
        parent.prompt_overrides = {"img_pr": "default", "anim_pr": "default"}
        parent.gpt_text_overrides = {"plan": "мастер-промт родителя"}
        await session.commit()

    monkeypatch.setattr(
        "app.services.gpt_client.get_gpt_client",
        lambda: _FakeGpt(
            '{"actions":[{"create_child":{"parent_title":"История ведьм"}}]}'
        ),
    )
    # чат с другого проекта — parent_title должен найти #project_id
    r = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/chat",
        json={"message": "найди история ведьм и сделай дочерний", "history": []},
    )
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["error"] is None, data
    assert "create_child" in data["actions_run"][0]
    assert f"parent #{project_id}" in data["actions_run"][0]["create_child"]
    assert data["ui_actions"][0]["kind"] == "open_project"
    child_id = data["ui_actions"][0]["project_id"]
    assert child_id != project_id
    # каталог проектов в системном промпте
    assert "ПРОЕКТЫ В СТУДИИ" in (_FakeGpt.last_prompt or "")
    assert "История ведьм" in (_FakeGpt.last_prompt or "")

    async with factory() as session:
        child = await session.get(Project, child_id)
        parent = await session.get(Project, project_id)
        assert child is not None and parent is not None
        assert (child.meta or {}).get("mass_parent_id") == project_id
        assert child.prompt_overrides == parent.prompt_overrides
        assert child.hero_mode == parent.hero_mode
        assert child.image_generator == parent.image_generator
        # контент кадров не копируем
        n_frames = (
            await session.execute(
                select(func.count()).select_from(Frame).where(Frame.project_id == child_id)
            )
        ).scalar_one()
        assert int(n_frames or 0) == 0


@pytest.mark.asyncio
async def test_orchestrator_chat_add_node_each_and_single(api_client, monkeypatch) -> None:
    """add_node: hitl_gate после каждой рабочей ноды; цепочка остаётся линейной."""
    from app.models import Workflow
    from app.web.routers.db_browser import _linear_order

    client, project_id, factory = api_client
    async with factory() as session:
        session.add(
            Workflow(
                name="default",
                is_default=True,
                nodes=[
                    {"id": "n_plan", "type": "plan", "position": {"x": 0.0, "y": 0.0}, "data": {"label": "Сценарий"}},
                    {"id": "n_script", "type": "script", "position": {"x": 290.0, "y": 0.0}, "data": {"label": "Закадровый"}},
                    {"id": "n_img", "type": "images", "position": {"x": 580.0, "y": 0.0}, "data": {"label": "Картинки"}},
                ],
                edges=[
                    {"id": "e_0", "source": "n_plan", "target": "n_script"},
                    {"id": "e_1", "source": "n_script", "target": "n_img"},
                ],
            )
        )
        await session.commit()

    monkeypatch.setattr(
        "app.services.gpt_client.get_gpt_client",
        lambda: _FakeGpt('{"actions":[{"add_node":{"node_type":"hitl_gate","after":"each"}}]}'),
    )
    r = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/chat",
        json={"message": "к каждой ноде добавь ноду проверки", "history": []},
    )
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["error"] is None
    assert "×3" in data["actions_run"][0]["add_node"]
    async with factory() as session:
        p = await session.get(Project, project_id)
        g = (p.meta or {}).get("canvas_graph") or {}
        types = [n["type"] for n in g["nodes"]]
        assert types.count("hitl_gate") == 3
        assert len(g["nodes"]) == 6
        assert len(_linear_order(g["nodes"], g["edges"])) == 6  # цепочка линейна

    # Одиночная вставка после конкретной ноды.
    monkeypatch.setattr(
        "app.services.gpt_client.get_gpt_client",
        lambda: _FakeGpt('{"actions":[{"add_node":{"node_type":"hitl_gate","after":"n_plan"}}]}'),
    )
    r2 = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/chat",
        json={"message": "добавь проверку после плана", "history": []},
    )
    assert r2.status_code == 200
    assert r2.json()["error"] is None
    async with factory() as session:
        p = await session.get(Project, project_id)
        g = (p.meta or {}).get("canvas_graph") or {}
        assert len(g["nodes"]) == 7
        assert len(_linear_order(g["nodes"], g["edges"])) == 7

    # Идемпотентность: повторное «each» не плодит дубли.
    monkeypatch.setattr(
        "app.services.gpt_client.get_gpt_client",
        lambda: _FakeGpt('{"actions":[{"add_node":{"node_type":"hitl_gate","after":"each"}}]}'),
    )
    r3 = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/chat",
        json={"message": "к каждой ноде добавь проверку ещё раз", "history": []},
    )
    assert r3.status_code == 200
    assert r3.json()["error"] is None
    assert "дублей нет" in r3.json()["actions_run"][0]["add_node"]
    async with factory() as session:
        p = await session.get(Project, project_id)
        g = (p.meta or {}).get("canvas_graph") or {}
        assert len(g["nodes"]) == 7  # не выросло


@pytest.mark.asyncio
async def test_orchestrator_chat_remove_node(api_client, monkeypatch) -> None:
    """remove_node: удаляет добавленные оркестратором ноды типа, цепочка перешита."""
    from app.models import Workflow
    from app.web.routers.db_browser import _linear_order

    client, project_id, factory = api_client
    async with factory() as session:
        session.add(
            Workflow(
                name="default",
                is_default=True,
                nodes=[
                    {"id": "n_plan", "type": "plan", "position": {"x": 0.0, "y": 0.0}, "data": {"label": "Сценарий"}},
                    {"id": "n_hitl_gate_old", "type": "hitl_gate", "position": {"x": 145.0, "y": 140.0}, "data": {"label": "Проверка"}},
                    {"id": "n_script", "type": "script", "position": {"x": 290.0, "y": 0.0}, "data": {"label": "Закадровый"}},
                ],
                edges=[
                    {"id": "e_0", "source": "n_plan", "target": "n_hitl_gate_old"},
                    {"id": "e_1", "source": "n_hitl_gate_old", "target": "n_script"},
                ],
            )
        )
        await session.commit()

    monkeypatch.setattr(
        "app.services.gpt_client.get_gpt_client",
        lambda: _FakeGpt('{"actions":[{"add_node":{"node_type":"hitl_gate","after":"each"}}]}'),
    )
    r1 = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/chat",
        json={"message": "к каждой ноде добавь проверку", "history": []},
    )
    assert r1.json()["error"] is None
    # за n_plan уже стоит hitl-нода → дубль не добавился, только после n_script
    assert "×1" in r1.json()["actions_run"][0]["add_node"]

    # Удаление — только с подтверждением: чат делает preview, НЕ удаляет.
    monkeypatch.setattr(
        "app.services.gpt_client.get_gpt_client",
        lambda: _FakeGpt('{"actions":[{"remove_node":{"node_type":"hitl_gate"}}]}'),
    )
    r2 = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/chat",
        json={"message": "удали все ноды проверки", "history": []},
    )
    assert r2.status_code == 200, r2.text
    assert r2.json()["error"] is None
    pend = r2.json()["pending_confirm"]
    assert pend and pend[0]["count"] == 1 and pend[0]["node_type"] == "hitl_gate"
    async with factory() as session:
        p = await session.get(Project, project_id)
        g = (p.meta or {}).get("canvas_graph") or {}
        assert any(
            n["id"].startswith("n_hitl_gate_") and n["id"] != "n_hitl_gate_old"
            for n in g["nodes"]
        )  # ещё НЕ удалено

    # Подтверждение — удаляет.
    r3 = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/confirm-remove",
        json={"node_type": "hitl_gate"},
    )
    assert r3.status_code == 200, r3.text
    assert "удалено 1" in r3.json()["remove_node"]
    async with factory() as session:
        p = await session.get(Project, project_id)
        g = (p.meta or {}).get("canvas_graph") or {}
        ids = [n["id"] for n in g["nodes"]]
        # родная (не оркестраторская) hitl-нода выжила, добавленная удалена
        assert "n_hitl_gate_old" in ids
        assert not any(i.startswith("n_hitl_gate_") and i != "n_hitl_gate_old" for i in ids)
        assert len(_linear_order(g["nodes"], g["edges"])) == 3

    # По конкретному ключу: preview → confirm; неизвестный ключ → ошибка сразу.
    monkeypatch.setattr(
        "app.services.gpt_client.get_gpt_client",
        lambda: _FakeGpt('{"actions":[{"remove_node":{"node_key":"n_hitl_gate_old"}}]}'),
    )
    r4 = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/chat",
        json={"message": "удали старую проверку", "history": []},
    )
    assert r4.json()["pending_confirm"][0]["node_key"] == "n_hitl_gate_old"
    r5 = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/confirm-remove",
        json={"node_key": "n_hitl_gate_old"},
    )
    assert r5.status_code == 200
    monkeypatch.setattr(
        "app.services.gpt_client.get_gpt_client",
        lambda: _FakeGpt('{"actions":[{"remove_node":{"node_key":"n_hack"}}]}'),
    )
    r6 = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/chat",
        json={"message": "удали n_hack", "history": []},
    )
    assert "неизвестный node_key" in (r6.json()["error"] or "")


@pytest.mark.asyncio
async def test_remove_node_consecutive_keeps_chain_linear(api_client) -> None:
    """Удаление подряд идущих нод не рвёт цепочку (регрессия «стопок» проверок)."""
    from app.web.routers.db_browser import _linear_order

    client, project_id, factory = api_client
    async with factory() as session:
        p = await session.get(Project, project_id)
        meta = dict(p.meta or {})
        meta["canvas_graph"] = {
            "nodes": [
                {"id": "n_plan", "type": "plan", "position": {"x": 0.0, "y": 0.0}, "data": {}},
                {"id": "n_hitl_gate_1", "type": "hitl_gate", "position": {"x": 100.0, "y": 160.0}, "data": {"description": "добавлено оркестратором"}},
                {"id": "n_hitl_gate_2", "type": "hitl_gate", "position": {"x": 200.0, "y": 160.0}, "data": {"description": "добавлено оркестратором"}},
                {"id": "n_script", "type": "script", "position": {"x": 290.0, "y": 0.0}, "data": {}},
            ],
            "edges": [
                {"id": "e0", "source": "n_plan", "target": "n_hitl_gate_1"},
                {"id": "e1", "source": "n_hitl_gate_1", "target": "n_hitl_gate_2"},
                {"id": "e2", "source": "n_hitl_gate_2", "target": "n_script"},
            ],
        }
        p.meta = meta
        await session.commit()

    r = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/confirm-remove",
        json={"node_type": "hitl_gate"},
    )
    assert r.status_code == 200, r.text
    assert "удалено 2" in r.json()["remove_node"]
    async with factory() as session:
        p = await session.get(Project, project_id)
        g = (p.meta or {})["canvas_graph"]
        assert _linear_order(g["nodes"], g["edges"]) == ["n_plan", "n_script"]


@pytest.mark.asyncio
async def test_repair_graph_rebuilds_broken_chain(api_client) -> None:
    """repair_graph: разорванная цепочка пересобирается слева направо."""
    from app.web.routers.db_browser import _linear_order

    client, project_id, factory = api_client
    async with factory() as session:
        p = await session.get(Project, project_id)
        meta = dict(p.meta or {})
        meta["canvas_graph"] = {
            "nodes": [
                {"id": "n_plan", "type": "plan", "position": {"x": 0.0, "y": 0.0}, "data": {}},
                {"id": "n_script", "type": "script", "position": {"x": 290.0, "y": 0.0}, "data": {}},
                {"id": "n_img", "type": "images", "position": {"x": 580.0, "y": 0.0}, "data": {}},
            ],
            # разрыв: plan ни с кем не связан
            "edges": [{"id": "e1", "source": "n_script", "target": "n_img"}],
        }
        p.meta = meta
        await session.commit()

    from app.web.routers.db_browser import _apply_repair_graph

    async with factory() as session:
        p = await session.get(Project, project_id)
        res = await _apply_repair_graph(session, p)
        await session.commit()
    assert "пересобрана" in res["repair_graph"]
    async with factory() as session:
        p = await session.get(Project, project_id)
        g = (p.meta or {})["canvas_graph"]
        assert _linear_order(g["nodes"], g["edges"]) == ["n_plan", "n_script", "n_img"]


@pytest.mark.asyncio
async def test_add_node_excel_gpt_gets_slot_index(api_client, monkeypatch) -> None:
    """excel_gpt-ноды от оркестратора получают slotIndex (для шагов enrich)."""
    from app.models import Workflow

    client, project_id, factory = api_client
    async with factory() as session:
        session.add(
            Workflow(
                name="default",
                is_default=True,
                nodes=[
                    {"id": "n_plan", "type": "plan", "position": {"x": 0.0, "y": 0.0}, "data": {}},
                    {"id": "n_script", "type": "script", "position": {"x": 290.0, "y": 0.0}, "data": {}},
                ],
                edges=[{"id": "e0", "source": "n_plan", "target": "n_script"}],
            )
        )
        await session.commit()
    monkeypatch.setattr(
        "app.services.gpt_client.get_gpt_client",
        lambda: _FakeGpt('{"actions":[{"add_node":{"node_type":"excel_gpt","after":"n_plan"}}]}'),
    )
    r = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/chat",
        json={"message": "добавь работу с гпт после плана", "history": []},
    )
    assert r.status_code == 200, r.text
    assert r.json()["error"] is None
    async with factory() as session:
        p = await session.get(Project, project_id)
        g = (p.meta or {})["canvas_graph"]
        new_gpt = [n for n in g["nodes"] if n["type"] == "excel_gpt"]
        assert len(new_gpt) == 1
        assert (new_gpt[0]["data"] or {}).get("slotIndex") == 1


@pytest.mark.asyncio
async def test_add_node_labels_by_parent_and_rename(api_client, monkeypatch) -> None:
    """Новые проверки подписаны по родителю; rename_node меняет подпись."""
    from app.models import Workflow

    client, project_id, factory = api_client
    async with factory() as session:
        session.add(
            Workflow(
                name="default",
                is_default=True,
                nodes=[
                    {"id": "n_plan", "type": "plan", "position": {"x": 0.0, "y": 0.0}, "data": {"label": "Сценарий"}},
                    {"id": "n_script", "type": "script", "position": {"x": 290.0, "y": 0.0}, "data": {"label": "Закадровый текст"}},
                ],
                edges=[{"id": "e0", "source": "n_plan", "target": "n_script"}],
            )
        )
        await session.commit()
    monkeypatch.setattr(
        "app.services.gpt_client.get_gpt_client",
        lambda: _FakeGpt('{"actions":[{"add_node":{"node_type":"excel_gpt","after":"each"}}]}'),
    )
    r = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/chat",
        json={"message": "добавь работу с гпт после каждой", "history": []},
    )
    assert r.status_code == 200, r.text
    assert r.json()["error"] is None
    async with factory() as session:
        p = await session.get(Project, project_id)
        g = (p.meta or {})["canvas_graph"]
        labels = [
            str((n.get("data") or {}).get("label") or "")
            for n in g["nodes"]
            if str(n.get("type")) == "excel_gpt"
        ]
        assert any("Сценарий" in lab for lab in labels)
        assert any("Закадровый текст" in lab for lab in labels)

    monkeypatch.setattr(
        "app.services.gpt_client.get_gpt_client",
        lambda: _FakeGpt('{"actions":[{"rename_node":{"node_key":"n_plan","label":"План ролика"}}]}'),
    )
    r2 = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/chat",
        json={"message": "переименуй n_plan", "history": []},
    )
    assert r2.status_code == 200, r2.text
    assert r2.json()["error"] is None
    async with factory() as session:
        p = await session.get(Project, project_id)
        g = (p.meta or {})["canvas_graph"]
        plan = next(n for n in g["nodes"] if n["id"] == "n_plan")
        assert (plan.get("data") or {}).get("label") == "План ролика"


@pytest.mark.asyncio
async def test_remove_node_only_duplicates(api_client) -> None:
    """only=duplicates: удаляется только вторая подряд проверка, остальные целы."""
    from app.web.routers.db_browser import _linear_order

    client, project_id, factory = api_client
    async with factory() as session:
        p = await session.get(Project, project_id)
        meta = dict(p.meta or {})
        meta["canvas_graph"] = {
            "nodes": [
                {"id": "n_plan", "type": "plan", "position": {"x": 0.0, "y": 0.0}, "data": {}},
                {"id": "n_hitl_gate_1", "type": "hitl_gate", "position": {"x": 100.0, "y": 160.0}, "data": {"description": "добавлено оркестратором"}},
                {"id": "n_hitl_gate_2", "type": "hitl_gate", "position": {"x": 200.0, "y": 160.0}, "data": {"description": "добавлено оркестратором"}},
                {"id": "n_script", "type": "script", "position": {"x": 290.0, "y": 0.0}, "data": {}},
            ],
            "edges": [
                {"id": "e0", "source": "n_plan", "target": "n_hitl_gate_1"},
                {"id": "e1", "source": "n_hitl_gate_1", "target": "n_hitl_gate_2"},
                {"id": "e2", "source": "n_hitl_gate_2", "target": "n_script"},
            ],
        }
        p.meta = meta
        await session.commit()

    r = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/confirm-remove",
        json={"node_type": "hitl_gate", "only": "duplicates"},
    )
    assert r.status_code == 200, r.text
    assert "удалено 1" in r.json()["remove_node"]
    async with factory() as session:
        p = await session.get(Project, project_id)
        g = (p.meta or {})["canvas_graph"]
        ids = [n["id"] for n in g["nodes"]]
        assert "n_hitl_gate_1" in ids and "n_hitl_gate_2" not in ids
        assert _linear_order(g["nodes"], g["edges"]) == ["n_plan", "n_hitl_gate_1", "n_script"]


@pytest.mark.asyncio
async def test_repair_graph_places_checks_after_parents(api_client) -> None:
    """Ряд проверок под пайплайном: после repair каждая за своим родителем."""
    from app.web.routers.db_browser import _apply_repair_graph, _linear_order

    client, project_id, factory = api_client
    async with factory() as session:
        p = await session.get(Project, project_id)
        meta = dict(p.meta or {})
        meta["canvas_graph"] = {
            "nodes": [
                {"id": "n_plan", "type": "plan", "position": {"x": 0.0, "y": 0.0}, "data": {"label": "Сценарий"}},
                {"id": "n_script", "type": "script", "position": {"x": 290.0, "y": 0.0}, "data": {"label": "Закадровый текст"}},
                {"id": "n_img", "type": "images", "position": {"x": 580.0, "y": 0.0}, "data": {"label": "Картинки"}},
                {"id": "n_excel_gpt_1", "type": "excel_gpt", "position": {"x": 0.0, "y": 160.0}, "data": {"label": "Работа с GPT — Сценарий"}},
                {"id": "n_excel_gpt_2", "type": "excel_gpt", "position": {"x": 200.0, "y": 160.0}, "data": {"label": "Работа с GPT — Картинки"}},
            ],
            "edges": [],
        }
        p.meta = meta
        await session.commit()
    async with factory() as session:
        p = await session.get(Project, project_id)
        await _apply_repair_graph(session, p)
        await session.commit()
    async with factory() as session:
        p = await session.get(Project, project_id)
        g = (p.meta or {})["canvas_graph"]
        assert _linear_order(g["nodes"], g["edges"]) == [
            "n_plan",
            "n_excel_gpt_1",
            "n_script",
            "n_img",
            "n_excel_gpt_2",
        ]


@pytest.mark.asyncio
async def test_repair_then_add_each_works(api_client, monkeypatch) -> None:
    """Сломанный граф: repair_graph ПЕРВЫМ действием, затем add_node each — ок."""
    from app.web.routers.db_browser import _linear_order

    client, project_id, factory = api_client
    async with factory() as session:
        p = await session.get(Project, project_id)
        meta = dict(p.meta or {})
        meta["canvas_graph"] = {
            "nodes": [
                {"id": "n_plan", "type": "plan", "position": {"x": 0.0, "y": 0.0}, "data": {"label": "Сценарий"}},
                {"id": "n_script", "type": "script", "position": {"x": 290.0, "y": 0.0}, "data": {"label": "Закадровый текст"}},
            ],
            "edges": [],  # разрыв
        }
        p.meta = meta
        await session.commit()

    monkeypatch.setattr(
        "app.services.gpt_client.get_gpt_client",
        lambda: _FakeGpt(
            '{"actions":[{"repair_graph":true},{"add_node":{"node_type":"excel_gpt","after":"each"}}]}'
        ),
    )
    r = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/chat",
        json={"message": "почини и добавь проверки", "history": []},
    )
    assert r.status_code == 200, r.text
    assert r.json()["error"] is None
    async with factory() as session:
        p = await session.get(Project, project_id)
        g = (p.meta or {})["canvas_graph"]
        order = _linear_order(g["nodes"], g["edges"])
        assert len(order) == 4
        types = [n["type"] for n in g["nodes"]]
        assert types.count("excel_gpt") == 2


@pytest.mark.asyncio
async def test_central_harness_gate_blocks_and_pauses(api_client, monkeypatch, tmp_path) -> None:
    """Центральный гейт: плохие данные → hold ×3 → paused; хорошие → пропуск."""
    from app.orchestrator.auto_advance import _harness_gate
    from app.settings import settings

    monkeypatch.setattr(settings, "harness_gate_disabled", False)
    client, project_id, factory = api_client
    async with factory() as session:
        p = await session.get(Project, project_id)
        p.status = ProjectStatus.plan_ready
        await session.commit()
        assert await _harness_gate(session, p, p.status) is True  # данные ок

    (tmp_path / "videos" / "dbv2-api" / "project.xlsx").unlink()
    async with factory() as session:
        p = await session.get(Project, project_id)
        await session.refresh(p)
        assert await _harness_gate(session, p, p.status) is False
        assert p.status is ProjectStatus.plan_ready  # hold, ещё не paused
        assert await _harness_gate(session, p, p.status) is False
        assert await _harness_gate(session, p, p.status) is False
        assert p.status is ProjectStatus.paused  # 3 фейла подряд → paused
        assert "harness gate" in (p.meta or {}).get("auto_paused_reason", "")


@pytest.mark.asyncio
async def test_add_node_each_gives_check_after_work_excel_gpt(api_client) -> None:
    """Рабочая нода excel_gpt («Тест») получает проверку; двух подряд — нет."""
    from app.web.routers.db_browser import _apply_add_node, _linear_order

    client, project_id, factory = api_client
    async with factory() as session:
        p = await session.get(Project, project_id)
        meta = dict(p.meta or {})
        meta["canvas_graph"] = {
            "nodes": [
                {"id": "n_plan", "type": "plan", "position": {"x": 0.0, "y": 0.0}, "data": {"label": "Сценарий"}},
                {"id": "n_script", "type": "script", "position": {"x": 290.0, "y": 0.0}, "data": {"label": "Закадровый"}},
                {"id": "n_excel_gpt_1", "type": "excel_gpt", "position": {"x": 580.0, "y": 0.0}, "data": {"label": "Тест", "slotIndex": 1}},
                {"id": "n_images", "type": "images", "position": {"x": 870.0, "y": 0.0}, "data": {"label": "Картинки"}},
            ],
            "edges": [
                {"id": "e0", "source": "n_plan", "target": "n_script"},
                {"id": "e1", "source": "n_script", "target": "n_excel_gpt_1"},
                {"id": "e2", "source": "n_excel_gpt_1", "target": "n_images"},
            ],
        }
        p.meta = meta
        await session.commit()

        res = await _apply_add_node(
            session, p, {"node_type": "excel_gpt", "after": "each"}
        )
        await session.commit()
        # ВСЕ ноды получают проверку: plan, script, «Тест», images (хвост).
        assert "×4" in res["add_node"]

        g = (p.meta or {})["canvas_graph"]
        order = _linear_order(g["nodes"], g["edges"])
        type_by_id = {n["id"]: n["type"] for n in g["nodes"]}
        marked = {
            n["id"]
            for n in g["nodes"]
            if str((n.get("data") or {}).get("description") or "")
            == "добавлено оркестратором"
        }
        # после n_script стоит проверка (даже если дальше рабочая excel_gpt)
        idx_s = order.index("n_script")
        assert order[idx_s + 1] in marked
        # после «Теста» (рабочая excel_gpt) стоит проверка
        idx = order.index("n_excel_gpt_1")
        assert type_by_id[order[idx + 1]] == "excel_gpt"
        assert order[idx + 1] in marked
        # двух ПРОВЕРОК подряд нет (проверка = маркер «добавлено оркестратором»;
        # рабочая «Тест» + её проверка — одинаковый тип, это законно)
        for a, b in zip(order, order[1:], strict=False):
            assert not (a in marked and b in marked)


@pytest.mark.asyncio
async def test_orchestrator_chat_run_harness_action(api_client, monkeypatch) -> None:
    """run_harness → свежий прогон проверок, итог в actions_run."""
    client, project_id, factory = api_client
    monkeypatch.setattr(
        "app.services.gpt_client.get_gpt_client",
        lambda: _FakeGpt('{"actions":[{"run_harness":true}]}'),
    )
    r = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/chat",
        json={"message": "прогони проверки", "history": []},
    )
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["error"] is None
    assert data["actions_run"] and "run_harness" in data["actions_run"][0]
    async with factory() as session:
        p = await session.get(Project, project_id)
        tel = (p.meta or {}).get("ops_telemetry") or {}
        assert tel.get("checks")  # телеметрия обновлена — контекст увидит итоги


@pytest.mark.asyncio
async def test_orchestrator_chat_diagnostics_in_context(api_client, monkeypatch) -> None:
    """Ошибки нод попадают в контекст (режим кодинг-агента для фикса багов)."""
    from app.models import NodeRun, NodeRunStatus, WorkflowRun, WorkflowRunStatus

    client, project_id, factory = api_client
    async with factory() as session:
        wr = WorkflowRun(
            workflow_id=1, project_id=project_id, status=WorkflowRunStatus.failed
        )
        session.add(wr)
        await session.flush()
        session.add(
            NodeRun(
                workflow_run_id=wr.id,
                node_key="n9",
                node_type="img",
                status=NodeRunStatus.failed,
                error="Outsee timeout",
            )
        )
        await session.commit()
    monkeypatch.setattr(
        "app.services.gpt_client.get_gpt_client",
        lambda: _FakeGpt("Смотрю ошибку…"),
    )
    r = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/chat",
        json={"message": "почему img упал?", "history": []},
    )
    assert r.status_code == 200
    assert "ДИАГНОСТИКА" in _FakeGpt.last_prompt
    assert "ОШИБКИ НОД" in _FakeGpt.last_prompt
    assert "Outsee timeout" in _FakeGpt.last_prompt
    assert "кодинг-агент" in _FakeGpt.last_prompt


@pytest.mark.asyncio
async def test_orchestrator_chat_writes_feedback_journal(api_client, monkeypatch) -> None:
    """Каждый диалог пишется в ops-журнал (контур самообучения) и читается API."""
    client, project_id, factory = api_client
    monkeypatch.setattr(
        "app.services.gpt_client.get_gpt_client",
        lambda: _FakeGpt("просто ответ"),
    )
    r = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/chat",
        json={"message": "привет оркестратор", "history": []},
    )
    assert r.status_code == 200
    fb = await client.get(f"/api/db/projects/{project_id}/orchestrator/feedback")
    assert fb.status_code == 200
    data = fb.json()
    assert data["count"] >= 1
    last = data["events"][-1]
    assert last["type"] == "orchestrator_chat"
    assert last["message"] == "привет оркестратор"
    assert last["reply"] == "просто ответ"


@pytest.mark.asyncio
async def test_orchestrator_chat_without_gpt_key_503(api_client, monkeypatch) -> None:
    """Без GPT-ключа — вежливая 503 с причиной, не 500/404."""
    client, project_id, factory = api_client
    for var in ("GPT_API_KEY", "TOKENROUTER_API_KEY"):
        monkeypatch.delenv(var, raising=False)
    r = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/chat",
        json={"message": "привет", "history": []},
    )
    assert r.status_code == 503
    assert "не настроен" in r.json()["detail"]


def test_extract_apply_ops_json_variants() -> None:
    raw = '{"ops":[{"frame_uuid":"u1","fields":{"закадр":"x"}}]}'
    assert db_apply.extract_apply_ops_json(raw)["ops"][0]["frame_uuid"] == "u1"
    fenced = 'Вот результат:\n```json\n{"ops":[{"target":"project","fields":{"план":"y"}}]}\n```\nГотово.'
    assert db_apply.extract_apply_ops_json(fenced)["ops"][0]["target"] == "project"
    assert db_apply.extract_apply_ops_json("просто проза без джейсона") is None
    assert db_apply.extract_apply_ops_json('{"not_ops": true}') is None
    # сревис и endpoint используют одни алиасы
    assert db_apply.FIELD_ALIASES["промт_картинки"] == "image_prompt"
    assert db_apply.FIELD_ALIASES["закадр"] == "voiceover_text"


@pytest.mark.asyncio
async def test_graph_includes_harness_summary(api_client) -> None:
    """Граф отдаёт сводку последних проверок (чип «Проверки» в «Базе»)."""
    client, project_id, factory = api_client
    async with factory() as session:
        p = await session.get(Project, project_id)
        meta = dict(p.meta or {})
        meta["ops_telemetry"] = {
            "updated_at": "t",
            "last_outcome": "verify_fail",
            "checks": [
                {"name": "scenes_png", "ok": False, "detail": "n=0"},
                {"name": "project_xlsx", "ok": True, "detail": "ok"},
            ],
            "next_action": "repair:img",
        }
        p.meta = meta
        await session.commit()
    r = await client.get(f"/api/db/projects/{project_id}/graph")
    assert r.status_code == 200
    h = r.json()["harness"]
    assert h["outcome"] == "verify_fail"
    assert h["failed"] == ["scenes_png"]
    assert h["total"] == 2


@pytest.mark.asyncio
async def test_export_xlsx_button_writes_plan_rows(api_client, tmp_path) -> None:
    client, project_id, factory = api_client
    r = await client.post(f"/api/db/projects/{project_id}/export-xlsx")
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["frames"] == 2
    assert data["cells"] > 0

    wb = load_workbook(tmp_path / "videos" / "dbv2-api" / "project.xlsx")
    try:
        ws = wb["план"]
        assert ws.cell(row=ROW_IMAGE_PROMPT_V8, column=3).value == "img 1"
        assert ws.cell(row=ROW_VOICEOVER_V8, column=4).value == "реплика 2"
    finally:
        wb.close()


@pytest.mark.asyncio
async def test_connect_edges_each_to_storage(api_client) -> None:
    """connect_edges from=each to_type=storage: создаёт storage и рёбра от всех нод."""
    from app.web.routers.db_browser import (
        _apply_add_node,
        _apply_connect_edges,
        _linear_order,
    )

    client, project_id, factory = api_client
    async with factory() as session:
        p = await session.get(Project, project_id)
        meta = dict(p.meta or {})
        meta["canvas_graph"] = {
            "nodes": [
                {
                    "id": "n_plan",
                    "type": "plan",
                    "position": {"x": 0.0, "y": 0.0},
                    "data": {"label": "Сценарий"},
                },
                {
                    "id": "n_script",
                    "type": "script",
                    "position": {"x": 290.0, "y": 0.0},
                    "data": {"label": "Закадровый"},
                },
                {
                    "id": "n_img",
                    "type": "images",
                    "position": {"x": 580.0, "y": 0.0},
                    "data": {"label": "Картинки"},
                },
            ],
            "edges": [
                {"id": "e0", "source": "n_plan", "target": "n_script"},
                {"id": "e1", "source": "n_script", "target": "n_img"},
            ],
        }
        p.meta = meta
        await session.commit()

        res = await _apply_connect_edges(
            session, p, {"from": "each", "to_type": "storage"}
        )
        assert "+3" in res["connect_edges"]
        g = (p.meta or {})["canvas_graph"]
        storage = [n for n in g["nodes"] if n["type"] == "storage"]
        assert len(storage) == 1
        sid = storage[0]["id"]
        to_storage = {
            e["source"] for e in g["edges"] if e["target"] == sid
        }
        assert to_storage == {"n_plan", "n_script", "n_img"}
        # Цепочка пайплайна всё ещё линейна (storage сбоку).
        assert _linear_order(g["nodes"], g["edges"]) == [
            "n_plan",
            "n_script",
            "n_img",
        ]

        # add_node each: старые рёбра к storage живы + новые ноды сами дотянуты.
        res_add = await _apply_add_node(
            session, p, {"node_type": "excel_gpt", "after": "each"}
        )
        assert "storage" in res_add["add_node"]
        g2 = (p.meta or {})["canvas_graph"]
        sid2 = next(n["id"] for n in g2["nodes"] if n["type"] == "storage")
        all_ids = {n["id"] for n in g2["nodes"] if n["id"] != sid2}
        wired = {e["source"] for e in g2["edges"] if e["target"] == sid2}
        assert wired == all_ids


def test_diagnose_only_question_blocks_code_dump() -> None:
    from app.web.routers.db_browser import (
        _human_reply_from_diagnostics,
        _is_diagnose_only_question,
    )

    assert _is_diagnose_only_question(
        "У МЕНЯ ОПЯТЬ ОШИБКА В РАБОТЕ НОДЫ РАБОТА ГПТ, НАЙДИ ПРИЧИНУ"
    )
    assert _is_diagnose_only_question("что не так с нодой?")
    assert not _is_diagnose_only_question("почини код в enrich_xlsx.py")
    reply = _human_reply_from_diagnostics(
        [
            "ДИАГНОСТИКА:",
            "СБОЙ ШАГА (step_failure):",
            "- шаг: enriching_5",
            "- ошибка: модель не вернула apply-ops JSON",
            "- active excel_gpt: n_excel_gpt_1",
        ]
    )
    assert "apply-ops" in reply
    assert "n_excel_gpt_1" in reply
    assert "read_file" not in reply.lower()


@pytest.mark.asyncio
async def test_canvas_nodes_context_shows_storage_side_edges(api_client) -> None:
    """Контекст оркестратора видит рёбра к storage, не только линейную цепочку."""
    from app.web.routers.db_browser import _canvas_nodes_context

    client, project_id, factory = api_client
    async with factory() as session:
        p = await session.get(Project, project_id)
        meta = dict(p.meta or {})
        meta["canvas_graph"] = {
            "nodes": [
                {
                    "id": "n_plan",
                    "type": "plan",
                    "position": {"x": 0.0, "y": 0.0},
                    "data": {"label": "Сценарий"},
                },
                {
                    "id": "n_excel_gpt_1",
                    "type": "excel_gpt",
                    "position": {"x": 200.0, "y": 0.0},
                    "data": {"label": "Проверка", "slotIndex": 1},
                },
                {
                    "id": "n_storage_1",
                    "type": "storage",
                    "position": {"x": 0.0, "y": 200.0},
                    "data": {"label": "Хранилище"},
                },
            ],
            "edges": [
                {"id": "e0", "source": "n_plan", "target": "n_excel_gpt_1"},
                {"id": "e1", "source": "n_plan", "target": "n_storage_1"},
                {"id": "e2", "source": "n_excel_gpt_1", "target": "n_storage_1"},
            ],
        }
        p.meta = meta
        await session.commit()

        lines, _km = await _canvas_nodes_context(session, p)
        blob = "\n".join(lines)
        assert "СВЯЗИ (цепочка):" in blob
        assert "n_plan → n_excel_gpt_1" in blob or "n_plan→n_excel_gpt_1" in blob
        assert "СВЯЗИ К ХРАНИЛИЩУ" in blob
        assert "n_plan→n_storage_1" in blob
        assert "n_excel_gpt_1→n_storage_1" in blob
        # Цепочка не должна притворяться что storage — часть пайплайна.
        chain_line = next(L for L in lines if L.startswith("СВЯЗИ (цепочка):"))
        assert "n_storage_1" not in chain_line


@pytest.mark.asyncio
async def test_orchestrator_chat_connect_edges_action(api_client, monkeypatch) -> None:
    """Оркестратор умеет connect_edges — не отмазывается «нет действия»."""
    from app.models import Workflow

    client, project_id, factory = api_client
    async with factory() as session:
        session.add(
            Workflow(
                name="default",
                is_default=True,
                nodes=[
                    {
                        "id": "n_plan",
                        "type": "plan",
                        "position": {"x": 0.0, "y": 0.0},
                        "data": {"label": "Сценарий"},
                    },
                    {
                        "id": "n_script",
                        "type": "script",
                        "position": {"x": 290.0, "y": 0.0},
                        "data": {"label": "Закадровый"},
                    },
                ],
                edges=[{"id": "e0", "source": "n_plan", "target": "n_script"}],
            )
        )
        await session.commit()

    monkeypatch.setattr(
        "app.services.gpt_client.get_gpt_client",
        lambda: _FakeGpt(
            '{"actions":[{"connect_edges":{"from":"each","to_type":"storage"}}]}'
        ),
    )
    r = await client.post(
        f"/api/db/projects/{project_id}/orchestrator/chat",
        json={"message": "подведи все ноды к хранилищу", "history": []},
    )
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["error"] is None
    assert data["actions_run"]
    assert "connect_edges" in data["actions_run"][0]
    async with factory() as session:
        p = await session.get(Project, project_id)
        g = (p.meta or {}).get("canvas_graph") or {}
        assert any(n.get("type") == "storage" for n in g.get("nodes") or [])
