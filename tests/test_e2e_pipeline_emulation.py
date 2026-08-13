"""T20: эмуляция полного прогона пайплайна на фейках (без GPT/Outsee/Chrome).

Гоняет НАСТОЯЩИЕ step run() по цепочке plan → script → split → img_pr →
img → anim_pr → video, после каждого шага — harness verify + сборщик
ошибок (step_failure meta, NodeRun failed). Хаос-кейсы — отдельными
тестами: partial ops, пустой ответ, wipe-без-воскрешения.
"""

from __future__ import annotations

import json
from pathlib import Path

import pytest
import pytest_asyncio
from sqlalchemy import select
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from app.models import Base, Frame, Project, ProjectStatus
from tests.fakes import FakeChaos, FakeScenario
from tests.fakes.install import install_fakes_monkeypatch


@pytest_asyncio.fixture
async def env(tmp_path: Path, monkeypatch):
    """Изолированная БД + data_dir + фейки GPT/медиа/оператор."""
    db_url = f"sqlite+aiosqlite:///{tmp_path / 'state.db'}"
    engine = create_async_engine(db_url, echo=False)
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
        from app.services import db_v2

        await db_v2.migrate_db_v2_schema(conn)
    factory = async_sessionmaker(engine, expire_on_commit=False)

    from app.settings import settings

    monkeypatch.setattr(settings, "data_dir", tmp_path)
    monkeypatch.setattr(settings, "sqlite_path", tmp_path / "state.db")
    # Медиа-шаги без браузера: HTTP-путь «включён», браузер не нужен.
    import app.orchestrator.steps.generate_images as gi
    import app.orchestrator.steps.generate_videos as gv

    monkeypatch.setattr(gi, "_img_http_primary", lambda: True)
    monkeypatch.setattr(gv, "_video_http_primary", lambda: True)

    yield {"factory": factory, "tmp": tmp_path, "monkeypatch": monkeypatch}
    await engine.dispose()


async def _new_project(factory, tmp_path: Path) -> int:
    async with factory() as s:
        p = Project(
            slug="e2e-emu",
            title="E2E эмуляция",
            topic="Рачок в неоне",
            status=ProjectStatus.new,
            meta={},
        )
        s.add(p)
        await s.flush()
        p.data_dir.mkdir(parents=True, exist_ok=True)
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "план"
        wb.create_sheet("Общий план")
        wb.save(p.data_dir / "project.xlsx")
        wb.close()
        await s.commit()
        return p.id


async def _get(factory, pid: int) -> Project:
    async with factory() as s:
        return await s.get(Project, pid)


async def _frames(factory, pid: int) -> list[Frame]:
    async with factory() as s:
        return list(
            (
                await s.execute(
                    select(Frame).where(Frame.project_id == pid).order_by(Frame.number)
                )
            ).scalars()
        )


def _collect_failures(p: Project) -> dict:
    """Сборщик ошибок: step_failure meta (NodeRun/логи — в отчёте e2e)."""
    meta = p.meta if isinstance(p.meta, dict) else {}
    return meta.get("step_failure") or {}


@pytest.mark.asyncio
async def test_full_pipeline_emulation_happy_path(env) -> None:
    factory = env["factory"]
    monkeypatch = env["monkeypatch"]
    scenario = FakeScenario()
    install_fakes_monkeypatch(
        monkeypatch, scenario=scenario, session_factory=factory
    )

    pid = await _new_project(factory, env["tmp"])
    report: list[dict] = []

    async def run_step(step_name: str, status: ProjectStatus, runner) -> None:
        async with factory() as s:
            p = await s.get(Project, pid)
            p.status = status
            await s.commit()
            await runner(s, p)
            await s.commit()
            await s.refresh(p)
            # harness verify после каждого шага (данные, без HTTP)
            from app.services.agent_harness import run_harness_verify

            rep = await run_harness_verify(s, p, allow_repair=False, include_http=False)
            await s.commit()
            fails = _collect_failures(p)
            report.append(
                {
                    "step": step_name,
                    "status": p.status.value,
                    "harness_ok": rep.ok,
                    "harness_bad": [c.name for c in rep.checks if not c.ok],
                    "step_failure": fails,
                }
            )

    from app.orchestrator.steps import (
        generate_image_prompts,
        generate_images,
        generate_videos,
        make_animation_prompts,
        make_plan,
        make_script,
        split_frames,
    )

    await run_step("plan", ProjectStatus.planning, lambda s, p: make_plan.run(s, p, None))
    await run_step("script", ProjectStatus.scripting, lambda s, p: make_script.run(s, p, None))
    await run_step("split", ProjectStatus.splitting, lambda s, p: split_frames.run(s, p, None))
    await run_step(
        "img_pr",
        ProjectStatus.generating_image_prompts,
        lambda s, p: generate_image_prompts.run(s, p, None),
    )
    await run_step(
        "img", ProjectStatus.generating_images, lambda s, p: generate_images.run(s, p, None)
    )
    await run_step(
        "anim_pr",
        ProjectStatus.generating_animation_prompts,
        lambda s, p: make_animation_prompts.run(s, p, None),
    )
    await run_step(
        "video", ProjectStatus.generating_videos, lambda s, p: generate_videos.run(s, p, None)
    )

    # ── итоговые инварианты ──
    p = await _get(factory, pid)
    assert p.status is ProjectStatus.videos_ready, json.dumps(report, ensure_ascii=False, indent=2)

    frames = await _frames(factory, pid)
    assert len(frames) == len(scenario.vo_blocks)
    for fr in frames:
        assert (fr.voiceover_text or "").strip(), f"frame {fr.number}: нет VO"
        assert (fr.image_prompt or "").strip(), f"frame {fr.number}: нет img prompt"
        assert (fr.animation_prompt or "").strip(), f"frame {fr.number}: нет anim prompt"

    # файлы на диске
    scenes = list((p.data_dir / "scenes").glob("*.png"))
    vids = list((p.data_dir / "videos").glob("*.mp4"))
    assert len(scenes) >= len(frames), f"scenes={len(scenes)}"
    assert len(vids) == len(frames), f"videos={len(vids)} — дубли/пропуски"

    # ни один шаг не оставил step_failure
    for entry in report:
        assert not entry["step_failure"], f"{entry['step']}: {entry['step_failure']}"
        assert entry["harness_ok"], f"{entry['step']}: harness {entry['harness_bad']}"

    # машиночитаемый отчёт (сборщик ошибок)
    report_path = p.data_dir / "e2e_report.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")


@pytest.mark.asyncio
async def test_img_pr_partial_ops_converge_via_checkpoint(env) -> None:
    """Хаос: GPT отдаёт 1/N ops за ответ → continue-раунды + checkpoint +
    внешний retry сходятся без часовых onesie-зависаний (13.08, #14)."""
    factory = env["factory"]
    monkeypatch = env["monkeypatch"]
    scenario = FakeScenario(
        vo_blocks=[f"Закадр кадра {i}: рачок и неоновый город." for i in range(1, 11)]
    )
    install_fakes_monkeypatch(
        monkeypatch,
        scenario=scenario,
        chaos=FakeChaos(img_pr_partial=True),
        session_factory=factory,
    )
    pid = await _new_project(factory, env["tmp"])

    async with factory() as s:
        p = await s.get(Project, pid)
        from app.services import db_apply

        await db_apply.apply_ops(
            s,
            p,
            [
                {
                    "target": "replace_frames",
                    "frames": [
                        {"закадр": b, "длительность": 3.0}
                        for b in scenario.vo_blocks
                    ],
                }
            ],
            export_xlsx=False,
        )
        p.status = ProjectStatus.generating_image_prompts
        await s.commit()

        from app.orchestrator.steps import generate_image_prompts

        await generate_image_prompts.run(s, p, None)
        await s.commit()
        await s.refresh(p)
        assert p.status is ProjectStatus.image_prompts_ready
        frames = await _frames(factory, pid)
        assert all((f.image_prompt or "").strip() for f in frames)


@pytest.mark.asyncio
async def test_img_pr_empty_replies_fail_loud_no_ready(env) -> None:
    """Хаос: GPT никогда не отдаёт ops → шаг падает явно, статус НЕ ready."""
    factory = env["factory"]
    monkeypatch = env["monkeypatch"]
    scenario = FakeScenario()
    install_fakes_monkeypatch(
        monkeypatch,
        scenario=scenario,
        chaos=FakeChaos(img_pr_empty_batches=99),
        session_factory=factory,
    )
    pid = await _new_project(factory, env["tmp"])

    async with factory() as s:
        p = await s.get(Project, pid)
        # мини-бутстрап до img_pr
        from app.services import db_apply

        await db_apply.apply_ops(
            s,
            p,
            [
                {
                    "target": "replace_frames",
                    "frames": [
                        {"закадр": b, "длительность": 3.0}
                        for b in scenario.vo_blocks
                    ],
                }
            ],
            export_xlsx=False,
        )
        p.status = ProjectStatus.generating_image_prompts
        await s.commit()

        from app.orchestrator.steps import generate_image_prompts

        with pytest.raises(RuntimeError):
            await generate_image_prompts.run(s, p, None)
        await s.rollback()
        p = await s.get(Project, pid)
        assert p.status is ProjectStatus.generating_image_prompts


@pytest.mark.asyncio
async def test_wipe_after_img_pr_leaves_no_resurrection(env) -> None:
    """Хаос→восстановление: wipe img_pr — ничего не воскресает ниоткуда."""
    factory = env["factory"]
    monkeypatch = env["monkeypatch"]
    scenario = FakeScenario()
    install_fakes_monkeypatch(
        monkeypatch, scenario=scenario, session_factory=factory
    )
    pid = await _new_project(factory, env["tmp"])

    async with factory() as s:
        p = await s.get(Project, pid)
        from app.services import db_apply

        await db_apply.apply_ops(
            s,
            p,
            [
                {
                    "target": "replace_frames",
                    "frames": [
                        {"закадр": b, "длительность": 3.0}
                        for b in scenario.vo_blocks
                    ],
                }
            ],
            export_xlsx=True,
        )
        await s.commit()
        frames = await _frames(factory, pid)
        ops = [
            {
                "frame_uuid": f.uuid,
                "fields": {"промт_картинки": f"prompt {f.number}"},
            }
            for f in frames
        ]
        await db_apply.apply_ops(s, p, ops, node_kind="img_pr")
        await s.commit()

        from app.services.reset_step import reset_step

        await reset_step(s, p, "img_pr")
        await s.commit()

        from app.models import PromptVersion

        pvs = list(
            (
                await s.execute(
                    select(PromptVersion).where(PromptVersion.project_id == pid)
                )
            ).scalars()
        )
        active = [pv for pv in pvs if pv.is_active]
        frames = await _frames(factory, pid)
        assert all(not (f.image_prompt or "").strip() for f in frames)
        assert not active, f"active PV после wipe: {len(active)}"

        # зеркало R45 пустое
        from openpyxl import load_workbook

        wb = load_workbook(p.data_dir / "project.xlsx")
        try:
            ws = wb["план"]
            r45 = [ws.cell(row=45, column=3 + i).value for i in range(len(frames))]
        finally:
            wb.close()
        assert all(v in (None, "") for v in r45), f"R45 не очищен: {r45}"

        # montage overlay не подсовывает stale
        from app.services.montage_board import _prompts_from_frame_db

        snapshot_prompts = _prompts_from_frame_db(frames)
        assert all(
            not v["image_prompt_shot1"] for v in snapshot_prompts.values()
        )
