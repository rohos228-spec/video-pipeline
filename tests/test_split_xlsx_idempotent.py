"""split_xlsx: не дергаем GPT если xlsx уже разбит или tmp download есть."""

from __future__ import annotations

from pathlib import Path

import pytest

from app.services import xlsx_step_runners as xsr
from app.services.xlsx_versioning import validate_xlsx


def test_count_v8_voiceover_blocks_empty(tmp_path: Path) -> None:
    from openpyxl import Workbook

    from app.services.xlsx_v8_import import SHEET_PLAN_V8

    p = tmp_path / "empty.xlsx"
    wb = Workbook()
    wb.active.title = SHEET_PLAN_V8
    wb.save(p)
    assert xsr._count_v8_voiceover_blocks(p) == 0


def test_try_reuse_split_download(tmp_path: Path) -> None:
    from openpyxl import Workbook

    from app.services.xlsx_v8_import import (
        ROW_VOICEOVER_V8,
        SHEET_PLAN_V8,
    )

    tmp_dir = tmp_path / "tmp_gpt"
    tmp_dir.mkdir()
    proj_xlsx = tmp_path / "project.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_PLAN_V8
    ws.cell(row=ROW_VOICEOVER_V8, column=3, value="block one")
    ws.cell(row=ROW_VOICEOVER_V8, column=4, value="block two")
    split_file = tmp_dir / "split_20260101_120000.xlsx"
    wb.save(split_file)
    wb.save(proj_xlsx)

    assert validate_xlsx(split_file) is None
    reused = xsr._try_reuse_split_download(tmp_dir, proj_xlsx)
    assert reused is not None
    assert reused.project_xlsx == proj_xlsx
    assert xsr._count_v8_voiceover_blocks(proj_xlsx) == 2
