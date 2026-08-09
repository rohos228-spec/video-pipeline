"""export_project_xlsx must wipe plan columns beyond the current frame count."""

from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook

from app.services.db_apply import export_project_xlsx
from app.services.xlsx_v8_import import ROW_DURATION_V8, ROW_VOICEOVER_V8


def test_export_clears_stale_voiceover_columns(tmp_path: Path) -> None:
    path = tmp_path / "project.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "план"
    for col in range(3, 12):
        ws.cell(row=ROW_VOICEOVER_V8, column=col, value=f"vo-{col}")
        ws.cell(row=ROW_DURATION_V8, column=col, value=3.0)
    wb.save(path)
    wb.close()

    project = SimpleNamespace(
        data_dir=tmp_path,
        meta={},
    )
    frames = [
        SimpleNamespace(
            number=i,
            image_prompt="",
            animation_prompt="",
            voiceover_text=f"keep-{i}",
            duration_seconds=2.0,
            start_ts=None,
            end_ts=None,
            attrs={},
        )
        for i in range(1, 4)
    ]
    out = export_project_xlsx(project, frames)  # type: ignore[arg-type]
    assert out["frames"] == 3

    wb2 = load_workbook(path)
    ws2 = wb2["план"]
    assert ws2.cell(ROW_VOICEOVER_V8, 3).value == "keep-1"
    assert ws2.cell(ROW_VOICEOVER_V8, 5).value == "keep-3"
    assert ws2.cell(ROW_VOICEOVER_V8, 6).value is None
    assert ws2.cell(ROW_VOICEOVER_V8, 11).value is None
    assert ws2.cell(ROW_DURATION_V8, 11).value is None
    wb2.close()
