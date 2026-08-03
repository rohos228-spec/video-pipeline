"""Точечная правка ячеек project.xlsx из «Базы»."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from app.models import Base, Entity, Frame, Project, ProjectStatus
from app.services import db_v2
from app.services.excel_characters import (
    ROW_ID,
    ROW_NAME,
    ROW_RULES,
    SHEET_PERSONS,
    parse_persons_sheet,
)
from app.services.sheet_cell_edit import write_sheet_cell
from app.services.xlsx_v8_import import ROW_VOICEOVER_V8, SHEET_PLAN_V8


@pytest.fixture
async def session() -> AsyncSession:
    engine = create_async_engine("sqlite+aiosqlite:///:memory:")
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
        await db_v2.migrate_db_v2_schema(conn)
    factory = async_sessionmaker(engine, expire_on_commit=False)
    async with factory() as s:
        yield s
    await engine.dispose()


def _mk_xlsx(data_dir: Path) -> Path:
    data_dir.mkdir(parents=True, exist_ok=True)
    path = data_dir / "project.xlsx"
    wb = Workbook()
    # plan
    ws = wb.active
    ws.title = SHEET_PLAN_V8
    ws.cell(row=1, column=1, value="метка")
    ws.cell(row=ROW_VOICEOVER_V8, column=3, value="старый закадр")
    # persons
    wp = wb.create_sheet(SHEET_PERSONS)
    wp["A1"] = "ID персонажа"
    wp["A3"] = "имя"
    wp["A4"] = "внешность"
    wp["A5"] = "одежда"
    wp["A6"] = "характер"
    wp["A7"] = "правила"
    wp.cell(row=ROW_ID, column=2, value="c01")
    wp.cell(row=ROW_NAME, column=2, value="Алиса")
    wp.cell(row=ROW_ID, column=3, value="c02")
    wp.cell(row=ROW_NAME, column=3, value="Боб")
    wp.cell(row=ROW_RULES, column=3, value="")
    # misc sheet
    wm = wb.create_sheet("Заметки")
    wm["B2"] = "hello"
    wb.save(path)
    wb.close()
    return path


async def _mk_project(session: AsyncSession, tmp_path: Path, monkeypatch) -> Project:
    import app.settings as app_settings

    monkeypatch.setattr(app_settings.settings, "data_dir", tmp_path / "data")
    p = Project(
        slug="cell-edit",
        topic="тест",
        status=ProjectStatus.frames_ready,
    )
    session.add(p)
    await session.flush()
    p.data_dir.mkdir(parents=True, exist_ok=True)
    _mk_xlsx(p.data_dir)
    session.add(
        Frame(
            project_id=p.id,
            number=1,
            voiceover_text="старый закадр",
            image_prompt="",
            animation_prompt="",
            attrs={},
        )
    )
    session.add(
        Entity(
            project_id=p.id,
            type="character",
            code="c01",
            name="Алиса",
            attrs={},
            sort_key=10.0,
        )
    )
    session.add(
        Entity(
            project_id=p.id,
            type="character",
            code="c02",
            name="Боб",
            attrs={},
            sort_key=20.0,
        )
    )
    await session.commit()
    await session.refresh(p)
    return p


@pytest.mark.asyncio
async def test_persons_rules_cell_syncs_entity(
    session: AsyncSession, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
):
    p = await _mk_project(session, tmp_path, monkeypatch)
    result = await write_sheet_cell(
        session,
        p,
        sheet=SHEET_PERSONS,
        row=ROW_RULES,
        col=3,  # c02
        value="c01",
    )
    await session.commit()
    assert result["ok"] is True
    assert result["synced"] == "entity:c02"

    en = (
        await session.execute(
            select(Entity).where(
                Entity.project_id == p.id,
                Entity.code == "c02",
            )
        )
    ).scalar_one()
    assert (en.attrs or {}).get("rules") == "c01"

    chars = parse_persons_sheet(p.data_dir / "project.xlsx")
    c02 = next(c for c in chars if c.id == "c02")
    assert c02.rules == "c01"
    assert c02.ref_ids == ["c01"]
    assert (p.meta or {}).get("excel_hero", {}).get("characters")


@pytest.mark.asyncio
async def test_plan_voiceover_cell_syncs_frame(
    session: AsyncSession, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
):
    p = await _mk_project(session, tmp_path, monkeypatch)
    result = await write_sheet_cell(
        session,
        p,
        sheet=SHEET_PLAN_V8,
        row=ROW_VOICEOVER_V8,
        col=3,  # frame 1
        value="новый закадр",
    )
    await session.commit()
    assert result["synced"] == "voiceover_text"
    fr = (
        await session.execute(
            select(Frame).where(Frame.project_id == p.id, Frame.number == 1)
        )
    ).scalar_one()
    assert fr.voiceover_text == "новый закадр"


@pytest.mark.asyncio
async def test_arbitrary_sheet_cell_xlsx_only(
    session: AsyncSession, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
):
    from openpyxl import load_workbook

    p = await _mk_project(session, tmp_path, monkeypatch)
    result = await write_sheet_cell(
        session,
        p,
        sheet="Заметки",
        row=2,
        col=2,
        value="updated",
    )
    await session.commit()
    assert result["ok"] is True
    assert result["synced"] is None
    wb = load_workbook(p.data_dir / "project.xlsx")
    assert wb["Заметки"]["B2"].value == "updated"
    wb.close()
