from contextlib import asynccontextmanager

import pytest
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from app.models import Base
from app.services.shot_menu import build_shot_menu, group_vo_cells


@pytest.fixture
async def mem_db():
    engine = create_async_engine("sqlite+aiosqlite:///:memory:")
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    factory = async_sessionmaker(engine, expire_on_commit=False, class_=AsyncSession)

    @asynccontextmanager
    async def _scope():
        async with factory() as session:
            yield session

    yield _scope
    await engine.dispose()


def test_adjacent_vo_cells_are_not_glued() -> None:
    frames = [
        {"id": 1, "number": 1, "sort_key": 10, "voiceover_text": "Первая ячейка", "duration_seconds": 2},
        {"id": 2, "number": 2, "sort_key": 20, "voiceover_text": "Вторая ячейка", "duration_seconds": 3},
    ]
    cells = group_vo_cells(frames)
    assert len(cells) == 2
    assert cells[0]["voiceover"] == "Первая ячейка"
    assert cells[1]["voiceover"] == "Вторая ячейка"
    assert [s["label"] for s in cells[0]["shots"]] == ["1.1"]
    assert [s["label"] for s in cells[1]["shots"]] == ["2.1"]


def test_duration_from_vo_when_seconds_missing() -> None:
    frames = [
        {
            "id": 1,
            "number": 1,
            "voiceover_text": "История Дарьи Салтыковой началась не с процесса",
            "duration_seconds": 0,
        }
    ]
    cells = group_vo_cells(frames)
    assert cells[0]["duration_sec"] > 0
    assert cells[0]["shots"][0]["duration_sec"] > 0


def test_camera_subdivide_fills_size_move_set() -> None:
    frames = [
        {
            "id": 1,
            "number": 1,
            "uuid": "a" * 24,
            "voiceover_text": "фрагмент",
            "duration_seconds": 3,
            "attrs": {
                "shot01_action": "кладёт жалобу",
                "camera_subdivide": {
                    "role": "vo_parent",
                    "parent_uuid": "a" * 24,
                    "крупность": "Средний план",
                    "движение": "статика",
                    "набор": "SET_08",
                },
            },
        }
    ]
    cells = group_vo_cells(frames)
    fields = cells[0]["shots"][0]["fields"]
    assert fields["action"] == "кладёт жалобу"
    assert fields["size"] == "Средний план"
    assert fields["move"] == "статика"
    assert "SET_08" in fields["set"]


def test_empty_vo_shots_attach_to_parent_cell() -> None:
    frames = [
        {
            "id": 1,
            "number": 1,
            "sort_key": 10,
            "voiceover_text": "Что, если самый страшный адрес",
            "duration_seconds": 2.4,
            "attrs": {"shot01_action": "Людмила заслоняет", "place": "SET_32"},
        },
        {
            "id": 2,
            "number": 2,
            "sort_key": 20,
            "voiceover_text": "",
            "duration_seconds": 2.1,
            "attrs": {"shot01_action": "Соседка хватает рукав", "place": "SET_37"},
        },
        {
            "id": 3,
            "number": 3,
            "sort_key": 30,
            "voiceover_text": "Новокузнецк, дом № 53",
            "duration_seconds": 3.0,
            "place": "SET_22",
        },
    ]
    cells = group_vo_cells(frames)
    assert len(cells) == 2
    assert len(cells[0]["shots"]) == 2
    assert cells[0]["shots"][0]["voiceover_in_shot"] != "—"
    # После перегенерации закадр часто только на родителе — лента режет по шотам.
    child_vo = cells[0]["shots"][1]["voiceover_in_shot"]
    assert child_vo != "—"
    joined = " ".join(
        s["voiceover_in_shot"]
        for s in cells[0]["shots"]
        if s["voiceover_in_shot"] != "—"
    )
    assert " ".join("Что, если самый страшный адрес".split()) == " ".join(joined.split())
    assert cells[0]["shots"][1]["fields"]["action"] == "Соседка хватает рукав"
    # SET показываем с человеческим именем из camera_sets.md
    assert cells[0]["loc"] == "SET_32 · Два человека"
    assert len(cells[1]["shots"]) == 1
    assert cells[1]["voiceover"].startswith("Новокузнецк")


def test_leading_empty_shots_do_not_eat_next_vo() -> None:
    frames = [
        {"id": 1, "number": 1, "voiceover_text": "", "duration_seconds": 1},
        {"id": 2, "number": 2, "voiceover_text": "Настоящий закадр", "duration_seconds": 2},
    ]
    cells = group_vo_cells(frames)
    assert len(cells) == 2
    assert cells[0]["voiceover"] == ""
    assert cells[1]["voiceover"] == "Настоящий закадр"


def test_shot_menu_is_menu_not_work_step() -> None:
    from app.orchestrator.graph.planner import is_side_sink_node_type
    from app.orchestrator.node_registry import (
        is_side_node_type,
        is_ui_menu_node_type,
        is_work_node_type,
    )

    assert is_ui_menu_node_type("shot_menu")
    assert is_side_node_type("shot_menu")
    assert is_side_sink_node_type("shot_menu")
    assert not is_work_node_type("shot_menu")


def test_subdivide_group_is_one_cell_and_vo_sum_equals_cell() -> None:
    """Ячейка = сцена: кадры с одним parent_uuid — одна ячейка; текст ячейки =
    сумма фрагментов кадров (после раздачи закадра по шотам)."""
    frames = [
        {
            "id": 1,
            "number": 1,
            "sort_key": 10,
            "voiceover_text": "Что, если самый",
            "duration_seconds": 2.0,
            "attrs": {"camera_subdivide": {"parent_uuid": "p1", "shot_index": 1}},
        },
        {
            "id": 2,
            "number": 2,
            "sort_key": 20,
            "voiceover_text": "страшный адрес —",
            "duration_seconds": 2.0,
            "attrs": {"camera_subdivide": {"parent_uuid": "p1", "shot_index": 2}},
        },
        {
            "id": 3,
            "number": 3,
            "sort_key": 30,
            "voiceover_text": "самый обычный?",
            "duration_seconds": 2.0,
            "attrs": {"camera_subdivide": {"parent_uuid": "p1", "shot_index": 3}},
        },
        {
            "id": 4,
            "number": 4,
            "sort_key": 40,
            "voiceover_text": "Следующая ячейка",
            "duration_seconds": 3.0,
        },
    ]
    cells = group_vo_cells(frames)
    assert len(cells) == 2
    assert [s["label"] for s in cells[0]["shots"]] == ["1.1", "1.2", "1.3"]
    assert cells[0]["voiceover"] == "Что, если самый страшный адрес — самый обычный?"
    assert cells[0]["voiceover"].split() == [
        "Что,", "если", "самый", "страшный", "адрес", "—", "самый", "обычный?",
    ]
    assert cells[0]["shots"][1]["voiceover_in_shot"] == "страшный адрес —"
    assert cells[1]["voiceover"] == "Следующая ячейка"


def test_build_shot_menu_summary() -> None:
    frames = [
        {"id": 1, "number": 1, "voiceover_text": "aaa", "duration_seconds": 1.5},
        {"id": 2, "number": 2, "voiceover_text": "", "duration_seconds": 2.5},
        {"id": 3, "number": 3, "voiceover_text": "bb", "duration_seconds": 4},
    ]
    menu = build_shot_menu(frames)
    assert menu["summary"]["vo_cells"] == 2
    assert menu["summary"]["shots"] == 3
    assert menu["summary"]["duration_sec"] == 8.0
    assert menu["summary"]["vo_chars"] == 5
    assert any(t["key"] == "vo" and t["pinned"] for t in menu["tracks"])
    assert "vo" in menu["default_tracks"]


def test_set_human_name() -> None:
    from app.services.shot_menu import set_human_name

    assert set_human_name("SET_08") == "SET_08 · Документ/свидетельство"
    assert set_human_name("set_8") == "SET_08 · Документ/свидетельство"
    assert set_human_name("SET_99") == "SET_99"  # нет в каталоге — как есть
    assert set_human_name("подъезд") == "подъезд"


def test_shot_dto_has_all_fields_for_custom_rows() -> None:
    frames = [
        {
            "id": 1,
            "number": 1,
            "voiceover_text": "текст",
            "duration_seconds": 2,
            "attrs": {"shot01_action": "идёт", "custom_field": "моё значение"},
        },
    ]
    menu = build_shot_menu(frames)
    shot = menu["cells"][0]["shots"][0]
    assert shot["all"]["custom_field"] == "моё значение"
    assert shot["all"]["voiceover_text"] == "текст"
    assert shot["all"]["number"] == "1"


@pytest.mark.asyncio
async def test_edit_cell_and_field_and_add_cell(mem_db) -> None:
    import uuid as _uuid

    from app.models import Frame, Project, ProjectStatus
    from app.services.shot_menu import add_cell, edit_cell_voiceover, edit_shot_field

    async with mem_db() as session:
        p = Project(
            slug=f"sm-{_uuid.uuid4().hex[:6]}",
            topic="t",
            status=ProjectStatus.new,
            meta={},
        )
        session.add(p)
        await session.flush()
        for i, txt in enumerate(("первая", "вторая"), start=1):
            session.add(
                Frame(
                    project_id=p.id,
                    number=i,
                    voiceover_text=txt,
                    status="planned",
                    uuid=f"u{i:03d}",
                    sort_key=float(i) * 10.0,
                    attrs={},
                )
            )
        await session.commit()

        # правка закадра ячейки
        out = await edit_cell_voiceover(session, p.id, "u001", "первая исправленная")
        assert out["ok"] and out["chars"] == len("первая исправленная")

        # правка поля шота (attrs) и колонки
        await edit_shot_field(session, p.id, "u001", "action", "смотрит вверх")
        await edit_shot_field(session, p.id, "u001", "img_prompt", "промт картинки")
        from sqlalchemy import select

        fr = (
            await session.execute(select(Frame).where(Frame.uuid == "u001"))
        ).scalars().first()
        assert fr.attrs["shot01_action"] == "смотрит вверх"
        assert fr.image_prompt == "промт картинки"

        # добавление ячейки между первой и второй
        add = await add_cell(session, p, before_index=2, voiceover="новая между")
        assert add["ok"]
        frames = list(
            (
                await session.execute(
                    select(Frame).where(Frame.project_id == p.id).order_by(
                        Frame.sort_key, Frame.number
                    )
                )
            ).scalars().all()
        )
        texts = [f.voiceover_text for f in frames]
        assert texts == ["первая исправленная", "новая между", "вторая"]
        # number перенумерованы 1..N без дыр
        assert [f.number for f in frames] == [1, 2, 3]

        # добавление в конец
        await add_cell(session, p, before_index=None, voiceover="в конец")
        frames = list(
            (
                await session.execute(
                    select(Frame).where(Frame.project_id == p.id).order_by(
                        Frame.sort_key, Frame.number
                    )
                )
            ).scalars().all()
        )
        assert frames[-1].voiceover_text == "в конец"


@pytest.mark.asyncio
async def test_shot_menu_xlsx_fallback(tmp_path, monkeypatch) -> None:
    """Пустая БД + project.xlsx с закадром → меню подтягивает кадры из xlsx."""
    import uuid as _uuid


    from app.models import Project, ProjectStatus
    from app.settings import settings
    from app.web.routers.db_browser import shot_menu

    monkeypatch.setattr(settings, "data_dir", tmp_path)
    engine = create_async_engine(f"sqlite+aiosqlite:///{tmp_path}/sm.db")
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    factory = async_sessionmaker(engine, expire_on_commit=False)

    async with factory() as session:
        p = Project(
            slug=f"smx-{_uuid.uuid4().hex[:6]}",
            topic="t",
            status=ProjectStatus.new,
            meta={},
        )
        session.add(p)
        await session.flush()
        await session.commit()
        pid = p.id
        slug = p.slug

    # project.xlsx из шаблона v8 + закадр в R49
    from openpyxl import load_workbook

    from app.project_root import find_project_root

    tpl = find_project_root() / "templates" / "project_template_v8.xlsx"
    d = tmp_path / "videos" / slug
    d.mkdir(parents=True)
    wb = load_workbook(tpl)
    ws = wb["план"]
    ws.cell(row=49, column=3, value="Первая ячейка из xlsx")
    ws.cell(row=49, column=4, value="Вторая ячейка из xlsx")
    wb.save(d / "project.xlsx")
    wb.close()

    async with factory() as session:
        menu = await shot_menu(pid, session)
    texts = [c["voiceover"] for c in menu["cells"]]
    assert texts == ["Первая ячейка из xlsx", "Вторая ячейка из xlsx"]

    await engine.dispose()


def test_character_ids_only() -> None:
    from app.services.shot_menu import character_ids

    assert character_ids("c02 Людмила Спесивцева, c01 Александр, соседка") == "c02, c01"
    assert character_ids("c02,c01") == "c02, c01"
    assert character_ids("") == ""


def test_plan_and_move_are_separate_tracks() -> None:
    frames = [
        {
            "id": 1,
            "number": 1,
            "voiceover_text": "фраза",
            "duration_seconds": 2,
            "attrs": {"shot01_notes": "средний план, eye-level, handheld"},
        }
    ]
    menu = build_shot_menu(frames)
    keys = [t["key"] for t in menu["tracks"]]
    assert "size" in keys
    assert "move" in keys
    assert keys.index("size") != keys.index("move")
    shot = menu["cells"][0]["shots"][0]
    assert "план" in shot["fields"]["size"].lower() or shot["fields"]["size"]
    assert "handheld" in shot["fields"]["move"].lower()
    assert shot["fields"]["size"] != shot["fields"]["move"]


def test_shot_plan_overlay_set_and_stitch_differ_inside_scene() -> None:
    """camera.json: разные SET/стык на фазах; attrs после expand — копия первого бита."""
    frames = [
        {
            "id": 1,
            "number": 1,
            "sort_key": 10,
            "voiceover_text": "Что если самый страшный адрес выглядел как ссора",
            "duration_seconds": 2.4,
            "attrs": {
                "shot01_id_scene": "scene_01",
                "place": "SET_32",
                "scene_transition": "eyeline",
                "camera_subdivide": {
                    "role": "vo_parent",
                    "parent_uuid": "u1",
                    "shot_index": 1,
                    "набор": "SET_32",
                    "крупность": "средний план",
                },
            },
            "uuid": "u1",
        },
        {
            "id": 2,
            "number": 2,
            "sort_key": 20,
            "voiceover_text": "",
            "duration_seconds": 2.0,
            "attrs": {
                "shot01_id_scene": "scene_01",
                "place": "SET_32",
                "scene_transition": "eyeline",
                "characters": "c02 Людмила Спесивцева, соседка",
                "camera_subdivide": {
                    "role": "shot",
                    "parent_uuid": "u1",
                    "shot_index": 2,
                    "набор": "SET_32",
                    "крупность": "средний план",
                },
            },
            "uuid": "u2",
        },
        {
            "id": 3,
            "number": 3,
            "sort_key": 30,
            "voiceover_text": "",
            "duration_seconds": 2.5,
            "attrs": {
                "shot01_id_scene": "scene_01",
                "place": "SET_32",
                "scene_transition": "eyeline",
                "camera_subdivide": {
                    "role": "shot",
                    "parent_uuid": "u1",
                    "shot_index": 3,
                    "набор": "SET_32",
                    "крупность": "средний план",
                },
            },
            "uuid": "u3",
        },
    ]
    plan = [
        {"id_scene": "scene_01", "phase_index": 1, "набор": "SET_32", "переход": "eyeline",
         "крупность": "средний план", "движение": "handheld"},
        {"id_scene": "scene_01", "phase_index": 2, "набор": "SET_37", "переход": "cut_on_action",
         "крупность": "средний крупный", "движение": "push-in"},
        {"id_scene": "scene_01", "phase_index": 3, "набор": "SET_29", "переход": "match_on_gesture",
         "крупность": "крупный план", "движение": "push-in"},
    ]
    cells = group_vo_cells(frames, shot_plan=plan)
    assert len(cells) == 1
    sets = [s["fields"]["set"] for s in cells[0]["shots"]]
    stitches = [s["fields"]["stitch"] for s in cells[0]["shots"]]
    assert "SET_32" in sets[0]
    assert "SET_37" in sets[1]
    assert "SET_29" in sets[2]
    assert stitches == ["eyeline", "cut_on_action", "match_on_gesture"]
    assert cells[0]["shots"][1]["fields"]["characters"] == "c02"
    vos = [s["voiceover_in_shot"] for s in cells[0]["shots"]]
    assert all(v != "—" for v in vos)
    assert " ".join("Что если самый страшный адрес выглядел как ссора".split()) == " ".join(
        " ".join(vos).split()
    )


def test_shot_plan_overlay_without_id_scene_uses_cell_index() -> None:
    """После replace_frames нет id_scene — меню берёт фазы из camera.json по индексу ячейки."""
    frames = [
        {
            "id": 1,
            "number": 1,
            "uuid": "u1",
            "voiceover_text": "Первая ячейка целиком для нарезки",
            "duration_seconds": 6.0,
        },
        {
            "id": 2,
            "number": 2,
            "uuid": "u2",
            "voiceover_text": "Вторая ячейка",
            "duration_seconds": 3.0,
        },
    ]
    plan = [
        {
            "id_scene": "scene_01",
            "phase_index": 1,
            "набор": "SET_32",
            "крупность": "средний план",
            "движение": "handheld",
            "переход": "eyeline",
        },
        {
            "id_scene": "scene_01",
            "phase_index": 2,
            "набор": "SET_37",
            "крупность": "крупный план",
            "движение": "push-in",
            "переход": "cut_on_action",
        },
        {
            "id_scene": "scene_02",
            "phase_index": 1,
            "набор": "SET_22",
            "крупность": "средний план",
            "движение": "static",
            "переход": "cut",
        },
    ]
    cells = group_vo_cells(frames, shot_plan=plan)
    assert len(cells) == 2
    assert len(cells[0]["shots"]) == 1
    assert "SET_32" in cells[0]["shots"][0]["fields"]["set"]
    assert cells[0]["shots"][0]["voiceover_in_shot"] != "—"
    assert "Первая ячейка" in cells[0]["voiceover"]
    assert len(cells[1]["shots"]) == 1
    assert "SET_22" in cells[1]["shots"][0]["fields"]["set"]


def test_stray_empty_frame_does_not_join_set_group() -> None:
    frames = [
        {
            "id": 1,
            "number": 1,
            "uuid": "u1",
            "voiceover_text": "ячейка",
            "duration_seconds": 2,
            "attrs": {
                "camera_subdivide": {
                    "role": "vo_parent",
                    "parent_uuid": "u1",
                    "shot_index": 1,
                }
            },
        },
        {
            "id": 2,
            "number": 2,
            "uuid": "u2",
            "voiceover_text": "",
            "duration_seconds": 1,
            "attrs": {},
        },
    ]
    cells = group_vo_cells(frames)
    assert len(cells) == 2
    assert len(cells[0]["shots"]) == 1


def test_display_vo_split_does_not_mutate_frames() -> None:
    """Нарезка VO по шотам — только DTO, исходные кадры не трогаем."""
    frames = [
        {
            "id": 1,
            "number": 1,
            "sort_key": 10,
            "voiceover_text": "один два три четыре пять шесть",
            "duration_seconds": 2,
            "attrs": {
                "camera_subdivide": {
                    "parent_uuid": "p1",
                    "shot_index": 1,
                }
            },
        },
        {
            "id": 2,
            "number": 2,
            "sort_key": 20,
            "voiceover_text": "",
            "duration_seconds": 2,
            "attrs": {
                "camera_subdivide": {
                    "parent_uuid": "p1",
                    "shot_index": 2,
                }
            },
        },
    ]
    before = frames[0]["voiceover_text"]
    cells = group_vo_cells(frames)
    assert frames[0]["voiceover_text"] == before
    assert frames[1]["voiceover_text"] == ""
    assert cells[0]["shots"][1]["voiceover_in_shot"] != "—"
    assert cells[0]["shots"][1]["voiceover_in_shot"] != ""

