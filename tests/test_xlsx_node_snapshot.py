"""Снимки Excel per-node для превью внутри ноды."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from app.models import Project
from app.services.xlsx_node_snapshot import (
    record_consume_for_node_keys,
    record_produce_for_node_keys,
    resolve_display_xlsx,
    save_node_xlsx_snapshot,
    snapshot_dir,
)


@pytest.fixture
def project(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Project:
    data_root = tmp_path / "data"
    data_root.mkdir()
    monkeypatch.setattr("app.settings.settings.data_dir", str(data_root))
    p = Project(id=7, slug="snap-test", topic="t", hero_mode="auto")
    p.data_dir.mkdir(parents=True, exist_ok=True)
    p.meta = {
        "canvas_graph": {
            "nodes": [
                {"id": "n_plan", "type": "plan"},
                {"id": "n_excel_gpt_1", "type": "excel_gpt"},
            ],
            "edges": [],
        }
    }
    return p


def _write_xlsx(path: Path, marker: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "план"
    ws["A1"] = marker
    wb.save(path)


def test_produce_snapshot_display_is_after(project: Project) -> None:
    before = project.data_dir / "before_src.xlsx"
    after = project.data_dir / "project.xlsx"
    _write_xlsx(before, "OLD")
    _write_xlsx(after, "NEW")

    save_node_xlsx_snapshot(
        project,
        "n_excel_gpt_1",
        role="produce",
        before_path=before,
        after_path=after,
        source="project_xlsx",
    )

    path, info = resolve_display_xlsx(project, node_key="n_excel_gpt_1")
    assert path == snapshot_dir(project, "n_excel_gpt_1") / "after.xlsx"
    assert info["role"] == "produce"
    assert info["resolved"] == "snapshot"

    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True)
    assert wb.active["A1"].value == "NEW"
    wb.close()


def test_consume_snapshot_display_is_before(project: Project) -> None:
    used = project.data_dir / "project.xlsx"
    _write_xlsx(used, "USED")
    record_consume_for_node_keys(
        project, ["n_hero"], used_path=used, source="project_xlsx"
    )
    path, info = resolve_display_xlsx(project, node_key="n_hero")
    assert path.name == "before.xlsx"
    assert info["role"] == "consume"


def test_live_fallback_without_snapshot(project: Project) -> None:
    live = project.data_dir / "project.xlsx"
    _write_xlsx(live, "LIVE")
    path, info = resolve_display_xlsx(project, node_key="n_missing")
    assert path == live
    assert info["resolved"] == "live"


def test_record_produce_for_plan_keys(project: Project) -> None:
    before = project.data_dir / "old.xlsx"
    after = project.data_dir / "project.xlsx"
    _write_xlsx(before, "B")
    _write_xlsx(after, "A")
    record_produce_for_node_keys(
        project,
        ["n_plan"],
        before_path=before,
        after_path=after,
    )
    path, _ = resolve_display_xlsx(project, node_key="n_plan")
    assert path.is_file()
    assert "xlsx_snapshots" in str(path)
