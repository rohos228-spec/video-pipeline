"""hero_quality: sheet checks without blocking ref-variations."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from PIL import Image

from app.services.excel_characters import (
    ROW_CLOTHES,
    ROW_ID,
    ROW_LOOK,
    ROW_NAME,
    ROW_RULES,
    SHEET_PERSONS,
)
from app.services.hero_quality import (
    collect_hero_quality_checks,
    png_corners_look_sheet_bg,
)


def _write_persons(xlsx: Path) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = SHEET_PERSONS
    ws.cell(ROW_ID, 2, "c01")
    ws.cell(ROW_NAME, 2, "A")
    ws.cell(ROW_LOOK, 2, "look a")
    ws.cell(ROW_CLOTHES, 2, "coat")
    ws.cell(ROW_ID, 3, "c02")
    ws.cell(ROW_NAME, 3, "B")
    ws.cell(ROW_LOOK, 3, "look b")
    ws.cell(ROW_CLOTHES, 3, "gown")
    ws.cell(ROW_RULES, 3, "c01")
    wb.save(xlsx)
    wb.close()


def test_png_corners_white_vs_scene(tmp_path: Path) -> None:
    sheet = tmp_path / "sheet.png"
    scene = tmp_path / "scene.png"
    Image.new("RGB", (64, 64), (255, 255, 255)).save(sheet)
    Image.new("RGB", (64, 64), (40, 60, 80)).save(scene)
    assert png_corners_look_sheet_bg(sheet) is True
    assert png_corners_look_sheet_bg(scene) is False


def test_ref_empty_prompt_does_not_fail_harness(tmp_path: Path, monkeypatch) -> None:
    """rules=c01 + пустой hero_prompt — норма, не блокер c02."""
    from app import settings as app_settings
    from app.services import hero_quality as hq

    monkeypatch.setattr(app_settings.settings, "sqlite_path", tmp_path / "missing.db")
    monkeypatch.setattr(
        hq,
        "_latest_hero_artifacts",
        lambda _pid: {
            "c02": {"hero_prompt": "", "path": "", "used_refs": ["c01"], "meta": {}},
            "c01": {
                "hero_prompt": "Professional character turnaround sheet ...",
                "path": "",
                "used_refs": None,
                "meta": {},
            },
        },
    )
    _write_persons(tmp_path / "project.xlsx")
    chars = tmp_path / "characters"
    chars.mkdir()
    Image.new("RGB", (128, 128), (255, 255, 255)).save(chars / "c01.png")
    # реф-png может быть любым — quality check рефы не валит
    Image.new("RGB", (128, 128), (30, 40, 50)).save(chars / "c02.png")

    checks = {c.name: c for c in collect_hero_quality_checks(1, tmp_path)}
    assert "hero_sheet_prompt" not in checks
    assert checks["hero_sheet_png"].ok is True
