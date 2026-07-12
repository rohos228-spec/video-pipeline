"""Тесты Scene Board: таймслоты, shot2, missing-флаги."""

from __future__ import annotations

from pathlib import Path

import pytest
import pytest_asyncio
from openpyxl import Workbook
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from app.models import (
    Artifact,
    ArtifactKind,
    Base,
    Frame,
    FrameStatus,
    Project,
    ProjectStatus,
)
from app.services.plan_shot2 import ROW_IMAGE_PROMPT_2_V8, ROW_SHOT2_ID_SHOT_V8
from app.services.scene_board import build_scene_board, validate_regen_draft
from app.services.xlsx_v8_import import ROW_VOICEOVER_V8, SHEET_PLAN_V8


@pytest_asyncio.fixture
async def session(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    db_url = f"sqlite+aiosqlite:///{tmp_path / 't.db'}"
    engine = create_async_engine(db_url, echo=False)
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    SessionLocal = async_sessionmaker(engine, expire_on_commit=False)
    async with SessionLocal() as s:
        yield s
    await engine.dispose()


def _write_plan(xlsx: Path, *, with_shot2: bool = False) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_PLAN_V8
    ws.cell(row=ROW_VOICEOVER_V8, column=3, value="Первая сцена закадровый")
    ws.cell(row=ROW_VOICEOVER_V8, column=4, value="Вторая сцена текст")
    ws.cell(row=8, column=3, value="c01")
    if with_shot2:
        ws.cell(row=ROW_SHOT2_ID_SHOT_V8, column=3, value="shot_02")
        ws.cell(row=ROW_IMAGE_PROMPT_2_V8, column=3, value="Close-up second shot prompt here")
    wb.save(xlsx)


@pytest.mark.asyncio
async def test_scene_board_timeslots_and_missing(tmp_path: Path, session, monkeypatch):
    from app import models as models_mod

    data_root = tmp_path / "data"
    videos = data_root / "videos" / "board1"
    (videos / "scenes").mkdir(parents=True)
    (videos / "videos").mkdir(parents=True)
    (videos / "audio").mkdir(parents=True)
    (videos / "characters").mkdir(parents=True)
    (videos / "music").mkdir(parents=True)

    # monkeypatch Project.data_dir via settings.data_dir
    monkeypatch.setattr("app.settings.settings.data_dir", data_root)

    xlsx = videos / "project.xlsx"
    _write_plan(xlsx, with_shot2=True)

    # character preview
    (videos / "characters" / "c01_ref.png").write_bytes(b"\x89PNG\r\n\x1a\n" + b"x" * 210_000)

    # shot1 image + shot2 image for scene 1
    (videos / "scenes" / "frame_001_aaa11111.png").write_bytes(
        b"\x89PNG\r\n\x1a\n" + b"x" * 210_000
    )
    (videos / "scenes" / "frame_001_s2_bbb22222.png").write_bytes(
        b"\x89PNG\r\n\x1a\n" + b"x" * 210_000
    )
    # video shot1 only
    (videos / "videos" / "clip_001_ccc33333.mp4").write_bytes(b"fake")
    # audio for frame 1
    (videos / "audio" / "frame_001.mp3").write_bytes(b"mp3")
    # music
    (videos / "music" / "bgm.mp3").write_bytes(b"bgm")

    p = Project(
        slug="board1",
        topic="Scene board test",
        hero_mode="full_auto",
        status=ProjectStatus.assembled,
        meta={
            "excel_hero": {"characters": [{"id": "c01", "name": "Герой"}]},
            "bgm_enabled": True,
            "bgm_level": 65,
        },
    )
    session.add(p)
    await session.flush()

    fr1 = Frame(
        project_id=p.id,
        number=1,
        voiceover_text="Первая сцена закадровый",
        status=FrameStatus.done,
        start_ts=0.0,
        end_ts=3.5,
        duration_seconds=3.5,
        image_prompt="shot1 prompt",
        animation_prompt="anim1",
        attrs={"image_prompt_shot2": "Close-up second shot prompt here"},
    )
    fr2 = Frame(
        project_id=p.id,
        number=2,
        voiceover_text="Вторая сцена текст",
        status=FrameStatus.planned,
        # no timeslot yet
    )
    session.add_all([fr1, fr2])
    await session.flush()

    import uuid

    session.add(
        Artifact(
            project_id=p.id,
            kind=ArtifactKind.audio,
            uuid=uuid.uuid4().hex,
            path=str(videos / "audio" / "voice_full.mp3"),
            meta={
                "mode": "per_frame",
                "clips": [
                    {
                        "frame_number": 1,
                        "start_ts": 0.0,
                        "end_ts": 3.5,
                        "duration": 3.5,
                        "text": "Первая сцена закадровый",
                    },
                    {
                        "frame_number": 2,
                        "start_ts": 3.5,
                        "end_ts": 7.0,
                        "duration": 3.5,
                        "text": "Вторая сцена текст",
                    },
                ],
            },
        )
    )
    (videos / "audio" / "voice_full.mp3").write_bytes(b"full")
    await session.commit()

    board = await build_scene_board(session, p)
    assert board["frame_count"] == 2
    assert board["music"]["present"] is True
    assert board["music"]["label"] == "bgm.mp3"

    s1 = board["scenes"][0]
    assert s1["number"] == 1
    assert s1["timeslot_label"] == "0:00.00 – 0:03.50"
    assert s1["image_shot1"]["present"] is True
    assert s1["has_shot2"] is True
    assert s1["image_shot2"]["present"] is True
    assert s1["video_shot1"]["present"] is True
    assert s1["video_shot2"]["present"] is False
    assert s1["audio"]["present"] is True
    assert any(c["id"] == "c01" for c in s1["characters"])
    assert "video_shot2" in s1["missing"]
    assert "timeslot" not in s1["missing"]

    s2 = board["scenes"][1]
    assert s2["start_ts"] == 3.5  # from audio.meta.clips fallback
    assert s2["end_ts"] == 7.0
    assert s2["image_shot1"]["present"] is False
    assert "image_shot1" in s2["missing"]
    assert "audio" in s2["missing"]


def test_validate_regen_draft_filters_bad_targets() -> None:
    draft = validate_regen_draft(
        {
            "note": "  Сделай светлее  ",
            "selections": [
                {
                    "frame_id": 10,
                    "number": 1,
                    "targets": ["image_shot1", "bogus", "audio"],
                    "regen_type": "full_scene",
                },
                {"frame_id": 11, "number": 2, "targets": [], "regen_type": "media"},
            ],
        }
    )
    assert draft["note"] == "Сделай светлее"
    assert len(draft["selections"]) == 1
    assert draft["selections"][0]["targets"] == ["image_shot1", "audio"]
    assert draft["selections"][0]["regen_type"] == "full_scene"
