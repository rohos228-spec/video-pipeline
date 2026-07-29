"""split_xlsx: повторный запуск шага всегда идёт в GPT (не skip по блокам в xlsx)."""

from __future__ import annotations

from pathlib import Path
from unittest.mock import AsyncMock, patch

import pytest
from openpyxl import Workbook, load_workbook

from app.models import Project
from app.services import xlsx_step_runners as xsr
from app.services.xlsx_v8_import import ROW_VOICEOVER_V8, SHEET_PLAN_V8


def _save_split_xlsx(path: Path, *, blocks: int, enrich: str | None = None) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_PLAN_V8
    if enrich is not None:
        ws["A2"] = "enrich"
        ws["C2"] = enrich
    for i in range(blocks):
        ws.cell(row=ROW_VOICEOVER_V8, column=3 + i, value=f"block {i + 1}")
    # validate_xlsx ждёт v8-набор листов (хотя бы «Общий план» + «план»)
    wb.create_sheet("Общий план")
    wb.save(path)


@pytest.mark.asyncio
async def test_split_runs_gpt_even_when_project_already_has_blocks(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    proj_dir = tmp_path / "data" / "videos" / "p9"
    proj_dir.mkdir(parents=True)
    proj_xlsx = proj_dir / "project.xlsx"
    voiceover = proj_dir / "voiceover.txt"
    voiceover.write_text("x" * 300, encoding="utf-8")
    _save_split_xlsx(proj_xlsx, blocks=4, enrich="must-keep")

    tmp_dir = proj_dir / "tmp_gpt"
    tmp_dir.mkdir()

    project = Project(id=9, topic="t", slug="p9")
    monkeypatch.setattr("app.models.settings.data_dir", tmp_path / "data")
    monkeypatch.setattr(type(project), "data_dir", property(lambda self: proj_dir))

    async def fake_gpt(_chat, _files, download_path, **_k) -> str:
        # GPT отдаёт только R49 — без enrich (раньше full replace затирал C2)
        _save_split_xlsx(Path(download_path), blocks=4)
        return "ok"

    async def _run_lock(_pid, _step, fn):
        return await fn()

    gpt_mock = AsyncMock(side_effect=fake_gpt)
    with (
        patch.object(xsr.cx, "tmp_gpt_dir", return_value=tmp_dir),
        patch.object(xsr.cx, "write_split_prompt_file", return_value=tmp_dir / "p.txt"),
        patch.object(xsr.cx, "chat_message", return_value="msg"),
        patch.object(xsr, "_ts", return_value="20260101_120000"),
        patch.object(xsr, "_try_reuse_split_download", return_value=None),
        patch.object(xsr.xgf, "run_under_xlsx_lock", new=_run_lock),
        patch.object(xsr.xgf, "telegram_style_ask_and_download", gpt_mock),
    ):
        (tmp_dir / "p.txt").write_text("prompt", encoding="utf-8")
        result = await xsr.run_split_xlsx(project)

    gpt_mock.assert_called_once()
    assert result.project_xlsx == proj_xlsx
    assert xsr._count_v8_voiceover_blocks(proj_xlsx) == 4
    got = load_workbook(proj_xlsx)
    assert got[SHEET_PLAN_V8]["C2"].value == "must-keep"
    got.close()


@pytest.mark.asyncio
async def test_split_does_not_replace_project_when_gpt_returns_empty(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    proj_dir = tmp_path / "data" / "videos" / "p9"
    proj_dir.mkdir(parents=True)
    proj_xlsx = proj_dir / "project.xlsx"
    voiceover = proj_dir / "voiceover.txt"
    voiceover.write_text("x" * 300, encoding="utf-8")
    _save_split_xlsx(proj_xlsx, blocks=0)

    tmp_dir = proj_dir / "tmp_gpt"
    tmp_dir.mkdir()

    project = Project(id=9, topic="t", slug="p9")
    monkeypatch.setattr("app.models.settings.data_dir", tmp_path / "data")
    monkeypatch.setattr(type(project), "data_dir", property(lambda self: proj_dir))

    async def fake_gpt(_chat, _files, download_path, **_k) -> str:
        _save_split_xlsx(Path(download_path), blocks=0)
        return "ok"

    async def _run_lock(_pid, _step, fn):
        return await fn()

    with (
        patch.object(xsr.cx, "tmp_gpt_dir", return_value=tmp_dir),
        patch.object(xsr.cx, "write_split_prompt_file", return_value=tmp_dir / "p.txt"),
        patch.object(xsr.cx, "chat_message", return_value="msg"),
        patch.object(xsr, "_ts", return_value="20260101_120000"),
        patch.object(xsr, "_try_reuse_split_download", return_value=None),
        patch.object(xsr.xgf, "run_under_xlsx_lock", new=_run_lock),
        patch.object(xsr.xgf, "telegram_style_ask_and_download", new=fake_gpt),
    ):
        (tmp_dir / "p.txt").write_text("prompt", encoding="utf-8")
        with pytest.raises(RuntimeError, match="разбивка не найдена"):
            await xsr.run_split_xlsx(project)

    assert xsr._count_v8_voiceover_blocks(proj_xlsx) == 0
