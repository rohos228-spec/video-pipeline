"""Очередь generate_images: промты из xlsx и завершение шага."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from sqlalchemy import select

from app.models import Frame, FrameStatus
from app.orchestrator.steps.generate_images import _all_frames_have_image_or_failed
from app.services.scan_frames import frame_needs_shot1_image
from app.services.plan_shot2 import ROW_IMAGE_PROMPT_2_V8, SHOT2_PROMPT_ATTR, SHOT2_STATUS_ATTR
from app.services.xlsx_v8_import import (
    ROW_IMAGE_PROMPT_V8,
    bootstrap_frames_for_image_step,
    read_image_prompts_from_project_xlsx,
    read_v8_image_prompts_from_path,
)


def _write_plan_xlsx(path: Path, *, prompts: list[str], voiceovers: list[str | None]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "план"
    for i, prompt in enumerate(prompts, start=1):
        col = i + 2
        ws.cell(row=ROW_IMAGE_PROMPT_V8, column=col, value=prompt)
        vo = voiceovers[i - 1] if i - 1 < len(voiceovers) else None
        if vo:
            ws.cell(row=49, column=col, value=vo)
    wb.save(path)


def test_read_v8_image_prompts_without_voiceover(tmp_path: Path) -> None:
    xlsx = tmp_path / "project.xlsx"
    _write_plan_xlsx(
        xlsx,
        prompts=["prompt one", "prompt two"],
        voiceovers=[None, None],
    )
    got = read_v8_image_prompts_from_path(xlsx)
    assert got == {1: "prompt one", 2: "prompt two"}
    assert read_image_prompts_from_project_xlsx(xlsx) == got


def test_read_image_prompts_plan_sheet_trailing_space(tmp_path: Path) -> None:
    xlsx = tmp_path / "trailing.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "план "  # trailing space — частый баг Excel
    ws.cell(row=ROW_IMAGE_PROMPT_V8, column=3, value="prompt with space sheet")
    wb.save(xlsx)
    assert read_image_prompts_from_project_xlsx(xlsx) == {1: "prompt with space sheet"}


def test_read_image_prompts_by_row_label(tmp_path: Path) -> None:
    xlsx = tmp_path / "labeled.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "план"
    ws.cell(row=47, column=1, value="промт для картинки 1")
    ws.cell(row=47, column=3, value="from row 47")
    ws.cell(row=47, column=4, value="second frame")
    wb.save(xlsx)
    assert read_image_prompts_from_project_xlsx(xlsx) == {
        1: "from row 47",
        2: "second frame",
    }


def test_read_image_prompts_v7_kadry_sheet(tmp_path: Path) -> None:
    from app.storage.project_sheet import ROW_HEADER, ROW_IMAGE_PROMPT, SHEET_FRAMES

    xlsx = tmp_path / "v7.xlsx"
    wb = Workbook()
    ws_plan = wb.active
    ws_plan.title = "план"
    ws = wb.create_sheet(SHEET_FRAMES)
    ws.cell(row=ROW_HEADER, column=2, value=1)
    ws.cell(row=ROW_HEADER, column=3, value=2)
    ws.cell(row=ROW_IMAGE_PROMPT, column=2, value="v7 prompt 1")
    ws.cell(row=ROW_IMAGE_PROMPT, column=3, value="v7 prompt 2")
    wb.save(xlsx)
    assert read_image_prompts_from_project_xlsx(xlsx) == {
        1: "v7 prompt 1",
        2: "v7 prompt 2",
    }


def test_read_image_prompts_merged_row45(tmp_path: Path) -> None:
    """Ручной xlsx: merged A45:E45 — раньше читалось 0 промтов."""
    xlsx = tmp_path / "merged.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "план"
    ws.merge_cells("A45:E45")
    ws["A45"] = "prompt across merged cells"
    ws.cell(row=ROW_IMAGE_PROMPT_V8, column=6, value="second scene prompt")
    wb.save(xlsx)
    assert read_image_prompts_from_project_xlsx(xlsx) == {
        1: "prompt across merged cells",
        2: "second scene prompt",
    }


def test_read_image_prompts_r45_only_no_voiceover(tmp_path: Path) -> None:
    """Ручной импорт: только R45, без R49 — кадры по порядку колонок."""
    xlsx = tmp_path / "manual.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "план"
    ws.cell(row=ROW_IMAGE_PROMPT_V8, column=3, value="scene one img")
    ws.cell(row=ROW_IMAGE_PROMPT_V8, column=4, value="scene two img")
    wb.save(xlsx)
    assert read_image_prompts_from_project_xlsx(xlsx) == {
        1: "scene one img",
        2: "scene two img",
    }


async def test_bootstrap_writes_skippable_marked_prompts(tmp_path: Path, monkeypatch) -> None:
    """Bootstrap не фильтрует is_skippable — xlsx на диске = истина."""
    from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

    from app import settings as app_settings
    from app.generation_options import is_skippable_empty_prompt
    from app.models import Base, Project

    placeholder = (
        "КАДР 1 / PROMPT_1:\nнет исходных данных для заполнения"
    )
    assert is_skippable_empty_prompt(placeholder)

    monkeypatch.setattr(app_settings.settings, "data_dir", tmp_path)
    proj_dir = tmp_path / "videos" / "ph"
    proj_dir.mkdir(parents=True)
    xlsx = proj_dir / "project.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "план"
    ws.cell(row=ROW_IMAGE_PROMPT_V8, column=3, value=placeholder)
    ws.cell(row=49, column=3, value="voice")
    wb.save(xlsx)

    engine = create_async_engine("sqlite+aiosqlite:///:memory:")
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    factory = async_sessionmaker(engine, expire_on_commit=False)
    async with factory() as session:
        project = Project(slug="ph", topic="")
        session.add(project)
        await session.flush()
        boot = await bootstrap_frames_for_image_step(session, project, xlsx)
        assert boot.frames_prompt_updated == [1]
        fr = (
            await session.execute(
                select(Frame).where(Frame.project_id == project.id, Frame.number == 1)
            )
        ).scalar_one()
        assert "нет исходных данных" in (fr.image_prompt or "")
    await engine.dispose()


async def test_apply_prompts_to_many_frames_sekty_like(
    tmp_path: Path, monkeypatch
) -> None:
    """124 кадра в БД, 62 промта в xlsx — apply пишет в кадры 1..62."""
    from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

    from app import settings as app_settings
    from app.models import Base, Frame, FrameStatus, Project
    from app.services.xlsx_v8_import import apply_image_prompts_from_xlsx_to_frames

    monkeypatch.setattr(app_settings.settings, "data_dir", tmp_path)
    xlsx = tmp_path / "project.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "план"
    for i in range(62):
        col = i + 3
        ws.cell(row=ROW_IMAGE_PROMPT_V8, column=col, value=f"prompt scene {i + 1}")
        ws.cell(row=49, column=col, value=f"voice {i + 1}")
    wb.save(xlsx)

    engine = create_async_engine("sqlite+aiosqlite:///:memory:")
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    factory = async_sessionmaker(engine, expire_on_commit=False)
    async with factory() as session:
        project = Project(slug="sekty", topic="")
        session.add(project)
        await session.flush()
        for n in range(1, 125):
            session.add(
                Frame(
                    project_id=project.id,
                    number=n,
                    voiceover_text=f"v{n}",
                    status=FrameStatus.failed,
                    attrs={"fail_reason": "no_image_prompt"},
                )
            )
        await session.flush()
        frames = (
            await session.execute(select(Frame).where(Frame.project_id == project.id))
        ).scalars().all()
        n = apply_image_prompts_from_xlsx_to_frames(frames, xlsx)
        assert n >= 62
        fr1 = (
            await session.execute(
                select(Frame).where(Frame.project_id == project.id, Frame.number == 1)
            )
        ).scalar_one()
        fr62 = (
            await session.execute(
                select(Frame).where(Frame.project_id == project.id, Frame.number == 62)
            )
        ).scalar_one()
        fr63 = (
            await session.execute(
                select(Frame).where(Frame.project_id == project.id, Frame.number == 63)
            )
        ).scalar_one()
        assert fr1.image_prompt == "prompt scene 1"
        assert fr62.image_prompt == "prompt scene 62"
        assert not fr63.image_prompt
        assert fr1.status is FrameStatus.image_prompt_ready
    await engine.dispose()


async def test_bootstrap_manual_xlsx_empty_db(tmp_path: Path, monkeypatch) -> None:
    """Пустая БД + вручную положенный project.xlsx → кадры с промтами."""
    from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

    from app import settings as app_settings
    from app.models import Base, Project

    monkeypatch.setattr(app_settings.settings, "data_dir", tmp_path)
    proj_dir = tmp_path / "videos" / "manual_only"
    proj_dir.mkdir(parents=True)
    xlsx = proj_dir / "project.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "план"
    ws.cell(row=ROW_IMAGE_PROMPT_V8, column=3, value="manual prompt A")
    ws.cell(row=ROW_IMAGE_PROMPT_V8, column=4, value="manual prompt B")
    wb.save(xlsx)

    engine = create_async_engine("sqlite+aiosqlite:///:memory:")
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    factory = async_sessionmaker(engine, expire_on_commit=False)
    async with factory() as session:
        project = Project(slug="manual_only", topic="")
        session.add(project)
        await session.flush()
        boot = await bootstrap_frames_for_image_step(session, project, xlsx)
        assert boot.prompts_in_xlsx == 2
        assert boot.frames_created == [1, 2]
        frames = (
            await session.execute(select(Frame).where(Frame.project_id == project.id))
        ).scalars().all()
        assert len(frames) == 2
        assert frames[0].image_prompt == "manual prompt A"
        assert frames[1].image_prompt == "manual prompt B"
    await engine.dispose()


def test_read_sekty_wide_sheet_124_prompts() -> None:
    """Реальный sekty: 150 колонок, 124 R45 — без искусственного лимита ширины."""
    from pathlib import Path

    p = Path("/home/ubuntu/.cursor/projects/workspace/uploads/sekty-project_54dc.xlsx")
    if not p.is_file():
        return
    from app.generation_options import is_skippable_empty_prompt

    prompts = read_image_prompts_from_project_xlsx(p)
    assert len(prompts) >= 118
    assert sum(1 for t in prompts.values() if not is_skippable_empty_prompt(t)) >= 118


async def test_all_frames_done_checks_prompted_frames_not_empty_ones(
    tmp_path: Path,
) -> None:
    from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

    from app.models import Base, Project

    scenes = tmp_path / "scenes"
    scenes.mkdir()
    engine = create_async_engine("sqlite+aiosqlite:///:memory:")
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    factory = async_sessionmaker(engine, expire_on_commit=False)
    async with factory() as session:
        project = Project(slug="t", topic="t")
        session.add(project)
        await session.flush()
        session.add(
            Frame(
                project_id=project.id,
                number=1,
                voiceover_text="v1",
                image_prompt="p1",
                status=FrameStatus.image_prompt_ready,
            )
        )
        session.add(
            Frame(
                project_id=project.id,
                number=2,
                voiceover_text="v2",
                image_prompt="",
                status=FrameStatus.planned,
            )
        )
        await session.commit()
        assert await _all_frames_have_image_or_failed(session, project.id, scenes) is False
        png = scenes / "frame_001_abcd1234.png"
        png.write_bytes(b"\x89PNG\r\n\x1a\n" + b"x" * 250_000)
        assert await _all_frames_have_image_or_failed(session, project.id, scenes) is True
    await engine.dispose()


def test_failed_frame_skipped_from_queue(tmp_path: Path) -> None:
    scenes = tmp_path / "scenes"
    scenes.mkdir()
    fr = Frame(project_id=1, number=1, voiceover_text="v", image_prompt="prompt")
    fr.status = FrameStatus.failed
    assert frame_needs_shot1_image(fr, scenes) is False


async def test_bootstrap_creates_frames_from_xlsx_only(
    tmp_path: Path, monkeypatch
) -> None:
    from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

    from app import settings as app_settings
    from app.models import Base, Project

    monkeypatch.setattr(app_settings.settings, "data_dir", tmp_path)
    proj_dir = tmp_path / "videos" / "manual"
    proj_dir.mkdir(parents=True)
    xlsx = proj_dir / "project.xlsx"
    _write_plan_xlsx(
        xlsx,
        prompts=["first prompt", "second prompt"],
        voiceovers=["vo1", "vo2"],
    )

    engine = create_async_engine("sqlite+aiosqlite:///:memory:")
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    factory = async_sessionmaker(engine, expire_on_commit=False)
    async with factory() as session:
        project = Project(slug="manual", topic="")
        session.add(project)
        await session.flush()
        boot = await bootstrap_frames_for_image_step(session, project, xlsx)
        assert boot.prompts_in_xlsx == 2
        assert boot.frames_created == [1, 2]
        frames = (
            await session.execute(
                select(Frame).where(Frame.project_id == project.id)
            )
        ).scalars().all()
        assert len(frames) == 2
        assert frames[0].image_prompt == "first prompt"
        assert frames[0].status is FrameStatus.image_prompt_ready
    await engine.dispose()


async def test_bootstrap_resets_failed_when_xlsx_has_prompt(
    tmp_path: Path, monkeypatch
) -> None:
    from sqlalchemy import select
    from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

    from app import settings as app_settings
    from app.models import Base, Project

    monkeypatch.setattr(app_settings.settings, "data_dir", tmp_path)
    proj_dir = tmp_path / "videos" / "t2"
    proj_dir.mkdir(parents=True)
    xlsx = proj_dir / "project.xlsx"
    _write_plan_xlsx(xlsx, prompts=["p1"], voiceovers=["v1"])

    engine = create_async_engine("sqlite+aiosqlite:///:memory:")
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    factory = async_sessionmaker(engine, expire_on_commit=False)
    async with factory() as session:
        project = Project(slug="t2", topic="")
        session.add(project)
        await session.flush()
        session.add(
            Frame(
                project_id=project.id,
                number=1,
                voiceover_text="v",
                image_prompt="stale",
                status=FrameStatus.failed,
                attrs={"fail_reason": "no_image_prompt"},
            )
        )
        await session.commit()
        boot = await bootstrap_frames_for_image_step(session, project, xlsx)
        assert 1 in boot.frames_status_reset
        fr = (
            await session.execute(
                select(Frame).where(Frame.project_id == project.id, Frame.number == 1)
            )
        ).scalar_one()
        assert fr.image_prompt == "p1"
        assert fr.status is FrameStatus.image_prompt_ready
        assert "fail_reason" not in (fr.attrs or {})
    await engine.dispose()


async def test_bootstrap_applies_shot2_from_r46(tmp_path: Path, monkeypatch) -> None:
    from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

    from app import settings as app_settings
    from app.models import Base, Project

    monkeypatch.setattr(app_settings.settings, "data_dir", tmp_path)
    proj_dir = tmp_path / "videos" / "shot2boot"
    proj_dir.mkdir(parents=True)
    xlsx = proj_dir / "project.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "план"
    ws.cell(row=ROW_IMAGE_PROMPT_V8, column=3, value="wide establishing shot")
    ws.cell(row=ROW_IMAGE_PROMPT_2_V8, column=3, value="tight reaction close-up")
    ws.cell(row=49, column=3, value="voice for scene one")
    wb.save(xlsx)

    engine = create_async_engine("sqlite+aiosqlite:///:memory:")
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    factory = async_sessionmaker(engine, expire_on_commit=False)
    async with factory() as session:
        project = Project(slug="shot2boot", topic="")
        session.add(project)
        await session.flush()
        boot = await bootstrap_frames_for_image_step(session, project, xlsx)
        assert boot.prompts_in_xlsx == 1
        assert boot.shot2_in_xlsx == 1
        assert boot.frames_shot2_updated == [1]
        fr = (
            await session.execute(
                select(Frame).where(Frame.project_id == project.id, Frame.number == 1)
            )
        ).scalar_one()
        assert fr.attrs[SHOT2_PROMPT_ATTR] == "tight reaction close-up"
        assert fr.attrs[SHOT2_STATUS_ATTR] == "image_prompt_ready"
    await engine.dispose()
