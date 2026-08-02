"""gpt_operator_client: project_file → только apply-ops JSON (без TSV)."""

from __future__ import annotations

from types import SimpleNamespace

import pytest
from openpyxl import Workbook, load_workbook

from app.services import gpt_operator_client as goc


@pytest.mark.asyncio
async def test_project_file_apply_ops_skips_tsv_writeback(tmp_path, monkeypatch) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "план"
    ws["A45"] = "промт для картинки 1"
    ws["C45"] = "old"
    wb.save(tmp_path / "project.xlsx")
    wb.close()

    reply = '{"ops":[{"frame_uuid":"u1","fields":{"закадр":"новый"}}]}'
    monkeypatch.setattr("app.services.gpt_api.gpt_api_enabled", lambda: True)

    async def fake_chat(**kw):
        return SimpleNamespace(text=reply)

    monkeypatch.setattr("app.services.gpt_api.chat", fake_chat)
    monkeypatch.setattr("app.services.gpt_api.collect_result_urls", lambda text: [])

    res = await goc.run_operator_api(
        project_dir=tmp_path,
        node_key="n_excel_gpt_1",
        role="assist",
        output_mode="project_file",
        prompt="заполни",
        accompanying="",
        input_paths=[tmp_path / "project.xlsx"],
    )
    assert res.apply_ops is not None
    assert res.apply_ops["ops"][0]["frame_uuid"] == "u1"
    # xlsx не тронут (TSV-writeback пропущен — запись пойдёт через db_apply).
    wb2 = load_workbook(tmp_path / "project.xlsx")
    try:
        assert wb2["план"]["C45"].value == "old"
    finally:
        wb2.close()


@pytest.mark.asyncio
async def test_project_file_tsv_rejected_retries_apply_ops(tmp_path, monkeypatch) -> None:
    """TSV без JSON — не пишем Excel; retry просит apply-ops; при успехе — ops."""
    wb = Workbook()
    ws = wb.active
    ws.title = "план"
    ws["A45"] = "промт для картинки 1"
    ws["C45"] = "old"
    wb.save(tmp_path / "project.xlsx")
    wb.close()

    calls: list[str] = []
    tsv = "# Лист: план\n@row=45\tпромт для картинки 1\t\tnew prompt\n"
    ops = '{"ops":[{"frame_uuid":"u1","fields":{"закадр":"из retry"}}]}'

    monkeypatch.setattr("app.services.gpt_api.gpt_api_enabled", lambda: True)

    async def fake_chat(**kw):
        calls.append(kw.get("prompt") or "")
        return SimpleNamespace(text=tsv if len(calls) == 1 else ops)

    monkeypatch.setattr("app.services.gpt_api.chat", fake_chat)
    monkeypatch.setattr("app.services.gpt_api.collect_result_urls", lambda text: [])

    res = await goc.run_operator_api(
        project_dir=tmp_path,
        node_key="n_excel_gpt_1",
        role="assist",
        output_mode="project_file",
        prompt="заполни",
        accompanying="",
        input_paths=[tmp_path / "project.xlsx"],
    )
    assert len(calls) == 2
    assert "apply-ops" in calls[1].lower() or "JSON" in calls[1]
    assert res.apply_ops is not None
    assert res.apply_ops["ops"][0]["fields"]["закадр"] == "из retry"
    wb2 = load_workbook(tmp_path / "project.xlsx")
    try:
        # TSV не должен был переписать книгу.
        assert wb2["план"]["C45"].value == "old"
    finally:
        wb2.close()


@pytest.mark.asyncio
async def test_project_file_without_apply_ops_fails(tmp_path, monkeypatch) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "план"
    wb.save(tmp_path / "project.xlsx")
    wb.close()

    monkeypatch.setattr("app.services.gpt_api.gpt_api_enabled", lambda: True)

    async def fake_chat(**kw):
        return SimpleNamespace(text="просто проза без JSON")

    monkeypatch.setattr("app.services.gpt_api.chat", fake_chat)
    monkeypatch.setattr("app.services.gpt_api.collect_result_urls", lambda text: [])

    with pytest.raises(RuntimeError, match="apply-ops JSON"):
        await goc.run_operator_api(
            project_dir=tmp_path,
            node_key="n_excel_gpt_1",
            role="assist",
            output_mode="project_file",
            prompt="заполни",
            accompanying="",
            input_paths=[tmp_path / "project.xlsx"],
        )
