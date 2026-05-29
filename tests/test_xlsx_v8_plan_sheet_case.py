"""v8 xlsx: лист «план» находится без учёта регистра."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from app.services.xlsx_v8_import import (
    ROW_VOICEOVER_V8,
    _read_voiceover_blocks,
    has_v8_plan_sheet,
)
from app.services import xlsx_step_runners as xsr


def test_voiceover_blocks_case_insensitive_sheet_name(tmp_path: Path) -> None:
    p = tmp_path / "split.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "План"
    ws.cell(row=ROW_VOICEOVER_V8, column=3, value="block one")
    ws.cell(row=ROW_VOICEOVER_V8, column=4, value="block two")
    wb.save(p)

    wb2 = __import__("openpyxl").load_workbook(p, data_only=True)
    assert has_v8_plan_sheet(wb2)
    assert _read_voiceover_blocks(wb2) == ["block one", "block two"]
    assert xsr._count_v8_voiceover_blocks(p) == 2
