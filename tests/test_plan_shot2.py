"""shot_02: чтение xlsx и имена файлов на диске."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from app.services.plan_shot2 import (
    ROW_IMAGE_PROMPT_2_V8,
    ROW_SHOT2_ACTION_V8,
    ROW_SHOT2_ID_SHOT_V8,
    ROW_VOICEOVER_V8,
    disk_has_shot2_image,
    find_shot1_image,
    read_shot2_columns,
)
from app.services.xlsx_v8_import import ROW_IMAGE_PROMPT_V8


def _write_min_v8(
    path: Path,
    *,
    voiceover: str = "Текст закадровки для сцены один.",
    prompt1: str = "wide shot hero",
    prompt2: str = "",
    shot2_id: str = "",
    shot2_action: str = "",
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "план"
    col = 3
    ws.cell(row=ROW_VOICEOVER_V8, column=col, value=voiceover)
    ws.cell(row=ROW_IMAGE_PROMPT_V8, column=col, value=prompt1)
    ws.cell(row=ROW_IMAGE_PROMPT_2_V8, column=col, value=prompt2)
    ws.cell(row=ROW_SHOT2_ID_SHOT_V8, column=col, value=shot2_id)
    ws.cell(row=ROW_SHOT2_ACTION_V8, column=col, value=shot2_action)
    wb.save(path)


def test_read_shot2_no_block(tmp_path: Path) -> None:
    p = tmp_path / "plan.xlsx"
    _write_min_v8(p, prompt2="", shot2_id="", shot2_action="")
    by_num = read_shot2_columns(p)
    assert 1 in by_num
    assert by_num[1].has_shot2 is False


def test_read_shot2_from_row46(tmp_path: Path) -> None:
    p = tmp_path / "plan.xlsx"
    _write_min_v8(
        p,
        prompt2="close-up on hands opening envelope",
        shot2_id="shot_02",
    )
    info = read_shot2_columns(p)[1]
    assert info.has_shot2
    assert info.prompt == "close-up on hands opening envelope"


def test_read_shot2_fallback_action_from_block(tmp_path: Path) -> None:
    p = tmp_path / "plan.xlsx"
    _write_min_v8(
        p,
        prompt2="",
        shot2_id="shot_02",
        shot2_action="Extreme close-up: trembling fingers on the diagnosis line",
    )
    info = read_shot2_columns(p)[1]
    assert info.has_shot2
    assert "trembling fingers" in info.prompt


def test_shot1_and_shot2_disk_helpers(tmp_path: Path) -> None:
    scenes = tmp_path / "scenes"
    scenes.mkdir()
    (scenes / "frame_001_abc12345.png").write_bytes(b"1")
    (scenes / "frame_001_s2_def67890.png").write_bytes(b"2")

    assert find_shot1_image(scenes, 1) is not None
    assert find_shot1_image(scenes, 1).name == "frame_001_abc12345.png"
    assert disk_has_shot2_image(scenes, 1)
    assert not disk_has_shot2_image(scenes, 2)
