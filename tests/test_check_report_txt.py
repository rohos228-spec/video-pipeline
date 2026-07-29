"""TXT-отчёт проверки + checkMode на excel_gpt."""

from __future__ import annotations

import json
from pathlib import Path

import pytest

from app.models import Project, ProjectStatus
from app.services.check_analysis import (
    CHECK_REPORT_NAME,
    SCHEMA_ID,
    append_txt_report_footer,
    parse_check_analysis,
    parse_check_report_txt,
    parse_gate_status,
    render_check_report_txt,
    write_check_report_txt,
)
from app.services.gpt_operator import (
    assemble_check_master_prompt,
    collect_source_prompts,
    gate_allows_successors,
    patch_operator_config,
    resolve_operator,
    save_operator_result,
    verdict_edge_blocks,
)
from app.services.gpt_operator_client import run_operator_api


SAMPLE_TXT = """
# ОТЧЁТ ПРОВЕРКИ
verdict: fail
mode: report_only
source_prompts: n_excel_gpt_1

## summary
Файл не соответствует промту заполнения.

## analysis
Проверяли колонки по исходному промту источника.

## findings
- [error] пустая ячейка R10
- [warn] странный тон

## related
- error_1 → промт:n_excel_gpt_1 | лист:Сцены | строка:10 | поле:R

## logic
Часть строк заполнена верно.
Пустые обязательные поля — ошибка.

## actions
Правки не делались (только отчёт).

## forward
file: original
path: —
""".strip()


def _project(tmp_path: Path, monkeypatch) -> Project:
    from app import settings as app_settings

    monkeypatch.setattr(app_settings.settings, "data_dir", tmp_path / "data")
    p = Project(
        slug="check-txt",
        topic="t",
        status=ProjectStatus.new,
        hero_mode="no_hero",
        meta={},
    )
    p.data_dir.mkdir(parents=True)
    return p


def test_parse_check_report_txt_verdict() -> None:
    a = parse_check_report_txt(SAMPLE_TXT)
    assert a is not None
    assert a.verdict == "fail"
    assert "не соответствует" in a.summary
    assert any("пустая" in c.note for c in a.checks)
    assert parse_gate_status(SAMPLE_TXT) == "fail"
    assert parse_check_analysis(SAMPLE_TXT).verdict == "fail"


def test_json_fallback_renders_txt(tmp_path: Path) -> None:
    raw = json.dumps(
        {
            "schema": SCHEMA_ID,
            "verdict": "pass",
            "summary": "ок",
            "checks": [{"id": "c1", "ok": True, "note": "норм"}],
            "forward": {"mode": "inherit", "paths": []},
            "fix": {"target": "none", "instructions": "", "rewrite_file": None},
        },
        ensure_ascii=False,
    )
    a = parse_check_analysis(raw)
    text = render_check_report_txt(a, mode="fix", source_prompts=["n1"])
    assert "# ОТЧЁТ ПРОВЕРКИ" in text
    assert "verdict: pass" in text
    path = write_check_report_txt(tmp_path, a, mode="report_only", source_prompts=["n1"])
    assert path.name == CHECK_REPORT_NAME
    assert "mode: report_only" in path.read_text(encoding="utf-8")


def test_txt_footer_appended() -> None:
    out = append_txt_report_footer("проверь по промту")
    assert "проверь по промту" in out
    assert "# ОТЧЁТ ПРОВЕРКИ" in out
    assert "бинарный" in out.lower() or "текстовый экспорт" in out.lower()
    assert "XLSX_WRITEBACK" in out
    assert append_txt_report_footer(out) == out


def test_split_check_reply_and_writeback() -> None:
    from app.services.check_analysis import split_check_reply_and_writeback

    report = SAMPLE_TXT
    wb = "# Лист: Общий план\nA\tB\n1\t2"
    full = f"{report}\n\n--- XLSX_WRITEBACK ---\n{wb}"
    r, w = split_check_reply_and_writeback(full)
    assert "verdict: fail" in r
    assert "# Лист: Общий план" in w
    assert "XLSX_WRITEBACK" not in w

    full2 = f"{report}\n\n# Лист: Общий план\nx\ty"
    r2, w2 = split_check_reply_and_writeback(full2)
    assert "## forward" in r2
    assert w2.startswith("# Лист:")


@pytest.mark.asyncio
async def test_check_fix_writeback_applies_tsv(tmp_path: Path, monkeypatch) -> None:
    from openpyxl import Workbook, load_workbook

    from app.services import gpt_api
    from app.services.gpt_api import GptChatResult

    src = tmp_path / "project.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Общий план"
    ws["A1"] = "Главная тема"
    ws["B1"] = "старая"
    wb.create_sheet("прочее")["A1"] = "keep"
    wb.save(src)

    reply = """
# ОТЧЁТ ПРОВЕРКИ
verdict: fail
mode: fix
source_prompts: n_plan

## summary
Исправил тему.

## analysis
Тема была пустой/битой.

## findings
- [error] Главная тема

## related
- error_1 → промт:n_plan | лист:Общий план | поле:Главная тема

## logic
Остальное ок.

## actions
Обновил лист Общий план.

## forward
file: fixed
path: excel_gpt_uploads/n_check/project_fixed.xlsx

--- XLSX_WRITEBACK ---
# Лист: Общий план
Главная тема\tновая тема ролика
""".strip()

    async def fake_chat(**kwargs):  # noqa: ANN003
        return GptChatResult(text=reply, model="test")

    monkeypatch.setattr(gpt_api, "gpt_api_enabled", lambda: True)
    monkeypatch.setattr(gpt_api, "chat", fake_chat)

    out = await run_operator_api(
        project_dir=tmp_path,
        node_key="n_check",
        role="assist",
        output_mode="text",
        prompt="проверь",
        accompanying="",
        input_paths=[src],
        check_mode=True,
        check_fix=True,
        source_prompt_keys=["n_plan"],
    )
    assert out.gate_status == "fail"
    assert out.analysis is not None
    assert out.analysis.fix.rewrite_file == "excel_gpt_uploads/n_check/project_fixed.xlsx"
    fixed = tmp_path / "excel_gpt_uploads" / "n_check" / "project_fixed.xlsx"
    assert fixed.is_file()
    got = load_workbook(fixed)
    assert got["Общий план"]["B1"].value == "новая тема ролика"
    assert got["прочее"]["A1"].value == "keep"


@pytest.mark.asyncio
async def test_stub_api_writes_check_report(tmp_path: Path, monkeypatch) -> None:
    from app.services import gpt_api

    monkeypatch.setattr(gpt_api, "gpt_api_enabled", lambda: False)
    out = await run_operator_api(
        project_dir=tmp_path,
        node_key="n_check",
        role="assist",
        output_mode="text",
        prompt="проверь",
        accompanying="",
        input_paths=[],
        check_mode=True,
        check_fix=False,
        source_prompt_keys=["n_src"],
    )
    report = tmp_path / "excel_gpt_uploads" / "n_check" / CHECK_REPORT_NAME
    assert report.is_file()
    body = report.read_text(encoding="utf-8")
    assert "verdict:" in body
    assert "mode: report_only" in body
    assert out.gate_status in ("pass", "fail")
    assert any(p.name == CHECK_REPORT_NAME for p in out.output_paths)


def test_check_mode_resolve_needs_source_prompt(tmp_path: Path, monkeypatch) -> None:
    p = _project(tmp_path, monkeypatch)
    src = "n_excel_gpt_1"
    chk = "n_excel_gpt_2"
    up = p.data_dir / "excel_gpt_uploads" / src
    up.mkdir(parents=True)
    xlsx = p.data_dir / "project.xlsx"
    xlsx.write_bytes(b"PK\x03\x04" + b"0" * 200)
    (up / "out.xlsx").write_bytes(b"PK\x03\x04" + b"1" * 200)
    p.meta = {
        "canvas_graph": {
            "nodes": [
                {"id": src, "type": "excel_gpt", "position": {"x": 0, "y": 0}},
                {"id": chk, "type": "excel_gpt", "position": {"x": 200, "y": 0}},
            ],
            "edges": [
                {"id": "e1", "source": src, "target": chk, "data": {"kind": "after"}}
            ],
        },
        "excel_gpt_nodes": {
            src: {"role": "assist", "transport": "api", "emitKinds": ["result"]},
            chk: {"role": "assist", "transport": "api", "checkMode": True},
        },
        "prompt_slot_variants": {
            src: {"main": "default"},
        },
        "gpt_operator_results": {
            src: {
                "outputPaths": [str(up / "out.xlsx")],
                "inputPaths": [str(xlsx)],
            }
        },
    }
    # Без реального текста промта — canRun false
    res = resolve_operator(p, chk)
    assert res["checkMode"] is True
    # default.md для excel_gpt обычно есть в репо — тогда ok
    sources = collect_source_prompts(p, chk)
    assert sources
    assert sources[0]["nodeKey"] == src

    patch_operator_config(p, chk, {"checkMode": True, "checkFix": False})
    res2 = resolve_operator(p, chk)
    assert res2["checkFix"] is False
    assert res2["branching"]["enabled"] is True


def test_assemble_prompt_includes_sources() -> None:
    text = assemble_check_master_prompt(
        [
            {
                "ok": True,
                "nodeKey": "n1",
                "text": "Заполни колонки A–C",
                "accompanying": "без общего плана",
            }
        ],
        check_fix=True,
        reviewer_notes="смотри R48",
    )
    assert "### source: n1" in text
    assert "Заполни колонки" in text
    assert "без общего плана" in text
    assert "смотри R48" in text
    assert "# ОТЧЁТ ПРОВЕРКИ" in text


def test_hydrate_result_from_disk_when_meta_stale(tmp_path: Path, monkeypatch) -> None:
    from app.services.gpt_operator import hydrate_check_result_from_disk, resolve_operator

    p = _project(tmp_path, monkeypatch)
    key = "n_check"
    up = p.data_dir / "excel_gpt_uploads" / key
    up.mkdir(parents=True)
    (up / "check_report.txt").write_text(SAMPLE_TXT, encoding="utf-8")
    (up / "analysis.json").write_text(
        json.dumps(parse_check_analysis(SAMPLE_TXT).to_dict(), ensure_ascii=False),
        encoding="utf-8",
    )
    (p.data_dir / "project.xlsx").write_bytes(b"PK\x03\x04" + b"0" * 200)
    p.meta = {
        "canvas_graph": {
            "nodes": [{"id": key, "type": "excel_gpt", "position": {"x": 0, "y": 0}}],
            "edges": [],
        },
        "excel_gpt_nodes": {
            key: {
                "role": "assist",
                "checkMode": True,
                "transport": "api",
                "uploadedFileNames": ["project.xlsx"],
                "inputSource": "upload",
            }
        },
        # stale meta без gate — как после web_get overwrite
        "gpt_operator_results": {
            key: {
                "gateStatus": None,
                "replyPreview": "старый эссе-текст",
                "outputPaths": [],
            }
        },
    }
    # upload path for project.xlsx under node
    import shutil

    shutil.copy(p.data_dir / "project.xlsx", up / "project.xlsx")
    p.meta["excel_gpt_nodes"][key]["uploadedFileNames"] = ["project.xlsx"]

    hydrated = hydrate_check_result_from_disk(p, key, {"gateStatus": None})
    assert hydrated["gateStatus"] == "fail"
    assert hydrated["analysis"]["verdict"] == "fail"

    res = resolve_operator(p, key)
    assert res["branching"]["verdict"] == "fail"
    assert res["analysis"]["verdict"] == "fail"


def test_sanitize_strips_old_json_contract() -> None:
    from app.services.gpt_operator import sanitize_check_reviewer_notes

    dirty = 'Верни ответ\n{\n  "schema": "vp.check.v1",\n  "verdict": "pass"\n}'
    assert sanitize_check_reviewer_notes(dirty) == ""
    assert sanitize_check_reviewer_notes("смотри только лист План") == "смотри только лист План"


def test_check_mode_gate_edges(tmp_path: Path, monkeypatch) -> None:
    p = _project(tmp_path, monkeypatch)
    key = "n_check"
    p.meta = {
        "canvas_graph": {
            "nodes": [{"id": key, "type": "excel_gpt", "position": {"x": 0, "y": 0}}],
            "edges": [
                {
                    "id": "ep",
                    "source": key,
                    "target": "n_ok",
                    "data": {"kind": "pass"},
                },
                {
                    "id": "ef",
                    "source": key,
                    "target": "n_fail",
                    "data": {"kind": "fail"},
                },
            ],
        },
        "excel_gpt_nodes": {
            key: {"role": "assist", "checkMode": True, "transport": "api"}
        },
    }
    save_operator_result(
        p,
        key,
        input_paths=[],
        output_paths=[],
        reply_text=SAMPLE_TXT,
        gate_status="fail",
        analysis=parse_check_analysis(SAMPLE_TXT).to_dict(),
    )
    assert gate_allows_successors(p, key) is False
    assert verdict_edge_blocks(p, key, "pass") is True
    assert verdict_edge_blocks(p, key, "fail") is False
