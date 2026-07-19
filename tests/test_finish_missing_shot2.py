"""Доделка shot_02: картинки, видео, anim_pr."""

from __future__ import annotations

from pathlib import Path
from unittest.mock import AsyncMock, MagicMock

import pytest
from openpyxl import Workbook

from app.models import Frame, FrameStatus, Project, ProjectStatus
from app.services.finish_missing import (
    trigger_finish_missing_images,
    trigger_finish_missing_videos,
)
from app.services.plan_shot2 import (
    ROW_IMAGE_PROMPT_2_V8,
    ROW_SHOT2_ID_SHOT_V8,
    ROW_VIDEO_PROMPT_2_V8,
    SHOT2_PROMPT_ATTR,
    SHOT2_STATUS_ATTR,
    SHOT2_VIDEO_STATUS_ATTR,
)
from app.services.scan_frames import (
    _disk_has_frame_video_shot1,
    reset_shot2_to_prompt_ready,
    scan_missing_shot2_frames,
    scan_missing_shot2_videos,
    scan_missing_videos_shot1,
)
from app.services.xlsx_v8_import import ROW_IMAGE_PROMPT_V8, ROW_VOICEOVER_V8


def _write_plan_with_shot2(
    path: Path,
    *,
    prompt2: str = "close-up hands",
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "план"
    col = 3
    ws.cell(row=ROW_VOICEOVER_V8, column=col, value="voiceover")
    ws.cell(row=ROW_IMAGE_PROMPT_V8, column=col, value="wide shot")
    ws.cell(row=ROW_IMAGE_PROMPT_2_V8, column=col, value=prompt2)
    wb.save(path)


def _valid_png(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(b"\x89PNG\r\n\x1a\n" + b"x" * 250_000)


def _project(data_dir: Path, *, status: ProjectStatus = ProjectStatus.images_ready) -> MagicMock:
    p = MagicMock(spec=Project)
    p.id = 17
    p.data_dir = data_dir
    p.status = status
    return p


@pytest.mark.asyncio
async def test_scan_missing_shot2_when_s1_ok_s2_absent(tmp_path: Path) -> None:
    data_dir = tmp_path / "p17"
    scenes = data_dir / "scenes"
    _valid_png(scenes / "frame_001_abc12345.png")
    xlsx = data_dir / "project.xlsx"
    _write_plan_with_shot2(xlsx)

    project = _project(data_dir)
    fr = Frame(project_id=17, number=1, voiceover_text="v", image_prompt="wide")
    session = AsyncMock()
    session.execute = AsyncMock(
        return_value=MagicMock(scalars=MagicMock(return_value=MagicMock(all=lambda: [fr])))
    )

    missing = await scan_missing_shot2_frames(session, project)
    assert missing == [1]


@pytest.mark.asyncio
async def test_scan_missing_shot2_ok_when_s2_on_disk(tmp_path: Path) -> None:
    data_dir = tmp_path / "p17"
    scenes = data_dir / "scenes"
    _valid_png(scenes / "frame_001_abc12345.png")
    _valid_png(scenes / "frame_001_s2_def67890.png")
    _write_plan_with_shot2(data_dir / "project.xlsx")

    project = _project(data_dir)
    fr = Frame(project_id=17, number=1, voiceover_text="v", image_prompt="wide")
    session = AsyncMock()
    session.execute = AsyncMock(
        return_value=MagicMock(scalars=MagicMock(return_value=MagicMock(all=lambda: [fr])))
    )

    missing = await scan_missing_shot2_frames(session, project)
    assert missing == []


@pytest.mark.asyncio
async def test_reset_shot2_sets_attrs(tmp_path: Path) -> None:
    data_dir = tmp_path / "p17"
    scenes = data_dir / "scenes"
    _valid_png(scenes / "frame_001_abc12345.png")
    _write_plan_with_shot2(data_dir / "project.xlsx", prompt2="macro eye")

    project = _project(data_dir)
    fr = Frame(project_id=17, number=1, voiceover_text="v", image_prompt="wide")
    session = AsyncMock()
    session.execute = AsyncMock(
        return_value=MagicMock(scalars=MagicMock(return_value=MagicMock(all=lambda: [fr])))
    )
    session.flush = AsyncMock()

    n = await reset_shot2_to_prompt_ready(session, project, [1])
    assert n == 1
    assert fr.attrs[SHOT2_PROMPT_ATTR] == "macro eye"
    assert fr.attrs[SHOT2_STATUS_ATTR] == "image_prompt_ready"


@pytest.mark.asyncio
async def test_trigger_finish_missing_only_shot2(tmp_path: Path, monkeypatch) -> None:
    data_dir = tmp_path / "p17"
    scenes = data_dir / "scenes"
    _valid_png(scenes / "frame_001_abc12345.png")
    _write_plan_with_shot2(data_dir / "project.xlsx")

    project = _project(data_dir, status=ProjectStatus.images_ready)
    fr = Frame(project_id=17, number=1, voiceover_text="v", image_prompt="wide")
    fr.status = FrameStatus.image_generated
    session = AsyncMock()
    session.execute = AsyncMock(
        return_value=MagicMock(scalars=MagicMock(return_value=MagicMock(all=lambda: [fr])))
    )
    session.flush = AsyncMock()

    info = await trigger_finish_missing_images(session, project)
    assert info["missing_shot2"] == [1]
    assert info["missing_shot1"] == []
    assert info["queued_shot2"] == 1
    assert project.status.value == "generating_images"


def test_video_shot1_glob_ignores_s2(tmp_path: Path) -> None:
    videos = tmp_path / "videos"
    videos.mkdir()
    (videos / "clip_001_s2_only.mp4").write_bytes(b"x" * 1000)
    assert _disk_has_frame_video_shot1(videos, 1) is False
    (videos / "clip_001_abc12345.mp4").write_bytes(b"x" * 1000)
    assert _disk_has_frame_video_shot1(videos, 1) is True


def _write_plan_shot2_video(
    path: Path,
    *,
    prompt2_img: str = "close-up",
    prompt2_video: str = "Slow dolly in on hands.",
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "план"
    col = 3
    ws.cell(row=ROW_VOICEOVER_V8, column=col, value="voiceover")
    ws.cell(row=ROW_IMAGE_PROMPT_V8, column=col, value="wide shot")
    ws.cell(row=ROW_IMAGE_PROMPT_2_V8, column=col, value=prompt2_img)
    ws.cell(row=ROW_SHOT2_ID_SHOT_V8, column=col, value="shot_02")
    ws.cell(row=ROW_VIDEO_PROMPT_2_V8, column=col, value=prompt2_video)
    wb.save(path)


@pytest.mark.asyncio
async def test_scan_missing_shot2_video(tmp_path: Path) -> None:
    data_dir = tmp_path / "p17"
    scenes = data_dir / "scenes"
    videos = data_dir / "videos"
    videos.mkdir(parents=True)
    _valid_png(scenes / "frame_001_abc12345.png")
    _valid_png(scenes / "frame_001_s2_def67890.png")
    (videos / "clip_001_aaa11111.mp4").write_bytes(b"x" * 1000)
    _write_plan_shot2_video(data_dir / "project.xlsx")

    project = _project(data_dir)
    fr = Frame(project_id=17, number=1, voiceover_text="v", image_prompt="wide")
    fr.animation_prompt = "Camera pan left slowly."
    session = AsyncMock()
    session.execute = AsyncMock(
        return_value=MagicMock(scalars=MagicMock(return_value=MagicMock(all=lambda: [fr])))
    )

    assert await scan_missing_videos_shot1(session, project) == []
    assert await scan_missing_shot2_videos(session, project) == [1]


@pytest.mark.asyncio
async def test_trigger_finish_missing_videos_only_shot2(tmp_path: Path) -> None:
    data_dir = tmp_path / "p17"
    scenes = data_dir / "scenes"
    videos = data_dir / "videos"
    videos.mkdir(parents=True)
    _valid_png(scenes / "frame_001_abc12345.png")
    _valid_png(scenes / "frame_001_s2_def67890.png")
    (videos / "clip_001_aaa11111.mp4").write_bytes(b"x" * 1000)
    _write_plan_shot2_video(data_dir / "project.xlsx")

    project = _project(data_dir, status=ProjectStatus.videos_ready)
    fr = Frame(project_id=17, number=1, voiceover_text="v", image_prompt="wide")
    fr.animation_prompt = "Camera pan."
    fr.status = FrameStatus.video_generated
    session = AsyncMock()
    session.execute = AsyncMock(
        return_value=MagicMock(scalars=MagicMock(return_value=MagicMock(all=lambda: [fr])))
    )
    session.flush = AsyncMock()

    info = await trigger_finish_missing_videos(session, project)
    assert info["missing_shot2"] == [1]
    assert info["queued_shot2"] == 1
    assert project.status.value == "generating_videos"
@pytest.mark.asyncio
async def test_finish_missing_clears_error_sleep(tmp_path: Path) -> None:
    from datetime import datetime, timedelta, timezone

    data_dir = tmp_path / "p17"
    scenes = data_dir / "scenes"
    _valid_png(scenes / "frame_001_abc12345.png")
    _write_plan_with_shot2(data_dir / "project.xlsx")

    project = _project(data_dir, status=ProjectStatus.generating_images)
    until = (datetime.now(timezone.utc) + timedelta(minutes=30)).isoformat()
    project.meta = {
        "step_failure": {
            "sleep_until": until,
            "total_fails": {"generating_images": 3},
        }
    }
    fr = Frame(project_id=17, number=1, voiceover_text="v", image_prompt="wide")
    session = AsyncMock()
    session.execute = AsyncMock(
        return_value=MagicMock(scalars=MagicMock(return_value=MagicMock(all=lambda: [fr])))
    )
    session.flush = AsyncMock()

    await trigger_finish_missing_images(session, project)
    assert "sleep_until" not in (project.meta.get("step_failure") or {})
    assert (project.meta.get("step_failure") or {}).get("total_fails", {}).get(
        "generating_images"
    ) is None


@pytest.mark.asyncio
async def test_finish_missing_videos_clears_user_stop(tmp_path: Path) -> None:
    data_dir = tmp_path / "p46"
    scenes = data_dir / "scenes"
    videos = data_dir / "videos"
    videos.mkdir(parents=True)
    _valid_png(scenes / "frame_001_abc12345.png")
    (videos / "clip_001_s2_done.mp4").write_bytes(b"x" * 1000)

    project = _project(data_dir, status=ProjectStatus.animation_prompts_ready)
    project.id = 46
    project.meta = {"user_stop": True, "mass_lane_user_stop": True}
    fr = Frame(
        project_id=46,
        number=1,
        voiceover_text="v",
        image_prompt="wide",
        animation_prompt="slow pan",
        status=FrameStatus.animation_prompt_ready,
    )
    session = AsyncMock()
    session.execute = AsyncMock(
        return_value=MagicMock(scalars=MagicMock(return_value=MagicMock(all=lambda: [fr])))
    )
    session.flush = AsyncMock()
    session.get = AsyncMock(return_value=None)

    info = await trigger_finish_missing_videos(session, project)
    assert info["missing_shot1"] == [1]
    assert info["queued_shot1"] == 1
    assert project.status is ProjectStatus.generating_videos
    assert not (project.meta or {}).get("user_stop")
    assert not (project.meta or {}).get("mass_lane_user_stop")

