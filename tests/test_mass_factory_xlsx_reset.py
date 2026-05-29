"""Mass factory: каждый lane получает чистый project.xlsx из шаблона."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from app.models import Base, Project, ProjectStatus
from app.services.mass_factory import (
    delete_new_mass_children,
    init_child_data_dir,
)
from app.services.xlsx_v8_import import SHEET_PLAN_V8, ROW_VOICEOVER_V8


@pytest.fixture
async def session() -> AsyncSession:
    engine = create_async_engine("sqlite+aiosqlite:///:memory:")
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    factory = async_sessionmaker(engine, expire_on_commit=False)
    async with factory() as s:
        yield s
    await engine.dispose()


def _write_filled_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_PLAN_V8
    ws.cell(row=ROW_VOICEOVER_V8, column=3, value="stale block from old run")
    wb.save(path)


@pytest.mark.asyncio
async def test_init_child_data_dir_resets_stale_xlsx(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    tpl_dir = tmp_path / "templates"
    tpl_dir.mkdir()
    tpl = tpl_dir / "project_template_v8.xlsx"
    wb = Workbook()
    wb.active.title = SHEET_PLAN_V8
    wb.save(tpl)

    root = tmp_path / "data"
    slug_dir = root / "videos" / "topic-a"
    slug_dir.mkdir(parents=True)
    stale = slug_dir / "project.xlsx"
    _write_filled_xlsx(stale)
    (slug_dir / "tmp_gpt").mkdir()
    (slug_dir / "tmp_gpt" / "split_old.xlsx").write_bytes(b"PK")

    monkeypatch.setattr("app.models.settings.data_dir", root)
    monkeypatch.setattr("app.storage.project_sheet._V8_TEMPLATE", tpl)
    monkeypatch.setattr("app.storage.project_sheet.DEFAULT_TEMPLATE_PATH", tpl)

    project = Project(id=9, topic="Topic A", slug="topic-a", status=ProjectStatus.new)
    monkeypatch.setattr(type(project), "data_dir", property(lambda self: slug_dir))

    await init_child_data_dir(project)

    wb2 = load_workbook(stale, data_only=True)
    ws = wb2[SHEET_PLAN_V8]
    assert ws.cell(row=ROW_VOICEOVER_V8, column=3).value in (None, "")
    assert not (slug_dir / "tmp_gpt").exists()


@pytest.mark.asyncio
async def test_delete_new_mass_children_removes_data_dir(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, session: AsyncSession
) -> None:
    root = tmp_path / "data"
    parent = Project(slug="factory", topic="Factory", status=ProjectStatus.new, auto_mode=True)
    parent.meta = {"mass_factory": True}
    session.add(parent)
    await session.flush()

    child = Project(
        slug="child-a",
        topic="Child",
        status=ProjectStatus.new,
        meta={"mass_parent_id": parent.id, "mass_lane_position": 1},
    )
    session.add(child)
    await session.flush()

    slug_dir = root / "videos" / "child-a"
    slug_dir.mkdir(parents=True)
    _write_filled_xlsx(slug_dir / "project.xlsx")
    monkeypatch.setattr("app.models.settings.data_dir", root)
    monkeypatch.setattr(type(child), "data_dir", property(lambda self: slug_dir))

    n = await delete_new_mass_children(session, parent.id)
    assert n == 1
    assert not slug_dir.exists()
