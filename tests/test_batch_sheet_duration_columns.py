"""Тесты колонок «Время ролика» (L) и «Закадровый текст» (M, формула)
в `topics.xlsx` массового проекта.

Проверяем что:
  1. write_subprojects_table пишет L и M, M = формула =L*13.5
  2. read_topics корректно парсит L (число) и M (formula → cached
     value берётся data_only=True если был calc'нут; иначе =L*13.5
     fallback'ом считается в коде).
  3. roundtrip — duration_seconds возвращается корректно.
  4. Excel-формула в M живая (не статичное число).
"""

from __future__ import annotations

import tempfile
from pathlib import Path

from app.storage.batch_sheet import (
    HEADERS,
    SERVICE_COL_INDICES,
    VOICEOVER_CHARS_PER_SECOND,
    init_topics_xlsx,
    read_topics,
    write_subprojects_table,
)


def _tmp_xlsx() -> Path:
    return Path(tempfile.mkdtemp()) / "topics.xlsx"


def test_headers_has_17_columns():
    """Структура: A..K карточные + L,M (новые) + N..Q сервисные."""
    assert len(HEADERS) == 17
    assert HEADERS[11] == "Время ролика (сек)"  # L (0-indexed: 11)
    assert HEADERS[12].startswith("Закадр.")  # M
    assert SERVICE_COL_INDICES == (14, 15, 16, 17)


def test_voiceover_multiplier_is_13_5():
    assert VOICEOVER_CHARS_PER_SECOND == 13.5


def test_write_and_read_duration_roundtrip():
    """Записали duration=30 → прочитали duration=30 и chars=405 (=30*13.5)."""
    path = _tmp_xlsx()
    init_topics_xlsx(path, "тест-батч")
    rows = [
        {"position": 1, "title": "Тема 1", "video_duration_sec": 30,
         "style": "Попаданец", "slug": "sub1", "status": "new"},
        {"position": 2, "title": "Тема 2", "video_duration_sec": 45.5,
         "slug": "sub2"},
        {"position": 3, "title": "Тема 3 без длительности"},
    ]
    write_subprojects_table(path, rows, "тест-батч")

    out = read_topics(path)
    assert len(out) == 3

    r1, r2, r3 = out
    assert r1["title"] == "Тема 1"
    assert r1["video_duration_sec"] == 30.0
    # M — формула, при чтении data_only=True openpyxl ВЕРНЁТ
    # last-cached значение из Excel (нет, если файл никогда не открывался).
    # read_topics fallback'ом сам считает chars из duration*13.5.
    assert r1["voiceover_chars_target"] == 405.0

    assert r2["video_duration_sec"] == 45.5
    assert r2["voiceover_chars_target"] == round(45.5 * 13.5, 1)

    assert r3["video_duration_sec"] is None
    assert r3["voiceover_chars_target"] is None


def test_formula_written_to_column_m():
    """M-ячейка должна быть формулой =L<row>*13.5, не статичным числом."""
    from openpyxl import load_workbook

    path = _tmp_xlsx()
    init_topics_xlsx(path, "тест-батч")
    rows = [
        {"position": 1, "title": "Тема 1", "video_duration_sec": 30},
        {"position": 2, "title": "Тема 2", "video_duration_sec": 60},
    ]
    write_subprojects_table(path, rows, "тест-батч")

    # data_only=False — получим формулы как строки.
    wb = load_workbook(path, data_only=False)
    ws = wb["Темы"]
    # Строка 3 — первая тема (1 — заголовок батча, 2 — headers).
    assert ws.cell(row=3, column=13).value == "=L3*13.5"
    assert ws.cell(row=4, column=13).value == "=L4*13.5"
    wb.close()


def test_read_topics_backward_compat_old_15_col_file():
    """Старый файл с 15 колонками (без L, M duration) должен читаться:
    L и M будут None.
    """
    from openpyxl import Workbook

    path = _tmp_xlsx()
    wb = Workbook()
    ws = wb.active
    ws.title = "Темы"
    # 15 колонок — старый формат: A..O без duration/chars.
    ws.cell(row=2, column=2, value="Название ролика")
    ws.cell(row=3, column=1, value=1)
    ws.cell(row=3, column=2, value="Старая тема")
    ws.cell(row=3, column=4, value="Попаданец")  # style
    ws.cell(row=3, column=12, value="old-slug")  # старая позиция slug (L)
    wb.save(path)

    out = read_topics(path)
    assert len(out) == 1
    r = out[0]
    assert r["title"] == "Старая тема"
    assert r["style"] == "Попаданец"
    assert r["video_duration_sec"] is None
    assert r["voiceover_chars_target"] is None


def test_topic_card_extracts_new_fields():
    """topic_card_from_row должна включать duration/chars если они есть."""
    from app.storage.batch_sheet import topic_card_from_row

    row = {
        "title": "T",
        "style": "S",
        "video_duration_sec": 30.0,
        "voiceover_chars_target": 405.0,
    }
    card = topic_card_from_row(row)
    assert card["title"] == "T"
    assert card["style"] == "S"
    assert card["video_duration_sec"] == 30.0
    assert card["voiceover_chars_target"] == 405.0

    # Пустые поля не попадают.
    row_empty = {"title": "T", "style": "S"}
    card2 = topic_card_from_row(row_empty)
    assert "video_duration_sec" not in card2
    assert "voiceover_chars_target" not in card2
