"""Проверка листов xlsx против шаблона."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from app.services.xlsx_v8_import import SHEET_GENERAL_V8, SHEET_PLAN_V8
from app.services.xlsx_versioning import (
    normalize_xlsx_to_reference_layout,
    validate_xlsx,
    validate_xlsx_sheets,
)
from app.storage.project_sheet import SHEET_FRAMES, SHEET_GENERAL


def _save_v8(path: Path) -> None:
    wb = Workbook()
    wb.active.title = SHEET_PLAN_V8
    wb.create_sheet(SHEET_GENERAL_V8)
    wb.save(path)


def test_validate_xlsx_sheets_accepts_v8_layout(tmp_path: Path) -> None:
    p = tmp_path / "ok.xlsx"
    _save_v8(p)
    assert validate_xlsx_sheets(p) is None
    assert validate_xlsx(p) is None


def test_validate_xlsx_sheets_rejects_wrong_layout(tmp_path: Path) -> None:
    p = tmp_path / "bad.xlsx"
    wb = Workbook()
    wb.active.title = "Лист1"
    wb.create_sheet("Лист2")
    wb.save(p)
    err = validate_xlsx_sheets(p)
    assert err is not None
    assert "ошибка формата эксель таблицы" in err


def test_validate_xlsx_sheets_accepts_v7_layout(tmp_path: Path) -> None:
    p = tmp_path / "v7.xlsx"
    wb = Workbook()
    wb.active.title = SHEET_FRAMES
    wb.create_sheet(SHEET_GENERAL)
    wb.save(p)
    assert validate_xlsx_sheets(p) is None


def test_validate_xlsx_sheets_rejects_v8_with_extra_gpt_sheets(
    tmp_path: Path,
) -> None:
    p = tmp_path / "gpt_plan.xlsx"
    wb = Workbook()
    wb.active.title = SHEET_PLAN_V8
    wb.create_sheet(SHEET_GENERAL_V8)
    wb.create_sheet("Персонажи")
    wb.create_sheet("Предметы")
    wb.create_sheet("Фоны")
    wb.create_sheet("Исследовательская база")
    wb.save(p)
    assert validate_xlsx_sheets(p) is not None


def test_normalize_xlsx_strips_extra_gpt_sheets(tmp_path: Path) -> None:
    reference = tmp_path / "reference.xlsx"
    wb_ref = Workbook()
    wb_ref.active.title = SHEET_PLAN_V8
    wb_ref.active["A1"] = "ref-plan"
    wb_ref.create_sheet(SHEET_GENERAL_V8)
    wb_ref[SHEET_GENERAL_V8]["A1"] = "ref-general"
    wb_ref.create_sheet("Персонажи")
    wb_ref.create_sheet("Предметы")
    wb_ref.create_sheet("Фоны")
    wb_ref.save(reference)

    gpt = tmp_path / "gpt.xlsx"
    wb_gpt = Workbook()
    wb_gpt.active.title = SHEET_PLAN_V8
    wb_gpt.active["A1"] = "gpt-plan"
    wb_gpt.create_sheet(SHEET_GENERAL_V8)
    wb_gpt[SHEET_GENERAL_V8]["A1"] = "gpt-general"
    wb_gpt.create_sheet("Персонажи")
    wb_gpt.create_sheet("Предметы")
    wb_gpt.create_sheet("Фоны")
    wb_gpt.create_sheet("Исследовательская база")
    wb_gpt.save(gpt)

    assert validate_xlsx_sheets(gpt) is not None
    assert normalize_xlsx_to_reference_layout(gpt, reference) is True
    assert validate_xlsx_sheets(gpt) is None
    assert validate_xlsx(gpt) is None

    from openpyxl import load_workbook

    wb = load_workbook(gpt, read_only=True)
    assert wb.sheetnames == [
        SHEET_PLAN_V8,
        SHEET_GENERAL_V8,
        "Персонажи",
        "Предметы",
        "Фоны",
    ]
    wb.close()
