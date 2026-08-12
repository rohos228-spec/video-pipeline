"""GPT API-клиент: успех, ретраи, фатальные ошибки, download, xlsx→text."""

from __future__ import annotations

import json
from pathlib import Path

import httpx
import pytest

from app.services import gpt_api
from app.services.gpt_api import (
    GptApiError,
    build_messages,
    chat,
    collect_result_urls,
    download_content,
    parse_responses_sse_lines,
    xlsx_to_text,
)


def _enable(monkeypatch) -> None:
    from app.settings import settings

    # Явно kie: иначе .env с TEXT_LLM_PROVIDER=tokenrouter перехватит путь/модель.
    monkeypatch.setattr(settings, "text_llm_provider", "kie")
    monkeypatch.setattr(settings, "tokenrouter_api_key", "")
    monkeypatch.setattr(settings, "gpt_api_key", "test-key")
    monkeypatch.setattr(settings, "gpt_base_url", "https://gw.test")
    monkeypatch.setattr(settings, "gpt_chat_path", "/v1/chat/completions")
    monkeypatch.setattr(settings, "gpt_api_mode", "chat")
    monkeypatch.setattr(settings, "gpt_max_retries", 3)


def _mock_httpx(monkeypatch, handler) -> None:
    real_client = httpx.AsyncClient

    def factory(*args, **kwargs):
        kwargs["transport"] = httpx.MockTransport(handler)
        return real_client(*args, **kwargs)

    monkeypatch.setattr(gpt_api.httpx, "AsyncClient", factory)

    async def _no_sleep(*_a, **_k):
        return None

    monkeypatch.setattr(gpt_api.asyncio, "sleep", _no_sleep)


def _completion(text: str) -> dict:
    return {
        "choices": [{"message": {"content": text}, "finish_reason": "stop"}],
        "usage": {"total_tokens": 10},
    }


@pytest.mark.asyncio
async def test_chat_success(monkeypatch) -> None:
    _enable(monkeypatch)
    calls: list[httpx.Request] = []

    def handler(request: httpx.Request) -> httpx.Response:
        calls.append(request)
        assert request.url.path == "/v1/chat/completions"
        assert request.headers["Authorization"] == "Bearer test-key"
        return httpx.Response(200, json=_completion("привет из API"))

    _mock_httpx(monkeypatch, handler)
    res = await chat(prompt="скажи привет")
    assert res.text == "привет из API"
    assert res.finish_reason == "stop"
    assert len(calls) == 1


@pytest.mark.asyncio
async def test_chat_templated_path_per_model(monkeypatch) -> None:
    """kie.ai: путь зависит от модели — /{model}/v1/chat/completions."""
    _enable(monkeypatch)
    from app.settings import settings

    monkeypatch.setattr(settings, "gpt_base_url", "https://api.kie.ai")
    monkeypatch.setattr(settings, "gpt_chat_path", "/{model}/v1/chat/completions")
    seen: list[str] = []

    def handler(request: httpx.Request) -> httpx.Response:
        seen.append(request.url.path)
        return httpx.Response(200, json=_completion("ok"))

    _mock_httpx(monkeypatch, handler)
    await chat(prompt="x", model="gemini-2.5-pro")
    assert seen == ["/gemini-2.5-pro/v1/chat/completions"]


@pytest.mark.asyncio
async def test_chat_retries_on_429_then_success(monkeypatch) -> None:
    _enable(monkeypatch)
    state = {"n": 0}

    def handler(request: httpx.Request) -> httpx.Response:
        state["n"] += 1
        if state["n"] < 3:
            return httpx.Response(429, text="rate limited")
        return httpx.Response(200, json=_completion("ok после ретраев"))

    _mock_httpx(monkeypatch, handler)
    res = await chat(prompt="x")
    assert res.text == "ok после ретраев"
    assert state["n"] == 3


@pytest.mark.asyncio
async def test_chat_fatal_401_not_retried(monkeypatch) -> None:
    _enable(monkeypatch)
    state = {"n": 0}

    def handler(request: httpx.Request) -> httpx.Response:
        state["n"] += 1
        return httpx.Response(401, text="unauthorized")

    _mock_httpx(monkeypatch, handler)
    with pytest.raises(GptApiError) as ei:
        await chat(prompt="x")
    assert ei.value.context.get("status_code") == 401
    assert ei.value.retryable is False
    assert state["n"] == 1  # без ретраев


@pytest.mark.asyncio
async def test_chat_exhausts_retries_on_500(monkeypatch) -> None:
    _enable(monkeypatch)
    state = {"n": 0}

    def handler(request: httpx.Request) -> httpx.Response:
        state["n"] += 1
        return httpx.Response(503, text="down")

    _mock_httpx(monkeypatch, handler)
    with pytest.raises(GptApiError):
        await chat(prompt="x", max_retries=2)
    assert state["n"] == 3  # 1 + 2 ретрая


@pytest.mark.asyncio
async def test_chat_responses_mode(monkeypatch) -> None:
    """kie.ai Responses API: input/output вместо messages/choices."""
    _enable(monkeypatch)
    from app.settings import settings

    monkeypatch.setattr(settings, "gpt_base_url", "https://api.kie.ai")
    monkeypatch.setattr(settings, "gpt_chat_path", "/codex/v1/responses")
    monkeypatch.setattr(settings, "gpt_api_mode", "auto")
    captured: dict = {}

    def handler(request: httpx.Request) -> httpx.Response:
        captured["path"] = request.url.path
        captured["body"] = json.loads(request.content)
        # Responses mode must request SSE — иначе CF рвёт длинный JSON.
        assert captured["body"].get("stream") is True
        sse = (
            'event: response.created\n'
            'data: {"type":"response.created","response":{"id":"resp_test123","status":"in_progress"}}\n'
            "\n"
            'event: response.output_text.delta\n'
            'data: {"type":"response.output_text.delta","delta":"работает"}\n'
            "\n"
            'event: response.completed\n'
            'data: {"type":"response.completed","response":{"id":"resp_test123","status":"completed",'
            '"output":[{"type":"message","role":"assistant","content":'
            '[{"type":"output_text","text":"работает"}]}],'
            '"usage":{"total_tokens":9}}}\n'
            "\n"
            "data: [DONE]\n"
        )
        return httpx.Response(
            200,
            content=sse.encode("utf-8"),
            headers={"content-type": "text/event-stream"},
        )

    _mock_httpx(monkeypatch, handler)
    res = await chat(prompt="скажи", model="gpt-5-6-sol")
    assert res.text == "работает"
    assert res.finish_reason == "completed"
    assert res.response_id == "resp_test123"
    assert captured["path"] == "/codex/v1/responses"
    assert "input" in captured["body"]  # responses-формат
    assert "messages" not in captured["body"]


def test_parse_responses_sse_lines_prefers_completed_payload() -> None:
    lines = [
        'data: {"type":"response.created","response":{"id":"resp_abc","status":"in_progress"}}',
        'data: {"type":"response.output_text.delta","delta":"hel"}',
        'data: {"type":"response.output_text.delta","delta":"lo"}',
        (
            'data: {"type":"response.completed","response":{"id":"resp_abc","status":"completed",'
            '"output":[{"type":"message","content":[{"type":"output_text","text":"hello world"}]}]}}'
        ),
    ]
    text, status, payload, rid = parse_responses_sse_lines(lines)
    assert text == "hello world"
    assert status == "completed"
    assert rid == "resp_abc"
    assert payload.get("id") == "resp_abc"


def test_parse_responses_sse_lines_salvages_deltas_without_completed() -> None:
    """CF рвёт на огромном completed — дельты уже есть, нельзя вернуть пусто."""
    lines = [
        'data: {"type":"response.created","response":{"id":"resp_x","status":"in_progress"}}',
        'data: {"type":"response.output_text.delta","delta":"часть "}',
        'data: {"type":"response.output_text.delta","delta":"ответа"}',
        'data: {"type":"response.output_text.done","text":"часть ответа"}',
        # битая/обрезанная строка completed — как при обрыве
        'data: {"type":"response.completed","response":{"id":"resp_x","status":"comp',
    ]
    text, status, _payload, rid = parse_responses_sse_lines(lines)
    assert text == "часть ответа"
    assert rid == "resp_x"
    assert status in {"completed", "stream_partial"}


def test_parse_responses_sse_lines_prefers_longer_deltas_over_stub_completed() -> None:
    lines = [
        'data: {"type":"response.created","response":{"id":"resp_y","status":"in_progress"}}',
        'data: {"type":"response.output_text.delta","delta":"полный длинный текст файла"}',
        (
            'data: {"type":"response.completed","response":{"id":"resp_y","status":"completed",'
            '"output":[{"type":"message","content":[{"type":"output_text","text":"ок"}]}]}}'
        ),
    ]
    text, status, _payload, rid = parse_responses_sse_lines(lines)
    assert text == "полный длинный текст файла"
    assert rid == "resp_y"
    assert status == "completed"


@pytest.mark.asyncio
async def test_chat_responses_salvages_after_stream_disconnect(monkeypatch) -> None:
    """RemoteProtocolError после дельт → текст salvage, не новая пустая ошибка."""
    _enable(monkeypatch)
    from app.settings import settings

    monkeypatch.setattr(settings, "gpt_chat_path", "/codex/v1/responses")
    monkeypatch.setattr(settings, "gpt_api_mode", "auto")

    class _BoomStream:
        def __init__(self) -> None:
            self.headers = {"cf-ray": "test-ray"}
            self.status_code = 200

        async def __aenter__(self):
            return self

        async def __aexit__(self, *args):
            return False

        async def aiter_lines(self):
            yield 'data: {"type":"response.created","response":{"id":"resp_salv","status":"in_progress"}}'
            yield 'data: {"type":"response.output_text.delta","delta":"живой "}'
            yield 'data: {"type":"response.output_text.delta","delta":"текст"}'
            raise httpx.RemoteProtocolError("Server disconnected without sending a response.")

    class _BoomClient:
        def __init__(self, *a, **k):
            pass

        async def __aenter__(self):
            return self

        async def __aexit__(self, *args):
            return False

        def stream(self, *a, **k):
            return _BoomStream()

    monkeypatch.setattr(gpt_api.httpx, "AsyncClient", _BoomClient)

    async def _no_sleep(*_a, **_k):
        return None

    monkeypatch.setattr(gpt_api.asyncio, "sleep", _no_sleep)

    res = await chat(prompt="длинный", model="gpt-5-6-sol", max_retries=0)
    assert res.text == "живой текст"
    assert res.response_id == "resp_salv"
    assert res.finish_reason == "stream_salvaged"
    assert res.raw.get("sse_salvaged") is True


@pytest.mark.asyncio
async def test_chat_provider_envelope_error(monkeypatch) -> None:
    """kie.ai отдаёт ошибку как HTTP 200 {code,msg} — ловим её понятно."""
    _enable(monkeypatch)
    state = {"n": 0}

    def handler(request: httpx.Request) -> httpx.Response:
        state["n"] += 1
        return httpx.Response(
            200,
            json={"code": 401, "msg": "The API key is not authorized to use this model.", "data": None},
        )

    _mock_httpx(monkeypatch, handler)
    with pytest.raises(GptApiError) as ei:
        await chat(prompt="x", max_retries=3)
    assert ei.value.context.get("provider_code") == 401
    assert state["n"] == 1  # 401 не ретраится


@pytest.mark.asyncio
async def test_chat_empty_choices_raises(monkeypatch) -> None:
    _enable(monkeypatch)

    def handler(request: httpx.Request) -> httpx.Response:
        return httpx.Response(200, json={"choices": []})

    _mock_httpx(monkeypatch, handler)
    with pytest.raises(GptApiError):
        await chat(prompt="x", max_retries=0)


@pytest.mark.asyncio
async def test_download_content(monkeypatch, tmp_path: Path) -> None:
    def handler(request: httpx.Request) -> httpx.Response:
        return httpx.Response(200, content=b"BINARYDATA")

    _mock_httpx(monkeypatch, handler)
    out = tmp_path / "sub" / "img.png"
    got = await download_content("https://cdn.test/img.png", out)
    assert got.read_bytes() == b"BINARYDATA"


@pytest.mark.asyncio
async def test_download_content_404_raises(monkeypatch, tmp_path: Path) -> None:
    def handler(request: httpx.Request) -> httpx.Response:
        return httpx.Response(404, text="nope")

    _mock_httpx(monkeypatch, handler)
    with pytest.raises(GptApiError):
        await download_content("https://cdn.test/x", tmp_path / "x")


def test_collect_result_urls() -> None:
    urls = collect_result_urls("готово: https://a.io/x.png, ещё http://b/y.mp4).")
    assert urls == ["https://a.io/x.png", "http://b/y.mp4"]


def test_xlsx_to_text(tmp_path: Path) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "План"
    ws.append(["кадр", "текст"])
    ws.append([1, "хук"])
    p = tmp_path / "project.xlsx"
    wb.save(p)
    text = xlsx_to_text(p)
    assert "Лист: План" in text
    assert "хук" in text
    assert "xlsx text-export" in text
    assert "@row=1" in text
    assert "@row=2" in text
    assert "# Лист:" in text or "для записи верни" in text


def test_xlsx_to_text_apply_ops_contract_forbids_tsv_refusal(tmp_path: Path) -> None:
    """project_file/DB SoT: баннер не толкает модель в отказ «нет xlsx» / TSV."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "план"
    ws["C49"] = "закадр"
    p = tmp_path / "project.xlsx"
    wb.save(p)
    text = xlsx_to_text(p, write_contract="apply_ops")
    assert "DB SoT" in text
    assert "apply-ops" in text.lower() or "JSON" in text
    assert "[SHEET: план]" in text
    assert "# Лист:" not in text
    assert "верни `# Лист:`" not in text
    # Не провоцировать parrot-отказ модели.
    assert "недоступен" not in text.casefold()


def test_xlsx_to_text_sparse_rows_keep_excel_numbers(tmp_path: Path) -> None:
    """Пустые ряды между R1 и R49 не должны сдвигать @row= в экспорте."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "план"
    ws["A1"] = "title"
    ws["C49"] = "voice cell"
    p = tmp_path / "project.xlsx"
    wb.save(p)
    text = xlsx_to_text(p)
    assert "@row=1" in text
    assert "@row=49" in text
    assert "voice cell" in text


def test_xlsx_to_text_keeps_wide_plan_columns(tmp_path: Path) -> None:
    """Кадры на «план» — столбцы; 150-й столбец должен попасть в TSV для API."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "план"
    ws.cell(row=48, column=1, value="animation_prompt")
    ws.cell(row=48, column=2, value="c01")
    ws.cell(row=48, column=150, value="frame-150-prompt")
    p = tmp_path / "wide.xlsx"
    wb.save(p)
    text = xlsx_to_text(p)
    assert "frame-150-prompt" in text
    assert "@row=48" in text


def test_xlsx_to_text_prioritizes_general_plan(tmp_path: Path) -> None:
    """Огромный лист «план» не должен вытеснять «Общий план» из контекста."""
    from openpyxl import Workbook

    from app.services.gpt_api import file_to_context

    wb = Workbook()
    plan = wb.active
    plan.title = "план"
    for i in range(300):
        plan.append([f"row{i}", "x" * 400])
    general = wb.create_sheet("Общий план")
    general.append(["Главная тема", "тема про Рим"])
    general.append(["Длительность", "12:00"])
    for i in range(1, 120):
        general.append([f"эпизод {i:03d}", f"таймкод 0:{i:02d}", "текст " * 20])
    p = tmp_path / "project.xlsx"
    wb.save(p)

    text = xlsx_to_text(p, max_chars=80_000)
    assert text.index("# Лист: Общий план") < text.index("# Лист: план")
    assert "эпизод 097" in text
    assert "эпизод 119" in text
    assert "тема про Рим" in text

    ctx = file_to_context(p)
    assert "эпизод 097" in ctx
    assert len(ctx) > 60_000


def test_build_messages_with_history() -> None:
    from app.services.gpt_api import build_messages, normalize_history

    hist = normalize_history(
        [
            {"role": "user", "content": "меня зовут Вася"},
            {"role": "assistant", "content": "Привет, Вася"},
            {"role": "system", "content": "игнор"},
        ]
    )
    assert len(hist) == 2
    msgs = build_messages(prompt="как меня зовут?", history=hist)
    assert [m["role"] for m in msgs] == ["user", "assistant", "user"]
    assert "Вася" in msgs[0]["content"]
    assert "как меня зовут" in msgs[-1]["content"]


def test_build_input_with_history() -> None:
    from app.services.gpt_api import build_input

    inp = build_input(
        prompt="второе сообщение",
        history=[
            {"role": "user", "content": "первое"},
            {"role": "assistant", "content": "ок"},
        ],
    )
    assert isinstance(inp, list)
    assert len(inp) == 3
    assert inp[0] == {"role": "user", "content": "первое"}
    assert inp[1]["role"] == "assistant"
    assert inp[2]["role"] == "user"
    assert "второе" in inp[2]["content"]


def test_build_input_plain_prompt_is_message_list() -> None:
    """kie.ai codex: голая строка input → code=500; нужен [{role,content}]."""
    from app.services.gpt_api import build_input

    inp = build_input(prompt="просто текст")
    assert isinstance(inp, list)
    assert len(inp) == 1
    assert inp[0]["role"] == "user"
    assert inp[0]["content"] == "просто текст"


def test_normalize_history_truncates_old() -> None:
    from app.services.gpt_api import normalize_history

    big = [{"role": "user", "content": f"m{i}"} for i in range(50)]
    out = normalize_history(big, max_messages=5)
    assert len(out) == 5
    assert out[0]["content"] == "m45"


def test_build_messages_with_file(tmp_path: Path) -> None:
    f = tmp_path / "voiceover.txt"
    f.write_text("закадровый текст", encoding="utf-8")
    msgs = build_messages(prompt="проверь", accompanying="важно", input_paths=[f], system="ты агент")
    assert msgs[0]["role"] == "system"
    user = msgs[-1]["content"]
    assert "проверь" in user
    assert "важно" in user
    assert "закадровый текст" in user


def _tiny_png(path: Path) -> Path:
    path.write_bytes(
        bytes.fromhex(
            "89504e470d0a1a0a0000000d49484452000000010000000108060000001f15c489"
            "0000000a49444154789c63000100000500010d0a2db40000000049454e44ae426082"
        )
    )
    return path


def test_build_messages_vision_parts(tmp_path: Path) -> None:
    img = _tiny_png(tmp_path / "hero.png")
    msgs = build_messages(prompt="проверь кадр", input_paths=[img])
    content = msgs[-1]["content"]
    assert isinstance(content, list)
    assert content[0]["type"] == "text"
    assert content[1]["type"] == "image_url"
    assert content[1]["image_url"]["url"].startswith("data:image/png;base64,")


def test_build_input_vision_responses(tmp_path: Path) -> None:
    from app.services.gpt_api import build_input

    img = _tiny_png(tmp_path / "scene.jpg")
    # .jpg with png bytes — mime by suffix; ok for structure test
    img.write_bytes(_tiny_png(tmp_path / "x.png").read_bytes())
    inp = build_input(prompt="aspect?", input_paths=[img])
    assert isinstance(inp, list)
    parts = inp[0]["content"]
    assert parts[0]["type"] == "input_text"
    assert parts[1]["type"] == "input_image"
    assert parts[1]["image_url"].startswith("data:image/jpeg;base64,")


@pytest.mark.asyncio
async def test_run_operator_api_real_check(monkeypatch, tmp_path: Path) -> None:
    """Интеграция: включённый API + review-роль → analysis.json + verdict."""
    _enable(monkeypatch)
    verdict_json = json.dumps(
        {
            "schema": "vp.check.v1",
            "verdict": "pass",
            "summary": "ок",
            "checks": [{"id": "x", "ok": True}],
            "forward": {"mode": "inherit", "paths": []},
            "fix": {"target": "none"},
        },
        ensure_ascii=False,
    )

    def handler(request: httpx.Request) -> httpx.Response:
        return httpx.Response(200, json=_completion(verdict_json))

    _mock_httpx(monkeypatch, handler)
    from app.services.gpt_operator_client import run_operator_api

    proj = tmp_path / "proj"
    proj.mkdir()
    res = await run_operator_api(
        project_dir=proj,
        node_key="n_check",
        role="review",
        output_mode="text",
        prompt="проверь план",
        accompanying="",
        input_paths=[],
    )
    assert res.gate_status == "pass"
    assert res.analysis is not None and res.analysis.verdict == "pass"
    assert (proj / "excel_gpt_uploads" / "n_check" / "analysis.json").is_file()


_MINIMAL_PDF = b"""%PDF-1.4
1 0 obj<< /Type /Catalog /Pages 2 0 R >>endobj
2 0 obj<< /Type /Pages /Kids [3 0 R] /Count 1 >>endobj
3 0 obj<< /Type /Page /Parent 2 0 R /MediaBox [0 0 300 144] /Contents 4 0 R /Resources<< /Font<< /F1 5 0 R >> >> >>endobj
4 0 obj<< /Length 55 >>stream
BT /F1 18 Tf 40 100 Td (Hello Filmmaking PDF) Tj ET
endstream
endobj
5 0 obj<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>endobj
xref
0 6
0000000000 65535 f 
0000000009 00000 n 
0000000058 00000 n 
0000000115 00000 n 
0000000266 00000 n 
0000000371 00000 n 
trailer<< /Size 6 /Root 1 0 R >>
startxref
444
%%EOF
"""


def test_pdf_to_text_and_no_base64_in_context(tmp_path: Path) -> None:
    from app.services.gpt_api import file_to_context, pdf_to_text

    p = tmp_path / "guide.pdf"
    p.write_bytes(_MINIMAL_PDF)
    text = pdf_to_text(p)
    assert "Filmmaking" in text
    ctx = file_to_context(p)
    assert "Filmmaking" in ctx
    assert "data:application/pdf;base64," not in ctx
    # старый баг: сырой бинарь PDF в тексте
    assert "%PDF" not in ctx


def test_build_input_attaches_pdf_as_input_file(tmp_path: Path, monkeypatch) -> None:
    """По умолчанию PDF = только текст; input_file — при GPT_PDF_INPUT_FILE=1."""
    import os

    from app.services.gpt_api import build_input

    p = tmp_path / "deck.pdf"
    p.write_bytes(_MINIMAL_PDF)

    inp_default = build_input(prompt="кратко о чём pdf", input_paths=[p])
    last = inp_default[-1]
    # text-only: content — строка, не multimodal list с input_file
    assert isinstance(last["content"], str)
    assert "Filmmaking" in last["content"] or "pdf" in last["content"].lower()
    assert "data:application/pdf;base64," not in last["content"]

    monkeypatch.setenv("GPT_PDF_INPUT_FILE", "1")
    try:
        inp = build_input(prompt="кратко о чём pdf", input_paths=[p])
    finally:
        monkeypatch.delenv("GPT_PDF_INPUT_FILE", raising=False)
    last = inp[-1]
    assert last["role"] == "user"
    content = last["content"]
    assert isinstance(content, list)
    types = [c.get("type") for c in content if isinstance(c, dict)]
    assert "input_text" in types
    assert "input_file" in types
    file_part = next(c for c in content if c.get("type") == "input_file")
    assert str(file_part.get("file_data") or "").startswith("data:application/pdf;base64,")
    text_part = next(c for c in content if c.get("type") == "input_text")
    assert "data:application/pdf;base64," not in str(text_part.get("text") or "")


def test_split_pdf_text_chunks_by_pages() -> None:
    from app.services.gpt_api import split_pdf_text_chunks

    pages = []
    for i in range(1, 6):
        pages.append(f"--- стр. {i}/5 ---\n" + ("word " * 400))
    text = "\n\n".join(pages)
    chunks = split_pdf_text_chunks(text, max_chars=2_500)
    assert len(chunks) >= 2
    assert all(len(c) <= 2_500 + 50 for c in chunks)  # page boundary soft
    assert "стр. 1/5" in chunks[0]
    assert sum("word" in c for c in chunks) == len(chunks)


def test_pdf_paths_need_chunking(tmp_path: Path) -> None:
    from app.services.gpt_api import pdf_paths_need_chunking

    p = tmp_path / "small.pdf"
    p.write_bytes(_MINIMAL_PDF)
    assert pdf_paths_need_chunking([p], threshold=10_000) is False
    assert pdf_paths_need_chunking([p], threshold=10) is True


def test_chat_pdf_in_chunks_calls_chat_per_piece(tmp_path: Path, monkeypatch) -> None:
    import asyncio

    from app.services import gpt_api

    calls: list[str] = []

    async def fake_chat(**kwargs):
        calls.append(kwargs.get("prompt") or "")
        return gpt_api.GptChatResult(text=f"OK{len(calls)}", model="test", finish_reason="stop")

    async def no_sleep(*_a, **_k):
        return None

    monkeypatch.setattr(gpt_api, "chat", fake_chat)
    monkeypatch.setattr(gpt_api.asyncio, "sleep", no_sleep)

    big = "\n\n".join(
        f"--- стр. {i}/8 ---\n" + ("alpha " * 300) for i in range(1, 9)
    )
    pdf = tmp_path / "guide.pdf"
    pdf.write_bytes(b"%PDF-1.4 fake")

    monkeypatch.setattr(gpt_api, "pdf_to_text", lambda path, max_chars=80_000: big)

    result = asyncio.run(
        gpt_api.chat_pdf_in_chunks(
            prompt="переведи",
            pdf_paths=[pdf],
            chunk_chars=2_000,
            max_retries=0,
        )
    )
    assert len(calls) >= 2
    assert result.finish_reason == "chunked"
    assert "OK1" in result.text
    assert "###" in result.text


def test_chat_pdf_in_chunks_continues_after_mid_500(tmp_path: Path, monkeypatch) -> None:
    """kie 500 на 2-м куске не должен убивать весь перевод — остальные куски идут."""
    import asyncio

    from app.services import gpt_api

    n = {"i": 0}

    async def fake_chat(**kwargs):
        n["i"] += 1
        if n["i"] == 2:
            raise gpt_api.GptApiError(
                "GPT провайдер code=500: Server exception, please try again later",
                context={"provider_code": 500, "retryable": True},
            )
        return gpt_api.GptChatResult(
            text=f"OK{n['i']}", model="test", finish_reason="stop"
        )

    async def no_sleep(*_a, **_k):
        return None

    monkeypatch.setattr(gpt_api, "chat", fake_chat)
    monkeypatch.setattr(gpt_api.asyncio, "sleep", no_sleep)

    big = "\n\n".join(
        f"--- стр. {i}/6 ---\n" + ("beta " * 200) for i in range(1, 7)
    )
    pdf = tmp_path / "deck.pdf"
    pdf.write_bytes(b"%PDF-1.4 fake")
    monkeypatch.setattr(gpt_api, "pdf_to_text", lambda path, max_chars=80_000: big)

    result = asyncio.run(
        gpt_api.chat_pdf_in_chunks(
            prompt="переведи",
            pdf_paths=[pdf],
            chunk_chars=1_500,
            max_retries=0,
        )
    )
    assert "OK1" in result.text
    assert "фрагмент не обработан" in result.text or "·a" in result.text
    assert result.finish_reason in {"chunked", "chunked_partial"}


def test_is_pdf_provider_failure() -> None:
    from app.services.gpt_api import GptApiError, is_pdf_provider_failure

    assert is_pdf_provider_failure(
        GptApiError("GPT провайдер code=500: Server exception", context={"provider_code": 500})
    )
    assert is_pdf_provider_failure(
        GptApiError("GPT timeout 90s", context={"error_kind": "timeout", "retryable": True})
    )
    assert not is_pdf_provider_failure(GptApiError("bad key", context={"status_code": 401}))

@pytest.mark.asyncio
async def test_download_content_html_renamed_off_xlsx(monkeypatch, tmp_path: Path) -> None:
    """Страница HTML, сохранённая как .xlsx, переименовывается в .html."""
    def handler(request: httpx.Request) -> httpx.Response:
        return httpx.Response(
            200,
            content=b"<!DOCTYPE html><html>err</html>",
            headers={"content-type": "text/html"},
        )

    _mock_httpx(monkeypatch, handler)
    out = tmp_path / "got.xlsx"
    got = await download_content("https://cdn.test/file", out)
    assert got.suffix == ".html"
    assert got.read_bytes().startswith(b"<!DOCTYPE")
