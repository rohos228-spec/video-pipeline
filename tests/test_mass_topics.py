"""Тесты парсера тем из Excel."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from app.storage.mass_topics import parse_topics_xlsx


def test_parse_flexible_first_column(tmp_path: Path) -> None:
    path = tmp_path / "topics.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Тема A"])
    ws.append(["Тема B"])
    ws.append([""])
    ws.append(["Тема C"])
    wb.save(path)
    assert parse_topics_xlsx(path) == ["Тема A", "Тема B", "Тема C"]


def test_parse_header_named_topic_column(tmp_path: Path) -> None:
    path = tmp_path / "topics.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["id", "Название", "note"])
    ws.append([1, "Видео один", "x"])
    ws.append([2, "Видео два", "y"])
    wb.save(path)
    assert parse_topics_xlsx(path) == ["Видео один", "Видео два"]


def test_parse_three_rows_with_notes_column(tmp_path: Path) -> None:
    """Три темы в A + заметки в B — первая строка не должна теряться."""
    path = tmp_path / "topics.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Сталин — биография", "note1"])
    ws.append(["Ленин — факты", "note2"])
    ws.append(["Хрущёв — мифы", "note3"])
    wb.save(path)
    assert parse_topics_xlsx(path) == [
        "Сталин — биография",
        "Ленин — факты",
        "Хрущёв — мифы",
    ]


def test_parse_topics_in_column_b_with_row_numbers(tmp_path: Path) -> None:
    path = tmp_path / "topics.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append([1, "Тема A"])
    ws.append([2, "Тема B"])
    ws.append([3, "Тема C"])
    wb.save(path)
    assert parse_topics_xlsx(path) == ["Тема A", "Тема B", "Тема C"]


def test_parse_topics_only_column_b(tmp_path: Path) -> None:
    path = tmp_path / "topics.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["", "Тема A"])
    ws.append(["", "Тема B"])
    ws.append(["", "Тема C"])
    wb.save(path)
    assert parse_topics_xlsx(path) == ["Тема A", "Тема B", "Тема C"]
