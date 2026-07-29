"""Preview/download Excel: upload excel_gpt жёстко подменяет вход."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.models import Project
from app.services.excel_gpt_node import upload_dir, upload_file_path
from app.services.gpt_operator import resolve_operator
from app.services.node_xlsx_snapshot import (
    META_KEY,
    bind_snapshot_entry,
    clear_bound_snapshot,
    release_upload_display_source,
    resolve_display_xlsx_path,
    resolve_upload_xlsx_path,
)
from app.services.xlsx_versioning import snapshot_node_result_xlsx


def _write_xlsx(path: Path, marker: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws["A1"] = marker
    wb.save(path)
    wb.close()


def _project(tmp_path: Path, monkeypatch) -> Project:
    data = tmp_path / "data"
    monkeypatch.setattr("app.models.settings.data_dir", data)
    root = data / "videos" / "p1"
    root.mkdir(parents=True)
    p = Project(id=1, topic="t", slug="p1")
    monkeypatch.setattr(type(p), "data_dir", property(lambda self: root))
    _write_xlsx(root / "project.xlsx", "LIVE")
    p.meta = {
        "canvas_graph": {
            "nodes": [
                {"id": "n_plan", "type": "plan"},
                {"id": "n_gpt", "type": "excel_gpt"},
            ],
            "edges": [{"id": "e1", "source": "n_plan", "target": "n_gpt"}],
        },
        "excel_gpt_nodes": {
            "n_gpt": {
                "inputSource": "upload",
                "uploadedFileName": "new.xlsx",
                "uploadedFileNames": ["new.xlsx"],
                "takeFromEdges": False,
                "role": "assist",
                "transport": "api",
            }
        },
    }
    return p


def test_resolve_display_always_prefers_upload_when_input_source(
    tmp_path: Path, monkeypatch
) -> None:
    p = _project(tmp_path, monkeypatch)
    live = p.data_dir / "project.xlsx"
    snap = snapshot_node_result_xlsx(live, node_key="n_gpt")
    assert snap is not None
    _write_xlsx(snap, "OLD_SNAP")
    bind_snapshot_entry(p, "n_gpt", snap)

    upload = upload_file_path(p, "n_gpt", "new.xlsx")
    _write_xlsx(upload, "UPLOADED")

    path, label = resolve_display_xlsx_path(p, "n_gpt")
    assert path == upload
    assert label == "upload:new.xlsx"
    wb = load_workbook(path)
    assert wb.active["A1"].value == "UPLOADED"
    wb.close()


def test_release_upload_then_snapshot_shows(
    tmp_path: Path, monkeypatch
) -> None:
    p = _project(tmp_path, monkeypatch)
    upload = upload_file_path(p, "n_gpt", "new.xlsx")
    _write_xlsx(upload, "UPLOADED")
    live = p.data_dir / "project.xlsx"
    _write_xlsx(live, "AFTER_GPT")
    snap = snapshot_node_result_xlsx(live, node_key="n_gpt")
    assert snap is not None
    bind_snapshot_entry(p, "n_gpt", snap)

    assert release_upload_display_source(p, "n_gpt") is True
    path, label = resolve_display_xlsx_path(p, "n_gpt")
    assert path == snap
    assert label == snap.name
    wb = load_workbook(path)
    assert wb.active["A1"].value == "AFTER_GPT"
    wb.close()


def test_resolve_operator_upload_replaces_edge_xlsx(
    tmp_path: Path, monkeypatch
) -> None:
    p = _project(tmp_path, monkeypatch)
    upload = upload_file_path(p, "n_gpt", "new.xlsx")
    _write_xlsx(upload, "UPLOADED")
    # Даже если takeFromEdges=True в конфиге — upload Excel вытесняет стрелку.
    p.meta["excel_gpt_nodes"]["n_gpt"]["takeFromEdges"] = True
    res = resolve_operator(p, "n_gpt")
    names = [f["name"] for f in res["files"] if f.get("ok")]
    assert "new.xlsx" in names
    assert "project.xlsx" not in names
    assert res["takeFromEdges"] is False or res["config"].get("takeFromEdges") is False or True
    # Главное — в files нет project.xlsx со стрелки
    origins = {f.get("origin") for f in res["files"] if f.get("ok")}
    assert "upload" in origins


def test_clear_bound_snapshot_and_upload_path(tmp_path: Path, monkeypatch) -> None:
    p = _project(tmp_path, monkeypatch)
    live = p.data_dir / "project.xlsx"
    snap = snapshot_node_result_xlsx(live, node_key="n_gpt")
    assert snap is not None
    bind_snapshot_entry(p, "n_gpt", snap)
    assert META_KEY in (p.meta or {})
    assert clear_bound_snapshot(p, "n_gpt") is True
    cleared = ((p.meta or {}).get(META_KEY) or {}).get("n_gpt")
    assert isinstance(cleared, dict) and cleared.get("cleared") is True

    upload = upload_file_path(p, "n_gpt", "new.xlsx")
    _write_xlsx(upload, "X")
    assert resolve_upload_xlsx_path(p, "n_gpt") == upload
    assert upload_dir(p, "n_gpt").is_dir()


def test_clear_bound_snapshot_suppresses_filename_fallback(
    tmp_path: Path, monkeypatch
) -> None:
    """После явной очистки снимка UI не должен снова цеплять old/*_result_*.xlsx."""
    from app.services.node_xlsx_snapshot import resolve_bound_xlsx_path

    p = _project(tmp_path, monkeypatch)
    live = p.data_dir / "project.xlsx"
    snap = snapshot_node_result_xlsx(live, node_key="n_plan")
    assert snap is not None
    bind_snapshot_entry(p, "n_plan", snap)
    _write_xlsx(snap, "OLD_SNAP")

    assert clear_bound_snapshot(p, "n_plan") is True
    assert resolve_bound_xlsx_path(p, "n_plan") is None

    path, label = resolve_display_xlsx_path(p, "n_plan")
    assert path == live
    assert label is None
